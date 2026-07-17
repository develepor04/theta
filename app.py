import sys
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace', line_buffering=True)
    sys.stderr.reconfigure(encoding='utf-8', errors='replace', line_buffering=True)
except Exception:
    pass

from flask import Flask, request, jsonify, send_file, send_from_directory, Response, stream_with_context
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
import os
import json
import hashlib
import secrets
import uuid
import threading
import smtplib
import re as _re
import tempfile
import shutil
import time as _time
import urllib.request as _urllib
import urllib.parse as _urlparse
import zipfile
import xml.etree.ElementTree as ET
import requests
from azure.identity import DefaultAzureCredential, get_bearer_token_provider
from anthropic import AnthropicFoundry
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta, timezone, date, time
from zoneinfo import ZoneInfo
from file_processor import process_file
import anthropic 

_FILES_API_CLIENT = None

def _get_files_api_client():
    global _FILES_API_CLIENT
    if _FILES_API_CLIENT:
        return _FILES_API_CLIENT
    api_key = os.getenv('AZURE_ANTHROPIC_API_KEY') or os.getenv('AZURE_AI_API_KEY')
    if not api_key:
        return None
    _FILES_API_CLIENT = anthropic.Anthropic(api_key=api_key)
    return _FILES_API_CLIENT

# In-memory cache: filepath -> file_id
_FILE_ID_CACHE = {}
_FILE_ID_CACHE_LOCK = threading.Lock()

def _upload_file_to_files_api(file_path: str) -> str | None:
    """Upload a file to the Anthropic Files API and return the file_id."""
    client = _get_files_api_client()
    if not client:
        return None
    with _FILE_ID_CACHE_LOCK:
        if file_path in _FILE_ID_CACHE:
            return _FILE_ID_CACHE[file_path]
    try:
        with open(file_path, 'rb') as f:
            uploaded = client.beta.files.upload(
                file=(os.path.basename(file_path), f, 'application/octet-stream'),
            )
        file_id = uploaded.id
        with _FILE_ID_CACHE_LOCK:
            _FILE_ID_CACHE[file_path] = file_id
        print(f"[FILES API] Uploaded {file_path} -> {file_id}")
        return file_id
    except Exception as e:
        print(f"[FILES API] Upload failed for {file_path}: {e}")
        return None
    

_APP_ROOT = os.path.dirname(os.path.abspath(__file__))
_TEMP_DIR = os.path.join(_APP_ROOT, 'uploads', '_tmp')
os.makedirs(_TEMP_DIR, exist_ok=True)
tempfile.tempdir = _TEMP_DIR
import jwt
from functools import wraps
from dotenv import load_dotenv
from kb_file_loader import get_kb_context, reload_kb_files, get_kb_file_list

JOB_STATUS = {}
JOB_STATUS_LOCK = threading.Lock()
AI_RESPONSE_CACHE_LOCK = threading.Lock()
BASE_ANALYTICS_CACHE_LOCK = threading.Lock()
BASE_ANALYTICS_CACHE = {}
_FILE_ID_CACHE: dict = {}
_FILE_ID_CACHE_LOCK = threading.Lock()
_FILES_API_CLIENT = None

_root_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
_db_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database', '.env')
load_dotenv(_root_env_path, override=False)
load_dotenv(_db_env_path, override=False)

app = Flask(__name__, static_folder=None)
_secret_key = os.getenv('SECRET_KEY', '')
if not _secret_key:
    import warnings
    warnings.warn(
        '[SECURITY] SECRET_KEY env var is not set — using insecure default. '
        'Set SECRET_KEY in your .env file before going to production.',
        stacklevel=2,
    )
    _secret_key = 'your-secret-key-change-in-production'
app.config['SECRET_KEY'] = _secret_key

try:
    _max_upload_mb = max(10, int(os.getenv('MAX_UPLOAD_MB', '150')))
except Exception:
    _max_upload_mb = 150
app.config['MAX_CONTENT_LENGTH'] = _max_upload_mb * 1024 * 1024


@app.errorhandler(413)
def request_entity_too_large(_error):
    max_mb = int(app.config.get('MAX_CONTENT_LENGTH', 0) / (1024 * 1024))
    return jsonify({'error': f'File too large. Max upload size is {max_mb}MB.'}), 413

try:
    AI_RESPONSE_CACHE_TTL_HOURS = max(1, int(os.getenv('AI_RESPONSE_CACHE_TTL_HOURS', '24')))
except Exception:
    AI_RESPONSE_CACHE_TTL_HOURS = 24

try:
    AZURE_AGENT_RUN_TIMEOUT_SEC = max(15, int(os.getenv('AZURE_AGENT_RUN_TIMEOUT_SEC', '600')))
except Exception:
    AZURE_AGENT_RUN_TIMEOUT_SEC = 600

# Initialize Azure Anthropic Foundry client
azure_ai_client = None
AZURE_ANTHROPIC_DEPLOYMENT = None
_azure_ai_init_error = None

def _init_azure_agent_client(force: bool = False):
    global azure_ai_client, AZURE_ANTHROPIC_DEPLOYMENT, _azure_ai_init_error
    if azure_ai_client is not None and not force:
        return azure_ai_client

    azure_endpoint = os.getenv('AZURE_ANTHROPIC_ENDPOINT')
    AZURE_ANTHROPIC_DEPLOYMENT = os.getenv('AZURE_ANTHROPIC_DEPLOYMENT')
    if not azure_endpoint or not AZURE_ANTHROPIC_DEPLOYMENT:
        _azure_ai_init_error = 'AZURE_ANTHROPIC_ENDPOINT or AZURE_ANTHROPIC_DEPLOYMENT is missing'
        print(f"[WARN] {_azure_ai_init_error}")
        azure_ai_client = None
        return None

    if '/anthropic/' not in azure_endpoint:
        _azure_ai_init_error = 'AZURE_ANTHROPIC_ENDPOINT must include /anthropic/ route'
        print(f"[WARN] {_azure_ai_init_error}")
        print(f"[WARN] Current AZURE_ANTHROPIC_ENDPOINT: {azure_endpoint}")
        azure_ai_client = None
        return None

    try:
        api_key = os.getenv('AZURE_ANTHROPIC_API_KEY') or os.getenv('AZURE_AI_API_KEY')

        if api_key:
            # Key-based auth — works everywhere without Azure CLI / managed identity
            print('[INFO] Using API key auth for Azure Anthropic Foundry')
            azure_ai_client = AnthropicFoundry(
                api_key=api_key,
                base_url=azure_endpoint
            )
        else:
            # Managed identity / DefaultAzureCredential — only works on Azure-hosted infra
            print('[INFO] Using DefaultAzureCredential for Azure Anthropic Foundry')
            _disable_interactive = os.getenv('AZURE_DISABLE_INTERACTIVE_LOGIN', 'true').lower() in ('1', 'true', 'yes')
            token_provider = get_bearer_token_provider(
                DefaultAzureCredential(exclude_interactive_browser_credential=_disable_interactive),
                'https://cognitiveservices.azure.com/.default'
            )
            azure_ai_client = AnthropicFoundry(
                azure_ad_token_provider=token_provider,
                base_url=azure_endpoint
            )

        _azure_ai_init_error = None
        print('[OK] Azure Anthropic Foundry client initialized successfully')
        return azure_ai_client
    except Exception as e:
        _azure_ai_init_error = str(e)
        azure_ai_client = None
        print(f"[WARN] Failed to initialize Azure Anthropic Foundry client: {_azure_ai_init_error}")
        return None

_init_azure_agent_client()


def _now_ist_iso() -> str:
    """Return timezone-aware ISO timestamp in India Standard Time."""
    return datetime.now(ZoneInfo('Asia/Kolkata')).isoformat()


def _to_ist_iso(value: str) -> str:
    """Convert ISO-like timestamp to IST ISO.

    If timezone is missing, treat it as UTC for backward compatibility.
    """
    if not value:
        return value
    try:
        dt = datetime.fromisoformat(str(value).replace('Z', '+00:00'))
    except Exception:
        return value

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    return dt.astimezone(ZoneInfo('Asia/Kolkata')).isoformat()


def _status_name(status) -> str:
    return str(getattr(status, 'value', status)).strip().upper()


def _extract_anthropic_text(message_obj) -> str:
    parts = []
    for block in getattr(message_obj, 'content', []) or []:
        text = getattr(block, 'text', None)
        if text:
            parts.append(str(text).strip())
    return '\n'.join([p for p in parts if p]).strip()

def _get_files_api_client():
    global _FILES_API_CLIENT
    if _FILES_API_CLIENT is not None:
        return _FILES_API_CLIENT
    try:
        import anthropic as _anthropic_sdk
        api_key = os.getenv('AZURE_ANTHROPIC_API_KEY') or os.getenv('AZURE_AI_API_KEY') or os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            print('[FILES API] No API key found — Files API disabled.')
            return None
        _FILES_API_CLIENT = _anthropic_sdk.Anthropic(api_key=api_key)
        print('[FILES API] Standard Anthropic client initialised.')
        return _FILES_API_CLIENT
    except Exception as e:
        print(f'[FILES API] Could not create Anthropic client: {e}')
        return None


def _upload_file_to_files_api(file_path: str) -> str | None:
    with _FILE_ID_CACHE_LOCK:
        cached = _FILE_ID_CACHE.get(file_path)
    if cached:
        return cached
    client = _get_files_api_client()
    if client is None:
        return None
    try:
        basename = os.path.basename(file_path)
        if file_path.lower().endswith('.pdf'):
            mime = 'application/pdf'
        elif file_path.lower().endswith(('.xlsx', '.xls')):
            mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif file_path.lower().endswith('.csv'):
            mime = 'text/csv'
        else:
            mime = 'application/octet-stream'
        with open(file_path, 'rb') as fh:
            uploaded = client.beta.files.upload(file=(basename, fh, mime))
        file_id = uploaded.id
        with _FILE_ID_CACHE_LOCK:
            _FILE_ID_CACHE[file_path] = file_id
        print(f'[FILES API] Uploaded {basename} -> {file_id}')
        return file_id
    except Exception as e:
        print(f'[FILES API] Upload failed for {file_path}: {e}')
        return None


def _invalidate_file_id_cache(file_path: str | None = None):
    with _FILE_ID_CACHE_LOCK:
        if file_path:
            _FILE_ID_CACHE.pop(file_path, None)
        else:
            _FILE_ID_CACHE.clear()


# In-memory cache of parsed Excel/CSV sheets, keyed by file path and
# invalidated on (mtime, size) change. The chat endpoint re-scans every
# completed job's output file plus every raw upload plus the whole KB folder
# on EVERY message — without this, openpyxl re-parses the same unchanged
# workbooks from disk on every single chat turn, which was the dominant
# source of Theta Project Advisor's multi-second latency.
_SHEET_CACHE: dict = {}
_SHEET_CACHE_LOCK = threading.Lock()

def _read_table_sheets_cached(file_path: str) -> dict:
    import pandas as pd
    try:
        stat = os.stat(file_path)
    except OSError:
        return {}
    cache_key = (stat.st_mtime, stat.st_size)
    with _SHEET_CACHE_LOCK:
        entry = _SHEET_CACHE.get(file_path)
    if entry and entry[0] == cache_key:
        return entry[1]
    if file_path.lower().endswith('.csv'):
        sheets = {'Sheet1': pd.read_csv(file_path)}
    else:
        sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
    with _SHEET_CACHE_LOCK:
        _SHEET_CACHE[file_path] = (cache_key, sheets)
    return sheets


def _generate_claude_response(prompt: str, max_tokens: int = 1024, document_blocks: list | None = None) -> str:
    client = _init_azure_agent_client(force=False)
    deployment = os.getenv('AZURE_ANTHROPIC_DEPLOYMENT')
    endpoint = os.getenv('AZURE_ANTHROPIC_ENDPOINT')
    fallback_api_key = os.getenv('AZURE_ANTHROPIC_API_KEY') or os.getenv('AZURE_AI_API_KEY')

    if not client or not deployment:
        raise RuntimeError('Azure Anthropic service is not available.')

    def _run_with_client(active_client):
        if document_blocks:
            user_content = document_blocks + [{'type': 'text', 'text': prompt}]
        else:
            user_content = prompt
        messages = [{'role': 'user', 'content': prompt}]
        system_msg = None
        if '\n\nUSER QUERY:' in prompt:
            parts = prompt.split('\n\nUSER QUERY:', 1)
            system_msg = parts[0].strip()
            user_msg = 'USER QUERY:' + parts[1]
            messages = [{'role': 'user', 'content': user_msg}]
        full_parts = []
        max_rounds = max(1, int(os.getenv('CLAUDE_MAX_CONTINUATIONS', '2')) + 1)

        for _ in range(max_rounds):
            response = active_client.messages.create(
                model=deployment,
                messages=messages,
                max_tokens=max_tokens,
                **({"system": system_msg} if system_msg else {}),
            )

            chunk = _extract_anthropic_text(response)
            if chunk:
                full_parts.append(chunk)

            stop_reason = str(getattr(response, 'stop_reason', '') or '').strip().lower()
            if stop_reason != 'max_tokens':
                break

            # Continue only when the model stopped because output token budget was exhausted.
            messages = messages + [
                {'role': 'assistant', 'content': chunk or ''},
                {'role': 'user', 'content': 'Continue exactly from where you stopped. Do not repeat previous text.'},
            ]

        full_text = '\n'.join([p for p in full_parts if p]).strip()
        return full_text or "I couldn't generate a response at this time."

    try:
        return _run_with_client(client)
    except Exception as primary_error:
        if 'PermissionDenied' not in str(primary_error) or not fallback_api_key or not endpoint:
            raise

        try:
            key_client = AnthropicFoundry(api_key=fallback_api_key, base_url=endpoint)
            return _run_with_client(key_client)
        except Exception as fallback_error:
            raise RuntimeError(
                'Claude authentication failed for both AAD and API key. '
                'Check AZURE_ANTHROPIC_ENDPOINT and AZURE_ANTHROPIC_API_KEY/AZURE_AI_API_KEY.'
            ) from fallback_error

# Non-Azure fallback has been removed.
# All AI generation flows now run through Azure Anthropic Foundry.
_pmo_engine = None

# Enable CORS with proper configuration
ALLOWED_ORIGINS = [
    "http://localhost:5173",
    "https://localhost:5173",
    "http://localhost:5174",        
    "https://localhost:5174",      
    "http://localhost",
    "https://localhost",
    "http://137.184.8.31:9000",
    "http://137.184.8.31:8000",
    "capacitor://localhost",
    "http://localhost:3000",
    "https://localhost:3000", 
    "https://app.pmo.thetadynamics.io",
    "https://pmo.thetadynamics.io",
]
 
CORS(
    app,
    resources={r"/*": {"origins": ALLOWED_ORIGINS}},
    supports_credentials=True,
    allow_headers=["Authorization", "Content-Type", "X-App-Source", "X-User-Id", "X-Company-Id"],
    expose_headers=["Content-Disposition"],
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
)



# Handle preflight requests
@app.after_request
def after_request(response):
    origin = request.headers.get('Origin', '')
    # Only set a specific origin header if it's in our allow-list.
    # Never overwrite with '*' because that breaks credentialed (Bearer) requests.
    if origin in ALLOWED_ORIGINS:
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Credentials'] = 'true'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type,Authorization,X-App-Source,X-User-Id,X-Company-Id'
    response.headers['Access-Control-Allow-Methods'] = 'GET,PUT,POST,DELETE,OPTIONS,PATCH'
    return response

# Directories
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
DB_FOLDER = 'database'

for folder in [UPLOAD_FOLDER, OUTPUT_FOLDER, DB_FOLDER]:
    os.makedirs(folder, exist_ok=True)

BASE_DASHBOARD_CACHE_DIR = os.path.join(_APP_ROOT, OUTPUT_FOLDER, '_base_dashboard_cache')
os.makedirs(BASE_DASHBOARD_CACHE_DIR, exist_ok=True)
BASE_MERGE_OUTPUT_DIR = os.path.join(_APP_ROOT, OUTPUT_FOLDER, '_base_merge_workbook')
os.makedirs(BASE_MERGE_OUTPUT_DIR, exist_ok=True)


def _build_base_file_signature(file_path):
    """Build a content-aware signature so any workbook change creates a new analytics version."""
    try:
        stat = os.stat(file_path)
        hasher = hashlib.sha256()
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b''):
                hasher.update(chunk)
        digest = hasher.hexdigest()
        return f"{os.path.abspath(file_path)}::{stat.st_mtime_ns}::{stat.st_size}::{digest}"
    except Exception:
        return os.path.abspath(file_path)


def _build_folder_signature(folder_path, max_files=4000):
    """Build a lightweight signature for folder state (names + mtimes + sizes)."""
    try:
        if not folder_path or not os.path.isdir(folder_path):
            return ''

        entries = []
        stop_scan = False
        for root, dirs, files in os.walk(folder_path):
            dirs.sort()
            files.sort()
            for name in files:
                full_path = os.path.join(root, name)
                try:
                    st = os.stat(full_path)
                    rel_path = os.path.relpath(full_path, folder_path).replace('\\', '/')
                    entries.append(f"{rel_path}::{st.st_mtime_ns}::{st.st_size}")
                except Exception:
                    continue

                if len(entries) >= max_files:
                    stop_scan = True
                    break

            if stop_scan:
                break

        digest = hashlib.sha256('\n'.join(entries).encode('utf-8')).hexdigest()
        return f"{os.path.abspath(folder_path)}::{len(entries)}::{digest}"
    except Exception:
        return os.path.abspath(folder_path or '')


def _is_base_analytics_folder_changed(file_path, company_id=None):
    """
    Compare current source-folder signature against persisted bundle metadata for
    the current base file token. Returns (changed, current_sig, previous_sig).
    """
    try:
        source_folder = os.path.dirname(os.path.abspath(file_path))
        current_sig = _build_folder_signature(source_folder)
        cache_key = _build_base_file_signature(file_path)
        cache_token = hashlib.sha256(cache_key.encode('utf-8')).hexdigest()[:32]
        _cache_ns = str(company_id) if company_id else '_global'
        bundle_path = os.path.join(BASE_DASHBOARD_CACHE_DIR, _cache_ns, cache_token, 'bundle.json')

        if not os.path.exists(bundle_path):
            return True, current_sig, ''

        with open(bundle_path, 'r', encoding='utf-8') as f:
            persisted = json.load(f)

        previous_sig = (((persisted or {}).get('cache_meta') or {}).get('source_folder_signature') or '')
        if not previous_sig:
            return True, current_sig, previous_sig

        return previous_sig != current_sig, current_sig, previous_sig
    except Exception:
        return True, '', ''

# JSON Database files
USERS_DB = os.path.join(DB_FOLDER, 'users.json')
HISTORY_DB = os.path.join(DB_FOLDER, 'history.json')
SUBSCRIPTIONS_DB = os.path.join(DB_FOLDER, 'subscriptions.json')
NOTIFICATIONS_DB = os.path.join(DB_FOLDER, 'notifications.json')
PASSWORD_RESET_TOKENS_DB = os.path.join(DB_FOLDER, 'password_reset_tokens.json')
ENGAGE_DB = os.path.join(DB_FOLDER, 'engage_posts.json')
ENGAGE_GROUPS_DB = os.path.join(DB_FOLDER, 'engage_groups.json')
AI_RESPONSE_CACHE_DB = os.path.join(DB_FOLDER, 'ai_response_cache.json')
INTELLIGENCE_INSIGHT_CACHE_DB = os.path.join(DB_FOLDER, 'intelligence_insight_cache.json')
BASE_FILE_CONFIG_DB = os.path.join(DB_FOLDER, 'base_file_config.json')
BASE_FILE_VERSIONS_DB = os.path.join(DB_FOLDER, 'base_file_versions.json')
PPTX_SLIDE_SECTIONS_DB = os.path.join(DB_FOLDER, 'pptx_slide_sections.json')
WHATIF_REALTIME_DATA_DB = os.path.join(DB_FOLDER, 'whatif_realtime_data.json')
WHATIF_CRITICAL_DASHBOARD_DB = os.path.join(DB_FOLDER, 'whatif_critical_dashboard_data.json')
WHATIF_PREDECESSOR_SUCCESSOR_DB = os.path.join(DB_FOLDER, 'whatif_predecessor_successor_cache.json')
WHATIF_CLAUDE_RESPONSES_DB = os.path.join(DB_FOLDER, 'whatif_claude_responses.json')
WHATIF_PROJECT_UPDATE_SUMMARY_DB = os.path.join(DB_FOLDER, 'whatif_project_update_summary.json')
RECOVERY_NARRATIVE_DB = os.path.join(DB_FOLDER, 'recovery_narrative.json')
UPDATE_CHAIN_DB = os.path.join(DB_FOLDER, 'update_chain.json')
UPDATE_CHAIN_DIR = os.path.join(_APP_ROOT, OUTPUT_FOLDER, '_update_chain')
os.makedirs(UPDATE_CHAIN_DIR, exist_ok=True)
PENDING_UPLOAD_APPROVALS_DB = os.path.join(DB_FOLDER, 'pending_upload_approvals.json')
# NOTE: deviations are now stored in SQLite (database/deviations.db)

# Intelligence Hub precompute state
BENCHMARK_PRECOMPUTE_LOCK = threading.Lock()
BENCHMARK_PRECOMPUTE_STATE = {
    'status': 'idle',
    'started_at': None,
    'finished_at': None,
    'processed': 0,
    'total': 0,
    'last_error': None,
}

try:
    INTELLIGENCE_CACHE_TTL_HOURS = max(1, int(os.getenv('INTELLIGENCE_CACHE_TTL_HOURS', '6')))
except Exception:
    INTELLIGENCE_CACHE_TTL_HOURS = 6

def init_db():
    pass  # Tables managed by PostgreSQL — no JSON files needed

init_db()

# Initialize SQLite deviations database
from deviation_db import (
    init_deviations_db,
    get_all_deviations,
    get_reviewed_deviations,
    get_pending_deviations,
    get_deviation_by_id,
    deviation_exists,
    insert_deviation,
    update_deviation_fields,
    delete_deviation,
    delete_all_deviations,
    delete_deviations_by_filename,
    delete_pending_cp_deviations_by_activity_ids,
    upsert_cp_deviation,
    is_deviation_locked,
    update_reminder_timestamp,
    get_deviations_due_reminder,
    get_deviations_to_expire,
    expire_deviation,
    get_deviation_report_by_month,
)
from db_postgres import (
    save_push_subscription,
    get_push_subscriptions,
    get_all_push_subscriptions,
    delete_push_subscription,
)
init_deviations_db()

# Activity log (cross-platform tracking)
from activity_log_db import (
    init_activity_log_db,
    log_activity,
    get_activities_for_user,
    get_activities_for_manager,
    get_all_activities_admin,
    get_activity_stats,
    get_knowledge_base_summary,
    get_deviation_activity_history,
    get_recursive_deviations,
    ACTION_LOGIN, ACTION_LOGOUT, ACTION_FILE_UPLOAD, ACTION_FILE_PROCESSED,
    ACTION_DEVIATION_VIEW, ACTION_DEVIATION_APPROVE, ACTION_DEVIATION_REJECT,
    ACTION_DEVIATION_COMMENT, ACTION_NOTIFICATION_VIEW, ACTION_NOTIFICATION_READ,
    ACTION_SETTINGS_UPDATE, ACTION_PASSWORD_CHANGE, ACTION_USER_CREATED,
    ACTION_AI_CHAT, ACTION_HISTORY_VIEW, ACTION_HISTORY_DELETE, ACTION_KB_VIEW,
    SOURCE_WEB, SOURCE_MOBILE,
    LEVEL_USER, LEVEL_MANAGER, LEVEL_ADMIN, LEVEL_SYSTEM,
)
init_activity_log_db()

# Database helper functions — backed by PostgreSQL via db_postgres.py
from db_postgres import (
    read_db,
    write_db,
    read_base_file_config,
    write_base_file_config,
    log_chat,
    get_chat_history,
    get_chat_stats,
    init_chat_history_table,
    get_monthly_reports,
    upsert_monthly_report,
    init_monthly_reports_table,
    init_engage_monthly_log_table,
    init_report_builder_tables,
    get_report_catalog,
    create_catalog_item,
    update_catalog_item,
    delete_catalog_item,
    get_report_builder_config,
    save_report_builder_config,
    get_company_features,
    is_monthly_summary_posted,
    log_monthly_summary_posted,
    pg_get_user_notifications,
    pg_create_notification,
    pg_mark_notification_read,
    pg_mark_all_read,
    pg_delete_notification,
    pg_delete_read_notifications,
    pg_delete_all_notifications,
    pg_unread_count,
    pg_exists_unread_deviation_reminder,
    pg_purge_old_read,
    pg_read_base_file_versions,
    pg_write_base_file_versions,
    pg_read_update_chain,
    pg_write_update_chain,
    pg_read_ai_cache,
    pg_write_ai_cache,
    pg_read_insight_cache,
    pg_write_insight_cache,
    pg_read_whatif_realtime,
    pg_write_whatif_realtime,
    pg_read_whatif_critical,
    pg_write_whatif_critical,
    pg_read_whatif_pred_succ,
    pg_write_whatif_pred_succ,
    pg_read_whatif_proj_summary,
    pg_write_whatif_proj_summary,
    pg_read_whatif_claude,
    pg_write_whatif_claude,
    pg_read_pptx_sections,
    pg_write_pptx_sections,
    pg_read_recovery_narrative,
    pg_write_recovery_narrative,
    pg_insert_engage_post,
    pg_get_engage_post,
    pg_delete_engage_post,
    pg_update_engage_post_likes,
    pg_update_engage_post_comments,
    pg_insert_engage_group,
    pg_delete_engage_group,
    pg_upsert_history_entry,
    pg_delete_history_entry,
    pg_get_history_entry,
    pg_read_history_for_company,
    pg_read_engage_posts_for_company,
    pg_read_engage_groups_for_company,
    pg_get_engage_group_by_id,
    pg_update_user_password,
    pg_update_user_name,
    pg_update_user_role,
    pg_create_reset_token,
    pg_get_valid_reset_token,
    pg_consume_reset_tokens_for_user,
    pg_update_subscription_daily_reset,
    pg_update_subscription_plan,
    pg_get_subscription,
    pg_increment_subscription_uploads,
    pg_get_admins_for_company,
)
init_chat_history_table()
init_monthly_reports_table()
init_engage_monthly_log_table()
init_report_builder_tables()

from theta_sheet_db import (
    init_theta_sheets_db,
    get_sheet_by_company,
    get_sheet_by_id,
    upsert_sheet,
    compute_metrics_from_sheet,
    validate_sheet_data,
    subscribe,
    unsubscribe,
)
init_theta_sheets_db()


def _auto_hash_plaintext_passwords():
    """
    On startup, detect any user passwords stored in plaintext (not a recognised
    werkzeug hash format) and hash them automatically.  This lets seed SQL files
    use plain passwords without requiring a separate hash-generation step.
    """
    from werkzeug.security import generate_password_hash
    _HASH_PREFIXES = ('scrypt:', 'pbkdf2:', 'bcrypt:', 'sha256$', 'sha512$')
    try:
        users = read_db(USERS_DB)
        changed = 0
        for u in users:
            pw = u.get('password', '')
            if pw and not any(pw.startswith(p) for p in _HASH_PREFIXES):
                u['password'] = generate_password_hash(pw)
                changed += 1
        if changed:
            write_db(USERS_DB, users)
            print(f'[STARTUP] Auto-hashed {changed} plaintext password(s)')
    except Exception as e:
        print(f'[STARTUP] _auto_hash_plaintext_passwords error: {e}')

_auto_hash_plaintext_passwords()


def _parse_delay_days(value):
    """Normalize delay variance where positive=late and negative=early."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    match = _re.search(r'-?\d+(?:\.\d+)?', text)
    if not match:
        return 0.0
    try:
        num = float(match.group(0))
        low = text.lower()

        # Workbook conventions:
        # - "-103d vs BL finish" means delayed by 103 days
        # - "+206d early vs BL" means 206 days early
        if 'early' in low:
            return -abs(num)
        if 'late' in low or 'overrun' in low:
            return abs(num)
        if 'finish' in low and num < 0:
            return abs(num)
        return num
    except Exception:
        return 0.0


def _extract_slide_title(slide_xml_text):
    """Best-effort extraction of first meaningful text run from slide XML."""
    try:
        root = ET.fromstring(slide_xml_text)
    except Exception:
        return None

    ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
    texts = []
    for node in root.findall('.//a:t', ns):
        text = (node.text or '').strip()
        if text:
            texts.append(text)

    for text in texts:
        if len(text) >= 3:
            return text
    return texts[0] if texts else None


def _build_pptx_slide_sections(pptx_path):
    """Build slide section metadata from PPTX. Falls back to generic names for image-only decks."""
    sections = []
    if not os.path.exists(pptx_path):
        return sections

    with zipfile.ZipFile(pptx_path, 'r') as zf:
        slide_files = sorted(
            [name for name in zf.namelist() if name.startswith('ppt/slides/slide') and name.endswith('.xml')],
            key=lambda n: int(_re.search(r'slide(\d+)\.xml$', n).group(1)) if _re.search(r'slide(\d+)\.xml$', n) else 0,
        )

        for idx, slide_name in enumerate(slide_files, start=1):
            title = None
            try:
                with zf.open(slide_name) as fh:
                    slide_xml = fh.read().decode('utf-8', errors='ignore')
                title = _extract_slide_title(slide_xml)
            except Exception:
                title = None

            sections.append({
                'id': f'slide_{idx}',
                'slide_number': idx,
                'heading': title or f'Slide {idx}',
                'source': os.path.basename(pptx_path),
                'image_based': bool(not title),
            })

    return sections


def _write_json_file(path, payload):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, indent=2)


def _get_knowledgebase_folder(company_id=None):
    if company_id:
        try:
            from db_postgres import get_company_by_id
            company = get_company_by_id(company_id)
            slug = (company.get('slug') or '').strip() if company else ''
        except Exception:
            slug = ''
        folder_name = f"{slug}_{company_id}" if slug else str(company_id)
        kb_dir = os.path.join(_APP_ROOT, 'Knowledgebase', folder_name)
        # Auto-migrate: if old UUID-only folder exists and new slug folder doesn't, rename it
        old_dir = os.path.join(_APP_ROOT, 'Knowledgebase', str(company_id))
        if old_dir != kb_dir and os.path.isdir(old_dir) and not os.path.isdir(kb_dir):
            try:
                import shutil
                shutil.move(old_dir, kb_dir)
                print(f'[KB] Migrated folder {old_dir} → {kb_dir}')
            except Exception as _mv_err:
                print(f'[KB] Folder migration failed: {_mv_err}')
                kb_dir = old_dir
    else:
        kb_dir = os.path.join(_APP_ROOT, 'Knowledgebase')
    os.makedirs(kb_dir, exist_ok=True)
    return kb_dir


def _load_cp_activity_ids(company_id=None):
    """
    Load critical-path activity identifiers from Knowledgebase/critical_path.json.
    Returns (ids_set, names_set_lowercase) — both are sets for O(1) lookup.
    Call this per-request (file is small and cached by the OS).
    """
    cp_path = os.path.join(_get_knowledgebase_folder(company_id), 'critical_path.json')
    try:
        with open(cp_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        activities = data.get('activities', [])
        ids   = {str(a['id']).strip()         for a in activities if a.get('id')}
        names = {str(a['name']).strip().lower() for a in activities if a.get('name')}
        return ids, names
    except Exception as e:
        print(f'[CP] Could not load critical_path.json: {e}')
        return set(), set()


def _is_cp_deviation(dev, cp_ids, cp_names):
    """Return True if this deviation belongs to a critical-path activity."""
    row = dev.get('row_data') or {}
    if isinstance(row, str):
        try:
            row = json.loads(row)
        except Exception:
            row = {}
    act_id   = str(row.get('activity_id')   or '').strip()
    act_name = str(row.get('activity_name') or '').strip().lower()
    return bool(act_id and act_id in cp_ids) or bool(act_name and act_name in cp_names)


def _load_ms_activity_ids(company_id=None):
    """
    Load milestone activity identifiers from Knowledgebase/milestones.json.
    Returns (activity_ids_set, labels_set_lowercase).
    """
    ms_path = os.path.join(_get_knowledgebase_folder(company_id), 'milestones.json')
    try:
        with open(ms_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        milestones = data.get('milestones', [])
        ids   = {str(m['activity_id']).strip()       for m in milestones if m.get('activity_id')}
        names = {str(m['label']).strip().lower()     for m in milestones if m.get('label')}
        return ids, names
    except Exception as e:
        print(f'[MS] Could not load milestones.json: {e}')
        return set(), set()


def _load_cp_and_ms_ids(company_id=None):
    """Return the union of CP and milestone activity ID/name sets for deviation filtering."""
    cp_ids, cp_names = _load_cp_activity_ids(company_id)
    ms_ids, ms_names = _load_ms_activity_ids(company_id)
    return cp_ids | ms_ids, cp_names | ms_names


def _ensure_default_base_file_registered(company_id=None):
    """
    Ensure the schedule-feb-update workbook is persisted in Knowledgebase and
    marked as dashboard base file by default.
    """
    kb_dir = _get_knowledgebase_folder(company_id)
    source_name = '2026.01.30_Borouge EU3 H2 Extraction Project PMS-Rev1 dates.xlsx'
    source_path = os.path.join(_APP_ROOT, source_name)
    kb_target_path = os.path.join(kb_dir, source_name)

    # Keep an existing config if it points to a valid file
    cfg = read_base_file_config(company_id)
    configured_filename = (cfg.get('filename') or '').strip()
    configured_path = os.path.join(kb_dir, configured_filename) if configured_filename else ''
    if configured_filename and os.path.exists(configured_path):
        return cfg

    if os.path.exists(source_path):
        try:
            shutil.copy2(source_path, kb_target_path)
        except Exception as copy_err:
            print(f"[BASE FILE] Could not copy default base file to Knowledgebase: {copy_err}")

    if os.path.exists(kb_target_path):
        new_cfg = {
            'filename': source_name,
            'sheet_name': 'Sheet2',
            'is_active': True,
            'updated_at': datetime.now().isoformat(),
            'updated_by': 'system',
        }
        write_base_file_config(new_cfg, company_id)
        return new_cfg

    return cfg


def _resolve_base_file_path_and_sheet(company_id=None):
    cfg = read_base_file_config(company_id)
    kb_dir = _get_knowledgebase_folder(company_id)
    fname = (cfg.get('filename') or '').strip()
    if not fname:
        cfg = _ensure_default_base_file_registered(company_id)
        fname = (cfg.get('filename') or '').strip()
    if not fname:
        return None, None, cfg

    merged_path = os.path.join(BASE_MERGE_OUTPUT_DIR, fname)
    kb_path = os.path.join(kb_dir, fname)

    # Always prefer the continuously updated merge workbook when available.
    if os.path.exists(merged_path):
        file_path = merged_path
    else:
        file_path = kb_path

    if not os.path.exists(file_path):
        return None, None, cfg
    preferred_sheet = (cfg.get('sheet_name') or '').strip() or 'Sheet2'
    return file_path, preferred_sheet, cfg


def _resolve_premerge_base_snapshot_path(base_filename=None, company_id=None):
    """
    Resolve the latest pre-merge snapshot path for the supplied (or active) base file.
    Returns (snapshot_path, version_meta) or (None, None) when unavailable.
    """
    cfg = read_base_file_config(company_id)
    target_name = (base_filename or cfg.get('filename') or '').strip().lower()
    if not target_name:
        return None, None

    versions = pg_read_base_file_versions(company_id)
    if not isinstance(versions, list):
        return None, None

    candidates = []
    for entry in versions:
        if not isinstance(entry, dict):
            continue
        name = str(entry.get('base_filename') or '').strip().lower()
        stage = str(entry.get('stage') or '').strip().lower()
        snap = str(entry.get('snapshot_abs_path') or '').strip()
        created = str(entry.get('created_at') or '')

        if name != target_name:
            continue
        if stage != 'pre_merge':
            continue
        if not snap or not os.path.exists(snap):
            continue

        candidates.append((created, entry))

    if not candidates:
        return None, None

    candidates.sort(key=lambda x: x[0], reverse=True)
    chosen = candidates[0][1]
    return chosen.get('snapshot_abs_path'), chosen


def _resolve_aftermerge_base_snapshot_path(base_filename=None, allow_cross_file_fallback=False, company_id=None):
    """
    Resolve the latest after-merge snapshot path.
    If base_filename is provided, prefer snapshots for that base file.
    Returns (snapshot_path, version_meta) or (None, None) when unavailable.
    """
    cfg = read_base_file_config(company_id)
    target_name = (base_filename or cfg.get('filename') or '').strip().lower()

    versions = pg_read_base_file_versions(company_id)
    if not isinstance(versions, list):
        return None, None

    candidates = []
    for entry in versions:
        if not isinstance(entry, dict):
            continue

        stage = str(entry.get('stage') or '').strip().lower()
        snap = str(entry.get('snapshot_abs_path') or '').strip()
        created = str(entry.get('created_at') or '')
        name = str(entry.get('base_filename') or '').strip().lower()

        if stage != 'after_merge':
            continue
        if not snap or not os.path.exists(snap):
            continue
        if target_name and name != target_name:
            continue

        candidates.append((created, entry))

    # Optional fallback to any after-merge snapshot when no target is provided.
    if allow_cross_file_fallback and (not candidates) and (not target_name):
        for entry in versions:
            if not isinstance(entry, dict):
                continue
            stage = str(entry.get('stage') or '').strip().lower()
            snap = str(entry.get('snapshot_abs_path') or '').strip()
            created = str(entry.get('created_at') or '')
            if stage != 'after_merge':
                continue
            if not snap or not os.path.exists(snap):
                continue
            candidates.append((created, entry))

    if not candidates:
        return None, None

    candidates.sort(key=lambda x: x[0], reverse=True)
    chosen = candidates[0][1]
    return chosen.get('snapshot_abs_path'), chosen


def _get_base_versions_root(company_id=None):
    root = os.path.join(_get_knowledgebase_folder(company_id), '_versions')
    os.makedirs(root, exist_ok=True)
    return root


def _snapshot_base_file_version(base_file_path, merge_summary=None, context=None, stage='snapshot', company_id=None):
    """
    Store a copy of the current base file version and append metadata audit log.
    Returns version metadata dict on success; otherwise None.
    """
    if not base_file_path or not os.path.exists(base_file_path):
        return None

    merge_summary = merge_summary or {}
    context = context or {}

    base_name = os.path.basename(base_file_path)
    base_stem = os.path.splitext(base_name)[0]
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
    version_id = f"{base_stem}_{stamp}"

    version_dir = os.path.join(_get_base_versions_root(company_id), base_stem)
    os.makedirs(version_dir, exist_ok=True)
    snapshot_name = f"{version_id}.xlsx"
    snapshot_path = os.path.join(version_dir, snapshot_name)

    try:
        shutil.copy2(base_file_path, snapshot_path)
    except Exception as snap_err:
        print(f"[BASE VERSION] Snapshot copy failed: {snap_err}")
        return None

    version_entry = {
        'version_id': version_id,
        'stage': stage,
        'base_filename': base_name,
        'snapshot_rel_path': os.path.relpath(snapshot_path, _APP_ROOT).replace('\\', '/'),
        'snapshot_abs_path': snapshot_path,
        'created_at': datetime.now().isoformat(),
        'snapshot_size_bytes': os.path.getsize(snapshot_path) if os.path.exists(snapshot_path) else 0,
        'merge_summary': merge_summary,
        'context': {
            'job_id': context.get('job_id'),
            'uploaded_filename': context.get('uploaded_filename'),
            'uploaded_by_user_id': context.get('user_id'),
            'uploaded_by_user_name': context.get('user_name'),
        },
    }

    versions = pg_read_base_file_versions(company_id)
    if not isinstance(versions, list):
        versions = []
    versions.insert(0, version_entry)
    pg_write_base_file_versions(versions[:1000], company_id)

    return version_entry


def _read_excel_sheet_for_dashboard(file_path, preferred_sheet='Sheet2'):
    """Read the base sheet and return analytics-compatible payload."""
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet

    def _load_workbook_safe(path, data_only=True):
        """
        openpyxl compatibility fallback for files that include chartsheets with
        workbook defined names on some openpyxl versions.
        """
        try:
            return openpyxl.load_workbook(path, data_only=data_only)
        except AttributeError as exc:
            if 'defined_names' not in str(exc):
                raise

            # First try a lightweight monkey patch for buggy openpyxl versions.
            try:
                from openpyxl.chartsheet.chartsheet import Chartsheet
                if not hasattr(Chartsheet, 'defined_names'):
                    Chartsheet.defined_names = {}
                return openpyxl.load_workbook(path, data_only=data_only)
            except AttributeError as retry_exc:
                if 'defined_names' not in str(retry_exc):
                    raise

            # Final fallback: remove workbook defined names from the XLSX package
            # and reload from memory. Dashboard read paths do not require them.
            import io
            import re
            import zipfile

            with zipfile.ZipFile(path, 'r') as zin:
                workbook_xml = zin.read('xl/workbook.xml').decode('utf-8')
                workbook_xml = re.sub(
                    r'<definedNames[^>]*>.*?</definedNames>',
                    '',
                    workbook_xml,
                    flags=re.DOTALL,
                )
                workbook_xml = re.sub(r'<definedNames\s*/>', '', workbook_xml)

                buffer = io.BytesIO()
                with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        data = zin.read(item.filename)
                        if item.filename == 'xl/workbook.xml':
                            data = workbook_xml.encode('utf-8')
                        zout.writestr(item, data)

            buffer.seek(0)
            return openpyxl.load_workbook(buffer, data_only=data_only)

    def _pick_worksheet(workbook, preferred=None):
        if preferred and preferred in workbook.sheetnames:
            candidate = workbook[preferred]
            if isinstance(candidate, Worksheet):
                return candidate

        if workbook.worksheets:
            return workbook.worksheets[0]

        raise ValueError('No worksheet found in workbook')

    wb = None
    try:
        wb = _load_workbook_safe(file_path, data_only=True)
        ws = _pick_worksheet(wb, preferred=preferred_sheet)
        sheet_name = ws.title

        headers = []
        for cell in ws[1]:
            headers.append(cell.value if cell.value is not None else '')

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None and str(v).strip() != '' for v in row):
                continue
            row_data = []
            for val in row:
                if isinstance(val, datetime):
                    row_data.append(val.strftime('%Y-%m-%d'))
                else:
                    row_data.append(val)
            rows.append(row_data)

        return {
            'sheet_name': sheet_name,
            'description': 'Knowledgebase Base File (Sheet2 baseline)',
            'headers': headers,
            'rows': rows,
            'row_count': len(rows),
        }
    finally:
        if wb:
            wb.close()


def _read_excel_sheet_for_view(file_path, preferred_sheet=None, max_rows=600):
    """Read one sheet from an Excel file for UI table preview."""
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet

    def _load_workbook_safe(path, data_only=True):
        try:
            return openpyxl.load_workbook(path, data_only=data_only)
        except AttributeError as exc:
            if 'defined_names' not in str(exc):
                raise

            try:
                from openpyxl.chartsheet.chartsheet import Chartsheet
                if not hasattr(Chartsheet, 'defined_names'):
                    Chartsheet.defined_names = {}
                return openpyxl.load_workbook(path, data_only=data_only)
            except AttributeError as retry_exc:
                if 'defined_names' not in str(retry_exc):
                    raise

            import io
            import re
            import zipfile

            with zipfile.ZipFile(path, 'r') as zin:
                workbook_xml = zin.read('xl/workbook.xml').decode('utf-8')
                workbook_xml = re.sub(
                    r'<definedNames[^>]*>.*?</definedNames>',
                    '',
                    workbook_xml,
                    flags=re.DOTALL,
                )
                workbook_xml = re.sub(r'<definedNames\s*/>', '', workbook_xml)

                buffer = io.BytesIO()
                with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        data = zin.read(item.filename)
                        if item.filename == 'xl/workbook.xml':
                            data = workbook_xml.encode('utf-8')
                        zout.writestr(item, data)

            buffer.seek(0)
            return openpyxl.load_workbook(buffer, data_only=data_only)

    def _pick_worksheet(workbook, preferred=None):
        if preferred and preferred in workbook.sheetnames:
            candidate = workbook[preferred]
            if isinstance(candidate, Worksheet):
                return candidate

        if workbook.worksheets:
            return workbook.worksheets[0]

        raise ValueError('No worksheet found in workbook')

    wb = None
    try:
        wb = _load_workbook_safe(file_path, data_only=True)
        available_sheets = list(wb.sheetnames)
        ws = _pick_worksheet(wb, preferred=preferred_sheet)
        selected_sheet = ws.title

        headers = []
        for cell in ws[1]:
            headers.append(cell.value if cell.value is not None else '')

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None and str(v).strip() != '' for v in row):
                continue
            row_data = []
            for val in row:
                if isinstance(val, datetime):
                    row_data.append(val.strftime('%Y-%m-%d'))
                else:
                    row_data.append(val)
            rows.append(row_data)
            if max_rows and len(rows) >= max_rows:
                break

        total_rows = max(0, ws.max_row - 1)
        has_more_rows = bool(max_rows and total_rows > len(rows))

        return {
            'sheet_name': selected_sheet,
            'available_sheets': available_sheets,
            'headers': headers,
            'rows': rows,
            'row_count': len(rows),
            'total_rows': total_rows,
            'has_more_rows': has_more_rows,
        }
    finally:
        if wb:
            wb.close()


def _normalize_header_key(value):
    s = str(value or '').strip().lower()
    s = _re.sub(r'[^a-z0-9]+', ' ', s)
    return ' '.join(s.split())


def _find_activity_id_col(headers):
    aliases = {
        'activity id', 'activity code', 'activity_id', 'task id',
        'target activity id', 'wbs code', 'wbs code activity id', 'wbs activity id'
    }
    normalized = [_normalize_header_key(h) for h in headers]

    # exact first
    for idx, h in enumerate(normalized):
        if h in aliases:
            return idx

    # partial fallback
    for idx, h in enumerate(normalized):
        if 'activity' in h and 'id' in h:
            return idx
    return None


def _find_activity_name_col(headers):
    aliases = {
        'activity name', 'wbs activity name', 'wbs / activity name',
        'task name', 'description', 'activity description'
    }
    normalized = [_normalize_header_key(h) for h in headers]

    for idx, h in enumerate(normalized):
        if h in aliases:
            return idx

    for idx, h in enumerate(normalized):
        if 'activity' in h and 'name' in h:
            return idx
    return None


def _clean_activity_id(value):
    if value is None:
        return ''
    return str(value).strip()


def _normalize_activity_id_for_match(value):
    """
    Normalize Activity ID for robust matching across Excel typing differences.
    Examples treated as same key: 123, 123.0, ' 123 ', 'abc-01', 'ABC-01'.
    """
    cleaned = _clean_activity_id(value)
    if not cleaned:
        return ''

    # Trim trailing .0 from numeric-like ids emitted by Excel.
    if _re.fullmatch(r'\d+\.0+', cleaned):
        cleaned = cleaned.split('.', 1)[0]

    # Matching should be case-insensitive and whitespace-normalized.
    cleaned = _re.sub(r'\s+', '', cleaned).lower()
    return cleaned


def _find_column_index_by_aliases(headers, aliases):
    normalized_headers = [_normalize_header_key(h) for h in (headers or [])]
    normalized_aliases = [_normalize_header_key(a) for a in (aliases or []) if _normalize_header_key(a)]
    if not normalized_headers or not normalized_aliases:
        return None

    # Exact match first
    for idx, header in enumerate(normalized_headers):
        if header in normalized_aliases:
            return idx

    # Token-based fallback (e.g., "early progress" in "early progress %")
    for idx, header in enumerate(normalized_headers):
        for alias in normalized_aliases:
            tokens = [t for t in alias.split() if t]
            if tokens and all(t in header for t in tokens):
                return idx
    return None


def _is_na_like(value):
    if value is None:
        return False
    v = str(value).strip().lower()
    return v in {'na', 'n/a', 'n.a.', '#n/a', 'nan', 'not available', 'not applicable'}


def _to_numeric(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return None

    s = str(value).strip()
    if not s:
        return None

    is_percent = s.endswith('%')
    s = s.replace('%', '').replace(',', '').strip()
    try:
        num = float(s)
        return num / 100.0 if is_percent else num
    except Exception:
        return None

overwrite_priority_keys = {
            'plannedstartdate',
            'plannedenddate',
            'actualstartdate',
            'actualcompletiondate',
            'actualdate',
            'startdate',
            'finishdate',
            'epdate',
            'lpdate',
            'ep',
            'lp',
            # Timeline deviation aliases
            'pstart',
            'pfinish',
            'lpfinish',
            'astart',
            'afinish',
            'inputplannedstartdate',
            'inputactualstart',
            'latestart',
        }

def _should_force_overwrite_upload_value(u_key):
            normalized = _normalize_header_key(u_key)
            if not normalized:
                return False
            if normalized in overwrite_priority_keys:
                return True
            if 'planned' in normalized and 'date' in normalized:
                return True
            if 'actual' in normalized and 'date' in normalized:
                return True
            if normalized.endswith('date'):
                return True
            if normalized.endswith('start') or normalized.endswith('finish'):
                if normalized.startswith('p') or normalized.startswith('a') or normalized.startswith('lp'):
                    return True
            return False


def _detect_upload_changes(upload_file_path, context=None, company_id=None):
    """
    Compare uploaded file against the latest update chain file (or original base if chain is empty).
    Returns a list of detected changes: {activity_id, column, old_value, new_value, change_type}.
    change_type: 'date_change' | 'empty_to_filled' | 'filled_to_empty'
    """
    import openpyxl

    # Resolve what to compare against: latest update chain file or base file
    chain = pg_read_update_chain(company_id)
    approved_chain = [e for e in chain if e.get('status') == 'approved']

    if approved_chain:
        latest = approved_chain[-1]
        compare_path = os.path.join(UPDATE_CHAIN_DIR, latest['filename'])
    else:
        base_path, _, _cfg = _resolve_base_file_path_and_sheet(company_id)
        compare_path = base_path

    if not compare_path or not os.path.exists(compare_path):
        return []

    changes = []
    try:
        upload_wb = _safe_load_workbook(upload_file_path, data_only=True)
        base_wb   = _safe_load_workbook(compare_path, data_only=True)

        upload_pick = _pick_sheet_with_activity_id_standalone(upload_wb)
        base_pick   = _pick_sheet_with_activity_id_standalone(base_wb)

        if not upload_pick or not base_pick:
            return []

        u_ws, u_header_row, u_headers, u_id_col = (
            upload_pick['ws'], upload_pick['header_row'],
            upload_pick['headers'], upload_pick['id_col'],
        )
        b_ws, b_header_row, b_headers, b_id_col = (
            base_pick['ws'], base_pick['header_row'],
            base_pick['headers'], base_pick['id_col'],
        )

        # Build base index: activity_id -> {col_key: value}
        base_index = {}
        for r in range(b_header_row + 1, b_ws.max_row + 1):
            bid = _normalize_activity_id_for_match(b_ws.cell(row=r, column=b_id_col + 1).value)
            if not bid:
                continue
            row_data = {}
            for c, h in enumerate(b_headers, start=1):
                k = _normalize_header_key(h)
                if k:
                    row_data[k] = b_ws.cell(row=r, column=c).value
            base_index[bid] = row_data

        for r in range(u_header_row + 1, u_ws.max_row + 1):
            uid_raw = u_ws.cell(row=r, column=u_id_col + 1).value
            uid = _normalize_activity_id_for_match(uid_raw)
            if not uid:
                continue

            base_row = base_index.get(uid, {})
            for c, h in enumerate(u_headers, start=1):
                k = _normalize_header_key(h)
                if not k:
                    continue
                u_val = u_ws.cell(row=r, column=c).value
                b_val = base_row.get(k)

                u_empty = u_val is None or str(u_val).strip() == ''
                b_empty = b_val is None or str(b_val).strip() == ''

                if u_empty and b_empty:
                    continue
                if str(u_val) == str(b_val):
                    continue

                # Determine change type
                if b_empty and not u_empty:
                    change_type = 'empty_to_filled'
                elif not b_empty and u_empty:
                    change_type = 'filled_to_empty'
                else:
                    change_type = 'date_change' if _should_force_overwrite_upload_value(k) else 'value_change'

                changes.append({
                    'activity_id': str(uid_raw or uid),
                    'column': h,
                    'column_key': k,
                    'old_value': str(b_val) if b_val is not None else '',
                    'new_value': str(u_val) if u_val is not None else '',
                    'change_type': change_type,
                })

        upload_wb.close()
        base_wb.close()
    except Exception as e:
        print(f"[CHANGE DETECT] Error: {e}")

    return changes


def _pick_sheet_with_activity_id_standalone(workbook):
    """Minimal sheet picker for change detection — reuse logic without closure dependency."""
    for ws in workbook.worksheets:
        name_norm = _normalize_header_key(ws.title)
        if any(x in name_norm for x in ('curve', 'summary', 'dashboard')):
            continue
        for row_idx in range(1, min(120, ws.max_row) + 1):
            headers = [ws.cell(row=row_idx, column=c).value or '' for c in range(1, ws.max_column + 1)]
            id_col = _find_activity_id_col(headers)
            if id_col is not None:
                return {'ws': ws, 'header_row': row_idx, 'headers': headers, 'id_col': id_col}
    return None

def _apply_update_to_chain(upload_file_path, upload_filename, approved_by, job_id, company_id=None):
    """
    Creates the next update chain file by overlaying the upload onto the latest chain file.
    Returns the new chain entry dict.
    """
    chain = pg_read_update_chain(company_id)
    approved_chain = [e for e in chain if e.get('status') == 'approved']

    if approved_chain:
        prev = approved_chain[-1]
        source_path = os.path.join(UPDATE_CHAIN_DIR, prev['filename'])
        next_index = prev['index'] + 1
    else:
        base_path, _, _ = _resolve_base_file_path_and_sheet(company_id)
        source_path = base_path
        next_index = 1

    if not source_path or not os.path.exists(source_path):
        return {'status': 'error', 'reason': 'No source file found for chain update'}

    new_filename = f"update_{next_index}.xlsx"
    new_path = os.path.join(UPDATE_CHAIN_DIR, new_filename)

    # Copy previous chain file / base as starting point
    shutil.copy2(source_path, new_path)

    # Now merge upload into this new file (same logic as _merge_uploaded_sheet_into_base_file
    # but targeting new_path instead of BASE_MERGE_OUTPUT_DIR)
    try:
        upload_wb = _safe_load_workbook(upload_file_path, data_only=True)
        chain_wb  = _safe_load_workbook(new_path)

        upload_pick = _pick_sheet_with_activity_id_standalone(upload_wb)
        chain_pick  = _pick_sheet_with_activity_id_standalone(chain_wb)

        if upload_pick and chain_pick:
            u_ws = upload_pick['ws']
            c_ws = chain_pick['ws']
            u_headers = upload_pick['headers']
            u_header_row = upload_pick['header_row']
            u_id_col = upload_pick['id_col']
            c_header_row = chain_pick['header_row']
            c_headers = chain_pick['headers']
            c_id_col = chain_pick['id_col']

            c_header_map = {_normalize_header_key(h): i + 1 for i, h in enumerate(c_headers) if _normalize_header_key(h)}

            # Add missing columns
            for h in u_headers:
                k = _normalize_header_key(h)
                if k and k not in c_header_map:
                    new_col = c_ws.max_column + 1
                    c_ws.cell(row=c_header_row, column=new_col, value=h)
                    c_header_map[k] = new_col

            # Index chain rows
            chain_id_map = {}
            for r in range(c_header_row + 1, c_ws.max_row + 1):
                bid = _normalize_activity_id_for_match(c_ws.cell(row=r, column=c_id_col + 1).value)
                if bid:
                    chain_id_map[bid] = r

            for r in range(u_header_row + 1, u_ws.max_row + 1):
                uid_raw = u_ws.cell(row=r, column=u_id_col + 1).value
                uid = _normalize_activity_id_for_match(uid_raw)
                if not uid:
                    continue
                target_row = chain_id_map.get(uid)
                if not target_row:
                    continue  # only update existing rows, don't append
                for c, h in enumerate(u_headers, start=1):
                    k = _normalize_header_key(h)
                    if not k:
                        continue
                    u_val = u_ws.cell(row=r, column=c).value
                    if u_val is None or str(u_val).strip() == '':
                        continue
                    if _should_force_overwrite_upload_value(k):
                        col_idx = c_header_map.get(k)
                        if col_idx:
                            c_ws.cell(row=target_row, column=col_idx, value=u_val)

        chain_wb.save(new_path)
        upload_wb.close()
        chain_wb.close()
    except Exception as e:
        return {'status': 'error', 'reason': f'Chain apply failed: {e}'}

    new_entry = {
        'index': next_index,
        'filename': new_filename,
        'source_upload': upload_filename,
        'job_id': job_id,
        'approved_by': approved_by,
        'created_at': datetime.now().isoformat(),
        'status': 'approved',
    }
    chain.append(new_entry)
    pg_write_update_chain(chain, company_id)
    return {'status': 'success', **new_entry}

def _merge_uploaded_sheet_into_base_file(upload_file_path, context=None, company_id=None):
    """
    Merge uploaded activities into the configured base file by Activity ID.

    Rules:
    - Find matching activities by Activity ID.
    - Update existing rows with uploaded values (non-empty only).
    - Append new activity rows if Activity ID does not exist in baseline.
    - Add any new uploaded columns into base header row before merging.
    """
    base_path, base_sheet_name, _cfg = _resolve_base_file_path_and_sheet(company_id)

    kb_dir = _get_knowledgebase_folder(company_id)
    configured_filename = (_cfg.get('filename') or '').strip() if isinstance(_cfg, dict) else ''
    # Merge writes must happen in outputs base workbook copy, not directly on starter Knowledgebase file.
    target_base_path = os.path.join(BASE_MERGE_OUTPUT_DIR, configured_filename) if configured_filename else ''
    if not target_base_path:
        target_base_path = os.path.join(BASE_MERGE_OUTPUT_DIR, os.path.basename(upload_file_path))

    # Compare against latest stored processed/base snapshot (after_merge), not original starter base.
    effective_base_filename = configured_filename or os.path.basename(target_base_path)
    comparison_base_path, comparison_meta = _resolve_aftermerge_base_snapshot_path(effective_base_filename, company_id=company_id)
    if not comparison_base_path and os.path.exists(target_base_path):
        comparison_base_path = target_base_path
    if not comparison_base_path and base_path and os.path.exists(base_path):
        comparison_base_path = base_path

    # If this is first run with no base and no snapshots, initialize base from first upload.
    if (not comparison_base_path) and (not os.path.exists(target_base_path)):
        try:
            shutil.copy2(upload_file_path, target_base_path)
            if not configured_filename:
                write_base_file_config({
                    'filename': os.path.basename(target_base_path),
                    'sheet_name': (base_sheet_name or 'Sheet2'),
                    'is_active': True,
                    'updated_at': datetime.now().isoformat(),
                    'updated_by': (context or {}).get('user_name') or 'system',
                }, company_id)
            try:
                reload_kb_files()
            except Exception as kb_err:
                print(f"[BASE MERGE] KB reload warning (init): {kb_err}")

            return {
                'status': 'initialized',
                'reason': 'Base initialized from first uploaded file',
                'base_file': os.path.basename(target_base_path),
                'base_sheet': base_sheet_name or 'Sheet2',
                'upload_sheet': '',
                'uploaded_rows': 0,
                'matched_count': 0,
                'updated_count': 0,
                'appended_count': 0,
                'new_columns_added': 0,
                'overwrite_applied_count': 0,
                'overwrite_skipped_count': 0,
                'comparison_base_file': '',
                'comparison_version_id': None,
            }
        except Exception as init_err:
            return {
                'status': 'error',
                'reason': f'Base initialization failed: {init_err}',
                'uploaded_rows': 0,
                'matched_count': 0,
                'updated_count': 0,
                'appended_count': 0,
                'new_columns_added': 0,
            }

    # If a processed snapshot exists, promote it as the working base before merge.
    if comparison_base_path and os.path.exists(comparison_base_path):
        try:
            if os.path.abspath(comparison_base_path) != os.path.abspath(target_base_path):
                shutil.copy2(comparison_base_path, target_base_path)
        except Exception as cmp_err:
            return {
                'status': 'error',
                'reason': f'Failed to prepare comparison base: {cmp_err}',
                'uploaded_rows': 0,
                'matched_count': 0,
                'updated_count': 0,
                'appended_count': 0,
                'new_columns_added': 0,
            }

    if not os.path.exists(target_base_path):
        return {
            'status': 'skipped',
            'reason': 'No comparison/base file available for merge',
            'uploaded_rows': 0,
            'matched_count': 0,
            'updated_count': 0,
            'appended_count': 0,
            'new_columns_added': 0,
        }

    context = context or {}
    upload_wb = None
    base_wb = None
    try:
        # Capture a true BEFORE-MERGE version for audit and rollback.
        pre_version_entry = _snapshot_base_file_version(
            base_file_path=target_base_path,
            merge_summary={
                'status': 'before_merge',
                'uploaded_file': os.path.basename(upload_file_path),
                'comparison_base_file': os.path.basename(comparison_base_path) if comparison_base_path else '',
                'comparison_version_id': comparison_meta.get('version_id') if isinstance(comparison_meta, dict) else None,
            },
            context=context,
            stage='before_merge',
            company_id=company_id,
        )

        upload_wb = _safe_load_workbook(upload_file_path, data_only=True)

        # Use chartsheet-compatible loader to avoid failures on workbooks with chart sheets.
        base_wb = _safe_load_workbook(target_base_path)

        def _detect_header_row(ws, max_scan_rows=120):
            """Detect header row by finding first row containing an Activity ID-like column."""
            scan_to = min(max_scan_rows, ws.max_row)
            for row_idx in range(1, scan_to + 1):
                headers = [ws.cell(row=row_idx, column=c).value or '' for c in range(1, ws.max_column + 1)]
                if _find_activity_id_col(headers) is not None:
                    return row_idx, headers
            # Fallback to row-1 for legacy templates
            return 1, [ws.cell(row=1, column=c).value or '' for c in range(1, ws.max_column + 1)]

        def _pick_sheet_with_activity_id(workbook, preferred_name=''):
            """Pick the best worksheet for merge by locating an Activity ID-like header."""
            candidates = []
            if preferred_name:
                preferred_ws = next((ws for ws in workbook.worksheets if ws.title == preferred_name), None)
                if preferred_ws is not None:
                    candidates.append(preferred_ws)

            for ws in workbook.worksheets:
                if ws not in candidates:
                    candidates.append(ws)

            best = None
            for ws in candidates:
                ws_name_norm = _normalize_header_key(ws.title)
                # Avoid selecting S-curve/summary style tabs for row-level merge.
                if ('curve' in ws_name_norm or 's curve' in ws_name_norm or
                        'summary' in ws_name_norm or 'dashboard' in ws_name_norm):
                    continue

                header_row, headers = _detect_header_row(ws)
                id_col = _find_activity_id_col(headers)
                if id_col is None:
                    continue

                name_col = _find_activity_name_col(headers)

                non_empty_id_count = 0
                scan_to = min(ws.max_row, header_row + 80)
                for rr in range(header_row + 1, scan_to + 1):
                    val = ws.cell(row=rr, column=id_col + 1).value
                    if val is not None and str(val).strip() != '':
                        non_empty_id_count += 1

                activity_like_bonus = 0
                if name_col is not None:
                    activity_like_bonus += 500
                if preferred_name and ws.title == preferred_name:
                    activity_like_bonus += 1000

                score = (activity_like_bonus + non_empty_id_count, -header_row)
                if best is None or score > best['score']:
                    best = {
                        'ws': ws,
                        'sheet_name': ws.title,
                        'header_row': header_row,
                        'headers': headers,
                        'id_col': id_col,
                        'score': score,
                    }

            if best is not None:
                return best

            fallback_ws = candidates[0] if candidates else (workbook.active if workbook.worksheets else None)
            if fallback_ws is None:
                return None
            header_row, headers = _detect_header_row(fallback_ws)
            return {
                'ws': fallback_ws,
                'sheet_name': fallback_ws.title,
                'header_row': header_row,
                'headers': headers,
                'id_col': _find_activity_id_col(headers),
                'score': (0, -header_row),
            }

        def _has_non_empty_in_cols(ws, row_idx, columns):
            for col_idx in columns:
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and str(val).strip() != '':
                    return True
            return False

        upload_pick = _pick_sheet_with_activity_id(upload_wb, preferred_name=base_sheet_name)
        base_pick = _pick_sheet_with_activity_id(base_wb, preferred_name=base_sheet_name)
        if not upload_pick or not base_pick:
            return {
                'status': 'skipped',
                'reason': 'Could not select upload/base worksheet for merge',
                'uploaded_rows': 0,
                'matched_count': 0,
                'updated_count': 0,
                'appended_count': 0,
                'new_columns_added': 0,
            }

        u_sheet_name = upload_pick['sheet_name']
        u_ws = upload_pick['ws']
        b_sheet_name = base_pick['sheet_name']
        b_ws = base_pick['ws']

        u_header_row, uploaded_headers, u_id_col = (
            upload_pick['header_row'],
            upload_pick['headers'],
            upload_pick['id_col'],
        )
        b_header_row, base_headers, b_id_col = (
            base_pick['header_row'],
            base_pick['headers'],
            base_pick['id_col'],
        )
        if u_id_col is None or b_id_col is None:
            return {
                'status': 'skipped',
                'reason': (
                    'Activity ID column not found in upload/base sheet '
                    f"(upload_sheet='{u_sheet_name}', base_sheet='{b_sheet_name}')"
                ),
                'uploaded_rows': 0,
                'matched_count': 0,
                'updated_count': 0,
                'appended_count': 0,
                'new_columns_added': 0,
            }

        # header maps (normalized header -> column index 1-based)
        b_header_map = {}
        for idx, h in enumerate(base_headers, start=1):
            key = _normalize_header_key(h)
            if key and key not in b_header_map:
                b_header_map[key] = idx

        # Add missing uploaded columns to base headers
        new_columns_added = 0
        for uh in uploaded_headers:
            uk = _normalize_header_key(uh)
            if not uk:
                continue
            if uk not in b_header_map:
                new_col_idx = b_ws.max_column + 1
                b_ws.cell(row=b_header_row, column=new_col_idx, value=uh)
                b_header_map[uk] = new_col_idx
                new_columns_added += 1

        u_data_start = u_header_row + 1
        b_data_start = b_header_row + 1

        # Index base rows by Activity ID
        base_id_to_row = {}
        for r in range(b_data_start, b_ws.max_row + 1):
            bid = _normalize_activity_id_for_match(b_ws.cell(row=r, column=b_id_col + 1).value)
            if bid:
                base_id_to_row[bid] = r

        # Find safe append position after last real data row (not just worksheet max_row).
        base_data_cols = sorted(set([b_id_col + 1, *b_header_map.values()]))
        next_append_row = b_data_start
        for r in range(b_data_start, b_ws.max_row + 1):
            if _has_non_empty_in_cols(b_ws, r, base_data_cols):
                next_append_row = r + 1

        uploaded_rows = 0
        matched_count = 0
        updated_count = 0
        appended_count = 0
        overwrite_skipped_count = 0
        overwrite_applied_count = 0

        # Inconsistency checks (per merge session / comparison)
        SAMPLE_LIMIT = 6
        inconsistency_summary = {
            'missing_activity_id_count': 0,
            'duplicate_activity_id_count': 0,
            'activity_not_found_in_base_count': 0,
            'ep_gt_lp_count': 0,
            'na_value_count': 0,
            'missing_ep_column': False,
            'missing_lp_column': False,
            'highlights': [],
            'samples': {
                'missing_activity_id': [],
                'duplicate_activity_id': [],
                'activity_not_found_in_base': [],
                'ep_gt_lp': [],
                'na_value': [],
            },
        }

        u_ep_col = _find_column_index_by_aliases(uploaded_headers, {
            'ep', 'early progress', 'earned progress', '%ep', 'ep %'
        })
        u_lp_col = _find_column_index_by_aliases(uploaded_headers, {
            'lp', 'late progress', 'latest progress', '%lp', 'lp %'
        })
        u_activity_name_col = _find_column_index_by_aliases(uploaded_headers, {
            'activity name', 'task name', 'description', 'activity description'
        })

        overwrite_priority_keys = {
            'plannedstartdate',
            'plannedenddate',
            'actualstartdate',
            'actualcompletiondate',
            'actualdate',
            'startdate',
            'finishdate',
            'epdate',
            'lpdate',
            'ep',
            'lp',
            # Timeline deviation aliases
            'pstart',
            'pfinish',
            'lpfinish',
            'astart',
            'afinish',
            'inputplannedstartdate',
            'inputactualstart',
            'latestart',
        }

        def _should_force_overwrite_upload_value(u_key):
            normalized = _normalize_header_key(u_key)
            if not normalized:
                return False
            if normalized in overwrite_priority_keys:
                return True
            if 'planned' in normalized and 'date' in normalized:
                return True
            if 'actual' in normalized and 'date' in normalized:
                return True
            if normalized.endswith('date'):
                return True
            if normalized.endswith('start') or normalized.endswith('finish'):
                if normalized.startswith('p') or normalized.startswith('a') or normalized.startswith('lp'):
                    return True
            return False

        inconsistency_summary['missing_ep_column'] = u_ep_col is None
        inconsistency_summary['missing_lp_column'] = u_lp_col is None

        seen_upload_ids = set()

        def _push_sample(key, text):
            target = inconsistency_summary['samples'].get(key)
            if isinstance(target, list) and len(target) < SAMPLE_LIMIT and text not in target:
                target.append(text)

        for r in range(u_data_start, u_ws.max_row + 1):
            uid_raw = _clean_activity_id(u_ws.cell(row=r, column=u_id_col + 1).value)
            uid = _normalize_activity_id_for_match(uid_raw)
            if not uid:
                inconsistency_summary['missing_activity_id_count'] += 1
                _push_sample('missing_activity_id', f"Row {r}: Activity ID missing")
                continue

            if uid in seen_upload_ids:
                inconsistency_summary['duplicate_activity_id_count'] += 1
                _push_sample('duplicate_activity_id', f"{uid_raw or uid} (row {r})")
            else:
                seen_upload_ids.add(uid)

            # NA-like values in key fields
            for label, col_idx in (('Activity Name', u_activity_name_col), ('EP', u_ep_col), ('LP', u_lp_col)):
                if col_idx is None:
                    continue
                cell_val = u_ws.cell(row=r, column=col_idx + 1).value
                if _is_na_like(cell_val):
                    inconsistency_summary['na_value_count'] += 1
                    _push_sample('na_value', f"{uid}: {label}=NA")

            # EP > LP consistency check
            if u_ep_col is not None and u_lp_col is not None:
                ep_val = u_ws.cell(row=r, column=u_ep_col + 1).value
                lp_val = u_ws.cell(row=r, column=u_lp_col + 1).value
                ep_num = _to_numeric(ep_val)
                lp_num = _to_numeric(lp_val)
                if ep_num is not None and lp_num is not None and ep_num > lp_num:
                    inconsistency_summary['ep_gt_lp_count'] += 1
                    _push_sample('ep_gt_lp', f"{uid}: EP ({ep_val}) > LP ({lp_val})")

            uploaded_rows += 1

            target_row = base_id_to_row.get(uid)
            is_existing = target_row is not None
            if is_existing:
                matched_count += 1
            else:
                target_row = next_append_row
                next_append_row += 1
                b_ws.cell(row=target_row, column=b_id_col + 1, value=uid_raw)
                base_id_to_row[uid] = target_row
                appended_count += 1
                inconsistency_summary['activity_not_found_in_base_count'] += 1
                _push_sample('activity_not_found_in_base', uid_raw or uid)

            row_changed = False
            for c in range(1, u_ws.max_column + 1):
                u_header = uploaded_headers[c - 1]
                u_key = _normalize_header_key(u_header)
                if not u_key:
                    continue
                b_col = b_header_map.get(u_key)
                if not b_col:
                    continue

                u_val = u_ws.cell(row=r, column=c).value
                # Keep existing base values unless uploaded value is explicitly non-empty
                if u_val is None or str(u_val).strip() == '':
                    continue
                old_val = b_ws.cell(row=target_row, column=b_col).value

                # Existing activity row: only fill base blanks from upload.
                # Do not overwrite already populated base values.
                if is_existing:
                    old_is_blank = (old_val is None) or (str(old_val).strip() == '')
                    if old_is_blank and old_val != u_val:
                        b_ws.cell(row=target_row, column=b_col, value=u_val)
                        row_changed = True
                    elif (not old_is_blank) and old_val != u_val:
                        if _should_force_overwrite_upload_value(u_key):
                            b_ws.cell(row=target_row, column=b_col, value=u_val)
                            row_changed = True
                            overwrite_applied_count += 1
                        else:
                            overwrite_skipped_count += 1
                else:
                    # New activity row: copy uploaded non-empty value.
                    if old_val != u_val:
                        b_ws.cell(row=target_row, column=b_col, value=u_val)
                        row_changed = True

            if row_changed and is_existing:
                updated_count += 1

        summary_lines = []
        if inconsistency_summary['ep_gt_lp_count'] > 0:
            summary_lines.append(f"EP > LP: {inconsistency_summary['ep_gt_lp_count']}")
        if inconsistency_summary['na_value_count'] > 0:
            summary_lines.append(f"NA values: {inconsistency_summary['na_value_count']}")
        if inconsistency_summary['missing_activity_id_count'] > 0:
            summary_lines.append(f"Missing Activity ID: {inconsistency_summary['missing_activity_id_count']}")
        if inconsistency_summary['activity_not_found_in_base_count'] > 0:
            summary_lines.append(f"Activity ID not found in base: {inconsistency_summary['activity_not_found_in_base_count']}")
        if inconsistency_summary['duplicate_activity_id_count'] > 0:
            summary_lines.append(f"Duplicate Activity ID: {inconsistency_summary['duplicate_activity_id_count']}")
        if inconsistency_summary['missing_ep_column'] or inconsistency_summary['missing_lp_column']:
            missing_cols = []
            if inconsistency_summary['missing_ep_column']:
                missing_cols.append('EP')
            if inconsistency_summary['missing_lp_column']:
                missing_cols.append('LP')
            summary_lines.append(f"Missing columns: {', '.join(missing_cols)}")

        inconsistency_summary['highlights'] = summary_lines[:5]
        inconsistency_summary['total_issues'] = (
            inconsistency_summary['missing_activity_id_count']
            + inconsistency_summary['duplicate_activity_id_count']
            + inconsistency_summary['activity_not_found_in_base_count']
            + inconsistency_summary['ep_gt_lp_count']
            + inconsistency_summary['na_value_count']
            + (1 if inconsistency_summary['missing_ep_column'] else 0)
            + (1 if inconsistency_summary['missing_lp_column'] else 0)
        )

        base_wb.save(target_base_path)

        version_entry = None
        try:
            reload_kb_files()
        except Exception as kb_err:
            print(f"[BASE MERGE] KB reload warning: {kb_err}")

        # Persist version snapshot for audit / rollback tracking
        try:
            version_entry = _snapshot_base_file_version(
                base_file_path=target_base_path,
                merge_summary={
                    'status': 'after_merge',
                    'uploaded_rows': uploaded_rows,
                    'matched_count': matched_count,
                    'updated_count': updated_count,
                    'appended_count': appended_count,
                    'new_columns_added': new_columns_added,
                    'overwrite_applied_count': overwrite_applied_count,
                    'overwrite_skipped_count': overwrite_skipped_count,
                    'comparison_base_file': os.path.basename(comparison_base_path) if comparison_base_path else '',
                    'comparison_version_id': comparison_meta.get('version_id') if isinstance(comparison_meta, dict) else None,
                    'inconsistency_summary': inconsistency_summary,
                },
                context=context,
                stage='after_merge',
                company_id=company_id,
            )
        except Exception as version_err:
            print(f"[BASE MERGE] Version snapshot warning: {version_err}")

        return {
            'status': 'success',
            'reason': '',
            'base_file': os.path.basename(target_base_path),
            'base_sheet': b_sheet_name,
            'upload_sheet': u_sheet_name,
            'base_header_row': b_header_row,
            'upload_header_row': u_header_row,
            'uploaded_rows': uploaded_rows,
            'matched_count': matched_count,
            'updated_count': updated_count,
            'appended_count': appended_count,
            'new_columns_added': new_columns_added,
            'overwrite_applied_count': overwrite_applied_count,
            'overwrite_skipped_count': overwrite_skipped_count,
            'comparison_base_file': os.path.basename(comparison_base_path) if comparison_base_path else '',
            'comparison_version_id': comparison_meta.get('version_id') if isinstance(comparison_meta, dict) else None,
            'inconsistency_summary': inconsistency_summary,
            'pre_version_id': pre_version_entry.get('version_id') if isinstance(pre_version_entry, dict) else None,
            'version_id': version_entry.get('version_id') if isinstance(version_entry, dict) else None,
            'version_path': version_entry.get('snapshot_rel_path') if isinstance(version_entry, dict) else None,
        }
    except Exception as merge_err:
        return {
            'status': 'error',
            'reason': str(merge_err),
            'uploaded_rows': 0,
            'matched_count': 0,
            'updated_count': 0,
            'appended_count': 0,
            'new_columns_added': 0,
        }
    finally:
        if upload_wb:
            upload_wb.close()
        if base_wb:
            base_wb.close()


# Startup bootstrap for Knowledgebase base file (must run after helper defs).
try:
    _ensure_default_base_file_registered()
    reload_kb_files()
except Exception as _kb_reload_err:
    print(f"[KB] reload on startup failed: {_kb_reload_err}")


def _is_generic_advisor_template(text: str) -> bool:
    """Detect non-query-specific canned advisor intro responses."""
    t = (text or '').strip().lower()
    if not t:
        return False
    is_module_menu_template = (
        'ai project advisor is ready' in t
        and 'available modules for instant analysis' in t
        and 'please specify which analysis module or query' in t
    )
    is_persona_ack_template = (
        ('understood. i am operating as' in t or "understood. i’ll operate as" in t or "understood. i'll operate as" in t)
        and 'theta pmo ai' in t
        and ('please proceed with your first analytical query' in t or 'please specify your next action' in t)
    )
    is_file_registration_intro = (
        'your uploaded files are now registered and ready for analysis' in t
        and 'once you specify the focus' in t
    )
    is_action_menu_intro = (
        'please specify your next action' in t
        and 'for example:' in t
        and ('show top 5 delayed activities' in t or 'analyze the critical path' in t)
    )
    is_role_identity_alignment_intro = (
        ('understood.' in t or t.startswith('understood'))
        and ('role, identity, and response framework' in t or 'role, identity and response framework' in t)
        and 'theta pmo ai' in t
        and 'all subsequent analyses will' in t
    )
    is_next_query_prompt_intro = (
        ('please specify your next query' in t or 'please specify your next action' in t)
        and ('you can now issue any project-level or tracker-specific command' in t or 'all subsequent analyses will' in t)
    )
    return (
        is_module_menu_template
        or is_persona_ack_template
        or is_file_registration_intro
        or is_action_menu_intro
        or is_role_identity_alignment_intro
        or is_next_query_prompt_intro
    )


def _is_cache_entry_expired(entry: dict) -> bool:
    created_at = (entry or {}).get('created_at')
    if not created_at:
        return True
    try:
        ts = str(created_at).replace('Z', '+00:00')
        dt = datetime.fromisoformat(ts)
        if dt.tzinfo is not None:
            now = datetime.now(dt.tzinfo)
        else:
            now = datetime.now()
        return (now - dt) > timedelta(hours=AI_RESPONSE_CACHE_TTL_HOURS)
    except Exception:
        return True


def _prune_cache_entries(entries):
    return [
        e for e in (entries or [])
        if isinstance(e, dict)
        and not _is_cache_entry_expired(e)
        and not _is_generic_advisor_template(e.get('response', ''))
    ]


def get_cached_ai_response(cache_key, company_id=None):
    with AI_RESPONSE_CACHE_LOCK:
        cache_entries = pg_read_ai_cache(company_id)
        pruned_entries = _prune_cache_entries(cache_entries)
        if len(pruned_entries) != len(cache_entries):
            pg_write_ai_cache(pruned_entries[:300], company_id)

        entry = next((item for item in pruned_entries if item.get('cache_key') == cache_key), None)
        if not entry:
            return None
        response = entry.get('response')
        # Never reuse stale generic template responses as answer for real queries.
        if _is_generic_advisor_template(response):
            return None
        return response


def save_cached_ai_response(cache_key, user_id, question, response, context_hash, company_id=None):
    # Do not persist generic template text; it can poison future cache hits.
    if _is_generic_advisor_template(response):
        return
    with AI_RESPONSE_CACHE_LOCK:
        cache_entries = _prune_cache_entries(pg_read_ai_cache(company_id))
        filtered_entries = [item for item in cache_entries if item.get('cache_key') != cache_key]
        filtered_entries.insert(0, {
            'cache_key': cache_key,
            'user_id': user_id,
            'company_id': company_id,
            'question': question[:500],
            'response': response,
            'context_hash': context_hash,
            'created_at': datetime.now().isoformat(),
        })
        pg_write_ai_cache(filtered_entries[:300], company_id)

# Notification helper function
def create_notification(user_id, title, message, notification_type='info', metadata=None):
    """Create a notification for a specific user — direct PostgreSQL insert."""
    nid = pg_create_notification(user_id, title, message, notification_type, metadata or {})
    return {'id': nid, 'user_id': user_id, 'title': title, 'message': message,
            'type': notification_type, 'metadata': metadata or {}, 'read': False}


def notify_admins_and_managers(title, message, notification_type='info', metadata=None, roles=None, company_id=None):
    """Send notifications only to admins/managers within the same company."""
    if roles is None:
        roles = ['admin', 'manager', 'company_admin']
    recipients = pg_get_admins_for_company(company_id) if company_id else []
    created = []
    for user in recipients:
        if user.get('role') in roles:
            nid = pg_create_notification(user['id'], title, message, notification_type, metadata or {})
            if nid:
                created.append(nid)
    return created

# JWT Authentication decorator
def token_required(f):
    def _candidate_jwt_secrets():
        secrets = []
        current = app.config.get('SECRET_KEY')
        legacy_env = os.getenv('JWT_LEGACY_SECRET_KEY')
        legacy_default = 'your-secret-key-change-in-production'

        for s in (current, legacy_env, legacy_default):
            if s and s not in secrets:
                secrets.append(s)
        return secrets

    def _decode_auth_token(raw_token: str):
        last_invalid = None
        for secret in _candidate_jwt_secrets():
            try:
                return jwt.decode(raw_token, secret, algorithms=["HS256"])
            except jwt.ExpiredSignatureError:
                raise
            except jwt.InvalidTokenError as e:
                last_invalid = e
                continue

        if last_invalid:
            raise last_invalid
        raise jwt.InvalidTokenError('Invalid token')

    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        
        if not token:
            return jsonify({'error': 'Token is missing'}), 401
        
        try:
            if token.startswith('Bearer '):
                token = token[7:]

            if not token or token.lower() in ('null', 'undefined'):
                return jsonify({'error': 'Token is missing'}), 401
            
            data = _decode_auth_token(token)
            from db_postgres import get_user_by_id
            current_user = get_user_by_id(data['user_id'])

            if not current_user:
                return jsonify({'error': 'User not found'}), 401

            # Block users who haven't been approved yet
            if current_user.get('status', 'approved') == 'pending':
                return jsonify({'error': 'Your account is pending approval by your company admin'}), 403
            if current_user.get('status') == 'rejected':
                return jsonify({'error': 'Your account registration was rejected'}), 403
            if current_user.get('status') == 'inactive':
                return jsonify({'error': 'Your account has been deactivated. Please contact your administrator.'}), 403

        except jwt.ExpiredSignatureError:
            return jsonify({'error': 'Token has expired'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'error': 'Invalid token'}), 401

        return f(current_user, *args, **kwargs)
    
    return decorated

# ==================== THETA ENGAGE (SOCIAL FEED) ====================

def _old_post_to_viva_engage_api(client_id, client_secret, tenant_id, group_email, summary_content, date_range):
    """
    Post summary to Viva Engage community using Microsoft Graph API.
    """
    try:
        # Step 1: Get access token using requests (more reliable than urllib)
        token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        token_data = {
            'client_id': client_id,
            'client_secret': client_secret,
            'scope': 'https://graph.microsoft.com/.default',
            'grant_type': 'client_credentials'
        }
        
        print(f"[VIVA ENGAGE] Getting access token for tenant: {tenant_id}")
        token_response = requests.post(token_url, data=token_data, timeout=30)
        
        if token_response.status_code != 200:
            print(f"[VIVA ENGAGE] Token request failed: {token_response.status_code}")
            print(f"[VIVA ENGAGE] Response: {token_response.text}")
            return False, f"Failed to obtain Microsoft access token (HTTP {token_response.status_code}). Verify MICROSOFT_CLIENT_ID, MICROSOFT_CLIENT_SECRET, and MICROSOFT_TENANT_ID in .env"
            
        token_result = token_response.json()
        access_token = token_result.get('access_token')
        
        if not access_token:
            print(f"[VIVA ENGAGE] No access token in response: {token_result}")
            return False, "No access token returned by Microsoft. Verify your Azure app credentials."

        print(f"[VIVA ENGAGE] ✅ Access token obtained successfully")

        # Step 2: Find the Viva Engage group by email
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }
        
        # Search for group by email (mail attribute) — escape single quotes to prevent OData injection
        safe_email = group_email.replace("'", "''")
        search_url = f"https://graph.microsoft.com/v1.0/groups?$filter=mail eq '{safe_email}'"
        print(f"[VIVA ENGAGE] Searching for group: {group_email}")
        
        search_response = requests.get(search_url, headers=headers, timeout=30)
        
        if search_response.status_code != 200:
            print(f"[VIVA ENGAGE] Group search failed: {search_response.status_code}")
            print(f"[VIVA ENGAGE] Response: {search_response.text}")
            
            # Handle specific permission errors
            if search_response.status_code == 403:
                return False, "Permission denied (403) when searching groups. Add Group.Read.All application permission in Azure Portal and grant admin consent."
                
            return False, f"Group search failed (HTTP {search_response.status_code}). Check Azure app permissions."
            
        search_result = search_response.json()
        groups = search_result.get('value', [])
        
        if not groups:
            print(f"[VIVA ENGAGE] ❌ No group found with email: {group_email}")

            # List available groups to build a helpful error message
            all_groups_url = "https://graph.microsoft.com/v1.0/groups"
            all_groups_response = requests.get(all_groups_url, headers=headers, timeout=30)
            available = []
            if all_groups_response.status_code == 200:
                all_groups = all_groups_response.json().get('value', [])
                for g in all_groups:
                    name = g.get('displayName', '')
                    mail = g.get('mail', '')
                    print(f"[VIVA ENGAGE] Group: {name} | Email: {mail}")
                    if mail:
                        available.append(f"{name}  →  {mail}")
            groups_list = "\n".join(f"  • {g}" for g in available) if available else "  (none found)"
            return False, (
                f"No group found with email '{group_email}'.\n"
                f"Available groups:\n{groups_list}\n\n"
                f"Go to Theta Engage, disconnect, and reconnect with the correct group email."
            )
            
        group_id = groups[0]['id']
        group_name = groups[0].get('displayName', 'Unknown Group')
        print(f"[VIVA ENGAGE] ✅ Found group: {group_name} (ID: {group_id})")

        # Step 3: Create the post content (formatted for Viva Engage)
        post_content = f"""🚀 **AI Project Summary Report**

📅 **Period**: {date_range}

{summary_content}

---
*Posted via Theta PMO AI Assistant*"""

        # Step 4: Try posting to Viva Engage via different endpoints
        # Method 1: Try conversations endpoint
        conversations_url = f"https://graph.microsoft.com/v1.0/groups/{group_id}/conversations"
        post_data = {
            "topic": f"AI Project Summary - {date_range}",
            "threads": [{
                "posts": [{
                    "body": {
                        "contentType": "text",
                        "content": post_content
                    }
                }]
            }]
        }
        
        print(f"[VIVA ENGAGE] Attempting to post via conversations endpoint...")
        post_response = requests.post(conversations_url, json=post_data, headers=headers, timeout=30)
        
        if post_response.status_code in [200, 201]:
            print(f"[VIVA ENGAGE] ✅ Successfully posted to group via conversations")
            return True, None
        else:
            print(f"[VIVA ENGAGE] Conversations post failed: {post_response.status_code}")
            print(f"[VIVA ENGAGE] Response: {post_response.text}")
            
            # Handle permission errors specifically
            if post_response.status_code == 403:
                print(f"[VIVA ENGAGE] ❌ MISSING WRITE PERMISSION: Need Group.ReadWrite.All permission in Azure")
        
        # Method 2: Try Microsoft Teams message posting
        teams_url = f"https://graph.microsoft.com/v1.0/teams/{group_id}/channels/19:general/messages"
        teams_data = {
            "body": {
                "contentType": "html",
                "content": f"<h3>🚀 AI Project Summary Report</h3><p><strong>📅 Period:</strong> {date_range}</p><pre>{summary_content}</pre><p><em>Posted via Theta PMO AI Assistant</em></p>"
            }
        }
        
        print(f"[VIVA ENGAGE] Attempting to post via Teams messages...")
        teams_response = requests.post(teams_url, json=teams_data, headers=headers, timeout=30)
        
        if teams_response.status_code in [200, 201]:
            print(f"[VIVA ENGAGE] ✅ Successfully posted via Teams messages")
            return True, None
        else:
            print(f"[VIVA ENGAGE] Teams post failed: {teams_response.status_code}")
            print(f"[VIVA ENGAGE] Response: {teams_response.text}")
        
        # Method 3: Try simplified group post (Outlook Groups) 
        outlook_url = f"https://graph.microsoft.com/v1.0/groups/{group_id}/threads"
        outlook_data = {
            "topic": f"AI Project Summary - {date_range}",
            "posts": [{
                "body": {
                    "contentType": "text",
                    "content": post_content
                }
            }]
        }
        
        print(f"[VIVA ENGAGE] Attempting to post via Outlook Groups...")
        outlook_response = requests.post(outlook_url, json=outlook_data, headers=headers, timeout=30)
        
        if outlook_response.status_code in [200, 201]:
            print(f"[VIVA ENGAGE] ✅ Successfully posted via Outlook Groups")
            return True, None
        else:
            print(f"[VIVA ENGAGE] Outlook Groups post failed: {outlook_response.status_code}")
            print(f"[VIVA ENGAGE] Response: {outlook_response.text}")
            
        # Method 4: Post to Teams — look up the actual General channel ID first
        print(f"[VIVA ENGAGE] Attempting to post via Teams channel...")
        channels_url = f"https://graph.microsoft.com/v1.0/teams/{group_id}/channels"
        channels_response = requests.get(channels_url, headers=headers, timeout=30)
        if channels_response.status_code == 200:
            channels = channels_response.json().get('value', [])
            general_channel = next(
                (c for c in channels if c.get('displayName', '').lower() == 'general'), None
            )
            if general_channel:
                channel_id = general_channel.get('id')
                teams_channel_url = f"https://graph.microsoft.com/v1.0/teams/{group_id}/channels/{channel_id}/messages"
                teams_channel_data = {
                    "body": {
                        "contentType": "html",
                        "content": (
                            f"<h3>🚀 AI Project Summary Report</h3>"
                            f"<p><strong>📅 Period:</strong> {date_range}</p>"
                            f"<pre>{summary_content}</pre>"
                            f"<p><em>Posted via Theta PMO AI Assistant</em></p>"
                        )
                    }
                }
                channel_response = requests.post(teams_channel_url, json=teams_channel_data, headers=headers, timeout=30)
                if channel_response.status_code in [200, 201]:
                    print(f"[VIVA ENGAGE] ✅ Successfully posted via Teams channel")
                    return True, None
                else:
                    print(f"[VIVA ENGAGE] Teams channel post failed: {channel_response.status_code}")
                    print(f"[VIVA ENGAGE] Response: {channel_response.text}")
            else:
                print(f"[VIVA ENGAGE] No 'General' channel found in team {group_id}")
        else:
            print(f"[VIVA ENGAGE] Teams channel lookup failed: {channels_response.status_code}")
            print(f"[VIVA ENGAGE] Response: {channels_response.text}")

        # Method 5: Try community-specific endpoint (if it's a Viva Engage community)
        print(f"[VIVA ENGAGE] Attempting direct community posting...")
        community_url = f"https://graph.microsoft.com/v1.0/employeeExperience/communities"
        
        # First, try to list communities to see if this group is a community
        communities_response = requests.get(community_url, headers=headers, timeout=30)
        if communities_response.status_code == 200:
            communities = communities_response.json().get('value', [])
            target_community = None
            for community in communities:
                if community.get('groupId') == group_id:
                    target_community = community
                    break
                    
            if target_community:
                community_id = target_community.get('id')
                community_post_url = f"https://graph.microsoft.com/v1.0/employeeExperience/communities/{community_id}/posts"
                community_data = {
                    "body": {
                        "content": post_content,
                        "contentType": "text"
                    }
                }
                
                community_response = requests.post(community_post_url, json=community_data, headers=headers, timeout=30)
                if community_response.status_code in [200, 201]:
                    print(f"[VIVA ENGAGE] ✅ Successfully posted via community endpoint")
                    return True, None
                else:
                    print(f"[VIVA ENGAGE] Community posting failed: {community_response.status_code}")
                    print(f"[VIVA ENGAGE] Response: {community_response.text}")

        # Method 6: Email-to-group (most reliable with app-only tokens)
        # M365 groups accept emails which appear as posts in the group conversation.
        # Requires: Mail.Send application permission in Azure + MICROSOFT_SENDER_EMAIL in .env
        sender_email = os.getenv('MICROSOFT_SENDER_EMAIL', '')
        if sender_email:
            print(f"[VIVA ENGAGE] Attempting email-to-group via {sender_email}...")
            send_mail_url = f"https://graph.microsoft.com/v1.0/users/{sender_email}/sendMail"
            email_body = (
                f"<h3>🚀 AI Project Summary Report</h3>"
                f"<p><strong>📅 Period:</strong> {date_range}</p>"
                f"<pre style='font-family:monospace;white-space:pre-wrap'>{summary_content}</pre>"
                f"<p><em>Posted via Theta PMO AI Assistant</em></p>"
            )
            mail_payload = {
                "message": {
                    "subject": f"AI Project Summary — {date_range}",
                    "body": {"contentType": "HTML", "content": email_body},
                    "toRecipients": [{"emailAddress": {"address": group_email}}]
                },
                "saveToSentItems": False
            }
            mail_response = requests.post(send_mail_url, json=mail_payload, headers=headers, timeout=30)
            if mail_response.status_code in [200, 202]:
                print(f"[VIVA ENGAGE] ✅ Successfully posted via email-to-group")
                return True, None
            else:
                print(f"[VIVA ENGAGE] Email-to-group failed: {mail_response.status_code}")
                print(f"[VIVA ENGAGE] Response: {mail_response.text}")
        else:
            print(f"[VIVA ENGAGE] Skipping email-to-group: MICROSOFT_SENDER_EMAIL not set in .env")

        # If all API methods fail — guide the user to the email-to-group fix
        error_msg = (
            "All direct Graph API posting methods failed. This is expected — Microsoft does not allow "
            "app-only tokens to post to Groups/Teams/Viva Engage directly.\n\n"
            "✅ RECOMMENDED FIX (2 steps):\n"
            "  1. In Azure Portal → App Registration → API Permissions:\n"
            "     Add  Mail.Send  (Application permission) and grant admin consent.\n"
            "  2. Add this line to your .env file:\n"
            "     MICROSOFT_SENDER_EMAIL=anurag@thetadynamics.io\n"
            "        (any valid Microsoft 365 user in your tenant)\n\n"
            "Then restart the backend — summaries will be emailed to the group and appear as posts."
        )
        print(f"[VIVA ENGAGE] {error_msg}")
        return False, error_msg
            
    except Exception as e:
        print(f"[VIVA ENGAGE] ❌ Unexpected error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, str(e)

# ==================== AUTH ENDPOINTS ====================

@app.route('/api/auth/signup', methods=['POST'])
def signup():
    try:
        data = request.get_json()
        print(f"[SIGNUP] Received data: {data}")
    except Exception as e:
        print(f"[SIGNUP] Error parsing request: {e}")
        return jsonify({'error': f'Invalid request format: {str(e)}'}), 400

    if not data or not data.get('email') or not data.get('password') or not data.get('name'):
        return jsonify({'error': 'Missing required fields'}), 400

    company_id = (data.get('company_id') or '').strip()
    if not company_id:
        return jsonify({'error': 'Please select a company to register under'}), 400

    try:
        from db_postgres import (
            get_user_by_email, create_user_direct, create_subscription_direct,
            get_company_by_id, get_users_by_company,
        )

        if get_user_by_email(data['email']):
            return jsonify({'error': 'Email already registered'}), 400

        company = get_company_by_id(company_id)
        if not company:
            return jsonify({'error': 'Selected company does not exist'}), 400

        user_id = str(uuid.uuid4())
        create_user_direct(
            user_id=user_id,
            email=data['email'],
            name=data['name'],
            role='user',
            password_hash=generate_password_hash(data['password']),
            company_id=company_id,
            status='pending',
        )
        create_subscription_direct(user_id)
        print(f"[SIGNUP] User created (pending): {data['email']} → company {company['name']}")

        # Notify company_admins of the pending registration (in-app + email)
        try:
            company_admins = [
                u for u in get_users_by_company(company_id)
                if u.get('role') in ('company_admin', 'admin')
            ]
            _frontend_url = os.getenv('FRONTEND_URL', 'http://localhost:3000').rstrip('/')
            _approval_url = f"{_frontend_url}/company-admin"
            for admin in company_admins:
                create_notification(
                    user_id=admin['id'],
                    title='New user registration pending',
                    message=f"{data['name']} ({data['email']}) has registered and is awaiting your approval.",
                    notification_type='user_pending',
                    metadata={'pending_user_id': user_id},
                )
                try:
                    _send_admin_signup_notification_email(
                        admin_email=admin['email'],
                        admin_name=admin.get('name', ''),
                        new_user_name=data['name'],
                        new_user_email=data['email'],
                        company_name=company['name'],
                        approval_url=_approval_url,
                    )
                except Exception as _email_err:
                    print(f"[SIGNUP] Admin email notification failed for {admin['email']}: {_email_err}")
        except Exception:
            pass

        try:
            _source = request.headers.get('X-App-Source', SOURCE_WEB)
            log_activity(
                action_type=ACTION_USER_CREATED,
                user_id=user_id,
                user_name=data['name'],
                user_role='user',
                company_id=company_id,
                description=f"New account registered (pending): {data['name']} ({data['email']}) for {company['name']}",
                source=_source,
                level=LEVEL_ADMIN,
                metadata={'email': data['email'], 'company_id': company_id},
                ip_address=request.remote_addr,
            )
        except Exception:
            pass

        return jsonify({
            'message': 'Registration successful. Your account is pending approval by your company admin.',
            'status': 'pending',
        }), 201
    except Exception as e:
        print(f"[SIGNUP] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        print(f"[LOGIN] Received data: {data}")
    except Exception as e:
        print(f"[LOGIN] Error parsing request: {e}")
        return jsonify({'error': f'Invalid request format: {str(e)}'}), 400
    
    if not data or not data.get('email') or not data.get('password'):
        print(f"[LOGIN] Missing fields - data: {data}")
        return jsonify({'error': 'Missing email or password'}), 400
    
    try:
        from db_postgres import get_user_by_email
        print(f"[LOGIN] Searching for user: {data['email']}")

        user = get_user_by_email(data['email'])

        if not user:
            print(f"[LOGIN] User not found: {data['email']}")
            return jsonify({'error': 'Invalid email or password'}), 401

        print(f"[LOGIN] User found, checking password")

        if not check_password_hash(user['password'], data['password']):
            print(f"[LOGIN] Invalid password for: {data['email']}")
            return jsonify({'error': 'Invalid email or password'}), 401
        
        # Enforce approval status
        user_status = user.get('status', 'approved')
        if user_status == 'pending':
            return jsonify({'error': 'Your account is pending approval by your company admin'}), 403
        if user_status == 'rejected':
            return jsonify({'error': 'Your account registration was rejected. Contact your company admin.'}), 403
        if user_status == 'inactive':
            return jsonify({'error': 'Your account has been deactivated. Please contact your administrator.'}), 403

        # JWT includes company_id so every protected route gets it automatically
        token = jwt.encode({
            'user_id': user['id'],
            'company_id': user.get('company_id'),
            'exp': datetime.utcnow() + timedelta(days=7)
        }, app.config['SECRET_KEY'], algorithm='HS256')

        print(f"[LOGIN] Login successful for: {user['email']}")

        try:
            _source = request.headers.get('X-App-Source', SOURCE_WEB)
            log_activity(
                action_type=ACTION_LOGIN,
                user_id=user['id'],
                user_name=user.get('name', ''),
                user_role=user.get('role', 'user'),
                company_id=user.get('company_id'),
                description=f"{user.get('name', 'User')} logged in via {_source}",
                source=_source,
                level=LEVEL_MANAGER if user.get('role') in ('admin', 'manager') else LEVEL_USER,
                metadata={'email': user['email']},
                ip_address=request.remote_addr,
            )
        except Exception:
            pass

        try:
            _start_benchmark_precompute_async(triggered_by=user['id'])
        except Exception:
            pass

        company_name = None
        if user.get('company_id'):
            try:
                from db_postgres import get_company_by_id as _get_company_by_id
                _company = _get_company_by_id(user['company_id'])
                company_name = _company.get('name') if _company else None
            except Exception:
                pass

        return jsonify({
            'token': token,
            'user': {
                'id': user['id'],
                'email': user['email'],
                'name': user['name'],
                'role': user.get('role', 'user'),
                'company_id': user.get('company_id'),
                'company_name': company_name,
                'must_change_password': _requires_first_login_password_change(user),
            }
        })
    except Exception as e:
        print(f"[LOGIN] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Server error: {str(e)}'}), 500


def _parse_iso_datetime(iso_text):
    try:
        if not iso_text:
            return None
        return datetime.fromisoformat(str(iso_text).replace('Z', '+00:00')).replace(tzinfo=None)
    except Exception:
        return None


def _requires_first_login_password_change(user):
    """Return True if this local-account user must change password after first login."""
    if not isinstance(user, dict):
        return False

    if (user.get('auth_type') or '').strip().lower() == 'microsoft':
        return False

    if not (user.get('password') or '').strip():
        return False

    return bool(user.get('must_change_password', True))


def _build_password_reset_link(raw_token, frontend_origin=None):
    reset_base = os.getenv('RESET_PASSWORD_URL', '').strip()
    if not reset_base:
        origin = (frontend_origin or '').strip().rstrip('/')
        if origin in ALLOWED_ORIGINS:
            frontend_url = origin
        else:
            frontend_url = os.getenv('FRONTEND_URL', 'http://localhost:3000').rstrip('/')
        reset_base = f"{frontend_url}/reset-password"
    sep = '&' if '?' in reset_base else '?'
    return f"{reset_base}{sep}token={_urlparse.quote(raw_token)}"


def _send_password_reset_email(recipient_email, recipient_name, reset_link):
    sender_email = os.getenv('SYSTEM_EMAIL', '').strip()
    sender_password = os.getenv('GMAIL_APP_PASSWORD', '').strip()
    if not sender_password:
        raise RuntimeError('GMAIL_APP_PASSWORD is not configured')

    subject = 'Reset your Theta PMO password'
    display_name = (recipient_name or 'User').strip() or 'User'

    text_body = (
        f"Hello {display_name},\n\n"
        "We received a request to reset your password for Theta PMO.\n"
        f"Reset link: {reset_link}\n\n"
        "This link expires in 30 minutes and can only be used once.\n"
        "If you did not request this, you can safely ignore this email.\n"
    )

    html_body = f"""\
    <html>
      <body style=\"font-family:Segoe UI,Arial,sans-serif;background:#f7faf9;padding:24px;\">
        <div style=\"max-width:560px;margin:0 auto;background:#ffffff;border-radius:12px;border:1px solid #e5e7eb;padding:24px;\">
          <h2 style=\"margin:0 0 12px;color:#064e3b;\">Password reset request</h2>
          <p style=\"margin:0 0 12px;color:#374151;line-height:1.6;\">Hello {display_name},</p>
          <p style=\"margin:0 0 16px;color:#374151;line-height:1.6;\">We received a request to reset your Theta PMO password.</p>
          <p style=\"margin:0 0 18px;\">
            <a href=\"{reset_link}\" style=\"display:inline-block;background:#059669;color:#ffffff;text-decoration:none;padding:10px 16px;border-radius:8px;font-weight:600;\">Reset Password</a>
          </p>
          <p style=\"margin:0 0 8px;color:#6b7280;font-size:13px;line-height:1.6;\">This link expires in 30 minutes and can only be used once.</p>
          <p style=\"margin:0;color:#6b7280;font-size:13px;line-height:1.6;\">If you did not request this, you can ignore this email.</p>
        </div>
      </body>
    </html>
    """

    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = f"Theta PMO AI <{sender_email}>"
    msg['To'] = recipient_email
    msg.attach(MIMEText(text_body, 'plain'))
    msg.attach(MIMEText(html_body, 'html'))

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, recipient_email, msg.as_string())


def _send_admin_signup_notification_email(admin_email, admin_name, new_user_name, new_user_email, company_name, approval_url):
    sender_email = os.getenv('SYSTEM_EMAIL', '').strip()
    sender_password = os.getenv('GMAIL_APP_PASSWORD', '').strip()
    if not sender_email or not sender_password:
        return

    display = (admin_name or 'Admin').strip() or 'Admin'
    subject = f'New user registration pending approval — {company_name}'

    text_body = (
        f"Hello {display},\n\n"
        f"{new_user_name} ({new_user_email}) has registered and is awaiting your approval for {company_name}.\n\n"
        f"Approve or reject at: {approval_url}\n\n"
        "This is an automated notification from Theta PMO AI.\n"
    )

    html_body = f"""\
    <html>
      <body style="font-family:Segoe UI,Arial,sans-serif;background:#f7faf9;padding:24px;">
        <div style="max-width:560px;margin:0 auto;background:#ffffff;border-radius:12px;border:1px solid #e5e7eb;padding:24px;">
          <h2 style="margin:0 0 12px;color:#064e3b;">New user registration pending</h2>
          <p style="margin:0 0 12px;color:#374151;line-height:1.6;">Hello {display},</p>
          <p style="margin:0 0 8px;color:#374151;line-height:1.6;">
            <strong>{new_user_name}</strong> (<a href="mailto:{new_user_email}">{new_user_email}</a>)
            has registered and is awaiting your approval for <strong>{company_name}</strong>.
          </p>
          <p style="margin:16px 0;">
            <a href="{approval_url}" style="display:inline-block;background:#059669;color:#ffffff;text-decoration:none;padding:10px 16px;border-radius:8px;font-weight:600;">Review Registration</a>
          </p>
          <p style="margin:0;color:#6b7280;font-size:13px;">Theta PMO AI — automated notification</p>
        </div>
      </body>
    </html>
    """

    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = f"Theta PMO AI <{sender_email}>"
    msg['To'] = admin_email
    msg.attach(MIMEText(text_body, 'plain'))
    msg.attach(MIMEText(html_body, 'html'))

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, admin_email, msg.as_string())


def _send_processing_inconsistency_email(
    *,
    job_id,
    filename,
    user_name,
    processing_result,
    base_merge_summary,
    deviations_detected,
):
    """Send a summary email for data-processing inconsistencies after upload completion.

    Uses environment variables (secure, no hardcoded secrets):
    - INCONSISTENCY_EMAIL_FROM (default: anuragkatre36@gmail.com)
    - INCONSISTENCY_EMAIL_TO (default: Abhishek@apliaglobal.com)
    - INCONSISTENCY_EMAIL_PASSWORD (fallback: GMAIL_APP_PASSWORD)
    - INCONSISTENCY_EMAIL_ENABLED (default: true)
    """
    try:
        enabled = os.getenv('INCONSISTENCY_EMAIL_ENABLED', 'true').strip().lower() in ('1', 'true', 'yes')
        if not enabled:
            return False, 'disabled'

        sender_email = os.getenv('INCONSISTENCY_EMAIL_FROM', 'anuragkatre36@gmail.com').strip()
        recipient_email = os.getenv('INCONSISTENCY_EMAIL_TO', 'Abhishek@apliaglobal.com').strip()
        sender_password = os.getenv('INCONSISTENCY_EMAIL_PASSWORD', '').strip() or os.getenv('GMAIL_APP_PASSWORD', '').strip()

        if not sender_email or not recipient_email:
            return False, 'sender/recipient missing'
        if not sender_password:
            return False, 'email password not configured'

        processing_result = processing_result or {}
        base_merge_summary = base_merge_summary or {}
        inconsistency = base_merge_summary.get('inconsistency_summary') or {}
        tracker_results = processing_result.get('results', []) or []
        failed_sheets = processing_result.get('failed_sheets', []) or []
        tracker_errors = [
            {
                'sheet_name': r.get('sheet_name', ''),
                'error': r.get('error', 'Unknown processing error'),
            }
            for r in tracker_results
            if str(r.get('status', '')).lower() == 'error'
        ]

        inconsistency_total = int(inconsistency.get('total_issues', 0) or 0)
        data_issue_present = bool(
            inconsistency_total > 0
            or tracker_errors
            or failed_sheets
            or str(base_merge_summary.get('status', '')).lower() in ('error', 'skipped')
        )
        if not data_issue_present:
            return False, 'no data inconsistency'

        subject = f"[PMO] Data Inconsistency Report — {filename} ({job_id[:8]})"
        detected_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        summary_lines = []
        if inconsistency_total > 0:
            summary_lines.append(f"- Inconsistency total: {inconsistency_total}")
        if tracker_errors:
            summary_lines.append(f"- Tracker errors: {len(tracker_errors)}")
        if failed_sheets:
            summary_lines.append(f"- Template mismatch sheets: {len(failed_sheets)}")
        if str(base_merge_summary.get('status', '')).lower() in ('error', 'skipped'):
            summary_lines.append(f"- Base merge status: {base_merge_summary.get('status')}")

        highlights = inconsistency.get('highlights', []) if isinstance(inconsistency.get('highlights', []), list) else []
        samples_map = inconsistency.get('samples', {}) if isinstance(inconsistency.get('samples', {}), dict) else {}
        sample_rows = []
        for _k, arr in samples_map.items():
            if isinstance(arr, list):
                sample_rows.extend(arr)
        sample_rows = sample_rows[:8]

        text_body = (
            f"Data inconsistency report\n\n"
            f"Detected at: {detected_at}\n"
            f"Job ID: {job_id}\n"
            f"Uploaded file: {filename}\n"
            f"Uploaded by: {user_name}\n\n"
            f"Summary:\n" + ('\n'.join(summary_lines) if summary_lines else '- No summary lines') + "\n\n"
            f"Highlights:\n" + ('\n'.join(f"- {h}" for h in highlights) if highlights else '- None') + "\n\n"
            f"Tracker errors:\n" + (
                '\n'.join(f"- {e.get('sheet_name')}: {e.get('error')}" for e in tracker_errors)
                if tracker_errors else '- None'
            ) + "\n\n"
            f"Template mismatch / failed sheets:\n" + (
                '\n'.join(f"- {s.get('sheet_name', '')}: {s.get('error', s.get('reason', 'Template mismatch'))}" for s in failed_sheets)
                if failed_sheets else '- None'
            ) + "\n\n"
            f"Inconsistency sample rows:\n" + (
                '\n'.join(f"- {s}" for s in sample_rows) if sample_rows else '- None'
            ) + "\n"
        )

        html_body = f"""
        <html>
          <body style=\"font-family:Segoe UI,Arial,sans-serif;background:#f8fafc;padding:20px;\">
            <div style=\"max-width:760px;margin:0 auto;background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:20px;\">
              <h2 style=\"margin:0 0 10px;color:#991b1b;\">⚠️ Data Inconsistency Report</h2>
              <p style=\"margin:0 0 8px;color:#334155;\"><strong>Detected at:</strong> {detected_at}</p>
              <p style=\"margin:0 0 8px;color:#334155;\"><strong>Job ID:</strong> {job_id}</p>
              <p style=\"margin:0 0 8px;color:#334155;\"><strong>Uploaded file:</strong> {filename}</p>
              <p style=\"margin:0 0 16px;color:#334155;\"><strong>Uploaded by:</strong> {user_name}</p>

              <h3 style=\"margin:16px 0 8px;color:#0f172a;\">Summary</h3>
              <ul>{''.join(f'<li>{line[2:]}</li>' for line in summary_lines) if summary_lines else '<li>No summary lines</li>'}</ul>

              <h3 style=\"margin:16px 0 8px;color:#0f172a;\">Highlights</h3>
              <ul>{''.join(f'<li>{h}</li>' for h in highlights) if highlights else '<li>None</li>'}</ul>

              <h3 style=\"margin:16px 0 8px;color:#0f172a;\">Tracker errors</h3>
              <ul>{''.join(f"<li>{e.get('sheet_name')}: {e.get('error')}</li>" for e in tracker_errors) if tracker_errors else '<li>None</li>'}</ul>

              <h3 style=\"margin:16px 0 8px;color:#0f172a;\">Template mismatch / failed sheets</h3>
              <ul>{''.join(f"<li>{s.get('sheet_name', '')}: {s.get('error', s.get('reason', 'Template mismatch'))}</li>" for s in failed_sheets) if failed_sheets else '<li>None</li>'}</ul>

              <h3 style=\"margin:16px 0 8px;color:#0f172a;\">Inconsistency sample rows</h3>
              <ul>{''.join(f'<li>{s}</li>' for s in sample_rows) if sample_rows else '<li>None</li>'}</ul>
            </div>
          </body>
        </html>
        """

        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg.attach(MIMEText(text_body, 'plain'))
        msg.attach(MIMEText(html_body, 'html'))

        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())

        return True, 'sent'
    except Exception as e:
        return False, str(e)


@app.route('/api/auth/forgot-password', methods=['POST'])
def forgot_password():
    data = request.get_json() or {}
    email = (data.get('email') or '').strip().lower()
    frontend_origin = (data.get('frontend_origin') or '').strip()

    if not email:
        return jsonify({'error': 'Email is required'}), 400

    generic_message = 'If the account exists, a password reset link has been sent.'

    try:
        from db_postgres import get_user_by_email as _get_user_by_email
        user = _get_user_by_email(email)

        if not user:
            return jsonify({'message': generic_message}), 200

        raw_token = secrets.token_urlsafe(32)
        token_hash = hashlib.sha256(raw_token.encode('utf-8')).hexdigest()
        expires_at = (datetime.utcnow() + timedelta(minutes=30)).isoformat()

        pg_create_reset_token(token_hash, user['id'], expires_at)

        reset_link = _build_password_reset_link(raw_token, frontend_origin=frontend_origin)
        try:
            _send_password_reset_email(user.get('email', email), user.get('name', 'User'), reset_link)
        except Exception as email_err:
            print(f"[FORGOT-PASSWORD] Email send failed: {email_err}")

        return jsonify({'message': generic_message}), 200

    except Exception as e:
        print(f"[FORGOT-PASSWORD] Unexpected error: {e}")
        return jsonify({'error': 'Failed to process password reset request'}), 500


@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    data = request.get_json() or {}
    raw_token = (data.get('token') or '').strip()
    new_password = data.get('new_password') or ''

    if not raw_token or not new_password:
        return jsonify({'error': 'Token and new_password are required'}), 400

    if len(new_password) < 6:
        return jsonify({'error': 'Password must be at least 6 characters'}), 400

    try:
        token_hash = hashlib.sha256(raw_token.encode('utf-8')).hexdigest()

        matched_token = pg_get_valid_reset_token(token_hash)
        if not matched_token:
            return jsonify({'error': 'Invalid or expired reset token'}), 400

        from db_postgres import get_user_by_id as _get_user_by_id
        user = _get_user_by_id(matched_token['user_id'])
        if not user:
            return jsonify({'error': 'User not found for this reset token'}), 404

        pg_update_user_password(user['id'], generate_password_hash(new_password))
        pg_consume_reset_tokens_for_user(user['id'])

        try:
            _source = request.headers.get('X-App-Source', SOURCE_WEB)
            log_activity(
                action_type=ACTION_PASSWORD_CHANGE,
                user_id=user['id'],
                user_name=user.get('name', ''),
                user_role=user.get('role', 'user'),
                company_id=user.get('company_id'),
                description=f"{user.get('name', 'User')} reset account password",
                source=_source,
                level=LEVEL_USER,
                metadata={'method': 'forgot-password-flow'},
                ip_address=request.remote_addr,
            )
        except Exception:
            pass

        return jsonify({'message': 'Password reset successful'}), 200

    except Exception as e:
        print(f"[RESET-PASSWORD] Unexpected error: {e}")
        return jsonify({'error': 'Failed to reset password'}), 500


@app.route('/api/auth/change-password', methods=['POST'])
@token_required
def change_password(current_user):
    data = request.get_json() or {}
    current_password = data.get('current_password') or ''
    new_password = data.get('new_password') or ''

    if not current_password or not new_password:
        return jsonify({'error': 'current_password and new_password are required'}), 400

    if len(new_password) < 6:
        return jsonify({'error': 'Password must be at least 6 characters'}), 400

    if current_password == new_password:
        return jsonify({'error': 'New password must be different from current password'}), 400

    try:
        from db_postgres import get_user_by_id as _get_user_by_id
        user = _get_user_by_id(current_user['id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404

        if (user.get('auth_type') or '').strip().lower() == 'microsoft':
            return jsonify({'error': 'Password change is not available for Microsoft sign-in users'}), 400

        user_password_hash = user.get('password') or ''
        if not user_password_hash or not check_password_hash(user_password_hash, current_password):
            return jsonify({'error': 'Current password is incorrect'}), 400

        pg_update_user_password(user['id'], generate_password_hash(new_password))

        try:
            _source = request.headers.get('X-App-Source', SOURCE_WEB)
            log_activity(
                action_type=ACTION_PASSWORD_CHANGE,
                user_id=user['id'],
                user_name=user.get('name', ''),
                user_role=user.get('role', 'user'),
                company_id=user.get('company_id'),
                description=f"{user.get('name', 'User')} changed account password",
                source=_source,
                level=LEVEL_USER,
                metadata={'method': 'authenticated-change-password'},
                ip_address=request.remote_addr,
            )
        except Exception:
            pass

        return jsonify({
            'message': 'Password changed successfully',
            'user': {
                'id': user['id'],
                'email': user['email'],
                'name': user['name'],
                'role': user.get('role', 'user'),
                'must_change_password': _requires_first_login_password_change(user),
            }
        }), 200

    except Exception as e:
        print(f"[CHANGE-PASSWORD] Unexpected error: {e}")
        return jsonify({'error': 'Failed to change password'}), 500


@app.route('/api/auth/microsoft', methods=['POST'])
def microsoft_login():
    """
    Microsoft / Teams SSO endpoint.
    User must already be registered and approved. No auto-provisioning.
    """
    import urllib.request as _urllib
    try:
        data         = request.get_json() or {}
        access_token = data.get('access_token', '').strip()
        if not access_token:
            return jsonify({'error': 'access_token is required'}), 400

        req = _urllib.Request(
            'https://graph.microsoft.com/v1.0/me',
            headers={'Authorization': f'Bearer {access_token}'},
        )
        try:
            with _urllib.urlopen(req, timeout=10) as resp:
                ms_user = json.loads(resp.read().decode())
        except Exception as graph_err:
            print(f"[MS-LOGIN] Graph call failed: {graph_err}")
            return jsonify({'error': 'Could not verify Microsoft identity. Please try again.'}), 401

        ms_email = (ms_user.get('mail') or ms_user.get('userPrincipalName') or '').lower()
        ms_name  = ms_user.get('displayName') or ms_user.get('givenName') or ms_email.split('@')[0]
        ms_id    = ms_user.get('id', '')

        if not ms_email:
            return jsonify({'error': 'Could not retrieve email from Microsoft account.'}), 400

        from db_postgres import get_user_by_email
        user = get_user_by_email(ms_email)

        if not user:
            return jsonify({
                'error': 'No account found for this Microsoft email. Please register first.'
            }), 404

        user_status = user.get('status', 'approved')
        if user_status == 'pending':
            return jsonify({'error': 'Your account is pending approval by your company admin'}), 403
        if user_status == 'rejected':
            return jsonify({'error': 'Your account registration was rejected.'}), 403
        if user_status == 'inactive':
            return jsonify({'error': 'Your account has been deactivated. Please contact your administrator.'}), 403

        token = jwt.encode({
            'user_id':    user['id'],
            'company_id': user.get('company_id'),
            'exp':        datetime.utcnow() + timedelta(days=7)
        }, app.config['SECRET_KEY'], algorithm='HS256')

        try:
            log_activity(
                action_type = ACTION_LOGIN,
                user_id     = user['id'],
                user_name   = user.get('name', ''),
                user_role   = user.get('role', 'user'),
                company_id  = user.get('company_id'),
                description = f"{user.get('name', 'User')} signed in via Microsoft",
                source      = request.headers.get('X-App-Source', SOURCE_WEB),
                level       = LEVEL_MANAGER if user.get('role') in ('admin', 'manager') else LEVEL_USER,
                metadata    = {'email': ms_email, 'auth_type': 'microsoft'},
                ip_address  = request.remote_addr,
            )
        except Exception:
            pass

        return jsonify({
            'token': token,
            'user': {
                'id':                  user['id'],
                'email':               user['email'],
                'name':                user['name'],
                'role':                user.get('role', 'user'),
                'company_id':          user.get('company_id'),
                'must_change_password': _requires_first_login_password_change(user),
            }
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': f'Microsoft sign-in failed: {str(e)}'}), 500


@app.route('/api/auth/google', methods=['POST'])
def google_login():
    """
    Google OAuth SSO endpoint.
    User must already be registered and approved. No auto-provisioning.
    """
    import urllib.request as _urllib
    try:
        data         = request.get_json() or {}
        access_token = data.get('access_token', '').strip()
        if not access_token:
            return jsonify({'error': 'access_token is required'}), 400

        req = _urllib.Request(
            'https://www.googleapis.com/oauth2/v3/userinfo',
            headers={'Authorization': f'Bearer {access_token}'},
        )
        try:
            with _urllib.urlopen(req, timeout=10) as resp:
                g_user = json.loads(resp.read().decode())
        except Exception as g_err:
            print(f"[GOOGLE-LOGIN] Userinfo call failed: {g_err}")
            return jsonify({'error': 'Could not verify Google identity. Please try again.'}), 401

        g_email  = (g_user.get('email') or '').lower()
        verified = g_user.get('email_verified', False)

        if not g_email:
            return jsonify({'error': 'Could not retrieve email from Google account.'}), 400
        if not verified:
            return jsonify({'error': 'Google account email is not verified.'}), 403

        from db_postgres import get_user_by_email
        user = get_user_by_email(g_email)

        if not user:
            return jsonify({
                'error': 'No account found for this Google email. Please register first.'
            }), 404

        user_status = user.get('status', 'approved')
        if user_status == 'pending':
            return jsonify({'error': 'Your account is pending approval by your company admin'}), 403
        if user_status == 'rejected':
            return jsonify({'error': 'Your account registration was rejected.'}), 403
        if user_status == 'inactive':
            return jsonify({'error': 'Your account has been deactivated. Please contact your administrator.'}), 403

        token = jwt.encode({
            'user_id':    user['id'],
            'company_id': user.get('company_id'),
            'exp':        datetime.utcnow() + timedelta(days=7)
        }, app.config['SECRET_KEY'], algorithm='HS256')

        try:
            log_activity(
                action_type = ACTION_LOGIN,
                user_id     = user['id'],
                user_name   = user.get('name', ''),
                user_role   = user.get('role', 'user'),
                company_id  = user.get('company_id'),
                description = f"{user.get('name', 'User')} signed in via Google",
                source      = request.headers.get('X-App-Source', SOURCE_WEB),
                level       = LEVEL_MANAGER if user.get('role') in ('admin', 'manager') else LEVEL_USER,
                metadata    = {'email': g_email, 'auth_type': 'google'},
                ip_address  = request.remote_addr,
            )
        except Exception:
            pass

        return jsonify({
            'token': token,
            'user': {
                'id':                  user['id'],
                'email':               user['email'],
                'name':                user['name'],
                'role':                user.get('role', 'user'),
                'company_id':          user.get('company_id'),
                'must_change_password': _requires_first_login_password_change(user),
            }
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': f'Google sign-in failed: {str(e)}'}), 500


@app.route('/api/auth/me', methods=['GET'])
@token_required
def get_current_user(current_user):
    cid = current_user.get('company_id')
    features = get_company_features(cid) if cid else {}
    company_name = None
    if cid:
        try:
            from db_postgres import get_company_by_id as _gcbi
            _c = _gcbi(cid)
            company_name = _c.get('name') if _c else None
        except Exception:
            pass
    return jsonify({
        'id':                  current_user['id'],
        'email':               current_user['email'],
        'name':                current_user['name'],
        'role':                current_user.get('role', 'user'),
        'company_id':          cid,
        'company_name':        company_name,
        'status':              current_user.get('status', 'approved'),
        'created_at':          current_user.get('created_at', ''),
        'must_change_password': _requires_first_login_password_change(current_user),
        'company_features':    features,
    })


@app.route('/api/auth/me', methods=['PUT'])
@token_required
def update_current_user(current_user):
    data = request.get_json() or {}
    new_name = (data.get('name') or '').strip()
    if not new_name:
        return jsonify({'error': 'Name is required'}), 400

    pg_update_user_name(current_user['id'], new_name)

    return jsonify({
        'id': current_user['id'],
        'email': current_user['email'],
        'name': new_name,
        'role': current_user.get('role', 'user'),
        'created_at': current_user.get('created_at', ''),
    })

# ==================== SUBSCRIPTION ENDPOINTS ====================

@app.route('/api/subscription', methods=['GET'])
@token_required
def get_subscription(current_user):
    subscription = pg_get_subscription(current_user['id'])

    if not subscription:
        return jsonify({'error': 'Subscription not found'}), 404

    # Reset daily limit if new day
    today = datetime.now().date().isoformat()
    if subscription.get('last_upload_date') != today:
        is_locked = subscription['plan'] == 'free' and subscription.get('total_uploads', 0) >= 3
        pg_update_subscription_daily_reset(current_user['id'], is_locked, today)
        subscription['uploads_today'] = 0
        subscription['last_upload_date'] = today
        subscription['is_locked'] = is_locked

    return jsonify(subscription)

@app.route('/api/subscription/upgrade', methods=['POST'])
@token_required
def upgrade_subscription(current_user):
    data = request.get_json()
    plan = data.get('plan')
    
    if plan not in ['pro', 'enterprise']:
        return jsonify({'error': 'Invalid plan'}), 400
    
    upgraded = pg_update_subscription_plan(current_user['id'], plan)
    if not upgraded:
        return jsonify({'error': 'Subscription not found'}), 404

    subscription = pg_get_subscription(current_user['id'])
    return jsonify({'message': 'Subscription upgraded successfully', 'subscription': subscription})

# ==================== ALGORITHM ENDPOINTS ====================

@app.route('/api/algorithms', methods=['GET'])
@token_required
def list_algorithms(current_user):
    """
    Get list of available processing algorithms.
    Automatically scans the algorithm folder for new processors.
    """
    try:
        from algorithm_scanner import get_available_algorithms
        algorithms = get_available_algorithms()
        
        return jsonify({
            'status': 'success',
            'algorithms': algorithms,
            'count': len(algorithms)
        })
    except Exception as e:
        print(f"[ERROR] Failed to list algorithms: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': 'Failed to load algorithms',
            'details': str(e)
        }), 500

@app.route('/api/algorithms/refresh', methods=['POST'])
@token_required
def refresh_algorithms(current_user):
    """
    Force refresh of algorithm list.
    Useful after adding new algorithm files to the algorithm folder.
    """
    try:
        from algorithm_scanner import refresh_algorithms as refresh_algo
        algorithms = refresh_algo()
        
        return jsonify({
            'status': 'success',
            'message': 'Algorithm list refreshed successfully',
            'algorithms': algorithms,
            'count': len(algorithms)
        })
    except Exception as e:
        print(f"[ERROR] Failed to refresh algorithms: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': 'Failed to refresh algorithms',
            'details': str(e)
        }), 500

# ==================== COMPANY & MULTI-TENANCY ====================

@app.route('/api/public/companies', methods=['GET'])
def list_companies_public():
    """Public endpoint — no auth required. Returns id + name for the signup dropdown."""
    from db_postgres import get_companies
    companies = get_companies()
    return jsonify([{'id': c['id'], 'name': c['name']} for c in companies])


@app.route('/api/super-admin/companies', methods=['GET'])
@token_required
def list_companies_admin(current_user):
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    from db_postgres import get_companies, get_all_users_with_company
    companies = get_companies()
    all_users = get_all_users_with_company()
    for company in companies:
        cid = company['id']
        company['user_count'] = sum(1 for u in all_users if u.get('company_id') == cid)
        company['pending_count'] = sum(
            1 for u in all_users if u.get('company_id') == cid and u.get('status') == 'pending'
        )
    return jsonify(companies)


@app.route('/api/super-admin/companies', methods=['POST'])
@token_required
def create_company_route(current_user):
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    slug = (data.get('slug') or '').strip().lower().replace(' ', '-')
    if not name or not slug:
        return jsonify({'error': 'name and slug are required'}), 400
    admin_email = (data.get('admin_email') or '').strip().lower()
    admin_name  = (data.get('admin_name')  or '').strip()

    from db_postgres import create_company
    try:
        company = create_company(name, slug)
    except Exception as e:
        if 'unique' in str(e).lower():
            return jsonify({'error': f'A company with slug "{slug}" already exists'}), 409
        return jsonify({'error': str(e)}), 500

    result = dict(company)

    if admin_email:
        if not _re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', admin_email):
            result['admin_warning'] = 'Invalid admin email — company created without auto-admin.'
        else:
            from db_postgres import get_user_by_email, create_user_direct, create_subscription_direct
            if get_user_by_email(admin_email):
                result['admin_warning'] = f'{admin_email} already exists. Company created without auto-admin.'
            else:
                temp_pw  = secrets.token_urlsafe(10)
                pw_hash  = generate_password_hash(temp_pw)
                display  = admin_name or admin_email.split('@')[0]
                try:
                    new_user = create_user_direct(
                        str(uuid.uuid4()), admin_email, display,
                        'company_admin', pw_hash, company['id'], status='approved'
                    )
                    try:
                        create_subscription_direct(new_user['id'])
                    except Exception:
                        pass
                    result['admin_created']       = True
                    result['admin_temp_password'] = temp_pw
                    result['admin_email']         = admin_email
                    result['admin_name']          = display
                except Exception as e2:
                    result['admin_warning'] = f'Company created but failed to create admin: {e2}'

    return jsonify(result), 201


@app.route('/api/super-admin/users', methods=['GET'])
@token_required
def list_all_users_admin(current_user):
    """Super admin: view all users across all companies."""
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    from db_postgres import get_all_users_with_company
    users = get_all_users_with_company()
    safe = [{k: v for k, v in u.items() if k != 'password'} for u in users]
    return jsonify(safe)


@app.route('/api/company/users', methods=['GET'])
@token_required
def list_company_users(current_user):
    """Company admin: list users in their company, optionally filter by status."""
    role = current_user.get('role', 'user')
    if role not in ('company_admin', 'admin', 'super_admin'):
        return jsonify({'error': 'Company admin access required'}), 403

    # Only super_admin is cross-company; admin is company-scoped
    is_global = role == 'super_admin'
    if is_global:
        company_id = request.args.get('company_id') or current_user.get('company_id')
    else:
        company_id = current_user.get('company_id')

    status_filter = request.args.get('status')  # pending | approved | rejected | None=all
    from db_postgres import get_users_by_company, get_all_users_with_company

    if is_global and not company_id:
        # Return all users across all companies
        all_users = get_all_users_with_company()
        if status_filter:
            all_users = [u for u in all_users if u.get('status') == status_filter]
        safe = [{k: v for k, v in u.items() if k != 'password'} for u in all_users]
        return jsonify(safe)

    users = get_users_by_company(company_id, status_filter)
    safe = [{k: v for k, v in u.items() if k != 'password'} for u in users]
    return jsonify(safe)


@app.route('/api/company/users/<user_id>/approve', methods=['POST'])
@token_required
def approve_user(current_user, user_id):
    role = current_user.get('role', 'user')
    if role not in ('company_admin', 'admin', 'super_admin'):
        return jsonify({'error': 'Company admin access required'}), 403

    from db_postgres import get_user_by_id, set_user_status
    target = get_user_by_id(user_id)
    if not target:
        return jsonify({'error': 'User not found'}), 404

    is_global = role == 'super_admin'
    if not is_global and target.get('company_id') != current_user.get('company_id'):
        return jsonify({'error': 'User is not in your company'}), 403

    set_user_status(user_id, 'approved', approved_by_id=current_user['id'])

    try:
        create_notification(
            user_id=target['id'],
            title='Account approved',
            message='Your account has been approved. You can now log in.',
            notification_type='account_approved',
            metadata={},
        )
    except Exception:
        pass

    try:
        log_activity(
            action_type='user_approved',
            user_id=current_user['id'],
            user_name=current_user.get('name', ''),
            user_role=role,
            company_id=current_user.get('company_id'),
            description=f"{current_user.get('name')} approved user {target.get('name')} ({target.get('email')})",
            source=request.headers.get('X-App-Source', SOURCE_WEB),
            level=LEVEL_ADMIN,
            metadata={'approved_user_id': user_id},
            ip_address=request.remote_addr,
        )
    except Exception:
        pass

    return jsonify({'message': 'User approved successfully'})


@app.route('/api/company/users/<user_id>/reject', methods=['POST'])
@token_required
def reject_user(current_user, user_id):
    role = current_user.get('role', 'user')
    if role not in ('company_admin', 'admin', 'super_admin'):
        return jsonify({'error': 'Company admin access required'}), 403

    from db_postgres import get_user_by_id, set_user_status
    target = get_user_by_id(user_id)
    if not target:
        return jsonify({'error': 'User not found'}), 404

    is_global = role == 'super_admin'
    if not is_global and target.get('company_id') != current_user.get('company_id'):
        return jsonify({'error': 'User is not in your company'}), 403

    set_user_status(user_id, 'rejected', approved_by_id=current_user['id'])

    try:
        log_activity(
            action_type='user_rejected',
            user_id=current_user['id'],
            user_name=current_user.get('name', ''),
            user_role=role,
            company_id=current_user.get('company_id'),
            description=f"{current_user.get('name')} rejected user {target.get('name')} ({target.get('email')})",
            source=request.headers.get('X-App-Source', SOURCE_WEB),
            level=LEVEL_ADMIN,
            metadata={'rejected_user_id': user_id},
            ip_address=request.remote_addr,
        )
    except Exception:
        pass

    return jsonify({'message': 'User rejected'})


@app.route('/api/company/users/<user_id>/role', methods=['PUT'])
@token_required
def update_user_role(current_user, user_id):
    """Company admin can promote/demote users within their company."""
    role = current_user.get('role', 'user')
    if role not in ('company_admin', 'admin', 'super_admin'):
        return jsonify({'error': 'Company admin access required'}), 403

    data = request.get_json() or {}
    new_role = (data.get('role') or '').strip()
    allowed_roles = ('user', 'manager', 'company_admin') if role != 'super_admin' else (
        'user', 'manager', 'company_admin', 'admin', 'super_admin'
    )
    if new_role not in allowed_roles:
        return jsonify({'error': f'Invalid role. Allowed: {allowed_roles}'}), 400

    from db_postgres import get_user_by_id
    target = get_user_by_id(user_id)
    if not target:
        return jsonify({'error': 'User not found'}), 404
    is_global = role == 'super_admin'
    if not is_global and target.get('company_id') != current_user.get('company_id'):
        return jsonify({'error': 'User is not in your company'}), 403

    pg_update_user_role(user_id, new_role)
    return jsonify({'message': f'Role updated to {new_role}'})


@app.route('/api/super-admin/companies/<company_id>', methods=['PUT'])
@token_required
def update_company_route(current_user, company_id):
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    data = request.get_json() or {}
    name = (data.get('name') or '').strip() or None
    slug = ((data.get('slug') or '').strip().lower().replace(' ', '-')) or None
    from db_postgres import update_company
    try:
        updated = update_company(company_id, name=name, slug=slug)
        if not updated:
            return jsonify({'error': 'Company not found'}), 404
        return jsonify(updated)
    except Exception as e:
        if 'unique' in str(e).lower():
            return jsonify({'error': 'Slug already taken by another company'}), 409
        return jsonify({'error': str(e)}), 500


@app.route('/api/super-admin/companies/<company_id>', methods=['DELETE'])
@token_required
def delete_company_route(current_user, company_id):
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    from db_postgres import delete_company
    try:
        ok = delete_company(company_id)
        if not ok:
            return jsonify({'error': 'Company not found'}), 404
        return jsonify({'message': 'Company deleted'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/super-admin/companies/<company_id>/suspend', methods=['POST'])
@token_required
def suspend_company_route(current_user, company_id):
    """Super admin: suspend or reactivate a company."""
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    data = request.get_json() or {}
    suspended = bool(data.get('suspended', True))
    from db_postgres import suspend_company
    ok = suspend_company(company_id, suspended)
    if not ok:
        return jsonify({'error': 'Company not found'}), 404
    action = 'suspended' if suspended else 'reactivated'
    return jsonify({'message': f'Company {action}'})


@app.route('/api/super-admin/companies/<company_id>/features', methods=['PUT'])
@token_required
def set_company_features_route(current_user, company_id):
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    data = request.get_json() or {}
    features = data.get('features')
    if not isinstance(features, dict):
        return jsonify({'error': 'features must be an object'}), 400
    from db_postgres import set_company_features
    try:
        updated = set_company_features(company_id, features)
        if not updated:
            return jsonify({'error': 'Company not found'}), 404
        return jsonify(updated)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Report Builder — Super Admin (catalog CRUD) ───────────────────────────────

@app.route('/api/super-admin/report-catalog', methods=['GET'])
@token_required
def list_report_catalog(current_user):
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    return jsonify(get_report_catalog())


@app.route('/api/super-admin/report-catalog', methods=['POST'])
@token_required
def add_catalog_item_route(current_user):
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    d = request.get_json() or {}
    label = (d.get('label') or '').strip()
    if not label:
        return jsonify({'error': 'label is required'}), 400
    item = create_catalog_item(
        label=label,
        description=(d.get('description') or '').strip(),
        type_=d.get('type', 'kpi_card'),
        data_key=(d.get('data_key') or '').strip(),
        unit=(d.get('unit') or '').strip(),
        is_suggested=bool(d.get('is_suggested', False)),
        sort_order=int(d.get('sort_order', 99)),
    )
    return jsonify(item), 201


@app.route('/api/super-admin/report-catalog/<item_id>', methods=['PUT'])
@token_required
def update_catalog_item_route(current_user, item_id):
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    d = request.get_json() or {}
    fields = {}
    for k in ('label', 'description', 'unit', 'data_key', 'is_suggested', 'sort_order'):
        if k in d:
            fields[k] = d[k]
    if 'type' in d:
        fields['type'] = d['type']
    updated = update_catalog_item(item_id, **fields)
    if not updated:
        return jsonify({'error': 'Item not found'}), 404
    return jsonify(updated)


@app.route('/api/super-admin/report-catalog/<item_id>', methods=['DELETE'])
@token_required
def delete_catalog_item_route(current_user, item_id):
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    ok = delete_catalog_item(item_id)
    if not ok:
        return jsonify({'error': 'Item not found'}), 404
    return jsonify({'deleted': item_id})


# ── Report Builder — Company (config get/save + catalog read) ─────────────────

@app.route('/api/report-builder/catalog', methods=['GET'])
@token_required
def company_get_catalog(current_user):
    return jsonify(get_report_catalog())


@app.route('/api/report-builder/config', methods=['GET'])
@token_required
def get_report_config(current_user):
    cid = current_user.get('company_id')
    if not cid:
        return jsonify([])
    return jsonify(get_report_builder_config(cid))


@app.route('/api/report-builder/config', methods=['PUT'])
@token_required
def save_report_config(current_user):
    if current_user.get('role') not in ('company_admin', 'admin', 'super_admin'):
        return jsonify({'error': 'Company admin access required'}), 403
    cid = current_user.get('company_id')
    if not cid:
        return jsonify({'error': 'No company assigned'}), 400
    layout = request.get_json()
    if not isinstance(layout, list):
        return jsonify({'error': 'layout must be an array'}), 400
    ok = save_report_builder_config(cid, layout)
    return jsonify({'ok': ok})


@app.route('/api/super-admin/platform-stats', methods=['GET'])
@token_required
def get_platform_stats(current_user):
    """Super admin: per-company last activity timestamp and event counts."""
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    from activity_log_db import get_company_last_activities
    company_activity = get_company_last_activities()
    return jsonify({'company_activity': company_activity})


@app.route('/api/super-admin/audit-log', methods=['GET'])
@token_required
def get_super_admin_audit_log(current_user):
    """Super admin: full audit log across all companies with optional filters."""
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    limit = min(int(request.args.get('limit', 200)), 1000)
    filter_keys = ('user_id', 'company_id', 'action_type', 'source', 'level', 'date_from', 'date_to')
    filters = {k: request.args.get(k) for k in filter_keys if request.args.get(k)}
    activities = get_all_activities_admin(limit=limit, filters=filters)
    return jsonify({'activities': activities, 'count': len(activities)})


@app.route('/api/super/users', methods=['POST'])
@token_required
def super_create_user(current_user):
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip().lower()
    password = (data.get('password') or '').strip()
    role = data.get('role', 'user')
    company_id = data.get('company_id') or None
    status = data.get('status', 'approved')
    if not name or not email or not password:
        return jsonify({'error': 'name, email and password are required'}), 400
    allowed_roles = ('user', 'manager', 'company_admin', 'admin', 'super_admin')
    if role not in allowed_roles:
        return jsonify({'error': 'Invalid role'}), 400
    from db_postgres import get_user_by_email, create_user_direct, create_subscription_direct
    if get_user_by_email(email):
        return jsonify({'error': 'A user with that email already exists'}), 409
    pw_hash = generate_password_hash(password)
    new_uid = str(uuid.uuid4())
    try:
        user = create_user_direct(
            user_id=new_uid, email=email, name=name, role=role,
            password_hash=pw_hash, company_id=company_id, status=status,
        )
        create_subscription_direct(user['id'])
        try:
            log_activity(
                action_type='user_created',
                user_id=current_user['id'],
                user_name=current_user.get('name', ''),
                user_role=current_user.get('role'),
                company_id=current_user.get('company_id'),
                description=f"Super admin created user {name} ({email}) with role {role}",
                source=request.headers.get('X-App-Source', SOURCE_WEB),
                level=LEVEL_ADMIN,
                metadata={'created_user_id': user['id'], 'role': role},
                ip_address=request.remote_addr,
            )
        except Exception:
            pass
        return jsonify({k: v for k, v in user.items() if k != 'password'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/super/users/<user_id>', methods=['PUT'])
@token_required
def super_update_user(current_user, user_id):
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    data = request.get_json() or {}
    allowed_roles = ('user', 'manager', 'company_admin', 'admin', 'super_admin')
    fields = {}
    if 'role' in data and data['role'] in allowed_roles:
        fields['role'] = data['role']
    if 'status' in data and data['status'] in ('pending', 'approved', 'rejected', 'inactive'):
        fields['status'] = data['status']
        if data['status'] == 'approved':
            fields['approved_by'] = current_user.get('id')
    if 'company_id' in data:
        fields['company_id'] = data['company_id'] or None
    if 'name' in data and data['name']:
        fields['name'] = data['name'].strip()
    if 'email' in data and data['email']:
        fields['email'] = data['email'].strip().lower()
    if not fields:
        return jsonify({'error': 'No valid fields to update'}), 400
    from db_postgres import update_user_full, get_user_by_id
    try:
        target = get_user_by_id(user_id)
        ok = update_user_full(user_id, **fields)
        if not ok:
            return jsonify({'error': 'User not found'}), 404
        try:
            changes = ', '.join(f"{k}={v}" for k, v in fields.items() if k != 'approved_by')
            log_activity(
                action_type='user_updated',
                user_id=current_user['id'],
                user_name=current_user.get('name', ''),
                user_role=current_user.get('role'),
                company_id=current_user.get('company_id'),
                description=f"Super admin updated {target.get('name', user_id) if target else user_id}: {changes}",
                source=request.headers.get('X-App-Source', SOURCE_WEB),
                level=LEVEL_ADMIN,
                metadata={'target_user_id': user_id, 'changes': fields},
                ip_address=request.remote_addr,
            )
        except Exception:
            pass
        return jsonify({'message': 'User updated'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/super/users/<user_id>', methods=['DELETE'])
@token_required
def super_delete_user(current_user, user_id):
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    if user_id == current_user.get('id'):
        return jsonify({'error': 'Cannot delete your own account'}), 400
    from db_postgres import delete_user, get_user_by_id, count_approved_super_admins
    target = get_user_by_id(user_id)
    if not target:
        return jsonify({'error': 'User not found'}), 404
    # Prevent deleting the last super_admin
    if target.get('role') == 'super_admin' and count_approved_super_admins() <= 1:
        return jsonify({'error': 'Cannot delete the last super admin account'}), 400
    try:
        ok = delete_user(user_id)
        if not ok:
            return jsonify({'error': 'User not found'}), 404
        try:
            log_activity(
                action_type='user_deleted',
                user_id=current_user['id'],
                user_name=current_user.get('name', ''),
                user_role=current_user.get('role'),
                company_id=current_user.get('company_id'),
                description=f"Super admin deleted user {target.get('name')} ({target.get('email')})",
                source=request.headers.get('X-App-Source', SOURCE_WEB),
                level=LEVEL_ADMIN,
                metadata={'deleted_user_id': user_id, 'deleted_user_email': target.get('email')},
                ip_address=request.remote_addr,
            )
        except Exception:
            pass
        return jsonify({'message': 'User deleted'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/super-admin/users/<user_id>/send-reset', methods=['POST'])
@token_required
def super_send_reset_link(current_user, user_id):
    if current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Super admin access required'}), 403
    from db_postgres import get_user_by_id
    user = get_user_by_id(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404
    email = (user.get('email') or '').strip()
    if not email:
        return jsonify({'error': 'User has no email address'}), 400
    raw_token = secrets.token_urlsafe(32)
    token_hash = hashlib.sha256(raw_token.encode('utf-8')).hexdigest()
    expires_at = (datetime.utcnow() + timedelta(minutes=30)).isoformat()
    pg_create_reset_token(token_hash, user['id'], expires_at)
    frontend_origin = os.getenv('FRONTEND_URL', 'http://localhost:3000').rstrip('/')
    reset_link = _build_password_reset_link(raw_token, frontend_origin=frontend_origin)
    try:
        _send_password_reset_email(email, user.get('name', 'User'), reset_link)
    except Exception as e:
        return jsonify({'error': f'Failed to send email: {str(e)}'}), 500
    return jsonify({'message': f'Reset link sent to {email}'}), 200



# ==================== FILE UPLOAD & PROCESSING ====================

def _run_processing_job(job_id, input_path, job_output_folder, filename, user_id, user_name, user_company_id, subscription_user_id):
    """Background thread: runs matching trackers, updates JOB_STATUS, saves history & deviations."""
    try:
        print(f"[BG:{job_id[:8]}] Starting processing for: {filename}")

        def progress_callback(idx, total_count, tracker_name, status, result):
            with JOB_STATUS_LOCK:
                entry = JOB_STATUS.get(job_id, {})
                entry['current_idx'] = idx
                entry['current_tracker'] = tracker_name
                entry['total'] = total_count
                entry['percent'] = int(idx / total_count * 100) if total_count else 0
                # Keep a running list of completed trackers for the frontend
                completed = entry.get('completed_trackers', [])
                if status in ('success', 'error') and result:
                    completed.append(result)
                    entry['completed_trackers'] = completed
                JOB_STATUS[job_id] = entry

        processing_result = process_file(input_path, job_output_folder, progress_callback=progress_callback)
        print(f"[BG:{job_id[:8]}] Processing done — {processing_result['success_count']} OK")

        # Compare uploaded sheet with baseline and merge changes into base file.
        base_merge_summary = {
            'status': 'skipped',
            'reason': 'Processing did not produce successful outputs',
        }
        if processing_result.get('success_count', 0) > 0:
            # Step 1: Detect changes vs latest chain / base file
            detected_changes = []
            try:
                detected_changes = _detect_upload_changes(input_path, context={
                    'job_id': job_id, 'user_id': user_id, 'user_name': user_name,
                }, company_id=user_company_id)
            except Exception as _dc_err:
                print(f"[BG:{job_id[:8]}] Change detection error: {_dc_err}")

            # Step 2: Auto-approve — apply changes immediately
            approval_id = str(uuid.uuid4())
            chain_result = _apply_update_to_chain(
                upload_file_path=input_path,
                upload_filename=filename,
                approved_by='auto',
                job_id=job_id,
                company_id=user_company_id,
            )

            # Audit trail — record as already approved
            pending_approvals = read_db(PENDING_UPLOAD_APPROVALS_DB) if os.path.exists(PENDING_UPLOAD_APPROVALS_DB) else []
            pending_approvals.append({
                'approval_id': approval_id,
                'job_id': job_id,
                'upload_filename': filename,
                'upload_path': input_path,
                'user_id': user_id,
                'user_name': user_name,
                'company_id': user_company_id,
                'submitted_at': datetime.now().isoformat(),
                'status': 'approved',
                'approved_by': 'auto',
                'approved_at': datetime.now().isoformat(),
                'detected_changes': detected_changes,
                'change_summary': {
                    'total': len(detected_changes),
                    'date_changes': sum(1 for c in detected_changes if c['change_type'] == 'date_change'),
                    'empty_to_filled': sum(1 for c in detected_changes if c['change_type'] == 'empty_to_filled'),
                    'filled_to_empty': sum(1 for c in detected_changes if c['change_type'] == 'filled_to_empty'),
                    'value_changes': sum(1 for c in detected_changes if c['change_type'] == 'value_change'),
                },
                'chain_result': chain_result,
            })
            write_db(PENDING_UPLOAD_APPROVALS_DB, pending_approvals)

            base_merge_summary = {
                'status': 'approved',
                'approval_id': approval_id,
                'change_count': len(detected_changes),
                'reason': 'Auto-approved and applied to update chain.',
                **chain_result,
            }
            print(
                f"[BG:{job_id[:8]}] Base merge: {base_merge_summary.get('status')} | "
                f"matched={base_merge_summary.get('matched_count', 0)} "
                f"updated={base_merge_summary.get('updated_count', 0)} "
                f"appended={base_merge_summary.get('appended_count', 0)}"
            )

        # Update subscription — atomic single-row increment (no race condition)
        pg_increment_subscription_uploads(subscription_user_id)

        # Save history entry (single-row upsert — no race condition with concurrent uploads)
        history_entry = {
            'id': job_id,
            'user_id': user_id,
            'company_id': user_company_id,
            'filename': filename,
            'processed_at': datetime.now().isoformat(),
            'status': 'completed',
            'total_sheets': processing_result['total_sheets'],
            'success_count': processing_result['success_count'],
            'error_count': processing_result['error_count'],
            'failed_sheets': processing_result.get('failed_sheets', []),
            'detected_sheets': processing_result['detected_sheets'],
            'results': processing_result['results'],
            'base_merge': base_merge_summary,
        }
        pg_upsert_history_entry(history_entry)

        # Auto-save monthly report extracted from the tracker output
        if processing_result.get('success_count', 0) > 0:
            try:
                _auto_save_monthly_report(
                    job_id=job_id,
                    results=processing_result['results'],
                    filename=filename,
                    processed_year=datetime.now().year,
                    company_id=user_company_id,
                )
            except Exception as _rpt_err:
                print(f'[BG:{job_id[:8]}] Monthly report auto-save error: {_rpt_err}')

            # Sync milestones / CP / one-pager JSON from the uploaded schedule XLSX
            try:
                if input_path.lower().endswith(('.xlsx', '.xls')):
                    _sync_schedule_json_from_xlsx(input_path, company_id=user_company_id)
                    print(f'[BG:{job_id[:8]}] Schedule JSON sync complete')
            except Exception as _sync_err:
                print(f'[BG:{job_id[:8]}] Schedule JSON sync error: {_sync_err}')

        # Notifications
        sheet_names = ', '.join(
            r['sheet_name'] for r in processing_result['results'] if r.get('status') == 'success'
        )
        create_notification(
            user_id=user_id,
            title='File Processed Successfully',
            message=(
                f"'{filename}' processed. {processing_result['success_count']} tracker(s) ready: {sheet_names}. "
                f"Base updated: {base_merge_summary.get('updated_count', 0)} updated, "
                f"{base_merge_summary.get('appended_count', 0)} new activity(ies)."
            ),
            notification_type='success',
            metadata={'job_id': job_id, 'filename': filename,
                      'success_count': processing_result['success_count'],
                      'error_count': processing_result['error_count'],
                      'base_merge': base_merge_summary}
        )
        notify_admins_and_managers(
            title='New File Upload',
            message=f"{user_name} uploaded '{filename}' – {processing_result['success_count']} tracker(s) processed",
            notification_type='info',
            metadata={'job_id': job_id, 'filename': filename, 'user_id': user_id,
                      'user_name': user_name,
                      'success_count': processing_result['success_count'],
                      'error_count': processing_result['error_count']},
            roles=['admin'],
            company_id=user_company_id,
        )
        # Track file processed activity
        try:
            log_activity(
                action_type=ACTION_FILE_PROCESSED,
                user_id=user_id,
                user_name=user_name,
                company_id=user_company_id,
                entity_type='job', entity_id=job_id,
                description=f"{user_name} processed '{filename}': {processing_result['success_count']} tracker(s)",
                source=SOURCE_WEB,
                level=LEVEL_MANAGER,
                metadata={'job_id': job_id, 'filename': filename,
                          'success_count': processing_result['success_count'],
                          'error_count': processing_result['error_count'],
                          'sheets': sheet_names},
            )
        except Exception:
            pass

        # Deviations - process and save in batches to reduce memory
        deviations_detected = []
        high_severity_count = 0
        try:
            from deviation_calculator import calculate_deviations
            import gc

            successful_results = [r for r in processing_result['results'] if r['status'] == 'success']
            print(f"[DEV:{job_id[:8]}] {len(successful_results)} successful sheet(s) to scan for deviations")
            for r in successful_results:
                exists = os.path.exists(r.get('output_path', ''))
                print(f"[DEV:{job_id[:8]}]   sheet={r.get('sheet_name')} processor={r.get('processor')} output_exists={exists} path={r.get('output_path')}")

            if successful_results:
                # Clear only PENDING deviations for this file — reviewed ones are preserved
                cleared, locked = delete_deviations_by_filename(filename, user_company_id)
                if cleared or locked:
                    print(f"[DEV:{job_id[:8]}] Re-upload '{filename}': cleared {cleared} pending, kept {locked} answered")

                # Process deviations in batches of 5 sheets to reduce memory usage
                BATCH_SIZE = 5
                for batch_start in range(0, len(successful_results), BATCH_SIZE):
                    batch_end = min(batch_start + BATCH_SIZE, len(successful_results))
                    batch = successful_results[batch_start:batch_end]

                    print(f"[DEV:{job_id[:8]}] Deviation batch {batch_start+1}-{batch_end}")

                    batch_deviations = calculate_deviations(
                        output_folder=job_output_folder,
                        job_id=job_id,
                        filename=filename,
                        processing_results=batch,
                        user_id=user_id,
                        company_id=user_company_id,
                        input_path=input_path,
                    )

                    print(f"[DEV:{job_id[:8]}] batch returned {len(batch_deviations)} deviation(s)")

                    # ── Filter to critical-path and milestone activities only ──
                    cp_ids, cp_names = _load_cp_and_ms_ids(user_company_id)
                    if cp_ids or cp_names:
                        before = len(batch_deviations)
                        batch_deviations = [
                            d for d in batch_deviations
                            if _is_cp_deviation(d, cp_ids, cp_names)
                        ]
                        print(f"[DEV:{job_id[:8]}] CP+MS filter: {before} → {len(batch_deviations)} deviation(s)")

                    # Immediately save to DB and clear from memory
                    # Uses upsert: UPDATE existing pending row by activity_id to
                    # preserve the row id (avoids burning the SERIAL sequence).
                    if batch_deviations:
                        new_high = 0
                        inserted = 0
                        skipped_locked = 0
                        for deviation in batch_deviations:
                            # Skip deviations a manager has already reviewed on this file
                            if is_deviation_locked(
                                deviation.get('sheet', ''),
                                deviation.get('description', ''),
                                deviation.get('company_id', ''),
                                filename,
                            ):
                                skipped_locked += 1
                                continue
                            new_id = upsert_cp_deviation(deviation)
                            deviation['id'] = new_id
                            deviations_detected.append(deviation)
                            inserted += 1
                            if deviation.get('severity') == 'High':
                                high_severity_count += 1
                                new_high += 1
                        print(f"[DEV:{job_id[:8]}] upserted={inserted} skipped_locked={skipped_locked}")
                        
                        # Notify for high-severity in this batch
                        if new_high > 0:
                            notify_admins_and_managers(
                                title='High-Severity Deviations Detected',
                                message=f"{new_high} high-severity deviation(s) in '{filename}' require your review.",
                                notification_type='warning',
                                metadata={'job_id': job_id, 'filename': filename,
                                          'high_severity_count': new_high},
                                roles=['manager'],
                                company_id=user_company_id,
                            )
                    
                    # Clear batch from memory
                    del batch_deviations
                    gc.collect()
                    
        except Exception as dev_err:
            import traceback
            traceback.print_exc()
            print(f"[BG:{job_id[:8]}] Deviation error: {dev_err}")

        # Auto-email data inconsistency details after completion (best-effort, non-blocking)
        inconsistency_email = {'sent': False, 'status': 'not_sent'}
        try:
            sent, status = _send_processing_inconsistency_email(
                job_id=job_id,
                filename=filename,
                user_name=user_name,
                processing_result=processing_result,
                base_merge_summary=base_merge_summary,
                deviations_detected=deviations_detected,
            )
            inconsistency_email = {'sent': bool(sent), 'status': str(status)}
            print(f"[BG:{job_id[:8]}] Inconsistency email status: {inconsistency_email}")
        except Exception as mail_err:
            inconsistency_email = {'sent': False, 'status': str(mail_err)}
            print(f"[BG:{job_id[:8]}] Inconsistency email error: {mail_err}")

        # Prune oldest output folders and raw uploads (keep most recent 30
        # jobs) rather than deleting each raw upload immediately after its
        # own processing run — Theta Sheets' "Browse Theta Sheets" picker
        # sources from these same raw uploads, so they need to survive past
        # the job that created them.
        _prune_old_outputs(keep=30)
        _prune_old_uploads(keep=30)

        # ── Push notification: deviations detected ────────────────────────────
        if deviations_detected:
            expiry_days = int(os.getenv('DEVIATION_EXPIRY_DAYS', '20'))
            pending_count = len([d for d in deviations_detected if d.get('severity') == 'High'])
            push_title = f'{len(deviations_detected)} Deviations Detected'
            push_body  = (
                f'{high_severity_count} high-severity deviations in \'{filename}\'. '
                f'Please review — each expires in {expiry_days} days.'
            )
            try:
                # Notify the uploader
                _send_push_to_user(user_id, push_title, push_body, '/deviations')
                # Notify admins & managers
                _send_push_to_all_users_with_role(['admin', 'manager'], push_title, push_body, '/deviations')
            except Exception as _push_err:
                print(f'[BG:{job_id[:8]}] Push notification error: {_push_err}')

        # Mark job as completed in status store
        with JOB_STATUS_LOCK:
            JOB_STATUS[job_id].update({
                'status': 'completed',
                'percent': 100,
                'current_tracker': None,
                'processing_result': processing_result,
                'deviations': {
                    'count': len(deviations_detected),
                    'high_severity': high_severity_count,
                    'medium_severity': sum(1 for d in deviations_detected if d.get('severity') == 'Medium'),
                    'low_severity': sum(1 for d in deviations_detected if d.get('severity') == 'Low'),
                },
                'message': processing_result['message'],
                'download_url': f'/api/download/{job_id}',
                'has_warnings': len(processing_result.get('failed_sheets', [])) > 0,
                'base_merge': base_merge_summary,
                'inconsistency_email': inconsistency_email,
            })
        print(f"[BG:{job_id[:8]}] Job marked completed")
        
        # Final memory cleanup
        import gc
        del processing_result
        del deviations_detected
        gc.collect()

    except Exception as e:
        import traceback
        traceback.print_exc()
        with JOB_STATUS_LOCK:
            entry = JOB_STATUS.get(job_id, {})
            entry['status'] = 'error'
            entry['error'] = str(e)
            JOB_STATUS[job_id] = entry
        print(f"[BG:{job_id[:8]}] Fatal error: {e}")


def _prune_old_outputs(keep: int = 30):
    """Remove oldest output job folders, keeping only the most recent `keep` jobs.
    Also removes orphaned job folders that have no matching history entry."""
    try:
        folders = [
            d for d in os.scandir(os.path.join(_APP_ROOT, OUTPUT_FOLDER))
            if d.is_dir()
        ]
        if len(folders) <= keep:
            return
        # Sort oldest-first by modification time
        folders.sort(key=lambda d: d.stat().st_mtime)
        to_delete = folders[:len(folders) - keep]
        deleted = 0
        for folder in to_delete:
            # Only prune if either it's too old OR not referenced in history
            try:
                shutil.rmtree(folder.path, ignore_errors=True)
                deleted += 1
            except Exception:
                pass
        if deleted:
            print(f"[PRUNE] Removed {deleted} old output folder(s), kept newest {keep}")
    except Exception as prune_err:
        print(f"[PRUNE] Error during output pruning: {prune_err}")


def _prune_old_uploads(keep: int = 30):
    """Remove the oldest raw uploaded files, keeping only the most recent
    `keep`. Raw uploads used to be deleted immediately after their own job
    finished processing; now they're kept (bounded by this prune) so Theta
    Sheets' "Browse Theta Sheets" picker can source from recently uploaded
    files, matching the retention pattern already used for output folders."""
    try:
        files = [
            d for d in os.scandir(os.path.join(_APP_ROOT, UPLOAD_FOLDER))
            if d.is_file()
        ]
        if len(files) <= keep:
            return
        files.sort(key=lambda d: d.stat().st_mtime)
        to_delete = files[:len(files) - keep]
        deleted = 0
        for f in to_delete:
            try:
                os.remove(f.path)
                deleted += 1
            except Exception:
                pass
        if deleted:
            print(f"[PRUNE] Removed {deleted} old raw upload(s), kept newest {keep}")
    except Exception as prune_err:
        print(f"[PRUNE] Error during upload pruning: {prune_err}")


@app.route('/api/upload', methods=['POST'])
@token_required
def upload_file(current_user):
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls') or file.filename.endswith('.xlsm')):
        return jsonify({'error': 'Invalid file type. Only Excel files allowed'}), 400

    # Load subscription (no upload limits enforced)
    subscription = pg_get_subscription(current_user['id'])

    try:
        job_id = str(uuid.uuid4())

        # Save uploaded file immediately
        input_path = os.path.join(UPLOAD_FOLDER, f"{job_id}_{file.filename}")
        file.save(input_path)

        job_output_folder = os.path.join(_APP_ROOT, OUTPUT_FOLDER, job_id)
        os.makedirs(job_output_folder, exist_ok=True)

        # Write a placeholder history entry so status endpoint can verify ownership
        pg_upsert_history_entry({
            'id': job_id,
            'user_id': current_user['id'],
            'company_id': current_user.get('company_id'),
            'filename': file.filename,
            'processed_at': datetime.now().isoformat(),
            'status': 'processing',
            'results': [],
        })

        # Initialise status store
        with JOB_STATUS_LOCK:
            JOB_STATUS[job_id] = {
                'status': 'processing',
                'job_id': job_id,
                'filename': file.filename,
                'current_idx': 0,
                'current_tracker': 'Detecting sheets…',
                'total': 0,
                'percent': 0,
                'completed_trackers': [],
            }

        # Launch background thread – return to client immediately
        t = threading.Thread(
            target=_run_processing_job,
            args=(
                job_id, input_path, job_output_folder,
                file.filename,
                current_user['id'],
                current_user.get('name', 'User'),
                current_user.get('company_id'),
                current_user['id'],
            ),
            daemon=True,
        )
        t.start()

        print(f"[UPLOAD] Job {job_id[:8]} started in background for: {file.filename}")

        return jsonify({
            'status': 'processing',
            'job_id': job_id,
            'message': f'Processing started for {file.filename}. Poll /api/status/{job_id} for progress.',
            'status_url': f'/api/status/{job_id}',
        }), 202

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Upload error: {str(e)}'}), 500


@app.route('/api/status/<job_id>', methods=['GET'])
@token_required
def get_job_status(current_user, job_id):
    """Poll processing status for a job. Returns live progress while running, full result when done."""
    entry = pg_get_history_entry(job_id, current_user['id'])
    if not entry:
        return jsonify({'error': 'Job not found'}), 404

    with JOB_STATUS_LOCK:
        job = JOB_STATUS.get(job_id)

    if not job:
        # Job may have been completed before server restart; check history
        if entry.get('status') == 'completed':
            return jsonify({
                'status': 'completed',
                'job_id': job_id,
                'processing_result': {
                    'results': entry.get('results', []),
                    'success_count': entry.get('success_count', 0),
                    'error_count': entry.get('error_count', 0),
                    'failed_sheets': entry.get('failed_sheets', []),
                },
                'base_merge': entry.get('base_merge', {}),
                'message': entry.get('message') or entry.get('status') or 'completed',
            })
        return jsonify({'status': 'unknown', 'job_id': job_id})

    return jsonify(job)

@app.route('/api/download/<job_id>', methods=['GET'])
@token_required
def download_file(current_user, job_id):
    """Download all output files for a job as a zip or individual file."""
    try:
        # Get sheet_name from query params (optional)
        sheet_name = request.args.get('sheet')

        entry = pg_get_history_entry(job_id, current_user['id'])
        if not entry:
            return jsonify({'error': 'File not found'}), 404

        job_output_folder = os.path.join(_APP_ROOT, OUTPUT_FOLDER, job_id)

        if not os.path.exists(job_output_folder):
            return jsonify({'error': 'Output files not found on server'}), 404

        # If specific sheet requested, download that file only
        if sheet_name:
            result = next((r for r in entry['results'] if r['sheet_name'] == sheet_name), None)
            if not result or result['status'] != 'success':
                return jsonify({'error': f'File for sheet "{sheet_name}" not found'}), 404
            
            file_path = os.path.join(job_output_folder, result['output_filename'])
            if not os.path.exists(file_path):
                return jsonify({'error': 'File not found on server'}), 404
            
            return send_file(file_path, as_attachment=True, download_name=result['output_filename'])
        
        # If multiple files, create a zip
        import zipfile
        from io import BytesIO
        
        memory_file = BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            for result in entry['results']:
                if result['status'] == 'success':
                    file_path = os.path.join(job_output_folder, result['output_filename'])
                    if os.path.exists(file_path):
                        zf.write(file_path, result['output_filename'])
        
        memory_file.seek(0)
        
        # Generate zip filename
        zip_filename = f"PMO_Outputs_{job_id[:8]}.zip"
        
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=zip_filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _raw_upload_path(entry):
    """Resolve a history entry's original (pre-processing) uploaded file
    path on disk, or None if it's missing/unresolvable. Raw uploads are
    saved as "{job_id}_{original_filename}" (see the /api/upload handler's
    file.save(...) call) — not just the plain filename, which would collide
    across jobs with the same name."""
    uploads_root = os.path.realpath(os.path.join(_APP_ROOT, UPLOAD_FOLDER))
    file_path = os.path.realpath(os.path.join(uploads_root, f"{entry['id']}_{entry['filename']}"))
    if not file_path.startswith(uploads_root + os.sep):
        return None
    if not os.path.exists(file_path):
        return None
    return file_path


@app.route('/api/history/<job_id>/raw', methods=['GET'])
@token_required
def download_raw_upload(current_user, job_id):
    """Download the original (pre-processing) uploaded file for a job —
    used by the Theta Sheets "Browse Theta Sheets" picker to source an
    already-uploaded workbook from server storage instead of a fresh
    local-device file dialog."""
    try:
        entry = pg_get_history_entry(job_id, current_user['id'])
        if not entry:
            return jsonify({'error': 'File not found'}), 404

        file_path = _raw_upload_path(entry)
        if not file_path:
            return jsonify({'error': 'Original upload is no longer available on the server'}), 404

        return send_file(file_path, as_attachment=True, download_name=entry['filename'])
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# ==================== HISTORY ENDPOINTS ====================

@app.route('/api/history', methods=['GET'])
@token_required
def get_history(current_user):
    try:
        limit = int(request.args.get('limit', 100))
    except Exception:
        limit = 100
    limit = min(max(limit, 1), 500)
    user_history = pg_read_history_for_company(user_id=current_user['id'], limit=limit)
    return jsonify(user_history)

@app.route('/api/history/<job_id>', methods=['GET'])
@token_required
def get_history_item(current_user, job_id):
    entry = pg_get_history_entry(job_id, current_user['id'])
    if not entry:
        return jsonify({'error': 'History item not found'}), 404
    return jsonify(entry)

@app.route('/api/history/<job_id>', methods=['DELETE'])
@token_required
def delete_history_item(current_user, job_id):
    """Delete a history entry and its associated output files."""
    try:
        entry = pg_get_history_entry(job_id, current_user['id'])
        if not entry:
            return jsonify({'error': 'History item not found'}), 404

        # Remove from DB — single-row delete, no race condition
        pg_delete_history_entry(job_id)

        # Clean up output files
        import shutil
        job_folder = os.path.join(_APP_ROOT, OUTPUT_FOLDER, job_id)
        if os.path.exists(job_folder):
            shutil.rmtree(job_folder, ignore_errors=True)

        print(f"[HISTORY] Deleted job {job_id} for user {current_user['id']}")
        return jsonify({'success': True, 'message': 'Session deleted'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to delete: {str(e)}'}), 500

# ==================== PREVIEW ENDPOINT ====================

@app.route('/api/preview/<job_id>', methods=['GET'])
@token_required
def preview_output(current_user, job_id):
    """Preview the output Excel file data."""
    try:
        sheet_name = request.args.get('sheet')
        try:
            max_rows = int(request.args.get('max_rows', 600))
        except Exception:
            max_rows = 600
        max_rows = min(max(max_rows, 50), 5000)

        entry = pg_get_history_entry(job_id, current_user['id'])
        if not entry:
            return jsonify({'error': 'File not found'}), 404

        job_output_folder = os.path.join(_APP_ROOT, OUTPUT_FOLDER, job_id)

        if not os.path.exists(job_output_folder):
            return jsonify({'error': 'Output files not found on server'}), 404

        # Find the result for the requested sheet
        if not sheet_name:
            # If no sheet specified, use the first successful result
            result = next((r for r in entry['results'] if r['status'] == 'success'), None)
        else:
            result = next((r for r in entry['results'] if r['sheet_name'] == sheet_name), None)
        
        if not result or result['status'] != 'success':
            return jsonify({'error': 'No successful output found'}), 404
        
        file_path = os.path.join(job_output_folder, result['output_filename'])
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found on server'}), 404
        
        # Read Excel file and convert to JSON
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active  # Get the first/active sheet
        
        # Extract headers (first row)
        headers = []
        for cell in ws[1]:
            headers.append(cell.value if cell.value is not None else '')
        
        # Extract capped data rows for fast preview rendering
        rows = []
        total_rows = max(0, ws.max_row - 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(rows) >= max_rows:
                break
            row_data = []
            for cell_value in row:
                # Convert datetime objects to strings
                if isinstance(cell_value, datetime):
                    row_data.append(cell_value.strftime('%Y-%m-%d %H:%M:%S'))
                else:
                    row_data.append(cell_value)
            rows.append(row_data)
        
        wb.close()
        
        return jsonify({
            'sheet_name': result['sheet_name'],
            'description': result['description'],
            'headers': headers,
            'rows': rows,
            'total_rows': total_rows,
            'row_count': len(rows),
            'has_more_rows': total_rows > len(rows),
            'is_preview': False,
            'output_filename': result['output_filename']
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to preview file: {str(e)}'}), 500


# ==================== ANALYTICS ENDPOINT ====================

@app.route('/api/analytics/<job_id>', methods=['GET'])
@token_required
def get_analytics(current_user, job_id):
    """Get all sheets data for S-curve analytics, overlaid with ML Forecast predictions."""
    try:
        entry = pg_get_history_entry(job_id, current_user['id'])
        if not entry:
            return jsonify({'error': 'File not found'}), 404

        job_output_folder = os.path.join(_APP_ROOT, OUTPUT_FOLDER, job_id)
        if not os.path.exists(job_output_folder):
            return jsonify({'error': 'Output files not found on server'}), 404

        import openpyxl

        include_formula = str(request.args.get('include_formula', '0')).lower() in ('1', 'true', 'yes')
        sheets_data = []
        for result in entry.get('results', []):
            if result.get('status') != 'success':
                continue
            file_path = os.path.join(job_output_folder, result['output_filename'])
            if not os.path.exists(file_path):
                continue
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)

                for ws_title in wb.sheetnames:
                    try:
                        ws = wb[ws_title]

                        # Auto-detect header row: first row with >= 3 non-empty cells
                        header_row_num = 1
                        for r_idx in range(1, min(6, ws.max_row + 1)):
                            filled = sum(1 for c in ws[r_idx] if c.value is not None and str(c.value).strip())
                            if filled >= 3:
                                header_row_num = r_idx
                                break

                        # Extract headers
                        headers = [str(cell.value) if cell.value is not None else '' for cell in ws[header_row_num]]

                        # Skip sheets with no meaningful headers
                        non_empty_headers = [h for h in headers if h.strip()]
                        if len(non_empty_headers) < 2:
                            continue

                        # Extract ALL data rows after header row (skip entirely empty rows)
                        rows = []
                        for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
                            if all(v is None for v in row):
                                continue
                            row_data = []
                            for cell_value in row:
                                if isinstance(cell_value, datetime):
                                    row_data.append(cell_value.strftime('%Y-%m-%d'))
                                else:
                                    row_data.append(cell_value)
                            rows.append(row_data)

                        if not rows:
                            continue

                        # Build display name: append sheet tab name when workbook has multiple sheets
                        if len(wb.sheetnames) > 1:
                            display_name = f"{result['sheet_name']} ({ws_title})"
                        else:
                            display_name = result['sheet_name']

                        sheets_data.append({
                            'sheet_name': display_name,
                            'description': result.get('description', ''),
                            'headers': headers,
                            'rows': rows,
                            'row_count': len(rows),
                        })
                    except Exception as ws_err:
                        print(f"[ANALYTICS] Error reading tab '{ws_title}' in {result['output_filename']}: {ws_err}")
                        continue

                wb.close()
            except Exception as sheet_err:
                print(f"[ANALYTICS] Error reading {result['output_filename']}: {sheet_err}")
                continue

        # Optional formula overlay (heavy). Disabled by default for faster analytics load.
        if include_formula:
            try:
                import pandas as pd
                from pathlib import Path
                import traceback as _tb

                project_root = Path(_APP_ROOT)
                print(f"[ANALYTICS] Formula overlay enabled for {len(sheets_data)} sheets...")

                for si, sheet_data in enumerate(sheets_data):
                    try:
                        headers = sheet_data['headers']
                        rows = sheet_data['rows']

                        if not rows:
                            continue

                        deduped_headers = []
                        seen_cols = {}
                        for h in headers:
                            h_str = str(h) if h else ''
                            if h_str in seen_cols:
                                seen_cols[h_str] += 1
                                deduped_headers.append(f"{h_str}_{seen_cols[h_str]}")
                            else:
                                seen_cols[h_str] = 0
                                deduped_headers.append(h_str)

                        df_eval = pd.DataFrame(rows, columns=deduped_headers)

                        try:
                            from dependency_formula_forecast import get_formula_forecast_for_dataframe

                            df_scored = get_formula_forecast_for_dataframe(df_eval, project_root)
                            formula_headers = [
                                "Formula Forecast Start Date",
                                "Formula Forecast Finish Date",
                                "Formula Forecast Date",
                                "Formula Constraint Source",
                                "Formula Predecessors",
                                "Formula Successors",
                                "Formula Lag Impact (days)",
                                "Formula Delay vs Plan (days)",
                                "Formula Engine Status",
                            ]
                            sheet_data['headers'] = headers + formula_headers

                            def _fmt_date(val):
                                if pd.isna(val) if hasattr(pd, 'isna') else (val is None or val != val):
                                    return ''
                                if isinstance(val, datetime):
                                    return val.strftime('%Y-%m-%d')
                                s = str(val)
                                return s[:10] if len(s) >= 10 else s

                            for i in range(len(rows)):
                                f_start = df_scored["formula_forecast_start_date"].iloc[i] if "formula_forecast_start_date" in df_scored.columns else None
                                f_finish = df_scored["formula_forecast_finish_date"].iloc[i] if "formula_forecast_finish_date" in df_scored.columns else None
                                f_fcst = df_scored["formula_forecast_date"].iloc[i] if "formula_forecast_date" in df_scored.columns else f_finish
                                f_source = df_scored["formula_constraint_source"].iloc[i] if "formula_constraint_source" in df_scored.columns else ""
                                f_preds = df_scored["formula_predecessors"].iloc[i] if "formula_predecessors" in df_scored.columns else ""
                                f_succs = df_scored["formula_successors"].iloc[i] if "formula_successors" in df_scored.columns else ""
                                f_lag = df_scored["formula_lag_impact_days"].iloc[i] if "formula_lag_impact_days" in df_scored.columns else 0
                                f_delay = df_scored["formula_delay_vs_plan_days"].iloc[i] if "formula_delay_vs_plan_days" in df_scored.columns else None
                                f_status = df_scored["formula_engine_status"].iloc[i] if "formula_engine_status" in df_scored.columns else ""

                                lag_text = '' if (f_lag is None or (hasattr(pd, 'isna') and pd.isna(f_lag))) else str(int(f_lag))
                                delay_text = '' if (f_delay is None or (hasattr(pd, 'isna') and pd.isna(f_delay))) else str(int(f_delay))

                                rows[i] = list(rows[i]) + [
                                    _fmt_date(f_start),
                                    _fmt_date(f_finish),
                                    _fmt_date(f_fcst),
                                    str(f_source or ''),
                                    str(f_preds or ''),
                                    str(f_succs or ''),
                                    lag_text,
                                    delay_text,
                                    str(f_status or ''),
                                ]

                            diag = df_scored.attrs.get('formula_diagnostics', {}) if hasattr(df_scored, 'attrs') else {}
                            print(f"[ANALYTICS] FORMULA OK sheet {si}: {sheet_data.get('sheet_name','?')} | mapped_edges={diag.get('mapped_edges_in_sheet',0)} nodes={diag.get('activity_nodes_in_sheet',0)}")

                        except Exception as formula_err:
                            print(f"[ANALYTICS] Formula overlay skipped for sheet {si}: {formula_err}")

                    except Exception as sheet_err:
                        print(f"[ANALYTICS] Sheet {si} overlay error: {sheet_err}")
                        continue

            except Exception as overlay_err:
                print(f"[ANALYTICS] Warning: Overlay failed -> {overlay_err}")
                _tb.print_exc()

        return jsonify({
            'job_id': job_id,
            'filename': entry.get('filename', ''),
            'processed_at': entry.get('processed_at', ''),
            'total_sheets': len(sheets_data),
            'sheets': sheets_data,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to load analytics: {str(e)}'}), 500


# ==================== NOTIFICATIONS ====================

@app.route('/api/notifications', methods=['GET'])
@token_required
def get_notifications(current_user):
    """Get notifications for the current user with 7-day deviation reminders."""
    try:
        # ---- 7-DAY REMINDER CHECK (direct SQL, no read-all) ----
        try:
            pending_devs = get_pending_deviations()
            now = datetime.now()
            users = read_db(USERS_DB)

            for d in pending_devs:
                if str(d.get('review_status', '')).strip().lower() != 'pending':
                    continue
                detected_at_str = d.get('detected_at')
                if not detected_at_str:
                    continue
                try:
                    detected_at = datetime.fromisoformat(detected_at_str.replace("Z", ""))
                    days_pending = (now - detected_at).days
                    if days_pending < 7:
                        continue

                    last_notified = d.get('last_reminder_notified_at')
                    should_notify = (
                        not last_notified or
                        (now - datetime.fromisoformat(last_notified.replace("Z", ""))).days >= 7
                    )
                    if should_notify and pg_exists_unread_deviation_reminder(d['id']):
                        should_notify = False

                    if should_notify:
                        update_reminder_timestamp(d['id'], now.isoformat())
                        filename = d.get('filename', 'Unknown File')
                        message = (
                            f"Reminder: Deviation #{d['id']} from file '{filename}' has been "
                            f"pending for {days_pending} days."
                        )
                        dev_company_id = d.get('company_id')
                        managers = [
                            u for u in users
                            if u.get('role') == 'manager'
                            and u.get('company_id') == dev_company_id
                        ]
                        for u in managers:
                            pg_create_notification(
                                u['id'], 'Action Required: Pending Activity Delay',
                                message, 'warning',
                                {'deviation_id': d['id'], 'filename': filename},
                            )
                except Exception as e:
                    print(f"[REMINDER ERROR] deviation {d.get('id')}: {e}")
        except Exception as outer_e:
            print(f"[REMINDER ERROR] {outer_e}")

        # Purge read notifications older than 30 days
        pg_purge_old_read(30)

        notifications = pg_get_user_notifications(current_user['id'])
        unread_count = sum(1 for n in notifications if not n.get('read'))
        return jsonify({'notifications': notifications, 'unread_count': unread_count})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/notifications/<notification_id>/read', methods=['PUT'])
@token_required
def mark_notification_read(current_user, notification_id):
    """Mark a single notification as read."""
    try:
        if pg_mark_notification_read(notification_id, current_user['id']):
            return jsonify({'message': 'Notification marked as read'})
        return jsonify({'error': 'Notification not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/notifications/read-all', methods=['PUT'])
@token_required
def mark_all_notifications_read(current_user):
    """Mark all notifications as read for the current user."""
    try:
        pg_mark_all_read(current_user['id'])
        return jsonify({'message': 'All notifications marked as read'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/notifications/clear-read', methods=['DELETE'])
@token_required
def clear_read_notifications(current_user):
    """Delete all read notifications for the current user."""
    try:
        deleted = pg_delete_read_notifications(current_user['id'])
        return jsonify({'deleted': deleted})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/notifications/clear-all', methods=['DELETE'])
@token_required
def clear_all_notifications(current_user):
    """Delete all notifications for the current user (admin only)."""
    if current_user.get('role') != 'admin':
        return jsonify({'error': 'Admin access required'}), 403
    try:
        deleted = pg_delete_all_notifications(current_user['id'])
        return jsonify({'deleted': deleted})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/notifications/<notification_id>', methods=['DELETE'])
@token_required
def delete_notification(current_user, notification_id):
    """Delete a single notification."""
    try:
        if pg_delete_notification(notification_id, current_user['id']):
            return jsonify({'message': 'Notification deleted'})
        return jsonify({'error': 'Notification not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== UPLOAD APPROVAL ENDPOINTS ====================

@app.route('/api/upload-approvals', methods=['GET'])
@token_required
def get_upload_approvals(current_user):
    """List upload approvals. Admin/manager scoped to own company; super_admin sees all."""
    role = current_user.get('role')
    if role not in ('admin', 'manager', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Unauthorized'}), 403
    approvals = read_db(PENDING_UPLOAD_APPROVALS_DB) if os.path.exists(PENDING_UPLOAD_APPROVALS_DB) else []
    if role != 'super_admin':
        company_id = current_user.get('company_id')
        approvals = [a for a in approvals if a.get('company_id') == company_id]
    status_filter = request.args.get('status', 'pending')
    if status_filter != 'all':
        approvals = [a for a in approvals if a.get('status') == status_filter]
    approvals.sort(key=lambda x: x.get('submitted_at', ''), reverse=True)
    return jsonify({'approvals': approvals, 'total': len(approvals)})


@app.route('/api/upload-approvals/<approval_id>', methods=['GET'])
@token_required
def get_upload_approval(current_user, approval_id):
    """Get one pending upload approval with full change list."""
    approvals = read_db(PENDING_UPLOAD_APPROVALS_DB) if os.path.exists(PENDING_UPLOAD_APPROVALS_DB) else []
    entry = next((a for a in approvals if a.get('approval_id') == approval_id), None)
    if not entry:
        return jsonify({'error': 'Approval not found'}), 404
    role = current_user.get('role')
    if role != 'super_admin' and entry.get('company_id') != current_user.get('company_id'):
        return jsonify({'error': 'Approval not found'}), 404
    return jsonify(entry)


@app.route('/api/upload-approvals/<approval_id>/approve', methods=['POST'])
@token_required
def approve_upload(current_user, approval_id):
    """Admin approves an upload — applies changes to update chain."""
    role = current_user.get('role')
    if role not in ('admin', 'manager', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Unauthorized'}), 403

    approvals = read_db(PENDING_UPLOAD_APPROVALS_DB) if os.path.exists(PENDING_UPLOAD_APPROVALS_DB) else []
    entry = next((a for a in approvals if a.get('approval_id') == approval_id), None)
    if not entry:
        return jsonify({'error': 'Approval not found'}), 404
    if role != 'super_admin' and entry.get('company_id') != current_user.get('company_id'):
        return jsonify({'error': 'Approval not found'}), 404
    if entry.get('status') != 'pending':
        return jsonify({'error': f"Already {entry.get('status')}"}), 400

    upload_path = entry.get('upload_path', '')
    if not os.path.exists(upload_path):
        return jsonify({'error': 'Original upload file no longer available on disk'}), 404

    # Apply to update chain
    result = _apply_update_to_chain(
        upload_file_path=upload_path,
        upload_filename=entry['upload_filename'],
        approved_by=current_user.get('name', current_user['id']),
        job_id=entry['job_id'],
        company_id=current_user.get('company_id'),
    )

    # Mark approval as done
    entry['status'] = 'approved'
    entry['approved_by'] = current_user.get('name', current_user['id'])
    entry['approved_at'] = datetime.now().isoformat()
    entry['chain_result'] = result
    write_db(PENDING_UPLOAD_APPROVALS_DB, approvals)

    # Notify the uploader
    create_notification(
        user_id=entry['user_id'],
        title='Upload Approved',
        message=(
            f"Your upload '{entry['upload_filename']}' was approved by {current_user.get('name','Admin')}. "
            f"Changes applied as {result.get('filename', 'update file')}."
        ),
        notification_type='success',
        metadata={'approval_id': approval_id, 'chain_result': result},
    )

    return jsonify({'status': 'approved', 'chain_result': result})


@app.route('/api/upload-approvals/<approval_id>/reject', methods=['POST'])
@token_required
def reject_upload(current_user, approval_id):
    """Admin rejects an upload — changes are NOT applied."""
    role = current_user.get('role')
    if role not in ('admin', 'manager', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.json or {}
    reason = str(data.get('reason', '')).strip() or 'No reason provided'

    approvals = read_db(PENDING_UPLOAD_APPROVALS_DB) if os.path.exists(PENDING_UPLOAD_APPROVALS_DB) else []
    entry = next((a for a in approvals if a.get('approval_id') == approval_id), None)
    if not entry:
        return jsonify({'error': 'Approval not found'}), 404
    if role != 'super_admin' and entry.get('company_id') != current_user.get('company_id'):
        return jsonify({'error': 'Approval not found'}), 404
    if entry.get('status') != 'pending':
        return jsonify({'error': f"Already {entry.get('status')}"}), 400

    entry['status'] = 'rejected'
    entry['rejected_by'] = current_user.get('name', current_user['id'])
    entry['rejected_at'] = datetime.now().isoformat()
    entry['rejection_reason'] = reason
    write_db(PENDING_UPLOAD_APPROVALS_DB, approvals)

    create_notification(
        user_id=entry['user_id'],
        title='Upload Rejected',
        message=(
            f"Your upload '{entry['upload_filename']}' was rejected. Reason: {reason}"
        ),
        notification_type='error',
        metadata={'approval_id': approval_id, 'reason': reason},
    )

    return jsonify({'status': 'rejected', 'reason': reason})


@app.route('/api/update-chain', methods=['GET'])
@token_required
def get_update_chain(current_user):
    """Return the full update chain history."""
    role = current_user.get('role')
    cid = None if role == 'super_admin' else current_user.get('company_id')
    chain = pg_read_update_chain(cid)
    return jsonify({'chain': chain, 'total': len(chain)})


@app.route('/api/update-chain/<int:index>/download', methods=['GET'])
@token_required
def download_update_chain_file(current_user, index):
    """Download a specific update chain file by index."""
    role = current_user.get('role')
    cid = None if role == 'super_admin' else current_user.get('company_id')
    chain = pg_read_update_chain(cid)
    entry = next((e for e in chain if e.get('index') == index), None)
    if not entry:
        return jsonify({'error': 'Update file not found'}), 404
    file_path = os.path.join(UPDATE_CHAIN_DIR, entry['filename'])
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found on disk'}), 404
    return send_file(file_path, as_attachment=True, download_name=entry['filename'])
# ==================== HEALTH CHECK ====================

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({
        'status': 'ok',
        'message': 'PMO Backend is running',
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/stats', methods=['GET'])
@token_required
def get_stats(current_user):
    user_history = pg_read_history_for_company(user_id=current_user['id'])

    subscription = pg_get_subscription(current_user['id'])

    return jsonify({
        'total_processed': len(user_history),
        'uploads_today': subscription['uploads_today'] if subscription else 0,
        'plan': subscription['plan'] if subscription else 'free'
    })


@app.route('/api/whatif/critical-dashboard', methods=['GET'])
@token_required
def get_whatif_critical_dashboard(current_user):
    """Return What-IF critical dashboard data from Critical_path_analysis-EPC_Borouge.json with DB cache."""
    try:
        refresh = str(request.args.get('refresh', '')).strip().lower() in ('1', 'true', 'yes', 'y')
        cached_only = str(request.args.get('cached_only', '')).strip().lower() in ('1', 'true', 'yes', 'y')

        knowledge = _load_critical_path_json_context(
            max_rows=max(80, int(os.getenv('CLAUDE_CRITICAL_JSON_MAX_ROWS', '220'))),
            max_chars=max(4000, int(os.getenv('CLAUDE_CRITICAL_JSON_MAX_CHARS', '18000'))),
            company_id=current_user.get('company_id'),
        )
        if not knowledge.get('ok'):
            return jsonify({'error': knowledge.get('error') or 'Critical knowledge JSON could not be loaded'}), 404

        cache_key = (
            f"critical_dashboard_json|{knowledge.get('signature', '')}|"
            f"{os.getenv('CLAUDE_CRITICAL_JSON_MAX_ROWS', '220')}|"
            f"{os.getenv('CLAUDE_CRITICAL_JSON_MAX_CHARS', '18000')}|"
            f"{os.getenv('CLAUDE_CRITICAL_DASHBOARD_MAX_TOKENS', '2400')}|v1"
        )

        critical_cache = pg_read_whatif_critical(current_user.get('company_id'))
        if not isinstance(critical_cache, dict):
            critical_cache = {}

        if not refresh:
            cached = critical_cache.get(cache_key)
            if isinstance(cached, dict):
                return jsonify({**cached, 'cached': True, 'cache_key': cache_key}), 200

            # Fallback: serve latest known critical cache entry (legacy key or previous signature)
            alias_cached = critical_cache.get('whatif_critical_activities|overall')
            if isinstance(alias_cached, dict):
                return jsonify({
                    **alias_cached,
                    'cached': True,
                    'cache_key': 'whatif_critical_activities|overall',
                    'fallback_cache': True,
                }), 200

            if cached_only:
                return jsonify({
                    'error': 'No cached critical dashboard data found. Run a non-cached request once to populate cache.',
                    'cached': False,
                    'cache_key': cache_key,
                }), 404

        summary = knowledge.get('summary', {}) if isinstance(knowledge.get('summary'), dict) else {}
        seeded_tracker = knowledge.get('threat_tracker_seed', []) if isinstance(knowledge.get('threat_tracker_seed'), list) else []

        prompt = _build_critical_dashboard_prompt(knowledge.get('context', ''))
        max_tokens = max(1000, int(os.getenv('CLAUDE_CRITICAL_DASHBOARD_MAX_TOKENS', '2400')))

        normalized = None
        json_repaired = False
        try:
            raw = _generate_claude_response(prompt, max_tokens=max_tokens)
            parsed = _extract_json_object(raw)

            if not parsed:
                repair_prompt = f"""Convert the following model output into one valid JSON object only.
No markdown and no commentary.

Required keys:
- summary_markdown
- ui_payload
- chart_plan
- endpoint_map

Model output:
{raw}
"""
                repaired_raw = _generate_claude_response(repair_prompt, max_tokens=max(600, min(1400, max_tokens // 2)))
                repaired = _extract_json_object(repaired_raw)
                if repaired:
                    parsed = repaired
                    json_repaired = True

            if parsed:
                normalized = _normalize_whatif_payload(parsed)
        except Exception:
            normalized = None

        if not normalized:
            max_delay = _to_int(summary.get('max_delay_days', 0), 0)
            delayed = _to_int(summary.get('delayed_rows', 0), 0)
            total = _to_int(summary.get('activity_rows', 0), 0)
            avg_delay = float(summary.get('average_delay_days', 0) or 0)

            opportunity_days = max(0, int(round(max_delay * 0.7)))
            residual_days = max(0, max_delay - opportunity_days)
            recovery_percent = int(round((opportunity_days / max_delay) * 100)) if max_delay > 0 else 0

            issue_dist = summary.get('issue_distribution', {}) if isinstance(summary.get('issue_distribution'), dict) else {}
            top_tracker = seeded_tracker[:8]
            if not top_tracker:
                top_tracker = [{
                    'id': 'A60780',
                    'name': 'Receipt of Approved vendor data for FG & OG compressor',
                    'baseline': 0,
                    'actual': max_delay,
                    'late': max_delay,
                }]

            normalized = _normalize_whatif_payload({
                'summary_markdown': (
                    f"### Critical Activities What-IF Snapshot\n\n"
                    f"- Data Source: **{knowledge.get('file', 'Critical_path_analysis-EPC_Borouge.json')}**\n"
                    f"- Delayed Activities: **{delayed}** of **{total}**\n"
                    f"- Max Delay: **{max_delay} days** | Avg Delay: **{avg_delay} days**\n"
                    f"- Estimated Recovery Opportunity: **{opportunity_days} days**"
                ),
                'ui_payload': {
                    'save_days': max_delay,
                    'opportunity_days': opportunity_days,
                    'residual_days': residual_days,
                    'recovery_percent': recovery_percent,
                    'core_threat_title': 'The Core Threat: Critical Path Delay Concentration',
                    'core_threat_summary': 'Critical-path milestones show concentrated finish variance that can push end-date certainty unless overlap actions are executed.',
                    'step1_title': 'Recovery Step 1: Sequence Compression',
                    'step1_summary': 'Prioritize overlap across engineering, procurement, and field readiness packages to recover schedule without compromising milestone gates.',
                    'step1_baseline_days': [40, 47, 46],
                    'step1_compressed_days': [18, 18, 17],
                    'step1_activity_chips': [t.get('id', '') for t in top_tracker[:3] if t.get('id')],
                    'step1_recovered_days': opportunity_days,
                    'step1_residual_to_sequence_days': residual_days,
                    'kpis': {
                        'total_activities': total,
                        'delayed_activities': delayed,
                        'max_delay_days': max_delay,
                        'average_delay_days': avg_delay,
                    },
                    'threat_tracker': top_tracker,
                    'compressor_savings': [
                        {'step': 'Engineering-Procurement overlap', 'days': f'+{max(0, int(opportunity_days * 0.45))}'},
                        {'step': 'Civil-Mechanical handoff compression', 'days': f'+{max(0, int(opportunity_days * 0.35))}'},
                        {'step': 'Commissioning prep pull-forward', 'days': f'+{max(0, opportunity_days - int(opportunity_days * 0.45) - int(opportunity_days * 0.35))}'},
                    ],
                    'recommendations': [
                        'Freeze top delayed critical activities into a daily recovery huddle owned by PM + Engineering.',
                        'Convert selected FS dependencies to controlled SS with explicit quality gate checkpoints.',
                        'Track delay burn-down weekly with activity-level owner accountability and milestone gating.',
                    ],
                },
                'chart_plan': [
                    {
                        'id': 'issue_distribution',
                        'type': 'pie',
                        'title': 'Critical Delay Issue Mix',
                        'data': [
                            {'name': k, 'value': _to_int(v, 0)} for k, v in issue_dist.items()
                        ],
                    },
                    {
                        'id': 'top_delays',
                        'type': 'bar',
                        'title': 'Top Delayed Activities',
                        'data': [
                            {'name': t.get('id', 'N/A'), 'value': _to_int(t.get('late', 0), 0)} for t in top_tracker
                        ],
                    },
                ],
                'endpoint_map': [
                    {
                        'name': 'Critical Dashboard',
                        'method': 'GET',
                        'path': '/api/whatif/critical-dashboard',
                        'purpose': 'Threat tracker and KPI baseline from critical path JSON',
                    }
                ],
            })

        ui_payload = normalized.get('ui_payload', {}) if isinstance(normalized.get('ui_payload', {}), dict) else {}
        response_payload = {
            'generated_at': datetime.now().isoformat(),
            'source': {
                'mode': 'critical_path_json',
                'file': knowledge.get('file', 'Critical_path_analysis-EPC_Borouge.json'),
                'path': knowledge.get('path', ''),
            },
            'summary_markdown': normalized.get('summary_markdown', ''),
            'ui_payload': ui_payload,
            'threat_tracker': ui_payload.get('threat_tracker', []),
            'compressor_savings': ui_payload.get('compressor_savings', []),
            'kpis': ui_payload.get('kpis', {}),
            'chart_plan': normalized.get('chart_plan', []),
            'endpoint_map': normalized.get('endpoint_map', []),
            'json_repaired': bool(json_repaired),
        }

        critical_cache[cache_key] = response_payload
        # Keep a stable alias for frontend consumers that read scenario/category keys.
        critical_cache['whatif_critical_activities|overall'] = response_payload

        if len(critical_cache) > 40:
            try:
                items = list(critical_cache.items())
                items.sort(key=lambda x: ((x[1] or {}).get('generated_at') or ''), reverse=True)
                critical_cache = dict(items[:40])
            except Exception:
                pass

        pg_write_whatif_critical(critical_cache, current_user.get('company_id'))
        return jsonify({**response_payload, 'cached': False, 'cache_key': cache_key}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to load critical dashboard data: {str(e)}'}), 500







def _build_project_update_frontend_prompt(
    workbook_name: str,
    sheet_name: str,
    summary: dict,
    top_updates: list,
    client_prompt: str = '',
    upload_context: dict = None,
):
    upload_context = upload_context if isinstance(upload_context, dict) else {}
    upload_file = str(upload_context.get('filename', '')).strip() or 'Unknown upload file'
    upload_job_id = str(upload_context.get('job_id', '')).strip()
    upload_processed_at = str(upload_context.get('processed_at', '')).strip()

    top_lines = []
    for idx, row in enumerate(top_updates[:6], start=1):
        if not isinstance(row, dict):
            continue
        top_lines.append(
            f"{idx}. {str(row.get('activity_id', '')).strip()} | {str(row.get('activity_name', '')).strip()} | "
            f"{str(row.get('update_type', '')).strip()} | {str(row.get('impact', '')).strip()}"
        )

    return f"""You are Theta PMO executive schedule analyst.
Return ONLY one valid JSON object. No markdown fences and no extra text.

Data source:
- Latest uploaded file: {upload_file}
- Job ID: {upload_job_id or 'N/A'}
- Processed At: {upload_processed_at or 'N/A'}
- Workbook: {workbook_name}
- Sheet: {sheet_name}

Client ask (show this context in the response strategy):
{client_prompt or 'N/A'}

Summary metrics:
- total_updates: {int(summary.get('total_updates', 0) or 0)}
- overwrites: {int(summary.get('overwrites', 0) or 0)}
- injects: {int(summary.get('injects', 0) or 0)}
- high_impact_overwrites: {int(summary.get('high_impact_overwrites', 0) or 0)}
- max_overwrite_delay_days: {int(summary.get('max_overwrite_delay_days', 0) or 0)}
- faster_activities: {int(summary.get('faster_activities', 0) or 0)}
- max_time_saved_days: {int(summary.get('max_time_saved_days', 0) or 0)}

Top update rows:
{chr(10).join(top_lines) if top_lines else 'No rows'}

Output JSON shape:
{{
  "headline": "<single line executive headline>",
  "executive_summary": "<2-3 line plain-English summary for frontend hero>",
  "risk_outlook": "<short risk outlook sentence>",
  "what_to_show": [
    "<frontend section 1>",
    "<frontend section 2>",
    "<frontend section 3>",
    "<frontend section 4>"
  ],
  "kpis": {{
    "total_updates": 0,
    "overwrites": 0,
    "injects": 0,
    "high_impact_overwrites": 0,
    "max_overwrite_delay_days": 0
  }},
  "insights": [
    {{"title": "<insight title>", "detail": "<insight detail>", "tone": "risk|watch|positive"}}
  ],
  "action_items": [
    "<action 1>",
    "<action 2>",
    "<action 3>"
  ],
  "chart_plan": [
    {{"type": "bar", "title": "Top Overwrite Delays", "x": "Activity", "y": "Delay Days"}},
    {{"type": "pie", "title": "Overwrite vs Inject Split", "x": "Type", "y": "Count"}}
  ]
}}"""


def _normalize_project_update_frontend_summary(payload: dict, fallback_summary: dict):
    payload = payload if isinstance(payload, dict) else {}
    kpis_in = payload.get('kpis') if isinstance(payload.get('kpis'), dict) else {}

    total_updates = _to_int(kpis_in.get('total_updates', fallback_summary.get('total_updates', 0)), 0)
    overwrites = _to_int(kpis_in.get('overwrites', fallback_summary.get('overwrites', 0)), 0)
    injects = _to_int(kpis_in.get('injects', fallback_summary.get('injects', 0)), 0)
    high_impact_overwrites = _to_int(kpis_in.get('high_impact_overwrites', fallback_summary.get('high_impact_overwrites', 0)), 0)
    max_overwrite_delay_days = _to_int(kpis_in.get('max_overwrite_delay_days', fallback_summary.get('max_overwrite_delay_days', 0)), 0)

    what_to_show = [str(x).strip() for x in (payload.get('what_to_show') or []) if str(x).strip()][:8]
    insights = [x for x in (payload.get('insights') or []) if isinstance(x, dict)][:6]
    action_items = [str(x).strip() for x in (payload.get('action_items') or []) if str(x).strip()][:6]
    chart_plan = [x for x in (payload.get('chart_plan') or []) if isinstance(x, dict)][:6]

    if not what_to_show:
        what_to_show = [
            'Executive summary headline and risk outlook',
            'KPI strip: total updates, overwrites, injects, high impact delays',
            'Top overwrite activities table with impact days',
            'Action checklist for recovery governance',
        ]

    if not insights:
        insights = [
            {
                'title': 'Delay Pressure Concentration',
                'detail': 'Most impact is concentrated in overwrite activities and should be prioritized by delay size.',
                'tone': 'risk',
            },
            {
                'title': 'Recovery Window',
                'detail': 'Inject opportunities indicate where timeline recovery actions can be activated quickly.',
                'tone': 'positive',
            },
        ]

    if not action_items:
        action_items = [
            'Review top overwrite delays in daily control meetings.',
            'Convert high-impact dependencies to parallel execution where feasible.',
            'Track inject opportunities as named recovery owners with due dates.',
        ]

    if not chart_plan:
        chart_plan = [
            {'type': 'bar', 'title': 'Top Overwrite Delays', 'x': 'Activity', 'y': 'Delay Days'},
            {'type': 'pie', 'title': 'Overwrite vs Inject Split', 'x': 'Type', 'y': 'Count'},
        ]

    headline = str(payload.get('headline', '')).strip() or f"{overwrites} overwrite updates are driving current schedule variance"
    executive_summary = str(payload.get('executive_summary', '')).strip() or (
        f"{total_updates} updates detected from the latest workbook scan, with {high_impact_overwrites} high-impact overwrite items requiring immediate mitigation focus."
    )
    risk_outlook = str(payload.get('risk_outlook', '')).strip() or (
        'Risk remains elevated until high-impact overwrite activities are either compressed or resequenced.'
    )

    return {
        'headline': headline,
        'executive_summary': executive_summary,
        'risk_outlook': risk_outlook,
        'what_to_show': what_to_show,
        'kpis': {
            'total_updates': total_updates,
            'overwrites': overwrites,
            'injects': injects,
            'high_impact_overwrites': high_impact_overwrites,
            'max_overwrite_delay_days': max_overwrite_delay_days,
        },
        'insights': insights,
        'action_items': action_items,
        'chart_plan': chart_plan,
    }





@app.route('/api/whatif/pptx-sections', methods=['GET'])
@token_required
def get_whatif_pptx_sections(current_user):
    """Return and persist slide section metadata from the WhatIF PPTX reference."""
    try:
        kb_dir = _get_knowledgebase_folder(current_user.get('company_id'))
        pptx_name = request.args.get('pptx', 'Borouge_EU3_Schedule_Optimization (1).pptx')
        pptx_path = os.path.join(kb_dir, pptx_name)

        sections = _build_pptx_slide_sections(pptx_path)
        payload = {
            'pptx': pptx_name,
            'generated_at': datetime.now().isoformat(),
            'sections': sections,
        }
        pg_write_pptx_sections(payload, current_user.get('company_id'))

        return jsonify(payload), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/whatif/realtime-analysis', methods=['POST'])
@token_required
def get_realtime_whatif_analysis(current_user):
    """Generate realtime WhatIF analysis from CP workbook and persist frontend JSON payload."""
    try:
        import pandas as pd

        data = request.json or {}
        category = (data.get('category') or '').strip()
        scenario = (data.get('scenario') or '').strip()
        analysis_request = (data.get('analysis_request') or '').strip()
        start_date = (data.get('start_date') or '').strip()
        end_date = (data.get('end_date') or '').strip()

        kb_dir = _get_knowledgebase_folder(current_user.get('company_id'))
        workbook_name = data.get('workbook') or 'CP_corrected (1) (1).xlsx'
        workbook_path = os.path.join(kb_dir, workbook_name)
        if not os.path.exists(workbook_path):
            return jsonify({'error': f'Workbook not found: {workbook_name}'}), 404

        xls = pd.ExcelFile(workbook_path)
        sheet_name = xls.sheet_names[0] if xls.sheet_names else None
        if not sheet_name:
            return jsonify({'error': 'Workbook has no sheets'}), 400

        df = pd.read_excel(workbook_path, sheet_name=sheet_name)

        # Keep activity-level rows where CP issue or activity name/id is present.
        cp_issue_col = 'CP Issue Type'
        delay_col = 'Duration vs BL (days)'
        float_col = 'Total Float'
        longest_col = 'Longest Path'
        activity_id_col = 'Activity ID'
        activity_name_col = 'Activity Name'
        note_col = 'Recovery / Risk Note'

        if cp_issue_col not in df.columns or delay_col not in df.columns:
            return jsonify({'error': 'Expected CP columns are missing in workbook'}), 400

        work_df = df.copy()
        work_df[activity_id_col] = work_df.get(activity_id_col, '').fillna('').astype(str)
        work_df[activity_name_col] = work_df.get(activity_name_col, '').fillna('').astype(str)
        work_df[note_col] = work_df.get(note_col, '').fillna('').astype(str)
        work_df[cp_issue_col] = work_df[cp_issue_col].fillna('').astype(str)
        work_df['delay_days'] = work_df[delay_col].apply(_parse_delay_days)
        work_df['float_days'] = pd.to_numeric(work_df.get(float_col), errors='coerce').fillna(0.0)
        work_df['is_longest_path'] = work_df.get(longest_col, '').fillna('').astype(str).str.lower().str.contains('yes|true|y|1')

        # Derive category match from activity text since CP sheet has no explicit category column.
        category_tokens = {
            'construction_precomm': ['construction', 'pre-comm', 'precomm', 'civil', 'site'],
            'civil_construction': ['civil', 'earthwork', 'foundation', 'concrete'],
            'bulk_materials': ['material', 'procure', 'delivery', 'bulk'],
            'engineering': ['engineering', 'design', 'drawing', 'review'],
            'mechanical_construction': ['mechanical', 'piping', 'equipment', 'erection'],
            'overall': [],
        }
        row_text = (work_df[activity_id_col] + ' ' + work_df[activity_name_col]).str.lower()
        key = category.lower().strip()
        tokens = category_tokens.get(key, [])
        if key and key != 'overall' and tokens:
            mask = row_text.apply(lambda txt: any(t in txt for t in tokens))
            filtered_df = work_df[mask].copy()
            if filtered_df.empty:
                filtered_df = work_df.copy()
        else:
            filtered_df = work_df.copy()

        # Optional date filtering using Start / Finish columns when parseable.
        if start_date or end_date:
            parsed_start = pd.to_datetime(start_date, errors='coerce') if start_date else None
            parsed_end = pd.to_datetime(end_date, errors='coerce') if end_date else None
            start_series = pd.to_datetime(filtered_df.get('Start'), errors='coerce')
            finish_series = pd.to_datetime(filtered_df.get('Finish'), errors='coerce')

            date_mask = pd.Series([True] * len(filtered_df), index=filtered_df.index)
            if parsed_start is not None and not pd.isna(parsed_start):
                date_mask = date_mask & ((start_series >= parsed_start) | (finish_series >= parsed_start))
            if parsed_end is not None and not pd.isna(parsed_end):
                date_mask = date_mask & ((start_series <= parsed_end) | (finish_series <= parsed_end))

            date_filtered = filtered_df[date_mask].copy()
            if not date_filtered.empty:
                filtered_df = date_filtered

        # KPI and chart datasets
        issue_counts = (
            filtered_df[filtered_df[cp_issue_col].str.strip() != ''][cp_issue_col]
            .value_counts()
            .head(8)
            .to_dict()
        )
        issue_chart = [{'name': k, 'value': int(v)} for k, v in issue_counts.items()]

        top_delay_df = filtered_df.sort_values('delay_days', ascending=False).head(5)
        top_delay_chart = [
            {
                'name': str(row[activity_id_col])[:30],
                'value': float(row['delay_days']),
            }
            for _, row in top_delay_df.iterrows()
        ]

        delayed_df = filtered_df[filtered_df['delay_days'] > 0].sort_values('delay_days', ascending=False)
        delayed_count = int(len(delayed_df))
        total_count = int(len(filtered_df))
        avg_delay = round(float(delayed_df['delay_days'].mean()), 2) if delayed_count else 0.0
        max_delay = round(float(delayed_df['delay_days'].max()), 2) if delayed_count else 0.0
        zero_float_count = int((filtered_df['float_days'] <= 0).sum())
        longest_path_count = int(filtered_df['is_longest_path'].sum())

        top_delayed_activities = []
        for _, row in delayed_df.head(10).iterrows():
            top_delayed_activities.append({
                'activity_id': row[activity_id_col],
                'activity_name': row[activity_name_col],
                'delay_days': float(row['delay_days']),
                'cp_issue_type': row[cp_issue_col],
                'risk_note': row[note_col],
            })

        markdown = (
            f"## Realtime WhatIF Summary ({category or 'overall'})\n\n"
            f"Scenario: **{scenario or 'N/A'}**\n\n"
            f"Request: {analysis_request or 'N/A'}\n\n"
            f"- Total activities considered: **{total_count}**\n"
            f"- Delayed activities: **{delayed_count}**\n"
            f"- Average delay (days): **{avg_delay}**\n"
            f"- Max delay (days): **{max_delay}**\n"
            f"- Zero/negative float activities: **{zero_float_count}**\n"
            f"- Longest path marked activities: **{longest_path_count}**\n\n"
            "```json_chart\n"
            + json.dumps(
                {
                    'type': 'pie',
                    'title': 'CP Issue Type Distribution',
                    'data': issue_chart,
                },
                indent=2,
            )
            + "\n```\n\n"
            "```json_chart\n"
            + json.dumps(
                {
                    'type': 'bar',
                    'title': 'Top Delays (Duration vs BL)',
                    'data': top_delay_chart,
                },
                indent=2,
            )
            + "\n```\n\n"
        )

        if top_delayed_activities:
            markdown += "### Top Delayed Activities\n"
            for item in top_delayed_activities[:5]:
                markdown += (
                    f"- **{item['activity_id']}** ({item['delay_days']} days): "
                    f"{item['cp_issue_type'] or 'No CP issue type'}"
                    + (f" | Note: {item['risk_note']}" if item['risk_note'] else '')
                    + "\n"
                )

        response_payload = {
            'generated_at': datetime.now().isoformat(),
            'source': {
                'workbook': workbook_name,
                'sheet': sheet_name,
            },
            'filters': {
                'category': category,
                'scenario': scenario,
                'analysis_request': analysis_request,
                'start_date': start_date,
                'end_date': end_date,
            },
            'kpis': {
                'total_activities': total_count,
                'delayed_activities': delayed_count,
                'average_delay_days': avg_delay,
                'max_delay_days': max_delay,
                'zero_or_negative_float_count': zero_float_count,
                'longest_path_count': longest_path_count,
            },
            'charts': {
                'cp_issue_distribution': issue_chart,
                'top_delays': top_delay_chart,
            },
            'top_delayed_activities': top_delayed_activities,
            'markdown': markdown,
        }

        pg_write_whatif_realtime(response_payload, current_user.get('company_id'))
        return jsonify(response_payload), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/chat', methods=['POST'])
@token_required
def chat_message(current_user):
    _chat_t0 = _time.time()
    try:
        data = request.json
        message = data.get('message', '').strip()
        selected_sheets = data.get('selected_sheets', [])
        
        if not message:
            return jsonify({'error': 'Message is required'}), 400

        # ── Quick Greeting Handling ──
        import re as _re_local
        greetings = {'hi', 'hello', 'hey', 'greetings', 'good morning', 'good afternoon', 'good evening'}
        msg_clean = _re_local.sub(r'[^a-zA-Z\s]', '', message.lower()).strip()
        if msg_clean in greetings:
            return jsonify({
                'message': f"Hello! How can I assist you today?",
                'suggestions': ['Show delayed activities', 'Analyze critical path', 'Compare with benchmark'],
                'timestamp': datetime.now().isoformat(),
                'model': 'fast-response',
                'context': 'Direct greeting reply'
            }), 200

# ── Fast Cache Check (skip for time-relative queries) ──
        import hashlib
        import re as _re
        _quick_time_words = {'today', 'tomorrow', 'yesterday', 'next', 'last', 'this',
                            'upcoming', 'coming', 'past', 'week', 'month', 'year',
                            'deadline', 'due', 'overdue', 'schedule'}
        _is_time_query = any(tw in message.lower().split() for tw in _quick_time_words)

        # Extract activity IDs early so cache check can use them
        activity_ids = list({
            m.strip() for m in _re.findall(
                r'\b([A-Za-z]{1,6}[\.\-_]?\d{3,8}(?:[\.\-_][A-Za-z0-9]+)*)\b',
                message
            ) if m.strip()
        })
        _has_activity_id = bool(activity_ids)

        cache_key = hashlib.sha256(f"{current_user['id']}:{message}".encode()).hexdigest()
        if not _is_time_query and not _has_activity_id:
            cached_resp = get_cached_ai_response(cache_key, current_user.get('company_id'))
            if cached_resp:
                return jsonify({
                    'message': cached_resp,
                    'suggestions': [],
                    'timestamp': datetime.now().isoformat(),
                    'model': 'cached-response',
                    'context': 'Retrieved from cache'
                }), 200

        from kb_file_loader import get_kb_context
        import pandas as pd
        import os
        import re as _re

        # ── Smarter keyword extraction ──
        _stop_words = {
            'what', 'show', 'tell', 'about', 'give', 'list', 'this', 'that',
            'with', 'from', 'have', 'does', 'which', 'where', 'when', 'will',
            'would', 'could', 'should', 'there', 'their', 'they', 'them',
            'been', 'being', 'before', 'after', 'between', 'some', 'please',
            'help', 'need', 'want', 'like', 'also', 'just', 'more', 'most',
            'much', 'many', 'make', 'made', 'know', 'into', 'only', 'very',
            'than', 'then', 'each', 'every', 'such', 'over', 'take', 'came',
            'come', 'your', 'mine', 'ours', 'next', 'last', 'month', 'week',
            'year', 'today', 'tomorrow', 'yesterday', 'date', 'time',
            'going', 'will', 'get', 'are', 'the', 'for', 'and', 'when',
        }
        raw_words = _re.findall(r'[A-Za-z0-9_.-]+', message.lower())
        keywords = [w for w in raw_words if len(w) > 2 and w not in _stop_words]

        # Also extract multi-word quoted phrases AND unquoted noun phrases (2-3 consecutive words)
        # This is what makes "ACS Pre Comm", "butterfly valves", "PAGA delivery" searchable as phrases
        raw_tokens = _re.findall(r'[A-Za-z0-9]+', message)
        noun_phrases = []
        for i in range(len(raw_tokens)):
            for j in range(2, 4):  # 2-word and 3-word phrases
                phrase = ' '.join(raw_tokens[i:i+j]).lower()
                if len(phrase) > 4:
                    noun_phrases.append(phrase)

        # Also extract quoted phrases and activity IDs (e.g. A60780)
        # quoted_phrases = _re.findall(r'"([^"]+)"', message) + _re.findall(r"'([^']+)'", message)
        # Also extract quoted phrases and activity IDs (e.g. A60780)
        quoted_phrases = _re.findall(r'"([^"]+)"', message) + _re.findall(r"'([^']+)'", message)

        # Delivery-intent detection: if user asks about delivery of something,
        # inject "delivery of <thing>" as an explicit high-priority search phrase.
        # This prevents broad single-word matches (e.g. "PAGA") from drowning out
        # the actual delivery activity row ("Delivery of PAGA + Telephone").
        _delivery_words = {'deliver', 'delivered', 'delivery', 'when', 'date', 'schedule'}
        _msg_lower = message.lower()
        if any(dw in _msg_lower for dw in _delivery_words):
            # Extract the subject being asked about (words that aren't question/delivery words)
            _filler = {'when', 'will', 'the', 'be', 'are', 'is', 'going', 'to', 'get',
                       'what', 'delivery', 'deliver', 'delivered', 'date', 'schedule', 'of'}
            _subject_words = [w for w in _re.findall(r'[A-Za-z0-9]+', message)
                              if w.lower() not in _filler and len(w) > 1]
            if _subject_words:
                # Build "delivery of <subject>" phrases and add as high-priority quoted phrases
                _delivery_phrases = [
                    f"delivery of {' '.join(_subject_words).lower()}",
                    f"delivery of {_subject_words[0].lower()}",
                ]
                quoted_phrases = _delivery_phrases + quoted_phrases
                noun_phrases = _delivery_phrases + noun_phrases

        activity_ids = list({
            m.strip() for m in _re.findall(
                r'\b([A-Za-z]{1,6}[\.\-_]?\d{3,8}(?:[\.\-_][A-Za-z0-9]+)*)\b',
                message
            ) if m.strip()
        })
        # ── Build user output context — search ALL completed jobs, ALL sheets ──
        _chat_role = current_user.get('role')
        _chat_cid  = current_user.get('company_id')
        if _chat_role in ('admin', 'company_admin', 'manager') and _chat_cid:
            user_jobs = pg_read_history_for_company(company_id=_chat_cid, status='completed')
        else:
            user_jobs = pg_read_history_for_company(user_id=current_user['id'], status='completed')
        
        context_lines = []
        document_blocks = []
        _total_chars = 0
        _CONTEXT_CHAR_LIMIT = 30_000

        # Limit to 3-5 most recent jobs to speed up reading multiple heavy Excel files
        target_jobs = user_jobs[:20]
        for job in target_jobs:
            job_id = job.get('id')
            for result in job.get('results', []):
                if result.get('status') != 'success':
                    continue
                output_filename = result.get('output_filename')
                if not output_filename:
                    continue
                file_path = os.path.join(_APP_ROOT, OUTPUT_FOLDER, job_id, output_filename)
                # Fallback: if output file missing, try the original uploaded file
                if not os.path.exists(file_path):
                    raw_upload = os.path.join(_APP_ROOT, UPLOAD_FOLDER, job.get('filename', ''))
                    if os.path.exists(raw_upload):
                        file_path = raw_upload
                    else:
                        continue
                # Full-document attachment via the Files API was removed here:
                # it re-attached every historical job's full workbook on
                # every single chat message (up to 20 jobs), which bloated
                # the model's input and was the dominant cause of slow
                # responses. The keyword-matched row search below already
                # surfaces the relevant data far more cheaply.
                try:
                    all_sheets = _read_table_sheets_cached(file_path)
                    for sheet_name, sheet_df in all_sheets.items():
                        # Skip legend/metadata sheets
                        _skip_sheet_keywords = ['legend', 'color', 'colour', 'key', 'guide', 'readme', 'instructions', 'notes', 'cover']
                        if any(sk in sheet_name.lower() for sk in _skip_sheet_keywords):
                            continue
                        cols_str = ', '.join(str(c) for c in sheet_df.columns)
                        overview = f"[Source: File={output_filename}, Sheet={sheet_name}, Rows={len(sheet_df)}, Columns={cols_str}]"

                        matched_df = pd.DataFrame()
                        if keywords or quoted_phrases or activity_ids:
                            all_search_terms = list({
                                t.lower() for t in (keywords + noun_phrases + [p.lower() for p in quoted_phrases] + activity_ids)
                                if t.strip()
                            })
                            # Vectorized search across ALL columns (not just 'name'/'id' columns)
                            pattern = '|'.join(_re.escape(t) for t in all_search_terms)
                            str_df = sheet_df.astype(str).apply(lambda col: col.str.strip().str.lower())
                            col_masks = [str_df[col].str.contains(pattern, regex=True, na=False) for col in str_df.columns]
                            combined = col_masks[0]
                            for m in col_masks[1:]:
                                combined = combined | m
                            matched_df = sheet_df[combined]

                        if not matched_df.empty:
                            # Score each row: exact multi-word phrase matches score higher than single-word matches
                            # This ensures "Delivery of Special GGC and Butterfly Valves" scores above
                            # "Datasheets for Butterfly Valves" when user asks about delivery
                            str_matched = matched_df.astype(str).apply(lambda col: col.str.strip().str.lower())
                            def _score_row(row):
                                row_text = ' '.join(str(v) for v in row.values)
                                score = 0
                                for t in all_search_terms:
                                    if t in row_text:
                                        score += len(t.split())  # longer phrase match = higher score
                                return score
                            matched_df = matched_df.copy()
                            matched_df['_score'] = str_matched.apply(_score_row, axis=1)
                            matched_df = matched_df.sort_values('_score', ascending=False).drop(columns=['_score'])
                            block = f"\n{overview}\n{matched_df.head(50).to_string()}"
                            context_lines.append(block)
                            _total_chars += len(block)
                        elif _total_chars < _CONTEXT_CHAR_LIMIT:
                            sample = sheet_df.head(50).to_string()
                            block = f"\n{overview}\n[Sample rows:]\n{sample}"
                            context_lines.append(block)
                            _total_chars += len(block)

                except Exception as e:
                    print(f"[CHAT] Error reading {output_filename}: {e}")

        user_context = ""
        if context_lines:
            user_context = (
                "\n\n=== USER PROJECT DATA (from processed output files) ===\n"
                + "\n".join(context_lines)
                + "\n=== END OF USER PROJECT DATA ===\n"
            )

        # ── Direct uploaded files search (raw uploads folder) ──
        upload_lines = []
        _upload_chars = 0
        _UPLOAD_CHAR_LIMIT = 20_000
        _uploads_dir = os.path.join(_APP_ROOT, UPLOAD_FOLDER)

        if os.path.isdir(_uploads_dir):
            import glob as _glob
            _upload_files = (
                _glob.glob(os.path.join(_uploads_dir, '*.xlsx')) +
                _glob.glob(os.path.join(_uploads_dir, '*.xls')) +
                _glob.glob(os.path.join(_uploads_dir, '*.csv'))
            )
            _upload_files.sort(key=lambda p: os.path.getmtime(p), reverse=True)

            for up_path in _upload_files:
                if _upload_chars >= _UPLOAD_CHAR_LIMIT:
                    break
                up_basename = os.path.basename(up_path)

                try:
                    up_sheets = _read_table_sheets_cached(up_path)

                    for up_sheet_name, up_df in up_sheets.items():
                        _skip_sheet_keywords = ['legend', 'color', 'colour', 'key', 'guide',
                                                 'readme', 'instructions', 'notes', 'cover']
                        if any(sk in up_sheet_name.lower() for sk in _skip_sheet_keywords):
                            continue
                        up_df = up_df.dropna(how='all').dropna(axis=1, how='all')
                        if up_df.empty:
                            continue

                        cols_str = ', '.join(str(c) for c in up_df.columns)
                        overview = (f"[Upload Source: File={up_basename}, Sheet={up_sheet_name}, "
                                    f"Rows={len(up_df)}, Columns={cols_str}]")

                        all_search_terms = list({
                            t.lower() for t in (
                                keywords + noun_phrases +
                                [p.lower() for p in quoted_phrases] + activity_ids
                            ) if t.strip()
                        })

                        matched_up = pd.DataFrame()
                        if all_search_terms:
                            pattern = '|'.join(_re.escape(t) for t in all_search_terms)
                            str_up = up_df.astype(str).apply(lambda col: col.str.strip().str.lower())
                            col_masks = [str_up[col].str.contains(pattern, regex=True, na=False)
                                         for col in str_up.columns]
                            combined = col_masks[0]
                            for cm in col_masks[1:]:
                                combined = combined | cm
                            matched_up = up_df[combined]

                        if not matched_up.empty:
                            str_matched_up = matched_up.astype(str).apply(
                                lambda col: col.str.strip().str.lower())
                            def _score_row_up(row):
                                row_text = ' '.join(str(v) for v in row.values)
                                return sum(len(t.split()) for t in all_search_terms if t in row_text)
                            matched_up = matched_up.copy()
                            matched_up['_score'] = str_matched_up.apply(_score_row_up, axis=1)
                            matched_up = (matched_up.sort_values('_score', ascending=False)
                                          .drop(columns=['_score']))
                            block = f"\n{overview}\n{matched_up.head(30).to_string()}"
                            upload_lines.append(block)
                            _upload_chars += len(block)
                        elif _upload_chars < _UPLOAD_CHAR_LIMIT:
                            sample = up_df.head(30).to_string()
                            block = f"\n{overview}\n[Sample rows:]\n{sample}"
                            upload_lines.append(block)
                            _upload_chars += len(block)

                except Exception as e:
                    print(f"[CHAT] Error reading upload file {up_basename}: {e}")

        upload_context = ""
        if upload_lines:
            upload_context = (
                "\n\n=== UPLOADED FILES DATA ===\n"
                + "\n".join(upload_lines)
                + "\n=== END OF UPLOADED FILES DATA ===\n"
            )

        # ── Direct KB file search ──
        KB_FOLDER = _get_knowledgebase_folder(current_user.get('company_id'))
        kb_lines = []
        _kb_chars = 0
        _KB_CHAR_LIMIT = 20_000

        if os.path.isdir(KB_FOLDER):
            import glob
            kb_files = glob.glob(os.path.join(KB_FOLDER, '*.xlsx')) + \
                       glob.glob(os.path.join(KB_FOLDER, '*.xls')) + \
                       glob.glob(os.path.join(KB_FOLDER, '*.csv'))

            for kb_path in kb_files:
                kb_basename = os.path.basename(kb_path)
                try:
                    kb_sheets = _read_table_sheets_cached(kb_path)

                    for kb_sheet_name, kb_df in kb_sheets.items():
                        # Skip legend/metadata/color-coding sheets — they never contain activity data
                        _skip_sheet_keywords = ['legend', 'color', 'colour', 'key', 'guide', 'readme', 'instructions', 'notes', 'cover']
                        if any(sk in kb_sheet_name.lower() for sk in _skip_sheet_keywords):
                            print(f"[CHAT] Skipping metadata sheet: {kb_basename} / {kb_sheet_name}")
                            continue
                        kb_df = kb_df.dropna(how='all').dropna(axis=1, how='all')
                        if kb_df.empty:
                            continue
                        # Skip sheets with no recognizable activity/schedule columns
                        col_names_lower = [str(c).lower() for c in kb_df.columns]
                        _useful_cols = ['activity', 'id', 'name', 'start', 'finish', 'date', 'successor', 'predecessor', 'duration', 'description']
                        if not any(uc in ' '.join(col_names_lower) for uc in _useful_cols):
                            print(f"[CHAT] Skipping non-schedule sheet: {kb_basename} / {kb_sheet_name} cols={col_names_lower[:5]}")
                            continue

                        cols_str = ', '.join(str(c) for c in kb_df.columns)
                        overview = f"[KB Source: File={kb_basename}, Sheet={kb_sheet_name}, Rows={len(kb_df)}, Columns={cols_str}]"

                        all_search_terms = list({
                            t.lower() for t in (keywords + noun_phrases + [p.lower() for p in quoted_phrases] + activity_ids)
                            if t.strip()
                        })
                        matched_kb = pd.DataFrame()
                        if all_search_terms:
                            pattern = '|'.join(_re.escape(t) for t in all_search_terms)
                            str_kb = kb_df.astype(str).apply(lambda col: col.str.strip().str.lower())
                            col_masks = [str_kb[col].str.contains(pattern, regex=True, na=False) for col in str_kb.columns]
                            combined = col_masks[0]
                            for cm in col_masks[1:]:
                                combined = combined | cm
                            matched_kb = kb_df[combined]

                        if not matched_kb.empty:
                            str_matched_kb = matched_kb.astype(str).apply(lambda col: col.str.strip().str.lower())
                            def _score_row_kb(row):
                                row_text = ' '.join(str(v) for v in row.values)
                                return sum(len(t.split()) for t in all_search_terms if t in row_text)
                            matched_kb = matched_kb.copy()
                            matched_kb['_score'] = str_matched_kb.apply(_score_row_kb, axis=1)
                            matched_kb = matched_kb.sort_values('_score', ascending=False).drop(columns=['_score'])
                            block = f"\n{overview}\n{matched_kb.head(30).to_string()}"
                            kb_lines.append(block)

                except Exception as e:
                    print(f"[CHAT] Error reading KB file {kb_basename}: {e}")

        kb_direct_context = ""
        if kb_lines:
            kb_direct_context = (
                "\n\n=== KNOWLEDGEBASE DATA ===\n"
                + "\n".join(kb_lines)
                + "\n=== END OF DATA ===\n"
            )
        
        whatif_context = _load_whatif_chat_context(current_user.get('company_id'))

        # ── Build full prompt — ALWAYS inject today's date, let AI handle the rest ──
        _now = datetime.now()
        current_date_str = _now.strftime('%B %d, %Y')
        current_date_iso = _now.strftime('%Y-%m-%d')

        system_prompt = (
            "You are Theta PMO AI — an expert project schedule analyst.\n\n"
            f"TODAY'S DATE: {current_date_str} ({current_date_iso}). Current year: {_now.year}.\n\n"
            "CORE RULES:\n"
            "1. ANSWER DIRECTLY: Lead with the direct answer. Do not say 'I searched...' or 'Based on my search...'. Just answer.\n"
            "2. SEARCH ALL DATA: The data blocks below contain your only source of truth. Search every block thoroughly.\n"
            "3. ACTIVITY NAME SEARCH: When the user mentions an activity name (e.g. 'butterfly valves', 'PAGA', 'ACS Pre Comm'), "
            "search for it in both Activity Name and Activity ID columns. The data may contain it even if you don't see it immediately.\n"
            "4. SUCCESSOR/PREDECESSOR: This data is in the 'Successor Details' and 'Predecessor Details' columns of PREDECESSOR-SUCCESSOR-LAG.xlsx "
            "and in the What-If Predecessor-Successor Scenarios block. Always report the exact activity IDs and relationship types (FS, FF, SS) found there.\n"
            "5. DELIVERY DATES: Delivery activities contain words like 'Delivery of' in the Activity Name. "
            "Report their Early Start and Early Finish as the delivery date range.\n"
            "6. SOURCE CITATION: Always cite the File name and Sheet name.\n"
            "7. NO FABRICATION: If data is truly not in the blocks below, say exactly what you did find and what is missing.\n"
            "8. DATE HANDLING: For time-period queries, calculate from "
            f"TODAY ({current_date_iso}) and only show activities within that range.\n"
            "9. WHATIF SCENARIOS: When the user asks 'what if we accelerate X', 'what happens if X finishes early', "
            "or 'which activities are blocking progress' — use the What-If Analysis Data block. "
            "Identify the bottleneck activity, list its successors, state how many days could be recovered, "
            "and recommend which predecessor to accelerate first. When given an activity ID, look it up in "
            "both the Predecessor-Successor Scenarios and the Critical Threat Tracker.\n"
        )

        # Separate matched blocks (contain the searched activity) from sample blocks
        _priority_terms = activity_ids if activity_ids else (noun_phrases + keywords + [p.lower() for p in quoted_phrases])
        matched_blocks = [b for b in context_lines if _priority_terms and any(t.lower() in b.lower() for t in _priority_terms)]
        other_blocks = [b for b in context_lines if b not in matched_blocks]
        
        priority_context = ""
        if matched_blocks:
            priority_context = (
                "\n\n=== DIRECT MATCHES FOR QUERY (CHECK HERE FIRST) ===\n"
                + "\n".join(matched_blocks)
                + "\n=== END DIRECT MATCHES ===\n"
            )
            user_context = (
                "\n\n=== OTHER PROJECT DATA ===\n"
                + "\n".join(other_blocks)
                + "\n=== END OTHER DATA ===\n"
            ) if other_blocks else ""

        MAX_PROMPT_CHARS = 120_000
        full_prompt = f"{system_prompt}\n\nUSER QUERY: {message}\n{priority_context}\n{user_context}\n{upload_context}\n{kb_direct_context}\n{whatif_context}"
        if len(full_prompt) > MAX_PROMPT_CHARS:
            allowed_other = MAX_PROMPT_CHARS - len(system_prompt) - len(priority_context) - len(kb_direct_context) - len(upload_context) - 2000
            trimmed_user = user_context[:max(0, allowed_other)]
            full_prompt = f"{system_prompt}\n\nUSER QUERY: {message}\n{priority_context}\n{trimmed_user}\n{upload_context}\n{kb_direct_context}\n{whatif_context}"
        print(f"[CHAT DEBUG] context_lines={len(context_lines)}, upload_lines={len(upload_lines)}, kb_lines={len(kb_lines)}, whatif_chars={len(whatif_context)}, prompt_chars={len(full_prompt)}, activity_ids={activity_ids}, matched_blocks={len(matched_blocks)}, context_build_sec={_time.time() - _chat_t0:.2f}")
        
        # ── Call Azure Anthropic Claude ──
        try:
            chat_max_tokens = max(512, int(os.getenv('CLAUDE_CHAT_MAX_TOKENS', '3072')))
            ai_response = _generate_claude_response(full_prompt, max_tokens=chat_max_tokens, document_blocks=document_blocks or None)
        except RuntimeError:
            return jsonify({'error': 'Azure Anthropic service is not available.'}), 503

        # Cache the response for faster future lookups
        try:
            save_cached_ai_response(cache_key, current_user['id'], message, ai_response, "", current_user.get('company_id'))
        except Exception as cache_err:
            print(f"[CHAT] Cache save error: {cache_err}")

        # Persist to chat history for admin reporting
        log_chat(
            user_id=current_user['id'],
            user_name=current_user.get('name', ''),
            user_role=current_user.get('role', 'user'),
            route='/api/chat',
            message=message,
            response=ai_response,
            model='azure-anthropic-claude',
            context_info=f"kb_lines={len(kb_lines)}, upload_lines={len(upload_lines)}",
        )

        # Log AI chat activity so it appears in the Knowledge Base history tab
        try:
            log_activity(
                action_type=ACTION_AI_CHAT,
                user_id=current_user['id'],
                user_name=current_user.get('name', ''),
                user_role=current_user.get('role', 'user'),
                company_id=current_user.get('company_id'),
                description=f"{current_user.get('name', 'User')} sent an AI chat message",
                source=request.headers.get('X-App-Source', SOURCE_WEB),
                level=LEVEL_USER,
                ip_address=request.remote_addr,
                metadata={
                    'prompt': message,
                    'response': ai_response,
                    'quick_prompt': data.get('quick_prompt', ''),
                },
            )
        except Exception as log_err:
            print(f"[CHAT] Activity log error: {log_err}")

        print(f"[CHAT DEBUG] total_request_sec={_time.time() - _chat_t0:.2f}")

        return jsonify({
            'message': ai_response,
            'suggestions': [],
            'timestamp': datetime.now().isoformat(),
            'model': 'azure-anthropic-claude',
            'context': f'Searched {len(context_lines)} output blocks + {len(upload_lines)} upload blocks + KB'
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'message': 'Error generating response.',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

# ==================== DEVIATION MANAGEMENT ====================

@app.route('/deviations', methods=['GET'])
@token_required
def get_deviations(current_user):
    """Get critical-path deviations scoped to the current user's company."""
    try:
        role = current_user.get('role', 'user')
        company_id = request.args.get('company_id') if role == 'super_admin' else current_user.get('company_id')
        user_id    = current_user['id']

        deviations = get_all_deviations(
            company_id=company_id or None,
            user_id=user_id or None,
        )

        cp_ids, cp_names = _load_cp_and_ms_ids(company_id)
        if cp_ids or cp_names:
            deviations = [d for d in deviations if _is_cp_deviation(d, cp_ids, cp_names)]

        return jsonify(deviations), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to fetch deviations: {str(e)}'}), 500


@app.route('/deviations/history', methods=['GET'])
@token_required
def get_deviations_history(current_user):
    """Get critical-path deviation history scoped to the current user's company."""
    try:
        role = current_user.get('role', 'user')
        company_id = request.args.get('company_id') if role == 'super_admin' else current_user.get('company_id')
        user_id    = current_user['id']
        limit      = request.args.get('limit', type=int)

        reviewed_deviations = get_reviewed_deviations(
            company_id=company_id or None,
            user_id=user_id or None,
            limit=limit,
        )

        cp_ids, cp_names = _load_cp_and_ms_ids(company_id)
        if cp_ids or cp_names:
            reviewed_deviations = [d for d in reviewed_deviations if _is_cp_deviation(d, cp_ids, cp_names)]

        return jsonify(reviewed_deviations), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to fetch deviation history: {str(e)}'}), 500

@app.route('/deviations', methods=['POST'])
@token_required
def create_deviation(current_user):
    """Create a new deviation entry (legacy route — requires auth)."""
    try:
        data = request.get_json()

        # Create new deviation — enforce company from token, not from caller
        new_deviation = {
            'sheet': data.get('sheet', ''),
            'flag': data.get('flag', ''),
            'severity': data.get('severity', ''),
            'description': data.get('description', ''),
            'row_data': data.get('row_data', {}),
            'detected_at': data.get('detected_at', datetime.now().isoformat()),
            'review_status': 'Pending',
            'review_reason': '',
            'reason_type': '',
            'user_id': current_user['id'],
            'company_id': current_user.get('company_id'),
        }
        
        deviation_id = insert_deviation(new_deviation)
        new_deviation['id'] = deviation_id
        
        print(f"[DEVIATION] Created new deviation #{deviation_id}")
        
        return jsonify({
            'status': 'success',
            'message': 'Deviation created successfully',
            'deviation': new_deviation
        }), 201
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to create deviation: {str(e)}'}), 500

@app.route('/deviations/update/<deviation_id>', methods=['POST'])
@token_required
def update_deviation(current_user, deviation_id):
    """Update a deviation with review status and reason (legacy route — requires auth)."""
    try:
        data = request.get_json()

        deviation = get_deviation_by_id(deviation_id)
        if not deviation:
            return jsonify({'error': 'Deviation not found'}), 404
        if current_user.get('role') != 'super_admin' and deviation.get('company_id') != current_user.get('company_id'):
            return jsonify({'error': 'Deviation not found'}), 404

        fields = {
            'review_status': data.get('review_status', 'Reviewed'),
            'review_reason': data.get('review_reason', ''),
            'reason_type': data.get('reason_type', ''),
            'reviewed_at': datetime.now().isoformat(),
            'reviewed_by_user_id': current_user['id'],
        }
        update_deviation_fields(deviation_id, fields)
        deviation.update(fields)
        
        print(f"[DEVIATION] Updated deviation #{deviation_id} with reason: {fields['review_reason']}")
        
        return jsonify({
            'status': 'success',
            'message': 'Deviation updated successfully',
            'deviation': deviation
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to update deviation: {str(e)}'}), 500

@app.route('/deviations/admin/approve/<deviation_id>', methods=['POST'])
@token_required
def admin_approve_deviation(current_user, deviation_id):
    """Admin approves a deviation (legacy route — requires auth)."""
    role = current_user.get('role')
    if role not in ('admin', 'manager', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Admin/manager access required'}), 403
    try:
        data = request.get_json() or {}

        deviation = get_deviation_by_id(deviation_id)
        if not deviation:
            return jsonify({'error': 'Deviation not found'}), 404
        if role != 'super_admin' and deviation.get('company_id') != current_user.get('company_id'):
            return jsonify({'error': 'Deviation not found'}), 404
        
        fields = {
            'review_status': 'Approved',
            'admin_comment': data.get('reason', ''),
            'reviewed_at': datetime.now().isoformat(),
            'reviewed_by_user_id': data.get('user_id'),
        }
        update_deviation_fields(deviation_id, fields)
        deviation.update(fields)
        
        print(f"[ADMIN] Approved deviation #{deviation_id}")

        # Track activity
        try:
            reviewer_id = data.get('user_id')
            log_activity(
                action_type=ACTION_DEVIATION_APPROVE,
                user_id=reviewer_id,
                user_role='admin',
                entity_type='deviation', entity_id=str(deviation_id),
                description=f"Admin approved deviation #{deviation_id}",
                source=request.headers.get('X-App-Source', SOURCE_WEB),
                level=LEVEL_ADMIN,
                metadata={
                    'deviation_id':  deviation_id,
                    'activity_name': (deviation.get('row_data') or {}).get('activity_name', '') if isinstance(deviation.get('row_data'), dict) else '',
                    'gate':          (deviation.get('row_data') or {}).get('stage_gate', '') if isinstance(deviation.get('row_data'), dict) else '',
                    'comment':       data.get('reason', ''),
                },
                ip_address=request.remote_addr,
            )
        except Exception:
            pass

        # Optionally notify the manager who submitted this
        try:
            if deviation.get('user_id'):
                create_notification(
                    user_id=deviation['user_id'],
                    title="Deviation Approved",
                    message=f"Your deviation report (#{deviation_id}) has been approved by admin.",
                    notification_type="approval"
                )
        except Exception as notif_error:
            print(f"[ADMIN] Failed to send approval notification: {notif_error}")

        return jsonify({
            'status': 'success',
            'message': 'Deviation approved successfully',
            'deviation': deviation
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to approve deviation: {str(e)}'}), 500

@app.route('/deviations/admin/reject/<deviation_id>', methods=['POST'])
@token_required
def admin_reject_deviation(current_user, deviation_id):
    """Admin rejects a deviation (legacy route — requires auth)."""
    role = current_user.get('role')
    if role not in ('admin', 'manager', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Admin/manager access required'}), 403
    try:
        data = request.get_json() or {}
        reason = data.get('reason', '')

        if not reason or not reason.strip():
            return jsonify({'error': 'Reason is required for rejection'}), 400

        deviation = get_deviation_by_id(deviation_id)
        if deviation and role != 'super_admin' and deviation.get('company_id') != current_user.get('company_id'):
            return jsonify({'error': 'Deviation not found'}), 404
        if not deviation:
            return jsonify({'error': 'Deviation not found'}), 404
        
        fields = {
            'review_status': 'Not Approved',
            'admin_comment': reason,
            'reviewed_at': datetime.now().isoformat(),
            'reviewed_by_user_id': data.get('user_id'),
        }
        update_deviation_fields(deviation_id, fields)
        deviation.update(fields)
        
        print(f"[ADMIN] Rejected deviation #{deviation_id} - Reason: {reason}")

        # Track activity
        try:
            reviewer_id = data.get('user_id')
            log_activity(
                action_type=ACTION_DEVIATION_REJECT,
                user_id=reviewer_id,
                user_role='admin',
                entity_type='deviation', entity_id=str(deviation_id),
                description=f"Admin rejected deviation #{deviation_id}: {reason}",
                source=request.headers.get('X-App-Source', SOURCE_WEB),
                level=LEVEL_ADMIN,
                metadata={
                    'deviation_id':  deviation_id,
                    'activity_name': (deviation.get('row_data') or {}).get('activity_name', '') if isinstance(deviation.get('row_data'), dict) else '',
                    'gate':          (deviation.get('row_data') or {}).get('stage_gate', '') if isinstance(deviation.get('row_data'), dict) else '',
                    'reason':        reason,
                },
                ip_address=request.remote_addr,
            )
        except Exception:
            pass

        # Notify the manager who submitted this
        try:
            if deviation.get('user_id'):
                create_notification(
                    user_id=deviation['user_id'],
                    title="Deviation Not Approved",
                    message=f"Your deviation report (#{deviation_id}) was not approved. Reason: {reason}",
                    notification_type="rejection"
                )
        except Exception as notif_error:
            print(f"[ADMIN] Failed to send rejection notification: {notif_error}")

        return jsonify({
            'status': 'success',
            'message': 'Deviation rejected successfully',
            'deviation': deviation
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to reject deviation: {str(e)}'}), 500

# ==================== DEVIATION NOTIFICATIONS ====================

@app.route('/api/deviation/submit', methods=['POST'])
@token_required
def submit_deviation(current_user):
    """
    When manager submits deviation reason, notify all admins
    """
    try:
        data = request.get_json()
        
        # Validate that user is a manager
        if current_user.get('role') != 'manager':
            return jsonify({'error': 'Only managers can submit deviations'}), 403
        
        # Extract deviation data
        deviation_reason = data.get('reason', '')
        deviation_type = data.get('type', 'General')
        deviation_id = data.get('deviation_id', '')
        job_id = data.get('job_id', '')
        sheet_name = data.get('sheet_name', '')
        filename = data.get('filename', '')
        
        # Try to get more context from deviation record if ID provided
        if deviation_id:
            deviation_record = get_deviation_by_id(deviation_id)
            if deviation_record:
                sheet_name = sheet_name or deviation_record.get('sheet', '')
                if not filename and deviation_record.get('description'):
                    filename = deviation_record.get('description', '')[:50]
        
        if not deviation_reason:
            return jsonify({'error': 'Deviation reason is required'}), 400
        
        # Create notification message
        message = f"{current_user.get('name', 'Manager')} submitted deviation reason"
        if deviation_id:
            message += f" for Deviation #{deviation_id}"
        if filename:
            message += f" - {filename}"
        if sheet_name:
            message += f" (Sheet: {sheet_name})"
        message += f" - Reason: {deviation_reason}"
        
        # Notify only admins/managers in the same company — no cross-company leakage
        company_id = current_user.get('company_id')
        notifications_created = notify_admins_and_managers(
            'Deviation Submitted',
            message,
            notification_type='warning',
            metadata={
                'job_id': job_id,
                'sheet_name': sheet_name,
                'filename': filename,
                'deviation_type': deviation_type,
                'deviation_reason': deviation_reason,
                'manager_id': current_user['id'],
                'manager_name': current_user.get('name', 'Unknown'),
            },
            company_id=company_id,
        )
        
        print(f"[DEVIATION] Created {len(notifications_created)} notification(s) for admins")
        
        return jsonify({
            'status': 'success',
            'message': 'Deviation submitted and admins notified',
            'notifications_sent': len(notifications_created)
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to submit deviation: {str(e)}'}), 500


# ==================== EMAIL ENDPOINT ====================

@app.route('/api/send-summary-email', methods=['POST'])
@token_required
def send_summary_email(current_user):
    """Send AI-generated project summary to the logged-in user's email."""
    try:
        data = request.get_json()
        summary_content = data.get('content', '')
        recipient_email = current_user.get('email')
        date_range = data.get('date_range', '')

        if not recipient_email:
            return jsonify({'error': 'No email address found for this account'}), 400

        if not summary_content:
            return jsonify({'error': 'No summary content provided'}), 400

        sender_email = 'anuragkatrerd@gmail.com'
        sender_password = os.getenv('GMAIL_APP_PASSWORD', '')

        if not sender_password:
            return jsonify({'error': 'Email service not configured. Please set GMAIL_APP_PASSWORD in .env'}), 500

        subject = f'Project Summary Report – AI Project Advisor'
        if date_range:
            subject = f'Project Summary Report ({date_range}) – AI Project Advisor'

        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = f'Theta PMO AI <{sender_email}>'
        msg['To'] = recipient_email

        # Plain text part
        text_part = MIMEText(summary_content, 'plain')

        # ── Markdown → HTML converter ──────────────────────────────────────
        def _inline(s):
            s = _re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', s)
            s = _re.sub(r'\*(.+?)\*',     r'<em>\1</em>',         s)
            return s

        def _flag_badge(cell):
            cell = cell.strip()
            if 'On Time' in cell:
                return f'<span style="background:#d1fae5;color:#065f46;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:700;white-space:nowrap;">{cell}</span>'
            if 'Not Started' in cell:
                return f'<span style="background:#f3f4f6;color:#6b7280;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:700;white-space:nowrap;">{cell}</span>'
            if any(w in cell for w in ('Delayed','Late','Overdue')):
                return f'<span style="background:#fee2e2;color:#dc2626;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:700;white-space:nowrap;">{cell}</span>'
            if 'At Risk' in cell:
                return f'<span style="background:#fef3c7;color:#d97706;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:700;white-space:nowrap;">{cell}</span>'
            return cell

        def flush_table(rows, out):
            if not rows: return
            header_cells = [c.strip() for c in rows[0].strip('|').split('|')]
            data_rows = []
            for row in rows[1:]:
                cells = [c.strip() for c in row.strip('|').split('|')]
                if all(_re.match(r'^:?-+:?$', c) for c in cells if c): continue  # separator
                data_rows.append(cells)
            h = ['<div style="overflow-x:auto;margin:18px 0;border-radius:10px;border:1px solid #d1fae5;">']
            h.append('<table style="width:100%;border-collapse:collapse;font-size:12px;">')
            h.append('<thead><tr style="background:linear-gradient(135deg,#064e3b 0%,#065f46 100%);">') 
            for col in header_cells:
                if col:
                    h.append(f'<th style="padding:10px 13px;text-align:left;color:#ffffff;font-weight:700;font-size:11.5px;white-space:nowrap;border-right:1px solid rgba(255,255,255,0.15);letter-spacing:0.2px;">{col}</th>')
            h.append('</tr></thead><tbody>')
            for r_idx, cells in enumerate(data_rows):
                bg = '#ffffff' if r_idx % 2 == 0 else '#f0fdf4'
                h.append(f'<tr style="background:{bg};">')
                for c_idx, cell in enumerate(cells):
                    is_first = (c_idx == 0)
                    style = 'padding:8px 13px;font-size:12px;border-right:1px solid #d1fae5;border-bottom:1px solid #ecfdf5;vertical-align:middle;'
                    style += ('font-weight:700;color:#065f46;white-space:nowrap;' if is_first else 'color:#374151;')
                    h.append(f'<td style="{style}">{_flag_badge(cell) if not is_first else cell}</td>')
                h.append('</tr>')
            h.append('</tbody></table></div>')
            out.extend(h)

        def md_to_html(text):
            lines = text.split('\n')
            out = []
            in_ul = in_ol = in_table = False
            table_rows = []

            def close_lists():
                nonlocal in_ul, in_ol
                if in_ul: out.append('</ul>'); in_ul = False
                if in_ol: out.append('</ol>'); in_ol = False

            for line in lines:
                s = line.strip()
                is_trow = s.startswith('|') and s.endswith('|') and len(s) > 2

                if is_trow:
                    close_lists()
                    in_table = True
                    table_rows.append(s)
                    continue
                elif in_table:
                    flush_table(table_rows, out)
                    table_rows = []; in_table = False

                if not s:
                    close_lists()
                    out.append('<div style="margin:8px 0"></div>')
                elif s.startswith('#### '):
                    close_lists()
                    out.append(f'<h4 style="margin:12px 0 3px;font-size:12.5px;color:#059669;font-weight:700;text-transform:uppercase;letter-spacing:0.4px;">{_inline(s[5:])}</h4>')
                elif s.startswith('### '):
                    close_lists()
                    out.append(f'<h3 style="margin:18px 0 6px;font-size:14px;color:#065f46;font-weight:700;padding-bottom:4px;border-bottom:1px solid #a7f3d0;">{_inline(s[4:])}</h3>')
                elif s.startswith('## '):
                    close_lists()
                    out.append(f'<h2 style="margin:22px 0 8px;font-size:17px;color:#064e3b;font-weight:800;">{_inline(s[3:])}</h2>')
                elif s.startswith('# '):
                    close_lists()
                    out.append(f'<h1 style="margin:22px 0 8px;font-size:19px;color:#064e3b;font-weight:800;">{_inline(s[2:])}</h1>')
                elif s.startswith(('- ', '* ', '\u2022 ')):
                    if in_ol: out.append('</ol>'); in_ol = False
                    if not in_ul: out.append('<ul style="margin:8px 0;padding-left:22px;">'); in_ul = True
                    out.append(f'<li style="margin:5px 0;color:#1f2937;font-size:13.5px;line-height:1.6;">{_inline(s[2:])}</li>')
                elif _re.match(r'^\d+\.\s', s):
                    if in_ul: out.append('</ul>'); in_ul = False
                    if not in_ol: out.append('<ol style="margin:8px 0;padding-left:22px;">'); in_ol = True
                    ol_text = _inline(_re.sub(r'^\d+\.\s', '', s))
                    out.append(f'<li style="margin:5px 0;color:#1f2937;font-size:13.5px;line-height:1.6;">{ol_text}</li>')
                elif s.startswith('---') or s.startswith('***'):
                    close_lists()
                    out.append('<hr style="border:none;border-top:1px solid #d1fae5;margin:16px 0;" />')
                else:
                    close_lists()
                    out.append(f'<p style="margin:6px 0;color:#374151;font-size:13.5px;line-height:1.75;">{_inline(s)}</p>')

            if in_table: flush_table(table_rows, out)
            if in_ul: out.append('</ul>')
            if in_ol: out.append('</ol>')
            return '\n'.join(out)

        formatted_body = md_to_html(summary_content)
        now_str = datetime.utcnow().strftime('%B %d, %Y at %H:%M UTC')
        period_html = f'<span style="font-weight:600;color:#065f46;">{date_range}</span>' if date_range else '<span style="color:#6b7280;">All periods</span>'

        html_content = f"""\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Project Summary Report</title>
</head>
<body style="margin:0;padding:0;background:#f0fdf4;font-family:'Segoe UI',Arial,sans-serif;">

  <!-- Outer wrapper -->
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f0fdf4;padding:32px 16px;">
    <tr><td align="center">

      <!-- Card -->
      <table width="100%" cellpadding="0" cellspacing="0" style="max-width:680px;background:#ffffff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">

        <!-- ── Header ── -->
        <tr>
          <td style="background:linear-gradient(135deg,#064e3b 0%,#065f46 55%,#047857 100%);padding:28px 36px;">
            <table width="100%" cellpadding="0" cellspacing="0">
              <tr>
                <td>
                  <div style="display:inline-block;background:rgba(255,255,255,0.15);border:1px solid rgba(255,255,255,0.25);border-radius:10px;padding:6px 14px;margin-bottom:10px;">
                    <span style="color:#a7f3d0;font-size:11px;font-weight:700;letter-spacing:1px;text-transform:uppercase;">&#9679; Theta PMO &nbsp;|&nbsp; AI Project Advisor</span>
                  </div>
                  <h1 style="margin:0 0 6px;color:#ffffff;font-size:24px;font-weight:800;letter-spacing:-0.3px;">&#128202; Project Summary Report</h1>
                  <p style="margin:0;color:rgba(255,255,255,0.7);font-size:13px;">Automatically generated by your AI Project Advisor</p>
                </td>
              </tr>
            </table>
          </td>
        </tr>

        <!-- ── Meta strip ── -->
        <tr>
          <td style="background:#f0fdf4;border-bottom:1px solid #d1fae5;padding:14px 36px;">
            <table width="100%" cellpadding="0" cellspacing="0">
              <tr>
                <td style="font-size:12.5px;color:#6b7280;">
                  &#128197;&nbsp; <strong style="color:#065f46;">Analysis Period:</strong>&nbsp; {period_html}
                </td>
                <td align="right" style="font-size:12px;color:#9ca3af;">
                  Generated: {now_str}
                </td>
              </tr>
            </table>
          </td>
        </tr>

        <!-- ── Body ── -->
        <tr>
          <td style="padding:28px 36px 8px;">
            <div style="font-size:13.5px;color:#374151;line-height:1.75;">
              {formatted_body}
            </div>
          </td>
        </tr>

        <!-- ── Divider ── -->
        <tr>
          <td style="padding:8px 36px 0;"><hr style="border:none;border-top:1px solid #d1fae5;" /></td>
        </tr>

        <!-- ── Footer ── -->
        <tr>
          <td style="padding:20px 36px 28px;">
            <table width="100%" cellpadding="0" cellspacing="0">
              <tr>
                <td style="font-size:11.5px;color:#9ca3af;line-height:1.7;">
                  <strong style="color:#6b7280;">Theta PMO Platform</strong><br />
                  This report was automatically generated and delivered to <strong>{recipient_email}</strong>.<br />
                  Sent from <a href="mailto:anuragkatrerd@gmail.com" style="color:#059669;text-decoration:none;">anuragkatrerd@gmail.com</a> &nbsp;&#8212;&nbsp; Do not reply to this email.
                </td>
                <td align="right" valign="bottom">
                  <div style="background:linear-gradient(135deg,#064e3b,#047857);border-radius:8px;padding:8px 14px;display:inline-block;">
                    <span style="color:#a7f3d0;font-size:10px;font-weight:700;letter-spacing:0.5px;">THETA PMO</span>
                  </div>
                </td>
              </tr>
            </table>
          </td>
        </tr>

      </table>
      <!-- end card -->

    </td></tr>
  </table>

</body>
</html>"""

        html_part = MIMEText(html_content, 'html')
        msg.attach(text_part)
        msg.attach(html_part)

        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())

        print(f"[EMAIL] Summary sent from {sender_email} to {recipient_email}")
        return jsonify({'success': True, 'message': f'Summary emailed to {recipient_email}'})

    except smtplib.SMTPAuthenticationError:
        return jsonify({'error': 'Gmail authentication failed. Check GMAIL_APP_PASSWORD in .env'}), 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to send email: {str(e)}'}), 500


# ==================== /api/deviations — authenticated, cross-platform ====================

@app.route('/api/deviations', methods=['GET'])
@token_required
def api_get_deviations(current_user):
    """Get critical-path deviations. Admin/manager see all; user sees own."""
    try:
        role = current_user.get('role', 'user')
        # Only super_admin may override company_id via query param
        if role == 'super_admin':
            company_id = request.args.get('company_id') or current_user.get('company_id')
        else:
            company_id = current_user.get('company_id')
        user_id = None if role in ('admin', 'manager', 'super_admin', 'company_admin') else current_user['id']
        _source    = request.headers.get('X-App-Source', SOURCE_WEB)

        deviations = get_all_deviations(
            company_id=company_id or None,
            user_id=user_id,
        )

        cp_ids, cp_names = _load_cp_and_ms_ids(company_id)
        if cp_ids or cp_names:
            deviations = [d for d in deviations if _is_cp_deviation(d, cp_ids, cp_names)]

        # Track view activity
        try:
            log_activity(
                action_type = ACTION_DEVIATION_VIEW,
                user_id     = current_user['id'],
                user_name   = current_user.get('name', ''),
                user_role   = role,
                company_id  = current_user.get('company_id'),
                description = f"{current_user.get('name', 'User')} viewed CP deviations via {_source}",
                source      = _source,
                level       = LEVEL_MANAGER if role in ('admin', 'manager') else LEVEL_USER,
                metadata    = {'count': len(deviations)},
                ip_address  = request.remote_addr,
            )
        except Exception:
            pass

        return jsonify(deviations), 200
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/deviations/history', methods=['GET'])
@token_required
def api_get_deviations_history(current_user):
    """Get only reviewed critical-path deviations."""
    try:
        role = current_user.get('role', 'user')
        # Only super_admin may override company_id via query param
        if role == 'super_admin':
            company_id = request.args.get('company_id') or current_user.get('company_id')
        else:
            company_id = current_user.get('company_id')
        user_id = None if role in ('admin', 'manager', 'super_admin', 'company_admin') else current_user['id']
        limit      = request.args.get('limit', type=int)
        reviewed   = get_reviewed_deviations(
            company_id=company_id or None,
            user_id=user_id,
            limit=limit,
        )
        cp_ids, cp_names = _load_cp_and_ms_ids(company_id)
        if cp_ids or cp_names:
            reviewed = [d for d in reviewed if _is_cp_deviation(d, cp_ids, cp_names)]
        return jsonify(reviewed), 200
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/deviations', methods=['POST'])
@token_required
def api_create_deviation(current_user):
    """Create a deviation (auto-fills user_id/company_id from token)."""
    try:
        data = request.get_json() or {}
        _source = request.headers.get('X-App-Source', SOURCE_WEB)
        new_deviation = {
            'sheet'       : data.get('sheet', ''),
            'flag'        : data.get('flag', ''),
            'severity'    : data.get('severity', ''),
            'description' : data.get('description', ''),
            'row_data'    : data.get('row_data', {}),
            'detected_at' : data.get('detected_at', datetime.now().isoformat()),
            'review_status': 'Pending',
            'review_reason': '',
            'reason_type' : '',
            'user_id'     : current_user['id'],
            'company_id'  : current_user.get('company_id'),   # always from token, never caller
            'filename'    : data.get('filename', ''),
            'job_id'      : data.get('job_id', ''),
        }
        deviation_id = insert_deviation(new_deviation)
        new_deviation['id'] = deviation_id
        try:
            log_activity(
                action_type = 'deviation_create',
                user_id     = current_user['id'],
                user_name   = current_user.get('name', ''),
                user_role   = current_user.get('role', 'user'),
                company_id  = current_user.get('company_id'),
                entity_type = 'deviation', entity_id=str(deviation_id),
                description = f"{current_user.get('name','User')} created deviation #{deviation_id} via {_source}",
                source      = _source,
                level       = LEVEL_MANAGER,
                ip_address  = request.remote_addr,
            )
        except Exception:
            pass
        return jsonify({'status': 'success', 'deviation': new_deviation}), 201
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/deviations/update/<deviation_id>', methods=['POST', 'PUT', 'PATCH'])
@token_required
def api_update_deviation(current_user, deviation_id):
    """Update review status / reason on a deviation."""
    try:
        data = request.get_json() or {}
        _source = request.headers.get('X-App-Source', SOURCE_WEB)
        deviation = get_deviation_by_id(deviation_id)
        if not deviation:
            return jsonify({'error': 'Deviation not found'}), 404
        role = current_user.get('role')
        if role != 'super_admin' and deviation.get('company_id') != current_user.get('company_id'):
            return jsonify({'error': 'Deviation not found'}), 404
        fields = {
            'review_status'      : data.get('review_status', 'Reviewed'),
            'review_reason'      : data.get('review_reason', ''),
            'reason_type'        : data.get('reason_type', ''),
            'reviewed_at'        : datetime.now().isoformat(),
            'reviewed_by_user_id': data.get('user_id') or current_user['id'],
        }
        update_deviation_fields(deviation_id, fields)
        deviation.update(fields)
        try:
            log_activity(
                action_type = ACTION_DEVIATION_COMMENT,
                user_id     = current_user['id'],
                user_name   = current_user.get('name', ''),
                user_role   = current_user.get('role', 'user'),
                company_id  = current_user.get('company_id'),
                entity_type = 'deviation', entity_id=str(deviation_id),
                description = f"{current_user.get('name','User')} updated deviation #{deviation_id} via {_source}",
                source      = _source,
                level       = LEVEL_MANAGER,
                ip_address  = request.remote_addr,
            )
        except Exception:
            pass
        return jsonify({'status': 'success', 'deviation': deviation}), 200
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/deviations/admin/approve/<deviation_id>', methods=['POST'])
@token_required
def api_admin_approve_deviation(current_user, deviation_id):
    """Admin approves a deviation (authenticated)."""
    role = current_user.get('role')
    if role not in ('admin', 'manager', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Admin/manager access required'}), 403
    try:
        data = request.get_json() or {}
        _source = request.headers.get('X-App-Source', SOURCE_WEB)
        deviation = get_deviation_by_id(deviation_id)
        if not deviation:
            return jsonify({'error': 'Deviation not found'}), 404
        if role != 'super_admin' and deviation.get('company_id') != current_user.get('company_id'):
            return jsonify({'error': 'Deviation not found'}), 404
        fields = {
            'review_status'      : 'Approved',
            'admin_comment'      : data.get('reason', ''),
            'reviewed_at'        : datetime.now().isoformat(),
            'reviewed_by_user_id': data.get('user_id') or current_user['id'],
        }
        update_deviation_fields(deviation_id, fields)
        deviation.update(fields)
        try:
            log_activity(
                action_type = ACTION_DEVIATION_APPROVE,
                user_id     = current_user['id'],
                user_name   = current_user.get('name', ''),
                user_role   = current_user.get('role', 'admin'),
                company_id  = current_user.get('company_id'),
                entity_type = 'deviation', entity_id=str(deviation_id),
                description = f"{current_user.get('name','Admin')} approved deviation #{deviation_id} via {_source}",
                source      = _source,
                level       = LEVEL_ADMIN,
                metadata    = {
                    'deviation_id':  deviation_id,
                    'activity_name': (deviation.get('row_data') or {}).get('activity_name', '') if isinstance(deviation.get('row_data'), dict) else '',
                    'gate':          (deviation.get('row_data') or {}).get('stage_gate', '') if isinstance(deviation.get('row_data'), dict) else '',
                    'comment':       data.get('reason', ''),
                },
                ip_address  = request.remote_addr,
            )
        except Exception:
            pass
        try:
            if deviation.get('user_id'):
                create_notification(deviation['user_id'], 'Deviation Approved',
                    f'Deviation #{deviation_id} approved.', 'approval')
        except Exception:
            pass
        return jsonify({'status': 'success', 'deviation': deviation}), 200
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/deviations/admin/reject/<deviation_id>', methods=['POST'])
@token_required
def api_admin_reject_deviation(current_user, deviation_id):
    """Admin rejects a deviation (authenticated)."""
    role = current_user.get('role')
    if role not in ('admin', 'manager', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Admin/manager access required'}), 403
    try:
        data   = request.get_json() or {}
        reason = data.get('reason', '')
        if not reason.strip():
            return jsonify({'error': 'Reason is required for rejection'}), 400
        _source = request.headers.get('X-App-Source', SOURCE_WEB)
        deviation = get_deviation_by_id(deviation_id)
        if not deviation:
            return jsonify({'error': 'Deviation not found'}), 404
        if role != 'super_admin' and deviation.get('company_id') != current_user.get('company_id'):
            return jsonify({'error': 'Deviation not found'}), 404
        fields = {
            'review_status'      : 'Not Approved',
            'admin_comment'      : reason,
            'reviewed_at'        : datetime.now().isoformat(),
            'reviewed_by_user_id': data.get('user_id') or current_user['id'],
        }
        update_deviation_fields(deviation_id, fields)
        deviation.update(fields)
        try:
            log_activity(
                action_type = ACTION_DEVIATION_REJECT,
                user_id     = current_user['id'],
                user_name   = current_user.get('name', ''),
                user_role   = current_user.get('role', 'admin'),
                company_id  = current_user.get('company_id'),
                entity_type = 'deviation', entity_id=str(deviation_id),
                description = f"{current_user.get('name','Admin')} rejected deviation #{deviation_id} via {_source}: {reason}",
                source      = _source,
                level       = LEVEL_ADMIN,
                metadata    = {
                    'deviation_id':  deviation_id,
                    'activity_name': (deviation.get('row_data') or {}).get('activity_name', '') if isinstance(deviation.get('row_data'), dict) else '',
                    'gate':          (deviation.get('row_data') or {}).get('stage_gate', '') if isinstance(deviation.get('row_data'), dict) else '',
                    'reason':        reason,
                },
                ip_address  = request.remote_addr,
            )
        except Exception:
            pass
        try:
            if deviation.get('user_id'):
                create_notification(deviation['user_id'], 'Deviation Not Approved',
                    f'Deviation #{deviation_id} not approved. Reason: {reason}', 'rejection')
        except Exception:
            pass
        return jsonify({'status': 'success', 'deviation': deviation}), 200
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/deviations/<deviation_id>', methods=['DELETE'])
@token_required
def api_delete_deviation(current_user, deviation_id):
    """Delete a single deviation. Admin scoped to own company; manager/user can only delete own."""
    try:
        role    = current_user.get('role', 'user')
        _source = request.headers.get('X-App-Source', SOURCE_WEB)
        deviation = get_deviation_by_id(deviation_id)
        if not deviation:
            return jsonify({'error': 'Deviation not found'}), 404
        is_own = deviation.get('user_id') == current_user['id']
        is_company_admin = role in ('admin', 'company_admin') and deviation.get('company_id') == current_user.get('company_id')
        is_super = role == 'super_admin'
        if not (is_own or is_company_admin or is_super):
            return jsonify({'error': 'Not authorised to delete this deviation'}), 403
        deleted = delete_deviation(deviation_id)
        if not deleted:
            return jsonify({'error': 'Delete failed'}), 500
        try:
            log_activity(
                action_type = 'deviation_delete',
                user_id     = current_user['id'],
                user_name   = current_user.get('name', ''),
                user_role   = role,
                company_id  = current_user.get('company_id'),
                entity_type = 'deviation', entity_id=str(deviation_id),
                description = f"{current_user.get('name','User')} deleted deviation #{deviation_id} via {_source}",
                source      = _source,
                level       = LEVEL_ADMIN,
                ip_address  = request.remote_addr,
            )
        except Exception:
            pass
        return jsonify({'status': 'success', 'message': f'Deviation #{deviation_id} deleted'}), 200
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/deviations/clear-all', methods=['DELETE'])
@token_required
def api_clear_all_deviations(current_user):
    """Admin-only: delete ALL deviations scoped to the caller's company."""
    role = current_user.get('role')
    if role not in ('admin', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Admin access required'}), 403
    try:
        # Only super_admin may target a different company via query param
        if role == 'super_admin':
            company_id = request.args.get('company_id') or current_user.get('company_id')
        else:
            company_id = current_user.get('company_id')
        _source    = request.headers.get('X-App-Source', SOURCE_WEB)
        count = delete_all_deviations(company_id=company_id or None)
        try:
            log_activity(
                action_type = 'deviation_clear_all',
                user_id     = current_user['id'],
                user_name   = current_user.get('name', ''),
                user_role   = 'admin',
                company_id  = current_user.get('company_id'),
                description = f"{current_user.get('name','Admin')} cleared all deviations via {_source} ({count} rows)",
                source      = _source,
                level       = LEVEL_ADMIN,
                metadata    = {'deleted_count': count, 'company_id': company_id},
                ip_address  = request.remote_addr,
            )
        except Exception:
            pass
        return jsonify({'status': 'success', 'deleted_count': count}), 200
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# ── ADD THIS BLOCK immediately after the api_clear_all_deviations route ──
# Paste after line 6408 in app.py (before the ACTIVITY LOG ENDPOINTS section)

@app.route('/api/deviations/overview', methods=['GET'])
@token_required
def api_get_deviations_overview(current_user):
    """
    Returns a dashboard overview of all pending deviations for the History page:
      - detected_date      : ISO date string of the snapshot
      - total              : total pending deviation count
      - high / medium / low: severity counts
      - by_work_area       : [ { area, count, high, medium, low, worst_delay_days,
                                  worst_delay_label, tags, notes } ]
      - top_delays         : top-7 deviations sorted by days_overdue descending
                             [ { activity, stage, severity, days_overdue } ]
      - critical_alert     : { activity, days_overdue, message } | null
    """
    try:
        role = current_user.get('role', 'user')
        # Only super_admin may request a different company's overview via query param
        if role == 'super_admin':
            company_id = request.args.get('company_id') or current_user.get('company_id')
        else:
            company_id = current_user.get('company_id')
        user_id = None if role in ('admin', 'manager', 'company_admin', 'super_admin') else current_user['id']

        deviations = get_all_deviations(company_id=company_id or None, user_id=user_id)

        # ── Filter to critical-path and milestone activities only ────────────
        cp_ids, cp_names = _load_cp_and_ms_ids(company_id)
        if cp_ids or cp_names:
            deviations = [d for d in deviations if _is_cp_deviation(d, cp_ids, cp_names)]

        # ── Optional: filter to specific job IDs (e.g. most recent file for month) ─
        job_ids = request.args.getlist('job_id')
        if job_ids:
            job_ids_set = set(str(j) for j in job_ids)
            deviations = [d for d in deviations if str(d.get('job_id', '')) in job_ids_set]

        # ── Filter to pending only ──────────────────────────────────────────
        pending = [d for d in deviations if str(d.get('review_status', '')).lower() == 'pending']

        # ── Deduplicate: per job_id, keep only one row per unique activity+sheet combo.
        #    This prevents multi-run re-processing from stacking duplicate deviation rows.
        #    Strategy: among duplicates (same job_id + sheet + activity_name), keep the
        #    one with the latest detected_at (or highest id as tie-breaker).
        def _dedup_pending(devs):
            import json as _j
            devs_sorted = sorted(devs, key=lambda d: d.get('detected_at') or '', reverse=True)
            seen = set()
            result = []
            for d in devs_sorted:
                row = d.get('row_data') or {}
                if isinstance(row, str):
                    try:
                        row = _j.loads(row)
                    except Exception:
                        row = {}
                act = (
                    row.get('activity_id')
                    or row.get('activity_name')
                    or d.get('description', '')
                ).strip().lower()
                sheet = str(d.get('sheet') or '').strip().lower()
                job   = str(d.get('job_id') or '').strip()
                key   = (sheet, act)
                if key not in seen:
                    seen.add(key)
                    result.append(d)
            return result

        pending = _dedup_pending(pending)

        # ── Severity counts ────────────────────────────────────────────────
        total  = len(pending)
        high   = sum(1 for d in pending if str(d.get('severity', '')).lower() == 'high')
        medium = sum(1 for d in pending if str(d.get('severity', '')).lower() == 'medium')
        low    = sum(1 for d in pending if str(d.get('severity', '')).lower() == 'low')

        # ── Helper: extract days_overdue from a deviation ──────────────────
        def _days_overdue(dev):
            row = dev.get('row_data') or {}
            if isinstance(row, str):
                try:
                    import json as _json
                    row = _json.loads(row)
                except Exception:
                    row = {}

            def _to_int(val):
                if val is None:
                    return None
                try:
                    return int(float(str(val).replace('d', '').strip()))
                except Exception:
                    return None

            # Check all known field names for overdue days
            for key in ('overdue_days', 'days_overdue', 'delay_days', 'variance_days', 'delay_vs_plan_days'):
                v = _to_int(row.get(key))
                if v is not None and v > 0:
                    return v

            # deviation_calculator.py fields: start_delay, duration_deviation
            sd  = _to_int(row.get('start_delay'))
            dd  = _to_int(row.get('duration_deviation'))
            best = max(abs(sd) if sd is not None else 0,
                       abs(dd) if dd is not None else 0)
            if best > 0:
                return best

            # Parse from description field: "... Overdue: 28d past planned finish"
            desc = dev.get('description') or ''
            import re as _re
            m = _re.search(r'Overdue:\s*(\d+)d', str(desc))
            if m:
                return int(m.group(1))

            return 0

        def _stage(dev):
            row = dev.get('row_data') or {}
            if isinstance(row, str):
                try:
                    import json as _json
                    row = _json.loads(row)
                except Exception:
                    row = {}
            return (
                row.get('stage_gate')
                or row.get('stage')
                or dev.get('flag')
                or ''
            )

        def _activity(dev):
            row = dev.get('row_data') or {}
            if isinstance(row, str):
                try:
                    import json as _json
                    row = _json.loads(row)
                except Exception:
                    row = {}
            return (
                row.get('activity_name')
                or dev.get('description')
                or dev.get('sheet', '')
            )

        # ── All delays sorted by days, top 50 returned ────────────────────
        ranked = sorted(pending, key=_days_overdue, reverse=True)
        top_delays = []
        for dev in ranked[:50]:
            row = dev.get('row_data') or {}
            if isinstance(row, str):
                try:
                    import json as _j2
                    row = _j2.loads(row)
                except Exception:
                    row = {}
            days = _days_overdue(dev)
            if days == 0:
                continue  # skip zero-day rows — they have no meaningful delay data
            top_delays.append({
                'activity_id'  : row.get('activity_id', ''),
                'activity'     : row.get('activity_name') or _activity(dev),
                'stage'        : _stage(dev),
                'severity'     : dev.get('severity', ''),
                'days_overdue' : days,
                'total_float'  : row.get('total_float', ''),
                'planned_end'  : row.get('planned_end', '') or row.get('planned_finish', ''),
                'actual_end'   : row.get('actual_end', '') or row.get('actual_finish', ''),
                'planned_start': row.get('planned_start', ''),
                'actual_start' : row.get('actual_start', ''),
                'description'  : dev.get('description', ''),
                'filename'     : dev.get('filename', ''),
            })

        # ── Critical alert (most overdue single item) ──────────────────────
        critical_alert = None
        if ranked:
            worst = ranked[0]
            worst_days = _days_overdue(worst)
            if worst_days > 0:
                critical_alert = {
                    'activity'     : _activity(worst),
                    'days_overdue' : worst_days,
                    'message'      : (
                        f"{_activity(worst)} is critically overdue — "
                        f"{worst_days} days behind schedule. "
                        "Requires immediate management attention."
                    ),
                }

        # ── By-work-area breakdown ─────────────────────────────────────────
        area_map = {}
        for dev in pending:
            area = dev.get('sheet') or 'Unknown'
            if area not in area_map:
                area_map[area] = {
                    'area'              : area,
                    'count'             : 0,
                    'high'              : 0,
                    'medium'            : 0,
                    'low'               : 0,
                    'worst_delay_days'  : 0,
                    'worst_delay_label' : '',
                    '_stages'           : set(),
                    '_notes'            : set(),
                }
            entry = area_map[area]
            entry['count'] += 1
            sev = str(dev.get('severity', '')).lower()
            if sev == 'high':
                entry['high'] += 1
            elif sev == 'medium':
                entry['medium'] += 1
            elif sev == 'low':
                entry['low'] += 1

            days = _days_overdue(dev)
            if days > entry['worst_delay_days']:
                entry['worst_delay_days'] = days
                st = _stage(dev)
                entry['worst_delay_label'] = f"{days} days ({st})" if st else f"{days} days"

            st = _stage(dev)
            if st:
                entry['_stages'].add(st)

            # Collect short description snippets as notes
            desc = str(dev.get('description') or '').strip()
            if desc:
                entry['_notes'].add(desc[:60])

        by_work_area = []
        for entry in area_map.values():
            stages = sorted(entry.pop('_stages'))
            notes_set = entry.pop('_notes')
            entry['tags']  = stages[:4]            # up to 4 stage tags shown in UI
            entry['notes'] = ' · '.join(list(notes_set)[:3])  # brief summary text
            by_work_area.append(entry)

        # Sort: most deviations first
        by_work_area.sort(key=lambda x: x['count'], reverse=True)

        # ── Detected date: latest detected_at among pending ────────────────
        detected_date = None
        dates = [d.get('detected_at') for d in pending if d.get('detected_at')]
        if dates:
            detected_date = max(dates)

        return jsonify({
            'detected_date' : detected_date,
            'total'         : total,
            'high'          : high,
            'medium'        : medium,
            'low'           : low,
            'by_work_area'  : by_work_area,
            'top_delays'    : top_delays,
            'critical_alert': critical_alert,
        }), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# ==================== ACTIVITY LOG ENDPOINTS ====================

@app.route('/api/activity-log', methods=['POST'])
@token_required
def post_activity_log(current_user):
    """
    Mobile / web client pushes one activity event.
    Body: { action_type, entity_type?, entity_id?, description?, source?, level?, metadata? }
    """
    try:
        body = request.get_json() or {}
        action_type = body.get('action_type', 'unknown')
        source      = body.get('source', request.headers.get('X-App-Source', SOURCE_WEB))
        role        = current_user.get('role', 'user')
        level       = body.get('level', LEVEL_MANAGER if role in ('admin', 'manager') else LEVEL_USER)
        new_id = log_activity(
            action_type  = action_type,
            user_id      = current_user['id'],
            user_name    = current_user.get('name', ''),
            user_role    = role,
            company_id   = current_user.get('company_id'),
            entity_type  = body.get('entity_type'),
            entity_id    = body.get('entity_id'),
            description  = body.get('description'),
            source       = source,
            level        = level,
            metadata     = body.get('metadata'),
            ip_address   = request.remote_addr,
            session_id   = body.get('session_id'),
        )
        return jsonify({'success': True, 'id': new_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/activity-log', methods=['GET'])
@token_required
def get_activity_log(current_user):
    """
    Returns activities based on the caller's role:
      user         -> own records only
      manager      -> company-wide user+manager records
      admin /
      company_admin-> company-wide records (cannot cross org boundary)
      super_admin  -> everything (with optional filters)
    """
    try:
        role       = current_user.get('role', 'user')
        company_id = current_user.get('company_id')
        limit      = min(int(request.args.get('limit', 500)), 2000)

        if role == 'super_admin':
            filters = {
                k: request.args.get(k)
                for k in ('user_id', 'company_id', 'action_type', 'source', 'level', 'date_from', 'date_to')
                if request.args.get(k)
            }
            activities = get_all_activities_admin(limit=limit, filters=filters)
        elif role in ('admin', 'company_admin', 'manager'):
            # Scoped to caller's own company — cannot query another org
            activities = get_activities_for_manager(company_id=company_id, limit=limit)
        else:
            activities = get_activities_for_user(user_id=current_user['id'], limit=limit)

        return jsonify({'activities': activities, 'count': len(activities), 'role': role})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/activity-log/stats', methods=['GET'])
@token_required
def get_activity_stats_endpoint(current_user):
    """Statistics summary — scoped to own company; super_admin sees global or specific company."""
    try:
        role = current_user.get('role', 'user')
        if role == 'super_admin':
            company_id = request.args.get('company_id') or None
        else:
            company_id = current_user.get('company_id')
        stats = get_activity_stats(company_id=company_id)
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/api/knowledge-base/base-file', methods=['GET'])
@token_required
def kb_get_base_file(current_user):
    """Get current Knowledgebase base file used for always-on dashboard view."""
    try:
        cid = current_user.get('company_id')
        cfg = _ensure_default_base_file_registered(cid)
        kb_dir = _get_knowledgebase_folder(cid)
        fname = (cfg.get('filename') or '').strip()
        fpath = os.path.join(kb_dir, fname) if fname else ''
        exists = bool(fname and os.path.exists(fpath))
        return jsonify({
            'base_file': cfg,
            'exists': exists,
            'path': fpath if exists else None,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/base-file/download', methods=['GET'])
@token_required
def kb_download_base_file(current_user):
    """Download the current active base file."""
    try:
        file_path, _sheet_name, cfg = _resolve_base_file_path_and_sheet(current_user.get('company_id'))
        if not file_path or not os.path.exists(file_path):
            return jsonify({'error': 'Base file not configured or missing'}), 404

        dl_name = os.path.basename(file_path)
        return send_file(file_path, as_attachment=True, download_name=dl_name)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/base-file/view', methods=['GET'])
@token_required
def kb_view_base_file(current_user):
    """View the current active base file content for UI preview."""
    try:
        file_path, preferred_sheet, cfg = _resolve_base_file_path_and_sheet(current_user.get('company_id'))
        if not file_path or not os.path.exists(file_path):
            return jsonify({'error': 'Base file not configured or missing'}), 404

        requested_sheet = (request.args.get('sheet_name') or '').strip() or preferred_sheet
        try:
            max_rows = int(request.args.get('max_rows', 600))
        except Exception:
            max_rows = 600
        max_rows = min(max(max_rows, 50), 5000)
        payload = _read_excel_sheet_for_view(file_path, preferred_sheet=requested_sheet, max_rows=max_rows)

        return jsonify({
            'base_file': cfg,
            'filename': os.path.basename(file_path),
            **payload,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/base-file', methods=['POST'])
@token_required
def kb_set_base_file(current_user):
    """
    Set/replace base file from existing Knowledgebase file list.
    Body: { filename: str, sheet_name?: str }
    """
    if current_user.get('role') not in ('admin', 'manager'):
        return jsonify({'error': 'Admin or manager access required'}), 403

    try:
        data = request.get_json() or {}
        filename = (data.get('filename') or '').strip()
        sheet_name = (data.get('sheet_name') or 'Sheet2').strip() or 'Sheet2'
        if not filename:
            return jsonify({'error': 'filename is required'}), 400

        kb_dir = _get_knowledgebase_folder(current_user.get('company_id'))
        target = os.path.join(kb_dir, filename)
        if not os.path.exists(target):
            return jsonify({'error': f'Knowledgebase file not found: {filename}'}), 404

        cfg = {
            'filename': filename,
            'sheet_name': sheet_name,
            'is_active': True,
            'updated_at': datetime.now().isoformat(),
            'updated_by': current_user.get('id'),
        }
        write_base_file_config(cfg, current_user.get('company_id'))

        # Rebuild KB cache so AI + dashboard stay in sync
        reload_kb_files()

        # Pre-generate a fresh base analytics S-curve snapshot for history and fast access.
        requested_scurve_sheets = [
            'Project Management',
            'Manufacture',
            'HO-Subcontract',
            'HO-Procurements',
            'HO-As Builts',
            'EDDR',
            'Const & Pre-Comm',
            'Commissioning RFSU',
        ]
        try:
            _read_processed_scurve_sheets_for_dashboard(
                target,
                requested_scurve_sheets,
                force_reprocess=True,
                company_id=current_user.get('company_id'),
            )
        except Exception as pregen_err:
            print(f"[BASE ANALYTICS] pre-generate failed for {filename}: {pregen_err}")

        return jsonify({'message': 'Base file updated', 'base_file': cfg})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/files', methods=['GET'])
@token_required
def kb_list_files(current_user):
    """List all files in the company's Knowledgebase folder."""
    try:
        role = current_user.get('role', 'user')
        if role == 'super_admin':
            cid = request.args.get('company_id') or current_user.get('company_id')
        else:
            cid = current_user.get('company_id')

        if role not in ('admin', 'manager', 'company_admin', 'super_admin'):
            return jsonify({'error': 'Admin access required'}), 403

        kb_dir = _get_knowledgebase_folder(cid)
        files = []
        for fname in sorted(os.listdir(kb_dir)):
            fpath = os.path.join(kb_dir, fname)
            if not os.path.isfile(fpath) or fname.startswith('.') or fname.startswith('_'):
                continue
            stat = os.stat(fpath)
            files.append({
                'name': fname,
                'size_bytes': stat.st_size,
                'modified_at': datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })
        return jsonify({'files': files, 'company_id': cid, 'folder': kb_dir})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/files/upload', methods=['POST'])
@token_required
def kb_upload_file(current_user):
    """Upload a file to the company's Knowledgebase folder."""
    role = current_user.get('role', 'user')
    if role not in ('admin', 'manager', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Admin access required'}), 403
    try:
        if role == 'super_admin':
            cid = request.form.get('company_id') or current_user.get('company_id')
        else:
            cid = current_user.get('company_id')

        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'error': 'Empty filename'}), 400

        safe_name = os.path.basename(f.filename)
        kb_dir = _get_knowledgebase_folder(cid)
        dest = os.path.join(kb_dir, safe_name)
        f.save(dest)

        log_activity(
            action_type='kb_file_upload',
            user_id=current_user['id'],
            user_name=current_user.get('name', ''),
            user_role=role,
            company_id=cid,
            description=f"{current_user.get('name', 'Admin')} uploaded KB file: {safe_name}",
            level=LEVEL_ADMIN,
        )
        return jsonify({'message': f'{safe_name} uploaded', 'name': safe_name, 'company_id': cid}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/files/<path:filename>', methods=['DELETE'])
@token_required
def kb_delete_file(current_user, filename):
    """Delete a file from the company's Knowledgebase folder."""
    role = current_user.get('role', 'user')
    if role not in ('admin', 'manager', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Admin access required'}), 403
    try:
        if role == 'super_admin':
            cid = request.args.get('company_id') or current_user.get('company_id')
        else:
            cid = current_user.get('company_id')

        safe_name = os.path.basename(filename)
        kb_dir = _get_knowledgebase_folder(cid)
        fpath = os.path.join(kb_dir, safe_name)
        if not os.path.exists(fpath):
            return jsonify({'error': f'File not found: {safe_name}'}), 404
        os.remove(fpath)

        log_activity(
            action_type='kb_file_delete',
            user_id=current_user['id'],
            user_name=current_user.get('name', ''),
            user_role=role,
            company_id=cid,
            description=f"{current_user.get('name', 'Admin')} deleted KB file: {safe_name}",
            level=LEVEL_ADMIN,
        )
        return jsonify({'message': f'{safe_name} deleted', 'name': safe_name})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/base-file/versions', methods=['GET'])
@token_required
def kb_get_base_file_versions(current_user):
    """List stored base-file versions for audit tracking."""
    try:
        company_id = current_user.get('company_id')
        limit = min(max(int(request.args.get('limit', 100)), 1), 1000)
        versions = pg_read_base_file_versions(company_id)
        if not isinstance(versions, list):
            versions = []

        cfg = read_base_file_config(company_id)
        base_name = (cfg.get('filename') or '').strip()

        total_all = len(versions)
        filtered_versions = versions
        filter_applied = False
        fallback_to_all = False

        if base_name:
            base_name_norm = base_name.lower().strip()
            filtered_versions = [
                v for v in versions
                if str(v.get('base_filename') or '').lower().strip() == base_name_norm
            ]
            filter_applied = True

            # Safety fallback: if config name does not match historical metadata
            # exactly, still return all versions so user can access history.
            if not filtered_versions and versions:
                filtered_versions = versions
                fallback_to_all = True

        return jsonify({
            'base_file': cfg,
            'total': len(filtered_versions),
            'total_all': total_all,
            'filter_applied': filter_applied,
            'fallback_to_all': fallback_to_all,
            'versions': filtered_versions[:limit],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/base-file/versions/<version_id>/download', methods=['GET'])
@token_required
def kb_download_base_file_version(current_user, version_id):
    """Download a specific historical base-file snapshot by version_id."""
    try:
        versions = pg_read_base_file_versions(current_user.get('company_id'))
        if not isinstance(versions, list):
            versions = []

        item = next((v for v in versions if v.get('version_id') == version_id), None)
        if not item:
            return jsonify({'error': 'Version not found'}), 404

        abs_path = item.get('snapshot_abs_path')
        if not abs_path:
            rel = item.get('snapshot_rel_path')
            abs_path = os.path.join(_APP_ROOT, rel) if rel else None

        if not abs_path or not os.path.exists(abs_path):
            return jsonify({'error': 'Snapshot file missing on server'}), 404

        dl_name = os.path.basename(abs_path)
        return send_file(abs_path, as_attachment=True, download_name=dl_name)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/base-file/versions/<version_id>/view', methods=['GET'])
@token_required
def kb_view_base_file_version(current_user, version_id):
    """View a historical base-file snapshot content for UI preview."""
    try:
        company_id = current_user.get('company_id')
        versions = pg_read_base_file_versions(company_id)
        if not isinstance(versions, list):
            versions = []

        item = next((v for v in versions if v.get('version_id') == version_id), None)
        if not item:
            return jsonify({'error': 'Version not found'}), 404

        abs_path = item.get('snapshot_abs_path')
        if not abs_path:
            rel = item.get('snapshot_rel_path')
            abs_path = os.path.join(_APP_ROOT, rel) if rel else None

        if not abs_path or not os.path.exists(abs_path):
            return jsonify({'error': 'Snapshot file missing on server'}), 404

        cfg = read_base_file_config(company_id)
        default_sheet = (cfg.get('sheet_name') or 'Sheet2').strip() or 'Sheet2'
        requested_sheet = (request.args.get('sheet_name') or '').strip() or default_sheet
        try:
            max_rows = int(request.args.get('max_rows', 600))
        except Exception:
            max_rows = 600
        max_rows = min(max(max_rows, 50), 5000)
        payload = _read_excel_sheet_for_view(abs_path, preferred_sheet=requested_sheet, max_rows=max_rows)

        return jsonify({
            'version': item,
            'filename': os.path.basename(abs_path),
            **payload,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/base-file/versions/<version_id>', methods=['DELETE'])
@token_required
def kb_delete_base_file_version(current_user, version_id):
    """Delete a specific historical base-file snapshot and metadata entry."""
    if current_user.get('role') not in ('admin', 'manager'):
        return jsonify({'error': 'Admin or manager access required'}), 403

    try:
        company_id = current_user.get('company_id')
        versions = pg_read_base_file_versions(company_id)
        if not isinstance(versions, list):
            versions = []

        idx = next((i for i, v in enumerate(versions) if v.get('version_id') == version_id), None)
        if idx is None:
            return jsonify({'error': 'Version not found'}), 404

        item = versions[idx]
        abs_path = item.get('snapshot_abs_path')
        if not abs_path:
            rel = item.get('snapshot_rel_path')
            abs_path = os.path.join(_APP_ROOT, rel) if rel else None

        file_deleted = False
        if abs_path and os.path.exists(abs_path):
            try:
                os.remove(abs_path)
                file_deleted = True

                # Remove empty version folder if possible
                parent_dir = os.path.dirname(abs_path)
                if parent_dir and os.path.isdir(parent_dir) and not os.listdir(parent_dir):
                    os.rmdir(parent_dir)
            except Exception as file_err:
                return jsonify({'error': f'Failed to delete snapshot file: {file_err}'}), 500

        # Remove metadata entry regardless of file existence (cleans stale records)
        del versions[idx]
        pg_write_base_file_versions(versions, company_id)

        return jsonify({
            'message': 'Version deleted successfully',
            'version_id': version_id,
            'file_deleted': file_deleted,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard/base-analytics', methods=['GET'])
@token_required
def dashboard_base_analytics(current_user):
    """
    Always-on dashboard dataset from Knowledgebase base file.
    This does not depend on uploaded processing jobs.
    """
    try:
        cid = current_user.get('company_id')
        file_path, preferred_sheet, cfg = _resolve_base_file_path_and_sheet(cid)
        use_premerge_base = str(request.args.get('use_premerge_base', '0')).lower() in ('1', 'true', 'yes')
        include_all_kb_files = str(request.args.get('include_all_kb_files', '0')).lower() in ('1', 'true', 'yes')
        premerge_meta = None

        if use_premerge_base and file_path:
            premerge_path, premerge_meta = _resolve_premerge_base_snapshot_path(cfg.get('filename'), company_id=cid)
            if premerge_path:
                file_path = premerge_path
        if not file_path:
            return jsonify({
                'error': 'Base file not configured or not found in Knowledgebase',
                'base_file': cfg,
                'sheets': [],
            }), 404

        requested_scurve_sheets = [
            'Project Management',
            'Manufacture',
            'HO-Subcontract',
            'HO-Procurements',
            'HO-As Builts',
            'EDDR',
            'Const & Pre-Comm',
            'Commissioning RFSU',
        ]

        force_reprocess_requested = str(request.args.get('force_reprocess', '0')).lower() in ('1', 'true', 'yes')
        folder_changed, current_folder_sig, previous_folder_sig = _is_base_analytics_folder_changed(file_path, company_id=cid)
        # Analytics should rerun only when source folder changed.
        force_reprocess = bool(folder_changed)

        if include_all_kb_files:
            kb_files = get_kb_file_list()
            kb_dir = _get_knowledgebase_folder(cid)
            all_sheets = []
            all_sheet_names = []
            matched_requested = []
            missing_requested = []
            blank_requested = []
            requested_sheet_statuses = []
            error_messages = []
            source_files = []

            for item in kb_files:
                if not isinstance(item, dict):
                    continue
                filename = (item.get('filename') or '').strip()
                if not filename:
                    continue

                source_path = os.path.join(kb_dir, filename)
                if not os.path.exists(source_path):
                    continue

                source_files.append(filename)
                try:
                    bundle = _read_processed_scurve_sheets_for_dashboard(
                        source_path,
                        requested_scurve_sheets,
                        force_reprocess=bool(folder_changed),
                        company_id=cid,
                    )
                except Exception as src_err:
                    error_messages.append(f"{filename}: {src_err}")
                    continue

                source_label = os.path.splitext(filename)[0]
                for sheet in bundle.get('sheets', []):
                    if not isinstance(sheet, dict):
                        continue
                    merged_sheet = dict(sheet)
                    merged_sheet['source_file'] = filename
                    merged_sheet['sheet_name'] = f"{source_label} :: {sheet.get('sheet_name') or 'Sheet'}"
                    desc = merged_sheet.get('description') or ''
                    suffix = f"Source file: {filename}"
                    merged_sheet['description'] = f"{desc} | {suffix}" if desc else suffix
                    all_sheets.append(merged_sheet)

                all_sheet_names.extend([
                    f"{source_label} :: {name}"
                    for name in bundle.get('all_sheet_names_found', [])
                    if name
                ])
                matched_requested.extend(bundle.get('matched_requested_sheets', []))
                missing_requested.extend(bundle.get('missing_requested_sheets', []))
                blank_requested.extend(bundle.get('blank_requested_sheets', []))
                requested_sheet_statuses.extend(bundle.get('requested_sheet_statuses', []))
                if bundle.get('warnings'):
                    error_messages.extend(bundle.get('warnings'))

            if not all_sheets and file_path:
                processed_bundle = _read_processed_scurve_sheets_for_dashboard(
                    file_path,
                    requested_scurve_sheets,
                    force_reprocess=bool(folder_changed),
                    company_id=cid,
                )
                all_sheets = processed_bundle.get('sheets', [])
                all_sheet_names = processed_bundle.get('all_sheet_names_found', [])
                matched_requested = processed_bundle.get('matched_requested_sheets', [])
                missing_requested = processed_bundle.get('missing_requested_sheets', [])
                blank_requested = processed_bundle.get('blank_requested_sheets', [])
                requested_sheet_statuses = processed_bundle.get('requested_sheet_statuses', [])
                error_messages = processed_bundle.get('warnings', [])
            else:
                processed_bundle = {
                    'mode': 'knowledgebase_base_file_processed_all_files',
                    'cache_meta': {
                        'include_all_kb_files': True,
                        'source_files': source_files,
                    },
                }

            if not all_sheets:
                all_sheets = [_read_excel_sheet_for_dashboard(file_path, preferred_sheet=preferred_sheet)] if file_path else []

            return jsonify({
                'status': 'success',
                'mode': processed_bundle.get('mode', 'knowledgebase_base_file_processed_all_files'),
                'source_base_version': {
                    'use_premerge_base': use_premerge_base,
                    'version_id': (premerge_meta or {}).get('version_id') if premerge_meta else None,
                    'stage': (premerge_meta or {}).get('stage') if premerge_meta else None,
                    'created_at': (premerge_meta or {}).get('created_at') if premerge_meta else None,
                },
                'base_file': {
                    'filename': os.path.basename(file_path) if file_path else (cfg.get('filename') or ''),
                    'sheet_name': preferred_sheet,
                    'updated_at': cfg.get('updated_at'),
                },
                'cache_meta': processed_bundle.get('cache_meta', {}),
                'rerun_decision': {
                    'requested_force_reprocess': force_reprocess_requested,
                    'effective_force_reprocess': bool(folder_changed),
                    'source_folder_changed': folder_changed,
                    'current_source_folder_signature': current_folder_sig,
                    'previous_source_folder_signature': previous_folder_sig,
                    'include_all_kb_files': True,
                },
                'all_sheet_names_found': sorted(set(all_sheet_names)),
                'requested_scurve_sheets': requested_scurve_sheets,
                'matched_requested_sheets': matched_requested,
                'missing_requested_sheets': missing_requested,
                'blank_requested_sheets': blank_requested,
                'requested_sheet_statuses': requested_sheet_statuses,
                'error_summary': error_messages,
                'sheets': all_sheets,
                'count': len(all_sheets),
                'sources': source_files,
            })

        processed_bundle = _read_processed_scurve_sheets_for_dashboard(
            file_path,
            requested_scurve_sheets,
            force_reprocess=force_reprocess,
            company_id=cid,
        )

        sheets = processed_bundle.get('sheets', [])
        if not sheets:
            # Last-resort fallback to configured sheet if processing yields nothing.
            sheets = [_read_excel_sheet_for_dashboard(file_path, preferred_sheet=preferred_sheet)]

        all_sheet_names = processed_bundle.get('all_sheet_names_found', [])
        matched_requested = processed_bundle.get('matched_requested_sheets', [])
        missing_requested = processed_bundle.get('missing_requested_sheets', [])
        blank_requested = processed_bundle.get('blank_requested_sheets', [])
        requested_sheet_statuses = processed_bundle.get('requested_sheet_statuses', [])
        mode_used = processed_bundle.get('mode', 'knowledgebase_base_file_processed')

        error_messages = []
        if missing_requested:
            error_messages.append(f"Missing requested sheets: {', '.join(missing_requested)}")
        if blank_requested:
            error_messages.append(f"Blank requested sheets: {', '.join(blank_requested)}")
        if processed_bundle.get('warnings'):
            error_messages.extend(processed_bundle.get('warnings'))

        return jsonify({
            'status': 'success',
            'mode': mode_used,
            'source_base_version': {
                'use_premerge_base': use_premerge_base,
                'version_id': (premerge_meta or {}).get('version_id') if premerge_meta else None,
                'stage': (premerge_meta or {}).get('stage') if premerge_meta else None,
                'created_at': (premerge_meta or {}).get('created_at') if premerge_meta else None,
            },
            'base_file': {
                'filename': os.path.basename(file_path),
                'sheet_name': sheets[0].get('sheet_name') if sheets else preferred_sheet,
                'updated_at': cfg.get('updated_at'),
            },
            'cache_meta': processed_bundle.get('cache_meta', {}),
            'rerun_decision': {
                'requested_force_reprocess': force_reprocess_requested,
                'effective_force_reprocess': force_reprocess,
                'source_folder_changed': folder_changed,
                'current_source_folder_signature': current_folder_sig,
                'previous_source_folder_signature': previous_folder_sig,
            },
            'all_sheet_names_found': all_sheet_names,
            'requested_scurve_sheets': requested_scurve_sheets,
            'matched_requested_sheets': matched_requested,
            'missing_requested_sheets': missing_requested,
            'blank_requested_sheets': blank_requested,
            'requested_sheet_statuses': requested_sheet_statuses,
            'error_summary': error_messages,
            'sheets': sheets,
            'count': len(sheets),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/base-analytics/history', methods=['GET'])
@token_required
def kb_base_analytics_history(current_user):
    """List persisted base analytics S-curve generations stored on disk."""
    try:
        limit = min(max(int(request.args.get('limit', 50)), 1), 500)
        _cid = current_user.get('company_id')
        cfg = read_base_file_config(_cid)
        current_base_name = (cfg.get('filename') or '').strip().lower()
        _cache_ns = str(_cid) if _cid else '_global'
        _company_cache_dir = os.path.join(BASE_DASHBOARD_CACHE_DIR, _cache_ns)

        entries = []
        if os.path.exists(_company_cache_dir):
            for item in os.scandir(_company_cache_dir):
                if not item.is_dir():
                    continue
                bundle_path = os.path.join(item.path, 'bundle.json')
                if not os.path.exists(bundle_path):
                    continue

                try:
                    with open(bundle_path, 'r', encoding='utf-8') as f:
                        bundle = json.load(f)
                except Exception:
                    continue

                cache_meta = bundle.get('cache_meta', {}) if isinstance(bundle, dict) else {}
                base_name = (cache_meta.get('base_filename') or '').strip()

                entries.append({
                    'cache_token': item.name,
                    'stored_at': cache_meta.get('stored_at', ''),
                    'base_filename': base_name,
                    'base_file_path': cache_meta.get('base_file_path', ''),
                    'cache_dir': cache_meta.get('cache_dir', ''),
                    'stored_output_files': cache_meta.get('stored_output_files', []),
                    'sheet_count': len(bundle.get('sheets', [])) if isinstance(bundle, dict) else 0,
                    'mode': bundle.get('mode') if isinstance(bundle, dict) else None,
                    'matched_requested_sheets': bundle.get('matched_requested_sheets', []) if isinstance(bundle, dict) else [],
                    'missing_requested_sheets': bundle.get('missing_requested_sheets', []) if isinstance(bundle, dict) else [],
                    'is_current_base_file': bool(base_name and current_base_name and base_name.lower() == current_base_name),
                })

        entries.sort(key=lambda x: x.get('stored_at', ''), reverse=True)

        return jsonify({
            'base_file': cfg,
            'total': len(entries),
            'history': entries[:limit],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/base-analytics/history/<cache_token>', methods=['GET'])
@token_required
def kb_base_analytics_history_item(current_user, cache_token):
    """Get one persisted base analytics bundle by cache token."""
    try:
        if not _re.fullmatch(r'[a-f0-9]{8,64}', str(cache_token or '')):
            return jsonify({'error': 'Invalid cache token'}), 400

        _cid = current_user.get('company_id')
        _cache_ns = str(_cid) if _cid else '_global'
        bundle_path = os.path.join(BASE_DASHBOARD_CACHE_DIR, _cache_ns, cache_token, 'bundle.json')
        if not os.path.exists(bundle_path):
            return jsonify({'error': 'History item not found'}), 404

        with open(bundle_path, 'r', encoding='utf-8') as f:
            bundle = json.load(f)

        return jsonify(bundle)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/base-analytics/history/<cache_token>/download/<path:filename>', methods=['GET'])
@token_required
def kb_base_analytics_history_download(current_user, cache_token, filename):
    """Download a previously generated base analytics output file from history."""
    try:
        if not _re.fullmatch(r'[a-f0-9]{8,64}', str(cache_token or '')):
            return jsonify({'error': 'Invalid cache token'}), 400

        _cid = current_user.get('company_id')
        _cache_ns = str(_cid) if _cid else '_global'
        cache_dir = os.path.join(BASE_DASHBOARD_CACHE_DIR, _cache_ns, cache_token)
        if not os.path.exists(cache_dir):
            return jsonify({'error': 'History item not found'}), 404

        safe_name = os.path.basename(filename)
        file_path = os.path.join(cache_dir, safe_name)
        if not os.path.exists(file_path):
            return jsonify({'error': 'Stored output file not found'}), 404

        return send_file(file_path, as_attachment=True, download_name=safe_name)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── POST /api/ai/chat  (required by History.jsx AIAnalysisPanel) ──────────────
@app.route('/api/ai/chat', methods=['POST'])
@token_required
def ai_chat(current_user):
    body    = request.get_json(silent=True) or {}
    message = str(body.get('message', '')).strip()
    context = str(body.get('context', 'history_dashboard')).strip()

    if not message:
        return jsonify({'error': 'message is required'}), 400

    if azure_ai_client is None or not AZURE_ANTHROPIC_DEPLOYMENT:
        return jsonify({'error': 'AI service unavailable'}), 503

    try:
        msg_obj = azure_ai_client.messages.create(
            model=AZURE_ANTHROPIC_DEPLOYMENT,
            max_tokens=1024,
            system=(
                "You are a senior PMO analyst specialising in EPC project schedule analysis. "
                "Provide concise, actionable insights in under 400 words."
            ),
            messages=[{'role': 'user', 'content': message}],
        )
        text = '\n'.join(
            getattr(b, 'text', '') for b in (msg_obj.content or [])
            if getattr(b, 'type', '') == 'text'
        ).strip()

        log_chat(
            user_id=current_user['id'],
            user_name=current_user.get('name', ''),
            user_role=current_user.get('role', 'user'),
            route='/api/ai/chat',
            message=message,
            response=text,
            model=AZURE_ANTHROPIC_DEPLOYMENT or 'azure-anthropic',
            context_info=f"{context}|company={current_user.get('company_id', '')}",
        )

        return jsonify({'response': text, 'context': context})
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500


# ── GET /api/download/<job_id>/<sheet_name>  (path-param variant used by History.jsx) ──
@app.route('/api/download/<job_id>/<path:sheet_name>', methods=['GET'])
@token_required
def download_file_by_sheet(current_user, job_id, sheet_name):
    from urllib.parse import unquote
    sheet_name = unquote(sheet_name)

    try:
        entry = pg_get_history_entry(job_id, current_user['id'])
        if not entry:
            return jsonify({'error': 'File not found'}), 404

        result = next(
            (r for r in (entry.get('results') or []) if r.get('sheet_name') == sheet_name),
            None,
        )
        if not result or result.get('status') != 'success':
            return jsonify({'error': f'Sheet "{sheet_name}" not found or not successful'}), 404

        file_path = os.path.join(_APP_ROOT, OUTPUT_FOLDER, job_id, result['output_filename'])
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found on server'}), 404

        return send_file(file_path, as_attachment=True, download_name=result['output_filename'])

    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500
    

def generate_schedule_executive_analysis(metrics):

    total = metrics.get("total_deviations", 0)
    high = metrics.get("high_severity", 0)

    high_pct = round((high / total) * 100, 1) if total else 0

    risk_level = (
        "HIGH" if high_pct >= 70
        else "MEDIUM" if high_pct >= 40
        else "LOW"
    )

    top_sheet = metrics.get("top_sheet", {})
    top_sheet_name = top_sheet.get("name", "Unknown")
    top_sheet_total = top_sheet.get("count", 0)

    prompt = f"""
Generate an EPC executive schedule analysis.

Metrics:
- Total deviations: {total}
- High severity deviations: {high}
- High severity percentage: {high_pct}%
- Risk level: {risk_level}
- Most affected sheet: {top_sheet_name}
- Deviations in top sheet: {top_sheet_total}

Provide:
1. Overall schedule health
2. Critical risk interpretation
3. Executive observations
4. Recommended actions

Professional PMO reporting tone.
No markdown.
"""

    try:
        return _generate_claude_response(prompt)
    except Exception as e:
        print(e)

        return f"""
Project risk level is currently {risk_level}.

{high_pct}% of all deviations are classified as high severity.

Primary risk concentration is observed in {top_sheet_name}.
"""


@app.route('/api/executive-analysis', methods=['POST'])
@token_required
def executive_analysis(current_user):

    try:
        data = request.json or {}

        analysis = generate_schedule_executive_analysis(data)

        return jsonify({
            "success": True,
            "analysis": analysis
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500    
    
def _read_selected_scurve_sheets_for_dashboard(file_path):
    """Read all data sheets from the base file, preserving curated labels when they match."""
    targets = [
        'Project Management',
        'Manufacture',
        'HO-Subcontract',
        'HO-Procurements',
        'HO-As Builts',
        'EDDR',
        'Const & Pre-Comm',
        'Commissioning RFSU',
    ]
    target_aliases = {
        'Project Management': ['projectmanagement', 'projectmgmt', 'projectmgt', 'project_management', 'projectmangement'],
        'Manufacture': ['manufacture', 'manufacturing'],
        'HO-Subcontract': ['hosubcontract', 'hosubcontracts', 'ho_subcontract'],
        'HO-Procurements': ['hoprocurements', 'hoprocurement', 'ho_procurements'],
        'HO-As Builts': ['hoasbuilts', 'hoasbuilt', 'ho_as_builts', 'hoasbuilts', 'asabuilt', 'asbuilt', 'ho as builts', 'ho-asbuilts'],
        'EDDR': ['eddr'],
        'Const & Pre-Comm': ['constprecomm', 'constprecommissioning', 'constructionprecomm', 'constandprecomm', 'const pre comm', 'const precomm'],
        'Commissioning RFSU': ['commissioningrfsu', 'commissioning_rfsu', 'commissioning rfsu', 'rfsu', 'commissioning'],
    }

    target_keys = {
        label: [_normalize_sheet_key(a) for a in aliases]
        for label, aliases in target_aliases.items()
    }

    wb = None
    try:
        wb = _safe_load_workbook(file_path, data_only=True)
        worksheet_names = [ws.title for ws in wb.worksheets]
        if not worksheet_names:
            return []

        chosen = []
        for ws_name in worksheet_names:
            ws_key = _normalize_sheet_key(ws_name)
            for t_label, aliases in target_keys.items():
                if ws_key in aliases:
                    chosen.append((t_label, ws_name))
                    break

        # Fallback: partial contains matching for messy naming conventions.
        # This runs for unmatched targets as well (not only when chosen is empty).
        matched_targets = {t_label for t_label, _ in chosen}
        for target_label in targets:
            if target_label in matched_targets:
                continue

            for ws_name in worksheet_names:
                ws_key = _normalize_sheet_key(ws_name)
                aliases = target_keys.get(target_label, [])
                if any(alias in ws_key or ws_key in alias for alias in aliases):
                    chosen.append((target_label, ws_name))
                    matched_targets.add(target_label)
                    break

        # Preserve target order and remove duplicates by worksheet name.
        unique_ws = set()
        ordered = []
        for target_label in targets:
            for t_label, ws_name in chosen:
                if t_label == target_label and ws_name not in unique_ws:
                    unique_ws.add(ws_name)
                    ordered.append((t_label, ws_name))

        # Only the matched configured targets will be returned.
        # Fallback for remaining sheets was requested to be removed by user.

        payloads = []
        for target_label, ws_name in ordered:
            ws = wb[ws_name]
            header_row = _detect_best_header_row(ws)

            headers = []
            for cell in ws[header_row]:
                val = cell.value if cell.value is not None else ''
                headers.append(_excel_value_to_json(val))

            rows = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(v is not None and str(v).strip() != '' for v in row):
                    continue
                rows.append([_excel_value_to_json(v) for v in row])

            payloads.append({
                'sheet_name': ws_name,
                'target_sheet': target_label,
                'description': 'Knowledgebase Base File (all available data sheets)',
                'header_row': header_row,
                'headers': headers,
                'rows': rows,
                'row_count': len(rows),
            })

        return payloads
    finally:
        if wb:
            wb.close()


def _normalize_requested_scurve_label(label):
    aliases = {
        'Project Management': ['projectmanagement', 'projectmgmt', 'projectmgt', 'project_management', 'projectmangement'],
        'Manufacture': ['manufacture', 'manufacturing'],
        'HO-Subcontract': ['hosubcontract', 'hosubcontracts', 'ho_subcontract'],
        'HO-Procurements': ['hoprocurements', 'hoprocurement', 'ho_procurements'],
        'HO-As Builts': ['hoasbuilts', 'hoasbuilt', 'ho_as_builts', 'ho as builts', 'ho-asbuilts', 'asabuilt', 'asbuilt'],
        'EDDR': ['eddr'],
        'Const & Pre-Comm': ['constprecomm', 'constprecommissioning', 'constructionprecomm', 'constandprecomm', 'const pre comm', 'const precomm'],
        'Commissioning RFSU': ['commissioningrfsu', 'commissioning_rfsu', 'commissioning rfsu', 'commissioningrfsustartup', 'rfsu', 'commissioning'],
    }
    norm = _normalize_sheet_key(label)
    for canonical, values in aliases.items():
        all_values = [canonical] + values
        normalized_values = [_normalize_sheet_key(v) for v in all_values]
        if any(v == norm for v in normalized_values):
            return canonical
        # Also allow tracker-style labels like "Project Management (Timeline Deviation)".
        if any(v and norm.startswith(v) for v in normalized_values):
            return canonical
    return label


def _read_processed_scurve_sheets_for_dashboard(file_path, requested_scurve_sheets, force_reprocess=False, company_id=None):
    """Run tracker algorithms on base file and return analytics-ready sheet payloads for requested S-curve tabs."""
    cache_key = _build_base_file_signature(file_path)
    source_folder = os.path.dirname(os.path.abspath(file_path))
    source_folder_signature = _build_folder_signature(source_folder)

    cache_token = hashlib.sha256(cache_key.encode('utf-8')).hexdigest()[:32]
    _cache_ns = str(company_id) if company_id else '_global'
    persistent_cache_dir = os.path.join(BASE_DASHBOARD_CACHE_DIR, _cache_ns, cache_token)
    persistent_bundle_file = os.path.join(persistent_cache_dir, 'bundle.json')

    if not force_reprocess:
        with BASE_ANALYTICS_CACHE_LOCK:
            cached = BASE_ANALYTICS_CACHE.get(cache_key)
            if cached:
                return cached

        # Survive process restarts: load a persisted bundle if available.
        if os.path.exists(persistent_bundle_file):
            try:
                with open(persistent_bundle_file, 'r', encoding='utf-8') as f:
                    persisted_bundle = json.load(f)
                if isinstance(persisted_bundle, dict) and isinstance(persisted_bundle.get('sheets', []), list):
                    with BASE_ANALYTICS_CACHE_LOCK:
                        BASE_ANALYTICS_CACHE.clear()
                        BASE_ANALYTICS_CACHE[cache_key] = persisted_bundle
                    return persisted_bundle
            except Exception:
                pass

    bundle = {
        'mode': 'knowledgebase_base_file_processed',
        'sheets': [],
        'all_sheet_names_found': [],
        'matched_requested_sheets': [],
        'missing_requested_sheets': [],
        'blank_requested_sheets': [],
        'requested_sheet_statuses': [],
        'warnings': [],
    }

    def _fallback_to_raw_mode(extra_warning=None):
        if extra_warning:
            bundle['warnings'].append(extra_warning)
        bundle['mode'] = 'knowledgebase_base_file_raw_fallback'
        raw_sheets = _read_selected_scurve_sheets_for_dashboard(file_path)
        bundle['sheets'] = raw_sheets
        bundle['all_sheet_names_found'] = [s.get('sheet_name') for s in raw_sheets if isinstance(s, dict)]
        bundle['matched_requested_sheets'] = []
        bundle['missing_requested_sheets'] = []
        bundle['blank_requested_sheets'] = []
        bundle['requested_sheet_statuses'] = []

        for req in requested_scurve_sheets:
            canonical = _normalize_requested_scurve_label(req)
            payload = next((s for s in raw_sheets if _normalize_requested_scurve_label(s.get('sheet_name')) == canonical), None)
            if payload:
                bundle['matched_requested_sheets'].append(canonical)
                status = 'matched' if payload.get('row_count', 0) > 0 else 'blank'
                if status == 'blank':
                    bundle['blank_requested_sheets'].append(canonical)
                bundle['requested_sheet_statuses'].append({
                    'requested_sheet': canonical,
                    'matched_sheet': payload.get('sheet_name'),
                    'status': status,
                })
            else:
                bundle['missing_requested_sheets'].append(canonical)
                bundle['requested_sheet_statuses'].append({
                    'requested_sheet': canonical,
                    'matched_sheet': None,
                    'status': 'missing',
                })

    tmp_output = tempfile.mkdtemp(prefix='base_scurve_', dir=_TEMP_DIR)
    try:
        processing_input = file_path
        output_filenames_to_persist = []

        category_tab_aliases = {
            'Project Management': ['projectmanagementscurve', 'projectmanagements curve', 'projectmanagements-curve'],
            'Manufacture': ['manufacturescurve', 'manufacturingscurve'],
            'HO-Procurements': ['procurementscurve', 'procurementsscurve'],
            'HO-Subcontract': ['subcontractsscurve', 'subcontractscurve'],
            'HO-As Builts': ['hoasbuiltsscurve', 'hoasbuiltscurve', 'asbuiltsscurve', 'asbuiltscurve'],
            'Const & Pre-Comm': ['constructionscurve', 'constprecommscurve', 'constprecommscurve', 'constandprecommscurve', 'const pre comm s curve'],
            'Commissioning RFSU': ['commissioningscurve', 'commissioningsscurve'],
        }

        def _make_payload_from_ws(ws_obj, canonical_label, source_sheet_name, description):
            header_row = _detect_best_header_row(ws_obj)
            headers = [_excel_value_to_json(c.value if c.value is not None else '') for c in ws_obj[header_row]]

            rows = []
            for row in ws_obj.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(v is not None and str(v).strip() != '' for v in row):
                    continue
                rows.append([_excel_value_to_json(v) for v in row])

            return {
                'sheet_name': canonical_label,
                'target_sheet': canonical_label,
                'source_sheet': source_sheet_name,
                'description': description,
                'header_row': header_row,
                'headers': headers,
                'rows': rows,
                'row_count': len(rows),
            }

        # Some Borouge workbooks include chartsheet-defined names that break direct openpyxl
        # reads in tracker processors. Create a sanitized temporary copy for processing.
        try:
            sanitized_input = os.path.join(tmp_output, 'base_input_sanitized.xlsx')
            with zipfile.ZipFile(file_path, 'r') as zin:
                workbook_xml = zin.read('xl/workbook.xml').decode('utf-8')
                workbook_xml = _re.sub(
                    r'<definedNames[^>]*>.*?</definedNames>',
                    '',
                    workbook_xml,
                    flags=_re.DOTALL,
                )
                workbook_xml = _re.sub(r'<definedNames\s*/>', '', workbook_xml)

                with zipfile.ZipFile(sanitized_input, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        data = zin.read(item.filename)
                        if item.filename == 'xl/workbook.xml':
                            data = workbook_xml.encode('utf-8')
                        zout.writestr(item, data)

            processing_input = sanitized_input
        except Exception as sanitize_err:
            bundle['warnings'].append(f'Workbook sanitize step skipped: {sanitize_err}')

        processing_result = process_file(processing_input, tmp_output)
        results = processing_result.get('results', []) if isinstance(processing_result, dict) else []
        
        print(f"[BASE-DEBUG] process_file results count: {len(results)}")
        for r in results:
            print(f"[BASE-DEBUG] sheet={r.get('sheet_name')} status={r.get('status')} out={r.get('output_filename')} err={r.get('error','')}")

        requested_norm = {_normalize_requested_scurve_label(s): s for s in requested_scurve_sheets}
        payload_by_requested = {}
        failed_by_requested = {}
        all_names = []

        for result in results:
            if not isinstance(result, dict):
                continue
            sheet_name = str(result.get('sheet_name') or '').strip()
            if not sheet_name:
                continue

            canonical_name = _normalize_requested_scurve_label(sheet_name)
            if canonical_name not in requested_norm:
                continue

            status_text = str(result.get('status', '')).lower()
            if status_text != 'success':
                failed_by_requested[canonical_name] = {
                    'sheet_name': sheet_name,
                    'error': str(result.get('error') or result.get('message') or '').strip(),
                    'status': status_text or 'error',
                }
                continue

            out_name = result.get('output_filename')
            if not out_name:
                continue
            out_path = os.path.join(tmp_output, out_name)
            if not os.path.exists(out_path):
                continue
            output_filenames_to_persist.append(out_name)

            wb = None
            try:
                wb = _safe_load_workbook(out_path, data_only=True)
                ws = wb.active if wb.worksheets else None
                if ws is None:
                    continue

                payload = _make_payload_from_ws(
                    ws,
                    canonical_name,
                    sheet_name,
                    'Knowledgebase Base File (algorithm processed)',
                )
                payload_by_requested[canonical_name] = payload
                all_names.append(canonical_name)

                # EDDR output embeds category-level S-curve tabs (e.g., Project Management S-Curve).
                # Harvest them so missing canonical tracks can still render charts.
                for ws_extra in wb.worksheets:
                    ws_extra_key = _normalize_sheet_key(ws_extra.title)
                    for canonical_label, aliases in category_tab_aliases.items():
                        if canonical_label not in requested_norm:
                            continue
                        if canonical_label in payload_by_requested and payload_by_requested[canonical_label].get('row_count', 0) > 0:
                            continue

                        alias_hit = any(_normalize_sheet_key(alias) in ws_extra_key for alias in aliases)
                        if not alias_hit:
                            continue

                        payload_by_requested[canonical_label] = _make_payload_from_ws(
                            ws_extra,
                            canonical_label,
                            ws_extra.title,
                            'Knowledgebase Base File (algorithm processed category S-curve)',
                        )
                        all_names.append(canonical_label)

            finally:
                if wb:
                    wb.close()

        ordered = []
        for req in requested_scurve_sheets:
            canonical = _normalize_requested_scurve_label(req)
            payload = payload_by_requested.get(canonical)
            if payload:
                ordered.append(payload)
                bundle['matched_requested_sheets'].append(canonical)
                if payload.get('row_count', 0) <= 0:
                    bundle['blank_requested_sheets'].append(canonical)
                    bundle['requested_sheet_statuses'].append({
                        'requested_sheet': canonical,
                        'matched_sheet': canonical,
                        'status': 'blank',
                    })
                else:
                    bundle['requested_sheet_statuses'].append({
                        'requested_sheet': canonical,
                        'matched_sheet': canonical,
                        'status': 'matched',
                    })
            else:
                bundle['missing_requested_sheets'].append(canonical)
                bundle['requested_sheet_statuses'].append({
                    'requested_sheet': canonical,
                    'matched_sheet': None,
                    'status': 'missing',
                })

        bundle['sheets'] = ordered
        bundle['all_sheet_names_found'] = sorted(set(all_names))

        for requested_label, fail_meta in failed_by_requested.items():
            err = fail_meta.get('error')
            if err:
                bundle['warnings'].append(
                    f"Processed tracker failed for {requested_label}: {err}"
                )

        if str(processing_result.get('status', '')).lower() == 'error':
            bundle['warnings'].append(processing_result.get('message', 'Base file processing returned no successful trackers.'))

        # If some requested sheets are not available from processed outputs,
        # fill them from raw workbook tabs so all requested tracks can still render.
        if bundle['missing_requested_sheets']:
            raw_sheets = _read_selected_scurve_sheets_for_dashboard(file_path)
            raw_by_canonical = {}
            for raw in raw_sheets:
                if not isinstance(raw, dict):
                    continue
                canonical_raw = _normalize_requested_scurve_label(raw.get('sheet_name'))
                if canonical_raw not in raw_by_canonical:
                    raw_by_canonical[canonical_raw] = raw

            still_missing = []
            augmented = []
            for missing_label in bundle['missing_requested_sheets']:
                raw_payload = raw_by_canonical.get(missing_label)
                if not raw_payload:
                    still_missing.append(missing_label)
                    continue

                fill_payload = {
                    'sheet_name': missing_label,
                    'target_sheet': missing_label,
                    'source_sheet': raw_payload.get('sheet_name'),
                    'description': 'Knowledgebase Base File (raw supplement for missing processed sheet)',
                    'header_row': raw_payload.get('header_row'),
                    'headers': raw_payload.get('headers', []),
                    'rows': raw_payload.get('rows', []),
                    'row_count': raw_payload.get('row_count', 0),
                }

                failure_meta = failed_by_requested.get(missing_label) or {}
                failure_reason = str(failure_meta.get('error') or '').strip()
                if failure_reason:
                    fill_payload['description'] = (
                        'Knowledgebase Base File (raw supplement after processed tracker failure: '
                        f"{failure_reason})"
                    )
                    fill_payload['fallback_reason'] = failure_reason
                augmented.append(fill_payload)

                existing_status = next(
                    (s for s in bundle['requested_sheet_statuses'] if s.get('requested_sheet') == missing_label),
                    None,
                )
                if existing_status:
                    existing_status['matched_sheet'] = raw_payload.get('sheet_name')
                    existing_status['status'] = 'matched_raw' if fill_payload.get('row_count', 0) > 0 else 'blank'
                    if failure_reason:
                        existing_status['processed_error'] = failure_reason

                bundle['matched_requested_sheets'].append(missing_label)
                if fill_payload.get('row_count', 0) <= 0:
                    bundle['blank_requested_sheets'].append(missing_label)

            if augmented:
                # Preserve requested order in final payload.
                by_name = {s.get('sheet_name'): s for s in (ordered + augmented)}
                ordered_full = []
                for req in requested_scurve_sheets:
                    canonical = _normalize_requested_scurve_label(req)
                    payload = by_name.get(canonical)
                    if payload:
                        ordered_full.append(payload)
                bundle['sheets'] = ordered_full
                bundle['all_sheet_names_found'] = sorted(set(bundle['all_sheet_names_found'] + [s.get('sheet_name') for s in augmented]))

            bundle['missing_requested_sheets'] = still_missing

        if not ordered:
            _fallback_to_raw_mode('Processed base analytics returned no usable S-curve sheets; using raw workbook tabs.')

        # Persist bundle + generated files so base analytics history stays accessible.
        try:
            os.makedirs(persistent_cache_dir, exist_ok=True)
            persisted_files = []
            for out_name in sorted(set(output_filenames_to_persist)):
                src = os.path.join(tmp_output, out_name)
                if not os.path.exists(src):
                    continue
                dst = os.path.join(persistent_cache_dir, out_name)
                shutil.copy2(src, dst)
                persisted_files.append(out_name)

            bundle['cache_meta'] = {
                'cache_key': cache_key,
                'cache_token': cache_token,
                'company_id': str(company_id) if company_id else None,
                'base_filename': os.path.basename(file_path),
                'base_file_path': os.path.abspath(file_path),
                'source_folder': source_folder,
                'source_folder_signature': source_folder_signature,
                'cache_dir': os.path.relpath(persistent_cache_dir, _APP_ROOT).replace('\\\\', '/'),
                'stored_output_files': persisted_files,
                'stored_at': datetime.now().isoformat(),
            }

            with open(persistent_bundle_file, 'w', encoding='utf-8') as f:
                json.dump(bundle, f, ensure_ascii=False, indent=2)
        except Exception as persist_err:
            bundle['warnings'].append(f'Could not persist base analytics cache: {persist_err}')

    except Exception as exc:
        _fallback_to_raw_mode(f'Processed base analytics fallback used: {exc}')
    finally:
        try:
            shutil.rmtree(tmp_output, ignore_errors=True)
        except Exception:
            pass

    with BASE_ANALYTICS_CACHE_LOCK:
        BASE_ANALYTICS_CACHE.clear()
        BASE_ANALYTICS_CACHE[cache_key] = bundle

    return bundle

def _normalize_sheet_key(name):
    return _re.sub(r'[^a-z0-9]+', '', str(name or '').lower())


def _safe_load_workbook(file_path, **kwargs):
    """Load workbook with compatibility fallback for chart-sheet defined names."""
    import openpyxl
    try:
        return openpyxl.load_workbook(file_path, **kwargs)
    except AttributeError as load_err:
        msg = str(load_err)
        if 'Chartsheet' in msg and 'defined_names' in msg:
            from openpyxl.chartsheet.chartsheet import Chartsheet
            if not hasattr(Chartsheet, 'defined_names'):
                Chartsheet.defined_names = {}
            return openpyxl.load_workbook(file_path, **kwargs)
        raise


def _excel_value_to_json(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, time):
        return value.strftime('%H:%M:%S')
    if isinstance(value, timedelta):
        return str(value)
    return value


def _detect_best_header_row(ws, max_scan_rows=80):
    """Detect most likely data header row in messy Excel tabs."""
    scan_to = min(max_scan_rows, ws.max_row)
    if scan_to < 1:
        return 1

    signal_keywords = (
        'activity', 'date', 'period', 'month',
        'ep', 'lp', 'baseline', 'revised',
        'actual', 'cumm', 'cumulative',
        'early planning', 'late planning',
        'forecast', 'progress', 'weight', 'eddr'
    )

    strong_header_aliases = (
        'activity id', 'wbs code', 'activity code', 'activity name',
        'early start', 'early finish', 'late start', 'late finish',
        'actual date', 'actual start', 'actual finish',
        'ep', 'lp', 'cumm', 'cumulative', 'stage gate',
        'discipline', 'package', 'doc class', 'period', 'month'
    )

    best_row = 1
    best_score = -1
    for row_idx in range(1, scan_to + 1):
        values = []
        for cell in ws[row_idx]:
            txt = str(cell.value).strip() if cell.value is not None else ''
            values.append(txt)

        non_empty = [v for v in values if v]
        if len(non_empty) < 2:
            continue

        lowered = [v.lower() for v in non_empty]
        keyword_hits = sum(1 for v in lowered if any(k in v for k in signal_keywords))
        alpha_hits = sum(1 for v in non_empty if any(ch.isalpha() for ch in v))

        alias_hits = sum(1 for v in lowered if any(a in v for a in strong_header_aliases))
        activity_like = any(
            ('activity' in v and 'id' in v)
            or ('wbs' in v and 'code' in v)
            or ('activity' in v and 'name' in v)
            for v in lowered
        )
        date_like = sum(
            1
            for v in lowered
            if any(t in v for t in ('date', 'start', 'finish', 'early', 'late', 'period', 'month', 'actual', 'ep', 'lp'))
        )
        numeric_only = sum(
            1
            for v in non_empty
            if _re.fullmatch(r'[\d\.,%\-]+', v or '')
        )

        score = (
            (keyword_hits * 4)
            + (alias_hits * 6)
            + min(len(non_empty), 15)
            + alpha_hits
            + (10 if activity_like else 0)
            + min(date_like, 6)
            - (numeric_only * 2)
        )

        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


@app.route('/api/knowledge-base/reload', methods=['POST'])
@token_required
def kb_reload(current_user):
    """Re-scan and reload Knowledgebase Excel files."""
    if current_user.get('role') not in ('admin', 'manager'):
        return jsonify({'error': 'Admin or manager access required'}), 403
    try:
        reload_kb_files()
        files = get_kb_file_list()
        return jsonify({'message': 'Knowledge base files reloaded', 'files': files, 'count': len(files)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base', methods=['GET'])
@token_required
def knowledge_base(current_user):
    """
    Full Knowledge Base endpoint.
    - manager : company-scoped summary
    - admin   : global summary (AI access)
    - user    : own activity only
    """
    try:
        role       = current_user.get('role', 'user')
        company_id = current_user.get('company_id')
        limit      = min(int(request.args.get('limit', 300)), 1000)

        # Track KB view
        try:
            log_activity(
                action_type = ACTION_KB_VIEW,
                user_id     = current_user['id'],
                user_name   = current_user.get('name', ''),
                user_role   = role,
                company_id  = company_id,
                description = f"{current_user.get('name', 'User')} viewed Knowledge Base",
                source      = request.headers.get('X-App-Source', SOURCE_WEB),
                level       = LEVEL_MANAGER if role in ('admin', 'manager') else LEVEL_USER,
                ip_address  = request.remote_addr,
            )
        except Exception:
            pass

        if role == 'super_admin':
            summary = get_knowledge_base_summary(company_id=None, limit=limit)
        elif role in ('admin', 'manager', 'company_admin'):
            summary = get_knowledge_base_summary(company_id=company_id, limit=limit)
        else:
            # Regular users get just their own events
            own = get_activities_for_user(user_id=current_user['id'], limit=200)
            summary = {'recent_activities': own, 'stats': {}, 'active_users': []}

        # Enrich approve/reject records that are missing activity_name
        # (records logged before the metadata fix)
        _dev_types = {'deviation_approve', 'deviation_reject'}
        for act in summary.get('recent_activities', []):
            if act.get('action_type') in _dev_types:
                meta = act.get('metadata') or {}
                if not meta.get('activity_name') and act.get('entity_id'):
                    try:
                        dev = get_deviation_by_id(act['entity_id'])
                        if dev:
                            row = dev.get('row_data') or {}
                            act['metadata'] = {
                                **meta,
                                'deviation_id':  dev.get('id', act['entity_id']),
                                'activity_name': row.get('activity_name') or dev.get('description', ''),
                                'gate':          row.get('stage_gate', ''),
                            }
                    except Exception:
                        pass

        summary['viewer_role'] = role
        return jsonify(summary)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge-base/ai', methods=['GET'])
@token_required
def knowledge_base_ai(current_user):
    """
    Full-access activity endpoint scoped to the caller's company (super_admin gets all).
    """
    role = current_user.get('role', 'user')
    if role not in ('admin', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Admin access required'}), 403
    try:
        limit  = min(int(request.args.get('limit', 1000)), 5000)
        cid = None if role == 'super_admin' else current_user.get('company_id')
        filters = {
            k: request.args.get(k)
            for k in ('user_id', 'action_type', 'source', 'level', 'date_from', 'date_to')
            if request.args.get(k)
        }
        if cid:
            filters['company_id'] = cid
        activities = get_all_activities_admin(limit=limit, filters=filters)
        stats      = get_activity_stats(company_id=cid)
        users_db   = read_db(USERS_DB)
        user_index = {u['id']: {'name': u.get('name'), 'role': u.get('role'), 'email': u.get('email')} for u in users_db}
        return jsonify({
            'activities'    : activities,
            'stats'         : stats,
            'user_directory': user_index,
            'generated_at'  : datetime.now().isoformat(),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== AI LEARNING DATA ENDPOINT ====================

@app.route('/api/knowledge-base/ai-learning', methods=['GET'])
@token_required
def kb_ai_learning(current_user):
    """
    Return structured AI learning data — deviation patterns, manager decisions,
    AI chat history, and processing context. Used by the AI Module sub-tab in
    the Knowledge Base and injected into AI chatbot context for better responses.
    """
    try:
        limit = min(int(request.args.get('limit', 200)), 1000)
        company_id = current_user.get('company_id')
        role = current_user.get('role', 'user')

        # Fetch relevant activities — super_admin gets all, everyone else scoped to own company
        filters = {} if role == 'super_admin' else ({'company_id': company_id} if company_id else {})
        all_activities = get_all_activities_admin(limit=limit, filters=filters)

        # ── Deviation decision patterns (what AI learns from) ──
        approved = []
        for a in all_activities:
            if a.get('action_type') != 'deviation_approve':
                continue
            meta = a.get('metadata') or {}
            approved.append({
                'activity_name': meta.get('activity_name', 'Unknown'),
                'reason': meta.get('reason', meta.get('comment', '—')),
                'approved_by': a.get('user_name', '—'),
                'date': (a.get('timestamp') or '')[:10],
                'severity': meta.get('severity', '—'),
                'gate': meta.get('gate', ''),
            })

        rejected = []
        for a in all_activities:
            if a.get('action_type') != 'deviation_reject':
                continue
            meta = a.get('metadata') or {}
            rejected.append({
                'activity_name': meta.get('activity_name', 'Unknown'),
                'reason': meta.get('reason', meta.get('comment', '—')),
                'rejected_by': a.get('user_name', '—'),
                'date': (a.get('timestamp') or '')[:10],
                'severity': meta.get('severity', '—'),
            })

        comments = []
        for a in all_activities:
            if a.get('action_type') != 'deviation_comment':
                continue
            meta = a.get('metadata') or {}
            comments.append({
                'activity_name': meta.get('activity_name', 'Unknown'),
                'comment': meta.get('reason', meta.get('comment', a.get('description', '—'))),
                'by': a.get('user_name', '—'),
                'role': a.get('user_role', '—'),
                'date': (a.get('timestamp') or '')[:10],
            })

        # ── AI Chat History ──
        ai_chats = []
        for a in all_activities:
            if a.get('action_type') != 'ai_chat':
                continue
            meta = a.get('metadata') or {}
            ai_chats.append({
                'user': a.get('user_name', '—'),
                'prompt': meta.get('prompt', '')[:500],
                'response': meta.get('response', '')[:1000],
                'quick_action': meta.get('quick_prompt', ''),
                'date': (a.get('timestamp') or '')[:10],
            })

        # ── Processing History ──
        processing = []
        for a in all_activities:
            if a.get('action_type') != 'file_processed':
                continue
            meta = a.get('metadata') or {}
            processing.append({
                'filename': meta.get('filename', 'Unknown'),
                'success_count': meta.get('success_count', 0),
                'error_count': meta.get('error_count', 0),
                'processed_by': a.get('user_name', '—'),
                'date': (a.get('timestamp') or '')[:10],
            })

        return jsonify({
            'deviation_patterns': {
                'approved': approved,
                'rejected': rejected,
                'comments': comments,
            },
            'ai_chat_history': ai_chats,
            'processing_history': processing,
            'summary': {
                'total_approvals': len(approved),
                'total_rejections': len(rejected),
                'total_comments': len(comments),
                'total_ai_chats': len(ai_chats),
                'total_processed': len(processing),
            },
            'generated_at': datetime.now().isoformat(),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== S-CURVE & CHART ARCHIVE FOR KB ====================

@app.route('/api/knowledge-base/scurve-archive', methods=['GET'])
@token_required
def kb_scurve_archive(current_user):
    """
    Return list of processed files available for S-Curve chart viewing.
    The frontend uses the SCurveAnalytics component which fetches chart
    data via /api/analytics/<job_id> — we only need the file listing here.
    """
    try:
        role = current_user.get('role', 'user')
        company_id = current_user.get('company_id')
        if role in ('admin', 'company_admin', 'manager', 'super_admin'):
            entries = pg_read_history_for_company(company_id=company_id, limit=20)
        else:
            entries = pg_read_history_for_company(user_id=current_user['id'], limit=20)

        archive = []
        for entry in entries:
            job_id = entry.get('id')
            job_output_folder = os.path.join(_APP_ROOT, OUTPUT_FOLDER, job_id)
            if not os.path.exists(job_output_folder):
                continue
            archive.append({
                'job_id': job_id,
                'filename': entry.get('filename', 'Unknown'),
                'processed_at': entry.get('processed_at', ''),
                'success_count': entry.get('success_count', 0),
            })

        return jsonify({'archive': archive, 'total': len(archive)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ==================== RECURSIVE DEVIATION HISTORY FOR KB ====================

@app.route('/api/knowledge-base/recursive-deviations', methods=['GET'])
@token_required
def kb_recursive_deviations(current_user):
    """
    Return deviations that went through multiple reject → resubmit rounds
    before final approval, with full activity trail for each.
    """
    try:
        role = current_user.get('role', 'user')
        company_id = current_user.get('company_id')

        if role == 'super_admin':
            recursive_list = get_recursive_deviations(company_id=None)
        else:
            recursive_list = get_recursive_deviations(company_id=company_id)

        results = []
        for item in recursive_list:
            dev_id = item['entity_id']
            # Get deviation details
            deviation = get_deviation_by_id(dev_id)
            # Get full activity trail
            trail = get_deviation_activity_history(dev_id)

            results.append({
                'deviation_id': dev_id,
                'approve_count': item['approve_count'],
                'reject_count': item['reject_count'],
                'comment_count': item['comment_count'],
                'total_actions': item['total_actions'],
                'iterations': item['reject_count'] + item['approve_count'],
                'last_action': item['last_action'],
                'deviation': deviation,
                'activity_trail': trail,
            })

        return jsonify({'recursive_deviations': results, 'total': len(results)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# ==================== THETA ENGAGE ENDPOINTS ====================

@app.route('/api/engage/posts', methods=['GET'])
@token_required
def get_engage_posts(current_user):
    """Get Theta Engage posts visible to this user.
    - Posts are company-scoped: users only see posts from their own company.
    - super_admin sees all posts across companies (filterable by ?company_id=).
    - Group posts are additionally filtered to groups the user belongs to.
    - ?group_id=<id> filters to that group only.
    - ?user_id=<id> filters to posts by that user (visible ones only).
    """
    uid        = current_user['id']
    role       = current_user.get('role', '')
    company_id = current_user.get('company_id', '')

    # Load posts and groups scoped to the right company at the DB level
    is_super = role == 'super_admin'
    if is_super:
        filter_company = request.args.get('company_id') or None
        posts  = pg_read_engage_posts_for_company(filter_company)
        groups = pg_read_engage_groups_for_company(filter_company)
    else:
        posts  = pg_read_engage_posts_for_company(company_id)
        groups = pg_read_engage_groups_for_company(company_id)

    # Build set of group ids this user belongs to
    my_group_ids = {g['id'] for g in groups if uid in g.get('member_ids', [])}

    # Filter: public posts + group posts where user is member
    visible = []
    for p in posts:
        gid = p.get('group_id', '')
        if not gid:
            visible.append(p)
        elif gid in my_group_ids or is_super:
            visible.append(p)

    # Optional filters
    filter_group = request.args.get('group_id', '')
    filter_user = request.args.get('user_id', '')
    if filter_group:
        visible = [p for p in visible if p.get('group_id') == filter_group]
    if filter_user:
        visible = [p for p in visible if p.get('user_id') == filter_user]

    for p in visible:
        p['created_at'] = _to_ist_iso(p.get('created_at', ''))
        for like in p.get('likes', []):
            like['created_at'] = _to_ist_iso(like.get('created_at', ''))
        for comment in p.get('comments', []):
            comment['created_at'] = _to_ist_iso(comment.get('created_at', ''))

    visible.sort(key=lambda p: p.get('created_at', ''), reverse=True)
    return jsonify({'posts': visible}), 200


@app.route('/api/engage/posts', methods=['POST'])
@token_required
def create_engage_post(current_user):
    """Create a new Theta Engage post (optionally in a group)."""
    data = request.get_json()
    content = (data.get('content') or '').strip()
    image_url = data.get('image_url', '')
    if not content and not image_url:
        return jsonify({'error': 'Post content or image is required'}), 400

    group_id = data.get('group_id', '')
    # Validate group membership if posting to a group
    if group_id:
        grp = pg_get_engage_group_by_id(group_id)
        if not grp:
            return jsonify({'error': 'Group not found'}), 404
        if current_user['id'] not in grp.get('member_ids', []):
            return jsonify({'error': 'You are not a member of this group'}), 403

    post = {
        'id': str(uuid.uuid4()),
        'user_id': current_user['id'],
        'user_name': current_user.get('name', 'Unknown'),
        'user_email': current_user.get('email', ''),
        'company_id': current_user.get('company_id', ''),
        'content': content,
        'image_url': image_url,
        'group_id': group_id,
        'source': data.get('source', 'manual'),
        'likes': [],
        'comments': [],
        'created_at': _now_ist_iso(),
    }
    pg_insert_engage_post(post)
    return jsonify({'post': post}), 201


@app.route('/api/engage/posts/<post_id>', methods=['DELETE'])
@token_required
def delete_engage_post(current_user, post_id):
    """Delete own post."""
    post = pg_get_engage_post(post_id)
    if not post:
        return jsonify({'error': 'Post not found'}), 404
    role = current_user.get('role')
    is_owner = post['user_id'] == current_user['id']
    is_company_admin = role in ('admin', 'company_admin') and post.get('company_id') == current_user.get('company_id')
    is_super = role == 'super_admin'
    if not (is_owner or is_company_admin or is_super):
        return jsonify({'error': 'Not authorised'}), 403
    pg_delete_engage_post(post_id)
    return jsonify({'success': True}), 200


@app.route('/api/engage/posts/<post_id>/like', methods=['POST'])
@token_required
def toggle_engage_like(current_user, post_id):
    """Toggle like on a post."""
    post = pg_get_engage_post(post_id)
    if not post:
        return jsonify({'error': 'Post not found'}), 404
    if post.get('company_id') != current_user.get('company_id'):
        return jsonify({'error': 'Post not found'}), 404

    uid = current_user['id']
    likes = post.get('likes') or []
    existing = next((l for l in likes if l['user_id'] == uid), None)
    if existing:
        likes = [l for l in likes if l['user_id'] != uid]
    else:
        likes.append({
            'user_id': uid,
            'user_name': current_user.get('name', 'Unknown'),
            'created_at': _now_ist_iso(),
        })
    pg_update_engage_post_likes(post_id, likes)
    post['likes'] = likes
    return jsonify({'post': post}), 200


@app.route('/api/engage/posts/<post_id>/comments', methods=['POST'])
@token_required
def add_engage_comment(current_user, post_id):
    """Add a comment to a post."""
    post = pg_get_engage_post(post_id)
    if not post:
        return jsonify({'error': 'Post not found'}), 404
    if post.get('company_id') != current_user.get('company_id'):
        return jsonify({'error': 'Post not found'}), 404

    data = request.get_json()
    text = (data.get('content') or '').strip()
    if not text:
        return jsonify({'error': 'Comment content is required'}), 400

    comment = {
        'id': str(uuid.uuid4()),
        'user_id': current_user['id'],
        'user_name': current_user.get('name', 'Unknown'),
        'content': text,
        'created_at': _now_ist_iso(),
    }
    comments = post.get('comments') or []
    comments.append(comment)
    pg_update_engage_post_comments(post_id, comments)
    post['comments'] = comments
    return jsonify({'post': post}), 200


@app.route('/api/engage/posts/<post_id>/comments/<comment_id>', methods=['DELETE'])
@token_required
def delete_engage_comment(current_user, post_id, comment_id):
    """Delete own comment (or admin can delete any within their company)."""
    post = pg_get_engage_post(post_id)
    if not post:
        return jsonify({'error': 'Post not found'}), 404
    if post.get('company_id') != current_user.get('company_id') and current_user.get('role') != 'super_admin':
        return jsonify({'error': 'Post not found'}), 404
    comments = post.get('comments') or []
    comment = next((c for c in comments if c['id'] == comment_id), None)
    if not comment:
        return jsonify({'error': 'Comment not found'}), 404
    role = current_user.get('role')
    is_owner = comment['user_id'] == current_user['id']
    is_company_admin = role in ('admin', 'company_admin') and post.get('company_id') == current_user.get('company_id')
    is_super = role == 'super_admin'
    if not (is_owner or is_company_admin or is_super):
        return jsonify({'error': 'Not authorised'}), 403
    comments = [c for c in comments if c['id'] != comment_id]
    pg_update_engage_post_comments(post_id, comments)
    post['comments'] = comments
    return jsonify({'post': post}), 200


# ==================== ENGAGE GROUPS ====================

@app.route('/api/engage/groups', methods=['GET'])
@token_required
def get_engage_groups(current_user):
    """Get groups the current user belongs to, scoped to their company."""
    uid        = current_user['id']
    company_id = current_user.get('company_id', '')
    groups     = pg_read_engage_groups_for_company(company_id)
    my_groups  = [g for g in groups if uid in g.get('member_ids', [])]
    return jsonify({'groups': my_groups}), 200


@app.route('/api/engage/groups', methods=['POST'])
@token_required
def create_engage_group(current_user):
    """Create a new group. Creator is automatically a member."""
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Group name is required'}), 400

    member_ids = data.get('member_ids', [])
    uid = current_user['id']
    if uid not in member_ids:
        member_ids.insert(0, uid)

    # Resolve member names from users DB
    users = read_db(USERS_DB)
    user_map = {u['id']: u for u in users}
    members = []
    for mid in member_ids:
        u = user_map.get(mid)
        if u:
            members.append({'id': mid, 'name': u.get('name', 'Unknown'), 'email': u.get('email', '')})

    group = {
        'id': str(uuid.uuid4()),
        'name': name,
        'created_by': uid,
        'company_id': current_user.get('company_id'),
        'member_ids': member_ids,
        'members': members,
        'created_at': datetime.now().isoformat(),
    }
    pg_insert_engage_group(group)
    return jsonify({'group': group}), 201


@app.route('/api/engage/groups/<group_id>', methods=['DELETE'])
@token_required
def delete_engage_group(current_user, group_id):
    """Delete a group (only creator or company admin)."""
    grp = pg_get_engage_group_by_id(group_id)
    if not grp:
        return jsonify({'error': 'Group not found'}), 404
    role = current_user.get('role')
    is_owner = grp['created_by'] == current_user['id']
    is_company_admin = role in ('admin', 'company_admin') and grp.get('company_id') == current_user.get('company_id')
    is_super = role == 'super_admin'
    if not (is_owner or is_company_admin or is_super):
        return jsonify({'error': 'Not authorised'}), 403
    pg_delete_engage_group(group_id)
    return jsonify({'success': True}), 200


@app.route('/api/engage/users', methods=['GET'])
@token_required
def get_engage_users(current_user):
    """Get list of users in the same company (id, name, email) for group member picker."""
    users = read_db(USERS_DB)
    company_id = current_user.get('company_id')
    same_company = [u for u in users if u.get('company_id') == company_id and u.get('status') == 'approved']
    safe = [{'id': u['id'], 'name': u.get('name', 'Unknown'), 'email': u.get('email', '')} for u in same_company]
    return jsonify({'users': safe}), 200


# ==================== ENGAGE IMAGE UPLOAD ====================

ENGAGE_IMAGES_FOLDER = os.path.join('uploads', 'engage_images')
os.makedirs(ENGAGE_IMAGES_FOLDER, exist_ok=True)

ALLOWED_IMAGE_EXT = {'.png', '.jpg', '.jpeg', '.gif', '.webp'}


@app.route('/api/engage/upload-image', methods=['POST'])
@token_required
def upload_engage_image(current_user):
    """Upload an image for a Theta Engage post."""
    if 'image' not in request.files:
        return jsonify({'error': 'No image file provided'}), 400

    file = request.files['image']
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_IMAGE_EXT:
        return jsonify({'error': 'File type not allowed. Use PNG, JPG, GIF, or WebP.'}), 400

    # Limit to 5 MB
    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > 5 * 1024 * 1024:
        return jsonify({'error': 'Image must be under 5 MB'}), 400

    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(ENGAGE_IMAGES_FOLDER, filename)
    file.save(filepath)

    image_url = f"/api/engage/images/{filename}"
    return jsonify({'image_url': image_url}), 201


@app.route('/api/engage/images/<filename>', methods=['GET'])
def serve_engage_image(filename):
    """Serve uploaded engage images."""
    # Sanitise filename to prevent path traversal
    safe_name = os.path.basename(filename)
    return send_from_directory(ENGAGE_IMAGES_FOLDER, safe_name)

def _normalize_key(value: str) -> str:
    txt = str(value or '').strip().lower()
    txt = _re.sub(r'[^a-z0-9]+', '-', txt)
    return txt.strip('-') or 'item'


def _read_intelligence_cache(company_id=None):
    data = pg_read_insight_cache(company_id)
    return data if isinstance(data, list) else []


def _save_intelligence_cache(entries, company_id=None):
    pg_write_insight_cache(entries[:1200], company_id)


def _parse_iso_dt(value: str):
    if not value:
        return None
    try:
        txt = str(value).replace('Z', '')
        return datetime.fromisoformat(txt)
    except Exception:
        return None


def _cache_age_hours(created_at: str):
    dt = _parse_iso_dt(created_at)
    if not dt:
        return None
    return max(0.0, (datetime.now() - dt).total_seconds() / 3600.0)


def _is_cache_entry_stale(entry: dict, ttl_hours: int = None):
    ttl = ttl_hours or INTELLIGENCE_CACHE_TTL_HOURS
    age = _cache_age_hours(entry.get('created_at'))
    if age is None:
        return True, None
    return age >= ttl, age


def _get_cached_intelligence_insight(cache_key: str, data_hash: str, company_id=None):
    entries = _read_intelligence_cache(company_id)
    for item in entries:
        if item.get('cache_key') == cache_key and item.get('data_hash') == data_hash:
            is_stale, age_hours = _is_cache_entry_stale(item)
            return {
                'insight': item.get('insight'),
                'created_at': item.get('created_at'),
                'is_stale': is_stale,
                'age_hours': age_hours,
            }
    return None


def _upsert_cached_intelligence_insight(cache_key: str, data_hash: str, section: str, title: str, insight: dict, generated_by='system', company_id=None):
    entries = _read_intelligence_cache(company_id)
    entries = [e for e in entries if not (e.get('cache_key') == cache_key and e.get('data_hash') == data_hash)]
    entries.insert(0, {
        'cache_key': cache_key,
        'data_hash': data_hash,
        'section': section,
        'title': title,
        'insight': insight,
        'generated_by': generated_by,
        'created_at': datetime.now().isoformat(),
    })
    _save_intelligence_cache(entries, company_id)


def _extract_json_object(raw_text: str):
    if not raw_text:
        return None
    match = _re.search(r'\{.*\}', str(raw_text), _re.DOTALL)
    if not match:
        return None
    try:
        parsed = json.loads(match.group())
        if isinstance(parsed, dict):
            return parsed
    except Exception:
        return None
    return None


def _clean_insight_payload(payload: dict):
    if not isinstance(payload, dict):
        payload = {}
    return {
        'summary_headline': str(payload.get('summary_headline', '')).strip(),
        'urgency': str(payload.get('urgency', 'medium')).strip().lower() if payload.get('urgency') else 'medium',
        'present_analysis': str(payload.get('present_analysis', '')).strip(),
        'future_impact': str(payload.get('future_impact', '')).strip(),
        'recommendations': [str(x).strip() for x in (payload.get('recommendations') or []) if str(x).strip()][:6],
        'do_list': [str(x).strip() for x in (payload.get('do_list') or []) if str(x).strip()][:6],
        'dont_list': [str(x).strip() for x in (payload.get('dont_list') or []) if str(x).strip()][:6],
    }


def _to_int(value, default=0):
    try:
        if value is None:
            return int(default)
        return int(float(str(value).strip()))
    except Exception:
        return int(default)


def _extract_docx_plain_text(docx_path: str) -> str:
    """Extract visible text content from a .docx file without external dependencies."""
    try:
        with zipfile.ZipFile(docx_path, 'r') as zf:
            xml_bytes = zf.read('word/document.xml')
    except Exception:
        return ''

    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return ''

    chunks = []
    for elem in root.iter():
        if str(elem.tag).endswith('}t') and elem.text:
            text = str(elem.text).strip()
            if text:
                chunks.append(text)

    if not chunks:
        return ''

    return _re.sub(r'\s+', ' ', ' '.join(chunks)).strip()


def _load_whatif_knowledge_context(max_docs: int = 8, max_chars_per_doc: int = 3500, total_max_chars: int = 16000, company_id=None):
    """Load and truncate What-IF knowledge from Knowledgebase/whatifKnowledge/*.docx."""
    kb_dir = _get_knowledgebase_folder(company_id)
    folder = os.path.join(kb_dir, 'whatifKnowledge')
    if not os.path.isdir(folder):
        return {'context': '', 'sources': [], 'folder': folder}

    files = [f for f in sorted(os.listdir(folder)) if f.lower().endswith('.docx')]
    if not files:
        return {'context': '', 'sources': [], 'folder': folder}

    sources = []
    context_blocks = []
    chars_used = 0

    for filename in files[:max_docs]:
        file_path = os.path.join(folder, filename)
        raw_text = _extract_docx_plain_text(file_path)
        if not raw_text:
            continue

        remaining = max(0, total_max_chars - chars_used)
        if remaining <= 0:
            break

        excerpt_limit = min(max_chars_per_doc, remaining)
        excerpt = raw_text[:excerpt_limit].strip()
        if not excerpt:
            continue

        context_blocks.append(f"[Source: {filename}]\n{excerpt}")
        chars_used += len(excerpt)
        sources.append({
            'file': filename,
            'chars_loaded': len(excerpt),
        })

    return {
        'context': '\n\n'.join(context_blocks),
        'sources': sources,
        'folder': folder,
    }

def _load_whatif_chat_context(company_id=None) -> str:
    """
    Build a concise What-If context block for the general chatbot.
    Pulls from:
      - whatif_realtime_data       (latest realtime analysis)
      - whatif_critical_dashboard  (critical dashboard cache)
      - whatif_project_update_summary (project update summary)
      - Knowledgebase/whatifKnowledge/*.docx (word docs)
    Returns a plain-text string ready to inject into the chat prompt.
    """
    sections = []

    # 1. Realtime analysis JSON
    try:
        realtime = pg_read_whatif_realtime(company_id)
        if isinstance(realtime, dict) and realtime.get('kpis'):
            kpis = realtime['kpis']
            top_acts = realtime.get('top_delayed_activities', [])[:5]
            lines = [
                f"[What-If Realtime Analysis | Workbook: {realtime.get('source', {}).get('workbook', 'N/A')}]",
                f"Total Activities: {kpis.get('total_activities', 'N/A')} | "
                f"Delayed: {kpis.get('delayed_activities', 'N/A')} | "
                f"Avg Delay: {kpis.get('average_delay_days', 'N/A')} days | "
                f"Max Delay: {kpis.get('max_delay_days', 'N/A')} days",
            ]
            if top_acts:
                lines.append("Top Delayed Activities:")
                for a in top_acts:
                    lines.append(
                        f"  - {a.get('activity_id', '')} | {a.get('activity_name', '')} | "
                        f"Delay: {a.get('delay_days', '')} days | Issue: {a.get('cp_issue_type', '')} | "
                        f"Note: {a.get('risk_note', '')}"
                    )
            sections.append('\n'.join(lines))
    except Exception as e:
        print(f"[CHAT WHATIF] Realtime load error: {e}")

    # 2. Critical dashboard cache — use the stable alias entry
    try:
        critical_cache = pg_read_whatif_critical(company_id)
        if isinstance(critical_cache, dict):
            entry = critical_cache.get('whatif_critical_activities|overall') or next(
                (v for v in critical_cache.values() if isinstance(v, dict) and v.get('kpis')), None
            )
            if entry:
                kpis = entry.get('kpis', {})
                tracker = entry.get('threat_tracker', [])[:5]
                md = entry.get('summary_markdown', '').strip()
                lines = [
                    f"[What-If Critical Dashboard | Source: {entry.get('source', {}).get('file', 'N/A')}]",
                ]
                if md:
                    # strip markdown headers to keep it plain
                    import re as _re_wi
                    lines.append(_re_wi.sub(r'[#*`]', '', md)[:800].strip())
                if kpis:
                    lines.append(
                        f"KPIs — Total: {kpis.get('total_activities', 'N/A')} | "
                        f"Delayed: {kpis.get('delayed_activities', 'N/A')} | "
                        f"Max Delay: {kpis.get('max_delay_days', 'N/A')} days | "
                        f"Avg Delay: {kpis.get('average_delay_days', 'N/A')} days"
                    )
                if tracker:
                    lines.append("Critical Threat Tracker (top items):")
                    for t in tracker:
                        lines.append(
                            f"  - {t.get('id', '')} | {t.get('name', '')} | "
                            f"Baseline: {t.get('baseline', '')} | Actual: {t.get('actual', '')} | "
                            f"Late: {t.get('late', '')} days"
                        )
                sections.append('\n'.join(lines))
    except Exception as e:
        print(f"[CHAT WHATIF] Critical dashboard load error: {e}")

    # 3. Project update summary
    try:
        proj_cache = pg_read_whatif_proj_summary(company_id)
        if isinstance(proj_cache, dict):
            entry = proj_cache.get('project_update|overall') or next(
                (v for v in proj_cache.values() if isinstance(v, dict)), None
            )
            if entry:
                summary = entry.get('summary', {}) or {}
                lines = [
                    f"[What-If Project Update | File: {entry.get('source', {}).get('file', 'N/A')}]",
                    f"Total Updates: {summary.get('total_updates', 'N/A')} | "
                    f"Overwrites: {summary.get('overwrites', 'N/A')} | "
                    f"Injects: {summary.get('injects', 'N/A')} | "
                    f"High Impact Overwrites: {summary.get('high_impact_overwrites', 'N/A')} | "
                    f"Max Overwrite Delay: {summary.get('max_overwrite_delay_days', 'N/A')} days",
                ]
                sections.append('\n'.join(lines))
    except Exception as e:
        print(f"[CHAT WHATIF] Project update load error: {e}")

    # 4. Predecessor-successor acceleration scenarios
    try:
        pred_cache = pg_read_whatif_pred_succ(company_id)
        if isinstance(pred_cache, dict):
            entry = (
                pred_cache.get('predecessor_successor|overall') or
                next((v for v in pred_cache.values() if isinstance(v, dict) and v.get('dependencies')), None)
            )
            if entry:
                deps = entry.get('dependencies', [])
                analysis = entry.get('analysis', {})
                summary = entry.get('summary', {})
                lines = [
                    f"[What-If Predecessor-Successor Scenarios | File: {entry.get('source', {}).get('file', 'N/A')}]",
                    f"Total Edges: {summary.get('total_edges', len(deps))} | "
                    f"Critical Chains: {summary.get('critical_chains', 'N/A')} | "
                    f"Max Chain Length: {summary.get('max_chain_length', 'N/A')}",
                ]
                # Top bottleneck activities (most successors / highest blocking impact)
                bottlenecks = analysis.get('bottlenecks', [])[:8]
                if bottlenecks:
                    lines.append("Top Bottleneck Activities (accelerating these recovers the most time):")
                    for b in bottlenecks:
                        lines.append(
                            f"  - {b.get('id', '')} | {b.get('name', '')} | "
                            f"Successors: {b.get('successor_count', '')} | "
                            f"Recoverable: {b.get('recoverable_days', '')} days"
                        )
                # Sample of dependency edges so the AI can answer activity-ID queries
                sample_deps = deps[:30]
                if sample_deps:
                    lines.append("Sample Dependencies (Activity → Successor):")
                    for d in sample_deps:
                        lines.append(
                            f"  {d.get('predecessor_id', '')} [{d.get('predecessor_name', '')}] "
                            f"→ {d.get('successor_id', '')} [{d.get('successor_name', '')}] "
                            f"Lag: {d.get('lag', 0)}d Type: {d.get('type', 'FS')}"
                        )
                sections.append('\n'.join(lines))
    except Exception as e:
        print(f"[CHAT WHATIF] Pred-succ load error: {e}")

    # 5. whatifKnowledge Word docs (truncated)
    try:
        wi_knowledge = _load_whatif_knowledge_context(max_docs=4, max_chars_per_doc=1500, total_max_chars=5000, company_id=company_id)
        if wi_knowledge.get('context'):
            sections.append(
                "[What-If Knowledge Docs]\n" + wi_knowledge['context'][:5000]
            )
    except Exception as e:
        print(f"[CHAT WHATIF] Knowledge docs load error: {e}")

    if not sections:
        return ""

    return (
        "\n\n=== WHAT-IF ANALYSIS DATA ===\n"
        + "\n\n".join(sections)
        + "\n=== END WHAT-IF DATA ===\n"
    )

def _load_critical_path_json_context(max_rows: int = 220, max_chars: int = 18000, company_id=None):
    """Load and summarize Knowledgebase/Critical_path_analysis-EPC_Borouge.json for What-IF critical dashboard."""
    kb_dir = _get_knowledgebase_folder(company_id)
    filename = 'Critical_path_analysis-EPC_Borouge.json'
    file_path = os.path.join(kb_dir, filename)

    if not os.path.exists(file_path):
        return {
            'ok': False,
            'error': f'Knowledge JSON not found: {filename}',
            'file': filename,
            'path': file_path,
            'rows': [],
            'summary': {},
            'context': '',
            'signature': '',
        }

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            payload = json.load(f)
    except Exception as e:
        return {
            'ok': False,
            'error': f'Failed to parse JSON: {e}',
            'file': filename,
            'path': file_path,
            'rows': [],
            'summary': {},
            'context': '',
            'signature': '',
        }

    rows = []
    if isinstance(payload, dict):
        rows = payload.get('CP analaysis') or payload.get('CP analysis') or payload.get('data') or []
    elif isinstance(payload, list):
        rows = payload

    if not isinstance(rows, list):
        rows = []

    def _variance_days(row):
        return _to_int((row or {}).get('Variance - BL Project Finish Date', 0), 0)

    def _is_critical(row):
        lp = str((row or {}).get('Longest Path', '')).strip().lower()
        return lp in ('yes', 'y', 'true', '1')

    def _activity_id(row):
        return str((row or {}).get('Activity ID', '')).strip()

    def _activity_name(row):
        return str((row or {}).get('Activity Name', '')).strip()

    activity_rows = [r for r in rows if isinstance(r, dict) and (_activity_id(r) or _activity_name(r))]
    delayed_rows = [r for r in activity_rows if _variance_days(r) > 0]
    critical_rows = [r for r in activity_rows if _is_critical(r)]

    top_delay_rows = sorted(delayed_rows, key=lambda r: _variance_days(r), reverse=True)[:10]
    top_delay_lines = []
    threat_tracker = []

    for r in top_delay_rows:
        act_id = _activity_id(r)[:40]
        act_name = _activity_name(r)[:110]
        variance = _variance_days(r)
        total_float = _to_int(r.get('Total Float', 0), 0)
        top_delay_lines.append(
            f"- {act_id or 'N/A'} | {act_name or 'Unknown Activity'} | finish_var_days={variance} | total_float={total_float}"
        )
        threat_tracker.append({
            'id': act_id,
            'name': act_name or 'Unknown Activity',
            'baseline': 0,
            'actual': variance,
            'late': variance,
        })

    issue_distribution = {
        'Engineering': 0,
        'Procurement': 0,
        'Civil': 0,
        'Mechanical': 0,
        'Commissioning': 0,
        'Project Management': 0,
        'Other': 0,
    }

    for r in delayed_rows:
        text = f"{_activity_id(r)} {_activity_name(r)}".lower()
        if 'engineering' in text or 'eddr' in text:
            issue_distribution['Engineering'] += 1
        elif 'procure' in text or 'vendor' in text or 'subcontract' in text:
            issue_distribution['Procurement'] += 1
        elif 'civil' in text or 'concrete' in text or 'rebar' in text:
            issue_distribution['Civil'] += 1
        elif 'mechanical' in text or 'compressor' in text or 'pipe' in text:
            issue_distribution['Mechanical'] += 1
        elif 'commission' in text or 'rfsu' in text or 'start-up' in text:
            issue_distribution['Commissioning'] += 1
        elif 'project management' in text or 'pmm' in text:
            issue_distribution['Project Management'] += 1
        else:
            issue_distribution['Other'] += 1

    positive_delays = [_variance_days(r) for r in delayed_rows if _variance_days(r) > 0]
    max_delay_days = max(positive_delays) if positive_delays else 0
    avg_delay_days = round(sum(positive_delays) / len(positive_delays), 2) if positive_delays else 0

    stat = os.stat(file_path)
    signature = f"{file_path}|{stat.st_mtime_ns}|{stat.st_size}"

    summary = {
        'total_rows': len(rows),
        'activity_rows': len(activity_rows),
        'critical_rows': len(critical_rows),
        'delayed_rows': len(delayed_rows),
        'max_delay_days': max_delay_days,
        'average_delay_days': avg_delay_days,
        'issue_distribution': issue_distribution,
        'top_delay_count': len(top_delay_rows),
    }

    context_lines = [
        f"File: {filename}",
        f"Rows: {summary['total_rows']}",
        f"Activity rows: {summary['activity_rows']}",
        f"Critical (Longest Path=Yes): {summary['critical_rows']}",
        f"Delayed rows (Variance Finish > 0): {summary['delayed_rows']}",
        f"Max finish delay days: {summary['max_delay_days']}",
        f"Average finish delay days: {summary['average_delay_days']}",
        "Issue distribution over delayed rows:",
    ]
    for k, v in issue_distribution.items():
        context_lines.append(f"- {k}: {v}")

    context_lines.append('Top delayed activities (sample):')
    context_lines.extend(top_delay_lines)

    compact_rows = []
    for r in activity_rows[:max_rows]:
        compact_rows.append({
            'Activity ID': _activity_id(r),
            'Activity Name': _activity_name(r),
            'Variance - BL Project Finish Date': _variance_days(r),
            'Total Float': _to_int(r.get('Total Float', 0), 0),
            'Longest Path': str(r.get('Longest Path', '')).strip(),
        })

    compact_json = json.dumps(compact_rows, ensure_ascii=True)
    context_blob = '\n'.join(context_lines) + '\n\nCompact activity sample JSON:\n' + compact_json
    context_blob = context_blob[:max_chars]

    return {
        'ok': True,
        'error': '',
        'file': filename,
        'path': file_path,
        'rows': rows,
        'summary': summary,
        'threat_tracker_seed': threat_tracker,
        'context': context_blob,
        'signature': signature,
    }


def _build_critical_dashboard_prompt(knowledge_context: str):
    return f"""You are a senior EPC PMO What-IF analyst.
Use ONLY the provided critical path JSON knowledge context.

GOAL:
- Build a response suitable for a dashboard where content balance is:
  - 50% concise executive UI information
  - 50% chart-ready data

KNOWLEDGE CONTEXT:
{knowledge_context}

STRICT RULES:
1. Do not use external sources or assumptions.
2. If a value is not present, infer conservatively from the supplied context and keep values internally consistent.
3. Threat tracker items must represent delayed critical activities.
4. Provide at least 4 charts with concrete data arrays.
5. Return ONLY one valid JSON object. No markdown fences.

Return JSON in exactly this shape:
{{
  "summary_markdown": "<short executive summary>",
  "ui_payload": {{
    "save_days": 0,
    "opportunity_days": 0,
    "residual_days": 0,
    "recovery_percent": 0,
    "core_threat_title": "<title>",
    "core_threat_summary": "<summary>",
    "step1_title": "<title>",
    "step1_summary": "<summary>",
    "step1_baseline_days": [40,47,46],
    "step1_compressed_days": [16,16,16],
    "step1_activity_chips": ["A1000","A1001","A1002"],
    "step1_recovered_days": 0,
    "step1_residual_to_sequence_days": 0,
    "kpis": {{
      "total_activities": 0,
      "delayed_activities": 0,
      "max_delay_days": 0,
      "average_delay_days": 0
    }},
    "threat_tracker": [
      {{"id":"A1000","name":"Activity","baseline":0,"actual":0,"late":0}}
    ],
    "compressor_savings": [
      {{"step":"Parallel engineering-procurement release","days":"+0"}},
      {{"step":"Civil/mechanical overlap","days":"+0"}},
      {{"step":"Commissioning prep pull-forward","days":"+0"}}
    ],
    "recommendations": ["<action 1>","<action 2>","<action 3>"]
  }},
  "chart_plan": [
    {{"id":"issue_distribution","type":"pie","title":"Critical Delay Issue Mix","data":[{{"name":"Engineering","value":0}}]}},
    {{"id":"top_delays","type":"bar","title":"Top Delayed Activities","data":[{{"name":"A1000","value":0}}]}},
    {{"id":"float_vs_delay","type":"scatter","title":"Float vs Delay Risk","data":[{{"name":"A1000","x":0,"y":0}}]}},
    {{"id":"recovery_waterfall","type":"waterfall","title":"Recovery Opportunity Waterfall","data":[{{"name":"Step","value":0}}]}}
  ],
  "endpoint_map": [
    {{"name":"Critical Dashboard","method":"GET","path":"/api/whatif/critical-dashboard","purpose":"Critical What-IF snapshot from JSON knowledge"}}
  ]
}}"""


def _parse_dependency_detail_tokens(raw_text: str):
    """Parse detail strings like 'A60780: FS 14, A18880: SS' into structured relation tokens."""
    text = str(raw_text or '').strip()
    if not text:
        return []

    tokens = []
    for part in text.split(','):
        item = str(part).strip()
        if not item or ':' not in item:
            continue

        left, right = item.split(':', 1)
        rel_id = str(left).strip()
        right_text = str(right).strip()
        if not rel_id or not right_text:
            continue

        logic_match = _re.search(r'\b(FS|SS|FF|SF)\b', right_text.upper())
        logic_type = logic_match.group(1) if logic_match else 'FS'

        lag_match = _re.search(r'(-?\d+)', right_text)
        lag_days = int(lag_match.group(1)) if lag_match else 0

        tokens.append({
            'id': rel_id,
            'logic_type': logic_type,
            'lag_days': lag_days,
        })

    return tokens


def _load_predecessor_successor_json_context(max_edges: int = 1200, company_id=None):
    """Load dependency edges from Knowledgebase/PREDECESSOR-SUCCESSOR-LAG.json."""
    kb_dir = _get_knowledgebase_folder(company_id)
    filename = 'PREDECESSOR-SUCCESSOR-LAG.json'
    file_path = os.path.join(kb_dir, filename)

    if not os.path.exists(file_path):
        return {
            'ok': False,
            'error': f'Knowledge JSON not found: {filename}',
            'file': filename,
            'path': file_path,
            'dependencies': [],
            'summary': {},
            'signature': '',
        }

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            payload = json.load(f)
    except Exception as e:
        return {
            'ok': False,
            'error': f'Failed to parse JSON: {e}',
            'file': filename,
            'path': file_path,
            'dependencies': [],
            'summary': {},
            'signature': '',
        }

    rows = payload.get('P6') if isinstance(payload, dict) else payload
    if not isinstance(rows, list):
        rows = []

    def _clean_id(value):
        return str(value or '').strip()

    def _is_activity_id(value):
        text = _clean_id(value)
        if not text:
            return False
        return bool(_re.match(r'^[A-Za-z0-9][A-Za-z0-9\-]*$', text))

    activity_names = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        act_id = _clean_id(row.get('Activity ID'))
        if not _is_activity_id(act_id):
            continue
        act_name = str(row.get('Activity Name') or '').strip()
        if act_name:
            activity_names[act_id] = act_name

    edges = []
    dedupe = set()
    for row in rows:
        if not isinstance(row, dict):
            continue

        src_id = _clean_id(row.get('Activity ID'))
        if not _is_activity_id(src_id):
            continue
        src_name = activity_names.get(src_id, str(row.get('Activity Name') or '').strip())

        successor_tokens = _parse_dependency_detail_tokens(row.get('Successor Details'))
        predecessor_tokens = _parse_dependency_detail_tokens(row.get('Predecessor Details'))

        for t in successor_tokens:
            to_id = _clean_id(t.get('id'))
            if not _is_activity_id(to_id):
                continue
            key = (src_id, to_id, t.get('logic_type', 'FS'), int(t.get('lag_days', 0)))
            if key in dedupe:
                continue
            dedupe.add(key)
            edges.append({
                'from_id': src_id,
                'from_name': src_name,
                'to_id': to_id,
                'to_name': activity_names.get(to_id, ''),
                'logic_type': t.get('logic_type', 'FS'),
                'lag_days': int(t.get('lag_days', 0)),
            })

        for t in predecessor_tokens:
            pred_id = _clean_id(t.get('id'))
            if not _is_activity_id(pred_id):
                continue
            key = (pred_id, src_id, t.get('logic_type', 'FS'), int(t.get('lag_days', 0)))
            if key in dedupe:
                continue
            dedupe.add(key)
            edges.append({
                'from_id': pred_id,
                'from_name': activity_names.get(pred_id, ''),
                'to_id': src_id,
                'to_name': src_name,
                'logic_type': t.get('logic_type', 'FS'),
                'lag_days': int(t.get('lag_days', 0)),
            })

    outgoing = {}
    incoming = {}
    lag_sum_by_pred = {}
    logic_counts = {'FS': 0, 'SS': 0, 'FF': 0, 'SF': 0}
    for edge in edges:
        from_id = edge.get('from_id', '')
        to_id = edge.get('to_id', '')
        lag_days = int(edge.get('lag_days', 0) or 0)
        logic_type = str(edge.get('logic_type', 'FS')).upper()

        outgoing[from_id] = outgoing.get(from_id, 0) + 1
        incoming[to_id] = incoming.get(to_id, 0) + 1
        lag_sum_by_pred[from_id] = lag_sum_by_pred.get(from_id, 0) + lag_days
        logic_counts[logic_type] = logic_counts.get(logic_type, 0) + 1

    # Hub nodes + high-lag links approximate dependency links with the highest schedule impact.
    top_hubs = {k for k, _v in sorted(outgoing.items(), key=lambda kv: kv[1], reverse=True)[:20]}
    for edge in edges:
        lag_days = int(edge.get('lag_days', 0) or 0)
        edge['on_critical_path'] = bool(
            lag_days >= 45
            or edge.get('from_id') in top_hubs
            or edge.get('to_id') in top_hubs
        )

    edges.sort(key=lambda r: (int(r.get('lag_days', 0) or 0), int(r.get('on_critical_path', False))), reverse=True)
    dependencies = edges[:max_edges]

    stat = os.stat(file_path)
    signature = f"{file_path}|{stat.st_mtime_ns}|{stat.st_size}|{len(rows)}"

    summary = {
        'total_rows': len(rows),
        'activity_nodes': len(activity_names),
        'total_dependencies': len(edges),
        'critical_dependencies': sum(1 for e in edges if e.get('on_critical_path')),
        'logic_counts': logic_counts,
        'avg_lag_days': round((sum(int(e.get('lag_days', 0) or 0) for e in edges) / len(edges)), 2) if edges else 0,
    }

    top_outgoing = sorted(outgoing.items(), key=lambda kv: kv[1], reverse=True)[:8]
    top_incoming = sorted(incoming.items(), key=lambda kv: kv[1], reverse=True)[:8]
    lag_hotspots = sorted(lag_sum_by_pred.items(), key=lambda kv: kv[1], reverse=True)[:8]

    chart_plan = [
        {
            'id': 'logic_type_mix',
            'type': 'bar',
            'title': 'Dependency Logic Mix (FS/SS/FF/SF)',
            'data': [
                {'name': rel, 'value': int(count)} for rel, count in logic_counts.items()
            ],
        },
        {
            'id': 'predecessor_hubs',
            'type': 'line',
            'title': 'Top Predecessor Hubs by Outgoing Links',
            'data': [
                {'name': node_id, 'value': int(count)} for node_id, count in top_outgoing
            ],
        },
        {
            'id': 'successor_load',
            'type': 'area',
            'title': 'Top Successor Load by Incoming Links',
            'data': [
                {'name': node_id, 'value': int(count)} for node_id, count in top_incoming
            ],
        },
        {
            'id': 'lag_hotspots',
            'type': 'bar',
            'title': 'Lag Hotspots by Predecessor Chain',
            'data': [
                {'name': node_id, 'value': int(total_lag)} for node_id, total_lag in lag_hotspots
            ],
        },
    ]

    total_time_saving_days = 0
    for rel_type, ratio, cap in (('FS', 0.25, 35), ('SS', 0.10, 12), ('FF', 0.05, 8)):
        total_time_saving_days += min(cap, int(round(logic_counts.get(rel_type, 0) * ratio)))

    analysis = {
        'executive_summary': (
            f"Dependency map loaded from {filename}: {summary['total_dependencies']} logic links across "
            f"{summary['activity_nodes']} activity nodes. Focus is on high fan-out predecessors and high-lag handoffs."
        ),
        'critical_path_review': (
            f"{summary['critical_dependencies']} links are marked as high-impact based on lag and dependency hub concentration. "
            f"Average lag across all links is {summary['avg_lag_days']} days."
        ),
        'total_time_saving_days': total_time_saving_days,
        'suggestions': [
            {
                'title': 'Convert High-Lag FS Links to Controlled SS Overlap',
                'time_saving_days': min(18, max(6, int(round(summary['avg_lag_days'] * 0.6)))) if summary['total_dependencies'] else 0,
                'reorder_action': 'For top lag hotspots, release successor enablement packs before predecessor finish where technically feasible.',
                'risk_guardrail': 'Apply hold-points at 30% and 70% completion to avoid quality rework from premature starts.',
                'dependencies_affected': [x.get('from_id', '') for x in dependencies[:4] if x.get('from_id')],
            },
            {
                'title': 'Reduce Fan-Out Bottlenecks on Predecessor Hubs',
                'time_saving_days': min(14, max(4, int(round(len(top_hubs) * 0.4)))),
                'reorder_action': 'Split outgoing successor packs from the busiest predecessor nodes into staged approval waves.',
                'risk_guardrail': 'Keep one owner per wave with daily unblock decisions to prevent partial-release drift.',
                'dependencies_affected': [node_id for node_id, _ in top_outgoing[:4]],
            },
            {
                'title': 'Lag Governance for Incoming Successor Queues',
                'time_saving_days': min(10, max(3, int(round(len(top_incoming) * 0.25)))),
                'reorder_action': 'Add weekly lag burn-down for high incoming-load successors and close stale lag offsets early.',
                'risk_guardrail': 'Any lag over 45 days requires PMO escalation and re-baseline decision in same reporting cycle.',
                'dependencies_affected': [node_id for node_id, _ in top_incoming[:4]],
            },
        ],
        'chart_plan': chart_plan,
    }

    return {
        'ok': True,
        'error': '',
        'file': filename,
        'path': file_path,
        'dependencies': dependencies,
        'summary': summary,
        'analysis': analysis,
        'chart_plan': chart_plan,
        'signature': signature,
    }


def _build_whatif_structured_prompt(
    scenario_id: str,
    category_label: str,
    analysis_request: str,
    raw_prompt: str,
    whatif_knowledge_context: str = '',
):
    knowledge_section = whatif_knowledge_context.strip() or 'No supplemental What-IF Word knowledge context available.'
    return f"""You are Theta PMO What-IF planning copilot.
Return ONLY one valid JSON object. No markdown fences, no extra text.

SCENARIO_ID: {scenario_id}
CATEGORY: {category_label}
USER_ANALYSIS_REQUEST: {analysis_request}

SOURCE PROMPT:
{raw_prompt}

WHATIF KNOWLEDGE (from Knowledgebase/whatifKnowledge Word docs):
{knowledge_section}

Rules:
1. Use ONLY the WHATIF KNOWLEDGE section (loaded from Knowledgebase/whatifKnowledge Word docs).
2. Ignore Excel/workbook/JSON/dashboard instructions even if they appear in SOURCE PROMPT.
3. Do not fabricate IDs, dates, or counts. If unavailable in Word docs, state uncertainty clearly.
4. Keep output aligned to executive What-IF dashboard storytelling.
5. Return charts with concrete "type", "title", and "data" arrays suitable for frontend rendering.
6. The UI has Save vs Opportunity + Core Threat + Recovery blocks. Fill them.

Return this exact JSON shape:
{{
  "summary_markdown": "<concise analysis in markdown>",
  "ui_payload": {{
    "save_days": 0,
    "opportunity_days": 0,
    "residual_days": 0,
    "recovery_percent": 0,
    "core_threat_title": "<title>",
    "core_threat_summary": "<2-3 sentence summary>",
        "step1_title": "<Recovery Step 1 title>",
        "step1_summary": "<How Step 1 reduces delay>",
        "step1_baseline_days": [40, 47, 46],
        "step1_compressed_days": [16, 16, 16],
        "step1_activity_chips": ["A15870 Formwork Removal", "A15880 Protective Coating", "A15890 Back Filling"],
        "step1_recovered_days": 85,
        "step1_residual_to_sequence_days": 7,
    "kpis": {{
      "total_activities": 0,
      "delayed_activities": 0,
      "max_delay_days": 0,
      "average_delay_days": 0
    }},
    "threat_tracker": [
      {{"id":"A1000","name":"Activity Name","baseline":0,"actual":0,"late":0}}
    ],
    "compressor_savings": [
      {{"step":"Step name","days":"+0"}}
    ],
    "recommendations": ["<action 1>", "<action 2>", "<action 3>"]
  }},
  "chart_plan": [
    {{"id":"cp_issue_distribution","type":"pie","title":"CP Issue Type Distribution","data":[{{"name":"Issue","value":0}}]}},
    {{"id":"top_delays","type":"bar","title":"Top Delays (Duration vs BL)","data":[{{"name":"A1000","value":0}}]}}
  ],
  "endpoint_map": [
    {{"name":"Critical Dashboard","method":"GET","path":"/api/whatif/critical-dashboard","purpose":"Threat tracker and KPI baseline"}},
    {{"name":"Realtime Analysis","method":"POST","path":"/api/whatif/realtime-analysis","purpose":"Scenario chart payload"}},
    {{"name":"Project Update","method":"GET","path":"/api/whatif/project-update","purpose":"Forecast panel data"}},
    {{"name":"Predecessor Successor","method":"GET","path":"/api/whatif/predecessor-successor","purpose":"Dependency chain data"}}
  ]
}}"""


def _normalize_whatif_payload(payload: dict):
    payload = payload if isinstance(payload, dict) else {}
    ui = payload.get('ui_payload') if isinstance(payload.get('ui_payload'), dict) else {}
    kpis = ui.get('kpis') if isinstance(ui.get('kpis'), dict) else {}

    threat_tracker = []
    for row in (ui.get('threat_tracker') or [])[:20]:
        if not isinstance(row, dict):
            continue
        threat_tracker.append({
            'id': str(row.get('id', '')).strip(),
            'name': str(row.get('name', '')).strip(),
            'baseline': _to_int(row.get('baseline'), 0),
            'actual': _to_int(row.get('actual'), 0),
            'late': _to_int(row.get('late'), 0),
        })

    compressor_savings = []
    for row in (ui.get('compressor_savings') or [])[:20]:
        if not isinstance(row, dict):
            continue
        step = str(row.get('step', '')).strip()
        days_val = str(row.get('days', '')).strip()
        if step:
            compressor_savings.append({'step': step, 'days': days_val or '0'})

    save_days = _to_int(ui.get('save_days', kpis.get('max_delay_days', 0)), 0)
    opportunity_days = _to_int(ui.get('opportunity_days', 0), 0)
    residual_days = _to_int(ui.get('residual_days', max(0, save_days - opportunity_days)), 0)
    recovery_percent = _to_int(ui.get('recovery_percent', (round((opportunity_days / save_days) * 100) if save_days > 0 else 0)), 0)

    baseline_days = ui.get('step1_baseline_days') if isinstance(ui.get('step1_baseline_days'), list) else []
    compressed_days = ui.get('step1_compressed_days') if isinstance(ui.get('step1_compressed_days'), list) else []
    activity_chips = ui.get('step1_activity_chips') if isinstance(ui.get('step1_activity_chips'), list) else []

    normalized_baseline_days = [_to_int(x, 0) for x in baseline_days if x is not None][:3]
    normalized_compressed_days = [_to_int(x, 0) for x in compressed_days if x is not None][:3]
    normalized_activity_chips = [str(x).strip() for x in activity_chips if str(x).strip()][:6]

    if not normalized_baseline_days:
        normalized_baseline_days = [40, 47, 46]
    if not normalized_compressed_days:
        normalized_compressed_days = [16, 16, 16]
    if not normalized_activity_chips:
        normalized_activity_chips = ['A15870 Formwork Removal', 'A15880 Protective Coating', 'A15890 Back Filling']

    step1_recovered_days = _to_int(ui.get('step1_recovered_days', max(0, sum(normalized_baseline_days) - sum(normalized_compressed_days))), 0)
    step1_residual_to_sequence_days = _to_int(ui.get('step1_residual_to_sequence_days', max(0, save_days - step1_recovered_days)), 0)

    normalized = {
        'summary_markdown': str(payload.get('summary_markdown', '')).strip(),
        'ui_payload': {
            'save_days': save_days,
            'opportunity_days': opportunity_days,
            'residual_days': residual_days,
            'recovery_percent': recovery_percent,
            'core_threat_title': str(ui.get('core_threat_title', 'The Core Threat: Cascading Delays')).strip(),
            'core_threat_summary': str(ui.get('core_threat_summary', '')).strip(),
            'step1_title': str(ui.get('step1_title', 'Recovery Step 1: Downstream Civil Compression')).strip(),
            'step1_summary': str(ui.get('step1_summary', 'Re-planning downstream civil works at shorter durations absorbs major delay before compressor installation.')).strip(),
            'step1_baseline_days': normalized_baseline_days,
            'step1_compressed_days': normalized_compressed_days,
            'step1_activity_chips': normalized_activity_chips,
            'step1_recovered_days': step1_recovered_days,
            'step1_residual_to_sequence_days': step1_residual_to_sequence_days,
            'kpis': {
                'total_activities': _to_int(kpis.get('total_activities', 0), 0),
                'delayed_activities': _to_int(kpis.get('delayed_activities', 0), 0),
                'max_delay_days': _to_int(kpis.get('max_delay_days', save_days), save_days),
                'average_delay_days': float(kpis.get('average_delay_days', 0) or 0),
            },
            'threat_tracker': threat_tracker,
            'compressor_savings': compressor_savings,
            'recommendations': [str(x).strip() for x in (ui.get('recommendations') or []) if str(x).strip()][:8],
        },
        'chart_plan': [c for c in (payload.get('chart_plan') or []) if isinstance(c, dict)][:8],
        'endpoint_map': [e for e in (payload.get('endpoint_map') or []) if isinstance(e, dict)][:8],
    }

    if not normalized['summary_markdown']:
        normalized['summary_markdown'] = (
            f"### What-IF Summary\n\n"
            f"- Save: **{save_days} days**\n"
            f"- Opportunity: **{opportunity_days} days**\n"
            f"- Net Residual: **{residual_days} days**\n"
            f"- Recovery Efficiency: **{recovery_percent}%**"
        )

    return normalized


def _load_latest_project_updates(
    max_items: int = 500,
    preferred_job_id: str = '',
    preferred_filename: str = '',
    include_merged: bool = False,
):
    """Load and analyze latest uploaded project update files.

    Selection priority:
    1) Explicit filename match (if provided)
    2) Explicit outputs/<job_id> folder (if provided)
     3) Newest non-MERGED workbook with source preference:
         uploads > outputs/<job_id> > outputs root > outputs/_base_merge_workbook
     4) Newest MERGED workbook as fallback only
    """
    uploads_dir = os.path.join(_APP_ROOT, 'uploads')
    outputs_dir = os.path.join(_APP_ROOT, 'outputs')
    preferred_job_id = str(preferred_job_id or '').strip()
    preferred_filename = str(preferred_filename or '').strip()

    if not os.path.exists(uploads_dir):
        return {
            'ok': False,
            'error': 'Uploads directory not found',
            'updates': [],
            'summary': {},
            'top_faster_activities': [],
            'client_prompt': 'Evaluate progress update and identify overruns, impacts, faster activities, and recovery opportunities.',
        }

    def _is_excel(name: str) -> bool:
        raw = str(name or '').strip()
        n = raw.lower()
        if raw.startswith('~$'):
            return False
        return n.endswith('.xlsx') or n.endswith('.xls')

    def _is_merged(name: str) -> bool:
        return 'merged' in str(name or '').lower()

    def _is_base_merge_folder_candidate(item: dict) -> bool:
        return (
            str(item.get('source_bucket', '')) == 'outputs'
            and str(item.get('job_id', '')) == '_base_merge_workbook'
        )

    def _source_rank(item: dict) -> int:
        bucket = str(item.get('source_bucket', ''))
        job = str(item.get('job_id', ''))
        if bucket == 'uploads':
            return 4
        if bucket == 'outputs' and job and job != '_base_merge_workbook':
            return 3
        if bucket == 'outputs' and not job:
            return 2
        if _is_base_merge_folder_candidate(item):
            return 1
        return 0

    # Gather candidates from uploads root
    candidates = []
    for fname in os.listdir(uploads_dir):
        if not _is_excel(fname):
            continue
        fpath = os.path.join(uploads_dir, fname)
        if not os.path.isfile(fpath):
            continue
        try:
            candidates.append({
                'path': fpath,
                'filename': fname,
                'mtime': os.path.getmtime(fpath),
                'source_bucket': 'uploads',
                'job_id': '',
            })
        except Exception:
            continue

    # Gather candidates from outputs root + outputs/<job_id>/
    if os.path.isdir(outputs_dir):
        # root level files
        for fname in os.listdir(outputs_dir):
            fpath = os.path.join(outputs_dir, fname)
            if os.path.isfile(fpath) and _is_excel(fname):
                try:
                    candidates.append({
                        'path': fpath,
                        'filename': fname,
                        'mtime': os.path.getmtime(fpath),
                        'source_bucket': 'outputs',
                        'job_id': '',
                    })
                except Exception:
                    continue

        # one-level deep job folders
        for child in os.listdir(outputs_dir):
            child_path = os.path.join(outputs_dir, child)
            if not os.path.isdir(child_path):
                continue
            for fname in os.listdir(child_path):
                if not _is_excel(fname):
                    continue
                fpath = os.path.join(child_path, fname)
                if not os.path.isfile(fpath):
                    continue
                try:
                    candidates.append({
                        'path': fpath,
                        'filename': fname,
                        'mtime': os.path.getmtime(fpath),
                        'source_bucket': 'outputs',
                        'job_id': child,
                    })
                except Exception:
                    continue

    if not candidates:
        return {
            'ok': False,
            'error': 'No project update Excel files found in uploads/outputs',
            'updates': [],
            'summary': {},
            'top_faster_activities': [],
            'client_prompt': 'Evaluate progress update and identify overruns, impacts, faster activities, and recovery opportunities.',
        }

    # Optional explicit filename filter
    if preferred_filename:
        explicit = [
            c for c in candidates
            if c.get('filename', '').lower() == preferred_filename.lower()
        ]
        if explicit:
            explicit.sort(key=lambda x: x.get('mtime', 0), reverse=True)
            chosen = explicit[0]
        else:
            return {
                'ok': False,
                'error': f"Requested filename not found: {preferred_filename}",
                'updates': [],
                'summary': {},
                'top_faster_activities': [],
                'client_prompt': 'Evaluate progress update and identify overruns, impacts, faster activities, and recovery opportunities.',
            }
    else:
        chosen = None

    # Optional explicit job folder preference
    if not chosen and preferred_job_id:
        by_job = [c for c in candidates if c.get('job_id') == preferred_job_id]
        if by_job:
            if include_merged:
                scoped = by_job
            else:
                scoped = [c for c in by_job if not _is_merged(c.get('filename', ''))]
                if not scoped:
                    scoped = by_job
            scoped.sort(key=lambda x: x.get('mtime', 0), reverse=True)
            chosen = scoped[0]
        else:
            return {
                'ok': False,
                'error': f"Requested job folder has no Excel files: {preferred_job_id}",
                'updates': [],
                'summary': {},
                'top_faster_activities': [],
                'client_prompt': 'Evaluate progress update and identify overruns, impacts, faster activities, and recovery opportunities.',
            }

    # Default: prefer non-merged workbook unless include_merged=true
    if not chosen:
        if include_merged:
            usable = list(candidates)
        else:
            usable = [c for c in candidates if not _is_merged(c.get('filename', ''))]
            if not usable:
                usable = list(candidates)

        # By default, avoid selecting the synthetic base-merge workbook when real uploads exist.
        if not preferred_job_id and not preferred_filename:
            non_base_merge = [c for c in usable if not _is_base_merge_folder_candidate(c)]
            if non_base_merge:
                usable = non_base_merge

        usable.sort(key=lambda x: (_source_rank(x), x.get('mtime', 0)), reverse=True)
        chosen = usable[0]

    latest_file_path = chosen.get('path', '')
    latest_filename = chosen.get('filename', '')
    selected_job_id = chosen.get('job_id', '')
    selected_bucket = chosen.get('source_bucket', 'uploads')
    try:
        _st = os.stat(latest_file_path)
        file_signature = f"{latest_file_path}|{_st.st_mtime_ns}|{_st.st_size}"
    except Exception:
        file_signature = latest_file_path

    try:
        import pandas as pd
        
        # Read Excel file
        excel_file = pd.ExcelFile(latest_file_path)
        sheet_names = excel_file.sheet_names
        
        # Find the main data sheet (usually first or 'Data', 'Activities', etc.)
        main_sheet = sheet_names[0] if sheet_names else None
        if not main_sheet:
            return {
                'ok': False,
                'error': 'No sheets found in Excel file',
                'updates': [],
                'summary': {},
                'top_faster_activities': [],
                'client_prompt': 'Evaluate progress update and identify overruns, impacts, faster activities, and recovery opportunities.',
            }

        # Read the main sheet
        df = pd.read_excel(latest_file_path, sheet_name=main_sheet)
        
        # Standardize column names
        df.columns = [str(c).strip().lower() for c in df.columns]
        
        # Identify key columns (flexible matching)
        activity_id_col = None
        activity_name_col = None
        baseline_finish_col = None
        actual_finish_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'activity id' in col_lower or 'activity_id' in col_lower or col_lower == 'id':
                activity_id_col = col
            elif 'activity name' in col_lower or 'activity_name' in col_lower or col_lower == 'name':
                activity_name_col = col
            elif (
                ('baseline' in col_lower and ('finish' in col_lower or 'end' in col_lower or 'date' in col_lower))
                or ('planned' in col_lower and ('end' in col_lower or 'finish' in col_lower or 'completion' in col_lower))
            ):
                baseline_finish_col = col
            elif (
                ('actual' in col_lower or 'current' in col_lower)
                and ('finish' in col_lower or 'end' in col_lower or 'date' in col_lower or 'completion' in col_lower)
            ):
                actual_finish_col = col

        duration_deviation_col = None
        for col in df.columns:
            col_lower = col.lower()
            if 'duration_deviation' in col_lower or 'duration deviation' in col_lower:
                duration_deviation_col = col
                break
        
        # Build updates list
        updates = []
        faster_activities = []
        overrunning_activities = []
        
        for idx, row in df.iterrows():
            if pd.isna(row.get(activity_id_col)) or pd.isna(row.get(activity_name_col)):
                continue
            
            act_id = str(row.get(activity_id_col, '')).strip()
            act_name = str(row.get(activity_name_col, '')).strip()
            
            if not act_id or not act_name:
                continue
            
            baseline_date = row.get(baseline_finish_col)
            actual_date = row.get(actual_finish_col)
            
            # Calculate variance
            variance_days = 0
            update_type = 'neutral'
            impact_desc = ''
            
            if baseline_date and actual_date:
                try:
                    if isinstance(baseline_date, str):
                        baseline_date = pd.to_datetime(baseline_date)
                    if isinstance(actual_date, str):
                        actual_date = pd.to_datetime(actual_date)
                    
                    variance = (actual_date - baseline_date).days
                    variance_days = variance
                    
                    if variance > 5:  # Activity is overrunning
                        update_type = 'overwrite'
                        impact_desc = f'Overrun by {variance} days'
                        overrunning_activities.append({
                            'id': act_id,
                            'name': act_name,
                            'variance': variance,
                        })
                    elif variance < -5:  # Activity is ahead of schedule
                        update_type = 'inject'
                        impact_desc = f'Ahead by {abs(variance)} days'
                        faster_activities.append({
                            'id': act_id,
                            'name': act_name,
                            'time_saved': abs(variance),
                        })
                except Exception:
                    pass

            if variance_days == 0 and duration_deviation_col:
                try:
                    raw_dev = row.get(duration_deviation_col)
                    if pd.notna(raw_dev):
                        variance_days = int(float(raw_dev))
                        if variance_days > 5:
                            update_type = 'overwrite'
                            impact_desc = f'Overrun by {variance_days} days'
                            overrunning_activities.append({
                                'id': act_id,
                                'name': act_name,
                                'variance': variance_days,
                            })
                        elif variance_days < -5:
                            update_type = 'inject'
                            impact_desc = f'Ahead by {abs(variance_days)} days'
                            faster_activities.append({
                                'id': act_id,
                                'name': act_name,
                                'time_saved': abs(variance_days),
                            })
                except Exception:
                    pass
            
            safe_variance_days = 0
            try:
                if pd.notna(variance_days):
                    safe_variance_days = int(round(float(variance_days)))
            except Exception:
                safe_variance_days = 0

            updates.append({
                'activity_id': act_id,
                'activity_name': act_name,
                'baseline_finish': str(baseline_date).split()[0] if baseline_date else '',
                'actual_finish': str(actual_date).split()[0] if actual_date else '',
                'update_type': update_type,
                'impact': f'{safe_variance_days:+d}' if safe_variance_days != 0 else '0',
                'impact_desc': impact_desc,
            })
        
        # Calculate summary
        total_updates = len(updates)
        overwrites = len([u for u in updates if u.get('update_type') == 'overwrite'])
        injects = len([u for u in updates if u.get('update_type') == 'inject'])
        
        # Find high-impact overwrites
        high_impact_overwrites = 0
        max_overwrite_delay = 0
        for u in updates:
            if u.get('update_type') == 'overwrite':
                try:
                    impact_val = int(u.get('impact', '0').replace('+', ''))
                    if impact_val >= 15:
                        high_impact_overwrites += 1
                    max_overwrite_delay = max(max_overwrite_delay, impact_val)
                except:
                    pass
        
        # Calculate max time saved
        max_time_saved = max([a['time_saved'] for a in faster_activities], default=0)
        
        summary = {
            'total_updates': total_updates,
            'overwrites': overwrites,
            'injects': injects,
            'high_impact_overwrites': high_impact_overwrites,
            'max_overwrite_delay_days': max_overwrite_delay,
            'faster_activities': len(faster_activities),
            'max_time_saved_days': max_time_saved,
        }
        
        # Sort faster activities by time saved
        faster_activities.sort(key=lambda x: x.get('time_saved', 0), reverse=True)
        top_faster = faster_activities[:10]
        
        # Convert to list of dicts for display
        top_faster_activities = [
            {
                'activity_id': f.get('id', ''),
                'activity_name': f.get('name', ''),
                'time_saved_days': f.get('time_saved', 0),
            }
            for f in top_faster
        ]
        
        return {
            'ok': True,
            'error': '',
            'file': latest_filename,
            'sheet': main_sheet,
            'source_bucket': selected_bucket,
            'source_job_id': selected_job_id,
            'file_signature': file_signature,
            'updates': sorted(updates, key=lambda x: abs(int(x.get('impact', '0').replace('+', '') or 0)), reverse=True)[:max_items],
            'summary': summary,
            'top_faster_activities': top_faster_activities,
            'client_prompt': 'Evaluate progress update and identify overruns, impacts, faster activities, and recovery opportunities.',
            'upload_context': {
                'filename': latest_filename,
                'processed_at': datetime.now().isoformat(),
                'job_id': selected_job_id or 'auto-loaded',
                'source_bucket': selected_bucket,
            },
        }

    except ImportError:
        return {
            'ok': False,
            'error': 'pandas not available - cannot parse Excel files',
            'updates': [],
            'summary': {},
            'top_faster_activities': [],
            'client_prompt': 'Evaluate progress update and identify overruns, impacts, faster activities, and recovery opportunities.',
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            'ok': False,
            'error': f'Failed to parse project updates: {str(e)}',
            'updates': [],
            'summary': {},
            'top_faster_activities': [],
            'client_prompt': 'Evaluate progress update and identify overruns, impacts, faster activities, and recovery opportunities.',
        }


@app.route('/api/whatif/project-update', methods=['GET'])
@token_required
def get_whatif_project_update(current_user):
    """Return project update analysis from latest uploaded files with cache support."""
    try:
        refresh = str(request.args.get('refresh', '')).strip().lower() in ('1', 'true', 'yes', 'y')
        cached_only = str(request.args.get('cached_only', '')).strip().lower() in ('1', 'true', 'yes', 'y')
        sticky_cache = str(request.args.get('sticky_cache', '1')).strip().lower() in ('1', 'true', 'yes', 'y')
        job_id = str(request.args.get('job_id', '')).strip()
        filename = str(request.args.get('filename', '')).strip()
        include_merged = str(request.args.get('include_merged', '')).strip().lower() in ('1', 'true', 'yes', 'y')

        proj_cache = pg_read_whatif_proj_summary(current_user.get('company_id'))
        if not isinstance(proj_cache, dict):
            proj_cache = {}

        # Optional sticky mode: always serve the pinned overall snapshot first,
        # even if request job_id/filename changes.
        if not refresh and sticky_cache:
            overall_cached = proj_cache.get('project_update|overall')
            if isinstance(overall_cached, dict):
                return jsonify({
                    **overall_cached,
                    'cached': True,
                    'cache_key': 'project_update|overall',
                    'sticky_cache': True,
                    'pinned_snapshot': True,
                }), 200

        knowledge = _load_latest_project_updates(
            preferred_job_id=job_id,
            preferred_filename=filename,
            include_merged=include_merged,
        )
        if not knowledge.get('ok'):
            return jsonify({'error': knowledge.get('error') or 'Project update files could not be loaded'}), 404

        cache_key = (
            f"project_update|{knowledge.get('file', '')}|"
            f"{knowledge.get('source_bucket', '')}|{knowledge.get('source_job_id', '')}|"
            f"sig={knowledge.get('file_signature', '')}|include_merged={int(include_merged)}|v4"
        )

        if not refresh:
            cached = proj_cache.get(cache_key)
            if isinstance(cached, dict):
                return jsonify({**cached, 'cached': True, 'cache_key': cache_key}), 200

            alias_cached = proj_cache.get('project_update|overall')
            alias_source = (alias_cached or {}).get('source', {}) if isinstance(alias_cached, dict) else {}
            if (
                isinstance(alias_cached, dict)
                and str(alias_source.get('file', '')) == str(knowledge.get('file', ''))
                and str(alias_source.get('source_bucket', 'uploads')) == str(knowledge.get('source_bucket', 'uploads'))
                and str(alias_source.get('source_job_id', '')) == str(knowledge.get('source_job_id', ''))
            ):
                return jsonify({
                    **alias_cached,
                    'cached': True,
                    'cache_key': 'project_update|overall',
                    'fallback_cache': True,
                }), 200

            if cached_only:
                return jsonify({
                    'error': 'No cached project update data found. Run a non-cached request once to populate cache.',
                    'cached': False,
                    'cache_key': cache_key,
                }), 404

        # Generate frontend summary using normalization
        summary = knowledge.get('summary', {})
        updates = knowledge.get('updates', [])
        top_updates = updates[:6]
        
        # Build the prompt for normalization
        prompt = _build_project_update_frontend_prompt(
            workbook_name=knowledge.get('file', 'Project Update'),
            sheet_name=knowledge.get('sheet', 'Data'),
            summary=summary,
            top_updates=top_updates,
            client_prompt=knowledge.get('client_prompt', ''),
            upload_context=knowledge.get('upload_context', {}),
        )

        # Generate analysis via Claude (optional - can be enhanced later)
        raw_analysis = ''
        try:
            raw_analysis = _generate_claude_response(prompt, max_tokens=2000)
            parsed_analysis = _extract_json_object(raw_analysis)
            if parsed_analysis and isinstance(parsed_analysis, dict):
                frontend_summary = _normalize_project_update_frontend_summary(parsed_analysis, {
                    'total_updates': summary.get('total_updates', 0),
                    'overwrites': summary.get('overwrites', 0),
                    'injects': summary.get('injects', 0),
                    'high_impact_overwrites': summary.get('high_impact_overwrites', 0),
                    'max_overwrite_delay_days': summary.get('max_overwrite_delay_days', 0),
                })
            else:
                # Fall back to default summary
                frontend_summary = _normalize_project_update_frontend_summary({}, summary)
        except Exception as e:
            print(f'Claude analysis failed: {e}')
            frontend_summary = _normalize_project_update_frontend_summary({}, summary)

        # Add chart plan with concrete data
        chart_plan = []
        
        # Chart 1: Top Overwrite Delays
        overwrite_data = [
            {'name': u.get('activity_id', ''), 'value': abs(int(u.get('impact', '0').replace('+', '') or 0))}
            for u in updates
            if u.get('update_type') == 'overwrite'
        ][:8]
        if overwrite_data:
            chart_plan.append({
                'type': 'bar',
                'title': 'Top Overwrite Delays (Days)',
                'data': overwrite_data,
            })
        
        # Chart 2: Overwrite vs Inject Split
        update_split = [
            {'name': 'Overwrite', 'value': summary.get('overwrites', 0)},
            {'name': 'Inject', 'value': summary.get('injects', 0)},
        ]
        chart_plan.append({
            'type': 'pie',
            'title': 'Update Split (Overwrite vs Inject)',
            'data': update_split,
        })
        
        # Chart 3: Top Faster Activities (Time Saved)
        faster_data = [
            {'name': f.get('activity_id', ''), 'value': f.get('time_saved_days', 0)}
            for f in knowledge.get('top_faster_activities', [])
        ][:6]
        if faster_data:
            chart_plan.append({
                'type': 'line',
                'title': 'Faster Activities (Days Ahead)',
                'data': faster_data,
            })
        
        # Chart 4: Impact Distribution
        impact_by_type = {
            'High Impact (≥15d)': summary.get('high_impact_overwrites', 0),
            'Medium Impact (5-14d)': max(0, summary.get('overwrites', 0) - summary.get('high_impact_overwrites', 0)),
            'Low Impact (<5d)': len([u for u in updates if u.get('update_type') != 'overwrite' and u.get('update_type') != 'inject']),
        }
        impact_data = [
            {'name': k, 'value': v} for k, v in impact_by_type.items() if v > 0
        ]
        if impact_data:
            chart_plan.append({
                'type': 'bar',
                'title': 'Impact Distribution',
                'data': impact_data,
            })

        # Always expose at least one chart with concrete points so frontend can render visuals.
        if not chart_plan:
            chart_plan = [{
                'type': 'bar',
                'title': 'KPI Snapshot',
                'data': [
                    {'name': 'Total Updates', 'value': summary.get('total_updates', 0)},
                    {'name': 'Overwrites', 'value': summary.get('overwrites', 0)},
                    {'name': 'Injects', 'value': summary.get('injects', 0)},
                ],
            }]

        if isinstance(frontend_summary, dict):
            # Keep narrative dominant while ensuring charts are available in same summary object.
            frontend_summary['chart_plan'] = chart_plan[:3]
            frontend_summary['content_mix'] = {
                'description_percent': 65,
                'chart_percent': 35,
            }

        response_payload = {
            'generated_at': datetime.now().isoformat(),
            'source': {
                'mode': 'project_update_excel',
                'file': knowledge.get('file', 'Project Update'),
                'sheet': knowledge.get('sheet', 'Data'),
                'source_bucket': knowledge.get('source_bucket', 'uploads'),
                'source_job_id': knowledge.get('source_job_id', ''),
            },
            'updates': updates,
            'summary': summary,
            'frontend_summary': frontend_summary,
            'client_prompt': {'text': knowledge.get('client_prompt', '')},
            'upload_context': knowledge.get('upload_context', {}),
            'upload_merge_summary': None,
            'top_faster_activities': knowledge.get('top_faster_activities', []),
            'chart_plan': chart_plan,
        }

        proj_cache[cache_key] = response_payload
        proj_cache['project_update|overall'] = response_payload

        if len(proj_cache) > 40:
            try:
                items = list(proj_cache.items())
                items.sort(key=lambda x: ((x[1] or {}).get('generated_at') or ''), reverse=True)
                proj_cache = dict(items[:40])
            except Exception:
                pass

        pg_write_whatif_proj_summary(proj_cache, current_user.get('company_id'))
        return jsonify({**response_payload, 'cached': False, 'cache_key': cache_key}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to load project update data: {str(e)}'}), 500


@app.route('/api/whatif/recovery-narrative', methods=['GET'])
@token_required
def get_recovery_narrative(current_user):
    """
    Return the live recovery narrative JSON.
    Seeded from the bundled snapshot; updated automatically on every monthly file upload.
    Query params:
      refresh=1  — force re-seed from the bundled file (admin reset)
    """
    try:
        cid = current_user.get('company_id')
        if str(request.args.get('refresh', '')).strip().lower() in ('1', 'true', 'yes'):
            pg_write_recovery_narrative({}, cid)
            _seed_recovery_narrative(cid)

        data = pg_read_recovery_narrative(cid)
        if not data:
            if cid is None:
                # Legacy / Descon (null company): seed the Borouge snapshot as default
                _seed_recovery_narrative(cid)
                data = pg_read_recovery_narrative(cid)
            else:
                # New company: no narrative yet — seed with the default snapshot
                _seed_recovery_narrative(cid)
                data = pg_read_recovery_narrative(cid)
        if not isinstance(data, dict):
            return jsonify({'error': 'Recovery narrative not found'}), 404
        return jsonify(data), 200
    except Exception as e:
        print(f'[recovery_narrative] GET error: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/whatif/predecessor-successor', methods=['GET'])
@token_required
def get_whatif_predecessor_successor(current_user):
    """Return predecessor-successor dependency dashboard from PREDECESSOR-SUCCESSOR-LAG.json with cache support."""
    try:
        refresh = str(request.args.get('refresh', '')).strip().lower() in ('1', 'true', 'yes', 'y')
        cached_only = str(request.args.get('cached_only', '')).strip().lower() in ('1', 'true', 'yes', 'y')
        max_edges = max(200, int(request.args.get('max_edges', os.getenv('PREDSUCC_MAX_EDGES', '1200'))))

        knowledge = _load_predecessor_successor_json_context(max_edges=max_edges, company_id=current_user.get('company_id'))
        if not knowledge.get('ok'):
            return jsonify({'error': knowledge.get('error') or 'Predecessor-successor knowledge JSON could not be loaded'}), 404

        cache_key = f"predsucc_json|{knowledge.get('signature', '')}|{max_edges}|v1"
        pred_cache = pg_read_whatif_pred_succ(current_user.get('company_id'))
        if not isinstance(pred_cache, dict):
            pred_cache = {}

        if not refresh:
            cached = pred_cache.get(cache_key)
            if isinstance(cached, dict):
                return jsonify({**cached, 'cached': True, 'cache_key': cache_key}), 200

            alias_cached = pred_cache.get('predecessor_successor|overall')
            if isinstance(alias_cached, dict):
                return jsonify({
                    **alias_cached,
                    'cached': True,
                    'cache_key': 'predecessor_successor|overall',
                    'fallback_cache': True,
                }), 200

            if cached_only:
                return jsonify({
                    'error': 'No cached predecessor-successor data found. Run a non-cached request once to populate cache.',
                    'cached': False,
                    'cache_key': cache_key,
                }), 404

        response_payload = {
            'generated_at': datetime.now().isoformat(),
            'source': {
                'mode': 'predecessor_successor_json',
                'file': knowledge.get('file', 'PREDECESSOR-SUCCESSOR-LAG.json'),
                'path': knowledge.get('path', ''),
            },
            'dependencies': knowledge.get('dependencies', []),
            'analysis': knowledge.get('analysis', {}),
            'summary': knowledge.get('summary', {}),
        }

        pred_cache[cache_key] = response_payload
        pred_cache['predecessor_successor|overall'] = response_payload

        if len(pred_cache) > 40:
            try:
                items = list(pred_cache.items())
                items.sort(key=lambda x: ((x[1] or {}).get('generated_at') or ''), reverse=True)
                pred_cache = dict(items[:40])
            except Exception:
                pass

        pg_write_whatif_pred_succ(pred_cache, current_user.get('company_id'))
        return jsonify({**response_payload, 'cached': False, 'cache_key': cache_key}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to load predecessor-successor data: {str(e)}'}), 500


@app.route('/api/whatif/claude-analysis', methods=['POST'])
@token_required
def whatif_claude_analysis(current_user):
    try:
        data = request.get_json() or {}
        scenario_id = str(data.get('scenario_id', '')).strip()
        category = str(data.get('category', '')).strip()
        analysis_request = str(data.get('analysis_request', '')).strip()
        raw_prompt = str(data.get('prompt', '')).strip()
        include_whatif_knowledge = str(data.get('include_whatif_knowledge', 'true')).strip().lower() in ('1', 'true', 'yes', 'y')

        if not scenario_id:
            return jsonify({'error': 'scenario_id is required'}), 400
        if not category:
            return jsonify({'error': 'category is required'}), 400
        if not raw_prompt:
            return jsonify({'error': 'prompt is required'}), 400

        knowledge_payload = _load_whatif_knowledge_context(company_id=current_user.get('company_id')) if include_whatif_knowledge else {'context': '', 'sources': [], 'folder': ''}

        structured_prompt = _build_whatif_structured_prompt(
            scenario_id=scenario_id,
            category_label=category,
            analysis_request=analysis_request,
            raw_prompt=raw_prompt,
            whatif_knowledge_context=knowledge_payload.get('context', ''),
        )

        max_tokens = max(1024, int(os.getenv('CLAUDE_WHATIF_MAX_TOKENS', '3500')))
        raw = _generate_claude_response(structured_prompt, max_tokens=max_tokens)
        parsed = _extract_json_object(raw)

        # If model returns non-JSON or mixed text, run one strict repair pass for stable frontend payloads.
        repaired = False
        if not parsed:
            repair_prompt = f"""Convert the following model output into ONE valid JSON object only.
No markdown, no prose outside JSON.

Required top-level keys:
- summary_markdown (string)
- ui_payload (object)
- chart_plan (array)
- endpoint_map (array)

Model output to repair:
{raw}
"""
            repair_max_tokens = max(600, min(1600, max_tokens // 2))
            repaired_raw = _generate_claude_response(repair_prompt, max_tokens=repair_max_tokens)
            repaired_parsed = _extract_json_object(repaired_raw)
            if repaired_parsed:
                parsed = repaired_parsed
                repaired = True

        normalized = _normalize_whatif_payload(parsed or {'summary_markdown': raw})

        entries = pg_read_whatif_claude(current_user.get('company_id'))
        if not isinstance(entries, list):
            entries = []

        response_id = str(uuid.uuid4())
        entries.insert(0, {
            'id': response_id,
            'user_id': current_user['id'],
            'scenario_id': scenario_id,
            'category': category,
            'analysis_request': analysis_request,
            'knowledge_sources': knowledge_payload.get('sources', []),
            'knowledge_folder': knowledge_payload.get('folder', ''),
            'payload': normalized,
            'json_repaired': bool(repaired),
            'created_at': datetime.now().isoformat(),
        })
        pg_write_whatif_claude(entries[:500], current_user.get('company_id'))

        if scenario_id == 'whatif_critical_activities':
            critical_cache = pg_read_whatif_critical(current_user.get('company_id'))
            if not isinstance(critical_cache, dict):
                critical_cache = {}
            ui_payload = normalized.get('ui_payload', {}) if isinstance(normalized.get('ui_payload', {}), dict) else {}
            critical_cache[f'{scenario_id}|{category}'] = {
                'generated_at': datetime.now().isoformat(),
                'scenario_id': scenario_id,
                'category': category,
                'analysis_request': analysis_request,
                'summary_markdown': normalized.get('summary_markdown', ''),
                'ui_payload': ui_payload,
                'threat_tracker': ui_payload.get('threat_tracker', []),
                'compressor_savings': ui_payload.get('compressor_savings', []),
                'kpis': ui_payload.get('kpis', {}),
                'chart_plan': normalized.get('chart_plan', []),
                'endpoint_map': normalized.get('endpoint_map', []),
                'cached': True,
                'source': 'claude-analysis',
            }
            pg_write_whatif_critical(critical_cache, current_user.get('company_id'))

        return jsonify({
            'id': response_id,
            'message': normalized.get('summary_markdown', ''),
            'ui_payload': normalized.get('ui_payload', {}),
            'chart_plan': normalized.get('chart_plan', []),
            'endpoint_map': normalized.get('endpoint_map', []),
            'knowledge_sources': knowledge_payload.get('sources', []),
            'knowledge_folder': knowledge_payload.get('folder', ''),
            'knowledge_sources_count': len(knowledge_payload.get('sources', [])),
            'knowledge_context_loaded': bool(knowledge_payload.get('context', '').strip()),
            'json_repaired': bool(repaired),
            'stored': True,
            'model': 'azure-anthropic-claude',
            'timestamp': datetime.now().isoformat(),
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/whatif/claude-analysis/history', methods=['GET'])
@token_required
def whatif_claude_analysis_history(current_user):
    try:
        limit = max(1, min(100, int(request.args.get('limit', 20))))
        scenario_id = str(request.args.get('scenario_id', '')).strip()

        entries = pg_read_whatif_claude(current_user.get('company_id'))
        if not isinstance(entries, list):
            entries = []

        filtered = [e for e in entries if e.get('user_id') == current_user['id']]
        if scenario_id:
            filtered = [e for e in filtered if str(e.get('scenario_id', '')).strip() == scenario_id]

        return jsonify({'items': filtered[:limit]}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/whatif/claude-analysis/<analysis_id>', methods=['GET'])
@token_required
def whatif_claude_analysis_item(current_user, analysis_id):
    try:
        entries = pg_read_whatif_claude(current_user.get('company_id'))
        if not isinstance(entries, list):
            entries = []

        item = next((e for e in entries if e.get('id') == analysis_id and e.get('user_id') == current_user['id']), None)
        if not item:
            return jsonify({'error': 'What-IF analysis not found'}), 404
        return jsonify(item), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _build_insight_prompt(section: str, title: str, context_text: str):
    return f"""You are a principal PMO intelligence advisor for large-scale EPC and capital construction programs.
You are generating an enterprise-grade insight card for a project intelligence dashboard.

SECTION : {section}
TITLE   : {title}

PROJECT CONTEXT DATA:
{context_text}

INSTRUCTIONS:
1. Base every finding ONLY on figures and facts present in the context data above.
   Do NOT invent numbers, dates, percentages, or activity names.
2. If data is partial, derive the most useful insight from what IS available and clearly note what is missing.
3. Use precise EPC/PMO language: S-curve, critical path, float, SPI, CPI, baseline, deviation, RFSU, EDDR, etc.
4. summary_headline: one punchy executive sentence — lead with the most critical finding or risk.
5. present_analysis: 2-3 sentences on the current measurable state (cite actual values where available).
6. future_impact: 2-3 sentences on the schedule, cost, or quality consequence if the current trend continues unchanged.
7. recommendations: 4 specific, actionable items — each must name a responsible party or discipline where possible.
8. do_list: 3 concrete "do NOW" actions for the project team (specific, not generic).
9. dont_list: 3 specific pitfalls or anti-patterns to avoid for this context.
10. urgency: set to "high" if the finding involves critical-path impact, cost overrun risk, or safety; "medium" for moderate risk; "low" otherwise.

Return ONLY valid JSON — no markdown, no commentary, no extra keys, no text outside the JSON object:
{{
  "summary_headline": "<one executive sentence>",
  "urgency": "high|medium|low",
  "present_analysis": "<2-3 sentences with specific values>",
  "future_impact": "<2-3 sentences on consequence if unchanged>",
  "recommendations": ["<Specific action 1>", "<Specific action 2>", "<Specific action 3>", "<Specific action 4>"],
  "do_list": ["<Do now 1>", "<Do now 2>", "<Do now 3>"],
  "dont_list": ["<Avoid 1>", "<Avoid 2>", "<Avoid 3>"]
}}"""


def _generate_intelligence_insight(section: str, title: str, context_text: str):
    prompt = _build_insight_prompt(section, title, context_text)

    try:
        insight_max_tokens = max(512, int(os.getenv('CLAUDE_INSIGHT_MAX_TOKENS', '1800')))
        raw = _generate_claude_response(prompt, max_tokens=insight_max_tokens)
        parsed = _extract_json_object(raw)
        if parsed:
            return _clean_insight_payload(parsed)
    except Exception as _az_err:
        print(f"[INTEL-INSIGHT] Azure generation failed: {_az_err}")

    return None


def _load_benchmark_dataset(company_id=None):
    """Reads benchmark workbook and returns parsed dashboard data + data hash for caching."""
    import pandas as pd
    import numpy as np

    file_path = os.path.join(_get_knowledgebase_folder(company_id), 'Project intel vs benchmark.xlsx')
    if not os.path.exists(file_path):
        raise FileNotFoundError('Benchmark file not found')

    stat = os.stat(file_path)
    data_hash = hashlib.sha256(f"{file_path}|{stat.st_mtime}|{stat.st_size}".encode('utf-8')).hexdigest()

    df = pd.read_excel(file_path, sheet_name=None)

    scorecard_df = df.get('Scorecard', pd.DataFrame()).replace({np.nan: None})
    score_data = scorecard_df.to_dict(orient='records')

    scorecard_items = []
    radar_items = []
    radar_map = {
        'Manufacturing & Delivery': 'Manufacturing',
        'Engineering / EDDR': 'Engineering',
        'Project Management': 'Project Mgmt',
        'Construction & Pre-commissioning Overall': 'Construction',
        'Electrical Construction': 'Electrical',
        'Mechanical Construction': 'Mechanical',
        'Civil Construction': 'Civil',
        'Bulk Materials': 'Bulk Materials'
    }

    if len(score_data) > 2:
        for row in score_data[2:]:
            area = row.get('Unnamed: 0')
            score = row.get('Unnamed: 1')
            basis = row.get('Unnamed: 3')
            if not area or score is None or not isinstance(score, (int, float)):
                continue
            status = 'amber'
            if score >= 80:
                status = 'green'
            elif score < 60:
                status = 'red'

            scorecard_items.append({
                'area': area,
                'score': score,
                'status': status,
                'basis': basis,
            })

            if area in radar_map:
                radar_items.append({
                    'subject': radar_map[area],
                    'A': score,
                    'fullMark': 100,
                })

    overall_info = {'index': None, 'summary': None}
    fixed_order = ['Manufacturing', 'Engineering', 'Project Mgmt', 'Construction', 'Electrical', 'Mechanical', 'Civil', 'Bulk Materials']
    radar_items.sort(key=lambda x: fixed_order.index(x['subject']) if x['subject'] in fixed_order else 99)

    for row in score_data:
        val = row.get('Unnamed: 0')
        if isinstance(val, str):
            v = val.strip()
            if v.startswith('Overall Project Benchmark Index:'):
                overall_info['index'] = v.split(':')[-1].strip()
            elif v.startswith('The project is showing'):
                overall_info['summary'] = v

    heatmap_df = df.get('Heatmap', pd.DataFrame()).replace({np.nan: None})
    heatmap_data = heatmap_df.to_dict(orient='records')

    top_drivers = []
    if len(heatmap_data) > 2:
        for row in heatmap_data[2:]:
            rank = row.get('Unnamed: 0')
            driver = row.get('Unnamed: 1')
            if isinstance(rank, (int, float)) and driver:
                top_drivers.append(driver)
            if len(top_drivers) >= 5:
                break

    strategic_insights = []
    sheet3_df = df.get('Sheet3', pd.DataFrame()).replace({np.nan: None})
    for row in sheet3_df.to_dict(orient='records'):
        val = row.get('Unnamed: 0')
        if val and isinstance(val, str):
            v_str = val.strip()
            if v_str and not v_str.startswith('Strategic') and v_str not in ['Meaning:', 'For future projects suggests:', 'For future projects this suggests:']:
                strategic_insights.append(v_str)

    management_review = {'strong': [], 'weak': [], 'conclusion': ''}
    mgmt_df = df.get('Management review', pd.DataFrame()).replace({np.nan: None})
    current_section = None
    for row in mgmt_df.to_dict(orient='records'):
        val = row.get('Unnamed: 0')
        if not val or not isinstance(val, str):
            continue
        v_str = val.strip()
        if not v_str:
            continue
        if 'Where the project is stronger' in v_str:
            current_section = 'strong'
            continue
        if 'Where the project is weaker' in v_str:
            current_section = 'weak'
            continue
        if 'Main conclusion' in v_str:
            current_section = 'conclusion'
            continue
        if 'Bottom-line management view' in v_str:
            continue
        if current_section == 'strong':
            management_review['strong'].append(v_str)
        elif current_section == 'weak':
            management_review['weak'].append(v_str)
        elif current_section == 'conclusion':
            management_review['conclusion'] += v_str + ' '

    management_review['conclusion'] = management_review['conclusion'].strip()

    benchmark_df = df.get('Benchmark', pd.DataFrame()).replace({np.nan: None})
    heatmap_df2 = df.get('Heatmap', pd.DataFrame()).replace({np.nan: None})

    benchmark_details = []
    for row in benchmark_df.to_dict(orient='records'):
        vals = list(row.values())
        if len(vals) >= 4 and any(v for v in vals):
            benchmark_details.append({
                'category': str(vals[0]) if vals[0] else '',
                'metric': str(vals[1]) if vals[1] else '',
                'range': str(vals[2]) if vals[2] else '',
                'implication': str(vals[3]) if vals[3] else '',
            })

    heatmap_details = []
    for row in heatmap_df2.to_dict(orient='records'):
        vals = [str(v) if v is not None else '' for v in row.values()]
        if any(v != '' for v in vals):
            heatmap_details.append(vals)

    return {
        'radarData': radar_items,
        'scorecard': scorecard_items,
        'topDrivers': top_drivers,
        'strategicInsights': strategic_insights,
        'managementReview': management_review,
        'overallInfo': overall_info,
        'benchmarkDetails': benchmark_details[1:] if len(benchmark_details) > 0 else [],
        'heatmapDetails': heatmap_details,
        '_data_hash': data_hash,
    }


def _build_benchmark_insight_items(benchmark_data):
    items = []

    for kpi in benchmark_data.get('scorecard', []):
        area = str(kpi.get('area') or '').strip()
        if not area:
            continue
        items.append({
            'section': 'kpi',
            'title': area,
            'cache_key': f"kpi|{_normalize_key(area)}",
            'context_text': f"Area: {area}\nScore: {kpi.get('score')}\nStatus: {kpi.get('status')}\nBasis: {kpi.get('basis') or ''}",
        })

    for r in benchmark_data.get('radarData', []):
        subject = str(r.get('subject') or '').strip()
        if not subject:
            continue
        items.append({
            'section': 'performance-variance',
            'title': f"Performance Variance - {subject}",
            'cache_key': f"performance-variance|{_normalize_key(subject)}",
            'context_text': f"Visual: Performance Variance vs Benchmark\nDimension: {subject}\nScore: {r.get('A')} / {r.get('fullMark')}",
        })

    for level in ['High', 'Medium-High', 'Medium']:
        for row in benchmark_data.get('heatmapDetails', [])[1:]:
            if len(row) < 3 or row[2] != level:
                continue
            title = str(row[1] or '').strip()
            if not title:
                continue
            items.append({
                'section': 'risk-delay',
                'title': f"{level} Risk - {title}",
                'cache_key': f"risk-delay|{_normalize_key(level)}|{_normalize_key(title)}",
                'context_text': (
                    f"Risk Level: {level}\nRisk Name: {title}\n"
                    f"Description: {row[3] if len(row) > 3 else ''}\n"
                    f"Impact Note: {row[4] if len(row) > 4 else ''}"
                ),
            })

    for driver in benchmark_data.get('topDrivers', []):
        title = str(driver or '').strip()
        if not title:
            continue
        items.append({
            'section': 'risk-delay',
            'title': f"Delay Driver - {title}",
            'cache_key': f"risk-delay|driver|{_normalize_key(title)}",
            'context_text': f"Top Delay Driver: {title}",
        })

    for insight in benchmark_data.get('strategicInsights', []):
        title = str(insight or '').strip()
        if not title:
            continue
        items.append({
            'section': 'strategic-intelligence',
            'title': title[:120],
            'cache_key': f"strategic-intelligence|{_normalize_key(title)[:100]}",
            'context_text': f"Strategic recommendation statement:\n{title}",
        })

    # Keep deterministic order and prevent accidental runaway
    dedup = {}
    for item in items:
        dedup[item['cache_key']] = item
    return list(dedup.values())[:180]


def _refresh_single_benchmark_insight(cache_key: str, section: str, title: str, context_text: str, data_hash: str, generated_by='system', company_id=None):
    try:
        generated = _generate_intelligence_insight(section, title, context_text)
        if generated:
            _upsert_cached_intelligence_insight(
                cache_key=cache_key,
                data_hash=data_hash,
                section=section,
                title=title,
                insight=generated,
                generated_by=generated_by,
                company_id=company_id,
            )
    except Exception as e:
        print(f"[INTEL-INSIGHT] Async refresh failed for {cache_key}: {e}")


def _refresh_single_benchmark_insight_async(cache_key: str, section: str, title: str, context_text: str, data_hash: str, generated_by='system', company_id=None):
    t = threading.Thread(
        target=_refresh_single_benchmark_insight,
        args=(cache_key, section, title, context_text, data_hash, generated_by, company_id),
        daemon=True,
    )
    t.start()


def _get_or_generate_benchmark_insight(
    section: str,
    title: str,
    context_text: str,
    data_hash: str,
    generated_by='system',
    allow_stale_fast: bool = True,
    refresh_stale_async: bool = True,
    company_id=None,
):
    cache_key = f"{_normalize_key(section)}|{_normalize_key(title)}"
    cached = _get_cached_intelligence_insight(cache_key, data_hash, company_id)
    if cached and cached.get('insight'):
        is_stale = bool(cached.get('is_stale'))
        if is_stale and refresh_stale_async:
            _refresh_single_benchmark_insight_async(
                cache_key=cache_key,
                section=section,
                title=title,
                context_text=context_text,
                data_hash=data_hash,
                generated_by=generated_by,
                company_id=company_id,
            )
        if allow_stale_fast or not is_stale:
            return cached.get('insight'), True, cache_key, is_stale, cached.get('age_hours')

    generated = _generate_intelligence_insight(section, title, context_text)
    if not generated:
        return None, False, cache_key, False, None

    _upsert_cached_intelligence_insight(
        cache_key=cache_key,
        data_hash=data_hash,
        section=section,
        title=title,
        insight=generated,
        generated_by=generated_by,
        company_id=company_id,
    )
    return generated, False, cache_key, False, 0.0


def _run_benchmark_precompute_job(triggered_by='system'):
    with BENCHMARK_PRECOMPUTE_LOCK:
        BENCHMARK_PRECOMPUTE_STATE.update({
            'status': 'running',
            'started_at': datetime.now().isoformat(),
            'finished_at': None,
            'processed': 0,
            'total': 0,
            'last_error': None,
        })

    try:
        benchmark_data = _load_benchmark_dataset()
        data_hash = benchmark_data.get('_data_hash')
        items = _build_benchmark_insight_items(benchmark_data)

        with BENCHMARK_PRECOMPUTE_LOCK:
            BENCHMARK_PRECOMPUTE_STATE['total'] = len(items)

        for idx, item in enumerate(items, start=1):
            cached = _get_cached_intelligence_insight(item['cache_key'], data_hash)
            stale = bool(cached.get('is_stale')) if cached else True
            if (not cached) or stale:
                generated = _generate_intelligence_insight(item['section'], item['title'], item['context_text'])
                if generated:
                    _upsert_cached_intelligence_insight(
                        cache_key=item['cache_key'],
                        data_hash=data_hash,
                        section=item['section'],
                        title=item['title'],
                        insight=generated,
                        generated_by=triggered_by,
                    )
            with BENCHMARK_PRECOMPUTE_LOCK:
                BENCHMARK_PRECOMPUTE_STATE['processed'] = idx

        with BENCHMARK_PRECOMPUTE_LOCK:
            BENCHMARK_PRECOMPUTE_STATE['status'] = 'completed'
            BENCHMARK_PRECOMPUTE_STATE['finished_at'] = datetime.now().isoformat()

    except Exception as e:
        print(f"[BENCHMARK-PRECOMPUTE] Error: {e}")
        with BENCHMARK_PRECOMPUTE_LOCK:
            BENCHMARK_PRECOMPUTE_STATE['status'] = 'failed'
            BENCHMARK_PRECOMPUTE_STATE['last_error'] = str(e)
            BENCHMARK_PRECOMPUTE_STATE['finished_at'] = datetime.now().isoformat()


def _start_benchmark_precompute_async(triggered_by='system'):
    with BENCHMARK_PRECOMPUTE_LOCK:
        if BENCHMARK_PRECOMPUTE_STATE.get('status') == 'running':
            return False
    t = threading.Thread(target=_run_benchmark_precompute_job, args=(triggered_by,), daemon=True)
    t.start()
    return True


@app.route('/api/benchmark/precompute-insights', methods=['POST'])
@token_required
def start_benchmark_precompute(current_user):
    started = _start_benchmark_precompute_async(triggered_by=current_user.get('id', 'system'))
    with BENCHMARK_PRECOMPUTE_LOCK:
        state = dict(BENCHMARK_PRECOMPUTE_STATE)
    state['cache_ttl_hours'] = INTELLIGENCE_CACHE_TTL_HOURS
    return jsonify({
        'started': started,
        'status': state,
    }), 202 if started else 200

# ==================== ENGAGE MONTHLY AUTO-SUMMARY ====================

# ── Push Notification Helpers ──────────────────────────────────────────────────

def _send_push_to_user(user_id: str, title: str, body: str, url: str = '/') -> int:
    """
    Send a Web Push notification to all registered devices of a user.
    Returns the number of successful sends.
    """
    try:
        from pywebpush import webpush, WebPushException
    except ImportError:
        print('[PUSH] pywebpush not installed — skipping push.')
        return 0

    vapid_private = os.getenv('VAPID_PRIVATE_KEY', '')
    vapid_public  = os.getenv('VAPID_PUBLIC_KEY', '')
    vapid_mailto  = os.getenv('VAPID_MAILTO', 'mailto:admin@example.com')

    if not vapid_private or not vapid_public:
        return 0

    import base64, json as _json

    # Decode the base64-PEM private key
    try:
        pem_bytes = base64.urlsafe_b64decode(vapid_private + '==')
        vapid_claims = {'sub': vapid_mailto}
        payload = _json.dumps({'title': title, 'body': body, 'url': url}).encode()
    except Exception as e:
        print(f'[PUSH] Key decode error: {e}')
        return 0

    subscriptions = get_push_subscriptions(user_id)
    sent = 0
    for sub in subscriptions:
        try:
            webpush(
                subscription_info=sub,
                data=payload,
                vapid_private_key=pem_bytes,
                vapid_claims=vapid_claims,
            )
            sent += 1
        except WebPushException as e:
            if e.response and e.response.status_code in (404, 410):
                delete_push_subscription(sub['endpoint'])
            else:
                print(f'[PUSH] WebPushException for {user_id}: {e}')
        except Exception as e:
            print(f'[PUSH] Send error for {user_id}: {e}')
    return sent


def _send_push_to_all_users_with_role(roles: list, title: str, body: str, url: str = '/') -> int:
    """Send push to all users matching the given roles."""
    from db_postgres import read_db
    import os as _os
    users = read_db(_os.path.join(DB_FOLDER, 'users.json'))
    total = 0
    for u in users:
        if u.get('role') in roles:
            total += _send_push_to_user(u['id'], title, body, url)
    return total


# ── Push Subscription Endpoints ────────────────────────────────────────────────

@app.route('/api/push/vapid-public-key', methods=['GET'])
def push_vapid_public_key():
    """Return the VAPID public key for the frontend to subscribe."""
    key = os.getenv('VAPID_PUBLIC_KEY', '')
    if not key:
        return jsonify({'error': 'Push notifications not configured'}), 503
    return jsonify({'publicKey': key}), 200


@app.route('/api/push/subscribe', methods=['POST'])
@token_required
def push_subscribe(current_user):
    """Register a push subscription for the current user."""
    body = request.get_json() or {}
    endpoint = body.get('endpoint', '')
    keys     = body.get('keys', {})
    p256dh   = keys.get('p256dh', '')
    auth     = keys.get('auth', '')
    if not endpoint or not p256dh or not auth:
        return jsonify({'error': 'endpoint, keys.p256dh and keys.auth are required'}), 400
    save_push_subscription(current_user['id'], endpoint, p256dh, auth)
    return jsonify({'status': 'subscribed'}), 200


@app.route('/api/push/unsubscribe', methods=['POST'])
@token_required
def push_unsubscribe(current_user):
    """Remove a push subscription."""
    body     = request.get_json() or {}
    endpoint = body.get('endpoint', '')
    if endpoint:
        delete_push_subscription(endpoint)
    return jsonify({'status': 'unsubscribed'}), 200


@app.route('/api/support/create', methods=['POST'])
def support_create():
    """Accept a contact or bug-report submission from any client (web/mobile)."""
    data = request.get_json() or {}
    ticket_type = (data.get('type') or 'contact').strip()
    subject     = (data.get('subject') or '').strip()
    message     = (data.get('message') or '').strip()
    name        = (data.get('name') or '').strip()
    email       = (data.get('email') or '').strip()
    priority    = (data.get('priority') or 'Medium').strip()

    if not subject or not message:
        return jsonify({'error': 'subject and message are required'}), 400

    try:
        import smtplib
        from email.mime.text import MIMEText

        smtp_host  = os.getenv('SMTP_HOST', '')
        smtp_port  = int(os.getenv('SMTP_PORT', 587))
        smtp_user  = os.getenv('SMTP_USER', '')
        smtp_pass  = os.getenv('SMTP_PASSWORD', '')
        support_to = os.getenv('SUPPORT_EMAIL', smtp_user)

        if smtp_host and smtp_user and support_to:
            body_lines = [
                f"Type: {ticket_type}",
                f"Priority: {priority}",
                f"From: {name} <{email}>",
                "",
                message,
            ]
            msg = MIMEText('\n'.join(body_lines))
            msg['Subject'] = f"[Support] {subject}"
            msg['From']    = smtp_user
            msg['To']      = support_to

            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.ehlo()
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, [support_to], msg.as_string())
    except Exception as e:
        print(f"[SUPPORT] Email send failed (non-fatal): {e}")

    try:
        auth_header = request.headers.get('Authorization', '')
        token_str   = auth_header.replace('Bearer ', '').strip()
        uid = None
        if token_str:
            import jwt as pyjwt
            payload = pyjwt.decode(token_str, app.config['SECRET_KEY'], algorithms=['HS256'])
            uid = payload.get('user_id') or payload.get('id')
        log_activity(
            action_type='support_ticket',
            user_id=uid or 'anonymous',
            user_name=name or 'anonymous',
            user_role='user',
            company_id=None,
            description=f"Support ticket ({ticket_type}): {subject}",
            source=request.headers.get('X-App-Source', SOURCE_WEB),
            level=LEVEL_USER,
            metadata={'type': ticket_type, 'priority': priority, 'email': email},
            ip_address=request.remote_addr,
        )
    except Exception:
        pass

    return jsonify({'message': 'Support ticket received. We will get back to you shortly.'}), 201


# ── Deviation Lifecycle Background Job ─────────────────────────────────────────

def _run_deviation_lifecycle():
    """
    Check for:
    1. Deviations due a 15-day reminder  → push notification to the manager/uploader
    2. Deviations past 20-day expiry     → auto-lock as 'Expired' + push to admin
    Runs every 6 hours.
    """
    import time as _t
    reminder_days = int(os.getenv('DEVIATION_REMINDER_DAYS', '15'))
    expiry_days   = int(os.getenv('DEVIATION_EXPIRY_DAYS', '20'))

    while True:
        try:
            # ── Reminders ───────────────────────────────────────────────────
            due = get_deviations_due_reminder(reminder_days)
            for dev in due:
                try:
                    uid = dev.get('user_id')
                    dev_id = dev.get('id')
                    if uid:
                        _send_push_to_user(
                            uid,
                            title='Deviation Reminder',
                            body=f'Deviation #{dev_id} has been pending for {reminder_days}+ days. Please review before it expires.',
                            url='/deviations',
                        )
                    update_reminder_timestamp(dev_id, datetime.now(tz=ZoneInfo('Asia/Kolkata')).isoformat())
                except Exception as e:
                    print(f'[LIFECYCLE] Reminder error dev #{dev.get("id")}: {e}')
            if due:
                print(f'[LIFECYCLE] Sent {len(due)} reminder notification(s)')

            # ── Expiry ───────────────────────────────────────────────────────
            expired = get_deviations_to_expire()
            for dev in expired:
                try:
                    dev_id = dev.get('id')
                    expire_deviation(dev_id)
                    # Notify the deviation owner
                    uid = dev.get('user_id')
                    if uid:
                        _send_push_to_user(
                            uid,
                            title='Deviation Expired',
                            body=f'Deviation #{dev_id} has expired after {expiry_days} days without a response and has been locked.',
                            url='/deviations',
                        )
                    # Notify admins
                    _send_push_to_all_users_with_role(
                        ['admin'],
                        title='Deviation Auto-Locked',
                        body=f'Deviation #{dev_id} was automatically locked after {expiry_days} days without a review.',
                        url='/deviations',
                    )
                    # In-app notification to admin
                    create_notification(
                        user_id=None,   # will notify all admins via notify_admins_and_managers
                        title='Deviation Auto-Locked',
                        message=f'Deviation #{dev_id} ({dev.get("sheet","")}) was auto-locked as unanswered after {expiry_days} days.',
                        notification_type='warning',
                        metadata={'deviation_id': dev_id, 'reason': 'expired'},
                    )
                    print(f'[LIFECYCLE] Expired deviation #{dev_id}')
                except Exception as e:
                    print(f'[LIFECYCLE] Expiry error dev #{dev.get("id")}: {e}')

        except Exception as e:
            print(f'[LIFECYCLE] Loop error: {e}')

        _t.sleep(6 * 3600)   # run every 6 hours


def _start_deviation_lifecycle_thread():
    t = threading.Thread(target=_run_deviation_lifecycle, daemon=True)
    t.start()
    print('[LIFECYCLE] Deviation lifecycle background thread started.')


# ── Admin Deviation Report Endpoint ───────────────────────────────────────────

@app.route('/api/admin/deviation-report', methods=['GET'])
@token_required
def admin_deviation_report(current_user):
    """
    Admin: month-wise deviation breakdown scoped to the caller's company.
    super_admin can pass ?company_id= to filter to a specific company.
    """
    role = current_user.get('role')
    if role not in ('admin', 'manager', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Admin or manager access required'}), 403
    if role == 'super_admin':
        company_id = request.args.get('company_id') or None
    else:
        company_id = current_user.get('company_id')
    rows = get_deviation_report_by_month(company_id=company_id)
    # Convert Decimal/int to plain int for JSON serialisation
    clean = []
    for r in rows:
        clean.append({k: int(v) if v is not None else 0 for k, v in r.items() if k != 'month'} | {'month': r['month']})
    return jsonify({'report': clean}), 200


def _generate_engage_monthly_summary(company_id=None):
    """
    Generate an AI monthly project summary using:
    - The latest monthly report KPIs (from monthly_reports table)
    - Recent deviation stats (pending / reviewed counts)
    - Latest processed output file context (timeline deviation sheet)
    Then post it to Engage and record in PostgreSQL so it never double-posts.
    """
    try:
        import pandas as pd

        _now = datetime.now(ZoneInfo('Asia/Kolkata'))
        month_label = _now.strftime('%B %Y')
        this_month  = _now.strftime('%Y-%m')
        print(f'[ENGAGE-AUTO] Generating monthly summary for {month_label} ...')

        # ── 1. Monthly report KPIs ─────────────────────────────────────────────
        kpi_block = ''
        try:
            reports = get_monthly_reports(year=_now.year, company_id=company_id)
            if reports:
                latest = reports[-1]
                d = latest['data']
                sc = d.get('scurves', {})
                kpi_block = (
                    f"MONTHLY REPORT KPIs — {latest['month']} {latest['year']}\n"
                    f"Total Activities : {d.get('totalActivities', 'N/A')}\n"
                    f"On Time          : {d.get('onTime', 'N/A')}\n"
                    f"Milestone Done   : {d.get('milestoneAchieved', 'N/A')}\n"
                    f"On Plan          : {d.get('onPlan', 'N/A')}\n"
                    f"Duration Missing : {d.get('durationMissing', 'N/A')}\n"
                    f"In Progress      : {d.get('inProgress', 'N/A')}\n"
                    f"Not Started      : {d.get('notStarted', 'N/A')}\n"
                    f"Avg Duration     : {d.get('avgPlannedDuration', 'N/A')}d\n"
                    f"S-Curves (planned% vs actual%):\n"
                    + '\n'.join(
                        f"  {k}: planned={v.get('planned', 0) if isinstance(v, dict) else v}% actual={v.get('actual', 0) if isinstance(v, dict) else v}%"
                        for k, v in sc.items()
                    )
                )
        except Exception as e:
            print(f'[ENGAGE-AUTO] KPI read error: {e}')

        # ── 2. Deviation stats ─────────────────────────────────────────────────
        dev_block = ''
        try:
            from deviation_db import get_all_deviations, get_pending_deviations
            all_devs     = get_all_deviations(company_id=company_id)
            pending_devs = get_pending_deviations(company_id=company_id)
            high_sev     = [d for d in all_devs if d.get('severity') == 'High']
            dev_block = (
                f"DEVIATION SUMMARY\n"
                f"Total deviations : {len(all_devs)}\n"
                f"Pending review   : {len(pending_devs)}\n"
                f"High severity    : {len(high_sev)}\n"
            )
            # Top 5 pending high-severity
            pending_high = [d for d in pending_devs if d.get('severity') == 'High'][:5]
            if pending_high:
                dev_block += "Top pending high-severity:\n"
                for d in pending_high:
                    dev_block += f"  #{d.get('id')} [{d.get('sheet','')}] {str(d.get('description',''))[:80]}\n"
        except Exception as e:
            print(f'[ENGAGE-AUTO] Deviation stats error: {e}')

        # ── 3. Output file context (timeline deviation sheet) ──────────────────
        context_lines = []
        _char_limit = 15_000
        _total_chars = 0
        try:
            completed = pg_read_history_for_company(company_id=company_id, status='completed', limit=5)
            for job in completed:
                job_id = job.get('id')
                for result in job.get('results', []):
                    if result.get('status') != 'success':
                        continue
                    out_file = result.get('output_filename', '')
                    if 'timeline_deviation_tracker' not in out_file:
                        continue
                    file_path = os.path.join(_APP_ROOT, OUTPUT_FOLDER, job_id, out_file)
                    if not os.path.exists(file_path):
                        continue
                    try:
                        df = pd.read_excel(file_path, sheet_name='Timeline Deviation', engine='openpyxl')
                        # Tag rows missing both Start and Early Start as milestones
                        start_missing = df.get('Start', pd.Series(dtype=object)).isna() | (df.get('Start', pd.Series(dtype=object)).astype(str).str.strip() == '')
                        early_start_missing = df.get('Early Start', pd.Series(dtype=object)).isna() | (df.get('Early Start', pd.Series(dtype=object)).astype(str).str.strip() == '')
                        df = df.copy()
                        df['Activity Type'] = 'Activity'
                        df.loc[start_missing & early_start_missing, 'Activity Type'] = 'Milestone'
                        # Delayed and in-progress activities only
                        delayed = df[df['Status'].astype(str).str.contains('Delay|Progress', case=False, na=False)]
                        sample  = delayed.head(30).to_string()
                        block   = f"\n[{out_file} — delayed/in-progress activities, {len(delayed)} rows]\n{sample}"
                        context_lines.append(block)
                        _total_chars += len(block)
                    except Exception as e:
                        print(f'[ENGAGE-AUTO] File read error {out_file}: {e}')
                    if _total_chars >= _char_limit:
                        break
                if _total_chars >= _char_limit:
                    break
        except Exception as e:
            print(f'[ENGAGE-AUTO] Output context error: {e}')

        if not kpi_block and not dev_block and not context_lines:
            print('[ENGAGE-AUTO] No data available — skipping post.')
            return

        data_block = '\n\n'.join(filter(None, [kpi_block, dev_block, '\n'.join(context_lines)]))

        # ── 4. Build prompt and call AI ────────────────────────────────────────
        prompt = (
            f"You are Theta PMO AI. Today is {_now.strftime('%B %d, %Y')}.\n\n"
            "Write a concise monthly project update post for the team's Engage feed using the data below.\n\n"
            "Use exactly these four bold headings, one short paragraph each:\n"
            "**Overall Health** — on-time rate and one-line project status.\n"
            "**Immediate Action Required** — the single most urgent overdue or at-risk item with its ID and day count.\n"
            "**Key Risks This Month** — top deviation area or systemic pattern with root cause.\n"
            "**Pending Reviews** — how many deviations are awaiting manager/admin action and a call-to-action.\n\n"
            "Rules: be specific (name activity IDs, day counts, percentages). No filler. Max 300 words. "
            "IMPORTANT: Use ONLY numbers, IDs, and facts explicitly present in the DATA block below — do NOT invent, estimate, or fabricate any values. "
            "Rows labelled 'Activity Type = Milestone' in the timeline data are project milestones (they have no Start or Early Start date), not regular activities — refer to them as milestones.\n\n"
            f"=== DATA ===\n{data_block}\n=== END DATA ==="
        )

        summary_text = _generate_claude_response(prompt, max_tokens=700)
        if not summary_text:
            print('[ENGAGE-AUTO] AI returned empty response — skipping.')
            return

        # ── 5. Post to Engage ──────────────────────────────────────────────────
        post_id = str(uuid.uuid4())
        post = {
            'id':         post_id,
            'user_id':    None,           # system post — no FK to users
            'company_id': company_id,
            'user_name':  'Theta PMO AI',
            'user_email': '',
            'content':    f"**Monthly Project Update — {month_label}**\n\n{summary_text}",
            'image_url':  '',
            'group_id':   '',
            'source':     'auto-monthly-summary',
            'likes':      [],
            'comments':   [],
            'created_at': _now_ist_iso(),
        }
        pg_insert_engage_post(post)

        # ── 6. Record in PostgreSQL so we never double-post ───────────────────
        log_monthly_summary_posted(this_month, post_id, company_id=company_id)
        print(f'[ENGAGE-AUTO] Monthly summary posted for {month_label}: {post_id}')

    except Exception as e:
        import traceback
        print(f'[ENGAGE-AUTO] Error: {e}')
        traceback.print_exc()


def _engage_monthly_summary_loop():
    """
    Background thread: wakes every 30 minutes and posts on the 1st of each month
    between 08:00 and 10:00 (IST) if not already posted this month.
    Using a 2-hour window + 30-minute checks ensures the post fires even if the
    app was restarted during that window.
    """
    import time as _time_mod
    while True:
        try:
            _now = datetime.now(ZoneInfo('Asia/Kolkata'))
            if _now.day == 1 and 8 <= _now.hour < 10:
                this_month = _now.strftime('%Y-%m')
                from db_postgres import get_companies
                for company in get_companies():
                    cid = company['id']
                    if not is_monthly_summary_posted(this_month, company_id=cid):
                        _generate_engage_monthly_summary(company_id=cid)
        except Exception as e:
            print(f'[ENGAGE-AUTO] Loop error: {e}')
        _time_mod.sleep(1800)   # check every 30 minutes


def _start_engage_monthly_summary_thread():
    t = threading.Thread(target=_engage_monthly_summary_loop, daemon=True)
    t.daemon = True
    t.start()
    print('[ENGAGE-AUTO] Monthly summary background thread started.')


@app.route('/api/engage/trigger-monthly-summary', methods=['POST'])
@token_required
def trigger_monthly_summary(current_user):
    """Admin: manually trigger the monthly summary post (ignores the already-posted guard)."""
    if current_user.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    force = str(request.get_json(silent=True) or {}).lower() != 'false'
    def _run():
        if force:
            # Temporarily clear this month's log entry so it re-posts
            from db_postgres import _conn as _pg
            try:
                _now = datetime.now(ZoneInfo('Asia/Kolkata'))
                with _pg() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "DELETE FROM engage_monthly_summary_log WHERE month = %s",
                            (_now.strftime('%Y-%m'),),
                        )
            except Exception:
                pass
        _generate_engage_monthly_summary(company_id=current_user.get('company_id'))
    threading.Thread(target=_run, daemon=True).start()
    return jsonify({'message': 'Monthly summary generation started in background.'}), 202


@app.route('/api/benchmark/precompute-status', methods=['GET'])
@token_required
def benchmark_precompute_status(current_user):
    with BENCHMARK_PRECOMPUTE_LOCK:
        state = dict(BENCHMARK_PRECOMPUTE_STATE)
    state['cache_ttl_hours'] = INTELLIGENCE_CACHE_TTL_HOURS
    return jsonify({'status': state})


@app.route('/api/benchmark/insight', methods=['POST'])
@token_required
def get_benchmark_insight(current_user):
    """Generic insight endpoint for KPI, performance variance, risks/delays, and strategic visuals."""
    try:
        data = request.json or {}
        section = str(data.get('section', 'kpi')).strip() or 'kpi'
        title = str(data.get('title', '')).strip()
        context_text = str(data.get('context_text', '')).strip()

        if not title:
            return jsonify({'error': 'Insight title is required'}), 400

        benchmark_data = _load_benchmark_dataset(current_user.get('company_id'))
        data_hash = benchmark_data.get('_data_hash')

        if not context_text:
            # Best-effort fallback from known benchmark items
            mapped = next(
                (x for x in _build_benchmark_insight_items(benchmark_data)
                 if _normalize_key(x.get('title')) == _normalize_key(title)
                 and _normalize_key(x.get('section')) == _normalize_key(section)),
                None,
            )
            context_text = mapped.get('context_text', title) if mapped else title

        insight, is_cached, cache_key, is_stale, cache_age_hours = _get_or_generate_benchmark_insight(
            section=section,
            title=title,
            context_text=context_text,
            data_hash=data_hash,
            generated_by=current_user.get('id', 'system'),
            company_id=current_user.get('company_id'),
        )

        if not insight:
            return jsonify({'error': 'AI insight generation failed. Please retry shortly.'}), 503

        return jsonify({
            **insight,
            'cache_key': cache_key,
            'cached': is_cached,
            'stale': is_stale,
            'cache_age_hours': cache_age_hours,
            'cache_ttl_hours': INTELLIGENCE_CACHE_TTL_HOURS,
            'section': section,
            'title': title,
        })
    except Exception as e:
        print(f"[BENCHMARK-INSIGHT] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/benchmark/kpi-insight', methods=['POST'])
@token_required
def get_kpi_ai_insight(current_user):
    """Backward-compatible KPI insight endpoint."""
    try:
        data = request.json or {}
        area = str(data.get('area', '')).strip()
        score = data.get('score', '')
        status = data.get('status', 'amber')
        basis = data.get('basis', '')

        if not area:
            return jsonify({'error': 'KPI area is required'}), 400

        benchmark_data = _load_benchmark_dataset()
        data_hash = benchmark_data.get('_data_hash')
        context_text = f"Area: {area}\nScore: {score}\nStatus: {status}\nBasis: {basis}"

        insight, is_cached, cache_key, is_stale, cache_age_hours = _get_or_generate_benchmark_insight(
            section='kpi',
            title=area,
            context_text=context_text,
            data_hash=data_hash,
            generated_by='legacy-kpi-endpoint',
        )

        if not insight:
            return jsonify({'error': 'AI service unavailable. Please try again shortly.'}), 503

        return jsonify({
            **insight,
            'cache_key': cache_key,
            'cached': is_cached,
            'stale': is_stale,
            'cache_age_hours': cache_age_hours,
            'cache_ttl_hours': INTELLIGENCE_CACHE_TTL_HOURS,
            'section': 'kpi',
            'title': area,
        })
    except Exception as e:
        print(f"[KPI-INSIGHT] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/benchmark', methods=['GET'])
@token_required
def get_benchmark_data(current_user):
    """Reads Benchmark Excel file and formats it into dashboard data."""
    try:
        benchmark_data = _load_benchmark_dataset()
        return jsonify({
            'radarData': benchmark_data.get('radarData', []),
            'scorecard': benchmark_data.get('scorecard', []),
            'topDrivers': benchmark_data.get('topDrivers', []),
            'strategicInsights': benchmark_data.get('strategicInsights', []),
            'managementReview': benchmark_data.get('managementReview', {}),
            'overallInfo': benchmark_data.get('overallInfo', {}),
            'benchmarkDetails': benchmark_data.get('benchmarkDetails', []),
            'heatmapDetails': benchmark_data.get('heatmapDetails', []),
            'precomputeStatus': BENCHMARK_PRECOMPUTE_STATE.get('status', 'idle'),
        })
    except FileNotFoundError:
        return jsonify({'error': 'Benchmark file not found'}), 404
    except Exception as e:
        print(f"[BENCHMARK] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


_start_engage_monthly_summary_thread()
_start_deviation_lifecycle_thread()


# ── Monthly Reports ────────────────────────────────────────────────────────────

# S-curve sheet name → discipline key used in History.jsx
_SCURVE_SHEET_MAP = {
    'Home_Office_Services_SCurve':     'homeOffice',
    'Manufacturing_and_Delive_SCurve': 'manufacturing',
    'Construction_and_Pre-Com_SCurve': 'construction',
    'Project_Management_SCurve':       'projectMgmt',
    'Commissioning,_RFSU,_Sta_SCurve': 'commissioning',
}

# Period label → lower-case 3-letter month key used in futurePlanned
_PERIOD_TO_KEY = {
    'Jan': 'jan', 'Feb': 'feb', 'Mar': 'mar', 'Apr': 'apr',
    'May': 'may', 'Jun': 'jun', 'Jul': 'jul', 'Aug': 'aug',
    'Sep': 'sep', 'Oct': 'oct', 'Nov': 'nov', 'Dec': 'dec',
}

_MONTH_FILE_PATTERNS = {
    'January':   ['%jan%'],
    'February':  ['%feb%'],
    'March':     ['%march%', '%mar%'],
    'April':     ['%apr%', '%17-04%'],
    'May':       ['%may%'],
    'June':      ['%jun%'],
    'July':      ['%jul%'],
    'August':    ['%aug%'],
    'September': ['%sep%'],
    'October':   ['%oct%'],
    'November':  ['%nov%'],
    'December':  ['%dec%'],
}

# Full month names and 3-letter abbreviations for filename detection (longest first to avoid partial matches)
_MONTH_FULLNAMES = [
    ('january', 'January'), ('february', 'February'), ('september', 'September'),
    ('november', 'November'), ('december', 'December'), ('october', 'October'),
    ('august', 'August'), ('march', 'March'), ('april', 'April'), ('june', 'June'),
    ('july', 'July'), ('may', 'May'),
]
_MONTH_ABBR3 = [
    ('jan', 'January'), ('feb', 'February'), ('mar', 'March'), ('apr', 'April'),
    ('may', 'May'), ('jun', 'June'), ('jul', 'July'), ('aug', 'August'),
    ('sep', 'September'), ('oct', 'October'), ('nov', 'November'), ('dec', 'December'),
]


def _detect_month_from_filename(filename: str) -> str | None:
    """Return a full month name detected in the filename, or None.

    Handles:
      - Full month names: 'april', 'january', …
      - 3-letter abbreviations: 'apr', 'jan', …
      - Numeric date patterns: DD-MM-YY / DD-MM-YYYY (e.g. '17-04-26' → April)
    """
    import re
    lower = filename.lower()
    for full, name in _MONTH_FULLNAMES:
        if full in lower:
            return name
    for abbr, name in _MONTH_ABBR3:
        if re.search(r'(?<![a-z])' + abbr + r'(?![a-z])', lower):
            return name
    # Numeric date patterns: DD-MM-YY or DD-MM-YYYY (European format — month is middle group)
    _num_to_month = {
        1: 'January', 2: 'February', 3: 'March', 4: 'April',
        5: 'May', 6: 'June', 7: 'July', 8: 'August',
        9: 'September', 10: 'October', 11: 'November', 12: 'December',
    }
    m = re.search(r'(?<!\d)(\d{1,2})[-./](\d{2})[-./](\d{2,4})(?!\d)', lower)
    if m:
        month_num = int(m.group(2))
        if 1 <= month_num <= 12:
            print(f'[reports] Detected month {month_num} from numeric date pattern in "{filename}"')
            return _num_to_month.get(month_num)
    return None


_MONTH_NUM_TO_NAME = {
    1: 'January', 2: 'February', 3: 'March', 4: 'April',
    5: 'May', 6: 'June', 7: 'July', 8: 'August',
    9: 'September', 10: 'October', 11: 'November', 12: 'December',
}


def _detect_month_from_tracker_dates(job_id: str, results: list, year: int) -> str | None:
    """Infer the report month from the majority of start/finish dates in the tracker file."""
    import openpyxl
    from collections import Counter
    import datetime as _dt

    tracker_path = None
    for res in results:
        out_file = res.get('output_filename', '')
        if 'timeline_deviation_tracker' in out_file:
            candidate = os.path.join(_APP_ROOT, OUTPUT_FOLDER, job_id, out_file)
            if os.path.exists(candidate):
                tracker_path = candidate
                break
    if not tracker_path:
        return None

    try:
        wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
        tdev_sheet_name = next(
            (s for s in wb.sheetnames if 'timeline deviation' in s.lower()),
            wb.sheetnames[0] if wb.sheetnames else None
        )
        if not tdev_sheet_name:
            wb.close()
            return None
        ws = wb[tdev_sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        print(f'[reports] _detect_month_from_tracker_dates open error: {e}')
        return None

    if not rows:
        return None
    h = {str(c).strip() if c else '': i for i, c in enumerate(rows[0])}
    data = rows[1:]

    # Prefer known date-column names; fall back to scanning ALL columns for datetime objects
    date_col_names = ['start', 'Start', 'early start', 'Early Start',
                      'finish', 'Finish', 'early finish', 'Early Finish']
    date_cols = [h[k] for k in date_col_names if k in h]
    # If no named date columns found, scan every column (openpyxl returns datetime objects
    # directly for date-formatted cells, so this works even without knowing column names)
    scan_all = not date_cols
    num_cols = len(rows[0]) if rows else 0

    month_counter = Counter()
    for row in data:
        col_indices = range(num_cols) if scan_all else date_cols
        for col_idx in col_indices:
            val = row[col_idx] if col_idx < len(row) else None
            if val is None:
                continue
            try:
                if isinstance(val, (_dt.datetime, _dt.date)):
                    d = val.date() if isinstance(val, _dt.datetime) else val
                else:
                    import pandas as _pd
                    ts = _pd.to_datetime(val, errors='coerce')
                    if _pd.isnull(ts):
                        continue
                    d = ts.date()
                if d.year == year:
                    month_counter[d.month] += 1
            except Exception:
                pass

    if not month_counter:
        return None
    most_common = month_counter.most_common(1)[0][0]
    print(f'[reports] Month inferred from tracker dates: {_MONTH_NUM_TO_NAME.get(most_common)} '
          f'(top months: {month_counter.most_common(3)})')
    return _MONTH_NUM_TO_NAME.get(most_common)


def _seed_recovery_narrative(company_id=None):
    """Write the default seed data to PostgreSQL if no row exists yet for this company."""
    if pg_read_recovery_narrative(company_id):
        return
    seed = {
        "last_updated": "2026-06-06T00:00:00",
        "source_month": "April", "source_year": 2026,
        "project": {"label": "Borouge EU3 - H2 Extraction EPC", "title": "Schedule Recovery and Calibration",
                    "mc_target": "20 Jan 2027", "status": "Zero Variance - On Track"},
        "kpis": [
            {"id": "target_completion", "label": "Target Completion", "value": "20 Jan 2027", "tone": "neutral"},
            {"id": "max_delay_absorbed", "label": "Max Delay Absorbed", "value": "+103",
             "sub": "Days upstream civil overrun fully neutralized", "tone": "amber"},
            {"id": "days_recovered", "label": "Days Recovered", "value": "98",
             "sub": "85 civil + 13 compressor compression embedded", "tone": "teal"},
            {"id": "residual_risk", "label": "Residual Risk", "value": "+7",
             "sub": "Days slip at C1000, mitigable via +1 shift/week", "tone": "red"},
        ],
        "recovery_architecture": {
            "waterfall": [
                {"label": "Baseline Zero", "days": 0, "tone": "slate"},
                {"label": "Max Civil Delay", "days": 103, "tone": "amber"},
                {"label": "Civil Compression", "days": -85, "tone": "teal"},
                {"label": "Compressor Chain", "days": -13, "tone": "teal_light"},
                {"label": "Net Position", "days": 0, "tone": "neutral"},
            ],
            "progress": {"civil_saved": 85, "civil_total": 103, "compressor_saved": 13, "compressor_total": 13},
            "civil_overruns": [
                {"id": "A15750", "name": "Waterproofing", "baseline_days": 19, "current_days": 136, "finish_variance_days": 103},
                {"id": "A15850", "name": "Concrete Pouring", "baseline_days": 40, "current_days": 80, "finish_variance_days": 81},
                {"id": "A15840", "name": "Rebar Works", "baseline_days": 40, "current_days": 75, "finish_variance_days": 79},
            ],
            "civil_overrun_summary": "5 activities, 243 total days overrun, 103-day finish variance",
        },
        "stage_recovery": {
            "stage1": {
                "title": "Stage 1 - Downstream Civil Compression",
                "summary": "Three downstream civil activities re-planned, embedding 85 days of recovery into the sequence.",
                "activities": [
                    {"id": "A15870", "name": "Formwork Removal", "baseline_days": 40, "planned_days": 16, "saved_days": 24},
                    {"id": "A15880", "name": "Protective Coating", "baseline_days": 47, "planned_days": 16, "saved_days": 31},
                    {"id": "A15890", "name": "Backfill and Compaction", "baseline_days": 46, "planned_days": 16, "saved_days": 30},
                ],
                "delay_before_days": 103, "delay_after_days": 7, "reduction_percent": 96,
            },
            "stage2": {
                "title": "Stage 2 - Compressor Chain Compression",
                "summary": "10 sequential activities tightened 1-2 days each, generating a 13-day buffer.",
                "chain": [
                    {"id": "C1000", "saved_days": -1}, {"id": "C1001", "saved_days": -1},
                    {"id": "C1002", "saved_days": -2}, {"id": "C1007", "saved_days": -2},
                    {"id": "C1008", "saved_days": -1}, {"id": "C1009", "saved_days": -2},
                    {"id": "C1010", "saved_days": -1}, {"id": "C1011", "saved_days": -1},
                    {"id": "C1013", "saved_days": -1}, {"id": "C1014", "saved_days": -1},
                ],
                "total_saved_days": 13,
                "outcome": "All downstream activities restored to baseline finish dates",
            },
        },
        "critical_path_calibration": {
            "summary": "Recalibration from 102 to 35 activities restores precision.",
            "false_positive": 67, "false_negative": 20, "true_critical": 35,
            "current_state_count": 102, "optimized_state_count": 35,
            "risk_register": [
                {"item": "C1000", "description": "Foundation Check - 7-day unrecovered slip", "impact": "HIGH", "status": "OPEN"},
                {"item": "Vendor Risk", "description": "Vendor delays could consume downstream buffer", "impact": "MED", "status": "MONITOR"},
            ],
        },
        "active_mitigation": {
            "title": "C1000 Residual Risk - Mitigation Lever",
            "lever": "+1 Extra Shift / Week", "recovered_days": 7, "window_days": 39,
            "detail": "Apply one additional working shift per week during C1000 Foundation Check and C1002 Baseplate Setting.",
        },
        "forward_operating_mandate": [
            {"id": "01", "title": "Lock the 35-Activity Path",
             "description": "Adopt the corrected Critical Path immediately and discontinue the 102-activity model."},
            {"id": "02", "title": "Protect the Compressions",
             "description": "Execute the 85-day civil and 13-day compressor compressions exactly as re-planned."},
            {"id": "03", "title": "Activate C1000 Shift Extension",
             "description": "Mandate the +1 shift/week protocol for foundation check and baseplate setting."},
        ],
        "trend_summary": {
            "headline": "Mechanical Completion has slipped 62 days from 20 Jan 2027 to 23 Mar 2027",
            "body": "All three terminal milestones have moved 62 calendar days. MC: 20 Jan to 23 Mar 2027. The slip is driven by the A15750 civil foundation chain (103-day finish variance), partially offset by 98 days of embedded compression.",
        },
        "trend_metrics": [
            {"value": "+62d", "label": "MC slip vs baseline", "change_note": "All milestones impacted", "severity": "red"},
            {"value": "+15d", "label": "C1000 float - improved from +4d (Mar)", "change_note": "Compressor chain healthier", "severity": "green"},
            {"value": "+112", "label": "New completions Mar to Apr - rate slowed 72%", "change_note": "Productivity concern", "severity": "amber"},
            {"value": "2", "label": "Vendor drawings still at 1-day float", "change_note": "A31580, A34220 stuck", "severity": "red"},
        ],
        "monthly_kpis": {
            "totalActivities": 0, "onTime": 0, "onLate": 0,
            "onEarly": 0, "notStarted": 0, "milestoneAchieved": 0,
            "avgPlannedDuration": 0, "maxPlannedDuration": 0,
        },
    }
    try:
        pg_write_recovery_narrative(seed, company_id)
        print('[recovery_narrative] Seeded default data.')
    except Exception as e:
        print(f'[recovery_narrative] Seed write error: {e}')


def _update_recovery_narrative(month: str, year: int, tracker_report: dict, company_id=None):
    """
    Merge tracker KPIs from the latest upload into the recovery_narrative PG row.
    Structural fields (stage_recovery, civil_overruns, etc.) are preserved;
    only the live-countable fields are overwritten.
    """
    try:
        existing = pg_read_recovery_narrative(company_id)
        if not isinstance(existing, dict) or not existing.get('active_mitigation'):
            # No full seed data yet — seed first so structural fields are present
            _seed_recovery_narrative(company_id)
            existing = pg_read_recovery_narrative(company_id)
            if not isinstance(existing, dict):
                existing = {}

        existing['last_updated'] = datetime.now().isoformat()
        existing['last_tracker_month'] = month
        existing['last_tracker_year'] = year

        existing['monthly_kpis'] = {
            'totalActivities':    tracker_report.get('totalActivities', 0),
            'onTime':             tracker_report.get('onTime', 0),
            'onLate':             tracker_report.get('onLate', 0),
            'onEarly':            tracker_report.get('onEarly', 0),
            'notStarted':         tracker_report.get('notStarted', 0),
            'milestoneAchieved':  tracker_report.get('milestoneAchieved', 0),
            'avgPlannedDuration': tracker_report.get('avgPlannedDuration', 0),
            'maxPlannedDuration': tracker_report.get('maxPlannedDuration', 0),
        }

        pg_write_recovery_narrative(existing, company_id)
        print(f'[recovery_narrative] Updated for {month} {year} '
              f'(totalActivities={tracker_report.get("totalActivities",0)}, '
              f'onLate={tracker_report.get("onLate",0)})')
    except Exception as e:
        print(f'[recovery_narrative] Update error: {e}')


def _auto_save_monthly_report(job_id: str, results: list, filename: str, processed_year: int, company_id=None):
    """Detect month from filename and persist the extracted report to monthly_reports."""
    month = _detect_month_from_filename(filename)
    if not month:
        print(f'[reports] Month not in filename "{filename}", trying tracker dates…')
        month = _detect_month_from_tracker_dates(job_id, results, processed_year)
    if not month:
        print(f'[reports] Could not detect month from filename or data — report not auto-saved')
        return
    print(f'[reports] Detected month={month} for "{filename}"')
    report = _extract_report_from_tracker(job_id, results, month, year=processed_year)
    if not report:
        print(f'[reports] Extraction failed for {month} {processed_year} (job {job_id[:8]})')
        return
    ok = upsert_monthly_report(month, processed_year, report, company_id)
    if ok:
        print(f'[reports] Auto-saved {month} {processed_year} report from job {job_id[:8]}')
        _update_recovery_narrative(month, processed_year, report, company_id)
    else:
        print(f'[reports] FAILED to save {month} {processed_year} report for job {job_id[:8]} company={company_id}')


def _find_best_job(month: str):
    """Return (job_id, results_list) for the most recent successful job matching month."""
    import psycopg2
    patterns = _MONTH_FILE_PATTERNS.get(month, [])
    try:
        with _conn_report() as conn:
            with conn.cursor() as cur:
                for pat in patterns:
                    cur.execute(
                        """SELECT id, results FROM history
                           WHERE status='completed' AND success_count>0
                             AND filename ILIKE %s
                           ORDER BY processed_at DESC LIMIT 1""",
                        (pat,),
                    )
                    row = cur.fetchone()
                    if row:
                        return str(row[0]), row[1] or []
    except Exception as e:
        print(f'[reports] _find_best_job error: {e}')
    return None, []


def _conn_report():
    """Reuse the db_postgres pool for report extraction."""
    from db_postgres import _conn as _pg_conn
    return _pg_conn()


def _extract_report_from_tracker(job_id: str, results: list, month: str, year: int = None) -> dict | None:
    """
    Read the timeline deviation tracker output for a job and return
    the full monthly report dict matching History.jsx's data shape.
    """
    import openpyxl
    from collections import Counter

    # Find the tracker file path
    tracker_path = None
    for res in results:
        out_file = res.get('output_filename', '')
        if 'timeline_deviation_tracker' in out_file:
            candidate = os.path.join(_APP_ROOT, OUTPUT_FOLDER, job_id, out_file)
            if os.path.exists(candidate):
                tracker_path = candidate
                break

    if not tracker_path:
        print(f'[reports] No tracker file found for job {job_id}')
        return None

    try:
        wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
    except Exception as e:
        print(f'[reports] Cannot open {tracker_path}: {e}')
        return None

    # ── KPIs from Timeline Deviation sheet ────────────────────────────────────
    # Find the sheet by partial name match (handles "Timeline Deviation", "Sheet1", etc.)
    tdev_sheet_name = next(
        (s for s in wb.sheetnames if 'timeline deviation' in s.lower()),
        wb.sheetnames[0] if wb.sheetnames else None
    )
    if not tdev_sheet_name:
        print(f'[reports] No sheets found in tracker for job {job_id}')
        wb.close()
        return None
    print(f'[reports] Reading sheet "{tdev_sheet_name}" from {os.path.basename(tracker_path)}')
    ws = wb[tdev_sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f'[reports] Sheet "{tdev_sheet_name}" is empty for job {job_id}')
        wb.close()
        return None
    h = {str(c).strip() if c else '': i for i, c in enumerate(rows[0])}
    data = rows[1:]

    status_idx     = h.get('Status', -1)
    tdev_idx       = h.get('Timeline_deviation_flag', -1)
    plan_start_idx = next((h[k] for k in ('planned start date', 'Early Start', 'planned start') if k in h), -1)
    statuses = Counter(r[status_idx] for r in data if status_idx >= 0 and r[status_idx])
    tflags   = Counter(r[tdev_idx]   for r in data if tdev_idx   >= 0 and r[tdev_idx])

    def _is_milestone(r):
        # Primary: Status column says "Milestone - ..."
        if status_idx >= 0 and r[status_idx] and str(r[status_idx]).startswith('Milestone'):
            return True
        # Fallback: planned start is empty (catches rows where milestone flag was mis-set)
        if plan_start_idx >= 0 and not r[plan_start_idx]:
            return True
        return False

    total = sum(1 for r in data if not _is_milestone(r))
    # Activities on time — strictly on schedule (not early, not late)
    on_time     = statuses.get('Activity - On Time', 0)
    # Activities running late (finish > planned finish)
    on_late     = statuses.get('Activity - Late', 0) + statuses.get('Activity - Delay', 0)
    # Activities that started early (ahead of planned start)
    on_early    = statuses.get('Activity - Early Start', 0)
    # All achieved milestones regardless of on-time/early/delay
    milestone_done = (
        statuses.get('Milestone - Achieved - On Time', 0) +
        statuses.get('Milestone - Achieved - Early', 0) +
        statuses.get('Milestone - Achieved - Delay', 0)
    )
    not_started      = statuses.get('Activity - Not Started', 0)
    on_plan          = tflags.get('Duration On Plan (matches planned duration)', 0)
    dur_missing      = tflags.get('Duration Check - Planned Duration Missing', 0)
    in_progress      = tflags.get('Duration Check - In Progress', 0)

    # planned_duration column — try multiple names
    dur_col = next((h[k] for k in ('planned_duration', 'Planned Duration', 'planned duration') if k in h), -1)
    dur_vals = [r[dur_col] for r in data
                if dur_col >= 0 and isinstance(r[dur_col], (int, float)) and r[dur_col] > 0]
    avg_dur = round(sum(dur_vals) / len(dur_vals), 2) if dur_vals else 0
    max_dur = int(max(dur_vals)) if dur_vals else 0
    print(f'[reports] KPIs: total={total} on_time={on_time} milestone={milestone_done} '
          f'not_started={not_started} on_plan={on_plan} dur_missing={dur_missing}')

    # ── S-curves from discipline sheets ───────────────────────────────────────
    if year is None:
        year = datetime.now().year
    current_period = f"{month[:3]}-{str(year)[2:]}"
    scurves      = {}
    future_rows  = {}  # discipline → list of (period_label, planned_pct)

    for sname in wb.sheetnames:
        disc = None
        for sheet_key, disc_key in _SCURVE_SHEET_MAP.items():
            if sheet_key in sname:
                disc = disc_key
                break
        if not disc:
            continue

        ws2 = wb[sname]
        sc_rows = list(ws2.iter_rows(values_only=True))
        sc_h = [str(c).strip() if c else '' for c in sc_rows[0]]
        if 'Period' not in sc_h or 'Actual %' not in sc_h:
            continue

        period_idx = sc_h.index('Period')
        actual_idx = sc_h.index('Actual %')
        planned_idx = sc_h.index('Planned %') if 'Planned %' in sc_h else actual_idx

        found_current = False
        future_periods = []
        for sc_row in sc_rows[1:]:
            period = sc_row[period_idx]
            if not period:
                continue
            actual  = sc_row[actual_idx]
            planned = sc_row[planned_idx]
            if period == current_period:
                scurves[disc] = {
                    'planned': round(float(planned or 0), 2),
                    'actual':  round(float(actual  or 0), 2),
                    'label':   _SCURVE_SHEET_MAP.get(sname, disc),
                }
                found_current = True
            elif found_current and len(future_periods) < 4:
                future_periods.append((str(period), round(float(planned or 0), 2)))
        future_rows[disc] = future_periods

    wb.close()

    # Fill missing disciplines with zeros
    for disc in ('homeOffice', 'manufacturing', 'construction', 'projectMgmt', 'commissioning'):
        if disc not in scurves:
            scurves[disc] = {'planned': 0.0, 'actual': 0.0, 'label': disc}

    # ── futurePlanned ─────────────────────────────────────────────────────────
    future_planned = {}
    for disc, periods in future_rows.items():
        fp = {}
        for period_label, pct in periods:
            month_abbr = period_label[:3]  # e.g. "May" from "May-26"
            key = _PERIOD_TO_KEY.get(month_abbr)
            if key:
                fp[key] = pct
        if fp:
            future_planned[disc] = fp

    return {
        'file':               os.path.basename(tracker_path),
        'totalActivities':    total,
        'onTime':             on_time,
        'onLate':             on_late,
        'onEarly':            on_early,
        'milestoneAchieved':  milestone_done,
        'notStarted':         not_started,
        'onPlan':             on_plan,
        'durationMissing':    dur_missing,
        'inProgress':         in_progress,
        'avgPlannedDuration': avg_dur,
        'maxPlannedDuration': max_dur,
        'scurves':            scurves,
        'futurePlanned':      future_planned,
    }


@app.route('/api/reports/delayed-activities', methods=['GET'])
@token_required
def api_get_delayed_activities(current_user):
    """
    Read the most recent processed output tracker file directly and return
    every activity flagged as delayed, with numeric delay days.

    This bypasses the deviations DB entirely — data comes straight from the
    timeline deviation tracker Excel sheet so there is no possibility of
    field-name mismatches or stale DB data.

    Query params:
      min_days  — minimum delay days to include (default 1)
      limit     — max rows to return (default 200)
    """
    import openpyxl, json as _json
    from datetime import datetime as _dt

    try:
        min_days = max(0, int(request.args.get('min_days', 1)))
        limit    = min(500, int(request.args.get('limit', 200)))
    except (ValueError, TypeError):
        min_days, limit = 1, 200

    # ── Find the most recent successful job with a tracker file ───────────────
    role = current_user.get('role')
    company_id = current_user.get('company_id')
    if role in ('admin', 'company_admin', 'manager'):
        completed = pg_read_history_for_company(company_id=company_id, status='completed', limit=20)
    elif role == 'super_admin':
        completed = pg_read_history_for_company(status='completed', limit=20)
    else:
        completed = pg_read_history_for_company(user_id=current_user['id'], status='completed', limit=20)
    completed = [h for h in completed if h.get('success_count', 0) > 0]

    tracker_path = None
    job_filename  = ''
    job_processed_at = ''
    for job in completed:
        job_id = job.get('id', '')
        for result in (job.get('results') or []):
            if result.get('status') != 'success':
                continue
            out_file = result.get('output_filename', '')
            if 'timeline_deviation_tracker' in out_file.lower():
                candidate = os.path.join(_APP_ROOT, OUTPUT_FOLDER, job_id, out_file)
                if os.path.exists(candidate):
                    tracker_path     = candidate
                    job_filename      = job.get('filename', '')
                    job_processed_at  = job.get('processed_at', '')
                    break
        if tracker_path:
            break

    if not tracker_path:
        return jsonify({
            'source': 'none',
            'message': 'No processed tracker file found. Upload a project file first.',
            'activities': [],
        }), 200

    # ── Read the tracker sheet ────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
    except Exception as e:
        return jsonify({'error': f'Cannot open tracker file: {e}'}), 500

    sheet_name = next(
        (s for s in wb.sheetnames if 'timeline deviation' in s.lower()),
        wb.sheetnames[0] if wb.sheetnames else None,
    )
    if not sheet_name:
        wb.close()
        return jsonify({'source': tracker_path, 'activities': []}), 200

    ws    = wb[sheet_name]
    rows  = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return jsonify({'source': tracker_path, 'activities': []}), 200

    h = {str(c).strip() if c else '': i for i, c in enumerate(rows[0])}

    def _col(name, *aliases):
        """Return index of first matching column name (case-insensitive)."""
        for n in (name,) + aliases:
            for k, v in h.items():
                if k.lower() == n.lower():
                    return v
        return -1

    act_id_idx     = _col('Activity ID',   'activity_id',   'activityid')
    act_name_idx   = _col('Activity Name', 'activity_name', 'activityname')
    status_idx     = _col('Status',        'status')
    tdev_idx       = _col('Timeline_Deviation_Flag', 'timeline deviation flag', 'timeline_deviation_flag')
    start_dev_idx  = _col('Start_Delay Deviation',   'start_delay deviation',  'start_delay', 'start delay')
    dur_dev_idx    = _col('Duration_Deviation',       'duration deviation (days)', 'duration deviation', 'duration_deviation')
    plan_start_idx = _col('Early Start',  'planned start date', 'planned start', 'early planned start')
    plan_end_idx   = _col('Early Finish', 'planned end date',   'planned end',   'late planning')
    act_start_idx  = _col('Start',        'actual start date',  'actual start')
    act_end_idx    = _col('Finish',       'actual completion date', 'actual finish', 'actual completion')
    float_idx      = _col('Total Float',  'total float', 'float')

    today = _dt.now().date()

    def _safe_float(val):
        try:
            return float(val) if val is not None else 0.0
        except (TypeError, ValueError):
            return 0.0

    def _safe_date(val):
        if val is None:
            return None
        if hasattr(val, 'date'):
            return val.date()
        try:
            import pandas as _pd
            ts = _pd.to_datetime(val, errors='coerce')
            return None if _pd.isnull(ts) else ts.date()
        except Exception:
            return None

    def _fmt_date(val):
        d = _safe_date(val)
        return d.isoformat() if d else ''

    activities = []
    for row in rows[1:]:
        status_val = str(row[status_idx]   or '').strip() if status_idx >= 0 else ''
        tdev_val   = str(row[tdev_idx]     or '').strip() if tdev_idx   >= 0 else ''
        start_dev  = _safe_float(row[start_dev_idx] if start_dev_idx >= 0 else None)
        dur_dev    = _safe_float(row[dur_dev_idx]   if dur_dev_idx   >= 0 else None)

        # ── Recompute delay from today if numeric columns are zero ────────────
        if dur_dev == 0 and 'in progress' in (status_val + tdev_val).lower():
            p_end = _safe_date(row[plan_end_idx] if plan_end_idx >= 0 else None)
            if p_end and p_end < today:
                dur_dev = (today - p_end).days

        if start_dev == 0 and 'not started' in (status_val + tdev_val).lower():
            p_start = _safe_date(row[plan_start_idx] if plan_start_idx >= 0 else None)
            if p_start and p_start < today:
                start_dev = (today - p_start).days

        best_days = int(max(abs(start_dev), abs(dur_dev)))

        # Include if: numeric delay > 0 OR text signals a delay
        is_delayed = (
            best_days > 0
            or 'delay' in status_val.lower()
            or 'delay' in tdev_val.lower()
        )
        if not is_delayed:
            continue
        if best_days < min_days:
            continue

        act_id   = str(row[act_id_idx]   or '').strip() if act_id_idx   >= 0 else ''
        act_name = str(row[act_name_idx] or '').strip() if act_name_idx >= 0 else ''
        tf       = _safe_float(row[float_idx] if float_idx >= 0 else None)

        severity = 'High' if best_days >= 7 else ('Medium' if best_days >= 3 else 'Low')

        activities.append({
            'activity_id'  : act_id,
            'activity'     : act_name or act_id or 'Unknown',
            'days_overdue' : best_days,
            'severity'     : severity,
            'status'       : status_val,
            'timeline_flag': tdev_val,
            'planned_end'  : _fmt_date(row[plan_end_idx]   if plan_end_idx   >= 0 else None),
            'planned_start': _fmt_date(row[plan_start_idx] if plan_start_idx >= 0 else None),
            'actual_start' : _fmt_date(row[act_start_idx]  if act_start_idx  >= 0 else None),
            'actual_end'   : _fmt_date(row[act_end_idx]    if act_end_idx    >= 0 else None),
            'total_float'  : tf,
        })

    # Sort by most overdue first
    activities.sort(key=lambda a: a['days_overdue'], reverse=True)

    return jsonify({
        'source_file'   : os.path.basename(tracker_path),
        'source_job'    : job_filename,
        'processed_at'  : job_processed_at,
        'total_delayed' : len(activities),
        'activities'    : activities[:limit],
    }), 200


@app.route('/api/reports/not-started-activities', methods=['GET'])
@token_required
def api_get_not_started_activities(current_user):
    """
    Return activities with Status = 'Activity - Not Started' from the most recent tracker file.
    """
    import openpyxl
    from datetime import datetime as _dt

    limit = min(500, int(request.args.get('limit', 200)))

    _role = current_user.get('role')
    _company_id = current_user.get('company_id')
    if _role in ('admin', 'company_admin', 'manager'):
        completed = pg_read_history_for_company(company_id=_company_id, status='completed', limit=20)
    elif _role == 'super_admin':
        completed = pg_read_history_for_company(status='completed', limit=20)
    else:
        completed = pg_read_history_for_company(user_id=current_user['id'], status='completed', limit=20)
    completed = [h for h in completed if h.get('success_count', 0) > 0]

    tracker_path = None
    for job in completed:
        job_id = job.get('id', '')
        for result in (job.get('results') or []):
            if result.get('status') != 'success':
                continue
            out_file = result.get('output_filename', '')
            if 'timeline_deviation_tracker' in out_file.lower():
                candidate = os.path.join(_APP_ROOT, OUTPUT_FOLDER, job_id, out_file)
                if os.path.exists(candidate):
                    tracker_path = candidate
                    break
        if tracker_path:
            break

    if not tracker_path:
        return jsonify({'activities': []}), 200

    try:
        wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    sheet_name = next(
        (s for s in wb.sheetnames if 'timeline deviation' in s.lower()),
        wb.sheetnames[0] if wb.sheetnames else None,
    )
    if not sheet_name:
        wb.close()
        return jsonify({'activities': []}), 200

    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return jsonify({'activities': []}), 200

    h = {str(c).strip() if c else '': i for i, c in enumerate(rows[0])}

    def _col(name, *aliases):
        for n in (name,) + aliases:
            for k, v in h.items():
                if k.lower() == n.lower():
                    return v
        return -1

    act_id_idx     = _col('Activity ID',   'activity_id',   'activityid')
    act_name_idx   = _col('Activity Name', 'activity_name', 'activityname')
    status_idx     = _col('Status',        'status')
    plan_start_idx = _col('Early Start',   'planned start date', 'planned start')
    plan_end_idx   = _col('Early Finish',  'planned end date',   'planned end')
    float_idx      = _col('Total Float',   'total float', 'float')

    def _fmt_date(val):
        if val is None:
            return ''
        if hasattr(val, 'date'):
            return val.date().isoformat()
        try:
            import pandas as _pd
            ts = _pd.to_datetime(val, errors='coerce')
            return '' if _pd.isnull(ts) else ts.date().isoformat()
        except Exception:
            return str(val)

    activities = []
    for row in rows[1:]:
        status_val = str(row[status_idx] or '').strip() if status_idx >= 0 else ''
        if status_val != 'Activity - Not Started':
            continue
        act_id   = str(row[act_id_idx]   or '').strip() if act_id_idx   >= 0 else ''
        act_name = str(row[act_name_idx] or '').strip() if act_name_idx >= 0 else ''
        tf_raw   = row[float_idx] if float_idx >= 0 else None
        try:
            tf = float(tf_raw) if tf_raw is not None else None
        except (TypeError, ValueError):
            tf = None
        activities.append({
            'activity_id'  : act_id,
            'activity'     : act_name or act_id or 'Unknown',
            'planned_start': _fmt_date(row[plan_start_idx] if plan_start_idx >= 0 else None),
            'planned_end'  : _fmt_date(row[plan_end_idx]   if plan_end_idx   >= 0 else None),
            'total_float'  : tf,
        })

    return jsonify({'total': len(activities), 'activities': activities[:limit]}), 200


@app.route('/api/reports/monthly-trend', methods=['GET'])
@token_required
def api_get_monthly_trend(current_user):
    """
    Return monthly report KPIs for the History dashboard.
    Query params:
      year — default 2026
    Response: { months: [...], data: { MonthName: {...} }, year: 2026 }
    """
    try:
        year = int(request.args.get('year', 2026))
    except (ValueError, TypeError):
        year = 2026

    rows = get_monthly_reports(year=year, company_id=current_user.get('company_id'))
    months = [r['month'] for r in rows]
    data   = {r['month']: r['data'] for r in rows}
    return jsonify({'months': months, 'data': data, 'year': year}), 200


@app.route('/api/reports/monthly-trend/compute', methods=['POST'])
@token_required
def api_compute_monthly_trend(current_user):
    """
    Admin: re-compute monthly reports from processed output tracker files
    and save them to the monthly_reports table.
    Body (optional): { months: ['February','March','April'], year: 2026 }
    If months is omitted, all known months are computed.
    """
    if current_user.get('role') != 'admin':
        return jsonify({'error': 'Admin access required'}), 403

    body = request.get_json() or {}
    months_to_compute = body.get('months') or list(_MONTH_FILE_PATTERNS.keys())
    try:
        year = int(body.get('year', 2026))
    except (ValueError, TypeError):
        year = 2026

    results_log = []
    for month in months_to_compute:
        job_id, job_results = _find_best_job(month)
        if not job_id:
            results_log.append({'month': month, 'status': 'no_job_found'})
            continue
        report = _extract_report_from_tracker(job_id, job_results, month, year=year)
        if not report:
            results_log.append({'month': month, 'status': 'extraction_failed', 'job_id': job_id})
            continue
        ok = upsert_monthly_report(month, year, report, current_user.get('company_id'))
        results_log.append({
            'month': month,
            'status': 'ok' if ok else 'save_failed',
            'job_id': job_id,
            'totalActivities': report.get('totalActivities'),
        })

    return jsonify({'results': results_log}), 200


@app.route('/api/reports/recompute', methods=['POST'])
@token_required
def api_recompute_reports(current_user):
    """
    Recompute monthly reports from this company's processed files.
    Accessible to company_admin, admin, and super_admin.
    """
    role = current_user.get('role')
    if role not in ('admin', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Access denied'}), 403

    company_id = current_user.get('company_id')
    year = datetime.now().year

    # Find all completed jobs for this company
    results_log = []
    try:
        from db_postgres import _conn as _pg_conn
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                if company_id:
                    cur.execute(
                        """SELECT id, filename, results FROM history
                           WHERE status='completed' AND success_count>0
                             AND company_id = %s
                           ORDER BY processed_at DESC LIMIT 20""",
                        (str(company_id),),
                    )
                else:
                    cur.execute(
                        """SELECT id, filename, results FROM history
                           WHERE status='completed' AND success_count>0
                             AND company_id IS NULL
                           ORDER BY processed_at DESC LIMIT 20"""
                    )
                jobs = cur.fetchall()
    except Exception as e:
        return jsonify({'error': f'DB error: {e}'}), 500

    seen_months = set()
    for job_id, filename, job_results in jobs:
        month = _detect_month_from_filename(filename or '')
        if not month:
            month = _detect_month_from_tracker_dates(str(job_id), job_results or [], year)
        if not month or month in seen_months:
            continue
        seen_months.add(month)

        report = _extract_report_from_tracker(str(job_id), job_results or [], month, year=year)
        if not report:
            results_log.append({'month': month, 'status': 'extraction_failed', 'filename': filename})
            continue
        ok = upsert_monthly_report(month, year, report, company_id)
        results_log.append({
            'month': month,
            'status': 'ok' if ok else 'save_failed',
            'filename': filename,
            'totalActivities': report.get('totalActivities'),
        })

    return jsonify({'results': results_log, 'year': year}), 200


@app.route('/api/reports/monthly-trend', methods=['POST'])
@token_required
def api_upsert_monthly_trend(current_user):
    """
    Admin: create or update a month's report data.
    Body: { month: 'May', year: 2026, data: { totalActivities: ..., ... } }
    """
    if current_user.get('role') != 'admin':
        return jsonify({'error': 'Admin access required'}), 403

    body = request.get_json() or {}
    month = (body.get('month') or '').strip()
    data  = body.get('data')
    try:
        year = int(body.get('year', 2026))
    except (ValueError, TypeError):
        year = 2026

    if not month or not isinstance(data, dict):
        return jsonify({'error': 'month (string) and data (object) are required'}), 400

    ok = upsert_monthly_report(month, year, data, current_user.get('company_id'))
    if not ok:
        return jsonify({'error': 'Failed to save report'}), 500

    return jsonify({'status': 'ok', 'month': month, 'year': year}), 200


# ── Reports: Status Update (phase deviation + recovery plan) ─────────────────

def _compute_phase_deviation(scurves: dict, total: int, avg_dur: float, max_dur: float) -> dict:
    """
    Derive EPC phase status from discipline S-curve data.
    Maps: Engineering = homeOffice + projectMgmt average
          Procurement = manufacturing
          Construction = construction + commissioning average
    Returns per-phase dict with actual%, planned%, gap%, status, estimated days lost/gained.
    """
    def _avg(*keys):
        vals = [scurves.get(k, {}) for k in keys]
        actuals  = [v.get('actual',  0) for v in vals if v]
        planneds = [v.get('planned', 0) for v in vals if v]
        return (
            round(sum(actuals)  / len(actuals),  2) if actuals  else 0,
            round(sum(planneds) / len(planneds), 2) if planneds else 0,
        )

    def _days_from_gap(gap_pct, max_d):
        # Each 1% of S-curve gap ≈ max_duration / 50 days (calibrated heuristic)
        return max(0, round(abs(gap_pct) * max_d / 50))

    def _status(gap):
        if gap <= 3:  return 'On Track'
        if gap <= 10: return 'Watchlist'
        return 'Critical'

    phases = {}
    for name, keys in [('Engineering', ('homeOffice', 'projectMgmt')),
                        ('Procurement', ('manufacturing',)),
                        ('Construction', ('construction', 'commissioning'))]:
        actual, planned = _avg(*keys)
        gap = round(planned - actual, 2)   # positive = behind plan
        days = _days_from_gap(gap, max_dur)
        phases[name] = {
            'actual':       actual,
            'planned':      planned,
            'gap':          gap,
            'status':       _status(gap),
            'daysLost':     days if gap > 0 else 0,
            'daysGained':   _days_from_gap(-gap, max_dur) if gap < 0 else 0,
            'atRisk':       round(days * 0.85),
            'recovered':    round(days * 0.15),
        }

    # Overall totals
    total_lost     = sum(p['daysLost']   for p in phases.values())
    total_gained   = sum(p['daysGained'] for p in phases.values())
    total_at_risk  = sum(p['atRisk']     for p in phases.values())
    total_recovered= sum(p['recovered']  for p in phases.values())

    return {
        'phases': phases,
        'overall': {
            'totalDaysLost':     total_lost,
            'totalDaysGained':   total_gained,
            'netScheduleSlip':   total_lost - total_gained,
            'alreadyRecovered':  total_recovered,
            'stillAtRisk':       total_at_risk,
        }
    }


@app.route('/api/reports/status-update', methods=['GET'])
@token_required
def api_get_status_update(current_user):
    """
    Return dynamic status-update data for the History dashboard:
      - Phase-wise EPC deviation (actual vs planned from S-curves)
      - Overall days gained/lost/recovered/at-risk
      - Recovery plan scenarios (configurable via POST)
    Query params: year (default 2026), month (optional, defaults to latest)
    """
    try:
        year = int(request.args.get('year', 2026))
    except (ValueError, TypeError):
        year = 2026
    filter_month = (request.args.get('month') or '').strip()

    _cid = current_user.get('company_id')
    rows = get_monthly_reports(year=year, company_id=_cid)
    if not rows:
        return jsonify({'error': 'No monthly reports found'}), 404

    # Pick requested or latest month
    if filter_month:
        row = next((r for r in rows if r['month'].lower() == filter_month.lower()), rows[-1])
    else:
        row = rows[-1]

    d = row.get('data', {})
    scurves    = d.get('scurves', {})
    total      = d.get('totalActivities', 1)
    avg_dur    = d.get('avgPlannedDuration', 45)
    max_dur    = d.get('maxPlannedDuration', 290)

    deviation = _compute_phase_deviation(scurves, total, avg_dur, max_dur)

    # Recovery plan — load from company-specific DB file if admin has saved one, else compute defaults
    import json as _json
    _rp_name = f'recovery_plan_{_cid}.json' if _cid else 'recovery_plan.json'
    recovery_db_path = os.path.join(_APP_ROOT, 'database', _rp_name)
    try:
        with open(recovery_db_path, 'r') as f:
            recovery_plan = _json.load(f)
    except Exception:
        # Default recovery plan derived from construction gap
        construction = deviation['phases'].get('Construction', {})
        days_lost = construction.get('daysLost', 100)
        recovery_plan = {
            'steps': [
                {
                    'step': '01',
                    'title': 'Downstream Civil Compression',
                    'tag': 'Uses Float',
                    'description': (
                        'Re-plan downstream civil works at dramatically shorter durations using available schedule float. '
                        'Mechanical works targeted at 6 months instead of 9 months planned, absorbing the bulk of delay '
                        'before entering the compressor installation sequence.'
                    ),
                    'baseline': round(days_lost * 1.3),
                    'compressed': round(days_lost * 0.47),
                    'recovered': round(days_lost * 0.83),
                },
                {
                    'step': '02',
                    'title': 'Procurement Fast-Track',
                    'tag': 'Vendor Action',
                    'description': (
                        'Expedite long-lead items currently lagging. Parallel procurement streams for bulk materials '
                        'to close the gap vs baseline. Supplier mobilisation fast-tracked for critical packages.'
                    ),
                    'baseline': round(days_lost * 0.66),
                    'compressed': round(days_lost * 0.53),
                    'recovered': round(days_lost * 0.13),
                },
                {
                    'step': '03',
                    'title': 'Scope Re-Sequencing',
                    'tag': 'Needs Review',
                    'description': (
                        'Re-sequence non-critical construction activities to free up resources for the critical path. '
                        'If float exhausted, this step escalates to project planning for a formal recovery schedule submission.'
                    ),
                    'baseline': round(days_lost * 0.51),
                    'compressed': round(days_lost * 0.39),
                    'recovered': round(days_lost * 0.12),
                },
            ],
            'ai_note': (
                'Recovery plan is computed from schedule float analysis. If compression via float is not achievable, '
                'a more detailed review by project planning is needed. AI can embed feedback from such reviews—including '
                'actual resource availability, vendor confirmations, or re-sequencing decisions—to improve future scenario '
                'modelling accuracy.'
            ),
            'customised': False,
        }

    return jsonify({
        'month':        row['month'],
        'year':         year,
        'deviation':    deviation,
        'recoveryPlan': recovery_plan,
    }), 200


@app.route('/api/reports/status-update', methods=['POST'])
@token_required
def api_save_recovery_plan(current_user):
    """Admin: save a customised recovery plan for the current project."""
    if current_user.get('role') not in ('admin', 'manager'):
        return jsonify({'error': 'Admin or manager access required'}), 403

    body = request.get_json() or {}
    steps = body.get('steps')
    ai_note = body.get('ai_note', '')
    if not isinstance(steps, list) or not steps:
        return jsonify({'error': 'steps array is required'}), 400

    import json as _json
    _cid = current_user.get('company_id')
    _rp_name = f'recovery_plan_{_cid}.json' if _cid else 'recovery_plan.json'
    recovery_db_path = os.path.join(_APP_ROOT, 'database', _rp_name)
    payload = {'steps': steps, 'ai_note': ai_note, 'customised': True}
    try:
        with open(recovery_db_path, 'w') as f:
            _json.dump(payload, f, indent=2)
    except Exception as e:
        return jsonify({'error': f'Failed to save: {e}'}), 500

    return jsonify({'status': 'ok'}), 200


@app.route('/api/reports/ai-insight', methods=['POST'])
@token_required
def api_ai_insight(current_user):
    """Generate a concise AI narrative for dashboard contexts."""
    if azure_ai_client is None or not AZURE_ANTHROPIC_DEPLOYMENT:
        return jsonify({'error': 'AI service unavailable'}), 503

    body = request.get_json(silent=True) or {}
    context = str(body.get('context', '')).strip()

    try:
        if context == 'intelligence_hub':
            slippage_count    = int(body.get('slippage_count', 0))
            improvement_count = int(body.get('improvement_count', 0))
            worst_slip_days   = int(body.get('worst_slip_days', 0))
            best_gain_days    = int(body.get('best_gain_days', 0))
            prompt = (
                f"You are a senior EPC project controls analyst. Write exactly 2 sentences: "
                f"The project has {slippage_count} activities slipping (worst: +{worst_slip_days} days overdue) "
                f"and {improvement_count} activities running ahead (best: -{best_gain_days} days saved). "
                f"State where the project stands overall and the single most important action the team should take now."
            )
        elif context == 'executive_brief':
            month             = str(body.get('month', ''))
            total_activities  = int(body.get('total_activities', 0))
            days_lost         = int(body.get('days_lost', 0))
            already_recovered = int(body.get('already_recovered', 0))
            still_at_risk     = int(body.get('still_at_risk', 0))
            top_action        = str(body.get('top_recovery_action', '')).strip()

            period      = month or 'the current reporting period'
            act_clause  = f"{total_activities:,} activities tracked" if total_activities > 0 else "activities being tracked"
            rec_clause  = (
                f"{days_lost}d lost, {already_recovered}d recovered, {still_at_risk}d still at risk"
                if days_lost > 0 or already_recovered > 0
                else "deviation quantification pending upload of Planned % data"
            )
            action_clause = f" Priority action: {top_action}." if top_action else ""

            prompt = (
                f"Write a 3-sentence executive project brief for an EPC PMO dashboard. "
                f"Reporting period: {period}. {act_clause}. Schedule position: {rec_clause}.{action_clause} "
                f"Write as a confident analyst who has reviewed the data — do not ask questions, "
                f"do not say you need more information. State the schedule position, what recovery action "
                f"is in progress, and what the team must prioritise. Output only the brief, nothing else."
            )
        else:
            month             = str(body.get('month', ''))
            total_activities  = int(body.get('total_activities', 0))
            days_lost         = int(body.get('days_lost', 0))
            already_recovered = int(body.get('already_recovered', 0))
            still_at_risk     = int(body.get('still_at_risk', 0))
            phase_actuals     = body.get('phase_actuals', {}) or {}
            eng  = float(phase_actuals.get('engineering', 0))
            proc = float(phase_actuals.get('procurement', 0))
            con  = float(phase_actuals.get('construction', 0))
            prompt = (
                f"You are a senior EPC project controls analyst. Write 2-3 sentences summarising: "
                f"schedule position for {month} ({total_activities:,} activities total), "
                f"{days_lost} days lost vs {already_recovered} days already recovered, "
                f"{still_at_risk} days still at risk, "
                f"engineering at {eng:.1f}%, procurement at {proc:.1f}%, construction at {con:.1f}%. "
                f"End with the single most important action needed to protect the finish date."
            )

        msg_obj = azure_ai_client.messages.create(
            model=AZURE_ANTHROPIC_DEPLOYMENT,
            max_tokens=300,
            system=(
                "You are a senior EPC project controls analyst generating dashboard copy. "
                "Always produce the requested output immediately using the data given. "
                "Never ask for clarification, never request more data, never say you need additional information. "
                "If a metric is zero or absent, acknowledge it briefly and move on. Be concise and direct."
            ),
            messages=[{'role': 'user', 'content': prompt}],
        )
        text = '\n'.join(
            getattr(b, 'text', '') for b in (msg_obj.content or [])
            if getattr(b, 'type', '') == 'text'
        ).strip()

        return jsonify({'ai_note': text}), 200

    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500


# ── Schedule JSON: Milestones, Critical Path, One Pager ───────────────────────
import json as _json


def _sync_schedule_json_from_xlsx(xlsx_path: str, company_id=None):
    """
    Parse the uploaded EPC schedule XLSX (columns: ActivityID, Activity Name,
    Start, Finish, Late Start, Late Finish, Early Start, Early Finish,
    Phase Definition, MS Definition, CP (LF), CP (BL))
    and update milestones.json, critical_path.json, and one_pager.json in-place.

    Called automatically after every successful upload.
    Only the date/status fields are overwritten — rich metadata (description,
    notes, responsible, risk_level) is preserved from the existing JSON.
    """
    try:
        import openpyxl as _xl
        from datetime import datetime as _dt
    except ImportError:
        print('[schedule-sync] openpyxl not installed — skipping schedule JSON sync')
        return

    if not xlsx_path or not os.path.exists(xlsx_path):
        print(f'[schedule-sync] File not found: {xlsx_path}')
        return

    try:
        wb = _xl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        print(f'[schedule-sync] Cannot open {xlsx_path}: {e}')
        return

    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print('[schedule-sync] Workbook is empty')
        return

    header = [str(c).strip() if c else '' for c in rows[0]]

    def _col(name):
        try:
            return header.index(name)
        except ValueError:
            return -1

    i_act   = _col('ActivityID')
    i_name  = _col('Activity Name')
    i_fin   = _col('Finish')
    i_lf    = _col('Late Finish')
    i_ef    = _col('Early Finish')
    i_es    = _col('Early Start')
    i_ms    = _col('MS Definition')
    i_cp    = _col('CP (LF)')

    if i_act < 0 or i_fin < 0 or i_lf < 0 or i_ef < 0:
        print(f'[schedule-sync] Unexpected columns {header} — skipping')
        return

    def _parse_date(val):
        """Return (datetime, is_actual) for a cell value."""
        if val is None:
            return None, False
        if isinstance(val, _dt):
            return val, False
        s = str(val).strip()
        if s.endswith(' A'):
            s = s[:-2].strip()
            try:
                return _dt.strptime(s, '%d-%b-%y'), True
            except ValueError:
                pass
        for fmt in ('%d-%b-%y', '%Y-%m-%d', '%d/%m/%Y'):
            try:
                return _dt.strptime(s, fmt), False
            except ValueError:
                pass
        return None, False

    def _iso(dt):
        return dt.strftime('%Y-%m-%d') if dt else None

    def _label(dt):
        return dt.strftime('%-d %b %Y').lstrip('0') if dt else None

    # ── Collect MS and CP rows ────────────────────────────────────────────────
    ms_rows = []
    cp_rows = []
    for r in rows[1:]:
        if len(r) <= max(i_act, i_fin, i_lf, i_ef):
            continue
        act_id = r[i_act]
        if not act_id:
            continue
        act_id = str(act_id).strip()
        name   = str(r[i_name]).strip() if i_name >= 0 and r[i_name] else ''

        fin_dt,  fin_actual  = _parse_date(r[i_fin])
        lf_dt,   _           = _parse_date(r[i_lf])
        ef_dt,   ef_actual   = _parse_date(r[i_ef])
        es_dt,   _           = _parse_date(r[i_es] if i_es >= 0 else None)

        is_ms = i_ms >= 0 and r[i_ms] and str(r[i_ms]).strip() == 'MS'
        is_cp = i_cp >= 0 and r[i_cp] and str(r[i_cp]).strip() == 'CP'

        if is_ms:
            done = fin_actual or ef_actual
            ms_rows.append({
                'activity_id': act_id,
                'name': name,
                'forecast_finish': _iso(fin_dt) or _iso(ef_dt),
                'early_finish':    _iso(ef_dt),
                'done': bool(done),
            })
        if is_cp:
            vd = int((lf_dt - ef_dt).days) if lf_dt and ef_dt else 0

            # Durations from early start → early finish (baseline) and start → finish (current)
            bl_dur = int((ef_dt - es_dt).days) if ef_dt and es_dt else None
            # Current duration: Finish - Early Start (or just baseline if no separate start given)
            cur_dur = bl_dur  # start col not always populated; keep symmetric for now

            cp_rows.append({
                'activity_id':     act_id,
                'name':            name,
                'forecast_finish': _label(lf_dt),
                'baseline_finish': _iso(ef_dt),
                'variance_days':   vd,
                'baseline_duration': bl_dur,
                'current_duration':  cur_dur,
                'float_days': 0,  # float not in this sheet; preserve existing
            })

    print(f'[schedule-sync] Found {len(ms_rows)} MS rows, {len(cp_rows)} CP rows')

    # ── Update milestones.json ────────────────────────────────────────────────
    ms_file = _read_schedule_json('milestones.json', company_id)
    if ms_file:
        existing_ms = {m['id']: m for m in ms_file.get('milestones', [])}

        # Build lookup: activity_id → xlsx row (MS rows first, then CP rows as fallback)
        ms_by_actid = {r['activity_id']: r for r in ms_rows}
        # CP rows provide dates for RFSU/PAC milestones flagged CP not MS
        cp_by_actid_ms = {}
        for r in cp_rows:
            cp_by_actid_ms[r['activity_id']] = {
                'activity_id':    r['activity_id'],
                'name':           r['name'],
                'forecast_finish': r['baseline_finish'],  # CP: early finish = baseline = our milestone forecast
                'early_finish':   r['baseline_finish'],
                'done':           False,
            }
        # Name-based lookup
        ms_by_name = {r['name'].lower(): r for r in ms_rows}

        changed = 0
        for ms_id, entry in existing_ms.items():
            xlsx_row = None
            aid = entry.get('activity_id')
            # 1. Direct activity_id match in MS rows
            if aid and aid in ms_by_actid:
                xlsx_row = ms_by_actid[aid]
            # 2. Fallback: CP rows (for RFSU, PAC which carry CP flag not MS)
            if not xlsx_row and aid and aid in cp_by_actid_ms:
                cp_r = cp_by_actid_ms[aid]
                # For CP-flagged milestones, forecast comes from Late Finish (already in cp_rows)
                matching_cp = next((r for r in cp_rows if r['activity_id'] == aid), None)
                if matching_cp:
                    xlsx_row = {
                        'activity_id':    aid,
                        'name':           matching_cp['name'],
                        'forecast_finish': matching_cp['baseline_finish'],  # early_finish = baseline
                        'early_finish':    matching_cp['baseline_finish'],
                        'done':            False,
                    }
                    # forecast_finish for the milestone should be the late finish (current forecast)
                    # The cp_rows store forecast_finish as the LF label — convert back
                    try:
                        lf_label = matching_cp['forecast_finish']
                        lf_dt = _dt.strptime(lf_label, '%d %b %Y') if lf_label else None
                        if lf_dt:
                            xlsx_row['forecast_finish'] = lf_dt.strftime('%Y-%m-%d')
                    except Exception:
                        pass
            # 3. Fallback: label substring match
            if not xlsx_row:
                label = entry.get('label', '').lower()
                for xname, xrow in ms_by_name.items():
                    if xname and label and (xname[:20] in label or label[:20] in xname):
                        xlsx_row = xrow
                        break
            if not xlsx_row:
                continue
            # Only update date fields — preserve metadata
            if xlsx_row.get('forecast_finish'):
                entry['forecast_finish'] = xlsx_row['forecast_finish']
            if xlsx_row.get('early_finish'):
                entry['early_finish'] = xlsx_row['early_finish']
            entry['done'] = xlsx_row['done']
            changed += 1

        ms_file['_meta']['version'] = _dt.now().strftime('%Y-%m-%d')
        _write_schedule_json('milestones.json', ms_file, company_id)
        print(f'[schedule-sync] Updated {changed}/{len(existing_ms)} milestones')

    # ── Update critical_path.json ─────────────────────────────────────────────
    cp_file = _read_schedule_json('critical_path.json', company_id)
    if cp_file:
        cp_by_actid = {r['activity_id']: r for r in cp_rows}
        changed = 0
        for entry in cp_file.get('activities', []):
            xlsx_row = cp_by_actid.get(entry['id'])
            if not xlsx_row:
                continue
            entry['forecast_finish']   = xlsx_row['forecast_finish']
            entry['baseline_finish']   = xlsx_row['baseline_finish']
            entry['variance_days']     = xlsx_row['variance_days']
            if xlsx_row['baseline_duration'] is not None:
                entry['baseline_duration'] = xlsx_row['baseline_duration']
                entry['current_duration']  = xlsx_row['current_duration']
            # float_days: preserve existing (not in this sheet)
            changed += 1

        cp_file['_meta']['version'] = _dt.now().strftime('%Y-%m-%d')
        _write_schedule_json('critical_path.json', cp_file, company_id)
        print(f'[schedule-sync] Updated {changed}/{len(cp_file["activities"])} CP activities')

    # ── Update one_pager.json (MC/RFSU/PAC dates + status card) ──────────────
    op_file = _read_schedule_json('one_pager.json', company_id)
    if op_file and ms_file:
        # Re-read just-written milestones for the 3 key dates
        ms_map = {m['id']: m for m in ms_file.get('milestones', [])}

        mc   = ms_map.get('MC')
        rfsu = ms_map.get('RFSU')
        pac  = ms_map.get('PAC')

        if mc and mc.get('forecast_finish') and mc.get('early_finish'):
            mc_forecast  = mc['forecast_finish']
            mc_baseline  = mc['early_finish']
            try:
                mc_slip = int((_dt.strptime(mc_forecast, '%Y-%m-%d') - _dt.strptime(mc_baseline, '%Y-%m-%d')).days)
            except Exception:
                mc_slip = None

            sc = op_file.get('status_cards', {})
            dgl = sc.get('days_gained_lost', {})
            dgl['mc_baseline'] = _dt.strptime(mc_baseline, '%Y-%m-%d').strftime('%-d %b %Y').lstrip('0')
            dgl['mc_forecast'] = _dt.strptime(mc_forecast, '%Y-%m-%d').strftime('%-d %b %Y').lstrip('0')
            if mc_slip is not None:
                sign = '+' if mc_slip >= 0 else ''
                dgl['mc_slip_display'] = f'{sign}{mc_slip}d slip to MC'
                sc['at_risk']['mc_slip_display'] = f'{sign}{mc_slip}d MC'

            rsum = op_file.get('recovery_summary', {})
            rsum['mc_baseline'] = dgl['mc_baseline']
            rsum['mc_forecast'] = dgl['mc_forecast']
            if mc_slip is not None:
                rsum['still_at_risk'] = mc_slip

        if pac and pac.get('forecast_finish') and pac.get('early_finish'):
            pac_forecast = pac['forecast_finish']
            pac_baseline = pac['early_finish']
            rsum = op_file.get('recovery_summary', {})
            rsum['pac_baseline'] = _dt.strptime(pac_baseline, '%Y-%m-%d').strftime('%-d %b %Y').lstrip('0')
            rsum['pac_forecast'] = _dt.strptime(pac_forecast, '%Y-%m-%d').strftime('%-d %b %Y').lstrip('0')

        # Update Net position waterfall bar to match actual MC slip
        if mc_slip is not None:
            for wf in op_file.get('waterfall', []):
                if wf.get('name') == 'Net position':
                    wf['bar']   = mc_slip
                    wf['delta'] = f'{sign}{mc_slip}d'
                    break

        op_file['_meta']['version'] = _dt.now().strftime('%Y-%m-%d')
        _write_schedule_json('one_pager.json', op_file, company_id)
        print(f'[schedule-sync] Updated one_pager.json (MC slip={mc_slip}d)')

    print('[schedule-sync] Done')

def _read_schedule_json(filename, company_id=None):
    path = os.path.join(_get_knowledgebase_folder(company_id), filename)
    if not os.path.exists(path):
        return None
    with open(path, 'r', encoding='utf-8') as f:
        return _json.load(f)

def _write_schedule_json(filename, data, company_id=None):
    path = os.path.join(_get_knowledgebase_folder(company_id), filename)
    with open(path, 'w', encoding='utf-8') as f:
        _json.dump(data, f, indent=2, ensure_ascii=False)


@app.route('/api/schedule/milestones', methods=['GET'])
@token_required
def api_get_milestones(current_user):
    """Return all project milestones from Knowledgebase/milestones.json."""
    data = _read_schedule_json('milestones.json', current_user.get('company_id'))
    if data is None:
        return jsonify({'error': 'milestones.json not found in Knowledgebase'}), 404
    return jsonify(data), 200


@app.route('/api/schedule/milestones', methods=['PUT'])
@token_required
def api_update_milestones(current_user):
    """Admin/manager: replace or patch milestones.json."""
    if current_user.get('role') not in ('admin', 'manager'):
        return jsonify({'error': 'Admin or manager access required'}), 403
    cid = current_user.get('company_id')
    body = request.get_json(silent=True) or {}
    existing = _read_schedule_json('milestones.json', cid) or {}
    if 'milestones' in body:
        existing['milestones'] = body['milestones']
    if '_meta' in body:
        existing['_meta'] = body['_meta']
    _write_schedule_json('milestones.json', existing, cid)
    return jsonify({'status': 'ok', 'count': len(existing.get('milestones', []))}), 200


@app.route('/api/schedule/milestones/<ms_id>', methods=['PATCH'])
@token_required
def api_patch_milestone(current_user, ms_id):
    """Admin/manager: update a single milestone by id."""
    if current_user.get('role') not in ('admin', 'manager'):
        return jsonify({'error': 'Admin or manager access required'}), 403
    cid = current_user.get('company_id')
    body = request.get_json(silent=True) or {}
    data = _read_schedule_json('milestones.json', cid)
    if data is None:
        return jsonify({'error': 'milestones.json not found'}), 404
    updated = False
    for ms in data.get('milestones', []):
        if ms.get('id') == ms_id:
            ms.update({k: v for k, v in body.items() if k != 'id'})
            updated = True
            break
    if not updated:
        return jsonify({'error': f'Milestone {ms_id} not found'}), 404
    _write_schedule_json('milestones.json', data, cid)
    return jsonify({'status': 'ok', 'id': ms_id}), 200


@app.route('/api/schedule/critical-path', methods=['GET'])
@token_required
def api_get_critical_path(current_user):
    """Return all CP activities from Knowledgebase/critical_path.json."""
    data = _read_schedule_json('critical_path.json', current_user.get('company_id'))
    if data is None:
        return jsonify({'error': 'critical_path.json not found in Knowledgebase'}), 404
    return jsonify(data), 200


@app.route('/api/schedule/critical-path', methods=['PUT'])
@token_required
def api_update_critical_path(current_user):
    """Admin/manager: replace CP activities list."""
    if current_user.get('role') not in ('admin', 'manager'):
        return jsonify({'error': 'Admin or manager access required'}), 403
    cid = current_user.get('company_id')
    body = request.get_json(silent=True) or {}
    existing = _read_schedule_json('critical_path.json', cid) or {}
    if 'activities' in body:
        existing['activities'] = body['activities']
    if '_meta' in body:
        existing['_meta'] = body['_meta']
    _write_schedule_json('critical_path.json', existing, cid)
    return jsonify({'status': 'ok', 'count': len(existing.get('activities', []))}), 200


@app.route('/api/schedule/critical-path/<act_id>', methods=['PATCH'])
@token_required
def api_patch_cp_activity(current_user, act_id):
    """Admin/manager: update a single CP activity by id."""
    if current_user.get('role') not in ('admin', 'manager'):
        return jsonify({'error': 'Admin or manager access required'}), 403
    cid = current_user.get('company_id')
    body = request.get_json(silent=True) or {}
    data = _read_schedule_json('critical_path.json', cid)
    if data is None:
        return jsonify({'error': 'critical_path.json not found'}), 404
    updated = False
    for act in data.get('activities', []):
        if act.get('id') == act_id:
            act.update({k: v for k, v in body.items() if k != 'id'})
            updated = True
            break
    if not updated:
        return jsonify({'error': f'Activity {act_id} not found'}), 404
    _write_schedule_json('critical_path.json', data, cid)
    return jsonify({'status': 'ok', 'id': act_id}), 200


@app.route('/api/schedule/one-pager', methods=['GET'])
@token_required
def api_get_one_pager(current_user):
    """Return One Pager data (waterfall, status cards) from Knowledgebase/one_pager.json."""
    data = _read_schedule_json('one_pager.json', current_user.get('company_id'))
    if data is None:
        return jsonify({'error': 'one_pager.json not found in Knowledgebase'}), 404
    return jsonify(data), 200


@app.route('/api/schedule/one-pager', methods=['PUT'])
@token_required
def api_update_one_pager(current_user):
    """Admin/manager: replace one_pager.json content."""
    if current_user.get('role') not in ('admin', 'manager'):
        return jsonify({'error': 'Admin or manager access required'}), 403
    cid = current_user.get('company_id')
    body = request.get_json(silent=True) or {}
    existing = _read_schedule_json('one_pager.json', cid) or {}
    for key in ('waterfall', 'status_cards', 'summary_text', 'recovery_summary', '_meta'):
        if key in body:
            existing[key] = body[key]
    _write_schedule_json('one_pager.json', existing, cid)
    return jsonify({'status': 'ok'}), 200


# ── Admin: Chat History ────────────────────────────────────────────────────────

@app.route('/api/admin/chat-history', methods=['GET'])
@token_required
def admin_get_chat_history(current_user):
    """
    Admin-only: return chat history scoped to the admin's company.
    super_admin can see all or pass ?company_id= to filter.
    Query params:
      user_id    — filter to a specific user (must be in the same company)
      route      — '/api/chat' or '/api/ai/chat'
      limit      — default 100, max 500
      offset     — default 0
      company_id — super_admin only: filter to a specific company
    """
    role = current_user.get('role')
    if role not in ('admin', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Admin access required'}), 403

    user_id = request.args.get('user_id') or None
    route   = request.args.get('route') or None
    try:
        limit  = min(int(request.args.get('limit',  100)), 500)
        offset = max(int(request.args.get('offset', 0)),   0)
    except (ValueError, TypeError):
        limit, offset = 100, 0

    if role == 'super_admin':
        filter_company = request.args.get('company_id') or None
        if filter_company:
            all_users = read_db(USERS_DB)
            company_user_ids = [u['id'] for u in all_users if u.get('company_id') == filter_company]
            rows = get_chat_history(user_ids=company_user_ids, limit=limit, offset=offset, route=route)
        else:
            rows = get_chat_history(user_id=user_id, limit=limit, offset=offset, route=route)
    else:
        company_id = current_user.get('company_id')
        all_users = read_db(USERS_DB)
        company_user_ids = [u['id'] for u in all_users if u.get('company_id') == company_id]
        if user_id and user_id not in company_user_ids:
            return jsonify({'error': 'User not in your company'}), 403
        rows = get_chat_history(user_id=user_id, user_ids=(None if user_id else company_user_ids),
                                limit=limit, offset=offset, route=route)

    stats = get_chat_stats()
    return jsonify({'history': rows, 'stats': stats, 'count': len(rows)}), 200


@app.route('/api/admin/chat-history/stats', methods=['GET'])
@token_required
def admin_chat_stats(current_user):
    """Admin-only: aggregate chat usage statistics."""
    if current_user.get('role') not in ('admin', 'company_admin', 'super_admin'):
        return jsonify({'error': 'Admin access required'}), 403
    return jsonify(get_chat_stats()), 200


@app.route('/api/me/chat-history', methods=['GET'])
@token_required
def my_chat_history(current_user):
    """Any user: view their own chat history."""
    try:
        limit  = min(int(request.args.get('limit',  50)), 200)
        offset = max(int(request.args.get('offset', 0)),  0)
    except (ValueError, TypeError):
        limit, offset = 50, 0

    route = request.args.get('route') or None
    rows  = get_chat_history(user_id=current_user['id'], limit=limit, offset=offset, route=route)
    return jsonify({'history': rows, 'count': len(rows)}), 200


# ==================== THETA SHEETS (live spreadsheet + dashboard sync) ====================

def _sheet_company_guard(current_user, sheet_record):
    if not sheet_record:
        return False
    user_cid = str(current_user.get('company_id') or '')
    sheet_cid = str(sheet_record.get('company_id') or '')
    if current_user.get('role') == 'super_admin':
        return True
    return bool(user_cid and user_cid == sheet_cid)


@app.route('/api/sheets/active', methods=['GET'])
@token_required
def get_active_sheet(current_user):
    cid = current_user.get('company_id')
    if not cid:
        return jsonify({'error': 'No company associated with this account'}), 400
    sheet = get_sheet_by_company(str(cid))
    if not sheet:
        return jsonify({'error': 'No active sheet found'}), 404
    return jsonify(sheet)


@app.route('/api/sheets/active', methods=['POST'])
@token_required
def create_active_sheet(current_user):
    cid = current_user.get('company_id')
    if not cid:
        return jsonify({'error': 'No company associated with this account'}), 400
    data = request.get_json() or {}
    name = (data.get('name') or 'Theta Sheets').strip()
    sheet_data = data.get('data')
    if not isinstance(sheet_data, dict):
        return jsonify({'error': 'data object is required'}), 400
    try:
        sheet = upsert_sheet(str(cid), name, sheet_data)
        errors = validate_sheet_data(sheet_data)
        sheet['validation'] = {'errors': errors, 'errorCount': len(errors)}
        return jsonify(sheet), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sheets/<sheet_id>', methods=['GET'])
@token_required
def get_sheet_by_id_route(current_user, sheet_id):
    sheet = get_sheet_by_id(sheet_id)
    if not _sheet_company_guard(current_user, sheet):
        return jsonify({'error': 'Sheet not found'}), 404
    return jsonify(sheet)


@app.route('/api/sheets/<sheet_id>', methods=['PUT', 'PATCH'])
@token_required
def update_sheet_route(current_user, sheet_id):
    sheet = get_sheet_by_id(sheet_id)
    if not _sheet_company_guard(current_user, sheet):
        return jsonify({'error': 'Sheet not found'}), 404
    body = request.get_json() or {}
    name = (body.get('name') or sheet.get('name') or 'Theta Sheets').strip()
    sheet_data = body.get('data')
    if not isinstance(sheet_data, dict):
        return jsonify({'error': 'data object is required'}), 400
    expected_version = body.get('version')
    try:
        updated = upsert_sheet(
            str(sheet['company_id']),
            name,
            sheet_data,
            expected_version=expected_version,
        )
        errors = validate_sheet_data(sheet_data)
        updated['validation'] = {'errors': errors, 'errorCount': len(errors)}
        return jsonify(updated)
    except ValueError as e:
        if str(e) == 'version_conflict':
            return jsonify({'error': 'Sheet was updated by another session. Please reload.'}), 409
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sheets/<sheet_id>/metrics', methods=['GET'])
@token_required
def get_sheet_metrics(current_user, sheet_id):
    sheet = get_sheet_by_id(sheet_id)
    if not _sheet_company_guard(current_user, sheet):
        return jsonify({'error': 'Sheet not found'}), 404
    data = sheet.get('data') or {}
    metrics = compute_metrics_from_sheet(data if isinstance(data, dict) else {})
    return jsonify({
        'sheet_id': sheet['id'],
        'version': sheet.get('version'),
        'updated_at': sheet.get('updated_at'),
        'metrics': metrics,
    })


@app.route('/api/sheets/active/metrics', methods=['GET'])
@token_required
def get_active_sheet_metrics(current_user):
    cid = current_user.get('company_id')
    if not cid:
        return jsonify({'error': 'No company associated with this account'}), 400
    sheet = get_sheet_by_company(str(cid))
    if not sheet:
        return jsonify({'error': 'No active sheet found'}), 404
    data = sheet.get('data') or {}
    metrics = compute_metrics_from_sheet(data if isinstance(data, dict) else {})
    return jsonify({
        'sheet_id': sheet['id'],
        'version': sheet.get('version'),
        'updated_at': sheet.get('updated_at'),
        'metrics': metrics,
    })


@app.route('/api/sheets/active/events', methods=['GET'])
@token_required
def active_sheet_events(current_user):
    """Server-Sent Events stream for live dashboard updates."""
    cid = current_user.get('company_id')
    if not cid:
        return jsonify({'error': 'No company associated with this account'}), 400

    company_id = str(cid)

    @stream_with_context
    def generate():
        event, payload = subscribe(company_id)
        token = (event, payload)
        try:
            yield f'data: {json.dumps({"type": "connected"})}\n\n'
            while True:
                if event.wait(timeout=25):
                    msg = payload.get('event')
                    event.clear()
                    if msg:
                        yield f'data: {json.dumps(msg)}\n\n'
                else:
                    yield f'data: {json.dumps({"type": "heartbeat"})}\n\n'
        finally:
            unsubscribe(company_id, token)

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection': 'keep-alive',
        },
    )

_seed_recovery_narrative()

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_react(path):
    dist_dir = os.path.join(_APP_ROOT, 'frontend', 'dist')
    full_path = os.path.join(dist_dir, path)
    if path != "" and os.path.exists(full_path) and os.path.isfile(full_path):
        return send_from_directory(dist_dir, path)
    return send_from_directory(dist_dir, 'index.html')

if __name__ == '__main__':
    def _env_bool(name: str, default: bool = False) -> bool:
        raw = str(os.getenv(name, str(default))).strip().lower()
        return raw in ('1', 'true', 'yes', 'y', 'on')

    try:
        backend_port = int(os.getenv('BACKEND_PORT', '8000'))
    except Exception:
        backend_port = 8000

    debug_mode = _env_bool('FLASK_DEBUG', False)
    use_waitress = _env_bool('USE_WAITRESS', False)
    print("=" * 60)
    print("PMO Backend Server Starting...")
    print(f"Server: http://localhost:{backend_port}")
    print(f"API Docs: http://localhost:{backend_port}/api/health")
    print(f"Mode: {'waitress' if use_waitress else 'flask-dev'}")
    print("=" * 60)

    if use_waitress:
        try:
            from waitress import serve

            try:
                waitress_threads = max(2, int(os.getenv('WAITRESS_THREADS', '8')))
            except Exception:
                waitress_threads = 8

            try:
                waitress_channel_timeout = max(120, int(os.getenv('WAITRESS_CHANNEL_TIMEOUT_SEC', '900')))
            except Exception:
                waitress_channel_timeout = 900

            print(
                f"Waitress config: threads={waitress_threads}, "
                f"channel_timeout={waitress_channel_timeout}s"
            )
            serve(
                app,
                host='0.0.0.0',
                port=backend_port,
                threads=waitress_threads,
                channel_timeout=waitress_channel_timeout,
            )
        except Exception as waitress_error:
            print(f"[WARN] Waitress start failed, falling back to Flask dev server: {waitress_error}")
            app.run(debug=debug_mode, port=backend_port, host='0.0.0.0', use_reloader=False)
    else:
        app.run(debug=debug_mode, port=backend_port, host='0.0.0.0', use_reloader=False)