#!/usr/bin/env python3
"""
Project Management Activity Timeline Tracker Generator
=========================================================
Processes Borouge Excel files - 'Project Mangement' sheet
with stage gate milestones and EP/LP/F/A date tracking.

Column Order (Output):
Activity ID | Activity Name | Stage Gate | Early Planning | Late Planning |
Actual Date | Duration Deviation | Timeline Flag

Stage Gates: Start, IDCs, IDCc, IFR, RCC, IFA, RCA, IFD/IFC

Author: Claude
Date: 2026-02-14
Version: 1.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys
from pathlib import Path
from collections import defaultdict


# ============================================================================
# DEFAULT FILE PATHS
# ============================================================================
DEFAULT_INPUT_FILE = r"2026.01.30_Borouge EU3 H2 Extraction Project PMS-Rev1 dates (1).xlsx"
DEFAULT_OUTPUT_FILE = r"01-03-26\project_management_tracker.xlsx"
# ============================================================================


class ProjectManagementProcessor:
    """Processes Project Management sheet and generates Activity Timeline Tracker."""
    
    # Column mappings for 'Project Mangement' sheet
    COL_ACTIVITY_CODE = 4      # Col D - Trim ID / Activity Code
    COL_ACTIVITY_NAME = 6      # Col F - WBS / Activity Name
    COL_STAGE_GATE = 11        # Col K - Stage Gate (EP/LP/F/A)
    COL_START = 12             # Col L - Start
    COL_IDCS = 13              # Col M - IDC Submission to EPC Contractor
    COL_IDCC = 14              # Col N - IDC Complete by EPC Contractor
    COL_IFR = 15               # Col O - Issued for Review
    COL_RCC = 16               # Col P - Receive COMPANY Comments
    COL_IFA = 17               # Col Q - Incorporate Comments and Resubmit (IFA/IFH)
    COL_RCA = 18               # Col R - Receive CLIENT Approval
    COL_IFC = 19               # Col S - Issued for Construction / Design / Info
    
    # Style definitions
    HEADER_COLOR = "366092"
    ON_TIME_COLOR = "C6EFCE"
    ON_TIME_FONT_COLOR = "006100"
    DELAYED_COLOR = "FFC7CE"
    DELAYED_FONT_COLOR = "9C0006"
    
    def __init__(self, input_file):
        self.input_file = Path(input_file)
        self.workbook = None
        self.pm_sheet = None
        # Key: (activity_code, group_index) -> Value: {activity_name, stages: {EP/LP/F/A: milestone_dates}}
        self.activities = {}
        self.activity_order = []  # To preserve order
        self.data_records = []
        
    def validate_input_file(self):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if not self.input_file.suffix.lower() in ['.xlsx', '.xlsm']:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}")
        print(f"✓ Input file validated: {self.input_file.name}")
        
    def load_workbook(self):
        try:
            print("Loading workbook...")
            self.workbook = openpyxl.load_workbook(self.input_file, data_only=True)
            
            # Sheet name has typo in source file: "Project Mangement"
            sheet_name = 'Project Mangement'
            if sheet_name not in self.workbook.sheetnames:
                # Try alternate spellings
                for name in self.workbook.sheetnames:
                    if 'project' in name.lower() and 'manage' in name.lower():
                        sheet_name = name
                        break
                else:
                    raise ValueError(f"'Project Mangement' sheet not found in workbook. Available: {self.workbook.sheetnames}")
            
            self.pm_sheet = self.workbook[sheet_name]
            print(f"✓ '{sheet_name}' sheet loaded: {self.pm_sheet.max_row} rows × {self.pm_sheet.max_column} columns")
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {str(e)}")
    
    def extract_data(self):
        print("\nExtracting activity data from Project Management sheet...")
        
        total_rows = self.pm_sheet.max_row
        processed_count = 0
        group_index = 0
        current_group_key = None  # Track current (activity_code, group_index)
        
        # Data starts from row 10 in Project Management sheet
        for row_idx in range(10, total_rows + 1):
            try:
                activity_code = self._get_cell_value(row_idx, self.COL_ACTIVITY_CODE)
                activity_name = self._get_cell_value(row_idx, self.COL_ACTIVITY_NAME)
                stage_gate = self._get_cell_value(row_idx, self.COL_STAGE_GATE)

                if not activity_code and activity_name:
                    activity_code = activity_name

                if not activity_name and activity_code:
                    activity_name = activity_code
                
                if not activity_code:
                    continue
                
                if not stage_gate:
                    continue
                
                activity_code_str = str(activity_code).strip()
                activity_name_str = str(activity_name).strip() if activity_name else ''
                stage_gate_str = str(stage_gate).strip()
                
                # Skip header/label rows (CLASS 1, CLASS 2, etc.)
                if stage_gate_str.startswith('CLASS'):
                    continue
                
                milestone_dates = {
                    'Start': self._get_cell_value(row_idx, self.COL_START),
                    'IDCs': self._get_cell_value(row_idx, self.COL_IDCS),
                    'IDCc': self._get_cell_value(row_idx, self.COL_IDCC),
                    'IFR': self._get_cell_value(row_idx, self.COL_IFR),
                    'RCC': self._get_cell_value(row_idx, self.COL_RCC),
                    'IFA': self._get_cell_value(row_idx, self.COL_IFA),
                    'RCA': self._get_cell_value(row_idx, self.COL_RCA),
                    'IFD/IFC': self._get_cell_value(row_idx, self.COL_IFC),
                }
                
                # EP row = start of a new activity group
                if stage_gate_str == 'EP':
                    group_index += 1
                    current_group_key = (activity_code_str, group_index)
                    self.activities[current_group_key] = {
                        'activity_code': activity_code_str,
                        'activity_name': activity_name_str,
                        'stages': {'EP': milestone_dates}
                    }
                    self.activity_order.append(current_group_key)
                elif current_group_key and current_group_key[0] == activity_code_str:
                    # LP/F/A row belongs to current EP group
                    self.activities[current_group_key]['stages'][stage_gate_str] = milestone_dates
                else:
                    # Stage gate row without a matching EP - create standalone entry
                    group_index += 1
                    standalone_key = (activity_code_str, group_index)
                    if standalone_key not in self.activities:
                        self.activities[standalone_key] = {
                            'activity_code': activity_code_str,
                            'activity_name': activity_name_str,
                            'stages': {}
                        }
                        self.activity_order.append(standalone_key)
                    self.activities[standalone_key]['stages'][stage_gate_str] = milestone_dates
                    current_group_key = standalone_key
                
                processed_count += 1
                
                if processed_count % 500 == 0:
                    print(f"  Processed {processed_count} rows...")
                    
            except Exception as e:
                continue
        
        print(f"✓ Data extraction complete: {processed_count} rows processed")
        print(f"  Unique activity groups: {len(self.activities)}")
        
        self._create_milestone_records()
        
    def _create_milestone_records(self):
        print("\nCreating milestone records...")
        record_count = 0
        
        for group_key in self.activity_order:
            data = self.activities[group_key]
            activity_code = data['activity_code']
            activity_name = data['activity_name']
            stages = data['stages']
            
            # Track all milestones
            milestones_to_track = ['Start', 'IDCs', 'IDCc', 'IFR', 'RCC', 'IFA', 'RCA', 'IFD/IFC']
            
            activity_has_any_record = False
            
            for milestone in milestones_to_track:
                # EP row date = Planned Start Date
                ep_date = stages.get('EP', {}).get(milestone)
                # LP row date = Planned End Date
                lp_date = stages.get('LP', {}).get(milestone)
                # A row date = Actual Date
                a_date = stages.get('A', {}).get(milestone)
                # Also check F (Forecast) stage
                f_date = stages.get('F', {}).get(milestone)
                
                # Create record if at least one date exists (datetime or any non-empty value)
                has_any_date = any(
                    d is not None and d != '' and d != 0
                    for d in [ep_date, lp_date, a_date, f_date]
                )
                
                if has_any_date:
                    duration_deviation, timeline_flag = self._calculate_deviation(
                        ep_date, lp_date, a_date
                    )
                    
                    record = {
                        'activity_id': activity_code,
                        'activity_name': activity_name,
                        'stage_gate': milestone,
                        'planned_start_date': ep_date,
                        'planned_end_date': lp_date,
                        'actual_date': a_date,
                        'duration_deviation': duration_deviation,
                        'timeline_flag': timeline_flag
                    }
                    
                    self.data_records.append(record)
                    record_count += 1
                    activity_has_any_record = True
            
            # If activity has no milestone records at all, still include it with one blank row
            if not activity_has_any_record:
                record = {
                    'activity_id': activity_code,
                    'activity_name': activity_name,
                    'stage_gate': '-',
                    'planned_start_date': None,
                    'planned_end_date': None,
                    'actual_date': None,                    'duration_deviation': None,
                    'timeline_flag': '-'
                }
                self.data_records.append(record)
                record_count += 1
        
        print(f"✓ Created {record_count} milestone records")
    
    def _get_cell_value(self, row, col):
        try:
            return self.pm_sheet.cell(row=row, column=col).value
        except:
            return None
    
    def _calculate_deviation(self, planned_start_ep, planned_end_lp, actual_date_a):
        """Calculate deviation between Planned End (LP) and Actual Date (A)."""
        duration_deviation = None
        
        # Validate: only treat real datetime objects as valid dates
        ep_valid = isinstance(planned_start_ep, datetime)
        lp_valid = isinstance(planned_end_lp, datetime)
        a_valid = isinstance(actual_date_a, datetime)
        
        # If no valid actual date → Not Started
        if not a_valid:
            timeline_flag = "Not Started"
        elif lp_valid:
            duration_deviation = (actual_date_a - planned_end_lp).days
            timeline_flag = "Delayed" if actual_date_a > planned_end_lp else "On Time"
        elif ep_valid:
            duration_deviation = (actual_date_a - planned_start_ep).days
            timeline_flag = "Delayed" if actual_date_a > planned_start_ep else "On Time"
        else:
            # Actual date exists but no planned dates to compare against
            timeline_flag = "On Time"
        
        # Manual Error: LP date is before EP date
        if (ep_valid and lp_valid and planned_end_lp < planned_start_ep):
            timeline_flag = "Manual Error"
        
        return duration_deviation, timeline_flag
    
    def _build_scurve_timeseries(self):
        """Build a proper monthly time-series for the S-Curve Data sheet."""
        from collections import OrderedDict

        ep_entries = []
        lp_entries = []
        actual_entries = []
        for rec in self.data_records:
            aid = rec.get('activity_id', '')
            aname = rec.get('activity_name', '')
            label = f"{aid} | {aname}" if aid else aname
            ep = rec.get('planned_start_date')
            lp = rec.get('planned_end_date')
            act = rec.get('actual_date')
            if isinstance(ep, datetime):
                ep_entries.append((ep, label))
            if isinstance(lp, datetime):
                lp_entries.append((lp, label))
            if isinstance(act, datetime):
                actual_entries.append((act, label))

        if not ep_entries and not lp_entries and not actual_entries:
            return []

        ep_entries.sort(key=lambda x: x[0])
        lp_entries.sort(key=lambda x: x[0])
        actual_entries.sort(key=lambda x: x[0])

        all_dates = ([d for d, _ in ep_entries] +
                     [d for d, _ in lp_entries] +
                     [d for d, _ in actual_entries])
        min_date = min(all_dates)
        max_date = max(all_dates)

        months = OrderedDict()
        cur = datetime(min_date.year, min_date.month, 1)
        end = datetime(max_date.year, max_date.month, 1)
        while cur <= end:
            key = cur.strftime('%Y-%m')
            months[key] = cur
            if cur.month == 12:
                cur = datetime(cur.year + 1, 1, 1)
            else:
                cur = datetime(cur.year, cur.month + 1, 1)

        total_planned = max(len(ep_entries), len(lp_entries), 1)

        def _monthly_cumulative(entries, month_keys):
            cum = 0
            idx = 0
            result = {}
            last_label = ''
            for mk in month_keys:
                while idx < len(entries):
                    d, lbl = entries[idx]
                    if d.strftime('%Y-%m') <= mk:
                        cum += 1
                        last_label = lbl
                        idx += 1
                    else:
                        break
                result[mk] = (cum, last_label)
            return result

        month_keys = list(months.keys())
        ep_monthly = _monthly_cumulative(ep_entries, month_keys)
        lp_monthly = _monthly_cumulative(lp_entries, month_keys)
        actual_monthly = _monthly_cumulative(actual_entries, month_keys)

        timeseries = []
        for mk in month_keys:
            ep_cum, ep_last = ep_monthly.get(mk, (0, ''))
            lp_cum, lp_last = lp_monthly.get(mk, (0, ''))
            act_cum, act_last = actual_monthly.get(mk, (0, ''))

            last_actual_month = actual_entries[-1][0].strftime('%Y-%m') if actual_entries else None
            if last_actual_month and mk > last_actual_month:
                act_pct = None
                act_cum_out = None
                act_last = ''
            else:
                act_pct = round((act_cum / total_planned) * 100, 2) if act_cum > 0 else 0
                act_cum_out = act_cum

            timeseries.append({
                'month_str': months[mk].strftime('%b-%Y'),
                'month_date': months[mk],
                'ep_cum': ep_cum,
                'lp_cum': lp_cum,
                'actual_cum': act_cum_out,
                'ep_pct': round((ep_cum / total_planned) * 100, 2),
                'lp_pct': round((lp_cum / total_planned) * 100, 2),
                'actual_pct': act_pct,
                'last_ep_activity': ep_last,
                'last_lp_activity': lp_last,
                'last_actual_activity': act_last,
            })

        print(f"\u2713 S-Curve time series: {len(timeseries)} months "
              f"({month_keys[0]} to {month_keys[-1]})")
        return timeseries

    def generate_output(self, output_file):
        print(f"\nGenerating output file: {output_file}")
        scurve_data = self._build_scurve_timeseries()
        
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "PM Activity Timeline Tracker"
        
        # Headers in specified order
        headers = [
            'Activity ID',
            'Activity Name',
            'Stage Gate',
            'Early Planning',
            'Late Planning',
            'Actual Date',
            'Duration Deviation (Days)',
            'Timeline Flag'
        ]
        
        # Styles
        header_fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        on_time_fill = PatternFill(start_color=self.ON_TIME_COLOR, end_color=self.ON_TIME_COLOR, fill_type="solid")
        delayed_fill = PatternFill(start_color=self.DELAYED_COLOR, end_color=self.DELAYED_COLOR, fill_type="solid")
        not_started_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        manual_error_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = new_ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Write data
        delayed_count = 0
        on_time_count = 0
        not_started_count = 0
        manual_error_count = 0
        
        for row_idx, record in enumerate(self.data_records, start=2):
            # Activity ID
            new_ws.cell(row=row_idx, column=1, value=record['activity_id']).border = border
            
            # Activity Name
            new_ws.cell(row=row_idx, column=2, value=record['activity_name']).border = border
            
            # Stage Gate
            new_ws.cell(row=row_idx, column=3, value=record['stage_gate']).border = border
            
            # Early Planning
            cell = new_ws.cell(row=row_idx, column=4, value=record['planned_start_date'])
            cell.border = border
            if record['planned_start_date'] and isinstance(record['planned_start_date'], datetime):
                cell.number_format = 'DD-MMM-YY'
            
            # Late Planning
            cell = new_ws.cell(row=row_idx, column=5, value=record['planned_end_date'])
            cell.border = border
            if record['planned_end_date'] and isinstance(record['planned_end_date'], datetime):
                cell.number_format = 'DD-MMM-YY'
            
            # Actual Date (A)
            cell = new_ws.cell(row=row_idx, column=6, value=record['actual_date'])
            cell.border = border
            if record['actual_date'] and isinstance(record['actual_date'], datetime):
                cell.number_format = 'DD-MMM-YY'

            # Duration Deviation
            new_ws.cell(row=row_idx, column=7, value=record['duration_deviation']).border = border

            # Timeline Flag
            cell = new_ws.cell(row=row_idx, column=8, value=record['timeline_flag'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            if record['timeline_flag'] == 'Delayed':
                cell.fill = delayed_fill
                cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                delayed_count += 1
            elif record['timeline_flag'] == 'On Time':
                cell.fill = on_time_fill
                cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                on_time_count += 1
            elif record['timeline_flag'] == 'Not Started':
                cell.fill = not_started_fill
                cell.font = Font(color="7F6000", bold=True)
                not_started_count += 1
            elif record['timeline_flag'] == 'Manual Error':
                cell.fill = manual_error_fill
                cell.font = Font(color="800000", bold=True)
                manual_error_count += 1
        
        # Column widths
        column_widths = {'A': 20, 'B': 50, 'C': 15, 'D': 22, 'E': 22, 'F': 18, 'G': 24, 'H': 15}
        for col, width in column_widths.items():
            new_ws.column_dimensions[col].width = width
        
        new_ws.freeze_panes = 'A2'

        # ================================================================
        # SHEET 2: S-Curve Data (proper monthly time series)
        # ================================================================
        if scurve_data:
            ws2 = new_wb.create_sheet(title="S-Curve Data")

            s2_headers = [
                'Date', 'EP Cumulative %', 'LP Cumulative %', 'Actual Cumulative %',
                'EP Count', 'LP Count', 'Actual Count',
                'Last EP Activity', 'Last LP Activity', 'Last Actual Activity'
            ]
            for col_idx, hdr in enumerate(s2_headers, start=1):
                cell = ws2.cell(row=1, column=col_idx, value=hdr)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

            for row_off, ts in enumerate(scurve_data):
                r = row_off + 2
                ws2.cell(row=r, column=1, value=ts['month_str']).border = border
                cell = ws2.cell(row=r, column=2, value=ts['ep_pct'])
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                cell = ws2.cell(row=r, column=3, value=ts['lp_pct'])
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                cell = ws2.cell(row=r, column=4, value=ts['actual_pct'])
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                ws2.cell(row=r, column=5, value=ts['ep_cum']).border = border
                ws2.cell(row=r, column=6, value=ts['lp_cum']).border = border
                ws2.cell(row=r, column=7, value=ts['actual_cum']).border = border
                ws2.cell(row=r, column=8, value=ts['last_ep_activity']).border = border
                ws2.cell(row=r, column=9, value=ts['last_lp_activity']).border = border
                ws2.cell(row=r, column=10, value=ts['last_actual_activity']).border = border

            s2_widths = {'A': 12, 'B': 18, 'C': 18, 'D': 22, 'E': 12, 'F': 12, 'G': 14, 'H': 50, 'I': 50, 'J': 50}
            for col, width in s2_widths.items():
                ws2.column_dimensions[col].width = width
            ws2.freeze_panes = 'A2'
            print(f"✓ S-Curve Data sheet added ({len(scurve_data)} months)")
        
        try:
            new_wb.save(output_file)
            print(f"✓ Output file created successfully")
        except Exception as e:
            raise RuntimeError(f"Failed to save: {str(e)}")
        finally:
            new_wb.close()
        
        print(f"\n{'='*60}")
        print("PROJECT MANAGEMENT - SUMMARY STATISTICS")
        print(f"{'='*60}")
        print(f"Total Records:      {len(self.data_records):,}")
        print(f"Delayed:            {delayed_count:,}")
        print(f"On Time:            {on_time_count:,}")
        print(f"Not Started:        {not_started_count:,}")
        print(f"Manual Error:       {manual_error_count:,}")
        print(f"Unique Activities:  {len(self.activity_order):,}")
        print(f"{'='*60}")
    
    def close(self):
        if self.workbook:
            self.workbook.close()
    
    def process(self, output_file):
        try:
            self.validate_input_file()
            self.load_workbook()
            self.extract_data()
            self.generate_output(output_file)
        finally:
            self.close()


def main():
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT_FILE
        print("="*60)
        print("PROJECT MANAGEMENT - ACTIVITY TIMELINE TRACKER GENERATOR")
        print("="*60)
        print()
    else:
        print("="*60)
        print("PROJECT MANAGEMENT - ACTIVITY TIMELINE TRACKER GENERATOR")
        print("="*60)
        print()
        print("DEFAULT INPUT FILE:")
        print(f"  {DEFAULT_INPUT_FILE}")
        print()
        user_input = input("Press ENTER to use default, or type new path: ").strip()
        input_file = user_input.strip('"').strip("'") if user_input else DEFAULT_INPUT_FILE
        
        print()
        print("DEFAULT OUTPUT FILE:")
        print(f"  {DEFAULT_OUTPUT_FILE}")
        print()
        user_output = input("Press ENTER to use default, or type new name: ").strip()
        output_file = user_output.strip('"').strip("'") if user_output else DEFAULT_OUTPUT_FILE
        print()
        print("="*60)
        print()
    
    try:
        print(f"Input:  {input_file}")
        print(f"Output: {output_file}")
        print()
        
        processor = ProjectManagementProcessor(input_file)
        processor.process(output_file)
        
        print()
        print("✓ Processing completed successfully!")
        print(f"✓ Output saved to: {output_file}")
        print()
    except Exception as e:
        print()
        print(f"✗ ERROR: {str(e)}")
        print()
        input("Press ENTER to exit...")
        sys.exit(1)
    
    if len(sys.argv) < 2:
        input("\nPress ENTER to exit...")


if __name__ == "__main__":
    main()
