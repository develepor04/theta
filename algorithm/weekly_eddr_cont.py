#!/usr/bin/env python3
"""
Weekly EDDR Cont. Activity Timeline Tracker Generator
=========================================================
Processes Borouge Excel files - 'Weekly EDDR Cont.' sheet
with discipline-wise milestone counts (EP vs Actual).

Column Order (Output):
Sr | Discipline | Stage Gate | Total Deliverables | Planned (EP) |
Actual (A) | Variance | Status

Stage Gates: Start, IDCs, IDCc, IFR, RCC, IFA, RCA, IFD/IFC

Author: Claude
Date: 2026-02-17
Version: 1.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import sys
from pathlib import Path


# ============================================================================
# DEFAULT FILE PATHS
# ============================================================================
DEFAULT_INPUT_FILE = r"2026.01.30_Borouge EU3 H2 Extraction Project PMS-Rev1 dates (1).xlsx"
DEFAULT_OUTPUT_FILE = r"01-03-26\weekly_eddr_cont_tracker.xlsx"
# ============================================================================


class WeeklyEDDRContProcessor:
    """Processes Weekly EDDR Cont. sheet and generates Activity Timeline Tracker."""

    # Column mappings for 'Weekly EDDR Cont.' sheet
    # Row 2 = Headers, Row 3 = EP/A sub-headers
    # Data rows: 4 to 14 (disciplines), Row 15 = Total, Row 16 = Progress %
    COL_SR = 5              # Col E - Sr number
    COL_DISCIPLINE = 6      # Col F - Discipline name
    COL_DELIVERABLES = 7    # Col G - Total Deliverables

    # Milestone columns: each has EP (plan) and A (actual) side by side
    MILESTONE_COLUMNS = {
        'Start':    {'EP': 8,  'A': 9},    # H, I
        'IDCs':     {'EP': 10, 'A': 11},   # J, K
        'IDCc':     {'EP': 12, 'A': 13},   # L, M
        'IFR':      {'EP': 14, 'A': 15},   # N, O
        'RCC':      {'EP': 16, 'A': 17},   # P, Q
        'IFA':      {'EP': 18, 'A': 19},   # R, S
        'RCA':      {'EP': 20, 'A': 21},   # T, U
        'IFD/IFC':  {'EP': 22, 'A': 23},   # V, W
    }

    DATA_START_ROW = 4
    DATA_END_ROW = 14      # Disciplines end at row 14
    TOTAL_ROW = 15
    PROGRESS_ROW = 16

    # Date range cells
    DATE_FROM_ROW = 6
    DATE_TO_ROW = 6
    DATE_FROM_COL = 1       # Col A
    DATE_TO_COL = 2         # Col B

    # Style definitions (same as EDDR/Project Management)
    HEADER_COLOR = "366092"
    ON_TIME_COLOR = "C6EFCE"
    ON_TIME_FONT_COLOR = "006100"
    DELAYED_COLOR = "FFC7CE"
    DELAYED_FONT_COLOR = "9C0006"

    def __init__(self, input_file):
        self.input_file = Path(input_file)
        self.workbook = None
        self.sheet = None
        self.data_records = []
        self.date_from = None
        self.date_to = None

    def validate_input_file(self):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if not self.input_file.suffix.lower() in ['.xlsx', '.xlsm']:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}")
        print(f"✓ Input file validated: {self.input_file.name}")

    def load_workbook(self):
        try:
            print("Loading workbook...")
            self.workbook = openpyxl.load_workbook(self.input_file, data_only=True)

            sheet_name = 'Weekly EDDR Cont.'
            if sheet_name not in self.workbook.sheetnames:
                # Try alternate matches
                for name in self.workbook.sheetnames:
                    if 'weekly' in name.lower() and 'eddr' in name.lower() and 'cont' in name.lower():
                        sheet_name = name
                        break
                else:
                    raise ValueError(
                        f"'Weekly EDDR Cont.' sheet not found in workbook. "
                        f"Available: {self.workbook.sheetnames}"
                    )

            self.sheet = self.workbook[sheet_name]
            print(f"✓ '{sheet_name}' sheet loaded: {self.sheet.max_row} rows × {self.sheet.max_column} columns")
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {str(e)}")

    def _get_cell_value(self, row, col):
        try:
            return self.sheet.cell(row=row, column=col).value
        except:
            return None

    def extract_data(self):
        print("\nExtracting discipline data from Weekly EDDR Cont. sheet...")

        # Get date range
        self.date_from = self._get_cell_value(self.DATE_FROM_ROW, self.DATE_FROM_COL)
        self.date_to = self._get_cell_value(self.DATE_TO_ROW, self.DATE_TO_COL)

        if self.date_from:
            print(f"  Date From: {self.date_from}")
        if self.date_to:
            print(f"  Date To:   {self.date_to}")

        processed_count = 0

        # Process discipline rows (4 to 14) + Total row (15)
        for row_idx in range(self.DATA_START_ROW, self.TOTAL_ROW + 1):
            sr = self._get_cell_value(row_idx, self.COL_SR)
            discipline = self._get_cell_value(row_idx, self.COL_DISCIPLINE)
            total_deliverables = self._get_cell_value(row_idx, self.COL_DELIVERABLES)

            if not discipline:
                continue

            discipline_str = str(discipline).strip()
            sr_str = str(sr).strip() if sr else ''
            total_del = total_deliverables if total_deliverables else 0

            # For each milestone stage, create a record
            for milestone_name, cols in self.MILESTONE_COLUMNS.items():
                ep_count = self._get_cell_value(row_idx, cols['EP'])
                a_count = self._get_cell_value(row_idx, cols['A'])

                # Normalize: convert timedelta → days, non-numeric → 0
                def _to_num(v):
                    if v is None:
                        return 0
                    if isinstance(v, timedelta):
                        return v.days
                    try:
                        return float(v)
                    except (TypeError, ValueError):
                        return 0

                ep_val = _to_num(ep_count)
                a_val = _to_num(a_count)

                # Calculate variance and status
                variance = a_val - ep_val
                if ep_val == 0 and a_val == 0:
                    status = "Not Started"
                elif a_val >= ep_val:
                    status = "On Time"
                else:
                    status = "Delayed"

                record = {
                    'sr': sr_str,
                    'discipline': discipline_str,
                    'stage_gate': milestone_name,
                    'total_deliverables': total_del,
                    'planned_ep': ep_val,
                    'actual_a': a_val,
                    'variance': variance,
                    'status': status
                }

                self.data_records.append(record)
                processed_count += 1

        print(f"✓ Data extraction complete: {processed_count} records created")

    def generate_output(self, output_file):
        print(f"\nGenerating output file: {output_file}")

        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "Weekly EDDR Cont. Tracker"

        # Headers
        headers = [
            'Sr',
            'Discipline',
            'Stage Gate',
            'Total Deliverables',
            'Planned (EP)',
            'Actual (A)',
            'Variance',
            'Status'
        ]

        # Styles (same as EDDR / Project Management)
        header_fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        on_time_fill = PatternFill(start_color=self.ON_TIME_COLOR, end_color=self.ON_TIME_COLOR, fill_type="solid")
        delayed_fill = PatternFill(start_color=self.DELAYED_COLOR, end_color=self.DELAYED_COLOR, fill_type="solid")
        not_started_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Write date range info
        if self.date_from or self.date_to:
            date_from_str = self.date_from.strftime('%d-%b-%Y') if isinstance(self.date_from, datetime) else str(self.date_from or '')
            date_to_str = self.date_to.strftime('%d-%b-%Y') if isinstance(self.date_to, datetime) else str(self.date_to or '')
            info_cell = new_ws.cell(row=1, column=1, value=f"Period: {date_from_str} to {date_to_str}")
            info_cell.font = Font(bold=True, size=12, color=self.HEADER_COLOR)
            new_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            header_start_row = 3
        else:
            header_start_row = 1

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = new_ws.cell(row=header_start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Write data
        delayed_count = 0
        on_time_count = 0
        not_started_count = 0

        for row_offset, record in enumerate(self.data_records):
            row_idx = header_start_row + 1 + row_offset
            is_total_row = (record['discipline'] == 'Total')

            # Sr
            cell = new_ws.cell(row=row_idx, column=1, value=record['sr'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_total_row:
                cell.fill = total_fill
                cell.font = Font(bold=True)

            # Discipline
            cell = new_ws.cell(row=row_idx, column=2, value=record['discipline'])
            cell.border = border
            if is_total_row:
                cell.fill = total_fill
                cell.font = Font(bold=True)

            # Stage Gate
            cell = new_ws.cell(row=row_idx, column=3, value=record['stage_gate'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_total_row:
                cell.fill = total_fill
                cell.font = Font(bold=True)

            # Total Deliverables
            cell = new_ws.cell(row=row_idx, column=4, value=record['total_deliverables'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_total_row:
                cell.fill = total_fill
                cell.font = Font(bold=True)

            # Planned (EP)
            cell = new_ws.cell(row=row_idx, column=5, value=record['planned_ep'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_total_row:
                cell.fill = total_fill
                cell.font = Font(bold=True)

            # Actual (A)
            cell = new_ws.cell(row=row_idx, column=6, value=record['actual_a'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_total_row:
                cell.fill = total_fill
                cell.font = Font(bold=True)

            # Variance
            cell = new_ws.cell(row=row_idx, column=7, value=record['variance'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_total_row:
                cell.fill = total_fill
                cell.font = Font(bold=True)
            elif record['variance'] < 0:
                cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
            elif record['variance'] > 0:
                cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)

            # Status
            cell = new_ws.cell(row=row_idx, column=8, value=record['status'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            if is_total_row:
                cell.fill = total_fill
                cell.font = Font(bold=True)
            elif record['status'] == 'Delayed':
                cell.fill = delayed_fill
                cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                delayed_count += 1
            elif record['status'] == 'On Time':
                cell.fill = on_time_fill
                cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                on_time_count += 1
            elif record['status'] == 'Not Started':
                cell.fill = not_started_fill
                cell.font = Font(color="7F6000", bold=True)
                not_started_count += 1


        # Column widths
        column_widths = {
            'A': 8, 'B': 35, 'C': 15, 'D': 22,
            'E': 16, 'F': 14, 'G': 14, 'H': 15
        }
        for col, width in column_widths.items():
            new_ws.column_dimensions[col].width = width

        new_ws.freeze_panes = f'A{header_start_row + 1}'

        try:
            new_wb.save(output_file)
            print(f"✓ Output file created successfully")
        except Exception as e:
            raise RuntimeError(f"Failed to save: {str(e)}")
        finally:
            new_wb.close()

        print(f"\n{'='*60}")
        print("WEEKLY EDDR CONT. - SUMMARY STATISTICS")
        print(f"{'='*60}")
        print(f"Total Records:      {len(self.data_records):,}")
        print(f"Delayed:            {delayed_count:,}")
        print(f"On Time:            {on_time_count:,}")
        print(f"Not Started:        {not_started_count:,}")
        print(f"{'='*60}")

    def close(self):
        if self.workbook:
            self.workbook.close()

    def process(self, output_file):
        try:
            self.validate_input_file()
            self.load_workbook()
            self.extract_data()
            self.generate_output(output_file)
        finally:
            self.close()


def main():
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT_FILE
        print("=" * 60)
        print("WEEKLY EDDR CONT. - ACTIVITY TIMELINE TRACKER GENERATOR")
        print("=" * 60)
        print()
    else:
        print("=" * 60)
        print("WEEKLY EDDR CONT. - ACTIVITY TIMELINE TRACKER GENERATOR")
        print("=" * 60)
        print()
        print("DEFAULT INPUT FILE:")
        print(f"  {DEFAULT_INPUT_FILE}")
        print()
        user_input = input("Press ENTER to use default, or type new path: ").strip()
        input_file = user_input.strip('"').strip("'") if user_input else DEFAULT_INPUT_FILE

        print()
        print("DEFAULT OUTPUT FILE:")
        print(f"  {DEFAULT_OUTPUT_FILE}")
        print()
        user_output = input("Press ENTER to use default, or type new name: ").strip()
        output_file = user_output.strip('"').strip("'") if user_output else DEFAULT_OUTPUT_FILE
        print()
        print("=" * 60)
        print()

    try:
        print(f"Input:  {input_file}")
        print(f"Output: {output_file}")
        print()

        processor = WeeklyEDDRContProcessor(input_file)
        processor.process(output_file)

        print()
        print("✓ Processing completed successfully!")
        print(f"✓ Output saved to: {output_file}")
        print()
    except Exception as e:
        print()
        print(f"✗ ERROR: {str(e)}")
        print()
        input("Press ENTER to exit...")
        sys.exit(1)

    if len(sys.argv) < 2:
        input("\nPress ENTER to exit...")


if __name__ == "__main__":
    main()
