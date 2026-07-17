#!/usr/bin/env python3
"""
Commissioning, RFSU, Start-up & Activity Timeline Tracker Generator
=========================================================
Processes Borouge Excel files - 'Commissioning, RFSU, Start-up &' sheet
with stage gate milestones and EP/LP/F/A date tracking.

Column Order (Output):
Activity ID | Activity Name | Stage Gate | Early Planning |
Late Planning | Actual Date | Duration Deviation | Timeline Flag

Milestones: Start, Finish

Author: Claude
Date: 2026-02-17
Version: 1.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys
from pathlib import Path


# ============================================================================
# DEFAULT FILE PATHS
# ============================================================================
DEFAULT_INPUT_FILE = r"2026.01.30_Borouge EU3 H2 Extraction Project PMS-Rev1 dates (1).xlsx"
DEFAULT_OUTPUT_FILE = r"01-03-26\commissioning_rfsu_tracker.xlsx"
# ============================================================================


class CommissioningRFSUProcessor:
    """Processes Commissioning, RFSU, Start-up & sheet and generates Activity Timeline Tracker."""

    # Default column mappings (will be auto-detected)
    COL_LEVEL = 1             # Col A - Level
    COL_ACTIVITY_CODE = 2     # Col B - WBS Code / Activity ID  
    COL_ACTIVITY_NAME = 3     # Col C - WBS / Activity Name
    COL_STAGE_GATE = 7        # Col G - Stage Gate (EP/LP/F/A)

    # Milestone date columns
    COL_START = 8             # Col H - Start
    COL_FINISH = 9            # Col I - Finish

    MILESTONES = [
        ('Start', 8),
        ('Finish', 9),
    ]

    # Data start (will be auto-detected)
    DATA_START_ROW = 6

    # Style definitions
    HEADER_COLOR = "366092"
    ON_TIME_COLOR = "C6EFCE"
    ON_TIME_FONT_COLOR = "006100"
    DELAYED_COLOR = "FFC7CE"
    DELAYED_FONT_COLOR = "9C0006"
    NOT_STARTED_COLOR = "FFF2CC"
    NOT_STARTED_FONT_COLOR = "7F6000"

    def __init__(self, input_file):
        self.input_file = Path(input_file)
        self.workbook = None
        self.sheet = None
        self.activities = {}
        self.activity_order = []
        self.data_records = []

    def validate_input_file(self):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if self.input_file.suffix.lower() not in ['.xlsx', '.xlsm']:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}")
        print(f"✓ Input file validated: {self.input_file.name}")

    def load_workbook(self):
        try:
            print("Loading workbook...")
            self.workbook = openpyxl.load_workbook(self.input_file, data_only=True, read_only=True)

            sheet_name = 'Commissioning, RFSU, Start-up &'
            if sheet_name not in self.workbook.sheetnames:
                for name in self.workbook.sheetnames:
                    nl = name.lower()
                    if 'commissioning' in nl and ('rfsu' in nl or 'rfs' in nl or 's-curve' in nl or 'startup' in nl or 'start-up' in nl):
                        sheet_name = name
                        break
                else:
                    # Last resort: any sheet with 'commissioning'
                    for name in self.workbook.sheetnames:
                        if 'commissioning' in name.lower():
                            sheet_name = name
                            break
                    else:
                        raise ValueError(
                            f"'Commissioning' sheet not found. "
                            f"Available: {self.workbook.sheetnames}"
                        )

            self.sheet = self.workbook[sheet_name]
            print(f"✓ '{sheet_name}' sheet loaded: {self.sheet.max_row} rows × {self.sheet.max_column} columns")
            
            # Pre-load all data for fast access in read-only mode
            print("  → Caching sheet data for fast access...")
            self.sheet_data = [list(row) for row in self.sheet.iter_rows(values_only=True)]
            print(f"  → Cached {len(self.sheet_data)} rows")
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {str(e)}")

    def _get_cell_value(self, row, col):
        try:
            # Access cached data (0-indexed)
            if 0 < row <= len(self.sheet_data) and 0 < col <= len(self.sheet_data[0]):
                return self.sheet_data[row - 1][col - 1]
            return None
        except:
            return None

    def extract_data(self):
        print("\nExtracting activity data...")
        
        # Auto-detect header row and data start
        self._detect_structure()

        total_rows = self.sheet.max_row
        processed_count = 0
        group_index = 0
        current_group_key = None

        for row_idx in range(self.DATA_START_ROW, total_rows + 1):
            try:
                level = self._get_cell_value(row_idx, self.COL_LEVEL)
                activity_code = self._get_cell_value(row_idx, self.COL_ACTIVITY_CODE)
                activity_name = self._get_cell_value(row_idx, self.COL_ACTIVITY_NAME)
                stage_gate = self._get_cell_value(row_idx, self.COL_STAGE_GATE)

                if not stage_gate:
                    continue

                sg_str = str(stage_gate).strip()

                # Skip non EP/LP/F/A rows
                if sg_str not in ('EP', 'LP', 'F', 'A'):
                    continue

                # Get milestone dates
                milestone_dates = {}
                for ms_name, ms_col in self.MILESTONES:
                    milestone_dates[ms_name] = self._get_cell_value(row_idx, ms_col)

                if sg_str == 'EP':
                    # EP row = start of new activity group
                    group_index += 1
                    act_name = str(activity_name or '').strip()
                    act_code = str(activity_code or level or '').strip()

                    # Use activity code if available, otherwise use level
                    if not act_name:
                        act_name = act_code

                    current_group_key = (act_code, group_index)
                    self.activities[current_group_key] = {
                        'activity_name': act_name,
                        'activity_code': act_code,
                        'stages': {'EP': milestone_dates}
                    }
                    self.activity_order.append(current_group_key)

                elif current_group_key:
                    self.activities[current_group_key]['stages'][sg_str] = milestone_dates

                    # Update activity code from LP row if it's more specific
                    if sg_str == 'LP':
                        lp_code = str(activity_code or level or '').strip()
                        if lp_code and (lp_code[0].upper() == 'A' or 'PA' in lp_code.upper()):
                            self.activities[current_group_key]['activity_code'] = lp_code

                processed_count += 1

            except Exception as e:
                continue

        print(f"✓ Data extraction complete: {processed_count} rows processed")
        print(f"  Unique activity groups: {len(self.activities)}")

        self._create_milestone_records()
    
    def _detect_structure(self):
        """Auto-detect header row and column positions."""
        print("  → Auto-detecting sheet structure...")
        
        # Look for header row in first 15 rows
        for row_idx in range(1, min(16, self.sheet.max_row + 1)):
            row_vals = [str(self.sheet.cell(row_idx, c).value or '').lower() 
                       for c in range(1, min(12, self.sheet.max_column + 1))]
            
            # Check if this row contains key headers
            has_level = any('level' in v for v in row_vals)
            has_activity = any('activity' in v or 'wbs' in v for v in row_vals)
            has_stage = any('stage' in v or 'gate' in v for v in row_vals)
            
            if has_level and has_activity:
                print(f"  → Found header row at row {row_idx}")
                # Try to find exact column positions
                for idx, val in enumerate(row_vals, 1):
                    if 'level' in val:
                        self.COL_LEVEL = idx
                    elif ('wbs code' in val or 'activity id' in val) and  'activity name' not in val:
                        self.COL_ACTIVITY_CODE = idx
                    elif 'activity name' in val or 'wbs / activity name' in val:
                        self.COL_ACTIVITY_NAME = idx
                    elif 'stage' in val and 'gate' in val:
                        self.COL_STAGE_GATE = idx
                
                self.DATA_START_ROW = row_idx + 1
                print(f"  → Data starts at row {self.DATA_START_ROW}")
                print(f"  → Columns: Level={self.COL_LEVEL}, Code={self.COL_ACTIVITY_CODE}, Name={self.COL_ACTIVITY_NAME}, Gate={self.COL_STAGE_GATE}")
                break

    def _create_milestone_records(self):
        print("\nCreating milestone records...")
        record_count = 0

        for group_key in self.activity_order:
            data = self.activities[group_key]
            activity_name = data['activity_name']
            activity_code = data['activity_code']
            stages = data['stages']

            for ms_name, ms_col in self.MILESTONES:
                ep_date = stages.get('EP', {}).get(ms_name)
                lp_date = stages.get('LP', {}).get(ms_name)
                a_date = stages.get('A', {}).get(ms_name)
                f_date = stages.get('F', {}).get(ms_name)

                has_any_date = any(
                    d is not None and isinstance(d, datetime)
                    for d in [ep_date, lp_date, a_date, f_date]
                )

                if has_any_date:
                    duration_deviation, timeline_flag = self._calculate_deviation(
                        ep_date, lp_date, a_date
                    )

                    record = {
                        'activity_id': activity_code,
                        'activity_name': activity_name,
                        'stage_gate': ms_name,
                        'planned_start_date': ep_date if isinstance(ep_date, datetime) else None,
                        'planned_end_date': lp_date if isinstance(lp_date, datetime) else None,
                        'actual_date': a_date if isinstance(a_date, datetime) else None,
                        'duration_deviation': duration_deviation,
                        'timeline_flag': timeline_flag
                    }

                    self.data_records.append(record)
                    record_count += 1

        print(f"✓ Created {record_count} milestone records")

    def _calculate_deviation(self, planned_start_ep, planned_end_lp, actual_date_a):
        """Calculate deviation between Planned End (LP) and Actual Date (A)."""
        duration_deviation = None
        today = datetime.now()

        ep_valid = isinstance(planned_start_ep, datetime)
        lp_valid = isinstance(planned_end_lp, datetime)
        a_valid = isinstance(actual_date_a, datetime)

        if not a_valid:
            if lp_valid and planned_end_lp < today:
                duration_deviation = (today - planned_end_lp).days
                timeline_flag = "Delayed"
            elif ep_valid and planned_start_ep < today:
                duration_deviation = (today - planned_start_ep).days
                timeline_flag = "Delayed"
            else:
                timeline_flag = "Not Started"
        elif lp_valid:
            duration_deviation = (actual_date_a - planned_end_lp).days
            timeline_flag = "Delayed" if actual_date_a > planned_end_lp else "On Time"
        elif ep_valid:
            duration_deviation = (actual_date_a - planned_start_ep).days
            timeline_flag = "Delayed" if actual_date_a > planned_start_ep else "On Time"
        else:
            timeline_flag = "Not Started"

        # LP earlier than EP is valid (schedule acceleration), not a data error
        # Only flag Manual Error if dates are clearly impossible
        # (e.g., dates before year 2000 or after 2100)
        if (ep_valid and lp_valid):
            min_year = 2000
            max_year = 2100
            if (planned_start_ep.year < min_year or planned_start_ep.year > max_year or
                planned_end_lp.year < min_year or planned_end_lp.year > max_year):
                timeline_flag = "Manual Error"

        return duration_deviation, timeline_flag

    def _format_date(self, date_value):
        if isinstance(date_value, datetime):
            return date_value.strftime('%d-%b-%Y')
        return None

    def generate_output(self, output_file):
        print(f"\nGenerating output file: {output_file}")

        out_wb = openpyxl.Workbook()
        ws = out_wb.active
        ws.title = "Commissioning RFSU Tracker"

        # ── Define Styles ──
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color=self.HEADER_COLOR,
                                  end_color=self.HEADER_COLOR, fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        on_time_fill = PatternFill(start_color=self.ON_TIME_COLOR,
                                   end_color=self.ON_TIME_COLOR, fill_type='solid')
        on_time_font = Font(name='Calibri', color=self.ON_TIME_FONT_COLOR, size=11)

        delayed_fill = PatternFill(start_color=self.DELAYED_COLOR,
                                   end_color=self.DELAYED_COLOR, fill_type='solid')
        delayed_font = Font(name='Calibri', color=self.DELAYED_FONT_COLOR, size=11)

        not_started_fill = PatternFill(start_color=self.NOT_STARTED_COLOR,
                                       end_color=self.NOT_STARTED_COLOR, fill_type='solid')
        not_started_font = Font(name='Calibri', color=self.NOT_STARTED_FONT_COLOR, size=11)

        manual_error_fill = PatternFill(start_color="FF9999",
                                        end_color="FF9999", fill_type='solid')
        manual_error_font = Font(name='Calibri', color="800000", bold=True, size=11)

        data_font = Font(name='Calibri', size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ── Write Headers ──
        headers = [
            'Activity ID',
            'Activity Name',
            'Stage Gate',
            'Early Planning',
            'Late Planning',
            'Actual Date',
            'Duration Deviation',
            'Timeline Flag'
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # ── Write Data ──
        for row_idx, record in enumerate(self.data_records, 2):
            ws.cell(row=row_idx, column=1, value=record['activity_id']).font = data_font
            ws.cell(row=row_idx, column=1).alignment = center_align
            ws.cell(row=row_idx, column=1).border = thin_border

            ws.cell(row=row_idx, column=2, value=record['activity_name']).font = data_font
            ws.cell(row=row_idx, column=2).alignment = left_align
            ws.cell(row=row_idx, column=2).border = thin_border

            ws.cell(row=row_idx, column=3, value=record['stage_gate']).font = data_font
            ws.cell(row=row_idx, column=3).alignment = center_align
            ws.cell(row=row_idx, column=3).border = thin_border

            ws.cell(row=row_idx, column=4,
                    value=self._format_date(record['planned_start_date'])).font = data_font
            ws.cell(row=row_idx, column=4).alignment = center_align
            ws.cell(row=row_idx, column=4).border = thin_border

            ws.cell(row=row_idx, column=5,
                    value=self._format_date(record['planned_end_date'])).font = data_font
            ws.cell(row=row_idx, column=5).alignment = center_align
            ws.cell(row=row_idx, column=5).border = thin_border

            ws.cell(row=row_idx, column=6,
                    value=self._format_date(record['actual_date'])).font = data_font
            ws.cell(row=row_idx, column=6).alignment = center_align
            ws.cell(row=row_idx, column=6).border = thin_border

            ws.cell(row=row_idx, column=7, value=record['duration_deviation']).font = data_font
            ws.cell(row=row_idx, column=7).alignment = center_align
            ws.cell(row=row_idx, column=7).border = thin_border

            flag = record['timeline_flag']
            flag_cell = ws.cell(row=row_idx, column=8, value=flag)
            flag_cell.alignment = center_align
            flag_cell.border = thin_border

            if flag == 'On Time':
                flag_cell.fill = on_time_fill
                flag_cell.font = on_time_font
            elif flag == 'Delayed':
                flag_cell.fill = delayed_fill
                flag_cell.font = delayed_font
            elif flag == 'Not Started':
                flag_cell.fill = not_started_fill
                flag_cell.font = not_started_font
            elif flag == 'Manual Error':
                flag_cell.fill = manual_error_fill
                flag_cell.font = manual_error_font
            else:
                flag_cell.font = data_font

        # ── Column Widths ──
        col_widths = {
            1: 16,   # Activity ID
            2: 45,   # Activity Name
            3: 18,   # Stage Gate
            4: 22,   # Planned Start Date
            5: 22,   # Planned End Date
            6: 18,   # Actual Date
            7: 20,   # Duration Deviation
            8: 16,   # Timeline Flag
        }

        for col, width in col_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:H{len(self.data_records) + 1}"

        try:
            out_wb.save(output_file)
            print(f"✓ Output file created successfully")
        except PermissionError:
            alt_file = output_file.replace('.xlsx', '_new.xlsx')
            out_wb.save(alt_file)
            print(f"⚠ Permission denied for {output_file}")
            print(f"✓ Saved to: {alt_file}")

    def print_summary(self):
        total = len(self.data_records)
        delayed = sum(1 for r in self.data_records if r['timeline_flag'] == 'Delayed')
        on_time = sum(1 for r in self.data_records if r['timeline_flag'] == 'On Time')
        not_started = sum(1 for r in self.data_records if r['timeline_flag'] == 'Not Started')
        manual_error = sum(1 for r in self.data_records if r['timeline_flag'] == 'Manual Error')

        print("\n" + "=" * 60)
        print("COMMISSIONING, RFSU, START-UP - SUMMARY STATISTICS")
        print("=" * 60)
        print(f"Total Records:      {total}")
        print(f"Delayed:            {delayed}")
        print(f"On Time:            {on_time}")
        print(f"Not Started:        {not_started}")
        print(f"Manual Error:       {manual_error}")
        print("=" * 60)

    def process(self, output_file):
        print("=" * 60)
        print("COMMISSIONING, RFSU, START-UP - TRACKER GENERATOR")
        print("=" * 60)
        print(f"\nInput:  {self.input_file}")
        print(f"Output: {output_file}\n")

        self.validate_input_file()
        self.load_workbook()
        self.extract_data()
        self.generate_output(output_file)
        self.print_summary()

        print(f"\n✓ Processing completed successfully!")
        print(f"✓ Output saved to: {output_file}")


def main():
    if len(sys.argv) >= 3:
        input_file = sys.argv[1]
        output_file = sys.argv[2]
    elif len(sys.argv) == 2:
        input_file = sys.argv[1]
        output_file = DEFAULT_OUTPUT_FILE
    else:
        input_file = DEFAULT_INPUT_FILE
        output_file = DEFAULT_OUTPUT_FILE

    processor = CommissioningRFSUProcessor(input_file)
    processor.process(output_file)


if __name__ == "__main__":
    main()
