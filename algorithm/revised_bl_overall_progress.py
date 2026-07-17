#!/usr/bin/env python3
"""
Revised BL Overall Progress Tracker Generator
=========================================================
Processes Borouge Excel files - 'Revised BL Overall Progress' sheet
with discipline-wise progress tracking (Plan vs Actual).

Column Order (Output):
SN | Discipline | Weight Factor (L1%) | Last Period Plan (EP) |
Last Period Actual | Reporting Period Plan | Reporting Period Actual |
Cumulative Plan (EP) | Cumulative Actual | Variance (EP) | Status

Author: Claude
Date: 2026-02-17
Version: 1.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys
from pathlib import Path


# ============================================================================
# DEFAULT FILE PATHS
# ============================================================================
DEFAULT_INPUT_FILE = r"2026.01.30_Borouge EU3 H2 Extraction Project PMS-Rev1 dates (1).xlsx"
DEFAULT_OUTPUT_FILE = r"01-03-26\revised_bl_overall_progress_tracker.xlsx"
# ============================================================================


class RevisedBLProgressProcessor:
    """Processes Revised BL Overall Progress sheet and generates Tracker."""

    # Column mappings for 'Revised BL Overall Progress' sheet
    COL_SN = 2                          # Col B - Serial Number
    COL_DISCIPLINE = 3                  # Col C - Discipline Name
    COL_WEIGHT_FACTOR = 4              # Col D - Weight Factors (L1 %)

    # Category-wise columns: each has Plan (EP) and Actual
    CATEGORY_COLUMNS = {
        'Last Period': {'EP': 5, 'A': 6},       # E, F
        'Reporting Period': {'EP': 7, 'A': 8},   # G, H
        'Cumulative To Date': {'EP': 9, 'A': 10, 'Variance': 11},  # I, J, K
    }

    # Row positions (will be auto-detected)
    HEADER_ROW_1 = 6
    HEADER_ROW_2 = 7
    OVERALL_ROW = 8        # EPC Overall summary row
    DATA_START_ROW = 9
    DATA_END_ROW = 25      # Extended to handle more rows

    # Info row (will be auto-detected)
    INFO_ROW = 5
    COL_CUTOFF_DATE = 3      # Col C
    COL_PREV_DATE = 5        # Col E
    COL_PROJ_COMM_DATE = 8   # Col H
    COL_WEEK_NUM = 11        # Col K

    # Style definitions (same as other trackers)
    HEADER_COLOR = "366092"
    ON_TIME_COLOR = "C6EFCE"
    ON_TIME_FONT_COLOR = "006100"
    DELAYED_COLOR = "FFC7CE"
    DELAYED_FONT_COLOR = "9C0006"

    def __init__(self, input_file):
        self.input_file = Path(input_file)
        self.workbook = None
        self.sheet = None
        self.data_records = []
        self.cutoff_date = None
        self.prev_date = None
        self.proj_comm_date = None
        self.week_num = None

    def validate_input_file(self):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if not self.input_file.suffix.lower() in ['.xlsx', '.xlsm']:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}")
        print(f"✓ Input file validated: {self.input_file.name}")

    def load_workbook(self):
        try:
            print("Loading workbook...")
            self.workbook = openpyxl.load_workbook(self.input_file, data_only=True, read_only=True)

            sheet_name = 'Revised BL Overall Progress'
            if sheet_name not in self.workbook.sheetnames:
                for name in self.workbook.sheetnames:
                    if 'revised' in name.lower() and 'overall' in name.lower() and 'progress' in name.lower():
                        sheet_name = name
                        break
                else:
                    raise ValueError(
                        f"'Revised BL Overall Progress' sheet not found. "
                        f"Available: {self.workbook.sheetnames}"
                    )

            self.sheet = self.workbook[sheet_name]
            print(f"✓ '{sheet_name}' sheet loaded: {self.sheet.max_row} rows × {self.sheet.max_column} columns")
            
            # Pre-load all data for fast access in read-only mode
            print("  → Caching sheet data for fast access...")
            self.sheet_data = [list(row) for row in self.sheet.iter_rows(values_only=True)]
            print(f"  → Cached {len(self.sheet_data)} rows")
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {str(e)}")

    def _get_cell_value(self, row, col):
        try:
            # Access cached data (0-indexed)
            if 0 < row <= len(self.sheet_data) and 0 < col <= len(self.sheet_data[0]):
                return self.sheet_data[row - 1][col - 1]
            return None
        except:
            return None

    def _fmt_pct(self, val):
        """Format value as percentage string."""
        if val is None:
            return None
        try:
            return round(float(val) * 100, 2)
        except (ValueError, TypeError):
            return val

    def extract_data(self):
        print("\nExtracting data from Revised BL Overall Progress sheet...")
        
        # Auto-detect structure
        self._detect_structure()

        # Get info data (search in first 10 rows for "Cut-Off Date")
        for row in range(1, 11):
            for col in range(1, 12):
                val = self._get_cell_value(row, col)
                if val and 'cut-off' in str(val).lower():
                    self.INFO_ROW = row
                    self.cutoff_date = self._get_cell_value(row, col + 1)
                    # Look for other dates in this row
                    for offset in range(1, 10):
                        next_val = self._get_cell_value(row, col + offset)
                        prev_label = self._get_cell_value(row, col + offset - 1)
                        if prev_label and 'previous' in str(prev_label).lower():
                            self.prev_date = next_val
                        elif prev_label and 'comm' in str(prev_label).lower():
                            self.proj_comm_date = next_val
                        elif prev_label and 'week' in str(prev_label).lower():
                            self.week_num = next_val
                    break

        if self.cutoff_date:
            print(f"  Cut-Off Date:     {self.cutoff_date}")
        if self.prev_date:
            print(f"  Previous Date:    {self.prev_date}")
        if self.proj_comm_date:
            print(f"  Project Comm:     {self.proj_comm_date}")
        if self.week_num:
            print(f"  Week:             {self.week_num}")

        processed_count = 0

        # Process data rows dynamically (stop at empty or totals)
        for row_idx in range(self.DATA_START_ROW, min(self.DATA_END_ROW + 1, self.sheet.max_row + 1)):
            sn = self._get_cell_value(row_idx, self.COL_SN)
            discipline = self._get_cell_value(row_idx, self.COL_DISCIPLINE)
            weight_factor = self._get_cell_value(row_idx, self.COL_WEIGHT_FACTOR)

            # Handle overall row (might have discipline in col 2)
            if not discipline and sn:
                discipline = sn
                sn = ''

            if not discipline:
                continue
            
            # Stop at total/summary rows
            discipline_str = str(discipline).strip()
            if discipline_str.upper() in ['TOTAL', 'GRAND TOTAL', 'SUMMARY']:
                break

            sn_str = str(sn).strip() if sn else ''
            
            # Check if this is an overall/summary row
            is_overall = 'OVERALL' in discipline_str.upper() or sn_str == ''

            # For each category, create a separate row
            for cat_name, cols in self.CATEGORY_COLUMNS.items():
                ep_val = self._get_cell_value(row_idx, cols['EP'])
                a_val = self._get_cell_value(row_idx, cols['A'])

                ep_pct = self._fmt_pct(ep_val)
                a_pct = self._fmt_pct(a_val)

                # Variance: use sheet value for Cumulative, else calculate
                if 'Variance' in cols:
                    variance_raw = self._get_cell_value(row_idx, cols['Variance'])
                    variance = self._fmt_pct(variance_raw)
                else:
                    if ep_pct is not None and a_pct is not None:
                        try:
                            variance = round(float(a_pct) - float(ep_pct), 2)
                        except (ValueError, TypeError):
                            variance = None
                    else:
                        variance = None

                # Determine status
                if variance is not None:
                    try:
                        var_val = float(variance)
                        if var_val >= 0:
                            status = "On Time"
                        else:
                            status = "Delayed"
                    except (ValueError, TypeError):
                        status = "-"
                elif ep_pct == 0 and a_pct == 0:
                    status = "Not Started"
                else:
                    status = "-"

                record = {
                    'sn': sn_str,
                    'discipline': discipline_str,
                    'stage_gate': cat_name,
                    'weight_factor': self._fmt_pct(weight_factor),
                    'planned_ep': ep_pct,
                    'actual_a': a_pct,
                    'variance': variance,
                    'status': status,
                    'is_overall': is_overall
                }

                self.data_records.append(record)
                processed_count += 1

        print(f"✓ Data extraction complete: {processed_count} records created")    
    def _detect_structure(self):
        """Auto-detect header rows and data start."""
        print("  → Auto-detecting sheet structure...")
        
        # Look for header row with "Discipline" and "Weight"
        for row_idx in range(1, min(12, self.sheet.max_row + 1)):
            row_vals = [str(self.sheet.cell(row_idx, c).value or '').lower() 
                       for c in range(1, min(15, self.sheet.max_column + 1))]
            
            has_sn = any('sn' in v or v.strip() == 'no' for v in row_vals)
            has_discipline = any('discipline' in v for v in row_vals)
            has_weight = any('weight' in v for v in row_vals)
            has_cumulative = any('cumulative' in v or 'progress' in v for v in row_vals)
            
            if has_discipline and (has_weight or has_cumulative):
                print(f"  → Found header area around row {row_idx}")
                
                # Find column positions
                for idx, val in enumerate(row_vals, 1):
                    if 'sn' in val or val.strip() in ['no', 'no.']:
                        self.COL_SN = idx
                    elif 'discipline' in val:
                        self.COL_DISCIPLINE = idx
                    elif 'weight' in val and 'factor' in val:
                        self.COL_WEIGHT_FACTOR = idx
                
                # Look for data row (first numeric SN after headers)
                for data_row in range(row_idx + 1, min(row_idx + 10, self.sheet.max_row + 1)):
                    first_col_val = self.sheet.cell(data_row, self.COL_SN).value
                    # Check if it's a data row (has EPC overall or number)
                    if first_col_val:
                        val_str = str(first_col_val).strip().upper()
                        if 'EPC' in val_str or val_str.isdigit():
                            self.DATA_START_ROW = data_row
                            print(f"  → Data starts at row {self.DATA_START_ROW}")
                            break
                
                print(f"  → Columns: SN={self.COL_SN}, Discipline={self.COL_DISCIPLINE}, Weight={self.COL_WEIGHT_FACTOR}")
                break
    def generate_output(self, output_file):
        print(f"\nGenerating output file: {output_file}")

        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "Revised BL Progress Tracker"

        # Styles (same as EDDR / Weekly EDDR Cont.)
        header_fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        on_time_fill = PatternFill(start_color=self.ON_TIME_COLOR, end_color=self.ON_TIME_COLOR, fill_type="solid")
        delayed_fill = PatternFill(start_color=self.DELAYED_COLOR, end_color=self.DELAYED_COLOR, fill_type="solid")
        not_started_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        overall_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Row 1: Period info
        cutoff_str = self.cutoff_date.strftime('%d-%b-%Y') if isinstance(self.cutoff_date, datetime) else str(self.cutoff_date or '')
        prev_str = self.prev_date.strftime('%d-%b-%Y') if isinstance(self.prev_date, datetime) else str(self.prev_date or '')
        comm_str = self.proj_comm_date.strftime('%d-%b-%Y') if isinstance(self.proj_comm_date, datetime) else str(self.proj_comm_date or '')
        week_str = str(self.week_num) if self.week_num else ''

        info_text = f"Cut-Off: {cutoff_str}  |  Previous: {prev_str}  |  Project Comm: {comm_str}  |  Week: {week_str}"
        info_cell = new_ws.cell(row=1, column=1, value=info_text)
        info_cell.font = Font(bold=True, size=12, color=self.HEADER_COLOR)
        new_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

        header_start_row = 3

        # Headers - same format as Weekly EDDR Cont. / EDDR
        headers = [
            'SN',
            'Discipline',
            'Stage Gate',
            'Weight Factor (L1 %)',
            'Planned (EP) %',
            'Actual %',
            'Variance %',
            'Status'
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = new_ws.cell(row=header_start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Write data
        delayed_count = 0
        on_time_count = 0
        not_started_count = 0

        for row_offset, record in enumerate(self.data_records):
            row_idx = header_start_row + 1 + row_offset
            is_overall = record['is_overall']

            # SN
            cell = new_ws.cell(row=row_idx, column=1, value=record['sn'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_overall:
                cell.fill = overall_fill
                cell.font = Font(bold=True)

            # Discipline
            cell = new_ws.cell(row=row_idx, column=2, value=record['discipline'])
            cell.border = border
            if is_overall:
                cell.fill = overall_fill
                cell.font = Font(bold=True)

            # Stage Gate (Category)
            cell = new_ws.cell(row=row_idx, column=3, value=record['stage_gate'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_overall:
                cell.fill = overall_fill
                cell.font = Font(bold=True)

            # Weight Factor
            cell = new_ws.cell(row=row_idx, column=4, value=record['weight_factor'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_overall:
                cell.fill = overall_fill
                cell.font = Font(bold=True)

            # Planned (EP)
            cell = new_ws.cell(row=row_idx, column=5, value=record['planned_ep'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_overall:
                cell.fill = overall_fill
                cell.font = Font(bold=True)

            # Actual
            cell = new_ws.cell(row=row_idx, column=6, value=record['actual_a'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_overall:
                cell.fill = overall_fill
                cell.font = Font(bold=True)

            # Variance
            cell = new_ws.cell(row=row_idx, column=7, value=record['variance'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_overall:
                cell.fill = overall_fill
                cell.font = Font(bold=True)
            elif record['variance'] is not None:
                try:
                    if float(record['variance']) < 0:
                        cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                    elif float(record['variance']) > 0:
                        cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                except (ValueError, TypeError):
                    pass

            # Status
            cell = new_ws.cell(row=row_idx, column=8, value=record['status'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            if is_overall:
                cell.fill = overall_fill
                cell.font = Font(bold=True)
            elif record['status'] == 'Delayed':
                cell.fill = delayed_fill
                cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                delayed_count += 1
            elif record['status'] == 'On Time':
                cell.fill = on_time_fill
                cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                on_time_count += 1
            elif record['status'] == 'Not Started':
                cell.fill = not_started_fill
                cell.font = Font(color="7F6000", bold=True)
                not_started_count += 1

        # Column widths
        column_widths = {
            'A': 8, 'B': 42, 'C': 22, 'D': 20,
            'E': 16, 'F': 14, 'G': 14, 'H': 15
        }
        for col, width in column_widths.items():
            new_ws.column_dimensions[col].width = width

        new_ws.freeze_panes = f'A{header_start_row + 1}'

        try:
            new_wb.save(output_file)
            print(f"✓ Output file created successfully")
        except Exception as e:
            raise RuntimeError(f"Failed to save: {str(e)}")
        finally:
            new_wb.close()

        print(f"\n{'='*60}")
        print("REVISED BL OVERALL PROGRESS - SUMMARY STATISTICS")
        print(f"{'='*60}")
        print(f"Total Records:      {len(self.data_records):,}")
        print(f"Delayed:            {delayed_count:,}")
        print(f"On Time:            {on_time_count:,}")
        print(f"Not Started:        {not_started_count:,}")
        print(f"{'='*60}")

    def close(self):
        if self.workbook:
            self.workbook.close()

    def process(self, output_file):
        try:
            self.validate_input_file()
            self.load_workbook()
            self.extract_data()
            self.generate_output(output_file)
        finally:
            self.close()


def main():
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT_FILE
        print("=" * 60)
        print("REVISED BL OVERALL PROGRESS - TRACKER GENERATOR")
        print("=" * 60)
        print()
    else:
        print("=" * 60)
        print("REVISED BL OVERALL PROGRESS - TRACKER GENERATOR")
        print("=" * 60)
        print()
        print("DEFAULT INPUT FILE:")
        print(f"  {DEFAULT_INPUT_FILE}")
        print()
        user_input = input("Press ENTER to use default, or type new path: ").strip()
        input_file = user_input.strip('"').strip("'") if user_input else DEFAULT_INPUT_FILE

        print()
        print("DEFAULT OUTPUT FILE:")
        print(f"  {DEFAULT_OUTPUT_FILE}")
        print()
        user_output = input("Press ENTER to use default, or type new name: ").strip()
        output_file = user_output.strip('"').strip("'") if user_output else DEFAULT_OUTPUT_FILE
        print()
        print("=" * 60)
        print()

    try:
        print(f"Input:  {input_file}")
        print(f"Output: {output_file}")
        print()

        processor = RevisedBLProgressProcessor(input_file)
        processor.process(output_file)

        print()
        print("✓ Processing completed successfully!")
        print(f"✓ Output saved to: {output_file}")
        print()
    except Exception as e:
        print()
        print(f"✗ ERROR: {str(e)}")
        print()
        input("Press ENTER to exit...")
        sys.exit(1)

    if len(sys.argv) < 2:
        input("\nPress ENTER to exit...")


if __name__ == "__main__":
    main()
