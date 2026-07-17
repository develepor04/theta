#!/usr/bin/env python3
"""
HO-As Builts Activity Timeline Tracker Generator
================================================
Dedicated processor for HO-As Builts sheet layout.

This processor reuses the HO-Subcontract extraction logic because the
underlying table structure and stage-gate logic are aligned, but it enforces
HO-As Builts-first sheet targeting.
"""

import openpyxl
from datetime import datetime

from .ho_subcontract import HOSubcontractProcessor


class HOAsBuiltsProcessor(HOSubcontractProcessor):
    """Processes HO-As Builts sheet and generates Activity Timeline Tracker."""

    def __init__(self, input_file):
        super().__init__(input_file)
        self._single_milestone_mode = False
        self._single_milestone_name = 'Issue RFQ to CONTRACTORs'
        self.COL_SINGLE_DATE = None

    def load_workbook(self):
        try:
            print("Loading workbook...")
            self.workbook = openpyxl.load_workbook(self.input_file, data_only=True, read_only=True)

            preferred_sheet = None
            for name in self.workbook.sheetnames:
                low = str(name).lower()
                if 'as built' in low or 'as-built' in low or 'asbuilt' in low:
                    preferred_sheet = name
                    break

            if preferred_sheet is None:
                # Fall back to the parent selection logic so behavior remains robust.
                return super().load_workbook()

            self.sheet = self.workbook[preferred_sheet]
            print(f"✓ '{preferred_sheet}' sheet loaded: {self.sheet.max_row} rows × {self.sheet.max_column} columns")

            print("  → Caching sheet data for fast access...")
            self.sheet_data = [list(row) for row in self.sheet.iter_rows(values_only=True)]
            print(f"  → Cached {len(self.sheet_data)} rows")
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {str(e)}")

    def _normalize_stage_gate(self, value):
        text = str(value or '').strip().upper()
        if not text:
            return None
        if text in ('EP', 'EARLY PLAN', 'EARLY PLANNING'):
            return 'EP'
        if text in ('LP', 'LATE PLAN', 'LATE PLANNING'):
            return 'LP'
        if text in ('A', 'ACT', 'ACTUAL'):
            return 'A'
        if text in ('F', 'FORECAST'):
            return 'F'
        return None

    def _detect_structure(self):
        """Detect HO-As Builts column layout with dynamic stage/date positions."""
        print("  -> Auto-detecting HO-As Builts structure...")

        max_scan_rows = min(30, self.sheet.max_row)
        header_row = None

        for row_idx in range(1, max_scan_rows + 1):
            row_vals = [str(self._get_cell_value(row_idx, c) or '').strip().lower()
                        for c in range(1, min(self.sheet.max_column, 40) + 1)]
            has_level = any('level' in v for v in row_vals)
            has_activity = any(('activity' in v or 'wbs' in v) for v in row_vals)
            if has_level and has_activity:
                header_row = row_idx
                break

        if header_row is None:
            header_row = 7

        row_vals = [str(self._get_cell_value(header_row, c) or '').strip().lower()
                    for c in range(1, min(self.sheet.max_column, 60) + 1)]

        for idx, val in enumerate(row_vals, 1):
            if 'level' in val:
                self.COL_LEVEL = idx
            elif ('wbs code' in val or 'activity id' in val) and 'name' not in val:
                self.COL_ACTIVITY_CODE = idx
            elif 'activity name' in val or 'wbs / activity name' in val:
                self.COL_ACTIVITY_NAME = idx
            elif 'stage' in val and 'gate' in val:
                self.COL_STAGE_GATE = idx

        # Try to find classic four milestone columns first.
        milestone_map = {
            'Issue RFQ to CONTRACTORs': ['issue rfq'],
            'Technical Bid Analysis': ['technical bid'],
            'Bid Analysis & Clarifications': ['bid analysis'],
            'Subcontract Awarded/Signed': ['subcontract awarded', 'subcontract signed'],
        }

        detected_milestones = []
        for ms_name, patterns in milestone_map.items():
            col_found = None
            for idx, val in enumerate(row_vals, 1):
                if any(p in val for p in patterns):
                    col_found = idx
                    break
            if col_found is not None:
                detected_milestones.append((ms_name, col_found))

        if detected_milestones:
            self.MILESTONES = detected_milestones
            self._single_milestone_mode = False
            self.COL_SINGLE_DATE = None
        else:
            # Book1-style layout: one milestone/date column near Stage Gate.
            self._single_milestone_mode = True
            date_col = None
            single_name = 'Issue RFQ to CONTRACTORs'
            for idx, val in enumerate(row_vals, 1):
                if idx <= self.COL_STAGE_GATE:
                    continue
                if ('rfq' in val or 'contractor' in val or 'baseline ep' in val or
                        'revised cum' in val or 'cumm' in val or 'date' in val):
                    date_col = idx
                    if val:
                        single_name = str(self._get_cell_value(header_row, idx) or single_name).strip()
                    break

            if date_col is None:
                date_col = self.COL_STAGE_GATE + 1

            self.COL_SINGLE_DATE = date_col
            self._single_milestone_name = single_name
            self.MILESTONES = [(single_name, date_col)]

        self.DATA_START_ROW = header_row + 1
        print(f"  -> Header row: {header_row}, data start: {self.DATA_START_ROW}")
        print(
            f"  -> Columns: Level={self.COL_LEVEL}, Code={self.COL_ACTIVITY_CODE}, "
            f"Name={self.COL_ACTIVITY_NAME}, Gate={self.COL_STAGE_GATE}"
        )
        if self._single_milestone_mode:
            print(f"  -> Single milestone mode at column {self.COL_SINGLE_DATE}: {self._single_milestone_name}")
        else:
            print(f"  -> Milestones detected: {', '.join(name for name, _ in self.MILESTONES)}")

    def _extract_row_milestone_dates(self, row_idx):
        milestone_dates = {}
        if self._single_milestone_mode:
            milestone_dates[self._single_milestone_name] = self._get_cell_value(row_idx, self.COL_SINGLE_DATE)
            return milestone_dates

        for ms_name, ms_col in self.MILESTONES:
            milestone_dates[ms_name] = self._get_cell_value(row_idx, ms_col)
        return milestone_dates

    def extract_data(self):
        print("\nExtracting activity data...")
        self._detect_structure()

        total_rows = self.sheet.max_row
        processed_count = 0
        group_index = 0
        current_group_key = None
        current_parent_name = None

        for row_idx in range(self.DATA_START_ROW, total_rows + 1):
            try:
                level = self._get_cell_value(row_idx, self.COL_LEVEL)
                activity_code = self._get_cell_value(row_idx, self.COL_ACTIVITY_CODE)
                activity_name = self._get_cell_value(row_idx, self.COL_ACTIVITY_NAME)
                stage_gate_raw = self._get_cell_value(row_idx, self.COL_STAGE_GATE)

                if level and str(level).strip().upper() == 'L3' and activity_name:
                    current_parent_name = str(activity_name or '').strip()

                stage_gate = self._normalize_stage_gate(stage_gate_raw)
                if not stage_gate:
                    continue

                code_str = str(activity_code or '').strip()
                name_str = str(activity_name or '').strip()
                milestone_dates = self._extract_row_milestone_dates(row_idx)

                if stage_gate == 'EP':
                    group_index += 1
                    display_name = current_parent_name if current_parent_name else (name_str or code_str)
                    current_group_key = (display_name, group_index)
                    self.activities[current_group_key] = {
                        'activity_name': display_name,
                        'activity_code': code_str,
                        'stages': {'EP': milestone_dates},
                    }
                    self.activity_order.append(current_group_key)
                elif current_group_key:
                    self.activities[current_group_key]['stages'][stage_gate] = milestone_dates
                else:
                    group_index += 1
                    standalone_name = current_parent_name or name_str or code_str
                    standalone_key = (standalone_name, group_index)
                    self.activities[standalone_key] = {
                        'activity_name': standalone_name,
                        'activity_code': code_str,
                        'stages': {stage_gate: milestone_dates},
                    }
                    self.activity_order.append(standalone_key)
                    current_group_key = standalone_key

                processed_count += 1
            except Exception:
                continue

        print(f"✓ Data extraction complete: {processed_count} stage rows processed")
        print(f"  Unique activity groups: {len(self.activities)}")
        self._create_milestone_records()

    def _create_milestone_records(self):
        print("\nCreating milestone records...")
        record_count = 0

        for group_key in self.activity_order:
            data = self.activities[group_key]
            activity_name = data['activity_name']
            activity_code = data['activity_code']
            stages = data['stages']

            activity_has_any_record = False

            for ms_name, _ in self.MILESTONES:
                ep_date = stages.get('EP', {}).get(ms_name)
                lp_date = stages.get('LP', {}).get(ms_name)
                a_date = stages.get('A', {}).get(ms_name)
                f_date = stages.get('F', {}).get(ms_name)

                actual_value = a_date if a_date not in (None, '', 0) else f_date

                has_any_date = any(
                    d is not None and d != '' and d != 0
                    for d in [ep_date, lp_date, actual_value]
                )

                if has_any_date:
                    duration_deviation, timeline_flag = self._calculate_deviation(
                        ep_date, lp_date, actual_value
                    )

                    record = {
                        'activity_id': activity_code,
                        'activity_name': activity_name,
                        'stage_gate': ms_name,
                        'planned_start_date': ep_date,
                        'planned_end_date': lp_date,
                        'actual_date': actual_value,
                        'duration_deviation': duration_deviation,
                        'timeline_flag': timeline_flag
                    }

                    self.data_records.append(record)
                    record_count += 1
                    activity_has_any_record = True

            if not activity_has_any_record:
                record = {
                    'activity_id': activity_code,
                    'activity_name': activity_name,
                    'stage_gate': '-',
                    'planned_start_date': None,
                    'planned_end_date': None,
                    'actual_date': None,
                    'duration_deviation': None,
                    'timeline_flag': '-'
                }
                self.data_records.append(record)
                record_count += 1

        print(f"✓ Created {record_count} milestone records")
