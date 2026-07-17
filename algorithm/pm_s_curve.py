#!/usr/bin/env python3
"""
PM S-Curve Tracker Generator
=========================================================
Processes Borouge Excel files - 'PM_S-Curve' sheet
with overall PM progress and sub-discipline breakdown.

Sheet 1 (PM Summary Tracker):
SN | Sub-Discipline | Last Month (Plan%|Actual%|Var%) |
This Month (Plan%|Actual%|Var%) | Status

Sheet 2 (PM Monthly S-Curve Data):
Period | Date | Baseline EP% | Revised BL-EP% | Revised BL-LP% | Cumm Actual%

Author: Claude
Date: 2026-02-17
Version: 1.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys
from pathlib import Path


# ============================================================================
# DEFAULT FILE PATHS
# ============================================================================
DEFAULT_INPUT_FILE = r"2026.01.30_Borouge EU3 H2 Extraction Project PMS-Rev1 dates (1).xlsx"
DEFAULT_OUTPUT_FILE = r"01-03-26\pm_s_curve_tracker.xlsx"
# ============================================================================


class PMSCurveProcessor:
    """Processes PM_S-Curve sheet and generates Tracker."""

    # ---- Monthly Time Series (cols 1-34) ----
    ROW_PERIOD_LABELS = 41
    ROW_DATES = 42
    ROW_BASELINE_EP = 43
    ROW_REVISED_BL_EP = 44
    ROW_REVISED_BL_LP = 45
    ROW_CUMM_ACTUAL_MONTHLY = 46
    MONTHLY_DATA_START_COL = 2

    # ---- Weekly Sub-Discipline Chart Data (cols 42+) ----
    WEEKLY_DATES_ROW = 1
    DISCIPLINE_NAME_COL = 42
    WEEKLY_DATA_START_COL = 44

    # Sub-discipline blocks: (name_row, ep_row, lp_row, actual_row)
    # Discipline name at (name_row, col 42), series label at (*, col 43), data cols 44+
    SUB_DISCIPLINE_BLOCKS = [
        {"name_row": 2,  "ep_row": 2,  "lp_row": 3,  "actual_row": 4},    # Project Management
        {"name_row": 10, "ep_row": 10, "lp_row": 11, "actual_row": 12},   # Detailed Design and Engineering
        {"name_row": 18, "ep_row": 18, "lp_row": 19, "actual_row": 20},   # Procurement
        {"name_row": 26, "ep_row": 26, "lp_row": 27, "actual_row": 28},   # Construction
        {"name_row": 34, "ep_row": 34, "lp_row": 35, "actual_row": 36},   # Commissioning
        {"name_row": 51, "ep_row": 51, "lp_row": 52, "actual_row": 53},   # QA/QC Management
        {"name_row": 59, "ep_row": 59, "lp_row": 60, "actual_row": 61},   # Contractor's Temporary Facilities
        {"name_row": 67, "ep_row": 67, "lp_row": 68, "actual_row": 69},   # Others
    ]

    # Style definitions
    HEADER_COLOR = "366092"
    ON_TIME_COLOR = "C6EFCE"
    ON_TIME_FONT_COLOR = "006100"
    DELAYED_COLOR = "FFC7CE"
    DELAYED_FONT_COLOR = "9C0006"
    NOT_STARTED_COLOR = "FFF2CC"
    NOT_STARTED_FONT_COLOR = "7F6000"
    OVERALL_FILL_COLOR = "D9E2F3"   # Light blue for overall PM row

    def __init__(self, input_file):
        self.input_file = Path(input_file)
        self.workbook = None
        self.sheet = None
        self.summary_records = []
        self.monthly_data = []
        self.cutoff_date = None
        self.last_month_date = None

    def validate_input_file(self):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if not self.input_file.suffix.lower() in ['.xlsx', '.xlsm']:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}")
        print(f"✓ Input file validated: {self.input_file.name}")

    def load_workbook(self):
        try:
            print("Loading workbook...")
            self.workbook = openpyxl.load_workbook(self.input_file, data_only=True, read_only=True)

            sheet_name = 'PM_S-Curve'
            if sheet_name not in self.workbook.sheetnames:
                for name in self.workbook.sheetnames:
                    if 'pm' in name.lower() and 'curve' in name.lower():
                        sheet_name = name
                        break
                else:
                    raise ValueError(
                        f"'PM_S-Curve' sheet not found. "
                        f"Available: {self.workbook.sheetnames}"
                    )

            self.sheet = self.workbook[sheet_name]
            print(f"✓ '{sheet_name}' sheet loaded: {self.sheet.max_row} rows × {self.sheet.max_column} columns")
            
            # Pre-load all data for fast access in read-only mode
            print("  → Caching sheet data for fast access...")
            self.sheet_data = [list(row) for row in self.sheet.iter_rows(values_only=True)]
            print(f"  → Cached {len(self.sheet_data)} rows")
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {str(e)}")

    def _get_cell_value(self, row, col):
        try:
            # Access cached data (0-indexed)
            if 0 < row <= len(self.sheet_data) and 0 < col <= len(self.sheet_data[0]):
                return self.sheet_data[row - 1][col - 1]
            return None
        except:
            return None

    def _fmt_pct(self, val):
        """Format value as percentage (multiply by 100 and round to 2 decimals)."""
        if val is None:
            return None
        try:
            return round(float(val) * 100, 2)
        except (ValueError, TypeError):
            return val

    def _find_date_column(self, target_date, date_row, start_col, end_col):
        """Find the column matching or closest to the target date in a given row."""
        if not isinstance(target_date, datetime):
            return None
        best_col = None
        best_diff = None
        for col in range(start_col, end_col + 1):
            date_val = self._get_cell_value(date_row, col)
            if isinstance(date_val, datetime):
                if date_val == target_date:
                    return col
                diff = abs((date_val - target_date).days)
                if best_diff is None or diff < best_diff:
                    best_diff = diff
                    best_col = col
        return best_col

    def _calc_variance_status(self, plan, actual):
        """Calculate variance and determine status."""
        if plan is None and actual is None:
            return None, "Not Started"
        if plan == 0 and (actual is None or actual == 0):
            return None, "Not Started"
        if plan is not None and actual is not None:
            try:
                variance = round(float(actual) - float(plan), 2)
                status = "On Time" if variance >= 0 else "Delayed"
                return variance, status
            except (ValueError, TypeError):
                return None, "-"
        return None, "-"

    def extract_data(self):
        print("\nExtracting data from PM_S-Curve sheet...")
        
        # Check if sheet has enough data
        if self.sheet.max_row < 10 or self.sheet.max_column < 10:
            print("  ⚠ Warning: Sheet appears to be empty or has minimal data")
            print("  → Skipping PM S-Curve processing (insufficient data)")
            return

        # ---- Get dates from Overall S-Curve sheet ----
        try:
            os_sheet = self.workbook['Overall S-Curve'] if 'Overall S-Curve' in self.workbook.sheetnames else None
            if not os_sheet:
                # Try Home Office_S-Curve
                os_sheet = self.workbook['Home Office_S-Curve'] if 'Home Office_S-Curve' in self.workbook.sheetnames else None
            if not os_sheet:
                print("  ⚠ Warning: Overall S-Curve sheet not found, using fallback date detection")
                raise Exception("Fallback to date detection")
            # Cache the os_sheet data for fast access
            os_data = [list(row) for row in os_sheet.iter_rows(values_only=True)]
            self.cutoff_date = os_data[0][8] if len(os_data) > 0 and len(os_data[0]) > 8 else None  # row 1, col 9
            self.last_month_date = os_data[0][10] if len(os_data) > 0 and len(os_data[0]) > 10 else None  # row 1, col 11
            print(f"  Cut-Off Date:     {self.cutoff_date}")
            print(f"  Last Month Date:  {self.last_month_date}")
        except Exception:
            # Fallback: determine from monthly data
            print("  Warning: Could not read dates from Overall S-Curve, detecting from data...")
            for col in range(34, 1, -1):
                val = self._get_cell_value(self.ROW_CUMM_ACTUAL_MONTHLY, col)
                if val is not None:
                    self.cutoff_date = self._get_cell_value(self.ROW_DATES, col)
                    # Try 3 columns back for last month
                    if col >= 5:
                        self.last_month_date = self._get_cell_value(self.ROW_DATES, col - 3)
                    break
            print(f"  Cut-Off Date (detected):    {self.cutoff_date}")
            print(f"  Last Month Date (detected): {self.last_month_date}")

        # ---- Find corresponding columns ----
        # Monthly columns for cut-off and last month dates
        tm_monthly_col = self._find_date_column(self.cutoff_date, self.ROW_DATES, 2, 34)
        lm_monthly_col = self._find_date_column(self.last_month_date, self.ROW_DATES, 2, 34)

        # Weekly columns for cut-off and last month dates
        tm_weekly_col = self._find_date_column(self.cutoff_date, self.WEEKLY_DATES_ROW, 44, 185)
        lm_weekly_col = self._find_date_column(self.last_month_date, self.WEEKLY_DATES_ROW, 44, 185)

        print(f"  Monthly cols: Last Month={lm_monthly_col}, This Month={tm_monthly_col}")
        print(f"  Weekly cols:  Last Month={lm_weekly_col}, This Month={tm_weekly_col}")

        # ---- Extract overall PM from monthly data ----
        print("\n  Extracting overall PM summary...")
        lm_plan_overall = self._fmt_pct(self._get_cell_value(self.ROW_REVISED_BL_EP, lm_monthly_col)) if lm_monthly_col else None
        lm_actual_overall = self._fmt_pct(self._get_cell_value(self.ROW_CUMM_ACTUAL_MONTHLY, lm_monthly_col)) if lm_monthly_col else None
        lm_var_overall, _ = self._calc_variance_status(lm_plan_overall, lm_actual_overall)

        tm_plan_overall = self._fmt_pct(self._get_cell_value(self.ROW_REVISED_BL_EP, tm_monthly_col)) if tm_monthly_col else None
        tm_actual_overall = self._fmt_pct(self._get_cell_value(self.ROW_CUMM_ACTUAL_MONTHLY, tm_monthly_col)) if tm_monthly_col else None
        tm_var_overall, tm_status_overall = self._calc_variance_status(tm_plan_overall, tm_actual_overall)

        self.summary_records.append({
            'sn': 1,
            'discipline': 'PM Overall',
            'is_overall': True,
            'lm_plan': lm_plan_overall,
            'lm_actual': lm_actual_overall,
            'lm_var': lm_var_overall,
            'tm_plan': tm_plan_overall,
            'tm_actual': tm_actual_overall,
            'tm_var': tm_var_overall,
            'status': tm_status_overall
        })

        # ---- Extract sub-disciplines from weekly chart data ----
        print("  Extracting sub-discipline data...")
        sn = 2
        for block in self.SUB_DISCIPLINE_BLOCKS:
            name = self._get_cell_value(block['name_row'], self.DISCIPLINE_NAME_COL)
            name_str = str(name or '').strip()

            if not name_str:
                continue

            # Last Month values
            lm_plan = None
            lm_actual = None
            if lm_weekly_col:
                lm_plan = self._fmt_pct(self._get_cell_value(block['ep_row'], lm_weekly_col))
                lm_actual = self._fmt_pct(self._get_cell_value(block['actual_row'], lm_weekly_col))
            lm_var, _ = self._calc_variance_status(lm_plan, lm_actual)

            # This Month values
            tm_plan = None
            tm_actual = None
            if tm_weekly_col:
                tm_plan = self._fmt_pct(self._get_cell_value(block['ep_row'], tm_weekly_col))
                tm_actual = self._fmt_pct(self._get_cell_value(block['actual_row'], tm_weekly_col))
            tm_var, tm_status = self._calc_variance_status(tm_plan, tm_actual)

            self.summary_records.append({
                'sn': sn,
                'discipline': name_str,
                'is_overall': False,
                'lm_plan': lm_plan,
                'lm_actual': lm_actual,
                'lm_var': lm_var,
                'tm_plan': tm_plan,
                'tm_actual': tm_actual,
                'tm_var': tm_var,
                'status': tm_status
            })
            sn += 1

        print(f"  ✓ Summary: {len(self.summary_records)} records extracted")

        # ---- Extract Monthly S-Curve Time Series Data ----
        print("  Extracting monthly S-Curve data...")
        col = self.MONTHLY_DATA_START_COL
        while True:
            period_label = self._get_cell_value(self.ROW_PERIOD_LABELS, col)
            if period_label is None:
                break

            date_val = self._get_cell_value(self.ROW_DATES, col)
            baseline_ep = self._fmt_pct(self._get_cell_value(self.ROW_BASELINE_EP, col))
            revised_bl_ep = self._fmt_pct(self._get_cell_value(self.ROW_REVISED_BL_EP, col))
            revised_bl_lp = self._fmt_pct(self._get_cell_value(self.ROW_REVISED_BL_LP, col))
            cumm_actual = self._fmt_pct(self._get_cell_value(self.ROW_CUMM_ACTUAL_MONTHLY, col))

            self.monthly_data.append({
                'period': str(period_label),
                'date': date_val,
                'baseline_ep': baseline_ep,
                'revised_bl_ep': revised_bl_ep,
                'revised_bl_lp': revised_bl_lp,
                'cumm_actual': cumm_actual
            })
            col += 1

        print(f"  ✓ Monthly S-Curve: {len(self.monthly_data)} periods extracted")
        print(f"\n✓ Data extraction complete")

    def generate_output(self, output_file):
        print(f"\nGenerating output file: {output_file}")

        new_wb = openpyxl.Workbook()

        # ---- Styles ----
        header_fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        on_time_fill = PatternFill(start_color=self.ON_TIME_COLOR, end_color=self.ON_TIME_COLOR, fill_type="solid")
        delayed_fill = PatternFill(start_color=self.DELAYED_COLOR, end_color=self.DELAYED_COLOR, fill_type="solid")
        not_started_fill = PatternFill(start_color=self.NOT_STARTED_COLOR, end_color=self.NOT_STARTED_COLOR, fill_type="solid")
        overall_fill = PatternFill(start_color=self.OVERALL_FILL_COLOR, end_color=self.OVERALL_FILL_COLOR, fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # ================================================================
        # SHEET 1: PM Summary Tracker
        # ================================================================
        ws1 = new_wb.active
        ws1.title = "PM S-Curve Summary Tracker"

        # Row 1: Info
        cutoff_str = self.cutoff_date.strftime('%d-%b-%Y') if isinstance(self.cutoff_date, datetime) else str(self.cutoff_date or '')
        lm_str = self.last_month_date.strftime('%d-%b-%Y') if isinstance(self.last_month_date, datetime) else str(self.last_month_date or '')

        info_text = f"PM S-Curve  |  Cut-Off: {cutoff_str}  |  Last Month: {lm_str}"
        info_cell = ws1.cell(row=1, column=1, value=info_text)
        info_cell.font = Font(bold=True, size=12, color=self.HEADER_COLOR)
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

        header_start_row = 3

        # Row 3: Top-level headers with merged cells
        # Cols 1-2: SN, Sub-Discipline (merged vertically)
        top_merged = {1: 'SN', 2: 'Sub-Discipline', 9: 'Status'}
        for col_idx, title in top_merged.items():
            ws1.merge_cells(start_row=header_start_row, start_column=col_idx,
                            end_row=header_start_row + 1, end_column=col_idx)
            cell = ws1.cell(row=header_start_row, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            ws1.cell(row=header_start_row + 1, column=col_idx).border = border

        # "Last Month" merged header (cols 3-5)
        ws1.merge_cells(start_row=header_start_row, start_column=3, end_row=header_start_row, end_column=5)
        lm_h = ws1.cell(row=header_start_row, column=3, value="Last Month")
        lm_h.font = header_font
        lm_h.fill = header_fill
        lm_h.alignment = Alignment(horizontal='center', vertical='center')
        lm_h.border = border
        for c in range(4, 6):
            ws1.cell(row=header_start_row, column=c).border = border

        # "This Month" merged header (cols 6-8)
        ws1.merge_cells(start_row=header_start_row, start_column=6, end_row=header_start_row, end_column=8)
        tm_h = ws1.cell(row=header_start_row, column=6, value="This Month")
        tm_h.font = header_font
        tm_h.fill = header_fill
        tm_h.alignment = Alignment(horizontal='center', vertical='center')
        tm_h.border = border
        for c in range(7, 9):
            ws1.cell(row=header_start_row, column=c).border = border

        # Row 4: Sub-headers
        sub_headers = {
            3: 'Plan %', 4: 'Actual %', 5: 'Var %',
            6: 'Plan %', 7: 'Actual %', 8: 'Var %'
        }
        for col_idx, title in sub_headers.items():
            cell = ws1.cell(row=header_start_row + 1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        data_start_row = header_start_row + 2  # Row 5

        # Write summary data
        delayed_count = 0
        on_time_count = 0
        not_started_count = 0

        for row_offset, record in enumerate(self.summary_records):
            row_idx = data_start_row + row_offset
            is_overall = record['is_overall']
            highlight_fill = overall_fill if is_overall else None

            # Col 1: SN
            cell = ws1.cell(row=row_idx, column=1, value=record['sn'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if highlight_fill:
                cell.fill = highlight_fill
                cell.font = Font(bold=True)

            # Col 2: Sub-Discipline
            cell = ws1.cell(row=row_idx, column=2, value=record['discipline'])
            cell.border = border
            if highlight_fill:
                cell.fill = highlight_fill
                cell.font = Font(bold=True)

            # Cols 3-5: Last Month (Plan, Actual, Var)
            for col_off, key in enumerate(['lm_plan', 'lm_actual', 'lm_var']):
                val = record[key]
                display = f"{val:.2f}%" if val is not None else ''
                cell = ws1.cell(row=row_idx, column=3 + col_off, value=display)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                if highlight_fill:
                    cell.fill = highlight_fill
                    cell.font = Font(bold=True)
                elif key == 'lm_var' and val is not None and not is_overall:
                    try:
                        if float(val) < 0:
                            cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                        elif float(val) > 0:
                            cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                    except (ValueError, TypeError):
                        pass

            # Cols 6-8: This Month (Plan, Actual, Var)
            for col_off, key in enumerate(['tm_plan', 'tm_actual', 'tm_var']):
                val = record[key]
                display = f"{val:.2f}%" if val is not None else ''
                cell = ws1.cell(row=row_idx, column=6 + col_off, value=display)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                if highlight_fill:
                    cell.fill = highlight_fill
                    cell.font = Font(bold=True)
                elif key == 'tm_var' and val is not None and not is_overall:
                    try:
                        if float(val) < 0:
                            cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                        elif float(val) > 0:
                            cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                    except (ValueError, TypeError):
                        pass

            # Col 9: Status
            cell = ws1.cell(row=row_idx, column=9, value=record['status'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            if highlight_fill:
                cell.fill = highlight_fill
                cell.font = Font(bold=True)
            elif record['status'] == 'Delayed':
                cell.fill = delayed_fill
                cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                delayed_count += 1
            elif record['status'] == 'On Time':
                cell.fill = on_time_fill
                cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                on_time_count += 1
            elif record['status'] == 'Not Started':
                cell.fill = not_started_fill
                cell.font = Font(color=self.NOT_STARTED_FONT_COLOR, bold=True)
                not_started_count += 1

        # Column widths for Sheet 1
        col_widths_s1 = {
            'A': 6, 'B': 42, 'C': 12, 'D': 12, 'E': 12,
            'F': 12, 'G': 12, 'H': 12, 'I': 15
        }
        for col, width in col_widths_s1.items():
            ws1.column_dimensions[col].width = width

        ws1.freeze_panes = f'A{data_start_row}'

        # ================================================================
        # SHEET 2: PM Monthly S-Curve Data
        # ================================================================
        ws2 = new_wb.create_sheet(title="PM Monthly S-Curve Data")

        # Info row
        ws2.cell(row=1, column=1, value=f"PM Monthly S-Curve Data  |  Cut-Off: {cutoff_str}").font = Font(bold=True, size=12, color=self.HEADER_COLOR)
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        s2_header_row = 3

        s2_headers = [
            'Period',
            'Date',
            'Baseline EP %',
            'Revised BL-EP %',
            'Revised BL-LP %',
            'Cumm. Actual %'
        ]

        for col_idx, header in enumerate(s2_headers, start=1):
            cell = ws2.cell(row=s2_header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Write monthly data
        actual_count = 0
        for row_offset, record in enumerate(self.monthly_data):
            row_idx = s2_header_row + 1 + row_offset
            has_actual = record['cumm_actual'] is not None

            # Period
            cell = ws2.cell(row=row_idx, column=1, value=record['period'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Date
            date_val = record['date']
            if isinstance(date_val, datetime):
                display_date = date_val.strftime('%b-%Y')
            else:
                display_date = str(date_val or '')
            cell = ws2.cell(row=row_idx, column=2, value=display_date)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Baseline EP %
            val = record['baseline_ep']
            display = f"{val:.2f}%" if val is not None else ''
            cell = ws2.cell(row=row_idx, column=3, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Revised BL-EP %
            val = record['revised_bl_ep']
            display = f"{val:.2f}%" if val is not None else ''
            cell = ws2.cell(row=row_idx, column=4, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Revised BL-LP %
            val = record['revised_bl_lp']
            display = f"{val:.2f}%" if val is not None else ''
            cell = ws2.cell(row=row_idx, column=5, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Cumm. Actual %
            val = record['cumm_actual']
            display = f"{val:.2f}%" if val is not None else ''
            cell = ws2.cell(row=row_idx, column=6, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if has_actual:
                actual_count += 1
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                cell.font = Font(bold=True)

        # Column widths for Sheet 2
        col_widths_s2 = {
            'A': 10, 'B': 14, 'C': 16, 'D': 16, 'E': 16, 'F': 16
        }
        for col, width in col_widths_s2.items():
            ws2.column_dimensions[col].width = width

        ws2.freeze_panes = f'A{s2_header_row + 1}'

        # Save
        try:
            new_wb.save(output_file)
            print(f"✓ Output file created successfully")
        except Exception as e:
            raise RuntimeError(f"Failed to save: {str(e)}")
        finally:
            new_wb.close()

        print(f"\n{'='*60}")
        print("PM S-CURVE - SUMMARY STATISTICS")
        print(f"{'='*60}")
        print(f"Summary Records:    {len(self.summary_records):,}")
        print(f"  Delayed:          {delayed_count:,}")
        print(f"  On Time:          {on_time_count:,}")
        print(f"  Not Started:      {not_started_count:,}")
        print(f"Monthly Periods:    {len(self.monthly_data):,}")
        print(f"  With Actuals:     {actual_count:,}")
        print(f"{'='*60}")

    def close(self):
        if self.workbook:
            self.workbook.close()

    def process(self, output_file):
        try:
            self.validate_input_file()
            self.load_workbook()
            self.extract_data()
            self.generate_output(output_file)
        finally:
            self.close()


def main():
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT_FILE
        print("=" * 60)
        print("PM S-CURVE - TRACKER GENERATOR")
        print("=" * 60)
        print()
    else:
        print("=" * 60)
        print("PM S-CURVE - TRACKER GENERATOR")
        print("=" * 60)
        print()
        print("DEFAULT INPUT FILE:")
        print(f"  {DEFAULT_INPUT_FILE}")
        print()
        user_input = input("Press ENTER to use default, or type new path: ").strip()
        input_file = user_input.strip('"').strip("'") if user_input else DEFAULT_INPUT_FILE

        print()
        print("DEFAULT OUTPUT FILE:")
        print(f"  {DEFAULT_OUTPUT_FILE}")
        print()
        user_output = input("Press ENTER to use default, or type new name: ").strip()
        output_file = user_output.strip('"').strip("'") if user_output else DEFAULT_OUTPUT_FILE
        print()
        print("=" * 60)
        print()

    try:
        print(f"Input:  {input_file}")
        print(f"Output: {output_file}")
        print()

        processor = PMSCurveProcessor(input_file)
        processor.process(output_file)

        print()
        print("✓ Processing completed successfully!")
        print(f"✓ Output saved to: {output_file}")
        print()
    except Exception as e:
        print()
        print(f"✗ ERROR: {str(e)}")
        print()
        input("Press ENTER to exit...")
        sys.exit(1)

    if len(sys.argv) < 2:
        input("\nPress ENTER to exit...")


if __name__ == "__main__":
    main()
