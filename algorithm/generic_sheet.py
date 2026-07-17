#!/usr/bin/env python3
"""
Generic Activity Sheet Processor
================================
Fallback processor for any sheet containing these columns:
- Activity Code
- Activity Name
- Early Start / Early Date (EP)
- Late Start / Late Date (LP)
- Actuals (A)
"""

import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class GenericSheetProcessor:
    """Processes generic activity sheets with EP/LP/A columns and generates tracker output."""

    HEADER_COLOR = "366092"
    ON_TIME_COLOR = "C6EFCE"
    ON_TIME_FONT_COLOR = "006100"
    DELAYED_COLOR = "FFC7CE"
    DELAYED_FONT_COLOR = "9C0006"
    NOT_STARTED_COLOR = "FFF2CC"

    def __init__(self, input_file):
        self.input_file = Path(input_file)
        self.workbook = None
        self.sheet = None
        self.data_records = []
        self.header_row_idx = None
        self.col_map = {}

    def validate_input_file(self):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if self.input_file.suffix.lower() not in [".xlsx", ".xlsm", ".xls"]:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}")

    def load_workbook(self):
        self.workbook = openpyxl.load_workbook(self.input_file, data_only=True)
        if not self.workbook.sheetnames:
            raise ValueError("Workbook has no sheets")
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        print(f"Loading workbook...\nUsing sheet '{self.sheet.title}': {self.sheet.max_row} rows × {self.sheet.max_column} columns")

    @staticmethod
    def _normalize_header(text):
        s = str(text or "").strip().lower()
        s = re.sub(r"[^a-z0-9]+", " ", s)
        return " ".join(s.split())

    @staticmethod
    def _is_date_like(value):
        if isinstance(value, datetime):
            return value
        if value is None:
            return None
        s = str(value).strip()
        if not s:
            return None

        fmts = [
            "%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d",
            "%m/%d/%Y", "%m/%d/%y", "%d.%m.%Y", "%d.%m.%y"
        ]
        for fmt in fmts:
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass
        return None

    @staticmethod
    def _find_col(headers, patterns):
        # exact first
        for idx, h in enumerate(headers, start=1):
            if h in patterns:
                return idx
        # contains fallback
        for idx, h in enumerate(headers, start=1):
            for p in patterns:
                if p in h:
                    return idx
        return None

    def _detect_columns(self):
        required_patterns = {
            "activity_code": ["activity code", "activity id", "trim id", "code"],
            "activity_name": ["activity name", "activity description", "wbs", "description"],
            "early": ["early start", "early date", "early planning", " ep "],
            "late": ["late start", "late date", "late planning", " lp "],
            "actual": ["actuals", "actual date", "actual"],
        }

        max_scan_rows = min(30, self.sheet.max_row)
        max_scan_cols = min(80, self.sheet.max_column)

        for r in range(1, max_scan_rows + 1):
            row_vals = [self._normalize_header(self.sheet.cell(row=r, column=c).value) for c in range(1, max_scan_cols + 1)]

            # Pad with spaces to improve short-token matching like " ep " / " lp "
            padded = [f" {v} " for v in row_vals]

            c_activity_code = self._find_col(row_vals, required_patterns["activity_code"])
            c_activity_name = self._find_col(row_vals, required_patterns["activity_name"])
            c_early = self._find_col(padded, required_patterns["early"]) or self._find_col(row_vals, required_patterns["early"])
            c_late = self._find_col(padded, required_patterns["late"]) or self._find_col(row_vals, required_patterns["late"])
            c_actual = self._find_col(row_vals, required_patterns["actual"])

            if all([c_activity_code, c_activity_name, c_early, c_late, c_actual]):
                self.header_row_idx = r
                self.col_map = {
                    "activity_code": c_activity_code,
                    "activity_name": c_activity_name,
                    "early": c_early,
                    "late": c_late,
                    "actual": c_actual,
                }
                return

        raise ValueError(
            "Generic fallback could not find required columns: "
            "Activity Code, Activity Name, Early Start/Early Date (EP), "
            "Late Start/Late Date (LP), Actuals (A)."
        )

    def extract_data(self):
        self._detect_columns()
        start_row = self.header_row_idx + 1

        for r in range(start_row, self.sheet.max_row + 1):
            activity_code = self.sheet.cell(row=r, column=self.col_map["activity_code"]).value
            activity_name = self.sheet.cell(row=r, column=self.col_map["activity_name"]).value
            early_raw = self.sheet.cell(row=r, column=self.col_map["early"]).value
            late_raw = self.sheet.cell(row=r, column=self.col_map["late"]).value
            actual_raw = self.sheet.cell(row=r, column=self.col_map["actual"]).value

            if not activity_code and not activity_name:
                continue

            early = self._is_date_like(early_raw)
            late = self._is_date_like(late_raw)
            actual = self._is_date_like(actual_raw)

            if actual is None:
                flag = "Not Started"
                deviation = None
            elif late is not None:
                deviation = (actual - late).days
                flag = "Delayed" if actual > late else "On Time"
            else:
                deviation = None
                flag = "On Time"

            self.data_records.append({
                "activity_code": str(activity_code).strip() if activity_code is not None else "",
                "activity_name": str(activity_name).strip() if activity_name is not None else "",
                "early": early,
                "late": late,
                "actual": actual,
                "deviation": deviation,
                "flag": flag,
            })

        if not self.data_records:
            raise ValueError("No data rows found after header detection for generic processor")

        print(f"✓ Generic processor extracted {len(self.data_records)} record(s)")

    def generate_output(self, output_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Generic Activity Tracker"

        headers = [
            "Activity Code",
            "Activity Name",
            "Early Start / Early Date (EP)",
            "Late Start / Late Date (LP)",
            "Actuals (A)",
            "Duration Deviation (Days)",
            "Timeline Flag",
        ]

        header_fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        for r_idx, rec in enumerate(self.data_records, start=2):
            ws.cell(r_idx, 1, rec["activity_code"]).border = border
            ws.cell(r_idx, 2, rec["activity_name"]).border = border

            c3 = ws.cell(r_idx, 3, rec["early"])
            c3.border = border
            if rec["early"]:
                c3.number_format = 'DD-MMM-YY'

            c4 = ws.cell(r_idx, 4, rec["late"])
            c4.border = border
            if rec["late"]:
                c4.number_format = 'DD-MMM-YY'

            c5 = ws.cell(r_idx, 5, rec["actual"])
            c5.border = border
            if rec["actual"]:
                c5.number_format = 'DD-MMM-YY'

            ws.cell(r_idx, 6, rec["deviation"]).border = border

            flag_cell = ws.cell(r_idx, 7, rec["flag"])
            flag_cell.border = border
            flag_cell.alignment = Alignment(horizontal='center')
            if rec["flag"] == "Delayed":
                flag_cell.fill = PatternFill(start_color=self.DELAYED_COLOR, end_color=self.DELAYED_COLOR, fill_type="solid")
                flag_cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
            elif rec["flag"] == "On Time":
                flag_cell.fill = PatternFill(start_color=self.ON_TIME_COLOR, end_color=self.ON_TIME_COLOR, fill_type="solid")
                flag_cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
            else:
                flag_cell.fill = PatternFill(start_color=self.NOT_STARTED_COLOR, end_color=self.NOT_STARTED_COLOR, fill_type="solid")
                flag_cell.font = Font(color="7F6000", bold=True)

        widths = {'A': 24, 'B': 50, 'C': 26, 'D': 26, 'E': 18, 'F': 24, 'G': 16}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'A2'

        wb.save(output_file)
        wb.close()
        print(f"✓ Generic output saved: {output_file}")

    def close(self):
        if self.workbook:
            self.workbook.close()

    def process(self, output_file):
        try:
            self.validate_input_file()
            self.load_workbook()
            self.extract_data()
            self.generate_output(output_file)
        finally:
            self.close()
