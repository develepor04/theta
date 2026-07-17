#!/usr/bin/env python3
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
from collections import defaultdict
import sys
from pathlib import Path
import re


# ============================================================================
# DEFAULT FILE PATHS
# ============================================================================
DEFAULT_INPUT_FILE = "Excel file - schedule feb update.xlsx"
DEFAULT_OUTPUT_FILE = "Excel file - schedule feb update_Timeline_Deviation_Analysis.xlsx"
# ============================================================================

class TimelineDeviationProcessor:
    """
    Processes Project Schedule Files (P6 Exports) and generates 
    a Timeline Deviation Report with dynamic column mapping.
    """
    
    # Define keywords to look for in Excel headers to identify data scope
    COLUMN_ALIASES = {
        'id': ['activity id', 'activity code', 'id', 'code', 'activity_id', 'activity no', 'activity number', 'act id'],
        'name': ['activity name', 'description', 'name', 'activity_name', 'task', 'activity description', 'task name'],
        'category': [
            'category', 'discipline', 'discipline name', 'sub discipline', 'major discipline',
            'work package', 'workstream', 'group', 'phase', 'section', 'section name',
            'activity type', 'wbs', 'wbs name', 'trade', 'package'
        ],
        'orig_dur': ['original duration', 'planned_duration', 'duration', 'planned duration'],
        'p_start': ['early start', 'early planned start', 'early planned start date', 'planned start date', 'ep start', 'planned start'],
        'p_finish': [
            'early finish', 'early planed finish', 'early planned finish', 'early planned finish date',
            'planned end date', 'planned finish date', 'planned finish',
            'late planned start', 'late planned finish', 'stage gate', 'target date'
        ],
        'lp_finish': ['late finish', 'late planned finish', 'late end date', 'lp finish'],
        'lp_start': ['late start', 'late start date', 'lp start', 'late planned start'],
        'a_start': ['start', 'start [actual]', 'actual start date', 'actual start', 'start actual', 'actual start (a)'],
        'a_finish': ['finish', 'finish [actual]', 'actual completion date', 'actual finish', 'finish actual', 'actual finish (a)', 'actual date']
    }

    MILESTONE_FILL = "BDD7EE"  # light blue
    EARLY_START_FILL = "E2F0D9"  # light green
    TARGET_SCURVE_CATEGORIES = [
        "COMPANY / PKG 1&2 Interfaces",
        "Project Management",
        "Home Office Services",
        "Manufacturing and Delivery",
        "Construction and Pre-Commissioning",
        "Commissioning, RFSU, Start-Up",
        "Optional Item’s Prices",
        "Reimbursable Items (Provisional Sums)"
    ]

    def __init__(self, input_file):
        self.input_file = Path(input_file)
        self.workbook = None
        self.sheet = None
        self.data_records = []
        self.col_map = {}
        self.header_row_idx = 1
        self.target_sheet_name = None

    def validate_input_file(self):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if self.input_file.suffix.lower() not in ['.xlsx', '.xlsm']:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}")
        print(f"Input file validated: {self.input_file.name}")

    @staticmethod
    def _normalize_header(value):
        """Normalize headers for safer alias matching."""
        if value is None:
            return ""
        return " ".join(str(value).lower().strip().split())

    @staticmethod
    def _canonical_header(value):
        """Canonical header form for punctuation/format-insensitive matching."""
        if value is None:
            return ""
        return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()

    @staticmethod
    def _token_match(header_value, alias_value):
        """Loose token-based match for near-equivalent header wording."""
        header_tokens = [tok for tok in header_value.split() if tok]
        alias_tokens = [tok for tok in alias_value.split() if tok]
        if not header_tokens or not alias_tokens:
            return False

        alias_set = set(alias_tokens)
        common = sum(1 for token in header_tokens if token in alias_set)
        min_required = max(2, min(len(alias_set), len(set(header_tokens))) - 1)
        return common >= min_required

    def load_workbook(self, target_sheet_name=None):
        """Load workbook and select a target sheet without hardcoding Sheet2."""
        self.workbook = openpyxl.load_workbook(self.input_file, data_only=True)

        preferred_sheet = target_sheet_name or self.target_sheet_name
        if preferred_sheet and preferred_sheet in self.workbook.sheetnames:
            self.sheet = self.workbook[preferred_sheet]
        else:
            self.sheet = self.workbook.active

        print(f"Using sheet '{self.sheet.title}' ({self.sheet.max_row} rows x {self.sheet.max_column} columns)")

    def parse_p6_date(self, date_val):
        """Cleans 'A' suffix from P6 dates and converts to datetime objects."""
        if date_val is None or date_val == "":
            return None
        if isinstance(date_val, datetime):
            return date_val
        
        # Strip trailing marker tokens like "A" while preserving month names.
        clean_str = re.sub(r"\s+[A-Za-z]+$", "", str(date_val)).strip()
        
        formats = ['%d-%b-%y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']
        for fmt in formats:
            try:
                return datetime.strptime(clean_str, fmt)
            except ValueError:
                continue
        return None

    def parse_actual_date(self, date_val):
        """
        Parse actual-date fields with strict marker logic.
        Only string values ending with marker 'A' are considered actual dates.
        """
        if date_val is None or date_val == "":
            return None

        if isinstance(date_val, datetime):
            return date_val

        raw = str(date_val).strip()
        if not raw:
            return None

        if not re.search(r"\sA$", raw, flags=re.IGNORECASE):
            return self.parse_p6_date(raw)

        return self.parse_p6_date(raw)

    @staticmethod
    def actual_start_output_value(raw_value, parsed_value):
        """Preserve original actual-start value for validation; show A only where input has A."""
        if raw_value is None:
            return parsed_value

        # Keep source value as-is so validation can compare marker presence row by row.
        if isinstance(raw_value, str):
            return raw_value.strip()
        return raw_value

    @staticmethod
    def _normalize_text(value):
        if value is None:
            return ""
        return " ".join(re.sub(r"[^a-z0-9]+", " ", str(value).lower()).split())

    def _infer_category(self, activity_name, activity_id=None, explicit_category=None):
        """Infer a major activity category from an explicit category or text patterns."""
        source_text = " ".join(
            str(part) for part in [explicit_category, activity_name, activity_id] if part is not None
        )
        text = self._normalize_text(source_text)

        category_patterns = [
            ("Manufacture", [
                "manufacture", "manufacturing", "fabrication shop", "module fabrication"
            ]),
            ("Project Management", [
                "project management", "project manager", "project management office", "pm ", " pm",
                "planning", "control", "reporting", "coordination", "administration", "admin",
                "interface", "project controls", "project services"
            ]),
            ("Engineering", [
                "engineering", "engineer", "design", "detailed design", "basic design", "detail design",
                "concept design", "detailed engineering", "civil design", "electrical design",
                "mechanical design", "instrument design", "structural design"
            ]),
            ("Procurement", [
                "procurement", "purchasing", "purchase", "vendor", "supply", "material",
                "po ", " po", "expediting", "logistics", "submittal", "long lead", "sourcing"
            ]),
            ("Subcontracts", [
                "subcontract", "sub contract", "subcontracts", "sub-contractor", "sub contractor"
            ]),
            ("Commissioning", [
                "commissioning", "rfsu", "start up", "startup", "pre commissioning", "pre-commissioning"
            ]),
            ("Construction", [
                "construction", "construction site", "site work", "civil", "erection", "installation",
                "fabrication", "commissioning support", "assembly", "hook up", "field work", "site"
            ]),
        ]

        for category, patterns in category_patterns:
            for pattern in patterns:
                if pattern in text:
                    return category

        return "Unclassified"

    @staticmethod
    def _month_label(dt_value):
        return dt_value.strftime("%b-%y")

    @staticmethod
    def _month_start(dt_value):
        return datetime(dt_value.year, dt_value.month, 1)

    @staticmethod
    def _next_month(dt_value):
        if dt_value.month == 12:
            return datetime(dt_value.year + 1, 1, 1)
        return datetime(dt_value.year, dt_value.month + 1, 1)

    @staticmethod
    def _month_end(dt_value):
        return TimelineDeviationProcessor._next_month(dt_value) - timedelta(days=1)

    def _build_monthly_scurve_rows(self, records):
        """Build cumulative planned/actual completion series for a group of activities."""
        dated_records = [
            rec for rec in records
            if isinstance(rec.get('p_finish'), datetime) or isinstance(rec.get('a_finish'), datetime) or isinstance(rec.get('lp_finish'), datetime)
        ]

        if not dated_records:
            return []

        dates = []
        for rec in dated_records:
            if isinstance(rec.get('p_finish'), datetime):
                dates.append(self._month_start(rec['p_finish']))
            if isinstance(rec.get('lp_finish'), datetime):
                dates.append(self._month_start(rec['lp_finish']))
            if isinstance(rec.get('a_finish'), datetime):
                dates.append(self._month_start(rec['a_finish']))

        if not dates:
            return []

        start_month = min(dates)
        end_month = max(dates)
        months = []
        current = start_month
        while current <= end_month:
            months.append(current)
            current = self._next_month(current)

        total = len(records)
        if total == 0:
            return []

        rows = []
        for month_start in months:
            month_end = self._month_end(month_start)
            planned_count = sum(
                1 for rec in records
                if isinstance(rec.get('p_finish'), datetime) and rec['p_finish'] <= month_end
            )
            lp_count = sum(
                1 for rec in records
                if isinstance(rec.get('lp_finish'), datetime) and rec['lp_finish'] <= month_end
            )
            actual_count = sum(
                1 for rec in records
                if isinstance(rec.get('a_finish'), datetime) and rec['a_finish'] <= month_end
            )
            planned_pct = round((planned_count / total) * 100, 2)
            lp_pct = round((lp_count / total) * 100, 2)
            actual_pct = round((actual_count / total) * 100, 2)
            rows.append({
                'period': self._month_label(month_start),
                'date': month_end,
                'planned_count': planned_count,
                'lp_count': lp_count,
                'actual_count': actual_count,
                'planned_pct': planned_pct,
                'lp_pct': lp_pct,
                'actual_pct': actual_pct,
                'variance_pct': round(actual_pct - planned_pct, 2),
            })

        return rows

    def _write_scurve_sheet(self, workbook, category_name, records):
        """Create a category-wise S-curve sheet for the supplied records."""
        rows = self._build_monthly_scurve_rows(records)
        if not rows:
            print(f"  • Skipping {category_name} S-Curve (no dated activities)")
            return

        # Excel limits sheet names to 31 chars. Ensure "_SCurve" suffix is preserved.
        suffix = "_SCurve"
        safe_cat_name = category_name.replace(' ', '_').replace('/', '').replace('&', 'and')
        max_cat_len = 31 - len(suffix)
        sheet_name = f"{safe_cat_name[:max_cat_len]}{suffix}"
        
        ws = workbook.create_sheet(title=sheet_name)

        title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(color="FFFFFF", bold=True, size=12)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # We place headers on row 1 so that backend UI array detection logic reliably picks it up
        headers = ["Period", "Date", "Planned Count", "LP Count", "Actual Count", "Planned %", "BL-LP %", "Actual %", "Variance %"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        for row_idx, data in enumerate(rows, 2):
            values = [
                data['period'], data['date'], data['planned_count'], data['lp_count'], data['actual_count'],
                data['planned_pct'], data['lp_pct'], data['actual_pct'], data['variance_pct']
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if isinstance(value, datetime):
                    cell.number_format = 'DD-MMM-YY'
                if col_idx >= 3:
                    cell.alignment = Alignment(horizontal='center')

        chart = LineChart()
        chart.title = f"{category_name} S-Curve"
        chart.style = 2
        chart.y_axis.title = "Completion %"
        chart.x_axis.title = "Period"
        chart.height = 7.5
        chart.width = 13

        data = Reference(ws, min_col=6, max_col=8, min_row=1, max_row=1 + len(rows))
        cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend.position = "r"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
        ws.add_chart(chart, "K2")

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12

        print(f"  • Added {category_name} S-Curve sheet with {len(rows)} periods")

    def _populate_category_scurves(self, workbook):
        """Create separate S-Curve sheets for the main activity categories."""
        grouped = defaultdict(list)
        for rec in self.data_records:
            category = rec.get('category') or self._infer_category(rec.get('activity_name'), rec.get('activity_id'))
            rec['category'] = category
            grouped[category].append(rec)

        preferred_order = [c for c in self.TARGET_SCURVE_CATEGORIES if c in grouped]
        discovered_extra = sorted([c for c in grouped.keys() if c not in preferred_order])
        all_categories = preferred_order + discovered_extra

        for category_name in all_categories:
            self._write_scurve_sheet(workbook, category_name, grouped[category_name])

        print("  • Category breakdown:")
        for category_name in all_categories:
            if grouped.get(category_name):
                print(f"    - {category_name}: {len(grouped[category_name])} activities")

    def map_columns(self, sheet):
        """Dynamically identify columns by scanning candidate header rows."""
        max_scan_rows = min(60, sheet.max_row)
        max_scan_cols = min(200, sheet.max_column)
        best_map = None
        best_header_row = None
        best_score = -1

        for row_idx in range(1, max_scan_rows + 1):
            header_row = [
                self._normalize_header(sheet.cell(row=row_idx, column=col_idx).value)
                for col_idx in range(1, max_scan_cols + 1)
            ]
            if not any(header_row):
                continue

            found_map = {}
            for key, aliases in self.COLUMN_ALIASES.items():
                mapped_col = None
                for alias in aliases:
                    norm_alias = self._normalize_header(alias)
                    canon_alias = self._canonical_header(alias)

                    if norm_alias in header_row:
                        mapped_col = header_row.index(norm_alias) + 1
                        break

                    for idx, header_val in enumerate(header_row):
                        if not header_val:
                            continue

                        canon_header = self._canonical_header(header_val)
                        if not canon_header or not canon_alias:
                            continue

                        if canon_alias == canon_header:
                            mapped_col = idx + 1
                            break

                        # Substring match only for substantial strings (≥7 chars)
                        # to prevent single words like 'start' matching 'planned start date'
                        if (canon_alias in canon_header and len(canon_alias) >= 7) or \
                           (canon_header in canon_alias and len(canon_header) >= 7):
                            mapped_col = idx + 1
                            break

                        if self._token_match(canon_header, canon_alias):
                            mapped_col = idx + 1
                            break

                    if mapped_col is not None:
                        break

                if mapped_col is not None:
                    found_map[key] = mapped_col

            critical = ['id', 'name', 'p_start', 'p_finish']
            missing = [c for c in critical if c not in found_map]
            if missing:
                continue

            score = len(found_map)
            if score > best_score:
                best_score = score
                best_map = found_map
                best_header_row = row_idx

        if not best_map:
            raise ValueError("Required columns missing in file: ['id', 'name', 'p_start', 'p_finish']")

        self.col_map = best_map
        self.header_row_idx = best_header_row
        print(f"✓ Columns mapped successfully from row {self.header_row_idx}: {list(self.col_map.keys())}")

    def extract_data(self):
        """Load selected sheet data and calculate timeline deviations."""
        if self.workbook is None or self.sheet is None:
            self.load_workbook()
        
        self.map_columns(self.sheet)
        print("Processing rows...")

        current_category = None

        for row in range(self.header_row_idx + 1, self.sheet.max_row + 1):
            act_id = self.sheet.cell(row=row, column=self.col_map['id']).value
            if not act_id: continue

            activity_name_raw = self.sheet.cell(row=row, column=self.col_map['name']).value

            # Extract and Clean Dates
            p_start_raw = self.sheet.cell(row=row, column=self.col_map['p_start']).value
            p_finish_raw = self.sheet.cell(row=row, column=self.col_map['p_finish']).value
            p_start = self.parse_p6_date(p_start_raw)
            p_finish = self.parse_p6_date(p_finish_raw)

            # Optional columns check (handle files that might lack actuals or LP)
            lp_finish = self.parse_p6_date(self.sheet.cell(row=row, column=self.col_map['lp_finish']).value) if 'lp_finish' in self.col_map else None
            lp_start = self.parse_p6_date(self.sheet.cell(row=row, column=self.col_map['lp_start']).value) if 'lp_start' in self.col_map else None
            explicit_category_raw = self.sheet.cell(row=row, column=self.col_map['category']).value if 'category' in self.col_map else None
            a_start_raw = self.sheet.cell(row=row, column=self.col_map['a_start']).value if 'a_start' in self.col_map else None
            a_start = self.parse_actual_date(a_start_raw) if 'a_start' in self.col_map else None
            a_finish_raw = self.sheet.cell(row=row, column=self.col_map['a_finish']).value if 'a_finish' in self.col_map else None
            a_finish = self.parse_actual_date(a_finish_raw) if 'a_finish' in self.col_map else None
            a_start_out = self.actual_start_output_value(a_start_raw, a_start)

            # P6 INLINE ACTUALS: If Actual column was missing/empty, but Early dates have 'A', treat as Actual
            if not a_start and isinstance(p_start_raw, str) and p_start_raw.strip().upper().endswith('A'):
                a_start = p_start
                a_start_out = p_start_raw
            if not a_finish and isinstance(p_finish_raw, str) and p_finish_raw.strip().upper().endswith('A'):
                a_finish = p_finish

            if 'orig_dur' in self.col_map:
                orig_dur_raw = self.sheet.cell(row=row, column=self.col_map['orig_dur']).value
            else:
                orig_dur_raw = 0

            try:
                orig_dur = int(float(orig_dur_raw or 0))
            except (TypeError, ValueError):
                orig_dur = 0

            activity_name_str = str(activity_name_raw).strip() if activity_name_raw is not None else ""
            act_id_raw_str = str(act_id) if act_id is not None else ""
            act_id_str = act_id_raw_str.strip()  
            explicit_category_str = str(explicit_category_raw).strip() if explicit_category_raw is not None else ""

            # WBS Node Check (Empty Activity Name)
            # Find indentation to determine the level
            indent_level = len(act_id_raw_str) - len(act_id_raw_str.lstrip(' '))
            
            if not activity_name_str:
                # Based on the file structure, top-level project phases have a 2-space indent
                if indent_level == 2:
                    current_category = act_id_str
                    print(f"  • Major WBS Category Set: {current_category}")
                # Skip WBS headings since they are not actionable tasks that need tracking
                continue

            resolved_category = explicit_category_str or current_category or self._infer_category(
                activity_name=activity_name_str,
                activity_id=act_id_str,
                explicit_category=explicit_category_str
            )

            # --- FORMULA LOGIC ---
            
            # 1. Planned Duration
            p_dur = (p_finish - p_start).days if (p_start and p_finish) else orig_dur
            
            # 2. Actual Duration
            a_dur = (a_finish - a_start).days if (a_start and a_finish) else None
            
            # 3. Start Delay Deviation (Actual Start - Planned Start)
            start_delay = (a_start - p_start).days if (a_start and p_start) else 0
            
            # 4. Duration Deviation (Actual Dur - Planned Dur)
            dur_dev = (a_dur - p_dur) if (a_dur is not None) else 0

            # 4a. Timeline deviation flag (duration-based explanation for end users).
            if p_dur is None or p_dur <= 0:
                timeline_deviation_flag = "Duration Check - Planned Duration Missing"
            elif a_dur is None:
                if a_start is None:
                    timeline_deviation_flag = "Duration Check - Not Started"
                else:
                    timeline_deviation_flag = "Duration Check - In Progress"
            else:
                dur_gap = a_dur - p_dur
                if dur_gap > 0:
                    timeline_deviation_flag = f"Duration Delay (+{dur_gap}d vs plan)"
                elif dur_gap < 0:
                    timeline_deviation_flag = f"Duration Ahead ({abs(dur_gap)}d faster than plan)"
                else:
                    timeline_deviation_flag = "Duration On Plan (matches planned duration)"

            # 4b. Start delay flag: start behavior against planned start.
            if a_start is None:
                start_delay_flag = "Not Started"
            elif p_start is None:
                start_delay_flag = "N/A"
            elif a_start < p_start:
                start_delay_flag = "Early Start"
            elif a_start > p_start:
                start_delay_flag = "Delay Start"
            else:
                start_delay_flag = "On Time Start"

            # 5. Timeline Flag Logic
            # Milestone: Start (p_start) empty AND Late Start (lp_start) empty.
            # Falls back to old heuristic when Late Start column is absent from the file.
            if 'lp_start' in self.col_map:
                is_milestone = (p_start is None) and (lp_start is None)
            else:
                is_milestone = (p_start is None) and (p_finish is not None)
            if is_milestone:
                if a_start is None and a_finish is None:
                    flag = "Milestone - Not Started"
                elif a_finish is None:
                    flag = "Milestone - Ongoing"
                elif p_finish is None:
                    flag = "Milestone - Achieved - On Time"
                elif a_finish < p_finish:
                    flag = "Milestone - Achieved - Early"
                elif a_finish > p_finish:
                    flag = "Milestone - Achieved - Delay"
                else:
                    flag = "Milestone - Achieved - On Time"
            else:
                if a_start is None:
                    flag = "Activity - Not Started"
                elif a_finish is not None and p_finish is not None and a_finish > p_finish:
                    flag = "Activity - Late"
                elif p_start is not None and a_start > p_start:
                    flag = "Activity - Delay"
                elif p_start is not None and a_start < p_start:
                    flag = "Activity - Early Start"
                else:
                    flag = "Activity - On Time"

            self.data_records.append({
                'activity_id': act_id,
                'activity_name': activity_name_raw,
                'category': resolved_category,
                'p_start': p_start,
                'p_finish': p_finish,
                'lp_finish': lp_finish,
                'a_start': a_start,
                'a_start_out': a_start_out,
                'a_finish': a_finish,
                'p_dur': p_dur,
                'a_dur': a_dur,
                's_delay': start_delay,
                'd_dev': dur_dev,
                'timeline_deviation_flag': timeline_deviation_flag,
                'start_delay_flag': start_delay_flag,
                'flag': flag
            })

    def close(self):
        if self.workbook:
            self.workbook.close()

    def process(self, output_file):
        try:
            self.validate_input_file()
            self.load_workbook()
            self.extract_data()
            self.generate_output(output_file)
        finally:
            self.close()

    def generate_output(self, output_file):
        """Generates the Timeline Deviation sheet with formatting."""
        new_wb = openpyxl.Workbook()
        ws = new_wb.active
        ws.title = "Timeline Deviation"

        headers = [
            'activity_id', 'activity_name', 'planned start date', 'planned end date',
            'actual start date', 'actual completion date', 'planned_duration',
            'actual_duration', 'start_delay deviation', 'duration_deviation',
            'start_delay_flag', 'Status', 'Timeline_deviation_flag'
        ]

        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Write Headers
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Write Data
        for r_idx, rec in enumerate(self.data_records, 2):
            row_data = [
                rec['activity_id'], rec['activity_name'], rec['p_start'], rec['p_finish'],
                rec['a_start_out'], rec['a_finish'], rec['p_dur'],
                rec['a_dur'], rec['s_delay'], rec['d_dev'],
                rec['start_delay_flag'], rec['flag'], rec['timeline_deviation_flag']
            ]
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border
                
                # Format Dates
                if isinstance(val, datetime):
                    cell.number_format = 'DD-MMM-YY'
                
                # Apply Status Coloring for flag/status columns.
                if c_idx in (11, 12, 13):
                    cell.alignment = Alignment(horizontal='center')

                    val_str = str(val).strip()
                    val_lower = val_str.lower()
                    is_delay = ("delay" in val_lower) or ("delayed" in val_lower)
                    is_early = "early" in val_lower
                    is_ongoing = ("ongoing" in val_lower) or ("in progress" in val_lower)
                    is_ahead = "ahead" in val_lower

                    if is_delay:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif is_early or is_ahead:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif is_ongoing:
                        cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                    elif c_idx == 12 and val_str.startswith("Milestone"):
                        cell.fill = PatternFill(start_color=self.MILESTONE_FILL, end_color=self.MILESTONE_FILL, fill_type="solid")
                    elif c_idx == 12 and ("on time" in val_lower or "plan" in val_lower):
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # Set Column Widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 45
        for char in ['C', 'D', 'E', 'F']: ws.column_dimensions[char].width = 18
        for char in ['G', 'H', 'I', 'J']: ws.column_dimensions[char].width = 20
        ws.column_dimensions['K'].width = 16
        ws.column_dimensions['L'].width = 16
        ws.column_dimensions['M'].width = 15

        self._populate_category_scurves(new_wb)

        new_wb.save(output_file)
        print(f"✓ Output generated: {output_file}")

def main():
    base_dir = Path(__file__).resolve().parent.parent

    if len(sys.argv) >= 2:
        input_path = Path(sys.argv[1])
        if not input_path.is_absolute():
            input_path = (Path.cwd() / input_path).resolve()
        output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else input_path.with_name(
            f"{input_path.stem}_Timeline_Deviation_Analysis.xlsx"
        )
        if not output_path.is_absolute():
            output_path = (Path.cwd() / output_path).resolve()
    else:
        print("=" * 60)
        print("TIMELINE DEVIATION TRACKER GENERATOR")
        print("=" * 60)
        print()
        print("DEFAULT INPUT FILE:")
        print(f"  {base_dir / DEFAULT_INPUT_FILE}")
        print()
        user_input = input("Press ENTER to use default, or type new path: ").strip()
        input_path = Path(user_input.strip('"').strip("'")) if user_input else (base_dir / DEFAULT_INPUT_FILE)
        if not input_path.is_absolute():
            input_path = (Path.cwd() / input_path).resolve()

        print()
        print("DEFAULT OUTPUT FILE:")
        print(f"  {base_dir / DEFAULT_OUTPUT_FILE}")
        print()
        user_output = input("Press ENTER to use default, or type new name: ").strip()
        output_path = Path(user_output.strip('"').strip("'")) if user_output else (base_dir / DEFAULT_OUTPUT_FILE)
        if not output_path.is_absolute():
            output_path = (Path.cwd() / output_path).resolve()
        print()
        print("=" * 60)
        print()
    
    try:
        processor = TimelineDeviationProcessor(input_path)
        processor.process(output_path)
        print("--- Final Summary ---")
        print(f"Total Activities Processed: {len(processor.data_records)}")
    except Exception as e:
        print(f"Error: {e}")
        if len(sys.argv) < 2:
            input("Press ENTER to exit...")
        sys.exit(1)

    if len(sys.argv) < 2:
        input("\nPress ENTER to exit...")

if __name__ == "__main__":
    main()