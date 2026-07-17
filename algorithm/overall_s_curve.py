#!/usr/bin/env python3
"""
Overall S-Curve Tracker Generator
=========================================================
Processes Borouge Excel files - 'Overall S-Curve' sheet
with discipline-wise progress and monthly S-Curve data.

Sheet 1 (Summary Tracker):
SN | Level | WBS Code | Discipline | L1 Weight% |
Last Month Plan% | Last Month Actual% | Last Month Var% |
This Month Plan% | This Month Actual% | This Month Var% | Status

Sheet 2 (Monthly S-Curve Data):
Period | Date | Baseline EP% | Revised BL-EP% | Revised BL-LP% | Cumm Actual%

Author: Claude
Date: 2026-02-17
Version: 1.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys
from pathlib import Path


# ============================================================================
# DEFAULT FILE PATHS
# ============================================================================
DEFAULT_INPUT_FILE = r"2026.01.30_Borouge EU3 H2 Extraction Project PMS-Rev1 dates (1).xlsx"
DEFAULT_OUTPUT_FILE = r"01-03-26\overall_s_curve_tracker.xlsx"
# ============================================================================


class OverallSCurveProcessor:
    """Processes Overall S-Curve sheet and generates Tracker."""

    # ---- Summary Section (Rows 1-8) ----
    # Info Row 1
    COL_CUTOFF_DATE = 9       # Col I
    COL_LAST_MONTH_DATE = 11  # Col K

    # Header Row 2 (column labels)
    # Data Rows 3-8
    SUMMARY_DATA_START = 3
    SUMMARY_DATA_END = 8

    # Column mappings for summary
    COL_LEVEL = 2             # Col B - Level (L0/L1)
    COL_WBS_CODE = 3          # Col C - WBS Code
    COL_WBS_NAME = 5          # Col E - WBS Name
    COL_L1_WEIGHT = 11        # Col K - L1 w.r.t Whole Project (weight)

    # Last Month columns
    COL_LM_PLAN = 12          # Col L - BL Cumm. Early Plan (last month)
    COL_LM_ACTUAL = 13        # Col M - Cumm. Actual (last month)
    COL_LM_VAR = 14           # Col N - Variance (last month)

    # This Month columns
    COL_TM_PLAN = 15          # Col O - BL Cumm. Early Plan (this month)
    COL_TM_ACTUAL = 16        # Col P - Cumm. Actual (this month)
    COL_TM_VAR = 17           # Col Q - Variance (this month)

    # ---- Monthly Time Series Section (Rows 35-44) ----
    ROW_PERIOD_LABELS = 35    # M1, M2, ... M34
    ROW_DATES = 36            # Nov 2024, Dec 2024, ...
    ROW_BASELINE_EP = 37      # Baseline EP cumulative
    ROW_REVISED_BL_EP = 38    # Revised BL-EP cumulative
    ROW_REVISED_BL_LP = 39    # Revised BL-LP cumulative
    ROW_CUMM_ACTUAL = 40      # Cumm. Actual
    MONTHLY_DATA_START_COL = 2  # Col B

    # Style definitions
    HEADER_COLOR = "366092"
    ON_TIME_COLOR = "C6EFCE"
    ON_TIME_FONT_COLOR = "006100"
    DELAYED_COLOR = "FFC7CE"
    DELAYED_FONT_COLOR = "9C0006"
    NOT_STARTED_COLOR = "FFF2CC"
    NOT_STARTED_FONT_COLOR = "7F6000"
    L0_FILL_COLOR = "D9E2F3"   # Light blue for L0 row
    L1_FILL_COLOR = "E2EFDA"   # Light green for L1 header rows

    def __init__(self, input_file):
        self.input_file = Path(input_file)
        self.workbook = None
        self.sheet = None
        self.summary_records = []
        self.monthly_data = []
        self.cutoff_date = None
        self.last_month_date = None

    def validate_input_file(self):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if not self.input_file.suffix.lower() in ['.xlsx', '.xlsm']:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}")
        print(f"✓ Input file validated: {self.input_file.name}")

    def load_workbook(self):
        try:
            print("Loading workbook...")
            self.workbook = openpyxl.load_workbook(self.input_file, data_only=True)

            sheet_name = 'Overall S-Curve'
            if sheet_name not in self.workbook.sheetnames:
                for name in self.workbook.sheetnames:
                    if 's-curve' in name.lower() or 's curve' in name.lower():
                        sheet_name = name
                        break
                else:
                    raise ValueError(
                        f"'Overall S-Curve' sheet not found. "
                        f"Available: {self.workbook.sheetnames}"
                    )

            self.sheet = self.workbook[sheet_name]
            print(f"✓ '{sheet_name}' sheet loaded: {self.sheet.max_row} rows × {self.sheet.max_column} columns")
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {str(e)}")

    def _get_cell_value(self, row, col):
        try:
            return self.sheet.cell(row=row, column=col).value
        except:
            return None

    def _fmt_pct(self, val):
        """Format value as percentage (multiply by 100 and round to 2 decimals)."""
        if val is None:
            return None
        try:
            return round(float(val) * 100, 2)
        except (ValueError, TypeError):
            return val

    def extract_data(self):
        print("\nExtracting data from Overall S-Curve sheet...")

        # ---- Extract info ----
        self.cutoff_date = self._get_cell_value(1, self.COL_CUTOFF_DATE)
        self.last_month_date = self._get_cell_value(1, self.COL_LAST_MONTH_DATE)

        if self.cutoff_date:
            print(f"  Cut-Off Date:     {self.cutoff_date}")
        if self.last_month_date:
            print(f"  Last Month Date:  {self.last_month_date}")

        # ---- Extract Summary Data (Rows 3-8) ----
        print("\n  Extracting summary data...")
        sn = 0
        for row_idx in range(self.SUMMARY_DATA_START, self.SUMMARY_DATA_END + 1):
            level = self._get_cell_value(row_idx, self.COL_LEVEL)
            wbs_code = self._get_cell_value(row_idx, self.COL_WBS_CODE)
            wbs_name = self._get_cell_value(row_idx, self.COL_WBS_NAME)
            l1_weight = self._get_cell_value(row_idx, self.COL_L1_WEIGHT)

            if not wbs_name and not level:
                continue

            sn += 1
            level_str = str(level or '').strip()
            wbs_code_str = str(wbs_code or '').strip()
            wbs_name_str = str(wbs_name or '').strip()
            is_l0 = (level_str == 'L0')

            lm_plan = self._fmt_pct(self._get_cell_value(row_idx, self.COL_LM_PLAN))
            lm_actual = self._fmt_pct(self._get_cell_value(row_idx, self.COL_LM_ACTUAL))
            lm_var = self._fmt_pct(self._get_cell_value(row_idx, self.COL_LM_VAR))

            tm_plan = self._fmt_pct(self._get_cell_value(row_idx, self.COL_TM_PLAN))
            tm_actual = self._fmt_pct(self._get_cell_value(row_idx, self.COL_TM_ACTUAL))
            tm_var = self._fmt_pct(self._get_cell_value(row_idx, self.COL_TM_VAR))

            l1_weight_pct = self._fmt_pct(l1_weight)

            # Determine status based on This Month variance
            if tm_plan is None and tm_actual is None:
                status = "Not Started"
            elif tm_plan == 0 and tm_actual == 0:
                status = "Not Started"
            elif tm_var is not None:
                try:
                    if float(tm_var) >= 0:
                        status = "On Time"
                    else:
                        status = "Delayed"
                except (ValueError, TypeError):
                    status = "-"
            else:
                status = "-"

            record = {
                'sn': sn,
                'level': level_str,
                'wbs_code': wbs_code_str,
                'discipline': wbs_name_str,
                'l1_weight': l1_weight_pct,
                'lm_plan': lm_plan,
                'lm_actual': lm_actual,
                'lm_var': lm_var,
                'tm_plan': tm_plan,
                'tm_actual': tm_actual,
                'tm_var': tm_var,
                'status': status,
                'is_l0': is_l0
            }
            self.summary_records.append(record)

        print(f"  ✓ Summary: {len(self.summary_records)} records extracted")

        # ---- Extract Monthly S-Curve Data (Rows 35-40) ----
        print("  Extracting monthly S-Curve data...")

        # Find how many monthly periods exist
        col = self.MONTHLY_DATA_START_COL
        while True:
            period_label = self._get_cell_value(self.ROW_PERIOD_LABELS, col)
            if period_label is None:
                break

            date_val = self._get_cell_value(self.ROW_DATES, col)
            baseline_ep = self._fmt_pct(self._get_cell_value(self.ROW_BASELINE_EP, col))
            revised_bl_ep = self._fmt_pct(self._get_cell_value(self.ROW_REVISED_BL_EP, col))
            revised_bl_lp = self._fmt_pct(self._get_cell_value(self.ROW_REVISED_BL_LP, col))
            cumm_actual = self._fmt_pct(self._get_cell_value(self.ROW_CUMM_ACTUAL, col))

            month_record = {
                'period': str(period_label),
                'date': date_val,
                'baseline_ep': baseline_ep,
                'revised_bl_ep': revised_bl_ep,
                'revised_bl_lp': revised_bl_lp,
                'cumm_actual': cumm_actual
            }
            self.monthly_data.append(month_record)
            col += 1

        print(f"  ✓ Monthly S-Curve: {len(self.monthly_data)} periods extracted")
        print(f"\n✓ Data extraction complete")

    def generate_output(self, output_file):
        print(f"\nGenerating output file: {output_file}")

        new_wb = openpyxl.Workbook()

        # ---- Styles ----
        header_fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        on_time_fill = PatternFill(start_color=self.ON_TIME_COLOR, end_color=self.ON_TIME_COLOR, fill_type="solid")
        delayed_fill = PatternFill(start_color=self.DELAYED_COLOR, end_color=self.DELAYED_COLOR, fill_type="solid")
        not_started_fill = PatternFill(start_color=self.NOT_STARTED_COLOR, end_color=self.NOT_STARTED_COLOR, fill_type="solid")
        l0_fill = PatternFill(start_color=self.L0_FILL_COLOR, end_color=self.L0_FILL_COLOR, fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # ================================================================
        # SHEET 1: Summary Tracker
        # ================================================================
        ws1 = new_wb.active
        ws1.title = "S-Curve Summary Tracker"

        # Row 1: Period info
        cutoff_str = self.cutoff_date.strftime('%d-%b-%Y') if isinstance(self.cutoff_date, datetime) else str(self.cutoff_date or '')
        lm_str = self.last_month_date.strftime('%d-%b-%Y') if isinstance(self.last_month_date, datetime) else str(self.last_month_date or '')

        info_text = f"Cut-Off: {cutoff_str}  |  Last Month: {lm_str}"
        info_cell = ws1.cell(row=1, column=1, value=info_text)
        info_cell.font = Font(bold=True, size=12, color=self.HEADER_COLOR)
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

        header_start_row = 3

        # Row 3: Main headers (merged for Last Month / This Month groups)
        # Merge "Last Month" over cols 6-8, "This Month" over cols 9-11
        ws1.merge_cells(start_row=header_start_row, start_column=6, end_row=header_start_row, end_column=8)
        lm_header = ws1.cell(row=header_start_row, column=6, value="Last Month")
        lm_header.font = header_font
        lm_header.fill = header_fill
        lm_header.alignment = Alignment(horizontal='center', vertical='center')
        lm_header.border = border
        for c in range(7, 9):
            cell = ws1.cell(row=header_start_row, column=c)
            cell.border = border

        ws1.merge_cells(start_row=header_start_row, start_column=9, end_row=header_start_row, end_column=11)
        tm_header = ws1.cell(row=header_start_row, column=9, value="This Month")
        tm_header.font = header_font
        tm_header.fill = header_fill
        tm_header.alignment = Alignment(horizontal='center', vertical='center')
        tm_header.border = border
        for c in range(10, 12):
            cell = ws1.cell(row=header_start_row, column=c)
            cell.border = border

        # Fill header row cells for cols 1-5 and 12 (merged vertically)
        top_headers = {1: 'SN', 2: 'Level', 3: 'WBS Code', 4: 'Discipline', 5: 'L1 Weight %', 12: 'Status'}
        for col_idx, title in top_headers.items():
            ws1.merge_cells(start_row=header_start_row, start_column=col_idx, end_row=header_start_row + 1, end_column=col_idx)
            cell = ws1.cell(row=header_start_row, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            # Bottom cell of merge also needs border
            ws1.cell(row=header_start_row + 1, column=col_idx).border = border

        # Row 4: Sub-headers for Last Month / This Month
        sub_headers = {
            6: 'Plan %', 7: 'Actual %', 8: 'Var %',
            9: 'Plan %', 10: 'Actual %', 11: 'Var %'
        }
        for col_idx, title in sub_headers.items():
            cell = ws1.cell(row=header_start_row + 1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        data_start_row = header_start_row + 2  # Row 5

        # Write summary data
        delayed_count = 0
        on_time_count = 0
        not_started_count = 0

        for row_offset, record in enumerate(self.summary_records):
            row_idx = data_start_row + row_offset
            is_l0 = record['is_l0']
            highlight_fill = l0_fill if is_l0 else None

            # Column 1: SN
            cell = ws1.cell(row=row_idx, column=1, value=record['sn'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if highlight_fill:
                cell.fill = highlight_fill
                cell.font = Font(bold=True)

            # Column 2: Level
            cell = ws1.cell(row=row_idx, column=2, value=record['level'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if highlight_fill:
                cell.fill = highlight_fill
                cell.font = Font(bold=True)

            # Column 3: WBS Code
            cell = ws1.cell(row=row_idx, column=3, value=record['wbs_code'])
            cell.border = border
            if highlight_fill:
                cell.fill = highlight_fill
                cell.font = Font(bold=True)

            # Column 4: Discipline
            cell = ws1.cell(row=row_idx, column=4, value=record['discipline'])
            cell.border = border
            if highlight_fill:
                cell.fill = highlight_fill
                cell.font = Font(bold=True)

            # Column 5: L1 Weight %
            val = record['l1_weight']
            display = f"{val:.2f}%" if val is not None else ''
            cell = ws1.cell(row=row_idx, column=5, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if highlight_fill:
                cell.fill = highlight_fill
                cell.font = Font(bold=True)

            # Columns 6-8: Last Month (Plan, Actual, Var)
            for col_off, key in enumerate(['lm_plan', 'lm_actual', 'lm_var']):
                val = record[key]
                display = f"{val:.2f}%" if val is not None else ''
                cell = ws1.cell(row=row_idx, column=6 + col_off, value=display)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                if highlight_fill:
                    cell.fill = highlight_fill
                    cell.font = Font(bold=True)
                elif key == 'lm_var' and val is not None and not is_l0:
                    try:
                        if float(val) < 0:
                            cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                        elif float(val) > 0:
                            cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                    except (ValueError, TypeError):
                        pass

            # Columns 9-11: This Month (Plan, Actual, Var)
            for col_off, key in enumerate(['tm_plan', 'tm_actual', 'tm_var']):
                val = record[key]
                display = f"{val:.2f}%" if val is not None else ''
                cell = ws1.cell(row=row_idx, column=9 + col_off, value=display)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                if highlight_fill:
                    cell.fill = highlight_fill
                    cell.font = Font(bold=True)
                elif key == 'tm_var' and val is not None and not is_l0:
                    try:
                        if float(val) < 0:
                            cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                        elif float(val) > 0:
                            cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                    except (ValueError, TypeError):
                        pass

            # Column 12: Status
            cell = ws1.cell(row=row_idx, column=12, value=record['status'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            if highlight_fill:
                cell.fill = highlight_fill
                cell.font = Font(bold=True)
            elif record['status'] == 'Delayed':
                cell.fill = delayed_fill
                cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                delayed_count += 1
            elif record['status'] == 'On Time':
                cell.fill = on_time_fill
                cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                on_time_count += 1
            elif record['status'] == 'Not Started':
                cell.fill = not_started_fill
                cell.font = Font(color=self.NOT_STARTED_FONT_COLOR, bold=True)
                not_started_count += 1

        # Column widths for Sheet 1
        column_widths_s1 = {
            'A': 6, 'B': 8, 'C': 16, 'D': 45, 'E': 14,
            'F': 12, 'G': 12, 'H': 12,
            'I': 12, 'J': 12, 'K': 12,
            'L': 15
        }
        for col, width in column_widths_s1.items():
            ws1.column_dimensions[col].width = width

        ws1.freeze_panes = f'A{data_start_row}'

        # ================================================================
        # SHEET 2: Monthly S-Curve Data
        # ================================================================
        ws2 = new_wb.create_sheet(title="Monthly S-Curve Data")

        # Info row
        ws2.cell(row=1, column=1, value=f"Overall S-Curve Monthly Data  |  Cut-Off: {cutoff_str}").font = Font(bold=True, size=12, color=self.HEADER_COLOR)
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        s2_header_row = 3

        s2_headers = [
            'Period',
            'Date',
            'Baseline EP %',
            'Revised BL-EP %',
            'Revised BL-LP %',
            'Cumm. Actual %'
        ]

        for col_idx, header in enumerate(s2_headers, start=1):
            cell = ws2.cell(row=s2_header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Write monthly data
        actual_count = 0
        for row_offset, record in enumerate(self.monthly_data):
            row_idx = s2_header_row + 1 + row_offset
            has_actual = record['cumm_actual'] is not None

            # Period
            cell = ws2.cell(row=row_idx, column=1, value=record['period'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Date
            date_val = record['date']
            if isinstance(date_val, datetime):
                display_date = date_val.strftime('%b-%Y')
            else:
                display_date = str(date_val or '')
            cell = ws2.cell(row=row_idx, column=2, value=display_date)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Baseline EP %
            val = record['baseline_ep']
            display = f"{val:.2f}%" if val is not None else ''
            cell = ws2.cell(row=row_idx, column=3, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Revised BL-EP %
            val = record['revised_bl_ep']
            display = f"{val:.2f}%" if val is not None else ''
            cell = ws2.cell(row=row_idx, column=4, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Revised BL-LP %
            val = record['revised_bl_lp']
            display = f"{val:.2f}%" if val is not None else ''
            cell = ws2.cell(row=row_idx, column=5, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Cumm. Actual %
            val = record['cumm_actual']
            display = f"{val:.2f}%" if val is not None else ''
            cell = ws2.cell(row=row_idx, column=6, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if has_actual:
                actual_count += 1
                # Highlight actual data rows with light background
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                cell.font = Font(bold=True)

        # Column widths for Sheet 2
        column_widths_s2 = {
            'A': 10, 'B': 14, 'C': 16, 'D': 16, 'E': 16, 'F': 16
        }
        for col, width in column_widths_s2.items():
            ws2.column_dimensions[col].width = width

        ws2.freeze_panes = f'A{s2_header_row + 1}'

        # Save
        try:
            new_wb.save(output_file)
            print(f"✓ Output file created successfully")
        except Exception as e:
            raise RuntimeError(f"Failed to save: {str(e)}")
        finally:
            new_wb.close()

        print(f"\n{'='*60}")
        print("OVERALL S-CURVE - SUMMARY STATISTICS")
        print(f"{'='*60}")
        print(f"Summary Records:    {len(self.summary_records):,}")
        print(f"  Delayed:          {delayed_count:,}")
        print(f"  On Time:          {on_time_count:,}")
        print(f"  Not Started:      {not_started_count:,}")
        print(f"Monthly Periods:    {len(self.monthly_data):,}")
        print(f"  With Actuals:     {actual_count:,}")
        print(f"{'='*60}")

    def close(self):
        if self.workbook:
            self.workbook.close()

    def process(self, output_file):
        try:
            self.validate_input_file()
            self.load_workbook()
            self.extract_data()
            self.generate_output(output_file)
        finally:
            self.close()


def main():
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT_FILE
        print("=" * 60)
        print("OVERALL S-CURVE - TRACKER GENERATOR")
        print("=" * 60)
        print()
    else:
        print("=" * 60)
        print("OVERALL S-CURVE - TRACKER GENERATOR")
        print("=" * 60)
        print()
        print("DEFAULT INPUT FILE:")
        print(f"  {DEFAULT_INPUT_FILE}")
        print()
        user_input = input("Press ENTER to use default, or type new path: ").strip()
        input_file = user_input.strip('"').strip("'") if user_input else DEFAULT_INPUT_FILE

        print()
        print("DEFAULT OUTPUT FILE:")
        print(f"  {DEFAULT_OUTPUT_FILE}")
        print()
        user_output = input("Press ENTER to use default, or type new name: ").strip()
        output_file = user_output.strip('"').strip("'") if user_output else DEFAULT_OUTPUT_FILE
        print()
        print("=" * 60)
        print()

    try:
        print(f"Input:  {input_file}")
        print(f"Output: {output_file}")
        print()

        processor = OverallSCurveProcessor(input_file)
        processor.process(output_file)

        print()
        print("✓ Processing completed successfully!")
        print(f"✓ Output saved to: {output_file}")
        print()
    except Exception as e:
        print()
        print(f"✗ ERROR: {str(e)}")
        print()
        input("Press ENTER to exit...")
        sys.exit(1)

    if len(sys.argv) < 2:
        input("\nPress ENTER to exit...")


if __name__ == "__main__":
    main()
