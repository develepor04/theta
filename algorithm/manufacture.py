#!/usr/bin/env python3
"""
Manufacture Activity Timeline Tracker Generator
=========================================================
Processes Borouge Excel files - 'Manufacture' sheet
with stage gate milestones and EP/LP/F/A date tracking.

Column Order (Output):
Activity ID | Activity Name | Stage Gate | Early Planning |
Late Planning | Actual Date | Duration Deviation | Timeline Flag

Stage Gates (9 milestones):
  1. Start PO
  2. Key Vendor Drawings / Data Approved
  3. PO for Major Sub-ordered Materials
  4. Major Sub-ordered Materials Received
  5. Start Manufacturing / PIM
  6. Manufacturing
  7. Complete Manufacturing & ready for Final Inspection
  8. Materials Successfully Inspected / Ex Work Delivered
  9. All Materials Received on SITE

Author: Claude
Date: 2026-02-17
Version: 1.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys
from pathlib import Path


# ============================================================================
# DEFAULT FILE PATHS
# ============================================================================
DEFAULT_INPUT_FILE = r"2026.01.30_Borouge EU3 H2 Extraction Project PMS-Rev1 dates (1).xlsx"
DEFAULT_OUTPUT_FILE = r"01-03-26\manufacture_tracker.xlsx"
# ============================================================================


class ManufactureProcessor:
    """Processes Manufacture sheet and generates Activity Timeline Tracker."""

    # Column mappings from source sheet
    COL_LEVEL = 2             # Col B - Level (L0-L5 or Activity ID)
    COL_CODE_NAME = 3         # Col C - WBS Code / Activity ID (EP rows: equipment name)
    COL_ACTIVITY_NAME = 4     # Col D - WBS / Activity Name
    COL_STAGE_GATE = 8        # Col H - Stage Gate (EP/LP/F/A)

    # Milestone date columns (9 milestones)
    MILESTONES = [
        ('Start PO', 9),
        ('Key Vendor Drawings / Data Approved', 10),
        ('PO for Major Sub-ordered Materials', 11),
        ('Major Sub-ordered Materials Received', 12),
        ('Start Manufacturing / PIM', 13),
        ('Manufacturing', 14),
        ('Complete Manufacturing & Ready for Inspection', 15),
        ('Materials Inspected / Ex Work Delivered', 16),
        ('All Materials Received on SITE', 17),
    ]

    # Header row and data start
    HEADER_ROW = 5
    DATA_START_ROW = 6

    # Style definitions
    HEADER_COLOR = "366092"
    ON_TIME_COLOR = "C6EFCE"
    ON_TIME_FONT_COLOR = "006100"
    DELAYED_COLOR = "FFC7CE"
    DELAYED_FONT_COLOR = "9C0006"
    NOT_STARTED_COLOR = "FFF2CC"
    NOT_STARTED_FONT_COLOR = "7F6000"

    def __init__(self, input_file):
        self.input_file = Path(input_file)
        self.workbook = None
        self.sheet = None
        self.activities = {}
        self.activity_order = []
        self.data_records = []

    def validate_input_file(self):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if self.input_file.suffix.lower() not in ['.xlsx', '.xlsm']:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}")
        print(f"✓ Input file validated: {self.input_file.name}")

    def load_workbook(self):
        try:
            print("Loading workbook...")
            self.workbook = openpyxl.load_workbook(self.input_file, data_only=True)

            sheet_name = 'Manufacture'
            if sheet_name not in self.workbook.sheetnames:
                for name in self.workbook.sheetnames:
                    if 'manufactur' in name.lower():
                        sheet_name = name
                        break
                else:
                    raise ValueError(
                        f"'Manufacture' sheet not found. "
                        f"Available: {self.workbook.sheetnames}"
                    )

            self.sheet = self.workbook[sheet_name]
            print(f"✓ '{sheet_name}' sheet loaded: {self.sheet.max_row} rows × {self.sheet.max_column} columns")
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {str(e)}")

    def _get_cell_value(self, row, col):
        try:
            return self.sheet.cell(row=row, column=col).value
        except:
            return None

    def extract_data(self):
        print("\nExtracting activity data...")

        total_rows = self.sheet.max_row
        processed_count = 0
        group_index = 0
        current_group_key = None

        for row_idx in range(self.DATA_START_ROW, total_rows + 1):
            try:
                level = self._get_cell_value(row_idx, self.COL_LEVEL)
                code_name = self._get_cell_value(row_idx, self.COL_CODE_NAME)
                activity_name = self._get_cell_value(row_idx, self.COL_ACTIVITY_NAME)
                stage_gate = self._get_cell_value(row_idx, self.COL_STAGE_GATE)

                if not stage_gate:
                    continue

                stage_gate_str = str(stage_gate).strip()
                if stage_gate_str not in ('EP', 'LP', 'F', 'A'):
                    continue

                # Get milestone dates from cols 9-17
                milestone_dates = {}
                for ms_name, ms_col in self.MILESTONES:
                    milestone_dates[ms_name] = self._get_cell_value(row_idx, ms_col)

                level_str = str(level or '').strip()

                if stage_gate_str == 'EP':
                    # EP row = start of new activity group
                    group_index += 1
                    equipment_name = str(code_name or '').strip()

                    current_group_key = (equipment_name, group_index)
                    self.activities[current_group_key] = {
                        'activity_name': equipment_name,
                        'activity_code': equipment_name,  # Will be updated from LP row
                        'stages': {'EP': milestone_dates}
                    }
                    self.activity_order.append(current_group_key)

                elif current_group_key:
                    # LP/F/A rows
                    self.activities[current_group_key]['stages'][stage_gate_str] = milestone_dates

                    # Use LP row's activity code as Activity ID
                    if stage_gate_str == 'LP' and level_str and level_str.startswith('A'):
                        self.activities[current_group_key]['activity_code'] = level_str

                processed_count += 1

            except Exception as e:
                continue

        print(f"✓ Data extraction complete: {processed_count} rows processed")
        print(f"  Unique activity groups: {len(self.activities)}")

        self._create_milestone_records()

    def _create_milestone_records(self):
        print("\nCreating milestone records...")
        record_count = 0

        for group_key in self.activity_order:
            data = self.activities[group_key]
            activity_name = data['activity_name']
            activity_code = data['activity_code']
            stages = data['stages']

            for ms_name, ms_col in self.MILESTONES:
                ep_date = stages.get('EP', {}).get(ms_name)
                lp_date = stages.get('LP', {}).get(ms_name)
                a_date = stages.get('A', {}).get(ms_name)
                f_date = stages.get('F', {}).get(ms_name)

                has_any_date = any(
                    d is not None and d != '' and d != 0
                    for d in [ep_date, lp_date, a_date, f_date]
                )

                if has_any_date:
                    duration_deviation, timeline_flag = self._calculate_deviation(
                        ep_date, lp_date, a_date
                    )

                    record = {
                        'activity_id': activity_code,
                        'activity_name': activity_name,
                        'stage_gate': ms_name,
                        'planned_start_date': ep_date,
                        'planned_end_date': lp_date,
                        'actual_date': a_date,
                        'duration_deviation': duration_deviation,
                        'timeline_flag': timeline_flag
                    }

                    self.data_records.append(record)
                    record_count += 1

        print(f"✓ Created {record_count} milestone records")

    def _calculate_deviation(self, planned_start_ep, planned_end_lp, actual_date_a):
        """Calculate deviation between Planned End (LP) and Actual Date (A)."""
        duration_deviation = None
        today = datetime.now()

        if not actual_date_a or not isinstance(actual_date_a, datetime):
            # No actual date - check if LP is in the past (overdue)
            if planned_end_lp and isinstance(planned_end_lp, datetime) and planned_end_lp < today:
                duration_deviation = (today - planned_end_lp).days
                timeline_flag = "Delayed"
            elif planned_start_ep and isinstance(planned_start_ep, datetime) and planned_start_ep < today:
                duration_deviation = (today - planned_start_ep).days
                timeline_flag = "Delayed"
            else:
                timeline_flag = "Not Started"
        elif planned_end_lp and isinstance(planned_end_lp, datetime):
            duration_deviation = (actual_date_a - planned_end_lp).days
            timeline_flag = "Delayed" if actual_date_a > planned_end_lp else "On Time"
        elif planned_start_ep and isinstance(planned_start_ep, datetime):
            duration_deviation = (actual_date_a - planned_start_ep).days
            timeline_flag = "Delayed" if actual_date_a > planned_start_ep else "On Time"
        else:
            timeline_flag = "Not Started"

        # Manual Error: LP date is before EP date
        if (planned_start_ep and isinstance(planned_start_ep, datetime) and
            planned_end_lp and isinstance(planned_end_lp, datetime) and
            planned_end_lp < planned_start_ep):
            timeline_flag = "Manual Error"

        return duration_deviation, timeline_flag

    def _format_date(self, date_value):
        if isinstance(date_value, datetime):
            return date_value.strftime('%d-%b-%Y')
        return None

    def generate_output(self, output_file):
        print(f"\nGenerating output file: {output_file}")

        out_wb = openpyxl.Workbook()
        ws = out_wb.active
        ws.title = "Manufacture Tracker"

        # ── Define Styles ──
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color=self.HEADER_COLOR,
                                  end_color=self.HEADER_COLOR, fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        on_time_fill = PatternFill(start_color=self.ON_TIME_COLOR,
                                   end_color=self.ON_TIME_COLOR, fill_type='solid')
        on_time_font = Font(name='Calibri', color=self.ON_TIME_FONT_COLOR, size=11)

        delayed_fill = PatternFill(start_color=self.DELAYED_COLOR,
                                   end_color=self.DELAYED_COLOR, fill_type='solid')
        delayed_font = Font(name='Calibri', color=self.DELAYED_FONT_COLOR, size=11)

        not_started_fill = PatternFill(start_color=self.NOT_STARTED_COLOR,
                                       end_color=self.NOT_STARTED_COLOR, fill_type='solid')
        not_started_font = Font(name='Calibri', color=self.NOT_STARTED_FONT_COLOR, size=11)

        manual_error_fill = PatternFill(start_color="FF9999",
                                        end_color="FF9999", fill_type='solid')
        manual_error_font = Font(name='Calibri', color="800000", bold=True, size=11)

        data_font = Font(name='Calibri', size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ── Write Headers ──
        headers = [
            'Activity ID',
            'Activity Name',
            'Stage Gate',
            'Early Planning',
            'Late Planning',
            'Actual Date',
            'Duration Deviation',
            'Timeline Flag'
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # ── Write Data ──
        for row_idx, record in enumerate(self.data_records, 2):
            ws.cell(row=row_idx, column=1, value=record['activity_id']).font = data_font
            ws.cell(row=row_idx, column=1).alignment = center_align
            ws.cell(row=row_idx, column=1).border = thin_border

            ws.cell(row=row_idx, column=2, value=record['activity_name']).font = data_font
            ws.cell(row=row_idx, column=2).alignment = left_align
            ws.cell(row=row_idx, column=2).border = thin_border

            ws.cell(row=row_idx, column=3, value=record['stage_gate']).font = data_font
            ws.cell(row=row_idx, column=3).alignment = center_align
            ws.cell(row=row_idx, column=3).border = thin_border

            # Dates
            ws.cell(row=row_idx, column=4,
                    value=self._format_date(record['planned_start_date'])).font = data_font
            ws.cell(row=row_idx, column=4).alignment = center_align
            ws.cell(row=row_idx, column=4).border = thin_border

            ws.cell(row=row_idx, column=5,
                    value=self._format_date(record['planned_end_date'])).font = data_font
            ws.cell(row=row_idx, column=5).alignment = center_align
            ws.cell(row=row_idx, column=5).border = thin_border

            ws.cell(row=row_idx, column=6,
                    value=self._format_date(record['actual_date'])).font = data_font
            ws.cell(row=row_idx, column=6).alignment = center_align
            ws.cell(row=row_idx, column=6).border = thin_border

            # Duration Deviation
            ws.cell(row=row_idx, column=7, value=record['duration_deviation']).font = data_font
            ws.cell(row=row_idx, column=7).alignment = center_align
            ws.cell(row=row_idx, column=7).border = thin_border

            # Timeline Flag with color
            flag = record['timeline_flag']
            flag_cell = ws.cell(row=row_idx, column=8, value=flag)
            flag_cell.alignment = center_align
            flag_cell.border = thin_border

            if flag == 'On Time':
                flag_cell.fill = on_time_fill
                flag_cell.font = on_time_font
            elif flag == 'Delayed':
                flag_cell.fill = delayed_fill
                flag_cell.font = delayed_font
            elif flag == 'Not Started':
                flag_cell.fill = not_started_fill
                flag_cell.font = not_started_font
            elif flag == 'Manual Error':
                flag_cell.fill = manual_error_fill
                flag_cell.font = manual_error_font
            else:
                flag_cell.font = data_font

        # ── Column Widths ──
        col_widths = {
            1: 16,   # Activity ID
            2: 50,   # Activity Name
            3: 48,   # Stage Gate
            4: 22,   # Planned Start Date
            5: 22,   # Planned End Date
            6: 18,   # Actual Date
            7: 20,   # Duration Deviation
            8: 16,   # Timeline Flag
        }

        for col, width in col_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # Freeze top row
        ws.freeze_panes = 'A2'

        # Auto-filter
        ws.auto_filter.ref = f"A1:H{len(self.data_records) + 1}"

        try:
            out_wb.save(output_file)
            print(f"✓ Output file created successfully")
        except PermissionError:
            alt_file = output_file.replace('.xlsx', '_new.xlsx')
            out_wb.save(alt_file)
            print(f"⚠ Permission denied for {output_file}")
            print(f"✓ Saved to: {alt_file}")

    def print_summary(self):
        total = len(self.data_records)
        delayed = sum(1 for r in self.data_records if r['timeline_flag'] == 'Delayed')
        on_time = sum(1 for r in self.data_records if r['timeline_flag'] == 'On Time')
        not_started = sum(1 for r in self.data_records if r['timeline_flag'] == 'Not Started')
        manual_error = sum(1 for r in self.data_records if r['timeline_flag'] == 'Manual Error')

        print("\n" + "=" * 60)
        print("MANUFACTURE - SUMMARY STATISTICS")
        print("=" * 60)
        print(f"Total Records:      {total}")
        print(f"Delayed:            {delayed}")
        print(f"On Time:            {on_time}")
        print(f"Not Started:        {not_started}")
        print(f"Manual Error:       {manual_error}")
        print("=" * 60)

    def process(self, output_file):
        print("=" * 60)
        print("MANUFACTURE - TRACKER GENERATOR")
        print("=" * 60)
        print(f"\nInput:  {self.input_file}")
        print(f"Output: {output_file}\n")

        self.validate_input_file()
        self.load_workbook()
        self.extract_data()
        self.generate_output(output_file)
        self.print_summary()

        print(f"\n✓ Processing completed successfully!")
        print(f"✓ Output saved to: {output_file}")


def main():
    if len(sys.argv) >= 3:
        input_file = sys.argv[1]
        output_file = sys.argv[2]
    elif len(sys.argv) == 2:
        input_file = sys.argv[1]
        output_file = DEFAULT_OUTPUT_FILE
    else:
        input_file = DEFAULT_INPUT_FILE
        output_file = DEFAULT_OUTPUT_FILE

    processor = ManufactureProcessor(input_file)
    processor.process(output_file)


if __name__ == "__main__":
    main()
