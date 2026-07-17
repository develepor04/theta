#!/usr/bin/env python3
"""
HO-Subcontract Activity Timeline Tracker Generator
=========================================================
Processes Borouge Excel files - 'HO-Subcontract' sheet
with stage gate milestones and EP/LP/F/A date tracking.

Column Order (Output):
Activity ID | Activity Name | Stage Gate | Early Planning |
Late Planning | Actual Date | Duration Deviation | Timeline Flag

Stage Gates: Issue RFQ to CONTRACTORs, Technical Bid analysis,
             Bid analysis, Subcontract Awarded/Signed

Author: Claude
Date: 2026-02-17
Version: 1.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys
from pathlib import Path


# ============================================================================
# DEFAULT FILE PATHS
# ============================================================================
DEFAULT_INPUT_FILE = r"2026.01.30_Borouge EU3 H2 Extraction Project PMS-Rev1 dates (1).xlsx"
DEFAULT_OUTPUT_FILE = r"01-03-26\ho_subcontract_tracker.xlsx"
# ============================================================================


class HOSubcontractProcessor:
    """Processes HO-Subcontract sheet and generates Activity Timeline Tracker."""

    # Default column mappings (will be auto-detected)
    COL_LEVEL = 1             # Col A - Level
    COL_ACTIVITY_CODE = 2     # Col B - Activity ID
    COL_ACTIVITY_NAME = 3     # Col C - Activity Name
    COL_STAGE_GATE = 7        # Col G - Stage Gate (EP/LP/F/A)

    # Milestone date columns (relative to stage gate column)
    COL_ISSUE_RFQ = 8         # Col H - Issue RFQ to CONTRACTORs
    COL_TECH_BID = 9          # Col I - Technical Bid analysis and clarifications
    COL_BID_ANALYSIS = 10     # Col J - Bid analysis and clarifications
    COL_SUBCONTRACT = 11      # Col K - Subcontract Awarded/Signed

    # Data rows (will be auto-detected)
    DATA_START_ROW = 8

    # Milestone names
    MILESTONES = [
        ('Issue RFQ to CONTRACTORs', 8),
        ('Technical Bid Analysis', 9),
        ('Bid Analysis & Clarifications', 10),
        ('Subcontract Awarded/Signed', 11),
    ]

    # Style definitions
    HEADER_COLOR = "366092"
    ON_TIME_COLOR = "C6EFCE"
    ON_TIME_FONT_COLOR = "006100"
    DELAYED_COLOR = "FFC7CE"
    DELAYED_FONT_COLOR = "9C0006"

    def __init__(self, input_file):
        self.input_file = Path(input_file)
        self.workbook = None
        self.sheet = None
        self.activities = {}
        self.activity_order = []
        self.data_records = []

    def validate_input_file(self):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if self.input_file.suffix.lower() not in ['.xlsx', '.xlsm']:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}")
        print(f"✓ Input file validated: {self.input_file.name}")

    def load_workbook(self):
        try:
            print("Loading workbook...")
            self.workbook = openpyxl.load_workbook(self.input_file, data_only=True, read_only=True)

            sheet_name = 'HO-Subcontract'
            if sheet_name not in self.workbook.sheetnames:
                for name in self.workbook.sheetnames:
                    if 'ho' in name.lower() and 'subcontract' in name.lower():
                        sheet_name = name
                        break
                else:
                    raise ValueError(
                        f"'HO-Subcontract' sheet not found. "
                        f"Available: {self.workbook.sheetnames}"
                    )

            self.sheet = self.workbook[sheet_name]
            print(f"✓ '{sheet_name}' sheet loaded: {self.sheet.max_row} rows × {self.sheet.max_column} columns")
            
            # Pre-load all data for fast access in read-only mode
            print("  → Caching sheet data for fast access...")
            self.sheet_data = [list(row) for row in self.sheet.iter_rows(values_only=True)]
            print(f"  → Cached {len(self.sheet_data)} rows")
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {str(e)}")

    def _get_cell_value(self, row, col):
        try:
            # Access cached data (0-indexed)
            if 0 < row <= len(self.sheet_data) and 0 < col <= len(self.sheet_data[0]):
                return self.sheet_data[row - 1][col - 1]
            return None
        except:
            return None

    def extract_data(self):
        print("\nExtracting activity data...")
        
        # Auto-detect structure
        self._detect_structure()

        total_rows = self.sheet.max_row
        processed_count = 0
        group_index = 0
        current_group_key = None
        current_parent_name = None

        # First pass: identify parent L3 names for each EP group
        for row_idx in range(self.DATA_START_ROW, total_rows + 1):
            try:
                level = self._get_cell_value(row_idx, self.COL_LEVEL)
                activity_code = self._get_cell_value(row_idx, self.COL_ACTIVITY_CODE)
                activity_name = self._get_cell_value(row_idx, self.COL_ACTIVITY_NAME)
                stage_gate = self._get_cell_value(row_idx, self.COL_STAGE_GATE)

                # L3 header rows (parent name)
                if level and str(level).strip() == 'L3' and activity_name:
                    current_parent_name = str(activity_name or '').strip()
                    continue

                if not stage_gate:
                    continue

                stage_gate_str = str(stage_gate).strip()
                if stage_gate_str not in ('EP', 'LP', 'F', 'A'):
                    continue

                code_str = str(activity_code or '').strip()
                name_str = str(activity_name or '').strip()

                # Get milestone dates
                milestone_dates = {}
                for ms_name, ms_col in self.MILESTONES:
                    milestone_dates[ms_name] = self._get_cell_value(row_idx, ms_col)

                # EP row = start of a new activity group
                if stage_gate_str == 'EP':
                    group_index += 1
                    # Use parent name if available, otherwise use name from row
                    display_name = current_parent_name if current_parent_name else (name_str or code_str)
                    current_group_key = (display_name, group_index)
                    self.activities[current_group_key] = {
                        'activity_name': display_name,
                        'activity_code': code_str,
                        'stages': {'EP': milestone_dates}
                    }
                    self.activity_order.append(current_group_key)
                elif current_group_key:
                    self.activities[current_group_key]['stages'][stage_gate_str] = milestone_dates
                else:
                    # Standalone stage (no EP found yet)
                    group_index += 1
                    standalone_key = (name_str or code_str, group_index)
                    self.activities[standalone_key] = {
                        'activity_name': name_str or code_str,
                        'activity_code': code_str,
                        'stages': {stage_gate_str: milestone_dates}
                    }
                    self.activity_order.append(standalone_key)
                    current_group_key = standalone_key

                processed_count += 1

            except Exception as e:
                continue

        print(f"✓ Data extraction complete: {processed_count} rows processed")
        print(f"  Unique activity groups: {len(self.activities)}")

        self._create_milestone_records()
    
    def _detect_structure(self):
        """Auto-detect header row and column positions."""
        print("  → Auto-detecting sheet structure...")
        
        # Look for header row in first 10 rows
        for row_idx in range(1, min(11, self.sheet.max_row + 1)):
            row_vals = [str(self.sheet.cell(row_idx, c).value or '').lower() 
                       for c in range(1, min(15, self.sheet.max_column + 1))]
            
            # Check for key headers
            has_level = any('level' in v for v in row_vals)
            has_activity = any('activity' in v or 'wbs' in v for v in row_vals)
            has_stage = any('stage' in v and 'gate' in v for v in row_vals)
            
            if has_level and has_activity:
                print(f"  → Found header row at row {row_idx}")
                # Find column positions
                for idx, val in enumerate(row_vals, 1):
                    if 'level' in val:
                        self.COL_LEVEL = idx
                    elif ('wbs code' in val or 'activity id' in val) and 'name' not in val:
                        self.COL_ACTIVITY_CODE = idx
                    elif 'activity name' in val or 'wbs / activity name' in val:
                        self.COL_ACTIVITY_NAME = idx
                    elif 'stage' in val and 'gate' in val:
                        self.COL_STAGE_GATE = idx
                
                self.DATA_START_ROW = row_idx + 1
                print(f"  → Data starts at row {self.DATA_START_ROW}")
                print(f"  → Columns: Level={self.COL_LEVEL}, Code={self.COL_ACTIVITY_CODE}, Name={self.COL_ACTIVITY_NAME}, Gate={self.COL_STAGE_GATE}")
                break

    def _create_milestone_records(self):
        print("\nCreating milestone records...")
        record_count = 0

        for group_key in self.activity_order:
            data = self.activities[group_key]
            activity_name = data['activity_name']
            activity_code = data['activity_code']
            stages = data['stages']

            activity_has_any_record = False

            for ms_name, ms_col in self.MILESTONES:
                ep_date = stages.get('EP', {}).get(ms_name)
                lp_date = stages.get('LP', {}).get(ms_name)
                a_date = stages.get('A', {}).get(ms_name)
                f_date = stages.get('F', {}).get(ms_name)

                has_any_date = any(
                    d is not None and d != '' and d != 0
                    for d in [ep_date, lp_date, a_date, f_date]
                )

                if has_any_date:
                    duration_deviation, timeline_flag = self._calculate_deviation(
                        ep_date, lp_date, a_date
                    )

                    record = {
                        'activity_id': activity_code,
                        'activity_name': activity_name,
                        'stage_gate': ms_name,
                        'planned_start_date': ep_date,
                        'planned_end_date': lp_date,
                        'actual_date': a_date,
                        'duration_deviation': duration_deviation,
                        'timeline_flag': timeline_flag
                    }

                    self.data_records.append(record)
                    record_count += 1
                    activity_has_any_record = True

            if not activity_has_any_record:
                record = {
                    'activity_id': activity_code,
                    'activity_name': activity_name,
                    'stage_gate': '-',
                    'planned_start_date': None,
                    'planned_end_date': None,
                    'actual_date': None,
                    'duration_deviation': None,
                    'timeline_flag': '-'
                }
                self.data_records.append(record)
                record_count += 1

        print(f"✓ Created {record_count} milestone records")

    def _calculate_deviation(self, planned_start_ep, planned_end_lp, actual_date_a):
        """Calculate deviation between Planned End (LP) and Actual Date (A)."""
        duration_deviation = None
        today = datetime.now()

        if not actual_date_a or not isinstance(actual_date_a, datetime):
            # No actual date - check if LP is in the past (overdue)
            if planned_end_lp and isinstance(planned_end_lp, datetime) and planned_end_lp < today:
                duration_deviation = (today - planned_end_lp).days
                timeline_flag = "Delayed"
            elif planned_start_ep and isinstance(planned_start_ep, datetime) and planned_start_ep < today:
                duration_deviation = (today - planned_start_ep).days
                timeline_flag = "Delayed"
            else:
                timeline_flag = "Not Started"
        elif planned_end_lp and isinstance(planned_end_lp, datetime):
            duration_deviation = (actual_date_a - planned_end_lp).days
            timeline_flag = "Delayed" if actual_date_a > planned_end_lp else "On Time"
        elif planned_start_ep and isinstance(planned_start_ep, datetime):
            duration_deviation = (actual_date_a - planned_start_ep).days
            timeline_flag = "Delayed" if actual_date_a > planned_start_ep else "On Time"
        else:
            timeline_flag = "-"

        # Manual Error: LP date is before EP date
        if (planned_start_ep and isinstance(planned_start_ep, datetime) and
            planned_end_lp and isinstance(planned_end_lp, datetime) and
            planned_end_lp < planned_start_ep):
            timeline_flag = "Manual Error"

        return duration_deviation, timeline_flag

    def generate_output(self, output_file):
        print(f"\nGenerating output file: {output_file}")

        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "HO Subcontract Tracker"

        # Styles
        header_fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        on_time_fill = PatternFill(start_color=self.ON_TIME_COLOR, end_color=self.ON_TIME_COLOR, fill_type="solid")
        delayed_fill = PatternFill(start_color=self.DELAYED_COLOR, end_color=self.DELAYED_COLOR, fill_type="solid")
        not_started_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        manual_error_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        header_start_row = 1

        headers = [
            'Activity ID',
            'Activity Name',
            'Stage Gate',
            'Early Planning',
            'Late Planning',
            'Actual Date',
            'Duration Deviation',
            'Timeline Flag'
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = new_ws.cell(row=header_start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Write data
        delayed_count = 0
        on_time_count = 0
        not_started_count = 0
        manual_error_count = 0

        for row_offset, record in enumerate(self.data_records):
            row_idx = header_start_row + 1 + row_offset

            # Activity ID
            cell = new_ws.cell(row=row_idx, column=1, value=record['activity_id'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Activity Name
            cell = new_ws.cell(row=row_idx, column=2, value=record['activity_name'])
            cell.border = border

            # Stage Gate
            cell = new_ws.cell(row=row_idx, column=3, value=record['stage_gate'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

            # Early Planning
            val = record['planned_start_date']
            if isinstance(val, datetime):
                display = val.strftime('%d-%b-%Y')
            else:
                display = str(val) if val else ''
            cell = new_ws.cell(row=row_idx, column=4, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Late Planning
            val = record['planned_end_date']
            if isinstance(val, datetime):
                display = val.strftime('%d-%b-%Y')
            else:
                display = str(val) if val else ''
            cell = new_ws.cell(row=row_idx, column=5, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Actual Date
            val = record['actual_date']
            if isinstance(val, datetime):
                display = val.strftime('%d-%b-%Y')
            else:
                display = str(val) if val else ''
            cell = new_ws.cell(row=row_idx, column=6, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Duration Deviation
            cell = new_ws.cell(row=row_idx, column=7, value=record['duration_deviation'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if record['duration_deviation'] is not None:
                try:
                    if int(record['duration_deviation']) > 0:
                        cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                    elif int(record['duration_deviation']) <= 0:
                        cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                except (ValueError, TypeError):
                    pass

            # Timeline Flag
            cell = new_ws.cell(row=row_idx, column=8, value=record['timeline_flag'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            if record['timeline_flag'] == 'Delayed':
                cell.fill = delayed_fill
                cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                delayed_count += 1
            elif record['timeline_flag'] == 'On Time':
                cell.fill = on_time_fill
                cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                on_time_count += 1
            elif record['timeline_flag'] == 'Not Started':
                cell.fill = not_started_fill
                cell.font = Font(color="7F6000", bold=True)
                not_started_count += 1
            elif record['timeline_flag'] == 'Manual Error':
                cell.fill = manual_error_fill
                cell.font = Font(color="800000", bold=True)
                manual_error_count += 1

        # Column widths
        column_widths = {
            'A': 15, 'B': 45, 'C': 30,
            'D': 18, 'E': 18, 'F': 18, 'G': 16, 'H': 15
        }
        for col, width in column_widths.items():
            new_ws.column_dimensions[col].width = width

        new_ws.freeze_panes = 'A2'

        try:
            new_wb.save(output_file)
            print("✓ Output file created successfully")
        except Exception as e:
            raise RuntimeError(f"Failed to save: {str(e)}")
        finally:
            new_wb.close()

        print(f"\n{'='*60}")
        print("HO-SUBCONTRACT - SUMMARY STATISTICS")
        print(f"{'='*60}")
        print(f"Total Records:      {len(self.data_records):,}")
        print(f"Delayed:            {delayed_count:,}")
        print(f"On Time:            {on_time_count:,}")
        print(f"Not Started:        {not_started_count:,}")
        print(f"Manual Error:       {manual_error_count:,}")
        print(f"{'='*60}")

    def close(self):
        if self.workbook:
            self.workbook.close()

    def process(self, output_file):
        try:
            self.validate_input_file()
            self.load_workbook()
            self.extract_data()
            self.generate_output(output_file)
        finally:
            self.close()


def main():
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT_FILE
        print("=" * 60)
        print("HO-SUBCONTRACT - TRACKER GENERATOR")
        print("=" * 60)
        print()
    else:
        print("=" * 60)
        print("HO-SUBCONTRACT - TRACKER GENERATOR")
        print("=" * 60)
        print()
        print("DEFAULT INPUT FILE:")
        print(f"  {DEFAULT_INPUT_FILE}")
        print()
        user_input = input("Press ENTER to use default, or type new path: ").strip()
        input_file = user_input.strip('"').strip("'") if user_input else DEFAULT_INPUT_FILE

        print()
        print("DEFAULT OUTPUT FILE:")
        print(f"  {DEFAULT_OUTPUT_FILE}")
        print()
        user_output = input("Press ENTER to use default, or type new name: ").strip()
        output_file = user_output.strip('"').strip("'") if user_output else DEFAULT_OUTPUT_FILE
        print()
        print("=" * 60)
        print()

    try:
        print(f"Input:  {input_file}")
        print(f"Output: {output_file}")
        print()

        processor = HOSubcontractProcessor(input_file)
        processor.process(output_file)

        print()
        print("✓ Processing completed successfully!")
        print(f"✓ Output saved to: {output_file}")
        print()
    except Exception as e:
        print()
        print(f"✗ ERROR: {str(e)}")
        print()
        input("Press ENTER to exit...")
        sys.exit(1)

    if len(sys.argv) < 2:
        input("\nPress ENTER to exit...")


if __name__ == "__main__":
    main()
