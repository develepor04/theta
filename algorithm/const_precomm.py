
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys
from pathlib import Path


# ============================================================================
# DEFAULT FILE PATHS
# ============================================================================
DEFAULT_INPUT_FILE = r"2026.01.30_Borouge EU3 H2 Extraction Project PMS-Rev1 dates (1).xlsx"
DEFAULT_OUTPUT_FILE = r"01-03-26\const_precomm_tracker.xlsx"
# ============================================================================


class ConstPreCommProcessor:
    """Processes Const & Pre-Comm sheet and generates Activity Timeline Tracker."""

    # Column mappings from source sheet
    COL_CATEGORY = 1          # Col A - Category (Accommodation, Civil, etc.)
    COL_LEVEL = 2             # Col B - Level (L0-L5 or Activity ID)
    COL_CODE_NAME = 3         # Col C - WBS Code / Activity ID (also holds name for L4/L5)
    COL_ACTIVITY_NAME = 4     # Col D - WBS / Activity Name
    COL_STAGE_GATE = 13       # Col M - Stage Gate (EP/LP/F/A/Stage gate/% Percentage)

    # Milestone date columns (max 6 milestones in cols 14-19)
    MILESTONE_COLS = [14, 15, 16, 17, 18, 19]

    # Data start
    DATA_START_ROW = 6

    # Style definitions
    HEADER_COLOR = "366092"
    ON_TIME_COLOR = "C6EFCE"
    ON_TIME_FONT_COLOR = "006100"
    DELAYED_COLOR = "FFC7CE"
    DELAYED_FONT_COLOR = "9C0006"
    NOT_STARTED_COLOR = "FFF2CC"
    NOT_STARTED_FONT_COLOR = "7F6000"

    def __init__(self, input_file):
        self.input_file = Path(input_file)
        self.workbook = None
        self.sheet = None
        self.data_records = []

    def validate_input_file(self):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if self.input_file.suffix.lower() not in ['.xlsx', '.xlsm']:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}")
        print(f"✓ Input file validated: {self.input_file.name}")

    def load_workbook(self):
        try:
            print("Loading workbook...")
            self.workbook = openpyxl.load_workbook(self.input_file, data_only=True)

            sheet_name = 'Const & Pre-Comm'
            if sheet_name not in self.workbook.sheetnames:
                for name in self.workbook.sheetnames:
                    if 'const' in name.lower() and 'pre' in name.lower() and 'curve' not in name.lower():
                        sheet_name = name
                        break
                else:
                    raise ValueError(
                        f"'Const & Pre-Comm' sheet not found. "
                        f"Available: {self.workbook.sheetnames}"
                    )

            self.sheet = self.workbook[sheet_name]
            print(f"✓ '{sheet_name}' sheet loaded: {self.sheet.max_row} rows × {self.sheet.max_column} columns")
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {str(e)}")

    def _get_cell_value(self, row, col):
        try:
            return self.sheet.cell(row=row, column=col).value
        except:
            return None

    def _is_valid_milestone_name(self, val):
        """Check if a value is a valid milestone name (not '-' or empty)."""
        if val is None:
            return False
        s = str(val).strip()
        return s != '' and s != '-' and s != '0'

    def extract_data(self):
        print("\nExtracting activity data...")

        total_rows = self.sheet.max_row
        processed_count = 0
        group_count = 0

        # Current milestone names (updated from 'Stage gate' sub-header rows)
        current_milestones = []

        # Current EP group data
        current_group = None

        row_idx = self.DATA_START_ROW
        while row_idx <= total_rows:
            try:
                stage_gate = self._get_cell_value(row_idx, self.COL_STAGE_GATE)
                level = self._get_cell_value(row_idx, self.COL_LEVEL)
                code_name = self._get_cell_value(row_idx, self.COL_CODE_NAME)
                activity_name = self._get_cell_value(row_idx, self.COL_ACTIVITY_NAME)

                if not stage_gate:
                    row_idx += 1
                    continue

                sg_str = str(stage_gate).strip()

                # ── "Stage gate" sub-header row: update milestone names ──
                if sg_str == 'Stage gate':
                    new_milestones = []
                    for mc in self.MILESTONE_COLS:
                        val = self._get_cell_value(row_idx, mc)
                        if self._is_valid_milestone_name(val):
                            new_milestones.append((str(val).strip(), mc))
                    if new_milestones:
                        current_milestones = new_milestones
                    row_idx += 1
                    continue

                # ── Skip % Percentage / Equipment / Bulk rows ──
                if sg_str in ('% Percentage', 'Equipment', 'Bulk'):
                    row_idx += 1
                    continue

                # ── EP row: start new activity group ──
                if sg_str == 'EP':
                    # Determine activity name
                    level_str = str(level or '').strip().lower()

                    if level_str == 'l3':
                        # L3: name in col D (activity_name), WBS code in col C
                        act_name = str(activity_name or code_name or '').strip()
                    elif level_str in ('l4', 'l5'):
                        # L4/L5: name is in col C (code_name), col D is empty
                        act_name = str(code_name or '').strip()
                    elif activity_name:
                        act_name = str(activity_name).strip()
                    else:
                        act_name = str(code_name or '').strip()

                    # Get EP dates
                    ep_dates = {}
                    for ms_name, ms_col in current_milestones:
                        val = self._get_cell_value(row_idx, ms_col)
                        if val is not None and str(val).strip() != '-':
                            ep_dates[ms_name] = val
                        else:
                            ep_dates[ms_name] = None

                    # Look ahead for LP, F, A rows
                    lp_dates = {}
                    f_dates = {}
                    a_dates = {}
                    activity_id = str(code_name or '').strip()
                    scan_row = row_idx + 1

                    while scan_row <= min(row_idx + 10, total_rows):
                        scan_sg = self._get_cell_value(scan_row, self.COL_STAGE_GATE)
                        if not scan_sg:
                            scan_row += 1
                            continue

                        scan_sg_str = str(scan_sg).strip()

                        if scan_sg_str == 'LP':
                            lp_code = self._get_cell_value(scan_row, self.COL_CODE_NAME)
                            if lp_code and str(lp_code).strip():
                                lp_code_str = str(lp_code).strip()
                                # Use LP activity code as Activity ID if it starts with 'A' or similar
                                if lp_code_str[0].isalpha() and lp_code_str[0].upper() == 'A' and any(c.isdigit() for c in lp_code_str):
                                    activity_id = lp_code_str

                            for ms_name, ms_col in current_milestones:
                                val = self._get_cell_value(scan_row, ms_col)
                                if val is not None and str(val).strip() != '-':
                                    lp_dates[ms_name] = val
                                else:
                                    lp_dates[ms_name] = None

                        elif scan_sg_str == 'F':
                            for ms_name, ms_col in current_milestones:
                                val = self._get_cell_value(scan_row, ms_col)
                                if val is not None and str(val).strip() != '-':
                                    f_dates[ms_name] = val
                                else:
                                    f_dates[ms_name] = None

                        elif scan_sg_str == 'A':
                            for ms_name, ms_col in current_milestones:
                                val = self._get_cell_value(scan_row, ms_col)
                                if val is not None and str(val).strip() != '-':
                                    a_dates[ms_name] = val
                                else:
                                    a_dates[ms_name] = None

                        elif scan_sg_str == 'EP' or scan_sg_str == 'Stage gate':
                            break  # Next group
                        elif scan_sg_str in ('% Percentage',):
                            break  # Next section

                        scan_row += 1

                    # Create records for each milestone
                    group_count += 1
                    for ms_name, ms_col in current_milestones:
                        ep_val = ep_dates.get(ms_name)
                        lp_val = lp_dates.get(ms_name)
                        f_val = f_dates.get(ms_name)
                        a_val = a_dates.get(ms_name)

                        has_any_date = any(
                            d is not None and isinstance(d, datetime)
                            for d in [ep_val, lp_val, f_val, a_val]
                        )

                        if has_any_date:
                            duration_deviation, timeline_flag = self._calculate_deviation(
                                ep_val, lp_val, a_val
                            )

                            record = {
                                'activity_id': activity_id,
                                'activity_name': act_name,
                                'stage_gate': ms_name,
                                'planned_start_date': ep_val if isinstance(ep_val, datetime) else None,
                                'planned_end_date': lp_val if isinstance(lp_val, datetime) else None,
                                'actual_date': a_val if isinstance(a_val, datetime) else None,
                                'duration_deviation': duration_deviation,
                                'timeline_flag': timeline_flag
                            }
                            self.data_records.append(record)

                    processed_count += 1

            except Exception as e:
                pass

            row_idx += 1

        print(f"✓ Data extraction complete: {processed_count} EP groups processed")
        print(f"  Total milestone records: {len(self.data_records)}")

    def _calculate_deviation(self, planned_start_ep, planned_end_lp, actual_date_a):
        """Calculate deviation between Planned End (LP) and Actual Date (A)."""
        duration_deviation = None
        today = datetime.now()

        ep_valid = isinstance(planned_start_ep, datetime)
        lp_valid = isinstance(planned_end_lp, datetime)
        a_valid = isinstance(actual_date_a, datetime)

        if not a_valid:
            # No actual date - check if LP is in the past (overdue)
            if lp_valid and planned_end_lp < today:
                duration_deviation = (today - planned_end_lp).days
                timeline_flag = "Delayed"
            elif ep_valid and planned_start_ep < today:
                duration_deviation = (today - planned_start_ep).days
                timeline_flag = "Delayed"
            else:
                timeline_flag = "Not Started"
        elif lp_valid:
            duration_deviation = (actual_date_a - planned_end_lp).days
            timeline_flag = "Delayed" if actual_date_a > planned_end_lp else "On Time"
        elif ep_valid:
            duration_deviation = (actual_date_a - planned_start_ep).days
            timeline_flag = "Delayed" if actual_date_a > planned_start_ep else "On Time"
        else:
            timeline_flag = "Not Started"

        # Manual Error: LP date is before EP date
        if (ep_valid and lp_valid and planned_end_lp < planned_start_ep):
            timeline_flag = "Manual Error"

        return duration_deviation, timeline_flag

    def _format_date(self, date_value):
        if isinstance(date_value, datetime):
            return date_value.strftime('%d-%b-%Y')
        return None

    def generate_output(self, output_file):
        print(f"\nGenerating output file: {output_file}")

        out_wb = openpyxl.Workbook()
        ws = out_wb.active
        ws.title = "Const & Pre-Comm Tracker"

        # ── Define Styles ──
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color=self.HEADER_COLOR,
                                  end_color=self.HEADER_COLOR, fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        on_time_fill = PatternFill(start_color=self.ON_TIME_COLOR,
                                   end_color=self.ON_TIME_COLOR, fill_type='solid')
        on_time_font = Font(name='Calibri', color=self.ON_TIME_FONT_COLOR, size=11)

        delayed_fill = PatternFill(start_color=self.DELAYED_COLOR,
                                   end_color=self.DELAYED_COLOR, fill_type='solid')
        delayed_font = Font(name='Calibri', color=self.DELAYED_FONT_COLOR, size=11)

        not_started_fill = PatternFill(start_color=self.NOT_STARTED_COLOR,
                                       end_color=self.NOT_STARTED_COLOR, fill_type='solid')
        not_started_font = Font(name='Calibri', color=self.NOT_STARTED_FONT_COLOR, size=11)

        manual_error_fill = PatternFill(start_color="FF9999",
                                        end_color="FF9999", fill_type='solid')
        manual_error_font = Font(name='Calibri', color="800000", bold=True, size=11)

        data_font = Font(name='Calibri', size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ── Write Headers ──
        headers = [
            'Activity ID',
            'Activity Name',
            'Stage Gate',
            'Early Planning',
            'Late Planning',
            'Actual Date',
            'Duration Deviation',
            'Timeline Flag'
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # ── Write Data ──
        for row_idx, record in enumerate(self.data_records, 2):
            ws.cell(row=row_idx, column=1, value=record['activity_id']).font = data_font
            ws.cell(row=row_idx, column=1).alignment = center_align
            ws.cell(row=row_idx, column=1).border = thin_border

            ws.cell(row=row_idx, column=2, value=record['activity_name']).font = data_font
            ws.cell(row=row_idx, column=2).alignment = left_align
            ws.cell(row=row_idx, column=2).border = thin_border

            ws.cell(row=row_idx, column=3, value=record['stage_gate']).font = data_font
            ws.cell(row=row_idx, column=3).alignment = center_align
            ws.cell(row=row_idx, column=3).border = thin_border

            # Dates
            ws.cell(row=row_idx, column=4,
                    value=self._format_date(record['planned_start_date'])).font = data_font
            ws.cell(row=row_idx, column=4).alignment = center_align
            ws.cell(row=row_idx, column=4).border = thin_border

            ws.cell(row=row_idx, column=5,
                    value=self._format_date(record['planned_end_date'])).font = data_font
            ws.cell(row=row_idx, column=5).alignment = center_align
            ws.cell(row=row_idx, column=5).border = thin_border

            ws.cell(row=row_idx, column=6,
                    value=self._format_date(record['actual_date'])).font = data_font
            ws.cell(row=row_idx, column=6).alignment = center_align
            ws.cell(row=row_idx, column=6).border = thin_border

            # Duration Deviation
            ws.cell(row=row_idx, column=7, value=record['duration_deviation']).font = data_font
            ws.cell(row=row_idx, column=7).alignment = center_align
            ws.cell(row=row_idx, column=7).border = thin_border

            # Timeline Flag with color
            flag = record['timeline_flag']
            flag_cell = ws.cell(row=row_idx, column=8, value=flag)
            flag_cell.alignment = center_align
            flag_cell.border = thin_border

            if flag == 'On Time':
                flag_cell.fill = on_time_fill
                flag_cell.font = on_time_font
            elif flag == 'Delayed':
                flag_cell.fill = delayed_fill
                flag_cell.font = delayed_font
            elif flag == 'Not Started':
                flag_cell.fill = not_started_fill
                flag_cell.font = not_started_font
            elif flag == 'Manual Error':
                flag_cell.fill = manual_error_fill
                flag_cell.font = manual_error_font
            else:
                flag_cell.font = data_font

        # ── Column Widths ──
        col_widths = {
            1: 18,   # Activity ID
            2: 55,   # Activity Name
            3: 35,   # Stage Gate
            4: 22,   # Planned Start Date
            5: 22,   # Planned End Date
            6: 18,   # Actual Date
            7: 20,   # Duration Deviation
            8: 16,   # Timeline Flag
        }

        for col, width in col_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # Freeze top row
        ws.freeze_panes = 'A2'

        # Auto-filter
        ws.auto_filter.ref = f"A1:H{len(self.data_records) + 1}"

        try:
            out_wb.save(output_file)
            print(f"✓ Output file created successfully")
        except PermissionError:
            alt_file = output_file.replace('.xlsx', '_new.xlsx')
            out_wb.save(alt_file)
            print(f"⚠ Permission denied for {output_file}")
            print(f"✓ Saved to: {alt_file}")

    def print_summary(self):
        total = len(self.data_records)
        delayed = sum(1 for r in self.data_records if r['timeline_flag'] == 'Delayed')
        on_time = sum(1 for r in self.data_records if r['timeline_flag'] == 'On Time')
        not_started = sum(1 for r in self.data_records if r['timeline_flag'] == 'Not Started')
        manual_error = sum(1 for r in self.data_records if r['timeline_flag'] == 'Manual Error')

        print("\n" + "=" * 60)
        print("CONST & PRE-COMM - SUMMARY STATISTICS")
        print("=" * 60)
        print(f"Total Records:      {total}")
        print(f"Delayed:            {delayed}")
        print(f"On Time:            {on_time}")
        print(f"Not Started:        {not_started}")
        print(f"Manual Error:       {manual_error}")
        print("=" * 60)

    def process(self, output_file):
        print("=" * 60)
        print("CONST & PRE-COMM - TRACKER GENERATOR")
        print("=" * 60)
        print(f"\nInput:  {self.input_file}")
        print(f"Output: {output_file}\n")

        self.validate_input_file()
        self.load_workbook()
        self.extract_data()
        self.generate_output(output_file)
        self.print_summary()

        print(f"\n✓ Processing completed successfully!")
        print(f"✓ Output saved to: {output_file}")


def main():
    if len(sys.argv) >= 3:
        input_file = sys.argv[1]
        output_file = sys.argv[2]
    elif len(sys.argv) == 2:
        input_file = sys.argv[1]
        output_file = DEFAULT_OUTPUT_FILE
    else:
        input_file = DEFAULT_INPUT_FILE
        output_file = DEFAULT_OUTPUT_FILE

    processor = ConstPreCommProcessor(input_file)
    processor.process(output_file)


if __name__ == "__main__":
    main()
