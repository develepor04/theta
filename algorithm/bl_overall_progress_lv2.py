#!/usr/bin/env python3
"""
BL Overall Project Progress Lv2 Tracker Generator
=========================================================
Processes Borouge Excel files - 'BL Overall Project Progress Lv2' sheet
with discipline-wise progress tracking (Planned vs Actual).

Column Order (Output):
SN | Discipline | Planned % | Actual % | Variance % | Status

Author: Claude
Date: 2026-02-17
Version: 2.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys
from pathlib import Path


# ============================================================================
# DEFAULT FILE PATHS
# ============================================================================
DEFAULT_INPUT_FILE = r"2026.01.30_Borouge EU3 H2 Extraction Project PMS-Rev1 dates (1).xlsx"
DEFAULT_OUTPUT_FILE = r"01-03-26\bl_overall_progress_lv2_tracker.xlsx"
# ============================================================================


class BLProgressLv2Processor:
    """Processes BL Overall Project Progress Lv2 sheet and generates Tracker."""

    # Column mappings
    COL_SN = 3              # Col C - Serial Number
    COL_DISCIPLINE = 4      # Col D - Discipline Name
    COL_PLANNED = 19        # Col S - Planned %
    COL_ACTUAL = 20         # Col T - Actual %

    # Row positions (will be auto-detected)
    DATA_START_ROW = 9
    DATA_END_ROW = 100  # Extended range

    # Info row (will be auto-detected)
    INFO_ROW = 5
    COL_CUTOFF_DATE = 4     # Col D
    COL_PREV_DATE = 7       # Col G

    # Main discipline rows (L1 parents)
    L1_ROWS = [9, 19, 24, 29, 37]

    # Style definitions
    HEADER_COLOR = "366092"
    ON_TIME_COLOR = "C6EFCE"
    ON_TIME_FONT_COLOR = "006100"
    DELAYED_COLOR = "FFC7CE"
    DELAYED_FONT_COLOR = "9C0006"

    def __init__(self, input_file):
        self.input_file = Path(input_file)
        self.workbook = None
        self.sheet = None
        self.data_records = []
        self.cutoff_date = None
        self.prev_date = None

    def validate_input_file(self):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if not self.input_file.suffix.lower() in ['.xlsx', '.xlsm']:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}")
        print(f"✓ Input file validated: {self.input_file.name}")

    def load_workbook(self):
        try:
            print("Loading workbook...")
            self.workbook = openpyxl.load_workbook(self.input_file, data_only=True, read_only=True)

            sheet_name = 'BL Overall Project Progress Lv2'
            if sheet_name not in self.workbook.sheetnames:
                for name in self.workbook.sheetnames:
                    if 'overall' in name.lower() and 'lv2' in name.lower() and 'bl' in name.lower():
                        sheet_name = name
                        break
                else:
                    raise ValueError(
                        f"'BL Overall Project Progress Lv2' sheet not found. "
                        f"Available: {self.workbook.sheetnames}"
                    )

            self.sheet = self.workbook[sheet_name]
            print(f"✓ '{sheet_name}' sheet loaded: {self.sheet.max_row} rows × {self.sheet.max_column} columns")
            
            # Pre-load all data for fast access in read-only mode
            print("  → Caching sheet data for fast access...")
            self.sheet_data = [list(row) for row in self.sheet.iter_rows(values_only=True)]
            print(f"  → Cached {len(self.sheet_data)} rows")
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {str(e)}")

    def _get_cell_value(self, row, col):
        try:
            # Access cached data (0-indexed)
            if 0 < row <= len(self.sheet_data) and 0 < col <= len(self.sheet_data[0]):
                return self.sheet_data[row - 1][col - 1]
            return None
        except:
            return None

    def _fmt_pct(self, val):
        """Format value as percentage (multiply by 100 and round)."""
        if val is None:
            return None
        try:
            return round(float(val) * 100, 2)
        except (ValueError, TypeError):
            return None
    
    def _detect_structure(self):
        """Auto-detect header rows and data start."""
        print("  -> Auto-detecting sheet structure...")
        
        # Look for header row with "Discipline", "Weight"
        for row_idx in range(1, min(12, self.sheet.max_row + 1)):
            row_vals = [str(self.sheet.cell(row_idx, c).value or '').lower() 
                       for c in range(1, min(25, self.sheet.max_column + 1))]
            
            has_sn = any('sr' in v and 'no' in v for v in row_vals) or any('sn' in v for v in row_vals)
            has_discipline = any('discipline' in v for v in row_vals)
            has_weight = any('weight' in v for v in row_vals)
            has_progress = any('progress' in v or 'cumulative' in v for v in row_vals)
            
            if has_discipline and (has_weight or has_progress):
                print(f"  -> Found header area around row {row_idx}")
                
                # Find columns
                for idx, val in enumerate(row_vals, 1):
                    if ('sr' in val and 'no' in val) or val.strip() in ['sn', 'no', 'no.']:
                        self.COL_SN = idx
                    elif 'discipline' in val and 'weight' not in val:
                        self.COL_DISCIPLINE = idx
                
                # Look for data start (after header)
                for data_row in range(row_idx + 1, min(row_idx + 8, self.sheet.max_row + 1)):
                    first_val = self.sheet.cell(data_row, self.COL_SN).value
                    if first_val and ('OVERALL' in str(first_val).upper() or str(first_val).strip().replace('.', '').isdigit()):
                        self.DATA_START_ROW = data_row
                        print(f"  -> Data starts at row {self.DATA_START_ROW}")
                        break
                
                print(f"  -> Columns: SN={self.COL_SN}, Discipline={self.COL_DISCIPLINE}")
                break

    def extract_data(self):
        print("\nExtracting data from BL Overall Project Progress Lv2 sheet...")
        
        # Auto-detect structure
        self._detect_structure()

        # Get info data (search for cut-off date)
        for row in range(1, 10):
            for col in range(1, 12):
                val = self._get_cell_value(row, col)
                if val and 'cut-off' in str(val).lower():
                    self.cutoff_date = self._get_cell_value(row, col + 1)
                    # Look for previous date nearby
                    for offset in range(1, 8):
                        label = self._get_cell_value(row, col + offset)
                        if label and 'previous' in str(label).lower():
                            self.prev_date = self._get_cell_value(row, col + offset + 1)
                            break
                    break

        if self.cutoff_date:
            print(f"  Cut-Off Date:     {self.cutoff_date}")
        if self.prev_date:
            print(f"  Previous Date:    {self.prev_date}")

        processed_count = 0

        # Process all data rows (stop at empty or totals)
        for row_idx in range(self.DATA_START_ROW, min(self.DATA_END_ROW + 1, self.sheet.max_row + 1)):
            sn = self._get_cell_value(row_idx, self.COL_SN)
            discipline = self._get_cell_value(row_idx, self.COL_DISCIPLINE)
            planned_raw = self._get_cell_value(row_idx, self.COL_PLANNED)
            actual_raw = self._get_cell_value(row_idx, self.COL_ACTUAL)

            if not discipline and not sn:
                continue
            
            # Stop at total/summary rows
            discipline_str = str(discipline or sn or '').strip()
            if discipline_str.upper() in ['TOTAL', 'GRAND TOTAL', 'SUMMARY']:
                break
            sn_str = str(sn).strip() if sn else ''
            is_l1 = (row_idx in self.L1_ROWS)

            planned_pct = self._fmt_pct(planned_raw)
            actual_pct = self._fmt_pct(actual_raw)

            # Calculate variance
            if planned_pct is not None and actual_pct is not None:
                try:
                    variance = round(float(actual_pct) - float(planned_pct), 2)
                except (ValueError, TypeError):
                    variance = None
            else:
                variance = None
            
            # Only add record if we have at least one value
            if planned_pct is None and actual_pct is None:
                continue

            # Determine status
            if planned_pct is None and actual_pct is None:
                status = "Not Started"
            elif variance is not None:
                if variance >= 0:
                    status = "On Time"
                else:
                    status = "Delayed"
            elif planned_pct == 0 and actual_pct == 0:
                status = "Not Started"
            else:
                status = "-"

            record = {
                'sn': sn_str,
                'discipline': discipline_str,
                'planned_ep': planned_pct,
                'actual_a': actual_pct,
                'variance': variance,
                'status': status,
                'is_l1': is_l1
            }

            self.data_records.append(record)
            processed_count += 1

        print(f"✓ Data extraction complete: {processed_count} records created")

    def generate_output(self, output_file):
        print(f"\nGenerating output file: {output_file}")

        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "BL Progress Lv2 Tracker"

        # Styles
        header_fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        on_time_fill = PatternFill(start_color=self.ON_TIME_COLOR, end_color=self.ON_TIME_COLOR, fill_type="solid")
        delayed_fill = PatternFill(start_color=self.DELAYED_COLOR, end_color=self.DELAYED_COLOR, fill_type="solid")
        not_started_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        l1_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Row 1: Period info
        cutoff_str = self.cutoff_date.strftime('%d-%b-%Y') if isinstance(self.cutoff_date, datetime) else str(self.cutoff_date or '')
        prev_str = self.prev_date.strftime('%d-%b-%Y') if isinstance(self.prev_date, datetime) else str(self.prev_date or '')

        info_text = f"Cut-Off: {cutoff_str}  |  Previous: {prev_str}"
        info_cell = new_ws.cell(row=1, column=1, value=info_text)
        info_cell.font = Font(bold=True, size=12, color=self.HEADER_COLOR)
        new_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        header_start_row = 3

        # Headers
        headers = [
            'SN',
            'Discipline',
            'Planned %',
            'Actual %',
            'Variance %',
            'Status'
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = new_ws.cell(row=header_start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Write data
        delayed_count = 0
        on_time_count = 0
        not_started_count = 0

        for row_offset, record in enumerate(self.data_records):
            row_idx = header_start_row + 1 + row_offset
            is_l1 = record['is_l1']

            # SN
            cell = new_ws.cell(row=row_idx, column=1, value=record['sn'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_l1:
                cell.fill = l1_fill
                cell.font = Font(bold=True)

            # Discipline
            cell = new_ws.cell(row=row_idx, column=2, value=record['discipline'])
            cell.border = border
            if is_l1:
                cell.fill = l1_fill
                cell.font = Font(bold=True)

            # Planned %
            val = record['planned_ep']
            display = f"{val:.2f}%" if val is not None else ''
            cell = new_ws.cell(row=row_idx, column=3, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_l1:
                cell.fill = l1_fill
                cell.font = Font(bold=True)

            # Actual %
            val = record['actual_a']
            display = f"{val:.2f}%" if val is not None else ''
            cell = new_ws.cell(row=row_idx, column=4, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_l1:
                cell.fill = l1_fill
                cell.font = Font(bold=True)

            # Variance %
            val = record['variance']
            display = f"{val:.2f}%" if val is not None else ''
            cell = new_ws.cell(row=row_idx, column=5, value=display)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if is_l1:
                cell.fill = l1_fill
                cell.font = Font(bold=True)
            elif val is not None:
                try:
                    if float(val) < 0:
                        cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                    elif float(val) > 0:
                        cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                except (ValueError, TypeError):
                    pass

            # Status
            cell = new_ws.cell(row=row_idx, column=6, value=record['status'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            if is_l1:
                cell.fill = l1_fill
                cell.font = Font(bold=True)
            elif record['status'] == 'Delayed':
                cell.fill = delayed_fill
                cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                delayed_count += 1
            elif record['status'] == 'On Time':
                cell.fill = on_time_fill
                cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                on_time_count += 1
            elif record['status'] == 'Not Started':
                cell.fill = not_started_fill
                cell.font = Font(color="7F6000", bold=True)
                not_started_count += 1

        # Column widths
        column_widths = {
            'A': 8, 'B': 45, 'C': 14, 'D': 14, 'E': 14, 'F': 15
        }
        for col, width in column_widths.items():
            new_ws.column_dimensions[col].width = width

        new_ws.freeze_panes = f'A{header_start_row + 1}'

        try:
            new_wb.save(output_file)
            print(f"✓ Output file created successfully")
        except Exception as e:
            raise RuntimeError(f"Failed to save: {str(e)}")
        finally:
            new_wb.close()

        print(f"\n{'='*60}")
        print("BL OVERALL PROJECT PROGRESS LV2 - SUMMARY STATISTICS")
        print(f"{'='*60}")
        print(f"Total Records:      {len(self.data_records):,}")
        print(f"Delayed:            {delayed_count:,}")
        print(f"On Time:            {on_time_count:,}")
        print(f"Not Started:        {not_started_count:,}")
        print(f"{'='*60}")

    def close(self):
        if self.workbook:
            self.workbook.close()

    def process(self, output_file):
        try:
            self.validate_input_file()
            self.load_workbook()
            self.extract_data()
            self.generate_output(output_file)
        finally:
            self.close()


def main():
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT_FILE
        print("=" * 60)
        print("BL OVERALL PROJECT PROGRESS LV2 - TRACKER GENERATOR")
        print("=" * 60)
        print()
    else:
        print("=" * 60)
        print("BL OVERALL PROJECT PROGRESS LV2 - TRACKER GENERATOR")
        print("=" * 60)
        print()
        print("DEFAULT INPUT FILE:")
        print(f"  {DEFAULT_INPUT_FILE}")
        print()
        user_input = input("Press ENTER to use default, or type new path: ").strip()
        input_file = user_input.strip('"').strip("'") if user_input else DEFAULT_INPUT_FILE

        print()
        print("DEFAULT OUTPUT FILE:")
        print(f"  {DEFAULT_OUTPUT_FILE}")
        print()
        user_output = input("Press ENTER to use default, or type new name: ").strip()
        output_file = user_output.strip('"').strip("'") if user_output else DEFAULT_OUTPUT_FILE
        print()
        print("=" * 60)
        print()

    try:
        print(f"Input:  {input_file}")
        print(f"Output: {output_file}")
        print()

        processor = BLProgressLv2Processor(input_file)
        processor.process(output_file)

        print()
        print("✓ Processing completed successfully!")
        print(f"✓ Output saved to: {output_file}")
        print()
    except Exception as e:
        print()
        print(f"✗ ERROR: {str(e)}")
        print()
        input("Press ENTER to exit...")
        sys.exit(1)

    if len(sys.argv) < 2:
        input("\nPress ENTER to exit...")


if __name__ == "__main__":
    main()
