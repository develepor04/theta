#!/usr/bin/env python3
"""
EDDR CNTR Tracker Generator
=========================================================
Processes Borouge Excel files - 'EDDR CNTR' sheet
with discipline-wise IFR/IFA/IFC tracking.

Output Columns:
SN | Discipline | Stage Gate | Total Documents | BL | BL-R1 | Actual | Variance | Status

Author: Claude
Date: 2026-02-17
Version: 1.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys
from pathlib import Path


# ============================================================================
# DEFAULT FILE PATHS
# ============================================================================
DEFAULT_INPUT_FILE = r"2026.01.30_Borouge EU3 H2 Extraction Project PMS-Rev1 dates (1).xlsx"
DEFAULT_OUTPUT_FILE = r"01-03-26\eddr_cntr_tracker.xlsx"
# ============================================================================


class EDDRCNTRProcessor:
    """Processes EDDR CNTR sheet and generates Tracker."""

    # Header rows
    INFO_ROW = 2
    INFO_DATE_ROW = 3
    COL_FROM = 4  # Col D
    COL_TO = 5    # Col E

    # Data rows
    DATA_START_ROW = 7
    DATA_END_ROW = 16

    # Column mappings
    COL_SN = 1
    COL_DISCIPLINE = 2
    COL_TOTAL_DOCS = 3

    # Stage gate columns (BL, BL-R1, Actual)
    STAGE_COLUMNS = {
        'IFR': (5, 6, 7),
        'IFA': (8, 9, 10),
        'IFC': (11, 12, 13)
    }

    # Style definitions
    HEADER_COLOR = "366092"
    ON_TIME_COLOR = "C6EFCE"
    ON_TIME_FONT_COLOR = "006100"
    DELAYED_COLOR = "FFC7CE"
    DELAYED_FONT_COLOR = "9C0006"
    NOT_STARTED_COLOR = "FFF2CC"
    NOT_STARTED_FONT_COLOR = "7F6000"

    def __init__(self, input_file):
        self.input_file = Path(input_file)
        self.workbook = None
        self.sheet = None
        self.data_records = []
        self.from_date = None
        self.to_date = None

    def validate_input_file(self):
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if self.input_file.suffix.lower() not in ['.xlsx', '.xlsm']:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}")
        print(f"✓ Input file validated: {self.input_file.name}")

    def load_workbook(self):
        try:
            print("Loading workbook...")
            self.workbook = openpyxl.load_workbook(self.input_file, data_only=True, read_only=True)

            sheet_name = 'EDDR CNTR'
            if sheet_name not in self.workbook.sheetnames:
                for name in self.workbook.sheetnames:
                    if 'eddr' in name.lower() and 'cntr' in name.lower():
                        sheet_name = name
                        break
                else:
                    raise ValueError(
                        f"'EDDR CNTR' sheet not found. "
                        f"Available: {self.workbook.sheetnames}"
                    )

            self.sheet = self.workbook[sheet_name]
            print(f"✓ '{sheet_name}' sheet loaded: {self.sheet.max_row} rows × {self.sheet.max_column} columns")
            
            # Pre-load all data for fast access in read-only mode
            print("  → Caching sheet data for fast access...")
            self.sheet_data = [list(row) for row in self.sheet.iter_rows(values_only=True)]
            print(f"  → Cached {len(self.sheet_data)} rows")
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {str(e)}")

    def _get_cell_value(self, row, col):
        try:
            # Access cached data (0-indexed)
            if 0 < row <= len(self.sheet_data) and 0 < col <= len(self.sheet_data[0]):
                return self.sheet_data[row - 1][col - 1]
            return None
        except:
            return None

    def extract_data(self):
        print("\nExtracting data from EDDR CNTR sheet...")
        
        # Auto-detect data start row
        self._detect_structure()

        # Info dates (look in first few rows)
        for row in range(1, 6):
            for col in range(1, 10):
                val = self._get_cell_value(row, col)
                if val and 'from' in str(val).lower():
                    self.from_date = self._get_cell_value(row + 1, col)
                    self.to_date = self._get_cell_value(row + 1, col + 1)
                    break

        if self.from_date:
            print(f"  From Date: {self.from_date}")
        if self.to_date:
            print(f"  To Date:   {self.to_date}")

        processed_count = 0
        
        # Process all rows from data start to end of sheet
        for row_idx in range(self.DATA_START_ROW, self.sheet.max_row + 1):
            sn = self._get_cell_value(row_idx, self.COL_SN)
            discipline = self._get_cell_value(row_idx, self.COL_DISCIPLINE)
            total_docs = self._get_cell_value(row_idx, self.COL_TOTAL_DOCS)

            if not discipline and not sn:
                continue
            
            # Stop at total/summary rows
            if discipline and str(discipline).strip().upper() in ['TOTAL', 'GRAND TOTAL', 'SUMMARY']:
                break

            discipline_str = str(discipline or '').strip()
            sn_str = str(sn).strip() if sn is not None else ''

            for stage_gate, (col_bl, col_blr1, col_actual) in self.STAGE_COLUMNS.items():
                bl = self._get_cell_value(row_idx, col_bl)
                blr1 = self._get_cell_value(row_idx, col_blr1)
                actual = self._get_cell_value(row_idx, col_actual)

                # Calculate variance (Actual - BL-R1)
                if blr1 is not None and actual is not None:
                    try:
                        variance = int(actual) - int(blr1)
                    except (ValueError, TypeError):
                        variance = None
                else:
                    variance = None

                # Determine status
                if blr1 is None and actual is None:
                    status = "Not Started"
                elif blr1 == 0 and (actual is None or actual == 0):
                    status = "Not Started"
                elif variance is not None:
                    status = "On Time" if variance >= 0 else "Delayed"
                else:
                    status = "-"

                record = {
                    'sn': sn_str,
                    'discipline': discipline_str,
                    'stage_gate': stage_gate,
                    'total_docs': total_docs,
                    'bl': bl,
                    'blr1': blr1,
                    'actual': actual,
                    'variance': variance,
                    'status': status
                }

                self.data_records.append(record)
                processed_count += 1

        print(f"✓ Data extraction complete: {processed_count} records created")    
    def _detect_structure(self):
        """Auto-detect header row and data start."""
        print("  → Auto-detecting sheet structure...")
        
        # Look for header row containing "No", "Discipline", "Total"
        for row_idx in range(1, min(10, self.sheet.max_row + 1)):
            row_vals = [str(self.sheet.cell(row_idx, c).value or '').lower() 
                       for c in range(1, min(20, self.sheet.max_column + 1))]
            
            has_sn = any('no' == v.strip() or 'sn' in v for v in row_vals)
            has_discipline = any('discpline' in v or 'discipline' in v for v in row_vals)
            has_total = any('total' in v and 'document' in v for v in row_vals)
            
            if has_discipline and (has_sn or has_total):
                print(f"  → Found header row at row {row_idx}")
                
                # Find column positions
                for idx, val in enumerate(row_vals, 1):
                    val_clean = val.strip()
                    if val_clean in ['no', 'sn']:
                        self.COL_SN = idx
                    elif 'discpline' in val or 'discipline' in val:
                        self.COL_DISCIPLINE = idx
                    elif 'total' in val and 'document' in val:
                        self.COL_TOTAL_DOCS = idx
                
                # Check next row for sub-headers (BL, BL-R1, Actual)
                next_row = row_idx + 1
                next_vals = [str(self.sheet.cell(next_row, c).value or '').upper() 
                            for c in range(1, min(20, self.sheet.max_column + 1))]
                
                # Find IFR, IFA columns
                ifr_cols = []
                ifa_cols = []
                ifc_cols = []
                
                for idx, val in enumerate(next_vals, 1):
                    if 'BL' in val:
                        # Determine which stage gate this belongs to
                        # Look for IFR/IFA/IFC context (check nearby cells)
                        header_above = str(self.sheet.cell(row_idx, idx).value or '').upper()
                        if 'IFR' in header_above:
                            ifr_cols.append(idx)
                        elif 'IFA' in header_above:
                            ifa_cols.append(idx)
                        elif 'IFC' in header_above or not header_above:
                            ifc_cols.append(idx)
                    elif 'ACTUAL' in val:
                        header_above = str(self.sheet.cell(row_idx, idx).value or '').upper()
                        if 'IFR' in header_above:
                            ifr_cols.append(idx)
                        elif 'IFA' in header_above:
                            ifa_cols.append(idx)
                        elif 'IFC' in header_above or not header_above:
                            ifc_cols.append(idx)
                
                # Update STAGE_COLUMNS if we found the columns
                if len(ifr_cols) >= 3:
                    self.STAGE_COLUMNS['IFR'] = tuple(ifr_cols[:3])
                if len(ifa_cols) >= 3:
                    self.STAGE_COLUMNS['IFA'] = tuple(ifa_cols[:3])
                if len(ifc_cols) >= 3:
                    self.STAGE_COLUMNS['IFC'] = tuple(ifc_cols[:3])
                
                self.DATA_START_ROW = next_row + 1
                print(f"  → Data starts at row {self.DATA_START_ROW}")
                print(f"  → Columns: SN={self.COL_SN}, Discipline={self.COL_DISCIPLINE}, Total={self.COL_TOTAL_DOCS}")
                break

    def generate_output(self, output_file):
        print(f"\nGenerating output file: {output_file}")

        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "EDDR CNTR Tracker"

        # Styles
        header_fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        on_time_fill = PatternFill(start_color=self.ON_TIME_COLOR, end_color=self.ON_TIME_COLOR, fill_type="solid")
        delayed_fill = PatternFill(start_color=self.DELAYED_COLOR, end_color=self.DELAYED_COLOR, fill_type="solid")
        not_started_fill = PatternFill(start_color=self.NOT_STARTED_COLOR, end_color=self.NOT_STARTED_COLOR, fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Row 1: Period info
        from_str = self.from_date.strftime('%d-%b-%Y') if isinstance(self.from_date, datetime) else str(self.from_date or '')
        to_str = self.to_date.strftime('%d-%b-%Y') if isinstance(self.to_date, datetime) else str(self.to_date or '')

        info_text = f"From: {from_str}  |  To: {to_str}"
        info_cell = new_ws.cell(row=1, column=1, value=info_text)
        info_cell.font = Font(bold=True, size=12, color=self.HEADER_COLOR)
        new_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

        header_start_row = 3

        headers = [
            'SN',
            'Discipline',
            'Stage Gate',
            'Total Documents',
            'BL',
            'BL-R1',
            'Actual',
            'Variance',
            'Status'
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = new_ws.cell(row=header_start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Write data
        delayed_count = 0
        on_time_count = 0
        not_started_count = 0

        for row_offset, record in enumerate(self.data_records):
            row_idx = header_start_row + 1 + row_offset

            # SN
            cell = new_ws.cell(row=row_idx, column=1, value=record['sn'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Discipline
            cell = new_ws.cell(row=row_idx, column=2, value=record['discipline'])
            cell.border = border

            # Stage Gate
            cell = new_ws.cell(row=row_idx, column=3, value=record['stage_gate'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Total Documents
            cell = new_ws.cell(row=row_idx, column=4, value=record['total_docs'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # BL
            cell = new_ws.cell(row=row_idx, column=5, value=record['bl'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # BL-R1
            cell = new_ws.cell(row=row_idx, column=6, value=record['blr1'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Actual
            cell = new_ws.cell(row=row_idx, column=7, value=record['actual'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Variance
            cell = new_ws.cell(row=row_idx, column=8, value=record['variance'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if record['variance'] is not None:
                try:
                    if float(record['variance']) < 0:
                        cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                    elif float(record['variance']) > 0:
                        cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                except (ValueError, TypeError):
                    pass

            # Status
            cell = new_ws.cell(row=row_idx, column=9, value=record['status'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            if record['status'] == 'Delayed':
                cell.fill = delayed_fill
                cell.font = Font(color=self.DELAYED_FONT_COLOR, bold=True)
                delayed_count += 1
            elif record['status'] == 'On Time':
                cell.fill = on_time_fill
                cell.font = Font(color=self.ON_TIME_FONT_COLOR, bold=True)
                on_time_count += 1
            elif record['status'] == 'Not Started':
                cell.fill = not_started_fill
                cell.font = Font(color=self.NOT_STARTED_FONT_COLOR, bold=True)
                not_started_count += 1


        # Column widths
        column_widths = {
            'A': 6, 'B': 30, 'C': 12, 'D': 16,
            'E': 10, 'F': 10, 'G': 10, 'H': 10, 'I': 14
        }
        for col, width in column_widths.items():
            new_ws.column_dimensions[col].width = width

        new_ws.freeze_panes = f'A{header_start_row + 1}'

        try:
            new_wb.save(output_file)
            print("✓ Output file created successfully")
        except Exception as e:
            raise RuntimeError(f"Failed to save: {str(e)}")
        finally:
            new_wb.close()

        print(f"\n{'='*60}")
        print("EDDR CNTR - SUMMARY STATISTICS")
        print(f"{'='*60}")
        print(f"Total Records:      {len(self.data_records):,}")
        print(f"Delayed:            {delayed_count:,}")
        print(f"On Time:            {on_time_count:,}")
        print(f"Not Started:        {not_started_count:,}")
        print(f"{'='*60}")

    def close(self):
        if self.workbook:
            self.workbook.close()

    def process(self, output_file):
        try:
            self.validate_input_file()
            self.load_workbook()
            self.extract_data()
            self.generate_output(output_file)
        finally:
            self.close()


def main():
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT_FILE
        print("=" * 60)
        print("EDDR CNTR - TRACKER GENERATOR")
        print("=" * 60)
        print()
    else:
        print("=" * 60)
        print("EDDR CNTR - TRACKER GENERATOR")
        print("=" * 60)
        print()
        print("DEFAULT INPUT FILE:")
        print(f"  {DEFAULT_INPUT_FILE}")
        print()
        user_input = input("Press ENTER to use default, or type new path: ").strip()
        input_file = user_input.strip('"').strip("'") if user_input else DEFAULT_INPUT_FILE

        print()
        print("DEFAULT OUTPUT FILE:")
        print(f"  {DEFAULT_OUTPUT_FILE}")
        print()
        user_output = input("Press ENTER to use default, or type new name: ").strip()
        output_file = user_output.strip('"').strip("'") if user_output else DEFAULT_OUTPUT_FILE
        print()
        print("=" * 60)
        print()

    try:
        print(f"Input:  {input_file}")
        print(f"Output: {output_file}")
        print()

        processor = EDDRCNTRProcessor(input_file)
        processor.process(output_file)

        print()
        print("✓ Processing completed successfully!")
        print(f"✓ Output saved to: {output_file}")
        print()
    except Exception as e:
        print()
        print(f"✗ ERROR: {str(e)}")
        print()
        input("Press ENTER to exit...")
        sys.exit(1)

    if len(sys.argv) < 2:
        input("\nPress ENTER to exit...")


if __name__ == "__main__":
    main()
