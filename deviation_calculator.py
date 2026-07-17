"""
Deviation Calculator for PMO Application
=========================================
Automatically analyzes processed Excel files to detect and categorize deviations.

Deviation Types:
- Timeline: Project schedule delays and duration deviations  
- Quantity: Material/resource quantity variances
- Cost: Budget vs actual cost overruns
- Fuel: Fuel consumption anomalies

Author: PMO Team
Date: 2026-02-21
Optimized for low-memory environments (1GB RAM)
"""

import sys
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace', line_buffering=True)
    sys.stderr.reconfigure(encoding='utf-8', errors='replace', line_buffering=True)
except Exception:
    pass

import openpyxl
import pandas as pd
from datetime import datetime
import os
import json
from pathlib import Path
import gc

# Configure pandas for low-memory environments
pd.set_option('mode.chained_assignment', None)  # Reduce warnings overhead


class DeviationCalculator:
    """Analyzes processed output files and calculates deviations."""
    
    # Deviation thresholds
    THRESHOLDS = {
        'timeline': {
            'low': 1,      # 1 day delay
            'medium': 3,   # 3 days delay
            'high': 7      # 7+ days delay
        },
        'cost': {
            'low': 5,      # 5% over budget
            'medium': 10,  # 10% over budget
            'high': 20     # 20%+ over budget
        },
        'quantity': {
            'low': 5,      # 5% variance
            'medium': 10,  # 10% variance
            'high': 20     # 20%+ variance
        },
        'fuel': {
            'low': 10,     # 10% over planned
            'medium': 20,  # 20% over planned
            'high': 30     # 30%+ over planned
        }
    }
    
    def __init__(self, output_folder, job_id, filename):
        """
        Initialize deviation calculator.
        
        Args:
            output_folder: Path to job-specific output folder
            job_id: Unique job identifier
            filename: Original uploaded filename
        """
        self.output_folder = Path(output_folder)
        self.job_id = job_id
        self.filename = filename
        self.deviations = []
        
    def calculate_severity(self, value, category):
        """
        Determine severity level based on value and category.
        
        Args:
            value: Deviation value (numeric)
            category: Deviation category (timeline/cost/quantity/fuel)
            
        Returns:
            str: 'Low', 'Medium', or 'High'
        """
        if category not in self.THRESHOLDS:
            return 'Medium'
        
        thresholds = self.THRESHOLDS[category]
        abs_value = abs(value)
        
        if abs_value >= thresholds['high']:
            return 'High'
        elif abs_value >= thresholds['medium']:
            return 'Medium'
        elif abs_value >= thresholds['low']:
            return 'Low'
        else:
            return None  # Below threshold, not a deviation
    
    def analyze_timeline_deviations(self, excel_path, sheet_name=None):
        """Analyze timeline deviations — handles both main1 and newalgo output formats."""
        wb = None
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            actual_sheets = wb.sheetnames
            if not actual_sheets:
                return
            target_sheet = sheet_name if (sheet_name and sheet_name in actual_sheets) else actual_sheets[0]
            print(f"[DEVIATION] Reading '{target_sheet}' from {Path(excel_path).name}")
            wb.close(); wb = None

            df = pd.read_excel(excel_path, sheet_name=target_sheet, engine='openpyxl')
            if df.empty:
                print(f"[DEVIATION] '{target_sheet}' is empty")
                return

            df.columns = df.columns.str.strip()
            col_map = {col.lower(): col for col in df.columns}

            def _find(*variants):
                for v in variants:
                    if v in col_map:
                        return col_map[v]
                return None

            # Column detection — covers both main1 and newalgo
            activity_id_col   = _find('activity_id', 'activity id', 'activityid')
            activity_name_col = _find('activity_name', 'activity name', 'activityname')
            # newalgo "Status" / main1 "Timeline Flag"
            status_col        = _find('status', 'timeline_flag', 'timeline flag', 'flag')
            # newalgo only: separate duration-delay description
            tdev_flag_col     = _find('timeline_deviation_flag', 'timeline deviation flag')
            # newalgo: numeric start delay (positive = started late)
            start_delay_col   = _find('start_delay deviation', 'start delay deviation', 'start_delay')
            # newalgo + main1: numeric duration overrun
            duration_col      = _find('duration_deviation', 'duration deviation (days)', 'duration deviation')
            # newalgo: textual start delay flag
            sdflag_col        = _find('start_delay_flag', 'start delay flag')

            print(f"[DEVIATION] {target_sheet!r}: {len(df)} rows | "
                  f"status={status_col!r} tdev={tdev_flag_col!r} "
                  f"start_delay={start_delay_col!r} duration={duration_col!r}")

            if status_col:
                print(f"[DEVIATION] Status dist:  {df[status_col].value_counts().to_dict()}")
            if tdev_flag_col:
                print(f"[DEVIATION] TDevFlag dist: {df[tdev_flag_col].value_counts().to_dict()}")
            if sdflag_col:
                print(f"[DEVIATION] SDFlag dist:   {df[sdflag_col].value_counts().to_dict()}")

            has_any_col = any([status_col, tdev_flag_col, start_delay_col, duration_col])
            if not has_any_col:
                print("[DEVIATION] No recognisable deviation columns — skipping")
                return

            today = datetime.now().date()

            p_start_col = _find('planned start date', 'early planning', 'early planned start')
            p_end_col   = _find('planned end date', 'late planning', 'planned end date')
            a_start_col = _find('actual start date', 'actual start', 'actual date')
            a_end_col   = _find('actual completion date', 'actual finish', 'actual completion')

            def _to_date(val):
                """Parse a cell value to a date, or return None. Handles NaT safely."""
                try:
                    if val is None or pd.isnull(val):
                        return None
                except (TypeError, ValueError):
                    pass
                if isinstance(val, datetime):
                    try:
                        return val.date()
                    except Exception:
                        return None
                try:
                    ts = pd.to_datetime(val, errors='coerce')
                    if pd.isnull(ts):
                        return None
                    return ts.date()
                except Exception:
                    return None

            for _, row in df.iterrows():
                status_val  = str(row.get(status_col,    '') or '').strip() if status_col    else ''
                tdev_val    = str(row.get(tdev_flag_col, '') or '').strip() if tdev_flag_col else ''
                sdflag_val  = str(row.get(sdflag_col,   '') or '').strip() if sdflag_col    else ''

                try:
                    start_days = float(row.get(start_delay_col, 0) or 0) if start_delay_col else 0.0
                except (TypeError, ValueError):
                    start_days = 0.0
                try:
                    dur_days = float(row.get(duration_col, 0) or 0) if duration_col else 0.0
                except (TypeError, ValueError):
                    dur_days = 0.0

                # --- Detect overdue activities that newalgo leaves at 0 ---
                # newalgo sets start_delay=0 when a_start is None, and dur_dev=0
                # when a_finish is None — even if the planned date has already passed.
                # We recompute using today's date for those cases.
                tdev_lower = tdev_val.lower()

                if dur_days == 0 and 'in progress' in tdev_lower:
                    # Activity started but not finished — check if planned end has passed
                    p_end = _to_date(row.get(p_end_col) if p_end_col else None)
                    if p_end and p_end < today:
                        dur_days = (today - p_end).days  # days overdue

                if start_days == 0 and ('not started' in tdev_lower or 'not started' in status_val.lower()):
                    # Activity hasn't started — check if planned start has passed
                    p_start = _to_date(row.get(p_start_col) if p_start_col else None)
                    if p_start and p_start < today:
                        start_days = (today - p_start).days  # days overdue

                # A row is a deviation when ANY signal fires:
                #  1. Status/flag contains "delay"
                #  2. Timeline_deviation_flag contains "delay"
                #  3. start_delay_flag == "Delay Start"
                #  4. Numeric start_delay > 0 (recorded or computed vs today)
                #  5. Numeric duration_deviation > 0 (recorded or computed vs today)
                is_delayed = (
                    'delay' in status_val.lower()
                    or 'delay' in tdev_lower
                    or sdflag_val.lower() == 'delay start'
                    or start_days > 0
                    or dur_days > 0
                )
                if not is_delayed:
                    continue

                best_days = max(abs(start_days), abs(dur_days))
                if best_days == 0:
                    best_days = 3

                severity = self.calculate_severity(best_days, 'timeline')
                if not severity:
                    continue

                activity_id   = str(row.get(activity_id_col,   '') or '') if activity_id_col   else ''
                activity_name = str(row.get(activity_name_col, '') or 'Unknown Activity') if activity_name_col else 'Unknown Activity'

                parts = []
                if start_days > 0 and 'not started' in (status_val + tdev_val).lower():
                    parts.append(f"Not started — {int(start_days)}d overdue")
                elif start_days > 0:
                    parts.append(f"Start delay: {int(start_days)}d")
                if dur_days > 0 and 'in progress' in tdev_lower:
                    parts.append(f"In progress — {int(dur_days)}d past planned end")
                elif dur_days > 0:
                    parts.append(f"Duration overrun: {int(dur_days)}d")
                if tdev_val and 'delay' in tdev_lower:
                    parts.append(tdev_val)
                if not parts:
                    parts.append(status_val or 'Flagged as delayed')

                description = f"{activity_name} — {', '.join(parts)}"

                self.deviations.append({
                    'sheet':       target_sheet,
                    'flag':        status_val or tdev_val or 'Delayed',
                    'severity':    severity,
                    'description': description,
                    'row_data': {
                        'activity_id':        activity_id,
                        'activity_name':      activity_name,
                        'start_delay':        start_days,
                        'duration_deviation': dur_days,
                        'planned_start': str(row.get(p_start_col, '') or '') if p_start_col else '',
                        'planned_end':   str(row.get(p_end_col,   '') or '') if p_end_col   else '',
                        'actual_start':  str(row.get(a_start_col, '') or '') if a_start_col else '',
                        'actual_end':    str(row.get(a_end_col,   '') or '') if a_end_col   else '',
                        'status':        status_val,
                        'timeline_flag': tdev_val,
                    },
                    'detected_at': datetime.now().isoformat()
                })

            del df
            gc.collect()

        except Exception as e:
            import traceback; traceback.print_exc()
            print(f"[DEVIATION] Timeline analysis error: {e}")
        finally:
            if wb:
                wb.close()
    
    def analyze_cost_deviations(self, excel_path, sheet_name=None):
        """Analyze cost deviations with memory optimization."""
        wb = None
        try:
            # Use read_only mode
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            
            possible_names = [
                'Cost vs Budget',
                'Cost Deviation',
                'Budget',
                'Financial'
            ]
            
            if sheet_name:
                possible_names.insert(0, sheet_name)
            
            target_sheet = None
            for name in possible_names:
                if name in wb.sheetnames:
                    target_sheet = name
                    break
            
            wb.close()
            wb = None
            
            if not target_sheet:
                return
            
            df = pd.read_excel(
                excel_path, 
                sheet_name=target_sheet,
                engine='openpyxl'
            )
            
            if df.empty:
                return
            
            for idx, row in df.iterrows():
                # Skip summary rows or rows without flags
                if pd.isna(row.get('Deviation Flag')) or row.get('Deviation Flag') == '':
                    continue
                
                if row.get('month') == 'ALL':
                    continue
                
                budget = float(row.get('Budgeted Cost', 0) or 0)
                actual = float(row.get('Actual Cost', 0) or 0)
                deviation = float(row.get('Cost Deviation', 0) or 0)
                
                if budget == 0:
                    continue
                
                # Calculate percentage deviation
                pct_deviation = (deviation / budget) * 100
                
                severity = self.calculate_severity(pct_deviation, 'cost')
                if not severity:
                    continue
                
                description = f"{row.get('Scope', 'Unknown Scope')} ({row.get('month', '')}) - Budget: ${budget:,.2f}, Actual: ${actual:,.2f}, Variance: ${deviation:,.2f} ({pct_deviation:+.1f}%)"
                
                deviation_entry = {
                    'sheet': target_sheet,
                    'flag': str(row.get('Deviation Flag', 'Over Budget')),
                    'severity': severity,
                    'description': description,
                    'row_data': {
                        'month': str(row.get('month', '')),
                        'scope': str(row.get('Scope', '')),
                        'budgeted_cost': budget,
                        'actual_cost': actual,
                        'cost_deviation': deviation,
                        'percentage': round(pct_deviation, 2)
                    },
                    'detected_at': datetime.now().isoformat()
                }
                
                self.deviations.append(deviation_entry)
                
            # Clean up memory
            del df
            gc.collect()
                
        except Exception as e:
            print(f"[DEVIATION] Cost analysis error: {e}")
        finally:
            if wb:
                wb.close()
    
    def analyze_quantity_deviations(self, excel_path, sheet_name=None):
        """Analyze quantity deviations with memory optimization."""
        wb = None
        try:
            # Use read_only mode
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            
            possible_names = [
                'Quantity Deviation',
                'Quantity',
                'Material',
                'Resources'
            ]
            
            if sheet_name:
                possible_names.insert(0, sheet_name)
            
            target_sheet = None
            for name in possible_names:
                if name in wb.sheetnames:
                    target_sheet = name
                    break
            
            wb.close()
            wb = None
            
            if not target_sheet:
                return
            
            df = pd.read_excel(
                excel_path, 
                sheet_name=target_sheet,
                engine='openpyxl'
            )
            
            if df.empty:
                return
            
            # Look for common deviation indicator columns
            flag_columns = [col for col in df.columns if 'flag' in col.lower() or 'deviation' in col.lower()]
            
            for idx, row in df.iterrows():
                # Check if any flag column has a value
                has_flag = False
                flag_value = ''
                
                for flag_col in flag_columns:
                    if pd.notna(row.get(flag_col)) and row.get(flag_col) != '':
                        has_flag = True
                        flag_value = str(row.get(flag_col))
                        break
                
                if not has_flag:
                    continue
                
                # Try to extract quantity values
                planned = float(row.get('Planned Quantity', row.get('planned_qty', 0)) or 0)
                actual = float(row.get('Actual Quantity', row.get('actual_qty', 0)) or 0)
                
                if planned == 0:
                    continue
                
                variance = actual - planned
                pct_variance = (variance / planned) * 100
                
                severity = self.calculate_severity(abs(pct_variance), 'quantity')
                if not severity:
                    continue
                
                item_name = str(row.get('Item', row.get('Material', row.get('Description', 'Unknown Item'))))
                description = f"{item_name} - Planned: {planned:.2f}, Actual: {actual:.2f}, Variance: {variance:+.2f} ({pct_variance:+.1f}%)"
                
                deviation_entry = {
                    'sheet': target_sheet,
                    'flag': flag_value,
                    'severity': severity,
                    'description': description,
                    'row_data': {
                        'item': item_name,
                        'planned_quantity': planned,
                        'actual_quantity': actual,
                        'variance': variance,
                        'percentage': round(pct_variance, 2)
                    },
                    'detected_at': datetime.now().isoformat()
                }
                
                self.deviations.append(deviation_entry)
                
            # Clean up memory
            del df
            gc.collect()
                
        except Exception as e:
            print(f"[DEVIATION] Quantity analysis error: {e}")
        finally:
            if wb:
                wb.close()
    
    def analyze_fuel_deviations(self, excel_path, sheet_name=None):
        """Analyze fuel deviations with memory optimization."""
        wb = None
        try:
            # Use read_only mode
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            
            possible_names = [
                'Fuel by Type (Daily)',
                'Fuel Deviation',
                'Fuel',
                'Consumption'
            ]
            
            if sheet_name:
                possible_names.insert(0, sheet_name)
            
            target_sheet = None
            for name in possible_names:
                if name in wb.sheetnames:
                    target_sheet = name
                    break
            
            wb.close()
            wb = None
            
            if not target_sheet:
                return
            
            df = pd.read_excel(
                excel_path, 
                sheet_name=target_sheet,
                engine='openpyxl'
            )
            
            if df.empty:
                return
            
            # Get summary row (date = 'ALL', Type = 'ALL')
            summary = df[(df['date'] == 'ALL') & (df['Type'] == 'ALL')]
            
            if summary.empty:
                return
            
            total_fuel = float(summary.iloc[0]['Fuel Used (L)'])
            
            # Get individual fuel types (excluding ALL rows)
            fuel_by_type = df[(df['date'] == 'ALL') & (df['Type'] != 'ALL')]
            
            for idx, row in fuel_by_type.iterrows():
                fuel_type = str(row['Type'])
                fuel_used = float(row['Fuel Used (L)'])
                
                # Calculate percentage of total
                pct_of_total = (fuel_used / total_fuel * 100) if total_fuel > 0 else 0
                
                # Flag if a single type uses > 60% of total fuel (configurable threshold)
                if pct_of_total > 60:
                    deviation_entry = {
                        'sheet': target_sheet,
                        'flag': 'High Consumption',
                        'severity': 'Medium',
                        'description': f"{fuel_type} fuel consumption is {pct_of_total:.1f}% of total ({fuel_used:,.2f}L)",
                        'row_data': {
                            'fuel_type': fuel_type,
                            'fuel_used': fuel_used,
                            'total_fuel': total_fuel,
                            'percentage': round(pct_of_total, 2)
                        },
                        'detected_at': datetime.now().isoformat()
                    }
                    
                    self.deviations.append(deviation_entry)
                    
            # Clean up memory
            del df
            gc.collect()
                    
        except Exception as e:
            print(f"[DEVIATION] Fuel analysis error: {e}")
        finally:
            if wb:
                wb.close()
    
    def analyze_all(self, processing_results):
        """
        Analyze all output files with memory optimization - process one at a time.
        
        Args:
            processing_results: List of processing result dictionaries
            
        Returns:
            list: List of detected deviations
        """
        print(f"\n[DEVIATION] Starting deviation analysis for job {self.job_id}")
        
        # Process one file at a time to minimize memory usage
        for idx, result in enumerate(processing_results, 1):
            if result['status'] != 'success':
                continue
            
            output_path = Path(result['output_path'])
            
            if not output_path.exists():
                continue
            
            print(f"[DEVIATION] Analyzing ({idx}/{len(processing_results)}): {output_path.name}")
            
            # Use the processor/description metadata to decide analysis type.
            actual_sheet = result['sheet_name']          # e.g. "Sheet1"
            description  = result.get('description', '').strip()
            processor    = result.get('processor', '').lower()
            combined     = description.lower() + ' ' + processor

            # Label is just the algo name
            algo_label = description if description else processor.replace('processor', '').strip().title()
            sheet_label = algo_label if algo_label else actual_sheet
            
            try:
                count_before = len(self.deviations)
                ran_any = False

                # Timeline / schedule analysis
                if any(k in combined for k in ['eddr', 'project', 'timeline', 'schedule', 'activity',
                                               'commissioning', 'rfsu', 'precomm', 'manufacture',
                                               'procurement', 'subcontract', 'overall', 'progress']):
                    print(f"[DEVIATION] Running timeline analysis on sheet: {actual_sheet}")
                    self.analyze_timeline_deviations(output_path, sheet_name=actual_sheet)
                    ran_any = True

                # Cost / budget analysis
                if any(k in combined for k in ['cost', 'budget', 'financial']):
                    print(f"[DEVIATION] Running cost analysis on sheet: {actual_sheet}")
                    self.analyze_cost_deviations(output_path, sheet_name=actual_sheet)
                    ran_any = True

                # Quantity / material analysis
                if any(k in combined for k in ['quantity', 'material', 'resource', 'subcontract']):
                    print(f"[DEVIATION] Running quantity analysis on sheet: {actual_sheet}")
                    self.analyze_quantity_deviations(output_path, sheet_name=actual_sheet)
                    ran_any = True

                # Fuel analysis
                if any(k in combined for k in ['fuel', 'consumption']):
                    print(f"[DEVIATION] Running fuel analysis on sheet: {actual_sheet}")
                    self.analyze_fuel_deviations(output_path, sheet_name=actual_sheet)
                    ran_any = True

                # Fallback: if nothing matched, always attempt timeline analysis
                if not ran_any:
                    print(f"[DEVIATION] No type matched via metadata — attempting timeline analysis on: {actual_sheet}")
                    self.analyze_timeline_deviations(output_path, sheet_name=actual_sheet)

                # Stamp algo label and deduplicate new deviations
                new_devs = self.deviations[count_before:]
                self.deviations = self.deviations[:count_before]  # trim back
                existing_keys = {(d['description'], d['sheet']) for d in self.deviations}
                for dev in new_devs:
                    dev['sheet'] = sheet_label
                    key = (dev['description'], dev['sheet'])
                    if key not in existing_keys:
                        self.deviations.append(dev)
                        existing_keys.add(key)
                
                # Force garbage collection after each file
                gc.collect()
                    
            except Exception as e:
                print(f"[DEVIATION] Error analyzing {output_path.name}: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        print(f"[DEVIATION] Analysis complete: {len(self.deviations)} deviation(s) detected")
        
        return self.deviations
    
    def format_for_database(self, user_id, company_id=None):
        """
        Format deviations for database storage.
        
        Args:
            user_id: ID of user who uploaded the file
            company_id: Optional company ID
            
        Returns:
            list: Formatted deviation records
        """
        formatted = []
        
        for deviation in self.deviations:
            record = {
                'id': None,  # Will be auto-assigned
                'sheet': deviation['sheet'],
                'flag': deviation['flag'],
                'severity': deviation['severity'],
                'description': deviation['description'],
                'row_data': deviation['row_data'],
                'detected_at': deviation['detected_at'],
                'review_status': 'Pending',
                'review_reason': '',
                'user_id': user_id,
                'company_id': company_id,
                'job_id': self.job_id,
                'filename': self.filename
            }
            
            formatted.append(record)
        
        return formatted


    def analyze_raw_input_deviations(self, input_path: str):
        """
        Scan ALL sheets in the original uploaded file for pre-computed deviation
        columns (Output for Deviation, Output for Start Delay, etc.).
        These columns are computed by P6 exports and ignored by newalgo.
        """
        if not input_path or not Path(input_path).exists():
            return

        print(f"[DEVIATION] Scanning raw input for pre-computed deviation columns: {Path(input_path).name}")
        wb = None
        try:
            wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close(); wb = None

            for sheet_name in sheets:
                try:
                    df = pd.read_excel(input_path, sheet_name=sheet_name, engine='openpyxl')
                    if df.empty:
                        continue

                    df.columns = df.columns.str.strip()
                    col_map = {c.lower(): c for c in df.columns}

                    # Look for pre-computed deviation columns (P6 export style)
                    print(f"[DEVIATION] Raw '{sheet_name}' cols: {list(col_map.keys())[:20]}")

                    # ── Column detection ─────────────────────────────────────
                    def _rc(*variants):
                        for v in variants:
                            if v in col_map:
                                return col_map[v]
                        return None

                    name_col        = _rc('activity name', 'activity_name', 'description', 'name')
                    id_col          = _rc('activity id', 'activity_id', 'activity code', 'id')

                    # P6-style: negative total float = behind schedule
                    float_col       = _rc('total float', 'total_float', 'float')
                    early_start_col = _rc('early start', 'early_start', 'planned start', 'ep start')
                    early_finish_col= _rc('early finish', 'early_finish', 'planned finish', 'ep finish')
                    late_finish_col = _rc('late finish', 'late_finish')
                    actual_start_col= _rc('actual start', 'start [actual]', 'actual start date')
                    actual_end_col  = _rc('actual finish', 'finish [actual]', 'actual completion date')

                    # P6 combined fields: 'start' = actual if started else early start
                    start_col       = _rc('start')
                    finish_col      = _rc('finish')

                    # Pre-computed deviation columns (some P6 exports include these)
                    dev_val_col     = _rc('output for deviation') or next(
                        (col_map[k] for k in col_map if 'output' in k and 'deviation' in k and 'flag' not in k), None)
                    dev_flag_col    = _rc('output for deviation flag') or next(
                        (col_map[k] for k in col_map if 'output' in k and 'deviation' in k and 'flag' in k), None)
                    sd_val_col      = _rc('output for start delay') or next(
                        (col_map[k] for k in col_map if 'output' in k and 'start delay' in k and 'flag' not in k), None)

                    has_p6 = float_col or early_finish_col
                    has_precomputed = dev_val_col or sd_val_col

                    print(f"[DEVIATION] Raw '{sheet_name}': float={float_col!r} "
                          f"early_finish={early_finish_col!r} dev_val={dev_val_col!r} rows={len(df)}")

                    if not has_p6 and not has_precomputed:
                        continue

                    today_raw = datetime.now().date()

                    # ── Diagnostics: float distribution ───────────────────────
                    if float_col:
                        float_series = pd.to_numeric(df[float_col], errors='coerce').dropna()
                        neg_count = int((float_series < -1).sum())
                        print(f"[DEVIATION] Raw '{sheet_name}' float stats: "
                              f"min={float_series.min():.1f} max={float_series.max():.1f} "
                              f"mean={float_series.mean():.1f} negative(<-1)={neg_count}/{len(float_series)}")
                    if early_finish_col:
                        ef_series = pd.to_datetime(df[early_finish_col], errors='coerce').dropna()
                        overdue_ef = int((ef_series.dt.date < today_raw).sum())
                        print(f"[DEVIATION] Raw '{sheet_name}' early_finish: "
                              f"{len(ef_series)} non-null, {overdue_ef} before today ({today_raw})")

                    found = 0

                    for _, row in df.iterrows():
                        activity_name = str(row.get(name_col, '') or '').strip() if name_col else ''
                        activity_id   = str(row.get(id_col,   '') or '').strip() if id_col   else ''
                        if not activity_name and not activity_id:
                            continue

                        # ── Signal 1: Negative total float (P6 critical delay) ──
                        float_days = 0.0
                        if float_col:
                            try:
                                float_days = float(row.get(float_col, 0) or 0)
                            except (TypeError, ValueError):
                                float_days = 0.0

                        # ── Signal 2: Pre-computed deviation columns ──────────
                        dev_val = sd_val = 0.0
                        if dev_val_col:
                            try: dev_val = float(row.get(dev_val_col, 0) or 0)
                            except (TypeError, ValueError): pass
                        if sd_val_col:
                            try: sd_val = float(row.get(sd_val_col, 0) or 0)
                            except (TypeError, ValueError): pass
                        dev_flag = str(row.get(dev_flag_col, '') or '').strip() if dev_flag_col else ''

                        # ── Signal 3: Overdue early finish ───────────────────
                        # Only use actual_end_col (not finish_col — P6 always populates
                        # 'finish' with the projected date, so it's not a completion signal)
                        overdue_days = 0.0
                        if early_finish_col:
                            try:
                                ef_val = row.get(early_finish_col)
                                ef_val = None if pd.isnull(ef_val) else ef_val
                                if ef_val is not None:
                                    ef_ts = pd.to_datetime(ef_val, errors='coerce')
                                    if not pd.isnull(ef_ts):
                                        ef_date = ef_ts.date()
                                        if ef_date < today_raw:
                                            # Only treat as done if actual_finish is present
                                            if actual_end_col:
                                                af_val = row.get(actual_end_col)
                                                af_done = af_val is not None and not pd.isnull(
                                                    pd.to_datetime(af_val, errors='coerce'))
                                            else:
                                                af_done = False
                                            if not af_done:
                                                overdue_days = (today_raw - ef_date).days
                            except Exception:
                                pass

                        # Determine if this row is a deviation
                        is_delayed = (
                            float_days < -1          # negative total float
                            or dev_val > 0           # pre-computed deviation
                            or sd_val > 0            # pre-computed start delay
                            or overdue_days > 0      # overdue but not finished
                            or 'delay' in dev_flag.lower()
                        )
                        if not is_delayed:
                            continue

                        # Best magnitude for severity
                        best_days = max(abs(float_days), abs(dev_val), abs(sd_val), overdue_days)
                        if best_days == 0:
                            best_days = 1
                        severity = self.calculate_severity(best_days, 'timeline')
                        if not severity:
                            continue

                        parts = []
                        if float_days < -1:
                            parts.append(f"Total float: {int(float_days)}d")
                        if sd_val > 0:
                            parts.append(f"Start delay: {int(sd_val)}d")
                        if dev_val > 0:
                            parts.append(f"Deviation: {int(dev_val)}d")
                        if overdue_days > 0:
                            parts.append(f"Overdue: {int(overdue_days)}d past planned finish")
                        if not parts:
                            parts.append('Flagged as delayed')

                        self.deviations.append({
                            'sheet':       sheet_name,
                            'flag':        'Negative Float' if float_days < -1 else (dev_flag or 'Delayed'),
                            'severity':    severity,
                            'description': f"{activity_name or activity_id} — {', '.join(parts)}",
                            'row_data': {
                                'activity_id':        activity_id,
                                'activity_name':      activity_name,
                                'total_float':        float_days,
                                'start_delay':        sd_val,
                                'duration_deviation': dev_val,
                                'overdue_days':       overdue_days,
                            },
                            'detected_at': datetime.now().isoformat()
                        })
                        found += 1

                    if found:
                        print(f"[DEVIATION] Raw '{sheet_name}': {found} deviation(s) detected")

                    del df
                    gc.collect()

                except Exception as e:
                    print(f"[DEVIATION] Raw sheet '{sheet_name}' error: {e}")

        except Exception as e:
            print(f"[DEVIATION] Raw input scan error: {e}")
        finally:
            if wb:
                wb.close()


def calculate_deviations(output_folder, job_id, filename, processing_results,
                         user_id, company_id=None, input_path=None):
    """
    Calculate deviations from processed output files AND from pre-computed
    deviation columns in the original uploaded file.
    """
    calculator = DeviationCalculator(output_folder, job_id, filename)
    calculator.analyze_all(processing_results)
    # Also scan the raw uploaded file for P6-style pre-computed deviation columns
    if input_path:
        calculator.analyze_raw_input_deviations(input_path)
    return calculator.format_for_database(user_id, company_id)
