

import openpyxl
from pathlib import Path
import os
from datetime import datetime
import gc  # For memory cleanup
import time
import multiprocessing
from concurrent.futures import ThreadPoolExecutor, as_completed
import re


def _patch_openpyxl_chartsheet_compat() -> None:
    """Monkey patch openpyxl.load_workbook with a safe retry for Chartsheet defined_names errors."""
    if getattr(openpyxl, "_chartsheet_defined_names_patch", False):
        return

    original_load_workbook = openpyxl.load_workbook

    def _safe_load_workbook(*args, **kwargs):
        try:
            return original_load_workbook(*args, **kwargs)
        except AttributeError as exc:
            msg = str(exc)
            if "Chartsheet" not in msg or "defined_names" not in msg:
                raise

            from openpyxl.chartsheet.chartsheet import Chartsheet

            if not hasattr(Chartsheet, "defined_names"):
                Chartsheet.defined_names = {}
            return original_load_workbook(*args, **kwargs)

    openpyxl.load_workbook = _safe_load_workbook
    openpyxl._chartsheet_defined_names_patch = True


_patch_openpyxl_chartsheet_compat()

# Import dynamic algorithm scanner
from algorithm_scanner import get_scanner

# Import all processors from algorithm package
from algorithm import (
    EDDRProcessor, ProjectManagementProcessor,
    HOSubcontractProcessor, HOAsBuiltsProcessor, CommissioningRFSUProcessor,
    ConstPreCommProcessor, HOProcurementsProcessor,
    ManufactureProcessor, TimelineDeviationProcessor,
)

# ============================================================================
# PERFORMANCE CONFIGURATION
# ============================================================================
# Number of parallel workers (auto-detect based on CPU cores, cap at 6)
MAX_WORKERS = min(multiprocessing.cpu_count(), 6)
# Minimum file size (bytes) to bother with parallel processing
PARALLEL_THRESHOLD_BYTES = 100_000  # 100 KB


class UnifiedFileProcessor:
    """Detects and processes Excel files with multiple sheets."""
    
    def __init__(self, input_file):
        """
        Initialize processor with input file.
        
        Args:
            input_file: Path to the input Excel file
        """
        self.input_file = Path(input_file)
        self.workbook = None
        self.detected_sheets = {}
        self.results = []
        self.scanner = get_scanner()
        
        # Build PROCESSORS dict dynamically from scanner
        self.PROCESSORS = self._build_processors_config()
        
    def _build_processors_config(self):
        """Build the PROCESSORS configuration from available algorithms."""
        processors = {}
        
        # Get all available algorithms
        algorithms = self.scanner.available_algorithms
        
        # Map to the old config format for compatibility
        for algo_id, algo_info in algorithms.items():
            # Extract detection config
            detection_config = algo_info.get('detection_config', {})
            
            processors[algo_id] = {
                'processor_class': algo_info['processor_class'],
                'output_suffix': detection_config.get('output_suffix', f"{algo_info['name']}_Output"),
                'description': algo_info['description'],
                'detection_columns': detection_config.get('detection_columns', []),
                'detection_row': detection_config.get('detection_row', 10),
                'expected_headers': detection_config.get('expected_headers', []),
                'name': algo_info['name']
            }
        
        print(f"[PROCESSOR] Loaded {len(processors)} algorithm(s) dynamically")
        return processors
        
    def validate_input_file(self):
        """Validate that input file exists and is an Excel file."""
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
        if not self.input_file.suffix.lower() in ['.xlsx', '.xlsm', '.xls']:
            raise ValueError(f"Invalid file type: {self.input_file.suffix}. Only Excel files allowed.")
        return True
    
    def detect_processor_type(self, sheet, sheet_name=None):
        """
        Detect which processor to use based on sheet name pattern matching and data structure.
        
        Args:
            sheet: openpyxl worksheet object
            sheet_name: Name of the sheet (optional, used for pattern matching)
            
        Returns:
            tuple: (processor_key, confidence_score) or (None, error_message)
        """
        print(f"  → Analyzing sheet structure...")
        
        # STEP 1: Try sheet name pattern matching first (more reliable)
        if sheet_name:
            sheet_lower = sheet_name.lower().strip()
            sheet_norm = _normalize_match_text(sheet_name)
            
            # Exact and pattern matches for each algorithm
            name_patterns = {
                'main1': ['eddr', 'eddr timeline'],  # EDDR (not EDDR CNTR or Weekly EDDR)
                'eddr_cntr': ['eddr cntr', 'eddr contractor'],
                'weekly_eddr_cont': ['weekly eddr cont', 'delayed docs', 'engg table'],
                'project_management': ['project mangement', 'project management'],  # Note: typo 'mangement' is in actual sheet
                'newalgo': ['activity timeline tracker', 'timeline deviation', 'timeline tracker'],
                'ho_procurements': ['ho-procurements', 'ho procurements'],
                'ho_subcontract': ['ho-subcontract', 'ho subcontract'],
                'ho_as_builts': ['ho-as builts', 'ho-as built', 'ho as builts', 'ho as built', 'as-built', 'as built'],
                'manufacture': ['manufacture', 'manufacturing'],
                'commissioning_rfsu': ['commissioning, rfsu', 'commissioning rfsu', 'rfsu'],
                'const_precomm': ['const & pre-comm', 'const. & precomm', 'construction pre'],
                'overall_s_curve': ['home office_s-curve', 'overall s-curve'],
                'pm_s_curve': ['pm_s-curve', 'pm s-curve', 'manufacturing_s-curve', 'const. & precomm._s-curve'],
                'bl_overall_progress_lv2': ['bl overall project progress lv2', 'bl progress lv2', 'mdr', 'comp. dates'],
                'revised_bl_overall_progress': ['revised bl overall progress', 'revised bl project progress'],
            }
            
            # Check for pattern matches
            for algo_key, patterns in name_patterns.items():
                for pattern in patterns:
                    pattern_norm = _normalize_match_text(pattern)
                    if pattern in sheet_lower or (pattern_norm and pattern_norm in sheet_norm):
                        # Const & Pre-Comm tracker must never bind to S-curve tabs.
                        if algo_key == 'const_precomm' and 'curve' in sheet_lower:
                            continue

                        # Special case: if sheet is "EDDR" and pattern is "eddr", make sure it's not EDDR CNTR
                        if pattern == 'eddr' and ('cntr' in sheet_lower or 'contractor' in sheet_lower or 'weekly' in sheet_lower):
                            continue
                        
                        if algo_key in self.PROCESSORS:
                            config = self.PROCESSORS[algo_key]
                            print(f"  ✓ Detected by name: {config['name']} (100% confidence)")
                            return algo_key, 1.0
        
        # STEP 2: Fall back to column detection
        best_match = None
        best_score = 0

        # Precompute normalized header row text for optional header-aware scoring
        header_candidates = []
        for c in range(1, min(sheet.max_column, 60) + 1):
            try:
                hv = sheet.cell(row=1, column=c).value
            except Exception:
                hv = None
            if hv is not None and str(hv).strip() != '':
                header_candidates.append(_normalize_match_text(str(hv)))
        
        for processor_key, config in self.PROCESSORS.items():
            detection_row = config['detection_row']
            detection_columns = config['detection_columns']
            
            if not detection_columns:
                continue
            
            # Check if expected columns have data at detection row
            has_data_count = 0
            for col_idx in detection_columns:
                try:
                    cell_value = sheet.cell(row=detection_row, column=col_idx).value
                    if cell_value is not None and str(cell_value).strip() != '':
                        has_data_count += 1
                except:
                    pass
            
            # Calculate confidence score
            confidence = has_data_count / len(detection_columns) if detection_columns else 0

            # Optional header-aware score boost for disambiguation (e.g., Sheet2-like names)
            expected_headers = config.get('expected_headers', [])
            header_score = 0
            if expected_headers and header_candidates:
                matched = 0
                for expected in expected_headers:
                    expected_norm = _normalize_match_text(expected)
                    if any(expected_norm in hc for hc in header_candidates):
                        matched += 1
                header_score = matched / len(expected_headers)

            # Combined score:
            # - For algorithms with expected headers, header match is mandatory/primary.
            # - For others, keep legacy structure-only scoring.
            if expected_headers:
                combined_score = (0.4 * confidence) + (0.8 * header_score)
                if header_score < 0.75:
                    combined_score = min(combined_score, 0.49)
            else:
                combined_score = confidence
            
            if combined_score > best_score:
                best_score = combined_score
                best_match = processor_key
        
        # If we have a partial match (at least 50%)
        if best_match and best_score >= 0.7:
            print(f"  ✓ Detected by structure: {self.PROCESSORS[best_match]['name']} (confidence: {best_score*100:.0f}%)")
            return best_match, best_score

        if best_match and best_score >= 0.5:
            print(f"  ⚠ Partial match: {self.PROCESSORS[best_match]['name']} (confidence: {best_score*100:.0f}%)")
            print(f"  ℹ Sheet may not follow standard template exactly")
            return best_match, best_score
        
        # No match found - sheet doesn't follow any known template
        print(f"  ✗ Sheet structure doesn't match any known template")
        error_msg = "Sheet doesn't follow standard PMO templates. Please ensure your Excel file matches one of the supported formats: EDDR, Project Management, Weekly EDDR Cont, or HO Subcontract."
        return None, error_msg
        
    def detect_sheets(self):
        """
        Detect which sheets exist in the workbook and determine their processor type.
        
        Returns:
            tuple: (detected_sheets_dict, failed_sheets_list)
        """
        try:
            print("Loading workbook for sheet detection...")
            self.workbook = openpyxl.load_workbook(self.input_file, data_only=True)
            available_sheets = self.workbook.sheetnames
            print(f"Found {len(available_sheets)} sheet(s): {', '.join(available_sheets)}")
            
            failed_sheets = []
            
            # Analyze each sheet to determine processor type
            for sheet_name in available_sheets:
                print(f"\n📋 Analyzing sheet: '{sheet_name}'")
                sheet = self.workbook[sheet_name]
                
                # Detect which processor should be used (pass sheet_name for pattern matching)
                result = self.detect_processor_type(sheet, sheet_name)
                
                if isinstance(result, tuple) and len(result) == 2:
                    processor_key, confidence_or_error = result
                    
                    if processor_key is None:
                        # Sheet doesn't match any template
                        failed_sheets.append({
                            'sheet_name': sheet_name,
                            'error': confidence_or_error,
                            'reason': 'Template mismatch'
                        })
                        print(f"  ✗ Cannot process: {confidence_or_error}")
                    else:
                        # Sheet matched a processor
                        config = self.PROCESSORS[processor_key]
                        self.detected_sheets[sheet_name] = {
                            'config': config,
                            'original_name': sheet_name,
                            'processor_key': processor_key,
                            'confidence': confidence_or_error
                        }
                        _cls = config['processor_class']
                        print(f"  ✓ Will process with: {_cls.__name__.replace('Processor', '')} processor")
                else:
                    # Fallback for old return format
                    processor_key = result
                    if processor_key:
                        config = self.PROCESSORS[processor_key]
                        self.detected_sheets[sheet_name] = {
                            'config': config,
                            'original_name': sheet_name,
                            'processor_key': processor_key
                        }
                        _cls = config['processor_class']
                        print(f"  ✓ Will process with: {_cls.__name__.replace('Processor', '')} processor")
            
            if not self.detected_sheets:
                print("  ⚠ No processable sheets found in this workbook.")
                return None, failed_sheets
            
            print(f"\n✓ Total sheets to process: {len(self.detected_sheets)}")
            if failed_sheets:
                print(f"⚠ Sheets that cannot be processed: {len(failed_sheets)}")
            
            return self.detected_sheets, failed_sheets
            
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {str(e)}")
        finally:
            if self.workbook:
                self.workbook.close()
                self.workbook = None
    
    def _create_patched_processor(self, processor_class, actual_sheet_name):
        """
        Create a processor instance with two patches:
        1. load_workbook() always loads `actual_sheet_name` regardless of what the
           processor expects internally (supports any sheet name like 'Sheet1').
        2. extract_data() raises a clear error when no data records are found,
           preventing silent blank-output generation.
        """
        import openpyxl as _opxl

        processor = processor_class(str(self.input_file))

        # --- Patch 1: load_workbook ---
        def patched_load_workbook():
            processor.workbook = _opxl.load_workbook(str(self.input_file), data_only=True)
            wb_sheets = processor.workbook.sheetnames

            # Prefer actual_sheet_name; fall back to first sheet
            target = actual_sheet_name if actual_sheet_name in wb_sheets else wb_sheets[0]
            ws = processor.workbook[target]
            print(f"Loading workbook...")
            print(f"Using sheet '{target}': {ws.max_row} rows x {ws.max_column} columns")

            # Assign to every known sheet attribute name used across all processors
            processor.sheet      = ws   # used by most processors
            processor.eddr_sheet = ws   # EDDR.py
            processor.pm_sheet   = ws   # project_management 1.py

        processor.load_workbook = patched_load_workbook

        # --- Patch 2: extract_data validation ---
        _original_extract = processor.extract_data

        def validated_extract_data():
            _original_extract()
            # Count data collected by any known storage attribute
            data_count = max(
                len(getattr(processor, 'data_records',   [])),
                len(getattr(processor, 'activities',     {})),
                len(getattr(processor, 'activity_order', [])),
            )
            if data_count == 0:
                raise ValueError(
                    f"No data extracted from sheet '{actual_sheet_name}'. "
                    f"The sheet columns/rows do not match the "
                    f"'{processor_class.__name__}' template. "
                    f"Check that data starts at the expected row and columns."
                )
            print(f"[OK] Extracted {data_count} record(s) from '{actual_sheet_name}'")

        processor.extract_data = validated_extract_data
        return processor

    def process_sheet(self, sheet_name, output_folder):
        """
        Process a single sheet with its appropriate processor.
        
        Args:
            sheet_name: Name of the sheet to process
            output_folder: Directory to save output file
            
        Returns:
            dict: Processing result with output file path and metadata
        """
        if sheet_name not in self.detected_sheets:
            raise ValueError(f"Sheet '{sheet_name}' not found in detected sheets")
        
        sheet_info = self.detected_sheets[sheet_name]
        config = sheet_info['config']
        processor_class = config['processor_class']
        output_suffix = config['output_suffix']
        # Build a short friendly label: "EDDRProcessor" → "EDDR", "ProjectManagementProcessor" → "ProjectManagement"
        import re as _re
        _raw = processor_class.__name__.replace('Processor', '')
        description = _re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', _raw) or processor_class.__name__
        
        # Generate output filename using the actual sheet name (sanitized for filesystem)
        safe_sheet_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in sheet_name).strip()
        safe_sheet_name = safe_sheet_name.replace(' ', '_')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{safe_sheet_name}_{timestamp}.xlsx"
        output_path = os.path.join(output_folder, output_filename)
        
        print(f"\n{'='*60}")
        print(f"Processing: '{sheet_name}'")
        print(f"Type: {description}")
        print(f"Processor: {processor_class.__name__}")
        print(f"Output: {output_filename}")
        print(f"{'='*60}")
        
        try:
            # Create processor patched to use actual sheet name (supports any sheet name)
            processor = self._create_patched_processor(processor_class, sheet_name)
            processor.process(output_path)

            # Get file size and guard against blank outputs
            file_size = os.path.getsize(output_path) if os.path.exists(output_path) else 0
            MIN_VALID_SIZE = 8192  # < 8 KB = headers only, no real data
            if file_size < MIN_VALID_SIZE:
                if os.path.exists(output_path):
                    os.remove(output_path)
                raise ValueError(
                    f"Processor ran but produced no data (output was {file_size} bytes). "
                    f"Sheet '{sheet_name}' structure may not match the "
                    f"'{processor_class.__name__}' template."
                )

            # Rename the output sheet tab to match the source sheet name
            try:
                import openpyxl as _opxl
                out_wb = _opxl.load_workbook(output_path)
                if out_wb.sheetnames:
                    out_wb.active.title = sheet_name
                out_wb.save(output_path)
                out_wb.close()
                print(f"[OK] Output sheet tab renamed to '{sheet_name}'")
            except Exception as rename_err:
                print(f"[WARN] Could not rename output sheet: {rename_err}")

            validation_issues, data_row_count = _collect_output_validation(output_path, sheet_name=sheet_name)
            if data_row_count == 0:
                if os.path.exists(output_path):
                    os.remove(output_path)
                raise ValueError("Output file contains no data rows (blank output sheet).")

            file_size = os.path.getsize(output_path)

            result = {
                'sheet_name': sheet_name,
                'description': description,
                'output_filename': output_filename,
                'output_path': output_path,
                'file_size': file_size,
                'status': 'success',
                'processor': processor_class.__name__,
                'validation_issues': validation_issues,
                'validation_issue_count': len(validation_issues),
            }
            
            print(f"✓ Successfully processed '{sheet_name}'")
            
            # Clean up processor memory
            del processor
            gc.collect()
            
            return result
            
        except Exception as e:
            print(f"✗ Error processing '{sheet_name}': {str(e)}")
            import traceback
            traceback.print_exc()
            
            result = {
                'sheet_name': sheet_name,
                'description': description,
                'status': 'error',
                'error': str(e),
                'processor': processor_class.__name__
            }
            return result
    
    def process_all(self, output_folder):
        """
        Process all detected sheets in the workbook.
        
        Args:
            output_folder: Directory to save output files
            
        Returns:
            dict: Summary of processing results including failed sheets
        """
        # Create output folder if needed
        os.makedirs(output_folder, exist_ok=True)
        
        print("\n" + "="*60)
        print("UNIFIED FILE PROCESSOR")
        print("="*60)
        
        # Validate input
        self.validate_input_file()
        print(f"✓ Input file validated: {self.input_file.name}\n")
        
        # Detect sheets
        detected, failed_sheets = self.detect_sheets()
        
        if not detected and not failed_sheets:
            return {
                'status': 'error',
                'message': 'No sheets found in the uploaded file',
                'detected_sheets': [],
                'failed_sheets': [],
                'results': []
            }
        
        if not detected:
            # All sheets failed validation
            return {
                'status': 'error',
                'message': f'{len(failed_sheets)} sheet(s) cannot be processed because they don\'t follow our standard templates',
                'detected_sheets': [],
                'failed_sheets': failed_sheets,
                'total_sheets': len(failed_sheets),
                'success_count': 0,
                'error_count': len(failed_sheets),
                'results': []
            }
        
        print(f"\n{len(detected)} sheet(s) detected for processing")
        if failed_sheets:
            print(f"{len(failed_sheets)} sheet(s) cannot be processed (template mismatch)\n")
        print("="*60)
        
        # Process each detected sheet
        results = []
        for sheet_name in detected.keys():
            result = self.process_sheet(sheet_name, output_folder)
            results.append(result)
        
        # Generate summary
        success_count = sum(1 for r in results if r['status'] == 'success')
        error_count = len(results) - success_count
        
        print("\n" + "="*60)
        print("PROCESSING SUMMARY")
        print("="*60)
        print(f"Total Sheets Attempted: {len(results)}")
        print(f"Successful: {success_count}")
        print(f"Failed: {error_count}")
        if failed_sheets:
            print(f"Skipped (Template Mismatch): {len(failed_sheets)}")
        
        for result in results:
            status_icon = "✓" if result['status'] == 'success' else "✗"
            algo_name = result.get('description', 'Unknown')
            print(f"{status_icon} {result['sheet_name']}: {result['status']} ({algo_name})")
        
        if failed_sheets:
            print("\nSkipped Sheets:")
            for failed in failed_sheets:
                print(f"✗ {failed['sheet_name']}: {failed['reason']}")
        
        print("="*60)
        
        # Build response message
        if success_count > 0 and (error_count > 0 or failed_sheets):
            failed_count = error_count + len(failed_sheets)
            message = f"Processed {success_count} sheet(s) successfully. {failed_count} sheet(s) cannot be processed because they don't follow our standard templates."
        elif success_count > 0:
            message = f"Processed {success_count} of {len(results)} sheet(s) successfully"
        else:
            message = "All sheets failed to process or don't follow standard templates"
        
        return {
            'status': 'success' if success_count > 0 else 'error',
            'message': message,
            'total_sheets': len(results) + len(failed_sheets),
            'success_count': success_count,
            'error_count': error_count,
            'failed_sheets': failed_sheets,
            'detected_sheets': list(detected.keys()),
            'results': results
        }


# ============================================================================
# ACTIVE TRACKERS (strict 7 requested by PMO)
# ============================================================================
ALL_TRACKERS = [
    ("HO-Procurements",             HOProcurementsProcessor,     "ho_procurements_tracker.xlsx"),
    ("Commissioning RFSU",          CommissioningRFSUProcessor,  "commissioning_rfsu_tracker.xlsx"),
    ("Const & Pre-Comm",            ConstPreCommProcessor,       "const_precomm_tracker.xlsx"),
    ("HO-As Builts",                HOAsBuiltsProcessor,         "ho_as_builts_tracker.xlsx"),
    ("HO-Subcontract",              HOSubcontractProcessor,      "ho_subcontract_tracker.xlsx"),
    ("Manufacture",                 ManufactureProcessor,        "manufacture_tracker.xlsx"),
    ("Project Management",          ProjectManagementProcessor,  "project_management_tracker.xlsx"),
    ("EDDR",                        EDDRProcessor,               "p6_consideration_output_updated_v2.xlsx"),
]

# When True, process only the seven explicit trackers above.
# No extra Generic outputs are created for unmatched sheets.
STRICT_SEVEN_ONLY = True


# ============================================================================
# PARALLEL PROCESSING HELPERS
# ============================================================================

def _preload_workbook_cache(input_file):
    """
    Load the Excel file ONCE and cache all sheet data in memory as plain
    Python lists.  This avoids 13 redundant openpyxl.load_workbook() calls.

    Returns:
        dict: { sheet_name: { 'rows': list[tuple], 'max_row': int, 'max_column': int } }
        Also returns the list of sheet names to preserve order.
    """
    t0 = time.time()
    print("[CACHE] Pre-loading workbook into memory...")

    wb = openpyxl.load_workbook(input_file, data_only=True, read_only=True)
    cache = {}
    sheet_names = wb.sheetnames

    for sname in sheet_names:
        ws = wb[sname]
        # Skip chart sheets and other non-data sheets
        if not hasattr(ws, 'iter_rows'):
            print(f"[CACHE] Skipping non-data sheet: '{sname}' ({type(ws).__name__})")
            continue
        # Materialize all rows as tuples of values (fast in read_only mode)
        rows = []
        try:
            for row in ws.iter_rows(values_only=True):
                rows.append(tuple(row))
        except Exception as e:
            print(f"[CACHE] Warning: could not cache sheet '{sname}': {e}")
            continue
        cache[sname] = {
            'rows': rows,
            'max_row': len(rows),
            'max_column': len(rows[0]) if rows else 0,
        }

    wb.close()
    elapsed = time.time() - t0
    total_cells = sum(c['max_row'] * c['max_column'] for c in cache.values())
    print(f"[CACHE] Done in {elapsed:.2f}s — {len(cache)}/{len(sheet_names)} sheets cached, ~{total_cells:,} cells")
    return cache, list(cache.keys())


class _CachedSheet:
    """
    A lightweight mock of an openpyxl Worksheet backed by pre-loaded data.
    Supports `.cell(row, column).value`, `.max_row`, `.max_column`,
    and `iter_rows(values_only=True)`.
    """
    def __init__(self, rows, max_row, max_column, name='Sheet'):
        self._rows = rows       # list of tuples (0-indexed internally)
        self.max_row = max_row
        self.max_column = max_column
        self.title = name

    def cell(self, row=None, column=None):
        """Return a lightweight cell-like object with a .value attribute."""
        return _CachedCell(self._rows, row, column)

    def __getitem__(self, key):
        """Support openpyxl-like row access such as ws[1]."""
        if isinstance(key, int):
            row_idx = key - 1
            if row_idx < 0 or row_idx >= len(self._rows):
                return tuple(_CachedCellObj(None) for _ in range(self.max_column))
            row_data = self._rows[row_idx]
            padded = list(row_data) + [None] * max(0, self.max_column - len(row_data))
            return tuple(_CachedCellObj(v) for v in padded[:self.max_column])
        raise TypeError(f"Unsupported key type for _CachedSheet: {type(key).__name__}")

    def iter_rows(self, min_row=None, max_row=None, min_col=None, max_col=None, values_only=False):
        """Yield rows from the cache, compatible with openpyxl API."""
        start = (min_row or 1) - 1
        end = min(max_row or self.max_row, self.max_row)
        col_start = (min_col or 1) - 1
        col_end = max_col or self.max_column
        for r in range(start, end):
            if r < len(self._rows):
                row_data = self._rows[r]
                sliced = row_data[col_start:col_end]
                if values_only:
                    yield sliced
                else:
                    yield tuple(
                        _CachedCellObj(v) for v in sliced
                    )
            else:
                if values_only:
                    yield tuple(None for _ in range(col_end - col_start))
                else:
                    yield tuple(
                        _CachedCellObj(None) for _ in range(col_end - col_start)
                    )

    @property
    def sheetnames(self):
        return [self.title]


class _CachedCellObj:
    """Minimal cell object with just a .value attribute."""
    __slots__ = ('value',)
    def __init__(self, value):
        self.value = value


class _CachedCell:
    """Provides .value for a specific (row, column) from cached data."""
    __slots__ = ('value',)
    def __init__(self, rows, row, column):
        try:
            self.value = rows[row - 1][column - 1] if row and column else None
        except (IndexError, TypeError):
            self.value = None


class _CachedWorkbook:
    """Minimal mock of openpyxl Workbook backed by cached data."""
    def __init__(self, cache, sheet_names):
        self._cache = cache
        self.sheetnames = sheet_names

    def __getitem__(self, sheet_name):
        info = self._cache.get(sheet_name)
        if info is None:
            raise KeyError(f"Sheet '{sheet_name}' not found in cached workbook")
        return _CachedSheet(info['rows'], info['max_row'], info['max_column'], name=sheet_name)

    def close(self):
        pass  # No-op; data is just Python lists


def _inject_cached_workbook(processor, cached_wb, target_sheet_override=None):
    """
    Replace the processor's load_workbook() with a version that uses the
    pre-cached in-memory data.  Never touches openpyxl – fully thread-safe.

    Every processor's load_workbook() does the same pattern:
      1. openpyxl.load_workbook(...) -> self.workbook
      2. Find target sheet by name (exact, then fuzzy)
      3. self.sheet = self.workbook[target]  (or self.eddr_sheet / self.pm_sheet)
      4. Optionally: self.sheet_data = list(self.sheet.iter_rows(values_only=True))

    This patched version lets openpyxl.load_workbook() return the cached wb,
    but does so per-processor by running the original method in a controlled way.
    """
    # For each processor, we know which attr it stores the sheet in.
    # - EDDRProcessor uses self.eddr_sheet
    # - ProjectManagementProcessor uses self.pm_sheet
    # - All others use self.sheet
    # After load, we also fix self.sheet_data if the processor sets it.

    proc_cls_name = processor.__class__.__name__

    def patched_load_workbook():
        processor.workbook = cached_wb

        # Find the best matching sheet name from the cache
        available = cached_wb.sheetnames

        # Try to determine what sheet the processor wants by checking class-level hints
        target_sheet = target_sheet_override if target_sheet_override in available else None

        hints = TRACKER_SHEET_HINTS.get(proc_cls_name, [])

        # Exact name match first
        if not target_sheet:
            for sname in available:
                sname_lower = sname.lower().strip()
                for hint in hints:
                    if hint in sname_lower:
                        # Const & Pre-Comm tracker must ignore S-curve style sheets.
                        if proc_cls_name == 'ConstPreCommProcessor' and 'curve' in sname_lower:
                            continue

                        # Disambiguation for EDDR
                        if proc_cls_name == 'EDDRProcessor':
                            if 'cntr' in sname_lower or 'weekly' in sname_lower or 'contractor' in sname_lower:
                                continue
                        target_sheet = sname
                        break
                if target_sheet:
                    break

        if not target_sheet and available:
            # Fallback: use first sheet
            target_sheet = available[0]

        ws = cached_wb[target_sheet] if target_sheet else cached_wb[available[0]]

        # Assign to the correct attribute based on processor type
        if proc_cls_name == 'EDDRProcessor':
            processor.eddr_sheet = ws
        elif proc_cls_name == 'ProjectManagementProcessor':
            processor.pm_sheet = ws
        else:
            processor.sheet = ws

        # For processors that pre-cache sheet_data (read_only=True pattern)
        if hasattr(processor, 'sheet_data') or proc_cls_name in (
            'HOAsBuiltsProcessor',
            'HOSubcontractProcessor', 'CommissioningRFSUProcessor',
            'EDDRCNTRProcessor', 'RevisedBLProgressProcessor',
            'BLProgressLv2Processor', 'PMSCurveProcessor',
        ):
            # Build self.sheet_data as list of lists (mimics iter_rows(values_only=True))
            # Processors use [list(row) for row in ws.iter_rows(values_only=True)]
            processor.sheet_data = [list(row) for row in ws.iter_rows(values_only=True)]

        print(f"Loading workbook...\nUsing sheet '{target_sheet}': {ws.max_row} rows × {ws.max_column} columns")

    processor.load_workbook = patched_load_workbook


# Mapping from processor class name -> sheet name keywords to match.
# Used to determine which trackers to run based on sheets present in the file.
TRACKER_SHEET_HINTS = {
    'EDDRProcessor':               ['eddr'],
    'ProjectManagementProcessor':  ['project mangement', 'project management'],
    'HOProcurementsProcessor':     ['ho-procurements', 'ho procurements'],
    'HOAsBuiltsProcessor':         ['ho-as builts', 'ho-as built', 'ho as built', 'ho as builts', 'as-built', 'as built'],
    'HOSubcontractProcessor':      [
        'ho-subcontract',
        'ho subcontract',
    ],
    'ManufactureProcessor':        ['manufacture'],
    'CommissioningRFSUProcessor':  ['commissioning', 'rfsu'],
    'ConstPreCommProcessor':       ['const', 'pre-comm', 'precomm', 'pre-com', 'construction pre commission'],
    'OverallSCurveProcessor':      ['overall s-curve', 'overall s curve', 'home office s-curve', 'home office s curve'],
    'PMSCurveProcessor':           ['pm_s-curve', 'pm s-curve', 'pm s curve', 'manufacturing_s-curve', 'manufacturing s-curve', 'const. & precomm._s-curve', 'const precomm s-curve', 'const pre com s curve'],
    'TimelineDeviationProcessor':  [],
}

# These processors should run for ALL matching sheets (not just first match)
MULTI_SHEET_TRACKERS = set()


def _normalize_header(text):
    s = str(text or '').strip().lower()
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return ' '.join(s.split())


def _normalize_match_text(text):
    s = str(text or '').strip().lower()
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return ' '.join(s.split())


def _find_header_col(headers, patterns):
    for idx, h in enumerate(headers, start=1):
        if h in patterns:
            return idx
    for idx, h in enumerate(headers, start=1):
        for p in patterns:
            if p in h:
                return idx
    return None


def _is_empty_cell(value):
    if value is None:
        return True
    text = str(value).strip()
    if text == '':
        return True
    return text.lower() in {'nan', 'none', 'null', 'nat'}


def _find_header_row_index(rows, max_scan_rows=40, min_non_empty=3):
    if not rows:
        return 0
    scan_limit = min(max_scan_rows, len(rows))
    for idx in range(scan_limit):
        row = rows[idx] if idx < len(rows) else ()
        non_empty = sum(1 for cell in row if not _is_empty_cell(cell))
        if non_empty >= min_non_empty:
            return idx
    return 0


def _find_first_matching_col(headers, patterns):
    # Exact
    for idx, h in enumerate(headers):
        if h in patterns:
            return idx
    # Partial
    for idx, h in enumerate(headers):
        for p in patterns:
            if p in h:
                return idx
    return None


def _to_comparable_date(value):
    """Try converting Excel/date-like values into comparable datetime values."""
    if _is_empty_cell(value):
        return None


def _collect_output_rows(output_path):
    """Read the first sheet from output workbook as plain rows."""
    wb = None
    try:
        wb = openpyxl.load_workbook(output_path, data_only=True, read_only=True)
        if not wb.sheetnames:
            return []
        ws = wb[wb.sheetnames[0]]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        return rows
    finally:
        if wb:
            wb.close()


def _collect_output_validation(output_path, sheet_name='Sheet'):
    """Validate produced output workbook and return issues + data row count."""
    rows = _collect_output_rows(output_path)
    if not rows:
        return ([{
            'type': 'blank_output_sheet',
            'severity': 'high',
            'message': 'Output sheet is blank (no rows found).',
            'count': 0,
            'samples': [],
        }], 0)

    header_idx = _find_header_row_index(rows)
    data_rows = []
    for r in rows[header_idx + 1:]:
        if any(not _is_empty_cell(c) for c in r):
            data_rows.append(r)

    issues = _collect_sheet_validation_issues(sheet_name, {
        'rows': rows,
        'max_row': len(rows),
        'max_column': max((len(r) for r in rows), default=0),
    })

    if len(data_rows) == 0:
        issues = issues + [{
            'type': 'blank_output_sheet',
            'severity': 'high',
            'message': 'Output sheet has headers but no data rows.',
            'count': 0,
            'samples': [],
        }]

    return issues, len(data_rows)
    try:
        import pandas as _pd
        dt = _pd.to_datetime(value, errors='coerce')
        if _pd.isna(dt):
            return None
        return dt
    except Exception:
        return None


def _collect_sheet_validation_issues(sheet_name, sheet_cache):
    """
    Detect manual/data-missing errors for a sheet before deviation analysis.

    Returns:
        list[dict]: [{type, severity, message, count}]
    """
    rows = (sheet_cache or {}).get('rows', []) or []
    issues = []

    if not rows:
        return [{
            'type': 'empty_sheet',
            'severity': 'high',
            'message': 'Sheet has no data rows to validate.',
            'count': 0,
        }]

    header_idx = _find_header_row_index(rows)
    header_row = rows[header_idx] if header_idx < len(rows) else ()
    headers = [_normalize_header(v) for v in header_row]

    activity_col = _find_first_matching_col(headers, [
        'activity id', 'activity code', 'activity_id', 'target activity id', 'trim id', 'wbs code', 'task id', 'id'
    ])

    name_col = _find_first_matching_col(headers, [
        'activity name', 'activity description', 'description', 'item', 'task name', 'milestone', 'wbs name', 'activity'
    ])

    ep_col = _find_first_matching_col(headers, ['ep dates', 'ep date', 'ep', 'early planning', 'early plan', 'early start', 'planned start', 'baseline start', 'bl start'])
    lp_col = _find_first_matching_col(headers, ['lp dates', 'lp date', 'lp', 'late planning', 'late plan', 'late start', 'planned finish', 'baseline finish', 'bl finish', 'late finish', 'last planned date'])

    planned_cols = [ep_col, lp_col, _find_first_matching_col(headers, ['planned date', 'bl date', 'baseline date'])]
    planned_cols = [c for c in planned_cols if c is not None]

    actual_cols = [
        _find_first_matching_col(headers, ['actual date', 'actual start date', 'actual finish date', 'actuals', 'actual start date as a', 'actual start', 'actual finish']),
        _find_first_matching_col(headers, ['actual start']),
        _find_first_matching_col(headers, ['actual finish']),
    ]
    actual_cols = [c for c in actual_cols if c is not None]

    # Be conservative: if a sheet has no recognizable activity/date structure,
    # don't emit generic warnings.
    if activity_col is None and not planned_cols and not actual_cols:
        return []

    data_rows = []
    for r in rows[header_idx + 1:]:
        if any(not _is_empty_cell(c) for c in r):
            data_rows.append(r)

    if not data_rows:
        return [{
            'type': 'no_data_rows',
            'severity': 'high',
            'message': 'No data rows found below header row.',
            'count': 0,
        }]

    missing_activity = 0
    missing_planned = 0
    missing_actual = 0
    deviation_unavailable = 0
    ep_after_lp = 0
    samples_missing_activity = []
    samples_missing_planned = []
    samples_missing_actual = []
    samples_deviation_unavailable = []
    samples_ep_after_lp = []

    for data_idx, row in enumerate(data_rows):
        has_any = any(not _is_empty_cell(c) for c in row)
        if not has_any:
            continue

        row_no = header_idx + 2 + data_idx
        activity_id_val = row[activity_col] if (activity_col is not None and activity_col < len(row)) else ''
        activity_name_val = row[name_col] if (name_col is not None and name_col < len(row)) else ''

        def _sample(reason):
            return {
                'row': row_no,
                'activity_id': '' if _is_empty_cell(activity_id_val) else str(activity_id_val),
                'activity_name': '' if _is_empty_cell(activity_name_val) else str(activity_name_val),
                'reason': reason,
            }

        has_activity = True
        # Only enforce activity id when timeline/date fields exist in the row.
        row_has_planned_value = any((idx < len(row)) and not _is_empty_cell(row[idx]) for idx in planned_cols) if planned_cols else False
        row_has_actual_value = any((idx < len(row)) and not _is_empty_cell(row[idx]) for idx in actual_cols) if actual_cols else False
        requires_activity = row_has_planned_value or row_has_actual_value

        if activity_col is not None:
            activity_val = row[activity_col] if activity_col < len(row) else None
            has_activity = not _is_empty_cell(activity_val)
            if requires_activity and not has_activity:
                missing_activity += 1
                if len(samples_missing_activity) < 25:
                    samples_missing_activity.append(_sample('Activity ID missing'))
        elif requires_activity:
            # Treat as missing activity only for rows where date data exists.
            missing_activity += 1
            has_activity = False
            if len(samples_missing_activity) < 25:
                samples_missing_activity.append(_sample('Activity ID column unavailable for row with date data'))

        if has_activity:
            has_planned = True
            if planned_cols:
                has_planned = any(
                    (idx < len(row)) and not _is_empty_cell(row[idx])
                    for idx in planned_cols
                )
                if not has_planned:
                    missing_planned += 1
                    if len(samples_missing_planned) < 25:
                        samples_missing_planned.append(_sample('Planned/EP/LP date missing'))
            else:
                # If planned column doesn't exist, don't force a missing issue.
                has_planned = row_has_planned_value

            has_actual = True
            if actual_cols:
                has_actual = any(
                    (idx < len(row)) and not _is_empty_cell(row[idx])
                    for idx in actual_cols
                )
                if not has_actual:
                    missing_actual += 1
                    if len(samples_missing_actual) < 25:
                        samples_missing_actual.append(_sample('Actual date missing'))
            else:
                # If actual column doesn't exist, don't force a missing issue.
                has_actual = row_has_actual_value

            # Deviation unavailable only when at least one side has data and the
            # counterpart is missing in the same row.
            if (row_has_planned_value and not has_actual) or (row_has_actual_value and not has_planned):
                deviation_unavailable += 1
                if len(samples_deviation_unavailable) < 25:
                    samples_deviation_unavailable.append(_sample('Deviation not possible due to missing planned/actual counterpart'))

            # EP > LP validation when both values are present and parseable
            if ep_col is not None and lp_col is not None:
                ep_val = row[ep_col] if ep_col < len(row) else None
                lp_val = row[lp_col] if lp_col < len(row) else None
                ep_dt = _to_comparable_date(ep_val)
                lp_dt = _to_comparable_date(lp_val)
                if ep_dt is not None and lp_dt is not None and ep_dt > lp_dt:
                    ep_after_lp += 1
                    if len(samples_ep_after_lp) < 25:
                        samples_ep_after_lp.append(_sample(f'EP ({ep_dt.date()}) greater than LP ({lp_dt.date()})'))

    if missing_activity > 0:
        issues.append({
            'type': 'activity_id_missing',
            'severity': 'high',
            'message': f'{missing_activity} row(s) have missing Activity ID.',
            'count': missing_activity,
            'samples': samples_missing_activity,
        })

    if missing_planned > 0:
        issues.append({
            'type': 'planned_date_missing',
            'severity': 'medium',
            'message': f'{missing_planned} row(s) have missing planned/baseline date values.',
            'count': missing_planned,
            'samples': samples_missing_planned,
        })

    if missing_actual > 0:
        issues.append({
            'type': 'actual_date_missing',
            'severity': 'medium',
            'message': f'{missing_actual} row(s) have missing actual date values.',
            'count': missing_actual,
            'samples': samples_missing_actual,
        })

    if deviation_unavailable > 0:
        issues.append({
            'type': 'deviation_not_possible',
            'severity': 'high',
            'message': f'Deviation cannot be calculated for {deviation_unavailable} row(s) because planned/actual date is missing.',
            'count': deviation_unavailable,
            'samples': samples_deviation_unavailable,
        })

    if ep_after_lp > 0:
        issues.append({
            'type': 'ep_greater_than_lp',
            'severity': 'high',
            'message': f'{ep_after_lp} row(s) have EP date greater than LP date.',
            'count': ep_after_lp,
            'samples': samples_ep_after_lp,
        })

    return issues


def _sheet_supports_timeline_processor(sheet_cache, strict=False):
    rows = sheet_cache.get('rows', []) or []
    if not rows:
        return False

    max_scan_rows = min(60, len(rows))
    max_scan_cols = min(80, sheet_cache.get('max_column', 0) or 0)
    if max_scan_cols <= 0:
        return False

    code_patterns = ['activity id', 'activity code', 'id', 'code', 'activity_id', 'activity no', 'activity number', 'act id']
    name_patterns = ['activity name', 'description', 'name', 'activity_name', 'task', 'activity description', 'task name']
    planned_start_patterns = [
        'early planned start', 'early planned start date', 'planned start date',
        'early start', 'ep start', 'planned start', 'ep date', 'ep dates',
        'early planning', 'early plan', 'baseline start', 'bl start',
    ]
    planned_finish_patterns = [
        'early planed finish', 'early planned finish', 'early planned finish date',
        'planned end date', 'planned finish date', 'early finish', 'planned finish'
    ]
    actual_start_patterns = ['start [actual]', 'actual start date', 'actual start', 'start actual']
    actual_finish_patterns = ['finish [actual]', 'actual completion date', 'actual finish', 'finish actual', 'actual date']
    if not strict:
        planned_finish_patterns += ['late planned start', 'late planned finish', 'stage gate', 'target date']

    for r in range(max_scan_rows):
        row_vals = rows[r]
        headers = [_normalize_header(v) for v in row_vals[:max_scan_cols]]
        if not any(headers):
            continue

        c_code = _find_header_col(headers, code_patterns)
        c_name = _find_header_col(headers, name_patterns)
        c_planned_start = _find_header_col(headers, planned_start_patterns)
        c_planned_finish = _find_header_col(headers, planned_finish_patterns)
        c_actual_start = _find_header_col(headers, actual_start_patterns)
        c_actual_finish = _find_header_col(headers, actual_finish_patterns)

        if strict:
            if all([c_code, c_name, c_planned_start, c_planned_finish, c_actual_start, c_actual_finish]):
                return True
        elif all([c_code, c_name, c_planned_start, c_planned_finish]):
            return True

    return False


def _sheet_has_data_rows(sheet_cache):
    """Return True when a sheet has at least one non-empty row below the detected header."""
    rows = (sheet_cache or {}).get('rows', []) or []
    if not rows:
        return False

    header_idx = _find_header_row_index(rows)
    for row in rows[header_idx + 1:]:
        if any(not _is_empty_cell(cell) for cell in row):
            return True
    return False


def _sheet_matches_known_tracker_name(sheet_name):
    """Return True when sheet title clearly maps to one of the explicit legacy trackers."""
    sheet_lower = str(sheet_name or '').lower().strip()
    sheet_norm = _normalize_match_text(sheet_name)

    for _, processor_class, _ in ALL_TRACKERS:
        cls_name = processor_class.__name__
        # TimelineDeviation is generic fallback; don't treat it as an explicit-name lock.
        if cls_name == 'TimelineDeviationProcessor':
            continue

        hints = TRACKER_SHEET_HINTS.get(cls_name, [])
        for hint in hints:
            hint_norm = _normalize_match_text(hint)
            if hint in sheet_lower or (hint_norm and hint_norm in sheet_norm):
                if cls_name == 'EDDRProcessor' and (
                    'cntr' in sheet_lower or 'weekly' in sheet_lower or 'contractor' in sheet_lower
                ):
                    continue
                return True

    return False


def _match_trackers_to_sheets(sheet_names, cache, strict_timeline_mode=False):
    """
    Given the list of sheet names in the uploaded workbook, return only
    the trackers from ALL_TRACKERS whose target sheet actually exists.

    Returns:
        matched: list of dict items with tracker metadata and target sheet
        skipped: list of tracker names that had no matching sheet
    """
    sheets_lower = [s.lower().strip() for s in sheet_names]
    sheets_norm = [_normalize_match_text(s) for s in sheet_names]
    matched = []
    skipped = []
    matched_sheet_names = set()

    # Sheets with generic timeline headers should be handled by TimelineDeviationProcessor,
    # not by legacy tracker processors matched only by sheet title.
    timeline_candidate_sheets = {
        sname for sname in sheet_names
        if (
            _sheet_has_data_rows(cache.get(sname))
            and _sheet_supports_timeline_processor(cache.get(sname), strict=strict_timeline_mode)
            and not _sheet_matches_known_tracker_name(sname)
        )
    }

    timeline_only_mode = strict_timeline_mode

    if not strict_timeline_mode:
        for tracker_name, processor_class, output_filename in ALL_TRACKERS:
            cls_name = processor_class.__name__
            hints = TRACKER_SHEET_HINTS.get(cls_name, [])
            matched_sheets_for_tracker = []

            for i, sname_lower in enumerate(sheets_lower):
                sname_norm = sheets_norm[i]
                sheet_match = False
                for hint in hints:
                    hint_norm = _normalize_match_text(hint)
                    if hint in sname_lower or (hint_norm and hint_norm in sname_norm):
                        # Const & Pre-Comm tracker should only target the non-S-curve sheet.
                        if cls_name == 'ConstPreCommProcessor' and 'curve' in sname_lower:
                            continue

                        # Disambiguate EDDR
                        if cls_name == 'EDDRProcessor':
                            if 'cntr' in sname_lower or 'weekly' in sname_lower or 'contractor' in sname_lower:
                                continue
                        sheet_match = True
                        break

                if sheet_match:
                    sheet_name = sheet_names[i]
                    sheet_cache = cache.get(sheet_name)
                    if not _sheet_has_data_rows(sheet_cache):
                        continue

                    matched_sheets_for_tracker.append(sheet_name)
                    if cls_name not in MULTI_SHEET_TRACKERS:
                        break  # single-sheet trackers keep first matching sheet only

            if matched_sheets_for_tracker:
                for idx_sheet, matched_sheet in enumerate(matched_sheets_for_tracker, start=1):
                    if cls_name in MULTI_SHEET_TRACKERS:
                        safe_sheet = ''.join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in matched_sheet).strip().replace(' ', '_')
                        out_name = f"{Path(output_filename).stem}_{safe_sheet}.xlsx"
                        display_name = f"{tracker_name} - {matched_sheet}"
                    else:
                        out_name = output_filename
                        display_name = tracker_name

                    matched.append({
                        'name': display_name,
                        'processor_class': processor_class,
                        'output_filename': out_name,
                        'target_sheet': matched_sheet,
                        'is_generic': False,
                    })
                    matched_sheet_names.add(matched_sheet)
            else:
                skipped.append(tracker_name)

    # In strict timeline mode, legacy trackers are intentionally skipped.
    # Add newalgo fallback for the sheets that actually contain the timeline schema.
    for sname in sheet_names:
        if sname in matched_sheet_names:
            continue
        if sname in timeline_candidate_sheets:
            safe = ''.join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in sname).strip().replace(' ', '_')
            matched.append({
                'name': f"{sname} (Timeline Deviation)",
                'processor_class': TimelineDeviationProcessor,
                'output_filename': f"{safe}_timeline_deviation_tracker.xlsx",
                'target_sheet': sname,
                'is_generic': False,
            })

    return matched, skipped


def _run_single_tracker(args):
    """
    Worker function for parallel execution of a single tracker.
    Runs in a separate thread. Each tracker gets its own processor instance.

    Args:
        args: tuple of (index, total, name, processor_class, output_path, input_file, cache, sheet_names, target_sheet)

    Returns:
        dict: tracker result
    """
    idx, total, name, processor_class, output_path, input_file, cache, sheet_names, target_sheet = args
    t0 = time.time()

    try:
        processor = processor_class(str(input_file))

        # Inject cached workbook to avoid redundant disk reads
        cached_wb = _CachedWorkbook(cache, sheet_names)
        _inject_cached_workbook(processor, cached_wb, target_sheet_override=target_sheet)

        processor.process(output_path)

        # Guard against blank outputs
        file_size = os.path.getsize(output_path) if os.path.exists(output_path) else 0
        MIN_VALID_SIZE = 2048
        if file_size < MIN_VALID_SIZE:
            if os.path.exists(output_path):
                os.remove(output_path)
            raise ValueError(
                f"Output file is too small ({file_size} bytes) – no data was written."
            )

        validation_issues, data_row_count = _collect_output_validation(output_path, sheet_name=name)
        if data_row_count == 0:
            if os.path.exists(output_path):
                os.remove(output_path)
            raise ValueError("Output file contains no data rows (blank output sheet).")

        elapsed = time.time() - t0
        print(f"  [{idx}/{total}] {name} [OK] {file_size // 1024} KB in {elapsed:.1f}s")

        return {
            'sheet_name': name,
            'description': name,
            'output_filename': os.path.basename(output_path),
            'output_path': output_path,
            'file_size': file_size,
            'status': 'success',
            'processor': processor_class.__name__,
            'validation_issues': validation_issues,
            'validation_issue_count': len(validation_issues),
            'elapsed': elapsed,
            'idx': idx,
        }

    except Exception as e:
        elapsed = time.time() - t0
        err_msg = str(e)
        print(f"  [{idx}/{total}] {name} [FAIL] {err_msg[:100]} ({elapsed:.1f}s)")
        return {
            'sheet_name': name,
            'description': name,
            'output_filename': os.path.basename(output_path),
            'status': 'error',
            'error': err_msg,
            'processor': processor_class.__name__,
            'validation_issues': [],
            'validation_issue_count': 0,
            'elapsed': elapsed,
            'idx': idx,
        }


def process_file(input_file, output_folder, progress_callback=None):
    """
    Main entry point for file processing.
    Runs active tracker processors on the input file using PARALLEL execution.

    Optimizations applied:
    - Pre-loads the Excel file ONCE into an in-memory cache (eliminates 12 redundant reads)
    - Runs matched trackers in parallel using ThreadPoolExecutor
    - Each tracker gets its own processor instance with injected cached workbook

    Args:
        input_file: Path to input Excel file
        output_folder: Directory to save output files
        progress_callback: optional callable(current_idx, total, tracker_name, status, result)
                           called after each tracker finishes.

    Returns:
        dict: Processing results summary
    """
    total_start = time.time()
    input_path = Path(input_file)

    # Validate input
    if not input_path.exists():
        return {
            'status': 'error',
            'message': f'Input file not found: {input_file}',
            'total_sheets': 0, 'success_count': 0, 'error_count': 0,
            'failed_sheets': [], 'detected_sheets': [], 'results': []
        }
    if input_path.suffix.lower() not in ['.xlsx', '.xls', '.xlsm']:
        return {
            'status': 'error',
            'message': 'Invalid file type. Only Excel files (.xlsx/.xls/.xlsm) are supported.',
            'total_sheets': 0, 'success_count': 0, 'error_count': 0,
            'failed_sheets': [], 'detected_sheets': [], 'results': []
        }

    os.makedirs(output_folder, exist_ok=True)
    strict_timeline_mode = 'standard format' in input_path.name.lower()

    # ── STEP 1: Pre-load workbook into memory (single read) ──
    try:
        cache, sheet_names = _preload_workbook_cache(str(input_file))
    except Exception as cache_err:
        print(f"[CACHE] Failed to pre-load: {cache_err}")
        print("[CACHE] Falling back to sequential processing...")
        return _process_file_sequential(input_file, output_folder, progress_callback)

    # ── STEP 2: Match trackers to sheets present in the file ──
    if strict_timeline_mode:
        matched_trackers = []
        skipped_trackers = []
        for sname in sheet_names:
            sheet_cache = cache.get(sname)
            if sheet_cache and _sheet_has_data_rows(sheet_cache) and _sheet_supports_timeline_processor(sheet_cache, strict=True):
                safe = ''.join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in sname).strip().replace(' ', '_')
                matched_trackers.append({
                    'name': f"{sname} (Timeline Deviation)",
                    'processor_class': TimelineDeviationProcessor,
                    'output_filename': f"{safe}_timeline_deviation_tracker.xlsx",
                    'target_sheet': sname,
                    'is_generic': False,
                })
            else:
                skipped_trackers.append(sname)
    else:
        matched_trackers, skipped_trackers = _match_trackers_to_sheets(sheet_names, cache, strict_timeline_mode=strict_timeline_mode)

    # Business rule for schedule-feb-update workbook:
    # ignore Sheet1 and process only the relevant schedule sheet(s).
    input_name = input_path.name.lower()
    if 'schedule feb update' in input_name:
        before_count = len(matched_trackers)
        matched_trackers = [
            t for t in matched_trackers
            if str(t.get('target_sheet', '')).strip().lower() != 'sheet1'
        ]
        if len(matched_trackers) < before_count:
            print("[RULE] Skipping 'Sheet1' for schedule-feb-update workbook")

    if not matched_trackers:
        print("[FALLBACK] Strict tracker match found no sheets. Trying dynamic sheet detection...")
        try:
            del cache
            fallback = UnifiedFileProcessor(str(input_file))
            dynamic_result = fallback.process_all(output_folder)
            dynamic_result['fallback_mode'] = 'dynamic_detection'
            return dynamic_result
        except Exception as fallback_err:
            return {
                'status': 'error',
                'message': 'No sheets in the uploaded file match any known tracker template.',
                'details': f'Dynamic fallback failed: {str(fallback_err)}',
                'total_sheets': 0, 'success_count': 0, 'error_count': 0,
                'failed_sheets': [], 'detected_sheets': [], 'results': [],
                'skipped_trackers': skipped_trackers,
            }

    total_matched = len(matched_trackers)
    print("\n" + "=" * 60)
    print(f"PROCESSING {total_matched} MATCHING TRACKER(S) (PARALLEL)")
    print(f"Input  : {input_path.name}")
    print(f"Output : {output_folder}")
    print(f"Workers: {min(MAX_WORKERS, total_matched)} threads")
    if skipped_trackers:
        print(f"Skipped: {len(skipped_trackers)} tracker(s) — no matching sheet")
        for sk in skipped_trackers:
            print(f"  • {sk}")
    print("=" * 60)

    # ── STEP 3: Build task list (only matched trackers) ──
    tasks = []
    for idx, tracker in enumerate(matched_trackers, 1):
        name = tracker['name']
        processor_class = tracker['processor_class']
        output_filename = tracker['output_filename']
        target_sheet = tracker.get('target_sheet')
        output_path = os.path.join(output_folder, output_filename)
        tasks.append((idx, total_matched, name, processor_class, output_path,
                       str(input_file), cache, sheet_names, target_sheet))

    # Notify that matched trackers are starting
    if progress_callback:
        for idx, (name, _, _, _, _, _, _, _, _) in enumerate(tasks):
            try:
                progress_callback(idx + 1, total_matched, name, 'running', None)
            except Exception:
                pass

    # ── STEP 4: Execute matched trackers in parallel ──
    results = [None] * total_matched
    completed_count = 0

    with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, total_matched)) as executor:
        future_to_idx = {}
        for task in tasks:
            future = executor.submit(_run_single_tracker, task)
            future_to_idx[future] = task[0] - 1  # 0-indexed position

        for future in as_completed(future_to_idx):
            pos = future_to_idx[future]
            try:
                result = future.result()
            except Exception as exc:
                name = matched_trackers[pos]['name']
                result = {
                    'sheet_name': name,
                    'description': name,
                    'output_filename': matched_trackers[pos]['output_filename'],
                    'status': 'error',
                    'error': str(exc),
                    'processor': matched_trackers[pos]['processor_class'].__name__,
                    'validation_issues': [],
                    'validation_issue_count': 0,
                    'idx': pos + 1,
                }
            results[pos] = result
            completed_count += 1

            # Progress callback (ordered by completion)
            if progress_callback:
                try:
                    status = result.get('status', 'error')
                    progress_callback(completed_count, total_matched,
                                      result['sheet_name'], status, result)
                except Exception:
                    pass

    # ── STEP 5: Summary ──
    del cache
    gc.collect()

    success_count = sum(1 for r in results if r and r['status'] == 'success')
    error_count = sum(1 for r in results if r and r['status'] == 'error')

    failed_sheets = [
        {'sheet_name': r['sheet_name'], 'error': r.get('error', ''), 'reason': r.get('error', '')}
        for r in results if r and r['status'] == 'error'
    ]
    detected_sheets = [r['sheet_name'] for r in results if r and r['status'] == 'success']

    total_elapsed = time.time() - total_start
    sum_individual = sum(r.get('elapsed', 0) for r in results if r)

    print("\n" + "=" * 60)
    print(f"DONE  {success_count} OK  |  {error_count} FAILED  |  {total_matched} matched  |  {len(skipped_trackers)} skipped")
    print(f"Wall time: {total_elapsed:.1f}s  |  Sum of individual: {sum_individual:.1f}s")
    if sum_individual > 0 and total_elapsed > 0:
        print(f"Parallel speedup: {sum_individual / total_elapsed:.1f}x")
    print("=" * 60)

    # Build user-friendly message
    if success_count > 0:
        msg = f'{success_count} of {total_matched} matching tracker(s) processed successfully'
        if skipped_trackers:
            msg += f' ({len(skipped_trackers)} tracker(s) skipped — sheets not in file)'
    else:
        msg = 'All matching trackers failed to process'

    return {
        'status': 'success' if success_count > 0 else 'error',
        'message': msg,
        'total_sheets': total_matched,
        'success_count': success_count,
        'error_count': error_count,
        'failed_sheets': failed_sheets,
        'detected_sheets': detected_sheets,
        'skipped_trackers': skipped_trackers,
        'results': results,
        'validation_issue_count': sum((r.get('validation_issue_count', 0) for r in results if r)),
        'sheets_with_validation_issues': sum((1 for r in results if r and (r.get('validation_issue_count', 0) > 0))),
    }


def _process_file_sequential(input_file, output_folder, progress_callback=None):
    """
    Fallback sequential processing (original behaviour).
    Used when cache pre-load fails.  Still only runs trackers whose
    target sheet exists in the workbook.
    """
    input_path = Path(input_file)
    os.makedirs(output_folder, exist_ok=True)
    strict_timeline_mode = 'standard format' in input_path.name.lower()

    # Detect which sheets exist
    try:
        wb = openpyxl.load_workbook(str(input_file), read_only=True, data_only=True)
        file_sheet_names = wb.sheetnames
    except Exception:
        file_sheet_names = None
        wb = None

    if file_sheet_names:
        cache = {}
        for sname in file_sheet_names:
            ws = wb[sname] if wb is not None else None
            rows = []
            if ws is not None:
                try:
                    for row in ws.iter_rows(values_only=True):
                        rows.append(tuple(row))
                except Exception:
                    rows = []
            cache[sname] = {
                'rows': rows,
                'max_row': len(rows),
                'max_column': len(rows[0]) if rows else 0,
            }
        if wb is not None:
            wb.close()
        if strict_timeline_mode:
            matched_trackers = []
            skipped_trackers = []
            for sname in file_sheet_names:
                sheet_cache = cache.get(sname)
                if sheet_cache and _sheet_has_data_rows(sheet_cache) and _sheet_supports_timeline_processor(sheet_cache, strict=True):
                    safe = ''.join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in sname).strip().replace(' ', '_')
                    matched_trackers.append({
                        'name': f"{sname} (Timeline Deviation)",
                        'processor_class': TimelineDeviationProcessor,
                        'output_filename': f"{safe}_timeline_deviation_tracker.xlsx",
                        'target_sheet': sname,
                        'is_generic': False,
                    })
                else:
                    skipped_trackers.append(sname)
        else:
            matched_trackers, skipped_trackers = _match_trackers_to_sheets(file_sheet_names, cache, strict_timeline_mode=strict_timeline_mode)
    else:
        matched_trackers = [
            {
                'name': name,
                'processor_class': cls,
                'output_filename': out,
                'target_sheet': None,
                'is_generic': False,
            }
            for name, cls, out in ALL_TRACKERS
        ]
        skipped_trackers = []

    if not matched_trackers:
        print("[FALLBACK][SEQ] Strict tracker match found no sheets. Trying dynamic sheet detection...")
        try:
            fallback = UnifiedFileProcessor(str(input_file))
            dynamic_result = fallback.process_all(output_folder)
            dynamic_result['fallback_mode'] = 'dynamic_detection'
            return dynamic_result
        except Exception as fallback_err:
            return {
                'status': 'error',
                'message': 'No sheets in the uploaded file match any known tracker template.',
                'details': f'Dynamic fallback failed: {str(fallback_err)}',
                'total_sheets': 0, 'success_count': 0, 'error_count': 0,
                'failed_sheets': [], 'detected_sheets': [], 'results': [],
                'skipped_trackers': skipped_trackers,
            }

    total_matched = len(matched_trackers)
    if skipped_trackers:
        print(f"[SEQ] Skipping {len(skipped_trackers)} tracker(s) — no matching sheet")

    results = []
    success_count = 0
    error_count = 0

    for idx, tracker in enumerate(matched_trackers, 1):
        name = tracker['name']
        processor_class = tracker['processor_class']
        output_filename = tracker['output_filename']
        target_sheet = tracker.get('target_sheet')
        output_path = os.path.join(output_folder, output_filename)
        print(f"\n[{idx}/{total_matched}] {name}")
        print(f"     -> {output_filename}")

        if progress_callback:
            try:
                progress_callback(idx, total_matched, name, 'running', None)
            except Exception:
                pass

        try:
            processor = processor_class(str(input_file))
            if target_sheet:
                def _patched_load_workbook_for_target():
                    processor.workbook = openpyxl.load_workbook(str(input_file), data_only=True)
                    target = target_sheet if target_sheet in processor.workbook.sheetnames else processor.workbook.sheetnames[0]
                    ws = processor.workbook[target]
                    processor.sheet = ws
                    processor.eddr_sheet = ws
                    processor.pm_sheet = ws
                    if hasattr(processor, 'sheet_data'):
                        processor.sheet_data = [list(row) for row in ws.iter_rows(values_only=True)]
                    print(f"Loading workbook...\nUsing sheet '{target}': {ws.max_row} rows × {ws.max_column} columns")
                processor.load_workbook = _patched_load_workbook_for_target
            processor.process(output_path)

            file_size = os.path.getsize(output_path) if os.path.exists(output_path) else 0
            MIN_VALID_SIZE = 2048
            if file_size < MIN_VALID_SIZE:
                if os.path.exists(output_path):
                    os.remove(output_path)
                raise ValueError(
                    f"Output file is too small ({file_size} bytes) – no data was written."
                )

            validation_issues, data_row_count = _collect_output_validation(output_path, sheet_name=name)
            if data_row_count == 0:
                if os.path.exists(output_path):
                    os.remove(output_path)
                raise ValueError("Output file contains no data rows (blank output sheet).")

            success_count += 1
            print(f"     [OK] {file_size // 1024} KB")
            tracker_result = {
                'sheet_name': name,
                'description': name,
                'output_filename': output_filename,
                'output_path': output_path,
                'file_size': file_size,
                'status': 'success',
                'processor': processor_class.__name__,
                'validation_issues': validation_issues,
                'validation_issue_count': len(validation_issues),
            }
            results.append(tracker_result)
            if progress_callback:
                try:
                    progress_callback(idx, total_matched, name, 'success', tracker_result)
                except Exception:
                    pass
            del processor
            gc.collect()

        except Exception as e:
            error_count += 1
            err_msg = str(e)
            print(f"     [FAIL] {err_msg[:120]}")
            tracker_result = {
                'sheet_name': name,
                'description': name,
                'output_filename': output_filename,
                'status': 'error',
                'error': err_msg,
                'processor': processor_class.__name__,
                'validation_issues': [],
                'validation_issue_count': 0,
            }
            results.append(tracker_result)
            if progress_callback:
                try:
                    progress_callback(idx, total_matched, name, 'error', tracker_result)
                except Exception:
                    pass

    failed_sheets = [
        {'sheet_name': r['sheet_name'], 'error': r.get('error', ''), 'reason': r.get('error', '')}
        for r in results if r['status'] == 'error'
    ]
    detected_sheets = [r['sheet_name'] for r in results if r['status'] == 'success']

    print("\n" + "=" * 60)
    print(f"DONE  {success_count} OK  |  {error_count} FAILED  |  {total_matched} matched  |  {len(skipped_trackers)} skipped")
    print("=" * 60)

    if success_count > 0:
        msg = f'{success_count} of {total_matched} matching tracker(s) processed successfully'
    else:
        msg = 'All matching trackers failed to process'

    return {
        'status': 'success' if success_count > 0 else 'error',
        'message': msg,
        'total_sheets': total_matched,
        'success_count': success_count,
        'error_count': error_count,
        'failed_sheets': failed_sheets,
        'detected_sheets': detected_sheets,
        'skipped_trackers': skipped_trackers,
        'results': results,
        'validation_issue_count': sum((r.get('validation_issue_count', 0) for r in results)),
        'sheets_with_validation_issues': sum((1 for r in results if r.get('validation_issue_count', 0) > 0)),
    }


# Command-line usage
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python file_processor.py <input_file> [output_folder]")
        print("\nExample:")
        print('  python file_processor.py "data.xlsx" "outputs"')
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_folder = sys.argv[2] if len(sys.argv) > 2 else "outputs"
    
    try:
        result = process_file(input_file, output_folder)
        
        if result['status'] == 'success':
            print(f"\n✓ Processing completed successfully!")
            print(f"Output files saved to: {output_folder}")
        else:
            print(f"\n✗ Processing failed: {result['message']}")
            sys.exit(1)
            
    except Exception as e:
        print(f"\n✗ Fatal error: {str(e)}")
        sys.exit(1)